# ----------------------------------------------------------------------
# views_admin.py — Admin panel, supervisores, Bitrix24 integration.
# ----------------------------------------------------------------------

import json
import logging
import requests
import mimetypes
import os
from django.conf import settings
import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.db import models
from .models import TodoItem, Cliente, Cotizacion, DetalleCotizacion, UserProfile, Contacto, PendingFileUpload, OportunidadProyecto, Volumetria, DetalleVolumetria, CatalogoCableado, OportunidadActividad, OportunidadComentario, OportunidadArchivo, OportunidadEstado, Notificacion, Proyecto, ProyectoComentario, ProyectoArchivo, Tarea, TareaComentario, TareaArchivo, Actividad, CarpetaProyecto, ArchivoProyecto, CompartirArchivo, IntercambioNavidad, ParticipanteIntercambio, HistorialIntercambio, SolicitudAccesoProyecto, ArchivoFacturacion, CarpetaOportunidad, ArchivoOportunidad, MensajeOportunidad, TareaOportunidad, ComentarioTareaOpp, PostMuro, ComentarioMuro, ProductoOportunidad, AsistenciaJornada, EficienciaMensual, SolicitudCambioPerfil, ProgramacionActividad
from . import views_exportar
from .views_tarea_comentarios import api_comentarios_tarea, api_agregar_comentario_tarea, api_editar_comentario_tarea, api_eliminar_comentario_tarea
from .forms import VentaForm, VentaFilterForm, CotizacionForm, ClienteForm, OportunidadModalForm, NuevaOportunidadForm
from django.db.models import Sum, Count, F, Q, Case, When, Value
from django.db.models.functions import Upper, Coalesce
from django.db.models import Value
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from django.utils import timezone
from decimal import Decimal
import decimal
from django.utils.html import json_script

# Helper function to detect lost opportunities
from django.views.decorators.clickjacking import xframe_options_exempt, xframe_options_sameorigin
from .views_utils import *
from .views_utils import _exportar_estadisticas_excel

@login_required
def api_admin_usuarios(request):
    # El GET lo permitimos a cualquier autenticado para poder seleccionar usuarios en tareas
    if request.method != 'GET' and not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    
    if not request.user.is_authenticated:
         return JsonResponse({'error': 'No autenticado'}, status=401)

    if request.method == 'GET':
        usuarios = User.objects.select_related('userprofile').all().order_by('first_name', 'last_name')
        data = []
        for u in usuarios:
            grupos = [g.name for g in u.groups.all()]
            profile = getattr(u, 'userprofile', None)
            data.append({
                'id': u.id,
                'username': u.username,
                'first_name': u.first_name,
                'last_name': u.last_name,
                'email': u.email,
                'is_active': u.is_active,
                'is_supervisor': 'Supervisores' in grupos,
                'meta_mensual': str(profile.meta_mensual) if profile else '0',
                'meta_oportunidades': str(getattr(profile, 'meta_oportunidades', 0)) if profile else '0',
                'meta_cotizado': str(getattr(profile, 'meta_cotizado', 0)) if profile else '0',
                'meta_cobrado': str(getattr(profile, 'meta_cobrado', 0)) if profile else '0',
                'rol': getattr(profile, 'rol', 'vendedor') if profile else 'vendedor',
            })
        return JsonResponse({'usuarios': data})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        first_name = data.get('first_name', '').strip()
        last_name = data.get('last_name', '').strip()
        email = data.get('email', '').strip()

        if not username or not password:
            return JsonResponse({'error': 'Username y password son requeridos'}, status=400)

        if User.objects.filter(username=username).exists():
            return JsonResponse({'error': 'El username ya existe'}, status=400)

        from django.contrib.auth.models import Group
        user = User.objects.create_user(
            username=username, password=password,
            first_name=first_name, last_name=last_name, email=email
        )
        profile, _ = UserProfile.objects.get_or_create(user=user)
        rol = data.get('rol', 'vendedor').strip()
        if rol in ('vendedor', 'ingeniero'):
            profile.rol = rol
            profile.save(update_fields=['rol'])

        if data.get('is_supervisor'):
            grupo, _ = Group.objects.get_or_create(name='Supervisores')
            user.groups.add(grupo)

        return JsonResponse({'success': True, 'id': user.id})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_admin_usuario_detalle(request, user_id):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    try:
        usuario = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        if 'first_name' in data:
            usuario.first_name = data['first_name']
        if 'last_name' in data:
            usuario.last_name = data['last_name']
        if 'email' in data:
            usuario.email = data['email']
        if 'is_active' in data:
            usuario.is_active = data['is_active']
        if data.get('password', '').strip():
            usuario.set_password(data['password'].strip())
        usuario.save()

        # Actualizar rol en perfil
        if 'rol' in data:
            profile, _ = UserProfile.objects.get_or_create(user=usuario)
            profile.rol = data['rol']
            profile.save(update_fields=['rol'])

        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_admin_clientes(request):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method == 'GET':
        clientes = Cliente.objects.select_related('asignado_a').all().order_by('nombre_empresa')
        data = []
        for c in clientes:
            data.append({
                'id': c.id,
                'nombre_empresa': c.nombre_empresa,
                'categoria': c.categoria,
                'asignado_a_id': c.asignado_a_id,
                'asignado_a_name': c.asignado_a.get_full_name() or c.asignado_a.username if c.asignado_a else '',
                'telefono': c.telefono or '',
                'email': c.email or '',
                'meta_mensual': str(c.meta_mensual or 0),
            })
        return JsonResponse({'clientes': data})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_admin_cliente_detalle(request, cliente_id):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)

    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        if 'asignado_a_id' in data:
            uid = data['asignado_a_id']
            if uid:
                try:
                    cliente.asignado_a = User.objects.get(id=uid)
                except User.DoesNotExist:
                    return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
            else:
                cliente.asignado_a = None
        if 'categoria' in data:
            cliente.categoria = data['categoria']
        if 'meta_mensual' in data:
            try:
                cliente.meta_mensual = Decimal(str(data['meta_mensual']))
            except Exception:
                pass
        cliente.save()

        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_admin_contactos(request):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method == 'GET':
        cliente_id = request.GET.get('cliente_id')
        qs = Contacto.objects.select_related('empresa').all().order_by('nombre')
        if cliente_id:
            qs = qs.filter(empresa_id=cliente_id)
        data = []
        for c in qs:
            data.append({
                'id': c.id,
                'nombre': c.nombre,
                'apellido': c.apellido or '',
                'email': c.email or '',
                'telefono': c.telefono or '',
                'empresa_id': c.empresa_id,
                'empresa_name': c.empresa.nombre_empresa if c.empresa else '',
            })
        return JsonResponse({'contactos': data})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        nombre = data.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'Nombre es requerido'}, status=400)

        empresa_id = data.get('empresa_id')
        if not empresa_id:
            return JsonResponse({'error': 'Cliente es requerido'}, status=400)

        try:
            empresa = Cliente.objects.get(id=empresa_id)
        except Cliente.DoesNotExist:
            return JsonResponse({'error': 'Cliente no encontrado'}, status=404)

        contacto = Contacto.objects.create(
            nombre=nombre,
            apellido=data.get('apellido', ''),
            email=data.get('email', ''),
            telefono=data.get('telefono', ''),
            empresa=empresa,
        )
        return JsonResponse({'success': True, 'id': contacto.id})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_admin_metas(request):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        metas = data.get('metas', {})  # {uid: {mensual: x, oportunidades: y...}} or {uid: val}
        updated = 0
        for uid_str, meta_data in metas.items():
            try:
                uid = int(uid_str)
                profile, _ = UserProfile.objects.get_or_create(user_id=uid)
                
                # Check if meta_data is dict or value (legacy)
                if isinstance(meta_data, dict):
                    if 'mensual' in meta_data: profile.meta_mensual = Decimal(str(meta_data['mensual']))
                    if 'oportunidades' in meta_data: profile.meta_oportunidades = Decimal(str(meta_data['oportunidades']))
                    if 'cotizado' in meta_data: profile.meta_cotizado = Decimal(str(meta_data['cotizado']))
                    if 'cobrado' in meta_data: profile.meta_cobrado = Decimal(str(meta_data['cobrado']))
                else:
                    # Legacy flat value = mensual
                    profile.meta_mensual = Decimal(str(meta_data))
                
                profile.save()
                updated += 1
            except (ValueError, User.DoesNotExist, decimal.InvalidOperation):
                continue

        return JsonResponse({'success': True, 'updated': updated})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_admin_permisos(request, user_id):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        try:
            usuario = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

        from django.contrib.auth.models import Group
        grupo, _ = Group.objects.get_or_create(name='Supervisores')

        if data.get('is_supervisor'):
            usuario.groups.add(grupo)
        else:
            usuario.groups.remove(grupo)

        return JsonResponse({'success': True, 'is_supervisor': data.get('is_supervisor', False)})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@xframe_options_exempt
@login_required
def bitrix_widget_launcher(request):
    return render(request, 'bitrix_widget_launcher.html')


@csrf_exempt
def bitrix_cotizador_redirect(request):
    """
    Renderiza una plantilla con JavaScript para redirigir al cotizador en una nueva pestaña.
    """
    cotizador_url = "https://nethive.mx/app/crear-cotizacion/"
    print(f"DEBUG: Renderizando plantilla de redirección para: {cotizador_url}", flush=True)
    return render(request, 'bitrix_redirect_cotizador.html', {'cotizador_url': cotizador_url})


@login_required
def usuario_redirect_view(request):
    """
    Muestra la página de configuración avanzada y maneja el envío del formulario de avatar.
    """
    if request.method == 'POST':
        # Obtener o crear el perfil del usuario
        profile, created = UserProfile.objects.get_or_create(user=request.user)
        
        # Manejar la subida de imagen
        if 'avatar' in request.FILES and request.FILES['avatar']:
            profile.avatar = request.FILES['avatar']
            profile.usar_animado = False
            profile.avatar_tipo = '1'  # Reset a humano por defecto
        # Si se está seleccionando un avatar animado
        elif 'avatar_choice' in request.POST and request.POST['avatar_choice']:
            profile.usar_animado = True
            profile.avatar_tipo = request.POST['avatar_choice']  # Guardar el tipo seleccionado
            # Limpiar avatar anterior si existía
            profile.avatar = None
        
        profile.save()
        
        # Agregar mensaje de éxito
        messages.success(request, 'Avatar actualizado correctamente')
        
        # Redirigir a la configuración avanzada
        return redirect('configuracion_avanzada')
    
    # Si es GET, redirigir a configuración avanzada
    return redirect('configuracion_avanzada')


@supervisor_required
def reporte_usuarios(request):
    usuarios = User.objects.all().order_by('username')
    return render(request, 'reporte_usuarios.html', {'usuarios': usuarios})


@login_required
def perfil_usuario(request, usuario_id):
    from datetime import date
    from django.db.models import Sum, Count
    usuario = get_object_or_404(User, id=usuario_id)

    today = date.today()

    # --- Filtros de fecha (igual que estadisticas_usuarios) ---
    filtro = request.GET.get('filtro', 'todas')
    fecha_especifica = request.GET.get('fecha', None)
    mes_especifico = request.GET.get('mes', None)
    anio_especifico = request.GET.get('anio', None)

    oportunidades = TodoItem.objects.filter(usuario=usuario).select_related('cliente')
    cotizaciones = Cotizacion.objects.filter(created_by=usuario)

    titulo_filtro = "Todas las oportunidades"
    if filtro == 'dia':
        if fecha_especifica:
            try:
                fecha = date.fromisoformat(fecha_especifica)
            except ValueError:
                fecha = today
        else:
            fecha = today
        oportunidades = oportunidades.filter(fecha_creacion__date=fecha)
        cotizaciones = cotizaciones.filter(fecha_creacion__date=fecha)
        titulo_filtro = f"Día: {fecha.strftime('%d/%m/%Y')}"
    elif filtro == 'mes':
        if mes_especifico:
            try:
                año, mes = mes_especifico.split('-')
                año = int(año)
                mes = int(mes)
            except (ValueError, AttributeError):
                año = today.year
                mes = today.month
        else:
            año = today.year
            mes = today.month
        oportunidades = oportunidades.filter(fecha_creacion__year=año, fecha_creacion__month=mes)
        cotizaciones = cotizaciones.filter(fecha_creacion__year=año, fecha_creacion__month=mes)
        from calendar import month_name
        import locale
        try:
            locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        except:
            pass
        titulo_filtro = f"Mes: {today.replace(year=año, month=mes, day=1).strftime('%B %Y').capitalize()}"
    elif filtro == 'anio':
        if anio_especifico:
            try:
                año = int(anio_especifico)
            except ValueError:
                año = today.year
        else:
            año = today.year
        oportunidades = oportunidades.filter(fecha_creacion__year=año)
        cotizaciones = cotizaciones.filter(fecha_creacion__year=año)
        titulo_filtro = f"Año: {año}"

    # --- Estadísticas principales (mismas que la vista general) ---
    total_vendido = oportunidades.filter(probabilidad_cierre=100).aggregate(
        total=Sum('monto')
    )['total'] or Decimal('0.00')

    total_pipeline = oportunidades.aggregate(
        total=Sum('monto')
    )['total'] or Decimal('0.00')

    num_oportunidades = oportunidades.count()
    num_cotizaciones = cotizaciones.count()

    # Cliente top por monto
    cliente_top = oportunidades.values(
        'cliente__id', 'cliente__nombre_empresa'
    ).annotate(
        total_cliente=Sum('monto'),
        num_ops=Count('id')
    ).order_by('-total_cliente').first()

    # Oportunidad más grande (1-99% probabilidad)
    oportunidad_mayor = oportunidades.filter(
        probabilidad_cierre__gte=1, probabilidad_cierre__lte=99
    ).order_by('-monto').first()

    context = {
        'usuario': usuario,
        'oportunidades': oportunidades.order_by('-monto'),
        'oportunidad_mayor': oportunidad_mayor,
        'total_vendido': total_vendido,
        'total_pipeline': total_pipeline,
        'num_oportunidades': num_oportunidades,
        'num_cotizaciones': num_cotizaciones,
        'cliente_top': cliente_top,
        'filtro_actual': filtro,
        'titulo_filtro': titulo_filtro,
        'fecha_actual': today.isoformat(),
        'mes_actual': today.strftime('%Y-%m'),
        'anio_actual': str(today.year),
    }
    return render(request, 'perfil_usuario.html', context)


@login_required
def estadisticas_usuarios(request):
    """
    Vista para mostrar estadísticas de usuarios:
    - Número de oportunidades/cotizaciones por usuario
    - Total vendido por usuario
    - Cliente al que más venden
    Filtros: día, mes, período, todas
    """
    from datetime import date, timedelta
    from django.db.models import Sum, Count
    from django.db.models.functions import TruncMonth, TruncDay

    # Obtener parámetro de filtro
    filtro = request.GET.get('filtro', 'todas')
    fecha_especifica = request.GET.get('fecha', None)
    mes_especifico = request.GET.get('mes', None)
    fecha_desde = request.GET.get('fecha_desde', None)
    fecha_hasta = request.GET.get('fecha_hasta', None)

    today = date.today()

    # Base queryset de oportunidades
    oportunidades = TodoItem.objects.all()

    # Base queryset de cotizaciones (PDFs generados)
    cotizaciones = Cotizacion.objects.all()

    # Aplicar filtros
    titulo_filtro = "Todas las oportunidades"
    if filtro == 'dia':
        if fecha_especifica:
            try:
                fecha = date.fromisoformat(fecha_especifica)
            except ValueError:
                fecha = today
        else:
            fecha = today
        oportunidades = oportunidades.filter(fecha_creacion__date=fecha)
        cotizaciones = cotizaciones.filter(fecha_creacion__date=fecha)
        titulo_filtro = f"Día: {fecha.strftime('%d/%m/%Y')}"
    elif filtro == 'mes':
        if mes_especifico:
            try:
                año, mes = mes_especifico.split('-')
                año = int(año)
                mes = int(mes)
            except (ValueError, AttributeError):
                año = today.year
                mes = today.month
        else:
            año = today.year
            mes = today.month
        oportunidades = oportunidades.filter(fecha_creacion__year=año, fecha_creacion__month=mes)
        cotizaciones = cotizaciones.filter(fecha_creacion__year=año, fecha_creacion__month=mes)
        from calendar import month_name
        import locale
        try:
            locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        except:
            pass
        titulo_filtro = f"Mes: {today.replace(year=año, month=mes, day=1).strftime('%B %Y').capitalize()}"
    elif filtro == 'periodo':
        if fecha_desde and fecha_hasta:
            try:
                f_desde = date.fromisoformat(fecha_desde)
                f_hasta = date.fromisoformat(fecha_hasta)
            except ValueError:
                f_desde = today.replace(day=1)
                f_hasta = today
            oportunidades = oportunidades.filter(fecha_creacion__date__gte=f_desde, fecha_creacion__date__lte=f_hasta)
            cotizaciones = cotizaciones.filter(fecha_creacion__date__gte=f_desde, fecha_creacion__date__lte=f_hasta)
            titulo_filtro = f"Período: {f_desde.strftime('%d/%m/%Y')} — {f_hasta.strftime('%d/%m/%Y')}"
        else:
            titulo_filtro = "Todas las oportunidades"

    # Exportar a Excel
    if request.GET.get('exportar') == 'excel':
        return _exportar_estadisticas_excel(request, oportunidades, cotizaciones, titulo_filtro)

    # Obtener todos los usuarios que tienen oportunidades
    usuarios_con_datos = []
    usuarios = User.objects.filter(oportunidades__isnull=False).distinct()

    for usuario in usuarios:
        oportunidades_usuario = oportunidades.filter(usuario=usuario)

        # Número de oportunidades
        num_oportunidades = oportunidades_usuario.count()

        if num_oportunidades == 0:
            continue

        # Total vendido (oportunidades con probabilidad 100%)
        total_vendido = oportunidades_usuario.filter(probabilidad_cierre=100).aggregate(
            total=Sum('monto')
        )['total'] or Decimal('0.00')

        # Total en pipeline (todas las oportunidades)
        total_pipeline = oportunidades_usuario.aggregate(
            total=Sum('monto')
        )['total'] or Decimal('0.00')

        # Cliente al que más vende (por monto total)
        cliente_top = oportunidades_usuario.values(
            'cliente__id', 'cliente__nombre_empresa'
        ).annotate(
            total_cliente=Sum('monto'),
            num_ops=Count('id')
        ).order_by('-total_cliente').first()

        # Número de cotizaciones PDF generadas por este usuario
        num_cotizaciones = cotizaciones.filter(created_by=usuario).count()

        usuarios_con_datos.append({
            'usuario': usuario,
            'num_oportunidades': num_oportunidades,
            'num_cotizaciones': num_cotizaciones,
            'total_vendido': total_vendido,
            'total_pipeline': total_pipeline,
            'cliente_top': cliente_top,
        })

    # Ordenar por total vendido descendente
    usuarios_con_datos.sort(key=lambda x: x['total_vendido'], reverse=True)

    # Totales generales
    total_general_vendido = sum(u['total_vendido'] for u in usuarios_con_datos)
    total_general_pipeline = sum(u['total_pipeline'] for u in usuarios_con_datos)
    total_oportunidades = sum(u['num_oportunidades'] for u in usuarios_con_datos)
    total_cotizaciones = sum(u['num_cotizaciones'] for u in usuarios_con_datos)

    context = {
        'usuarios_datos': usuarios_con_datos,
        'filtro_actual': filtro,
        'titulo_filtro': titulo_filtro,
        'fecha_actual': today.isoformat(),
        'mes_actual': today.strftime('%Y-%m'),
        'fecha_desde': fecha_desde or '',
        'fecha_hasta': fecha_hasta or '',
        'total_general_vendido': total_general_vendido,
        'total_general_pipeline': total_general_pipeline,
        'total_oportunidades': total_oportunidades,
        'total_cotizaciones': total_cotizaciones,
    }

    return render(request, 'estadisticas_usuarios.html', context)


@csrf_exempt
def bitrix_webhook_receiver(request):
    """
    Receives webhook notifications from Bitrix24.
    This view handles the 'ONCRMDEALADD' event to create a new opportunity (TodoItem)
    in the local database when a new deal is created in Bitrix24.
    """
    # Log EVERY request to debug
    print(f"BITRIX WEBHOOK: Received {request.method} request from {request.META.get('REMOTE_ADDR', 'unknown')}", flush=True)
    print(f"BITRIX WEBHOOK: Request headers: {dict(request.headers)}", flush=True)
    print(f"BITRIX WEBHOOK: Request body: {request.body}", flush=True)
    
    if request.method != 'POST':
        print("BITRIX WEBHOOK: Received non-POST request. Ignoring.", flush=True)
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

    try:
        data = request.POST
        event = data.get('event')

        print(f"BITRIX WEBHOOK: POST data: {dict(data)}", flush=True)
        print(f"BITRIX WEBHOOK: Received event '{event}'", flush=True)

        if event in ['ONCRMDEALADD', 'ONCRMDEALUPDATE', 'ONCRMDEALDEL']:
            deal_id = data.get('data[FIELDS][ID]')
            if not deal_id:
                print(f"BITRIX WEBHOOK: '{event}' event received but no deal ID found.", flush=True)
                return JsonResponse({'status': 'error', 'message': 'Deal ID missing'}, status=400)

            print(f"BITRIX WEBHOOK: Processing {event} for Deal ID: {deal_id}", flush=True)
            
            # Check if this is a deletion event
            if event == 'ONCRMDEALDEL':
                print(f"BITRIX WEBHOOK: Processing DELETE event for Deal ID: {deal_id}", flush=True)
                try:
                    existing_opportunity = TodoItem.objects.get(bitrix_deal_id=deal_id)
                    opportunity_name = existing_opportunity.oportunidad
                    existing_opportunity.delete()
                    print(f"BITRIX WEBHOOK: Successfully deleted opportunity '{opportunity_name}' (Deal ID: {deal_id})", flush=True)
                    return JsonResponse({'status': 'success', 'message': f'Opportunity deleted: {opportunity_name}'})
                except TodoItem.DoesNotExist:
                    print(f"BITRIX WEBHOOK: Deal ID {deal_id} not found locally for deletion. Nothing to delete.", flush=True)
                    return JsonResponse({'status': 'success', 'message': 'Opportunity not found locally, nothing to delete'})
                except Exception as delete_error:
                    print(f"BITRIX WEBHOOK: Error deleting opportunity for Deal ID {deal_id}: {delete_error}", flush=True)
                    return JsonResponse({'status': 'error', 'message': f'Error deleting opportunity: {str(delete_error)}'}, status=500)

            # Check if this is an update to an existing deal
            is_update = event == 'ONCRMDEALUPDATE'
            existing_opportunity = None
            
            if is_update:
                try:
                    existing_opportunity = TodoItem.objects.get(bitrix_deal_id=deal_id)
                    print(f"BITRIX WEBHOOK: Found existing opportunity '{existing_opportunity.oportunidad}' for update", flush=True)
                except TodoItem.DoesNotExist:
                    print(f"BITRIX WEBHOOK: Deal ID {deal_id} not found locally. Will create new opportunity.", flush=True)
                    is_update = False
            else:
                # For ONCRMDEALADD, check for duplicates
                if TodoItem.objects.filter(bitrix_deal_id=deal_id).exists():
                    print(f"BITRIX WEBHOOK: An opportunity for Deal ID {deal_id} already exists. Skipping creation.", flush=True)
                    return JsonResponse({'status': 'success', 'message': 'Duplicate ignored'})

            # Intentar obtener detalles de la oportunidad con una llamada simple
            deal_details = None
            if BITRIX_WEBHOOK_URL:
                try:
                    get_url = BITRIX_WEBHOOK_URL.replace("crm.deal.add.json", "crm.deal.get.json")
                    response = requests.post(get_url, json={
                        'id': deal_id,
                        'select': [
                            'ID', 'TITLE', 'OPPORTUNITY', 'CURRENCY_ID', 'COMMENTS',
                            'COMPANY_ID', 'CONTACT_ID', 'ASSIGNED_BY_ID', 'STAGE_ID', 'CATEGORY_ID',
                            'UF_CRM_1752859685662',  # Producto (Runrate)
                            'UF_CRM_1750723256972',  # Solución (Proyectos)
                            'UF_CRM_1752859525038',  # Área
                            'UF_CRM_1752859877756',  # Mes de Cobro
                            'UF_CRM_1752855787179',  # Probabilidad de cierre
                            'UF_CRM_1755615484859',  # Utilidad
                        ]
                    }, timeout=5)
                    if response.status_code == 200:
                        deal_data = response.json()
                        if 'result' in deal_data and deal_data['result']:
                            deal_details = deal_data['result']
                            print(f"BITRIX WEBHOOK: Successfully fetched deal details for ID: {deal_id}", flush=True)
                        else:
                            print(f"BITRIX WEBHOOK: No result in API response for Deal ID: {deal_id}", flush=True)
                    else:
                        print(f"BITRIX WEBHOOK: API returned status {response.status_code} for Deal ID: {deal_id}", flush=True)
                except requests.exceptions.RequestException as e:
                    print(f"BITRIX WEBHOOK: Request failed for Deal ID {deal_id}: {e}", flush=True)
                except Exception as e:
                    print(f"BITRIX WEBHOOK: Unexpected error fetching deal details for ID {deal_id}: {e}", flush=True)
            
            # Si no se pudieron obtener los detalles, usar valores por defecto
            if not deal_details:
                deal_details = {
                    'ID': deal_id,
                    'TITLE': f'Oportunidad #{deal_id}',  # Título por defecto
                    'OPPORTUNITY': 0.0,  # Monto por defecto
                    'COMPANY_ID': None,  # Se buscará si hay datos disponibles
                    'ASSIGNED_BY_ID': None,  # Se buscará si hay datos disponibles
                    'CONTACT_ID': None,  # Se buscará si hay datos disponibles
                }
                print(f"BITRIX WEBHOOK: Using default deal details for Deal ID: {deal_id}", flush=True)
            else:
                print(f"BITRIX WEBHOOK: Using fetched deal details: Title='{deal_details.get('TITLE', 'N/A')}', Amount={deal_details.get('OPPORTUNITY', 'N/A')}", flush=True)

            # --- Find or Create the Associated Client (Company) ---
            company_id = deal_details.get('COMPANY_ID')
            cliente = None
            if company_id:
                try:
                    cliente = Cliente.objects.get(bitrix_company_id=company_id)
                    print(f"BITRIX WEBHOOK: Found existing client '{cliente.nombre_empresa}' for Company ID: {company_id}", flush=True)
                except Cliente.DoesNotExist:
                    print(f"BITRIX WEBHOOK: Client with Company ID {company_id} not found. Creating new client.", flush=True)
                    company_details = get_bitrix_company_details(company_id, request=request)
                    if company_details and company_details.get('TITLE'):
                        cliente = Cliente.objects.create(
                            bitrix_company_id=company_details.get('ID'),
                            nombre_empresa=company_details.get('TITLE')
                        )
                        print(f"BITRIX WEBHOOK: Created new client '{cliente.nombre_empresa}'", flush=True)
                    else:
                        print(f"BITRIX WEBHOOK: Could not fetch details for Company ID: {company_id}", flush=True)
            else:
                print("BITRIX WEBHOOK: Deal has no associated Company ID.", flush=True)

            # --- Find the Assigned User ---
            assigned_by_id = deal_details.get('ASSIGNED_BY_ID')
            usuario = None
            
            if assigned_by_id:
                try:
                    usuario = User.objects.filter(userprofile__bitrix_user_id=assigned_by_id).first()
                    if usuario:
                        print(f"BITRIX WEBHOOK: Found user '{usuario.username}' for Assigned ID: {assigned_by_id}", flush=True)
                    else:
                        print(f"BITRIX WEBHOOK: User with Bitrix User ID {assigned_by_id} not found in local DB.", flush=True)
                except Exception as e:
                    print(f"BITRIX WEBHOOK: Error finding user for Bitrix ID {assigned_by_id}: {e}", flush=True)
                    usuario = None
            else:
                print("BITRIX WEBHOOK: Deal has no assigned user.", flush=True)
            
            # --- Find or Create the Contact ---
            contact_id = deal_details.get('CONTACT_ID')
            contacto_obj = None
            
            if contact_id:
                try:
                    # Buscar contacto existente por bitrix_contact_id
                    contacto_obj = Contacto.objects.filter(bitrix_contact_id=contact_id).first()
                    if contacto_obj:
                        print(f"BITRIX WEBHOOK: Found existing contact '{contacto_obj.nombre} {contacto_obj.apellido}' for Contact ID: {contact_id}", flush=True)
                    else:
                        print(f"BITRIX WEBHOOK: Contact with ID {contact_id} not found locally. Fetching from Bitrix24...", flush=True)
                        # Obtener detalles del contacto desde Bitrix24
                        from .bitrix_integration import get_bitrix_contact_details
                        contact_details = get_bitrix_contact_details(contact_id, request=request)
                        if contact_details:
                            try:
                                contacto_obj = Contacto.objects.create(
                                    nombre=contact_details.get('NAME', 'Sin nombre'),
                                    apellido=contact_details.get('LAST_NAME', ''),
                                    bitrix_contact_id=contact_id,
                                    company_id=contact_details.get('COMPANY_ID'),
                                    cliente=cliente  # Asociar con el cliente
                                )
                                print(f"BITRIX WEBHOOK: Created new contact '{contacto_obj.nombre} {contacto_obj.apellido}'", flush=True)
                            except Exception as e:
                                print(f"BITRIX WEBHOOK: Error creating contact: {e}", flush=True)
                                contacto_obj = None
                        else:
                            print(f"BITRIX WEBHOOK: Could not fetch contact details from Bitrix24 for ID {contact_id}", flush=True)
                except Exception as e:
                    print(f"BITRIX WEBHOOK: Error processing contact ID {contact_id}: {e}", flush=True)
                    contacto_obj = None
            else:
                print("BITRIX WEBHOOK: Deal has no associated Contact ID.", flush=True)

            print(f"BITRIX WEBHOOK: DEBUG - After user assignment. is_update={locals().get('is_update', 'NOT_DEFINED')}, existing_opportunity={locals().get('existing_opportunity', 'NOT_DEFINED')}", flush=True)

            # --- Create the Opportunity (TodoItem) ---
            # Always create the opportunity, using defaults if needed
            if not usuario:
                print(f"BITRIX WEBHOOK: User not found, using default user", flush=True)
                usuario, _ = User.objects.get_or_create(
                    username='default_user',
                    defaults={
                        'first_name': 'Usuario',
                        'last_name': 'Sin Asignar',
                        'is_active': True
                    }
                )
            
            if not cliente:
                print(f"BITRIX WEBHOOK: Client not found, using default client", flush=True)
                cliente, _ = Cliente.objects.get_or_create(
                    nombre_empresa='Cliente Desconocido',
                    defaults={'bitrix_company_id': None}
                )

            # Now create the opportunity
            # Mapeos de Bitrix ID a Django Value - igual que en import_bitrix_opportunities.py
            PRODUCTO_BITRIX_ID_TO_DJANGO_VALUE = {
                "176": "ZEBRA", "178": "PANDUIT", "180": "APC", "182": "AVIGILION",
                "184": "GENETEC", "186": "AXIS", "188": "SOFTWARE", "190": "RUNRATE",
                "192": "PÓLIZA", "194": "CISCO", "374": "RFID", "376": "CONSUMIBLE",
                "378": "IMPRESORA INDUSTRIAL", "380": "SCANNER", "382": "TABLETA",
                "582": "SERVICIO",
            }

            AREA_BITRIX_ID_TO_DJANGO_VALUE = {
                "164": "SISTEMAS", "166": "Recursos Humanos", "168": "Compras",
                "170": "Seguridad", "172": "Mantenimiento", "174": "Almacén",
            }

            MES_COBRO_BITRIX_ID_TO_DJANGO_VALUE = {
                "196": "Enero", "198": "Febrero", "200": "Marzo", "202": "Abril",
                "204": "Mayo", "206": "Junio", "208": "Julio", "210": "Agosto",
                "212": "Septiembre", "214": "Octubre", "216": "Noviembre", "218": "Diciembre",
            }

            PROBABILIDAD_BITRIX_ID_TO_VALUE_STRING = {
                "220": "0%", "124": "10%", "126": "20%", "128": "30%",
                "130": "40%", "132": "50%", "134": "60%", "136": "70%",
                "138": "80%", "140": "90%", "142": "100%",
            }

            UTILIDAD_BITRIX_ID_TO_VALUE_STRING = {
                "718": "10%", "720": "15%", "722": "20%", 
                "724": "25%", "726": "30%", "728": "35%",
            }

            # Obtener valores raw de Bitrix
            area_bitrix_id = deal_details.get('UF_CRM_1752859525038')
            mes_cierre_bitrix_id = deal_details.get('UF_CRM_1752859877756')
            probabilidad_bitrix_id = deal_details.get('UF_CRM_1752855787179')
            utilidad_bitrix_id = deal_details.get('UF_CRM_1755615484859')
            category_id = deal_details.get('CATEGORY_ID')

            print(f"BITRIX WEBHOOK: Raw area ID: {area_bitrix_id}", flush=True)
            print(f"BITRIX WEBHOOK: Raw mes_cierre ID: {mes_cierre_bitrix_id}", flush=True)
            print(f"BITRIX WEBHOOK: Raw probabilidad ID: {probabilidad_bitrix_id}", flush=True)
            print(f"BITRIX WEBHOOK: Raw utilidad ID: {utilidad_bitrix_id}", flush=True)
            print(f"BITRIX WEBHOOK: Raw category ID: {category_id}", flush=True)

            # Mapear tipo de negociación desde CATEGORY_ID
            tipo_negociacion = map_bitrix_category_to_tipo_negociacion(category_id)
            
            # Aplicar conversiones de mapeo
            area = AREA_BITRIX_ID_TO_DJANGO_VALUE.get(str(area_bitrix_id), 'SISTEMAS') # Default
            mes_cierre = MES_COBRO_BITRIX_ID_TO_DJANGO_VALUE.get(str(mes_cierre_bitrix_id), 'Enero') # Default
            
            # Mapear producto/solución según el tipo de negociación
            producto = get_producto_from_bitrix_deal(deal_details, tipo_negociacion)
            
            # Mapear etapa según el tipo de negociación
            stage_id = deal_details.get('STAGE_ID')
            etapa_corta, etapa_completa, etapa_color = get_etapa_from_bitrix_stage(stage_id, tipo_negociacion)

            probabilidad_cierre = 0 # Default value
            if probabilidad_bitrix_id is not None:
                prob_str = PROBABILIDAD_BITRIX_ID_TO_VALUE_STRING.get(str(probabilidad_bitrix_id))
                if prob_str:
                    try:
                        parsed_prob = int(prob_str.replace('%', ''))
                        probabilidad_cierre = min(parsed_prob, 100) # Cap at 100
                    except ValueError:
                        print(f"BITRIX WEBHOOK: Invalid probability string from Bitrix: {prob_str}. Setting to default (0).", flush=True)
                        probabilidad_cierre = 0
                else:
                    print(f"BITRIX WEBHOOK: Unknown probability ID from Bitrix: {probabilidad_bitrix_id}. Setting to default (0).", flush=True)
                    probabilidad_cierre = 0

            # Procesar porcentaje de utilidad
            porcentaje_utilidad = 0 # Default value
            if utilidad_bitrix_id is not None:
                utilidad_str = UTILIDAD_BITRIX_ID_TO_VALUE_STRING.get(str(utilidad_bitrix_id))
                if utilidad_str:
                    try:
                        parsed_utilidad = int(utilidad_str.replace('%', ''))
                        porcentaje_utilidad = min(parsed_utilidad, 100) # Cap at 100
                        print(f"BITRIX WEBHOOK: Porcentaje de utilidad configurado: {porcentaje_utilidad}%", flush=True)
                    except ValueError:
                        print(f"BITRIX WEBHOOK: Invalid utilidad string from Bitrix: {utilidad_str}. Setting to default (0).", flush=True)
                        porcentaje_utilidad = 0
                else:
                    print(f"BITRIX WEBHOOK: Unknown utilidad ID from Bitrix: {utilidad_bitrix_id}. Setting to default (0).", flush=True)
                    porcentaje_utilidad = 0

                print(f"BITRIX WEBHOOK: Mapped product: {producto}", flush=True)
                print(f"BITRIX WEBHOOK: Mapped area: {area}", flush=True)
                print(f"BITRIX WEBHOOK: Mapped mes_cierre: {mes_cierre}", flush=True)
                print(f"BITRIX WEBHOOK: Mapped probabilidad_cierre: {probabilidad_cierre}", flush=True)
                print(f"BITRIX WEBHOOK: About to create opportunity with usuario={usuario}, cliente={cliente}", flush=True)

                if is_update and existing_opportunity:
                    # Verificar si cambió el producto (para detectar cuando se agrega Zebra)
                    producto_anterior = existing_opportunity.producto
                    print(f"BITRIX WEBHOOK: PRODUCTO DEBUG - Anterior: '{producto_anterior}' | Nuevo: '{producto}' | Son diferentes: {producto_anterior != producto}", flush=True)
                    
                    # Obtener el estado de Bitrix24
                    bitrix_stage_id = deal_details.get('STAGE_ID')
                    print(f"BITRIX WEBHOOK: Raw STAGE_ID: {bitrix_stage_id}", flush=True)
                    
                    # Actualizar oportunidad existente
                    existing_opportunity.oportunidad = deal_details.get('TITLE', existing_opportunity.oportunidad)
                    existing_opportunity.monto = deal_details.get('OPPORTUNITY', existing_opportunity.monto) or existing_opportunity.monto
                    existing_opportunity.cliente = cliente or existing_opportunity.cliente
                    existing_opportunity.usuario = usuario or existing_opportunity.usuario
                    existing_opportunity.producto = producto
                    existing_opportunity.area = area
                    existing_opportunity.mes_cierre = mes_cierre
                    existing_opportunity.tipo_negociacion = tipo_negociacion  # Nuevo campo
                    existing_opportunity.etapa_corta = etapa_corta
                    existing_opportunity.etapa_completa = etapa_completa
                    existing_opportunity.etapa_color = etapa_color
                    existing_opportunity.probabilidad_cierre = probabilidad_cierre
                    existing_opportunity.bitrix_stage_id = bitrix_stage_id  # Guardar el estado
                    existing_opportunity.contacto = contacto_obj  # Asignar contacto
                    existing_opportunity.save()
                    print(f"BITRIX WEBHOOK: Successfully updated opportunity '{existing_opportunity.oportunidad}' with ID {existing_opportunity.id}", flush=True)
                    
                    # ========================================
                    # VERIFICAR SI CAMBIÓ EL PRODUCTO A UNA MARCA AUTOMÁTICA
                    # ========================================
                    if producto_anterior != producto:
                        print(f"BITRIX WEBHOOK: Producto cambió de '{producto_anterior}' a '{producto}' - Verificando cotización automática", flush=True)
                        
                        # Verificar si debe generar cotización automática
                        if es_cotizacion_automatica(deal_details):
                            print(f"BITRIX WEBHOOK: Oportunidad actualizada califica para cotización automática - Marca: {producto}", flush=True)
                            
                            # Verificar si ya tiene cotización para evitar duplicados
                            from app.models import Cotizacion
                            cotizacion_existente = Cotizacion.objects.filter(oportunidad=existing_opportunity).first()
                            
                            if cotizacion_existente:
                                print(f"BITRIX WEBHOOK: Ya existe cotización ID {cotizacion_existente.id} para esta oportunidad - No se creará duplicado", flush=True)
                            else:
                                # Obtener usuario final del contacto
                                contact_id = deal_details.get('CONTACT_ID')
                                usuario_final = ''
                                if contact_id:
                                    try:
                                        from .bitrix_integration import get_bitrix_contact_details
                                        contact_details = get_bitrix_contact_details(contact_id, request=request)
                                        if contact_details:
                                            usuario_final = f"{contact_details.get('NAME', '')} {contact_details.get('LAST_NAME', '')}".strip()
                                            print(f"BITRIX WEBHOOK: Usuario final obtenido: {usuario_final}", flush=True)
                                    except Exception as e:
                                        print(f"BITRIX WEBHOOK: Error obteniendo contacto: {e}", flush=True)
                                
                                # Crear cotización automática
                                cotizacion_automatica = crear_cotizacion_automatica_bitrix(deal_details, cliente, usuario, porcentaje_utilidad)
                                
                                if cotizacion_automatica:
                                    # Agregar usuario final si se obtuvo
                                    if usuario_final:
                                        cotizacion_automatica.usuario_final = usuario_final
                                        cotizacion_automatica.save()
                                    
                                    # Subir PDF a Bitrix24
                                    resultado_subida = subir_cotizacion_a_bitrix(cotizacion_automatica, deal_id, request)
                                    
                                    if resultado_subida:
                                        print(f"BITRIX WEBHOOK: ✅ Cotización automática completada exitosamente para deal actualizado {deal_id}")
                                    else:
                                        print(f"BITRIX WEBHOOK: ⚠️ Cotización creada pero falló la subida a Bitrix24 para deal actualizado {deal_id}")
                                else:
                                    print(f"BITRIX WEBHOOK: ❌ No se pudo crear la cotización automática para deal actualizado {deal_id}")
                        else:
                            print(f"BITRIX WEBHOOK: Oportunidad actualizada NO califica para cotización automática - Producto: {producto}", flush=True)
                else:
                    # Crear nueva oportunidad
                    print(f"BITRIX WEBHOOK: Creating new opportunity with data:", flush=True)
                    print(f"  - Title: {deal_details.get('TITLE', 'Oportunidad sin título')}", flush=True)
                    print(f"  - Amount: {deal_details.get('OPPORTUNITY', 0.0)}", flush=True)
                    print(f"  - Cliente: {cliente}", flush=True)
                    print(f"  - Usuario: {usuario}", flush=True)
                    print(f"  - Deal ID: {deal_id}", flush=True)
                    
                    # Obtener el estado de Bitrix24
                    bitrix_stage_id = deal_details.get('STAGE_ID')
                    print(f"BITRIX WEBHOOK: Raw STAGE_ID for new opportunity: {bitrix_stage_id}", flush=True)
                    
                    try:
                        new_opportunity = TodoItem.objects.create(
                            oportunidad=deal_details.get('TITLE', 'Oportunidad sin título'),
                            monto=deal_details.get('OPPORTUNITY', 0.0) or 0.0,
                            cliente=cliente,
                            usuario=usuario,
                            bitrix_deal_id=deal_id,
                            producto=producto,
                            area=area,
                            mes_cierre=mes_cierre,
                            tipo_negociacion=tipo_negociacion,  # Nuevo campo
                            etapa_corta=etapa_corta,
                            etapa_completa=etapa_completa,
                            etapa_color=etapa_color,
                            probabilidad_cierre=probabilidad_cierre,
                            bitrix_stage_id=bitrix_stage_id,  # Guardar el estado
                            contacto=contacto_obj,  # Asignar contacto
                        )
                        print(f"BITRIX WEBHOOK: Successfully created new opportunity '{new_opportunity.oportunidad}' with ID {new_opportunity.id}", flush=True)
                        
                        # ========================================
                        # LÓGICA DE COTIZACIÓN AUTOMÁTICA
                        # ========================================
                        
                        # Verificar si debe generar cotización automática
                        if es_cotizacion_automatica(deal_details):
                            print(f"BITRIX WEBHOOK: Oportunidad califica para cotización automática - Marca: {producto}", flush=True)
                            
                            # Obtener usuario final del contacto
                            contact_id = deal_details.get('CONTACT_ID')
                            usuario_final = ''
                            if contact_id:
                                try:
                                    from .bitrix_integration import get_bitrix_contact_details
                                    contact_details = get_bitrix_contact_details(contact_id, request=request)
                                    if contact_details:
                                        usuario_final = f"{contact_details.get('NAME', '')} {contact_details.get('LAST_NAME', '')}".strip()
                                        print(f"BITRIX WEBHOOK: Usuario final obtenido: {usuario_final}", flush=True)
                                except Exception as e:
                                    print(f"BITRIX WEBHOOK: Error obteniendo contacto: {e}", flush=True)
                            
                            # Crear cotización automática
                            cotizacion_automatica = crear_cotizacion_automatica_bitrix(deal_details, cliente, usuario, porcentaje_utilidad)
                            
                            if cotizacion_automatica:
                                # Agregar usuario final si se obtuvo
                                if usuario_final:
                                    cotizacion_automatica.usuario_final = usuario_final
                                    cotizacion_automatica.save()
                                
                                # Subir PDF a Bitrix24
                                resultado_subida = subir_cotizacion_a_bitrix(cotizacion_automatica, deal_id, request)
                                
                                if resultado_subida:
                                    print(f"BITRIX WEBHOOK: ✅ Cotización automática completada exitosamente para deal {deal_id}")
                                else:
                                    print(f"BITRIX WEBHOOK: ⚠️ Cotización creada pero falló la subida a Bitrix24 para deal {deal_id}")
                            else:
                                print(f"BITRIX WEBHOOK: ❌ No se pudo crear la cotización automática para deal {deal_id}")
                        else:
                            print(f"BITRIX WEBHOOK: Oportunidad NO califica para cotización automática - Producto: {producto}", flush=True)
                        
                    except Exception as create_error:
                        print(f"BITRIX WEBHOOK: Error creating opportunity: {create_error}", flush=True)
                        print(f"BITRIX WEBHOOK: Traceback: {traceback.format_exc()}", flush=True)

        return JsonResponse({'status': 'success'})

    except Exception as e:
        print(f"BITRIX WEBHOOK: An unexpected error occurred: {e}", flush=True)
        print(traceback.format_exc(), flush=True)
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
@login_required
def sync_bitrix_manual(request):
    """
    Vista para que los supervisores ejecuten sincronización manual con Bitrix
    """
    # Verificar que el usuario sea supervisor
    if not is_supervisor(request.user):
        return JsonResponse({
            'success': False,
            'error': 'Solo los supervisores pueden ejecutar sincronización'
        })
    
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        })
    
    try:
        from django.core.management import call_command
        import io
        import sys
        from contextlib import redirect_stdout, redirect_stderr
        
        # Capturar la salida de los comandos
        stdout_buffer = io.StringIO()
        stderr_buffer = io.StringIO()
        
        results = {
            'success': [],
            'errors': [],
            'output': []
        }
        
        sync_commands = [
            ('sync_bitrix_users', 'Usuarios de Bitrix'),
            ('sync_bitrix', 'Empresas/Clientes'),
            ('sync_bitrix_contacts', 'Contactos'),
            ('import_bitrix_opportunities', 'Oportunidades')
        ]
        
        for command, description in sync_commands:
            try:
                stdout_buffer.seek(0)
                stdout_buffer.truncate(0)
                stderr_buffer.seek(0) 
                stderr_buffer.truncate(0)
                
                with redirect_stdout(stdout_buffer), redirect_stderr(stderr_buffer):
                    call_command(command)
                
                output = stdout_buffer.getvalue()
                results['success'].append({
                    'command': command,
                    'description': description,
                    'output': output[:500]  # Limitar output
                })
                
            except Exception as e:
                error_output = stderr_buffer.getvalue()
                results['errors'].append({
                    'command': command,
                    'description': description,
                    'error': str(e),
                    'output': error_output[:500]
                })
        
        # Estadísticas finales
        from app.models import Cliente, TodoItem, UserProfile
        from django.contrib.auth.models import User
        
        stats = {
            'usuarios': User.objects.count(),
            'profiles_bitrix': UserProfile.objects.filter(bitrix_user_id__isnull=False).count(),
            'clientes': Cliente.objects.count(),
            'oportunidades': TodoItem.objects.count()
        }
        
        return JsonResponse({
            'success': True,
            'message': f'Sincronización completada: {len(results["success"])} exitosos, {len(results["errors"])} errores',
            'results': results,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error general en sincronización: {str(e)}'
        })


@login_required
def search_clientes_api(request):
    """
    API endpoint para buscar clientes por nombre de empresa.
    Permite búsqueda parcial insensible a mayúsculas y minúsculas.
    """
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({'clientes': []})
    
    try:
        # Filtrar clientes basado en permisos del usuario
        if is_supervisor(request.user):
            # Supervisores pueden ver todos los clientes
            clientes_queryset = Cliente.objects.filter(
                nombre_empresa__icontains=query
            ).order_by('nombre_empresa')[:20]  # Limitar a 20 resultados
        else:
            # Usuarios normales solo ven sus clientes asignados
            clientes_queryset = Cliente.objects.filter(
                asignado_a=request.user,
                nombre_empresa__icontains=query
            ).order_by('nombre_empresa')[:20]
        
        # Serializar los resultados
        clientes_data = []
        for cliente in clientes_queryset:
            clientes_data.append({
                'id': cliente.id,
                'nombre_empresa': cliente.nombre_empresa,
                'contacto_principal': cliente.contacto_principal,
                'email': cliente.email,
                'telefono': cliente.telefono,
            })
        
        return JsonResponse({
            'clientes': clientes_data,
            'total': len(clientes_data)
        })
        
    except Exception as e:
        print(f"ERROR: Error en búsqueda de clientes: {e}")
        return JsonResponse({
            'clientes': [],
            'error': 'Error interno del servidor'
        }, status=500)


def crear_cotizacion_automatica_bitrix(deal_details, cliente, usuario, porcentaje_utilidad=0):
    """
    Crea una cotización automática basada en los datos de Bitrix24
    """
    try:
        # Obtener la marca del campo producto
        PRODUCTO_BITRIX_ID_TO_DJANGO_VALUE = {
            "176": "ZEBRA", "178": "PANDUIT", "180": "APC", "182": "AVIGILON",
            "184": "GENETEC", "186": "AXIS", "194": "CISCO"
        }
        
        producto_bitrix_id = deal_details.get('UF_CRM_1752859685662')
        marca_seleccionada = PRODUCTO_BITRIX_ID_TO_DJANGO_VALUE.get(str(producto_bitrix_id))
        
        if not marca_seleccionada:
            print(f"COTIZACIÓN AUTOMÁTICA: Marca no válida para ID {producto_bitrix_id}")
            return None
            
        # Obtener el texto de requisición del cliente desde COMMENTS
        # Los ejecutivos pueden escribir números de parte aquí en cualquier formato
        requisicion_cliente = deal_details.get('COMMENTS', '')
        
        print(f"COTIZACIÓN AUTOMÁTICA: Texto de requisición obtenido: '{requisicion_cliente[:100]}...'", flush=True)
        
        # Extraer productos inteligentemente
        productos = extraer_productos_inteligente(requisicion_cliente, marca_seleccionada)
        
        if not productos:
            print(f"COTIZACIÓN AUTOMÁTICA: No se encontraron productos válidos en la requisición")
            return None
            
        # Crear la cotización
        from app.models import Cotizacion, DetalleCotizacion
        
        # Usar el título del deal directamente
        titulo_cotizacion = deal_details.get('TITLE', 'Sin título')

        # Encontrar la oportunidad local correspondiente
        oportunidad_local = None
        deal_id = deal_details.get('ID')
        if deal_id:
            oportunidad_local = TodoItem.objects.filter(bitrix_deal_id=deal_id).first()
        
        cotizacion = Cotizacion.objects.create(
            cliente=cliente,
            created_by=usuario,
            titulo=titulo_cotizacion,
            nombre_cotizacion=titulo_cotizacion,
            oportunidad=oportunidad_local, # Asociar la oportunidad local
            descripcion="", # Descripción vacía
            moneda='USD',
            iva_rate=0.16,  # 16% por defecto
            tipo_cotizacion='Iamet',  # IAMET por defecto
            subtotal=0,  # Se calculará después
            iva_amount=0,  # Se calculará después
            total=0,  # Se calculará después
            descuento_visible=False
        )
        
        # Agregar productos a la cotización
        subtotal = 0
        print(f"COTIZACIÓN AUTOMÁTICA: Aplicando {porcentaje_utilidad}% de utilidad a los precios", flush=True)
        
        for i, producto in enumerate(productos, 1):
            # Aplicar porcentaje de utilidad al precio base
            precio_base = producto['precio']
            if porcentaje_utilidad > 0:
                factor_utilidad = 1 + (porcentaje_utilidad / 100)
                precio_con_utilidad = precio_base * factor_utilidad
                total_con_utilidad = precio_con_utilidad * producto['cantidad']
                print(f"COTIZACIÓN AUTOMÁTICA: Producto {producto['no_parte']} - Precio base: ${precio_base:.2f} → Con {porcentaje_utilidad}% utilidad: ${precio_con_utilidad:.2f}", flush=True)
            else:
                precio_con_utilidad = precio_base
                total_con_utilidad = precio_base * producto['cantidad']
            
            DetalleCotizacion.objects.create(
                cotizacion=cotizacion,
                orden=i,
                marca=marca_seleccionada,
                nombre_producto=producto['no_parte'],
                no_parte=producto['no_parte'],
                descripcion=producto['descripcion'],
                cantidad=producto['cantidad'],
                precio_unitario=precio_con_utilidad,
                descuento_porcentaje=0,
                total=total_con_utilidad
            )
            subtotal += total_con_utilidad
        
        # Calcular totales
        iva_amount = Decimal(str(subtotal)) * Decimal('0.16')
        total = Decimal(str(subtotal)) + iva_amount
        
        # Actualizar cotización con totales
        cotizacion.subtotal = Decimal(str(subtotal))
        cotizacion.iva_amount = iva_amount
        cotizacion.total = total
        cotizacion.save()
        
        print(f"COTIZACIÓN AUTOMÁTICA: Creada cotización {cotizacion.id} con {len(productos)} productos. Total: ${total:.2f}")
        return cotizacion
        
    except Exception as e:
        print(f"COTIZACIÓN AUTOMÁTICA: Error creando cotización: {e}")
        return None


def subir_cotizacion_a_bitrix(cotizacion, deal_id, request=None):
    """
    Genera el PDF de la cotización y lo sube como comentario a Bitrix24
    """
    try:
        from django.test import RequestFactory
        import base64
        
        # Crear una petición temporal para generar el PDF
        factory = RequestFactory()
        pdf_request = factory.get(f'/app/cotizacion/{cotizacion.id}/pdf/')
        pdf_request.user = cotizacion.created_by
        
        # Generar el PDF
        pdf_response = generate_cotizacion_pdf(pdf_request, cotizacion.id)
        
        if pdf_response.status_code == 200:
            # Convertir PDF a base64
            pdf_content = pdf_response.content
            pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
            
            # Nombre del archivo
            pdf_name_raw = cotizacion.nombre_cotizacion or f"Cotizacion_{cotizacion.id}"
            sanitized_name = "".join(c for c in pdf_name_raw if c.isalnum() or c in ('_', '-')).strip().replace(' ', '_')
            pdf_filename = f"{sanitized_name}.pdf"
            
            # Comentario para Bitrix
            comentario_texto = f"""
🤖 Cotización Automática Generada

Cotización: {cotizacion.nombre_cotizacion}
Total: ${float(cotizacion.total):.2f} {cotizacion.moneda}
Productos: {cotizacion.detalles.count()} artículos
Generada por: {cotizacion.created_by.get_full_name() or cotizacion.created_by.username}
Fecha: {convert_to_tijuana_time(cotizacion.fecha_creacion).strftime('%d/%m/%Y %H:%M')}

⚡ Generada automáticamente desde el sistema Nethive
            """.strip()
            
            # Subir a Bitrix24
            from .bitrix_integration import add_comment_with_attachment_to_deal
            resultado_bitrix = add_comment_with_attachment_to_deal(
                deal_id=deal_id,
                file_name=pdf_filename,
                file_content_base64=pdf_base64,
                comment_text=comentario_texto,
                request=request
            )
            
            if resultado_bitrix:
                print(f"COTIZACIÓN AUTOMÁTICA: PDF subido exitosamente a Bitrix24 deal {deal_id}")
                return True
            else:
                print(f"COTIZACIÓN AUTOMÁTICA: Error subiendo PDF a Bitrix24 deal {deal_id}")
                return False
                
        else:
            print(f"COTIZACIÓN AUTOMÁTICA: Error generando PDF (status {pdf_response.status_code})")
            return False
            
    except Exception as e:
        print(f"COTIZACIÓN AUTOMÁTICA: Error subiendo a Bitrix24: {e}")
        return False


@login_required
@require_http_methods(["POST"])
def actualizar_avatar(request):
    """
    API para actualizar el avatar del usuario
    """
    try:
        # Obtener o crear el perfil del usuario
        user_profile, created = UserProfile.objects.get_or_create(user=request.user)
        
        # Verificar si se envió un archivo de avatar
        if 'avatar' in request.FILES:
            avatar_file = request.FILES['avatar']
            
            # Validar tipo de archivo
            allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp']
            if avatar_file.content_type not in allowed_types:
                return JsonResponse({
                    'success': False, 
                    'error': 'Tipo de archivo no permitido. Solo se permiten imágenes (JPG, PNG, GIF, WebP)'
                })
            
            # Validar tamaño de archivo (máximo 5MB)
            if avatar_file.size > 5 * 1024 * 1024:
                return JsonResponse({
                    'success': False, 
                    'error': 'El archivo es demasiado grande. Máximo 5MB permitido'
                })
            
            # Eliminar avatar anterior si existe
            if user_profile.avatar:
                try:
                    user_profile.avatar.delete(save=False)
                except:
                    pass  # Ignorar errores al eliminar archivo anterior
            
            # Guardar nuevo avatar
            user_profile.avatar = avatar_file
            user_profile.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Avatar actualizado correctamente',
                'avatar_url': user_profile.get_avatar_url()
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'No se envió ningún archivo de avatar'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar avatar: {str(e)}'
        })


@login_required
@user_passes_test(lambda u: u.is_superuser)
def bitrix_sync_admin(request):
    """
    Vista de administración para sincronización con Bitrix24
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'sync_deletions':
            # Sincronizar eliminaciones
            from .bitrix_integration import get_all_bitrix_deals
            
            try:
                # Obtener todas las oportunidades locales con bitrix_deal_id
                local_opportunities = TodoItem.objects.filter(bitrix_deal_id__isnull=False)
                
                # Obtener todos los deals de Bitrix24
                bitrix_deals = get_all_bitrix_deals()
                if not bitrix_deals:
                    messages.error(request, 'No se pudieron obtener deals de Bitrix24')
                    return redirect('bitrix-sync-admin')
                
                # Crear set de IDs que existen en Bitrix24
                bitrix_deal_ids = {str(deal['ID']) for deal in bitrix_deals}
                
                # Encontrar oportunidades locales que ya no existen en Bitrix24
                opportunities_to_delete = []
                for opportunity in local_opportunities:
                    if str(opportunity.bitrix_deal_id) not in bitrix_deal_ids:
                        opportunities_to_delete.append(opportunity)
                
                # Eliminar las oportunidades
                deleted_count = 0
                for opportunity in opportunities_to_delete:
                    try:
                        opportunity.delete()
                        deleted_count += 1
                    except Exception as e:
                        print(f"Error eliminando oportunidad {opportunity.id}: {e}")
                
                if deleted_count > 0:
                    messages.success(request, f'Se eliminaron {deleted_count} oportunidades que ya no existen en Bitrix24')
                else:
                    messages.info(request, 'No hay oportunidades para eliminar. Todo está sincronizado.')
                    
            except Exception as e:
                messages.error(request, f'Error durante la sincronización: {str(e)}')
        
        elif action == 'import_all':
            # Importar todas las oportunidades
            from django.core.management import call_command
            from io import StringIO
            
            try:
                out = StringIO()
                call_command('import_bitrix_opportunities', stdout=out)
                output = out.getvalue()
                messages.success(request, f'Importación completada: {output}')
            except Exception as e:
                messages.error(request, f'Error durante la importación: {str(e)}')
        
        elif action == 'sync_lost_opportunities':
            # Ejecutar sincronización de oportunidades perdidas
            try:
                from django.core.management import call_command
                from io import StringIO
                
                # Capturar la salida del comando
                out = StringIO()
                call_command('sync_lost_opportunities', stdout=out)
                output = out.getvalue()
                
                # Extraer estadísticas del output
                lines = output.split('\n')
                stats_found = False
                for line in lines:
                    if 'Oportunidades perdidas encontradas:' in line:
                        found_count = line.split(':')[-1].strip()
                        messages.success(
                            request, 
                            f'Sincronización completada. Se encontraron {found_count} oportunidades perdidas que fueron actualizadas.'
                        )
                        stats_found = True
                        break
                
                if not stats_found:
                    messages.success(request, 'Sincronización de oportunidades perdidas completada')
                    
            except Exception as e:
                messages.error(request, f'Error durante la sincronización: {str(e)}')
        
        elif action == 'full_bitrix_sync':
            # Ejecutar sincronización completa bidireccional
            try:
                from django.core.management import call_command
                from io import StringIO
                
                # Capturar la salida del comando
                out = StringIO()
                call_command('full_bitrix_sync', stdout=out)
                output = out.getvalue()
                
                # Extraer estadísticas del output
                lines = output.split('\n')
                for line in lines:
                    if 'Nuevas importadas:' in line:
                        new_count = line.split(':')[-1].strip()
                    elif 'Actualizadas:' in line:
                        updated_count = line.split(':')[-1].strip()
                    elif 'Perdidas detectadas:' in line:
                        lost_count = line.split(':')[-1].strip()
                
                messages.success(
                    request, 
                    f'Sincronización completa terminada. Nuevas: {new_count if "new_count" in locals() else "0"}, '
                    f'Actualizadas: {updated_count if "updated_count" in locals() else "0"}, '
                    f'Perdidas: {lost_count if "lost_count" in locals() else "0"}'
                )
                    
            except Exception as e:
                messages.error(request, f'Error durante la sincronización completa: {str(e)}')
        
        return redirect('bitrix-sync-admin')
    
    # GET request - mostrar estadísticas
    context = {
        'total_opportunities': TodoItem.objects.count(),
        'active_opportunities': TodoItem.objects.count() - len([item for item in TodoItem.objects.all() if is_lost_opportunity(item.bitrix_stage_id)]),
        'lost_opportunities': len([item for item in TodoItem.objects.all() if is_lost_opportunity(item.bitrix_stage_id)]),
        'bitrix_opportunities': TodoItem.objects.filter(bitrix_deal_id__isnull=False).count(),
        'local_opportunities': TodoItem.objects.filter(bitrix_deal_id__isnull=True).count(),
    }
    
    return render(request, 'bitrix_sync_admin.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
@login_required  
def bitrix_lost_opportunities(request):
    """
    Vista para mostrar y gestionar oportunidades marcadas como "Cerrado Perdido"
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        opportunity_id = request.POST.get('opportunity_id')
        
        if opportunity_id and action in ['delete', 'reactivate']:
            try:
                opportunity = TodoItem.objects.get(id=opportunity_id)
                
                if action == 'delete':
                    opportunity_name = opportunity.oportunidad
                    opportunity.delete()
                    messages.success(request, f'Oportunidad "{opportunity_name}" eliminada exitosamente')
                elif action == 'reactivate':
                    # Cambiar el estado a uno activo (ej: Cotizando)
                    opportunity.bitrix_stage_id = 'UC_YUQKW6'  # Cotizando
                    opportunity.save()
                    messages.success(request, f'Oportunidad "{opportunity.oportunidad}" reactivada')
                    
            except TodoItem.DoesNotExist:
                messages.error(request, 'Oportunidad no encontrada')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
        
        return redirect('bitrix-lost-opportunities')
    
    # Obtener oportunidades perdidas usando detección precisa
    all_opportunities = TodoItem.objects.select_related('cliente', 'usuario').all()
    lost_ids = [item.id for item in all_opportunities if is_lost_opportunity(item.bitrix_stage_id)]
    lost_opportunities = TodoItem.objects.filter(id__in=lost_ids).select_related('cliente', 'usuario').order_by('-fecha_actualizacion')
    
    context = {
        'lost_opportunities': lost_opportunities,
        'total_lost': lost_opportunities.count(),
    }
    
    return render(request, 'bitrix_lost_opportunities.html', context)


@login_required
def api_export_excel(request):
    """
    Exporta la tabla activa a Excel con metadata
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from datetime import datetime
        
        # Obtener parámetros
        tab = request.GET.get('tab', 'crm')
        mes = request.GET.get('mes', '')
        anio = request.GET.get('anio', '')
        usuario = request.GET.get('usuario', request.user.get_full_name() or request.user.username)
        filtros = request.GET.get('filtros', 'Sin filtros')
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Reporte {tab.upper()}"
        
        # Metadata
        ws['A1'] = 'Reporte CRM'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'Descargado por: {usuario}'
        ws['A3'] = f'Fecha: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A4'] = f'Periodo: {mes}/{anio}'
        ws['A5'] = f'Filtros aplicados: {filtros}'
        
        # Espacio
        row_start = 7
        
        # Obtener datos según el tab
        if tab == 'crm':
            headers = ['Oportunidad', 'Cliente', 'Contacto', 'Área', 'Zebra', 'Panduit', 'APC', 'Avigilon', 'Genetec', 'Axis', 'Software', 'Runrate', 'Póliza', 'Otros', 'Total']
            data = get_oportunidades_data(request, mes, anio)
        elif tab == 'facturado':
            headers = ['Cliente', 'Zebra', 'Panduit', 'APC', 'Avigilon', 'Genetec', 'Axis', 'Software', 'Runrate', 'Póliza', 'Otros', 'Total', 'Meta', 'Fact.']
            data = get_facturado_data(request, mes, anio)
        else:
            headers = ['Datos']
            data = []
        
        # Escribir headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_start, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0052D4", end_color="0052D4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Escribir datos
        for row_num, row_data in enumerate(data, row_start + 1):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=reporte_crm_{tab}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response
        
    except ImportError:
        return JsonResponse({'error': 'openpyxl no está instalado. Ejecute: pip install openpyxl'}, status=500)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_solicitar_cambio_perfil(request):
    """
    Recibe los datos propuestos por el usuario. 
    Idioma y Tema se aplican de inmediato. 
    Si es Admin, todo se aplica de inmediato.
    Si es usuario normal y cambia Nombre/Email/Foto, se crea una solicitud de aprobación.
    """
    try:
        user = request.user
        first_name = request.POST.get('first_name', user.first_name)
        last_name = request.POST.get('last_name', user.last_name)
        email = request.POST.get('email', user.email)
        language = request.POST.get('language', user.userprofile.language)
        theme = request.POST.get('theme', user.userprofile.theme)
        avatar_file = request.FILES.get('avatar')

        # 1. Aplicar cambios inmediatos (Idioma y Tema)
        profile = user.userprofile
        profile.language = language
        profile.theme = theme
        profile.save()

        # 2. Si es superusuario, aplicar todo de inmediato
        if user.is_superuser:
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.save()
            if avatar_file:
                profile.avatar = avatar_file
                profile.save()
            return JsonResponse({'success': True, 'message': 'Información actualizada correctamente.', 'immediate': True})

        # 3. Si es usuario normal, verificar cambios sensibles
        cambio_sensible = (
            first_name != user.first_name or 
            last_name != user.last_name or 
            email != user.email or 
            avatar_file
        )

        if cambio_sensible:
            # Crear solicitud
            solicitud = SolicitudCambioPerfil.objects.create(
                solicitante=user,
                first_name=first_name,
                last_name=last_name,
                email=email,
                language=language or 'es',
                avatar=avatar_file,
                old_first_name=user.first_name,
                old_last_name=user.last_name,
                old_email=user.email,
                old_language=getattr(user.userprofile, 'language', 'es')
            )

            # Notificar a administradores (supervisores)
            admins = User.objects.filter(is_superuser=True)
            if not admins.exists():
                 admins = User.objects.all()[:1] 

            for admin in admins:
                crear_notificacion(
                    usuario_destinatario=admin,
                    tipo='solicitud_cambio_perfil',
                    titulo='Solicitud de Cambio de Perfil',
                    mensaje=f'El usuario {user.username} ha solicitado cambiar su información sensible. Requiere autorización.',
                    usuario_remitente=user,
                    solicitud_perfil=solicitud
                )

            return JsonResponse({
                'success': True, 
                'message': 'Preferencias guardadas. Los cambios en nombre, email o foto han sido enviados para aprobación administrativa.',
                'immediate': True # Recargar para aplicar idioma/tema si cambiaron
            })

        return JsonResponse({'success': True, 'message': 'Preferencias actualizadas correctamente.', 'immediate': True})

    except Exception as e:
        logger.error(f"Error en api_solicitar_cambio_perfil: {e}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@user_passes_test(is_supervisor)
@require_http_methods(["POST"])
def api_procesar_solicitud_perfil(request, solicitud_id):
    """
    Permite que un administrador apruebe o rechace una solicitud de cambio de perfil.
    """
    try:
        solicitud = get_object_or_404(SolicitudCambioPerfil, id=solicitud_id)
        if solicitud.procesada:
            return JsonResponse({'success': False, 'message': 'Esta solicitud ya ha sido procesada.'}, status=400)

        data = json.loads(request.body)
        aprobado = data.get('aprobado', False)
        comentario = data.get('comentario', '')

        solicitud.procesada = True
        solicitud.aprobada = aprobado
        solicitud.procesado_por = request.user
        solicitud.fecha_procesamiento = timezone.now()
        solicitud.comentario_admin = comentario
        solicitud.save()

        user = solicitud.solicitante
        if aprobado:
            # Aplicar cambios al usuario
            user.first_name = solicitud.first_name
            user.last_name = solicitud.last_name
            user.email = solicitud.email
            user.save()

            # Aplicar cambios al perfil
            profile = user.userprofile
            profile.language = solicitud.language
            if solicitud.avatar:
                profile.avatar = solicitud.avatar
            profile.save()

            # Notificar al usuario que su solicitud fue aprobada
            crear_notificacion(
                usuario_destinatario=user,
                tipo='sistema',
                titulo='Perfil Actualizado',
                mensaje=f'Tu solicitud de cambio de perfil ha sido aprobada por {request.user.username}.'
            )
        else:
            # Notificar al usuario que su solicitud fue rechazada
            crear_notificacion(
                usuario_destinatario=user,
                tipo='sistema',
                titulo='Solicitud de Perfil Rechazada',
                mensaje=f'Tu solicitud de cambio de perfil ha sido rechazada. Motivo: {comentario}'
            )

        return JsonResponse({'success': True, 'message': 'Solicitud procesada correctamente.'})
    except Exception as e:
        logger.error(f"Error en api_procesar_solicitud_perfil: {e}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
