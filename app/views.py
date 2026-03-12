import json
import logging
import requests
import mimetypes
import os
from django.conf import settings
import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.db import models
from .models import TodoItem, Cliente, Cotizacion, DetalleCotizacion, UserProfile, Contacto, PendingFileUpload, OportunidadProyecto, Volumetria, DetalleVolumetria, CatalogoCableado, OportunidadActividad, OportunidadComentario, OportunidadArchivo, OportunidadEstado, Notificacion, Proyecto, ProyectoComentario, ProyectoArchivo, Tarea, TareaComentario, TareaArchivo, Actividad, CarpetaProyecto, ArchivoProyecto, CompartirArchivo, IntercambioNavidad, ParticipanteIntercambio, HistorialIntercambio, SolicitudAccesoProyecto, ArchivoFacturacion, CarpetaOportunidad, ArchivoOportunidad, MensajeOportunidad, TareaOportunidad, ComentarioTareaOpp, PostMuro, ComentarioMuro, ProductoOportunidad, AsistenciaJornada, EficienciaMensual, SolicitudCambioPerfil, ProgramacionActividad
from . import views_exportar
from .views_tarea_comentarios import api_comentarios_tarea, api_agregar_comentario_tarea, api_editar_comentario_tarea, api_eliminar_comentario_tarea
from .forms import VentaForm, VentaFilterForm, CotizacionForm, ClienteForm, OportunidadModalForm, NuevaOportunidadForm
from django.db.models import Sum, Count, F, Q, Case, When, Value
from django.db.models.functions import Upper, Coalesce
from django.db.models import Value
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from django.utils import timezone
from decimal import Decimal
import decimal
from django.utils.html import json_script

# Helper function to detect lost opportunities
def is_lost_opportunity(stage_id):
    """
    Detecta si una oportunidad está marcada como "Cerrado Perdido" en Bitrix24
    Maneja múltiples formatos de stage_id:
    - '2': Formato simple
    - 'C2:11': Formato con prefijo numérico (perdido)
    - 'C2:LOSE': Formato "Cerrar negociación" -> "Cerrado perdido"
    - 'LOSE': Formato texto
    
    NOTA: C2:NEW no es perdido, es nuevo
    """
    if not stage_id:
        return False
    
    stage_id = str(stage_id)
    
    # Casos específicos de estado perdido
    if stage_id == '2':  # Formato simple
        return True
    elif stage_id == 'LOSE':  # Formato texto
        return True
    elif stage_id == 'C2:LOSE':  # Formato "Cerrar negociación" -> "Cerrado perdido"
        return True
    elif stage_id.startswith('C2:'):
        # Solo considerar perdido si después de C2: hay un número
        # C2:NEW = nuevo, C2:11 = perdido, C2:LOSE = perdido (ya manejado arriba)
        try:
            suffix = stage_id[3:]  # Obtener la parte después de 'C2:'
            int(suffix)  # Si es un número, es perdido
            return True
        except ValueError:
            # Si no es un número (ej: C2:NEW), no es perdido
            return False
    
    return False
import json
from zoneinfo import ZoneInfo


def convert_to_tijuana_time(utc_datetime):
    """Convierte una fecha UTC a tiempo de Tijuana"""
    try:
        # Usar zoneinfo que es parte de la biblioteca estándar de Python
        tijuana_tz = ZoneInfo('America/Tijuana')
        if utc_datetime.tzinfo is None:
            # Si no tiene timezone info, asumir UTC
            utc_datetime = timezone.make_aware(utc_datetime, timezone.utc)
        return utc_datetime.astimezone(tijuana_tz)
    except Exception:
        return utc_datetime
from django.urls import reverse
from django.utils import timezone, translation
from django.utils.translation import gettext_lazy as _
from datetime import datetime
from django.utils.translation import activate, get_language

# Definir la constante manualmente para compatibilidad
LANGUAGE_SESSION_KEY = 'django_language'
from django.http import JsonResponse

# Importaciones para generación de PDF
from weasyprint import HTML
from django.template.loader import render_to_string

# Asegúrate de importar tu modelo Cotizacion y DetalleCotizacion
from .models import Cotizacion, DetalleCotizacion

# Importaciones para el logo Base64
import base64
import os

logger = logging.getLogger(__name__)

# Función auxiliar para comprobar si el usuario es supervisor
def is_supervisor(user):
    return user.is_superuser or user.groups.filter(name='Supervisores').exists()


# Función auxiliar para comprobar si el usuario es ingeniero
def is_ingeniero(user):
    try:
        from app.models import UserProfile
        profile = UserProfile.objects.get(user=user)
        return getattr(profile, 'rol', 'vendedor') == 'ingeniero'
    except Exception:
        return False

def is_engineer(user):
    return user.groups.filter(name='Ingenieros').exists()

# Función auxiliar para obtener el display de un valor de choice
def _get_display_for_value(value, choices_list):
    return dict(choices_list).get(value, value)

# Vistas principales y funcionales

@login_required
def get_oportunidades_por_cliente(request):
    cliente_id = request.GET.get('cliente_id')
    oportunidad_inicial_id = request.GET.get('oportunidad_inicial_id')  # Nueva línea para oportunidad específica
    
    print(f"DEBUG: get_oportunidades_por_cliente - cliente_id: {cliente_id}, oportunidad_inicial_id: {oportunidad_inicial_id}")

    if is_supervisor(request.user):
        if cliente_id:
            # Solo las 10 oportunidades más recientes del cliente para supervisores
            oportunidades = TodoItem.objects.filter(cliente_id=cliente_id).order_by('-fecha_creacion')[:10]
        else:
            # If no client_id, return the 20 most recent opportunities for supervisors
            oportunidades = TodoItem.objects.all().order_by('-fecha_creacion')[:20]
    else:
        if cliente_id:
            # Solo las 10 oportunidades más recientes del cliente del usuario actual
            oportunidades = TodoItem.objects.filter(cliente_id=cliente_id, usuario=request.user).order_by('-fecha_creacion')[:10]
        else:
            # If no client_id, return the 20 most recent opportunities for the current user
            oportunidades = TodoItem.objects.filter(usuario=request.user).order_by('-fecha_creacion')[:20]

    # Si hay una oportunidad inicial específica, asegurar que esté incluida
    if oportunidad_inicial_id:
        try:
            # Limpiar el ID eliminando comas y espacios
            clean_id = oportunidad_inicial_id.replace(',', '').replace(' ', '').strip()
            print(f"DEBUG: Buscando oportunidad inicial con ID: {clean_id}")
            oportunidad_inicial = TodoItem.objects.get(id=int(clean_id))
            print(f"DEBUG: Oportunidad inicial encontrada: {oportunidad_inicial.oportunidad}")
            
            # Convertir queryset a lista para manipulación
            oportunidades_list = list(oportunidades)
            oportunidades_ids = [op.id for op in oportunidades_list]
            
            # Verificar si ya está en la lista por ID
            if oportunidad_inicial.id not in oportunidades_ids:
                print(f"DEBUG: Oportunidad inicial NO estaba en la lista, agregándola al principio")
                # Agregar la oportunidad específica al principio de la lista
                oportunidades_list.insert(0, oportunidad_inicial)
                oportunidades = oportunidades_list
            else:
                print(f"DEBUG: Oportunidad inicial YA estaba en la lista")
                # Moverla al principio si ya estaba presente
                oportunidades_list = [op for op in oportunidades_list if op.id != oportunidad_inicial.id]
                oportunidades_list.insert(0, oportunidad_inicial)
                oportunidades = oportunidades_list
        except (TodoItem.DoesNotExist, ValueError, TypeError) as e:
            print(f"DEBUG: Error procesando oportunidad inicial {oportunidad_inicial_id}: {e}")
            pass  # Si no existe o hay error de conversión, continuar con la lista normal

    data = [{'id': op.id, 'nombre': op.oportunidad} for op in oportunidades]

    return JsonResponse(data, safe=False)

@login_required
def get_bitrix_contacts_api(request):
    query = request.GET.get('query', '')
    company_id = request.GET.get('company_id', None)

    contacts = get_all_bitrix_contacts(request=request, company_id=company_id)

    # Filter contacts by query if provided
    if query:
        contacts = [c for c in contacts if query.lower() in (c.get('NAME', '') + ' ' + c.get('LAST_NAME', '')).lower()]

    data = []
    for contact in contacts:
        full_name = f"{contact.get('NAME', '')} {contact.get('LAST_NAME', '')}".strip()
        data.append({
            'ID': contact['ID'],
            'NAME': full_name,
            'COMPANY_ID': contact.get('COMPANY_ID'),
        })
    return JsonResponse(data, safe=False)

@login_required
def ventas_fullscreen(request):
    user = request.user
    is_super = hasattr(user, 'perfil') and getattr(user.perfil, 'es_supervisor', False)
    qs = TodoItem.objects.filter(probabilidad_cierre=100).exclude(mes_cierre__isnull=True).exclude(mes_cierre='')
    if not is_super:
        qs = qs.filter(usuario=user)
    ventas_por_mes = {str(i).zfill(2): 0 for i in range(1, 13)}
    for item in qs:
        mes = item.mes_cierre
        if mes in ventas_por_mes:
            ventas_por_mes[mes] += item.monto or 0
    meses = ['01','02','03','04','05','06','07','08','09','10','11','12']
    ventas_por_mes_list = [
        {'mes': m, 'monto': float(ventas_por_mes[m])} for m in meses
    ]
    return render(request, 'ventas_fullscreen.html', {'ventas_por_mes_list': ventas_por_mes_list})


@login_required
def oportunidades_mes_actual(request):
    hoy = date.today()
    mes_actual_val = str(hoy.month).zfill(2)
    mes_actual_nombre = dict(TodoItem.MES_CHOICES).get(mes_actual_val, f"Mes {hoy.month}")
    if is_supervisor(request.user):
        oportunidades = TodoItem.objects.filter(mes_cierre=mes_actual_val)
        meta_mensual = 400000
    else:
        oportunidades = TodoItem.objects.filter(mes_cierre=mes_actual_val, usuario=request.user)
        meta_mensual = 130000

    # Monto por cobrar: oportunidades con probabilidad entre 1 and 99%
    monto_por_cobrar = oportunidades.filter(probabilidad_cierre__gte=1, probabilidad_cierre__lte=99).aggregate(suma=Sum('monto'))['suma'] or 0

    return render(request, 'oportunidades_mes_actual.html', {
        'oportunidades': oportunidades,
        'mes_actual_nombre': mes_actual_nombre,
        'meta_mensual': meta_mensual,
        'monto_por_cobrar': monto_por_cobrar,
    })

@login_required
def crm_home(request):
    """
    Vista principal del CRM - tabla pivotada por cliente/producto.
    """
    from datetime import datetime
    user = request.user

    # Asegurar perfil
    profile, _ = UserProfile.objects.get_or_create(user=user)

    # Filtros mes/año — por defecto mes actual y año actual
    now = datetime.now()
    mes_filter = request.GET.get('mes', str(now.month).zfill(2))
    anio_filter = request.GET.get('anio', str(now.year))
    tab_activo = request.GET.get('tab', 'crm')

    # Mapeo de código mes a nombre en español (Bitrix guarda el nombre)
    MES_CODE_TO_NAME = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre',
    }
    MES_NAME_TO_CODE = {v: k for k, v in MES_CODE_TO_NAME.items()}

    try:
        anio_int = int(anio_filter)
    except ValueError:
        anio_int = now.year

    # MES_CHOICES para template (con opción "Todos" al inicio)
    mes_choices = [('todos', 'Todos')] + list(TodoItem.MES_CHOICES)
    mes_nombre = dict(mes_choices).get(mes_filter, '')
    mes_nombre_db = MES_CODE_TO_NAME.get(mes_filter, mes_filter)

    # ── Supervisor / Vendedor / Ingeniero logic ──
    es_supervisor = is_supervisor(user)
    es_ingeniero = (getattr(profile, 'rol', 'vendedor') == 'ingeniero')
    vendedores_filter = request.GET.get('vendedores', '')  # "1,2,3" or ""
    vendedores_ids = []
    if vendedores_filter:
        vendedores_ids = [int(x) for x in vendedores_filter.split(',') if x.strip().isdigit()]

    # Base queryset - oportunidades filtradas por anio_cierre/mes_cierre
    base_qs = TodoItem.objects.select_related('cliente', 'usuario', 'contacto', 'usuario__userprofile').filter(
        anio_cierre=anio_int
    )
    if mes_filter != 'todos':
        base_qs = base_qs.filter(mes_cierre=mes_filter)

    # Si NO es supervisor, filtrar solo sus oportunidades
    if not es_supervisor:
        base_qs = base_qs.filter(usuario=user)
    elif vendedores_ids:
        # Supervisor con filtro de vendedores específicos
        base_qs = base_qs.filter(usuario_id__in=vendedores_ids)

    # Lista de vendedores para el filtro (solo para supervisores)
    vendedores_list = []
    if es_supervisor:
        vendedores_list = User.objects.filter(
            is_active=True
        ).exclude(
            groups__name='Supervisores'
        ).order_by('first_name', 'last_name')

    # ── Meta (calculada antes para usar en running meta) ──
    # Determinar qué campo de meta usar según el tab activo
    meta_field = 'meta_mensual'  # Default (Facturado)
    if tab_activo == 'crm':
        meta_field = 'meta_oportunidades'
    elif tab_activo == 'cotizado':
        meta_field = 'meta_cotizado'
    elif tab_activo == 'cobrado':
        meta_field = 'meta_cobrado'

    if es_supervisor:
        if vendedores_ids:
            meta = UserProfile.objects.filter(user_id__in=vendedores_ids).aggregate(
                t=Coalesce(Sum(meta_field), Value(Decimal('0')))
            )['t'] or Decimal('0')
        else:
            # Suma de todos los vendedores activos para el supervisor
            all_sellers_profiles = UserProfile.objects.filter(user__is_active=True).exclude(user__groups__name='Supervisores')
            meta = all_sellers_profiles.aggregate(t=Coalesce(Sum(meta_field), Value(Decimal('0'))))['t'] or Decimal('0')
    else:
        meta = getattr(profile, meta_field, Decimal('0')) or Decimal('0')

    # Si se seleccionó "Todos" los meses, la meta es anual (mensual × 12)
    if mes_filter == 'todos':
        meta = meta * 12

    # ── Tab CRM: Lista de oportunidades individuales ──
    if tab_activo == 'crm':
        tabla_data = base_qs.select_related('cliente', 'contacto', 'usuario').order_by('-fecha_actualizacion')

    # ── Tab Facturado: Datos del XLS por cliente + desglose por producto ──
    elif tab_activo == 'facturado':
        # Obtener datos de facturación del XLS subido
        facturado_por_cliente = {}  # {cliente_name: monto}
        if mes_filter == 'todos':
            # Sumar todos los meses del año
            for af in ArchivoFacturacion.objects.filter(anio=anio_int):
                for cname, monto_str in (af.datos_json or {}).items():
                    facturado_por_cliente[cname] = str(
                        Decimal(facturado_por_cliente.get(cname, '0')) + Decimal(str(monto_str))
                    )
        else:
            try:
                archivo_fact = ArchivoFacturacion.objects.get(mes=mes_filter, anio=anio_int)
                facturado_por_cliente = archivo_fact.datos_json or {}
            except ArchivoFacturacion.DoesNotExist:
                pass

        # Mapear nombres del XLS a objetos Cliente
        facturado_por_cliente_obj = {}  # {cliente_id: monto}
        for cliente_name, monto_str in facturado_por_cliente.items():
            monto_val = Decimal(str(monto_str))
            cliente_match = Cliente.objects.filter(nombre_empresa__iexact=cliente_name).first()
            if not cliente_match:
                cliente_match = Cliente.objects.filter(nombre_empresa__icontains=cliente_name).first()
            if not cliente_match:
                for cliente_obj in Cliente.objects.all():
                    if cliente_obj.nombre_empresa and cliente_obj.nombre_empresa.upper() in cliente_name.upper():
                        cliente_match = cliente_obj
                        break
            if cliente_match:
                facturado_por_cliente_obj[cliente_match.id] = (
                    facturado_por_cliente_obj.get(cliente_match.id, Decimal('0')) + monto_val
                )

        # Desglose por producto (de monto_facturacion en oportunidades)
        prod_data = base_qs.filter(monto_facturacion__gt=0).values('cliente').annotate(
            zebra=Coalesce(Sum('monto_facturacion', filter=Q(producto='ZEBRA')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            panduit=Coalesce(Sum('monto_facturacion', filter=Q(producto='PANDUIT')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            apc=Coalesce(Sum('monto_facturacion', filter=Q(producto='APC')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            avigilon=Coalesce(Sum('monto_facturacion', filter=Q(producto='AVIGILON') | Q(producto='AVIGILION')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            genetec=Coalesce(Sum('monto_facturacion', filter=Q(producto='GENETEC')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            axis=Coalesce(Sum('monto_facturacion', filter=Q(producto='AXIS')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            software=Coalesce(Sum('monto_facturacion', filter=Q(producto='SOFTWARE') | Q(producto='Desarrollo')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            runrate=Coalesce(Sum('monto_facturacion', filter=Q(producto='RUNRATE')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            poliza=Coalesce(Sum('monto_facturacion', filter=Q(producto='PÓLIZA') | Q(producto='POLIZA')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            total_prod=Coalesce(Sum('monto_facturacion'), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
        )
        prod_dict = {item['cliente']: item for item in prod_data}

        # Filtrar clientes según permisos
        if es_supervisor:
            if vendedores_ids:
                clientes_qs = Cliente.objects.filter(asignado_a_id__in=vendedores_ids).order_by('nombre_empresa')
            else:
                clientes_qs = Cliente.objects.all().order_by('nombre_empresa')
        else:
            clientes_qs = Cliente.objects.filter(asignado_a=user).order_by('nombre_empresa')

        raw_data = []
        for c in clientes_qs:
            pdat = prod_dict.get(c.id, {})
            fact_monto = facturado_por_cliente_obj.get(c.id, Decimal('0'))
            zb = pdat.get('zebra', Decimal('0'))
            pa = pdat.get('panduit', Decimal('0'))
            ap = pdat.get('apc', Decimal('0'))
            av = pdat.get('avigilon', Decimal('0'))
            ge = pdat.get('genetec', Decimal('0'))
            ax = pdat.get('axis', Decimal('0'))
            so = pdat.get('software', Decimal('0'))
            rr = pdat.get('runrate', Decimal('0'))
            po = pdat.get('poliza', Decimal('0'))
            tp = pdat.get('total_prod', Decimal('0'))
            otros = tp - (zb + pa + ap + av + ge + ax + so + rr + po)
            raw_data.append({
                'cliente': c,
                'zebra': zb, 'panduit': pa, 'apc': ap, 'avigilon': av,
                'genetec': ge, 'axis': ax, 'software': so, 'runrate': rr,
                'poliza': po, 'otros': otros,
                'facturado': fact_monto,
            })

        # Ordenar por facturado descendente
        raw_data.sort(key=lambda x: x['facturado'], reverse=True)

        # Meta por cliente: meta individual del cliente - su facturado
        for item in raw_data:
            cliente_meta = item['cliente'].meta_mensual or Decimal('0')
            item['meta_cliente'] = cliente_meta
            item['meta_restante'] = cliente_meta - item['facturado']
            item['total'] = item['facturado'] # Para que el template use item.total
        tabla_data = raw_data

    # ── Tab Cotizado: Cotizaciones PDF generadas ──
    elif tab_activo == 'cotizado':
        opp_ids = base_qs.values_list('id', flat=True)
        cotizaciones_qs = (
            Cotizacion.objects
            .select_related('oportunidad', 'created_by', 'cliente')
            .filter(
                Q(oportunidad_id__in=opp_ids) |
                Q(oportunidad__isnull=True, fecha_creacion__year=anio_int)
            )
            .order_by('-fecha_creacion')
        )
        tabla_data = cotizaciones_qs

    # ── Tab Cobrado: Oportunidades con 100% probabilidad ──
    elif tab_activo == 'cobrado':
        tabla_data = base_qs.filter(probabilidad_cierre=100).order_by('-monto')

    else:
        tabla_data = []

    # Stats generales
    total_general = base_qs.aggregate(t=Coalesce(Sum('monto'), Value(Decimal('0'))))['t']
    num_clientes = Cliente.objects.count() if es_supervisor else base_qs.values('cliente').distinct().count()
    num_deals = base_qs.count()
    num_cobradas = base_qs.filter(probabilidad_cierre=100).count()

    # ── Total facturado desde XLS ──
    total_facturado = Decimal('0')
    try:
        if mes_filter == 'todos':
            archivos_fact = ArchivoFacturacion.objects.filter(anio=anio_int)
            for af in archivos_fact:
                if es_supervisor and not vendedores_ids:
                    total_facturado += af.total_facturado
                elif es_supervisor and vendedores_ids:
                    fm = calcular_facturado_por_vendedor(af.datos_json)
                    total_facturado += sum(Decimal(str(v)) for uid, v in fm.items() if uid in vendedores_ids)
                else:
                    fm = calcular_facturado_por_vendedor(af.datos_json)
                    total_facturado += Decimal(str(fm.get(user.id, 0)))
        else:
            archivo_fact = ArchivoFacturacion.objects.get(mes=mes_filter, anio=anio_int)
            if es_supervisor and not vendedores_ids:
                total_facturado = archivo_fact.total_facturado
            elif es_supervisor and vendedores_ids:
                facturado_map = calcular_facturado_por_vendedor(archivo_fact.datos_json)
                total_facturado = sum(Decimal(str(v)) for uid, v in facturado_map.items() if uid in vendedores_ids)
            else:
                facturado_map = calcular_facturado_por_vendedor(archivo_fact.datos_json)
                total_facturado = Decimal(str(facturado_map.get(user.id, 0)))
    except ArchivoFacturacion.DoesNotExist:
        total_facturado = Decimal('0')

    progreso = min(int((total_facturado / meta * 100)) if meta > 0 else 0, 100)

    # Stats para tab Cobrado
    total_cobrado = Decimal('0')
    if tab_activo == 'cobrado':
        total_cobrado = base_qs.filter(probabilidad_cierre=100).aggregate(t=Coalesce(Sum('monto'), Value(Decimal('0'))))['t']

    # Stats para tab Cotizado
    num_cotizaciones = 0
    num_oportunidades_cotizadas = 0
    total_cotizado = Decimal('0')
    if tab_activo == 'cotizado':
        num_cotizaciones = cotizaciones_qs.count()
        num_oportunidades_cotizadas = cotizaciones_qs.exclude(oportunidad__isnull=True).values('oportunidad').distinct().count()
        total_cotizado = cotizaciones_qs.aggregate(t=Coalesce(Sum('total'), Value(Decimal('0'))))['t']

    # ── Widget Logic ──
    widget_label = 'Total Facturado'
    widget_metric = total_facturado 

    if tab_activo == 'crm':
        widget_label = 'Total Oportunidades'
        widget_metric = total_general
    elif tab_activo == 'cotizado':
        widget_label = 'Total Cotizado'
        widget_metric = total_cotizado
    elif tab_activo == 'cobrado':
        widget_label = 'Total Cobrado'
        widget_metric = total_cobrado

    # Recalculate progress based on the correct metric vs correct meta (sin cap para mostrar > 100%)
    progreso = int((widget_metric / meta * 100)) if meta > 0 else 0

    context = {
        'widget_label': widget_label,
        'widget_metric': widget_metric,
        'tab_activo': tab_activo,
        'tabla_data': tabla_data,
        'mes_filter': mes_filter,
        'anio_filter': anio_filter,
        'anio_int': anio_int,
        'mes_nombre': mes_nombre,
        'mes_choices': mes_choices,
        'total_general': total_general,
        'total_facturado': total_facturado,
        'num_clientes': num_clientes,
        'num_deals': num_deals,
        'num_cobradas': num_cobradas,
        'meta': meta,
        'progreso': progreso,
        'progreso_visual': min(progreso, 100),
        'usuario': user,
        'years_range': range(2024, now.year + 2),
        'num_cotizaciones': num_cotizaciones,
        'num_oportunidades_cotizadas': num_oportunidades_cotizadas,
        'total_cotizado': total_cotizado,
        'total_cobrado': total_cobrado,
        'es_supervisor': es_supervisor,
        'es_ingeniero': es_ingeniero,
        'vendedores_list': vendedores_list,
        'vendedores_filter': vendedores_filter,
    }
    return render(request, 'crm_home.html', context)


@login_required
def api_crm_table_data(request):
    """
    API endpoint que devuelve los datos de la tabla CRM en JSON
    para actualizar sin recargar la página.
    """
    from datetime import datetime
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)

    now = datetime.now()
    mes_filter = request.GET.get('mes', str(now.month).zfill(2))
    anio_filter = request.GET.get('anio', str(now.year))
    tab_activo = request.GET.get('tab', 'crm')
    desde_filter = request.GET.get('desde', '').strip()
    hasta_filter = request.GET.get('hasta', '').strip()
    usando_periodo = bool(desde_filter and hasta_filter)

    MES_CODE_TO_NAME = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre',
    }

    try:
        anio_int = int(anio_filter)
    except ValueError:
        anio_int = now.year

    mes_nombre_db = MES_CODE_TO_NAME.get(mes_filter, mes_filter)

    # Supervisor / Vendedor logic
    es_supervisor = is_supervisor(user)
    vendedores_filter = request.GET.get('vendedores', '')
    vendedores_ids = [int(x) for x in vendedores_filter.split(',') if x.strip().isdigit()] if vendedores_filter else []

    q_search = request.GET.get('q', '').strip()

    if q_search:
        # Búsqueda global: ignora mes/año, busca en nombre de oportunidad y cliente
        base_qs = TodoItem.objects.select_related('cliente', 'usuario', 'contacto', 'usuario__userprofile').filter(
            Q(oportunidad__icontains=q_search) | Q(cliente__nombre__icontains=q_search)
        )
    elif usando_periodo:
        base_qs = TodoItem.objects.select_related('cliente', 'usuario', 'contacto', 'usuario__userprofile').filter(
            fecha_creacion__date__gte=desde_filter,
            fecha_creacion__date__lte=hasta_filter,
        )
    else:
        base_qs = TodoItem.objects.select_related('cliente', 'usuario', 'contacto', 'usuario__userprofile').filter(
            anio_cierre=anio_int
        )
        if mes_filter != 'todos':
            base_qs = base_qs.filter(mes_cierre=mes_filter)

    if not es_supervisor:
        base_qs = base_qs.filter(usuario=user)
    elif vendedores_ids:
        base_qs = base_qs.filter(usuario_id__in=vendedores_ids)

    def format_money(val):
        if val is None:
            return '0'
        try:
            return '{:,.0f}'.format(val)
        except (ValueError, TypeError):
            return '0'

    # ── Calcular meta según tab (aplica a todos los tabs) ──
    meta_field_api = 'meta_mensual'
    if tab_activo == 'crm':
        meta_field_api = 'meta_oportunidades'
    elif tab_activo == 'cotizado':
        meta_field_api = 'meta_cotizado'
    elif tab_activo == 'cobrado':
        meta_field_api = 'meta_cobrado'

    if es_supervisor:
        if vendedores_ids:
            api_meta = UserProfile.objects.filter(user_id__in=vendedores_ids).aggregate(
                t=Coalesce(Sum(meta_field_api), Value(Decimal('0')))
            )['t'] or Decimal('0')
        else:
            all_sellers_profiles = UserProfile.objects.filter(user__is_active=True).exclude(user__groups__name='Supervisores')
            api_meta = all_sellers_profiles.aggregate(t=Coalesce(Sum(meta_field_api), Value(Decimal('0'))))['t'] or Decimal('0')
    else:
        api_meta = getattr(profile, meta_field_api, Decimal('0')) or Decimal('0')

    if mes_filter == 'todos' and not usando_periodo:
        api_meta = api_meta * 12

    if tab_activo == 'crm':
        from django.utils import timezone as _tz
        _now = _tz.now()
        items = base_qs.select_related('cliente', 'contacto', 'usuario').order_by('-fecha_actualizacion')
        rows = []
        for item in items:
            # Revisar si tiene actividades vencidas (fecha_limite pasada y no completada)
            tareas = item.tareas_oportunidad.all()
            tiene_vencida = False
            tiene_pendiente = False
            for t in tareas:
                if t.estado != 'completada':
                    tiene_pendiente = True
                    if t.fecha_limite and t.fecha_limite < _now:
                        tiene_vencida = True
            es_bitrix = item.tipo_negociacion == 'bitrix_proyecto'
            rows.append({
                'id': item.id,
                'oportunidad': (item.oportunidad or '')[:35],
                'cliente': (item.cliente.nombre_empresa if item.cliente else '- Sin Cliente -')[:35],
                'cliente_id': item.cliente.id if item.cliente else None,
                'contacto': (item.contacto.nombre[:18] if item.contacto else '-'),
                'area': item.area or '-',
                'producto': item.producto or '',
                'monto': '0' if es_bitrix else format_money(item.monto),
                'fecha_iso': item.fecha_creacion.strftime('%Y-%m-%d'),
                'fecha_ts': int(item.fecha_actualizacion.timestamp()) if item.fecha_actualizacion else 0,
                'etapa': item.etapa_corta or '',
                'tiene_actividad_vencida': tiene_vencida,
                'sin_actividad_pendiente': not tiene_pendiente,
                'tipo_negociacion': item.tipo_negociacion or 'runrate',
            })
        # Ordenar: primero las que tienen actividad vencida
        rows.sort(key=lambda x: (not x['tiene_actividad_vencida'], x.get('fecha_iso', '')), reverse=False)
        # Stats
        total_general = base_qs.aggregate(t=Coalesce(Sum('monto'), Value(Decimal('0'))))['t']
        num_clientes = base_qs.values('cliente').distinct().count()
        num_deals = base_qs.count()

        api_progreso = int((total_general / api_meta * 100)) if api_meta > 0 else 0

        return JsonResponse({
            'tab': 'crm',
            'rows': rows,
            'footer': {
                'left': f'{num_clientes} clientes / {num_deals} Deals',
                'right': f'Total: ${format_money(total_general)}',
            },
            'total_facturado': format_money(total_general),
            'widget_label': 'Total Oportunidades',
            'meta': format_money(api_meta),
            'progreso': api_progreso,
            'widget_left_stat': f'{num_deals} Oportunidades Creadas',
        })

    elif tab_activo == 'cotizado':
        opp_ids = base_qs.values_list('id', flat=True)
        if usando_periodo:
            cotizaciones_qs = (
                Cotizacion.objects
                .select_related('oportunidad', 'created_by', 'cliente')
                .filter(
                    Q(oportunidad_id__in=opp_ids) |
                    Q(oportunidad__isnull=True,
                      fecha_creacion__date__gte=desde_filter,
                      fecha_creacion__date__lte=hasta_filter)
                )
                .order_by('-fecha_creacion')
            )
        else:
            cotizaciones_qs = (
                Cotizacion.objects
                .select_related('oportunidad', 'created_by', 'cliente')
                .filter(
                    Q(oportunidad_id__in=opp_ids) |
                    Q(oportunidad__isnull=True, fecha_creacion__year=anio_int)
                )
                .order_by('-fecha_creacion')
            )
        rows = []
        for cot in cotizaciones_qs:
            rows.append({
                'id': cot.id,
                'oportunidad': (cot.oportunidad.oportunidad if cot.oportunidad else '—')[:35],
                'oportunidad_id': cot.oportunidad.id if cot.oportunidad else None,
                'cliente': (cot.cliente.nombre_empresa if cot.cliente else '- Sin Cliente -')[:35],
                'cliente_id': cot.cliente.id if cot.cliente else None,
                'usuario': (cot.created_by.get_full_name() or cot.created_by.username) if cot.created_by else '—',
                'subtotal': format_money(cot.subtotal),
                'total': format_money(cot.total),
                'pdf_url': f'/app/cotizacion/view/{cot.id}/',
                'fecha_iso': cot.fecha_creacion.strftime('%Y-%m-%d'),
            })
        num_cotizaciones = cotizaciones_qs.count()
        num_oportunidades_cotizadas = cotizaciones_qs.exclude(oportunidad__isnull=True).values('oportunidad').distinct().count()
        total_cotizado = cotizaciones_qs.aggregate(t=Coalesce(Sum('total'), Value(Decimal('0'))))['t']
        api_progreso_cot = int((total_cotizado / api_meta * 100)) if api_meta > 0 else 0
        return JsonResponse({
            'tab': 'cotizado',
            'rows': rows,
            'footer': {
                'left': f'{num_oportunidades_cotizadas} oportunidades / {num_cotizaciones} cotizaciones',
                'right': f'Total cotizado: ${format_money(total_cotizado)}',
            },
            'total_facturado': format_money(total_cotizado),
            'widget_label': 'Total Cotizado',
            'meta': format_money(api_meta),
            'progreso': api_progreso_cot,
            'widget_left_stat': f'{num_cotizaciones} Cotizaciones Creadas',
        })

    elif tab_activo == 'cobrado':
        items = base_qs.filter(probabilidad_cierre=100).order_by('-monto')
        rows = []
        for op in items:
            rows.append({
                'id': op.id,
                'oportunidad': (op.oportunidad or '')[:35],
                'cliente': (op.cliente.nombre_empresa if op.cliente else '- Sin Cliente -')[:35],
                'cliente_id': op.cliente.id if op.cliente else None,
                'producto_display': op.get_producto_display(),
                'usuario': (op.usuario.get_full_name() or op.usuario.username) if op.usuario else '—',
                'fecha': op.fecha_creacion.strftime('%d %b %Y'),
                'fecha_iso': op.fecha_creacion.strftime('%Y-%m-%d'),
                'monto': format_money(op.monto),
            })
        total_cobrado = items.aggregate(t=Coalesce(Sum('monto'), Value(Decimal('0'))))['t']
        num_deals = items.count()
        api_progreso_cob = int((total_cobrado / api_meta * 100)) if api_meta > 0 else 0
        return JsonResponse({
            'tab': 'cobrado',
            'rows': rows,
            'footer': {
                'left': f'{num_deals} Deals Cobrados',
                'right': f'Total cobrado: ${format_money(total_cobrado)}',
            },
            'total_facturado': format_money(total_cobrado),
            'widget_label': 'Total Cobrado',
            'meta': format_money(api_meta),
            'progreso': api_progreso_cob,
            'widget_left_stat': f'{num_deals} Oportunidades Cobradas',
        })

    elif tab_activo == 'facturado':
        # Reutilizamos la lógica del view principal para facturación
        facturado_por_cliente_obj = {}
        try:
            if usando_periodo:
                from datetime import date as _date
                cur = _date.fromisoformat(desde_filter).replace(day=1)
                end = _date.fromisoformat(hasta_filter).replace(day=1)
                while cur <= end:
                    mes_code = cur.strftime('%m')
                    try:
                        af = ArchivoFacturacion.objects.get(mes=mes_code, anio=cur.year)
                        data_af = af.datos_json.get('datos', {})
                        for c_name, val in data_af.items():
                            facturado_por_cliente_obj[c_name] = facturado_por_cliente_obj.get(c_name, Decimal('0')) + Decimal(str(val))
                    except ArchivoFacturacion.DoesNotExist:
                        pass
                    cur = (cur.replace(month=cur.month % 12 + 1, day=1) if cur.month < 12
                           else cur.replace(year=cur.year + 1, month=1, day=1))
            elif mes_filter == 'todos':
                for af in ArchivoFacturacion.objects.filter(anio=anio_int):
                    data_af = af.datos_json.get('datos', {})
                    for c_name, val in data_af.items():
                        facturado_por_cliente_obj[c_name] = facturado_por_cliente_obj.get(c_name, Decimal('0')) + Decimal(str(val))
            else:
                af = ArchivoFacturacion.objects.get(mes=mes_filter, anio=anio_int)
                data_af = af.datos_json.get('datos', {})
                for c_name, val in data_af.items():
                    facturado_por_cliente_obj[c_name] = Decimal(str(val))
        except ArchivoFacturacion.DoesNotExist:
            pass

        # Mapear nombres a IDs de clientes
        fact_by_id = {}
        for name, monto in facturado_por_cliente_obj.items():
            cliente = Cliente.objects.filter(nombre_empresa__icontains=name).first()
            if cliente:
                fact_by_id[cliente.id] = fact_by_id.get(cliente.id, Decimal('0')) + monto

        # Desglose prod
        prod_data = base_qs.filter(monto_facturacion__gt=0).values('cliente').annotate(
            zebra=Coalesce(Sum('monto_facturacion', filter=Q(producto='ZEBRA')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            panduit=Coalesce(Sum('monto_facturacion', filter=Q(producto='PANDUIT')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            apc=Coalesce(Sum('monto_facturacion', filter=Q(producto='APC')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            avigilon=Coalesce(Sum('monto_facturacion', filter=Q(producto='AVIGILON') | Q(producto='AVIGILION')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            genetec=Coalesce(Sum('monto_facturacion', filter=Q(producto='GENETEC')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            axis=Coalesce(Sum('monto_facturacion', filter=Q(producto='AXIS')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            software=Coalesce(Sum('monto_facturacion', filter=Q(producto='SOFTWARE') | Q(producto='Desarrollo')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            runrate=Coalesce(Sum('monto_facturacion', filter=Q(producto='RUNRATE')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            poliza=Coalesce(Sum('monto_facturacion', filter=Q(producto='PÓLIZA') | Q(producto='POLIZA')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            total_prod=Coalesce(Sum('monto_facturacion'), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
        )
        prod_dict = {item['cliente']: item for item in prod_data}

        if es_supervisor:
            clientes_qs = Cliente.objects.filter(asignado_a_id__in=vendedores_ids) if vendedores_ids else Cliente.objects.all()
        else:
            clientes_qs = Cliente.objects.filter(asignado_a=user)

        rows = []
        total_facturado_acum = Decimal('0')
        for c in clientes_qs.order_by('nombre_empresa'):
            p = prod_dict.get(c.id, {})
            fact = fact_by_id.get(c.id, Decimal('0'))
            meta_c = c.meta_mensual or Decimal('0')
            rows.append({
                'cliente_id': c.id,
                'cliente': c.nombre_empresa[:35],
                'zebra': format_money(p.get('zebra')),
                'panduit': format_money(p.get('panduit')),
                'apc': format_money(p.get('apc')),
                'avigilon': format_money(p.get('avigilon')),
                'genetec': format_money(p.get('genetec')),
                'axis': format_money(p.get('axis')),
                'software': format_money(p.get('software')),
                'runrate': format_money(p.get('runrate')),
                'poliza': format_money(p.get('poliza')),
                'otros': format_money(p.get('total_prod', Decimal('0')) - sum(p.get(k, 0) for k in ['zebra','panduit','apc','avigilon','genetec','axis','software','runrate','poliza'] if k in p)),
                'total': format_money(fact),
                'meta_cliente': format_money(meta_c),
                'meta_restante': format_money(meta_c - fact),
            })
            total_facturado_acum += fact

        api_progreso_fact = int((total_facturado_acum / api_meta * 100)) if api_meta > 0 else 0
        num_clientes_fact = clientes_qs.count()
        return JsonResponse({
            'tab': 'facturado',
            'rows': rows,
            'footer': {
                'left': f'{num_clientes_fact} clientes',
                'right': f'Total facturado: ${format_money(total_facturado_acum)}',
            },
            'total_facturado': format_money(total_facturado_acum),
            'widget_label': 'Total Facturado',
            'meta': format_money(api_meta),
            'progreso': api_progreso_fact,
            'widget_left_stat': f'{num_clientes_fact} Clientes',
        })

    return JsonResponse({'tab': tab_activo, 'rows': [], 'footer': {'left': '', 'right': ''}})


@login_required
@require_http_methods(["POST"])
def api_subir_facturacion(request):
    """
    API para subir archivo XLS de facturación.
    Solo supervisores pueden subir.
    Parsea el XLS y extrae total de pagos por cliente.
    """
    if not is_supervisor(request.user):
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)

    from datetime import datetime as dt_now
    archivo = request.FILES.get('archivo')
    mes = request.POST.get('mes', '')
    anio = request.POST.get('anio', '')

    if not archivo:
        return JsonResponse({'success': False, 'error': 'Falta el archivo'})

    # Defaults si no vienen mes/anio
    now = dt_now.now()
    if not mes:
        mes = str(now.month).zfill(2)

    try:
        anio_int = int(anio) if anio else now.year
    except (ValueError, TypeError):
        anio_int = now.year

    try:
        import xlrd
        from xlrd import xldate_as_datetime
        content = archivo.read()
        wb = xlrd.open_workbook(file_contents=content)
        sheet = wb.sheet_by_index(0)

        # Agrupar por mes de emisión: {(mes, anio): {cliente: monto}}
        datos_por_mes = {}  # { 'MM': { 'YYYY': { cliente_name: monto_str } } }
        totales_por_mes = {}  # { (mes, anio): Decimal }

        for row_idx in range(1, sheet.nrows):
            try:
                estatus = str(sheet.cell_value(row_idx, 40)).strip().lower()
                if estatus == 'cancelada':
                    continue

                cliente_name = str(sheet.cell_value(row_idx, 5)).strip()
                nombre_comercial = str(sheet.cell_value(row_idx, 6)).strip()
                if not cliente_name:
                    continue

                # Extraer mes/año de la fecha de emisión (col D, idx 3)
                try:
                    date_val = sheet.cell_value(row_idx, 3)
                    fecha = xldate_as_datetime(date_val, wb.datemode)
                    row_mes = str(fecha.month).zfill(2)
                    row_anio = fecha.year
                except Exception:
                    continue  # Sin fecha válida, saltar

                # Facturado = (Col L Subtotal idx 11 - Col O Descuento idx 14) * Col AQ T.C. idx 42
                subtotal_str = str(sheet.cell_value(row_idx, 11)).replace(',', '').strip()
                descuento_str = str(sheet.cell_value(row_idx, 14)).replace(',', '').strip()
                tc_str = str(sheet.cell_value(row_idx, 42)).replace(',', '').strip()
                try:
                    subtotal = Decimal(subtotal_str) if subtotal_str else Decimal('0')
                except Exception:
                    subtotal = Decimal('0')
                try:
                    descuento = Decimal(descuento_str) if descuento_str else Decimal('0')
                except Exception:
                    descuento = Decimal('0')
                try:
                    tc = Decimal(tc_str) if tc_str else Decimal('1')
                except Exception:
                    tc = Decimal('1')
                monto = (subtotal - descuento) * tc

                if monto > 0:
                    key = (row_mes, row_anio)
                    if key not in datos_por_mes:
                        datos_por_mes[key] = {}
                        totales_por_mes[key] = Decimal('0')

                    clientes_mes = datos_por_mes[key]
                    clientes_mes[cliente_name] = str(
                        Decimal(clientes_mes.get(cliente_name, '0')) + monto
                    )
                    if nombre_comercial and nombre_comercial != cliente_name:
                        clientes_mes[nombre_comercial] = str(
                            Decimal(clientes_mes.get(nombre_comercial, '0')) + monto
                        )
                    totales_por_mes[key] += monto
            except (IndexError, ValueError):
                continue

        # Guardar un ArchivoFacturacion por cada mes encontrado
        archivo.seek(0)
        meses_guardados = []
        for (m, a), clientes_data in datos_por_mes.items():
            obj, created = ArchivoFacturacion.objects.update_or_create(
                mes=m, anio=a,
                defaults={
                    'archivo': archivo,
                    'total_facturado': totales_por_mes[(m, a)],
                    'datos_json': clientes_data,
                    'subido_por': request.user,
                }
            )
            meses_guardados.append(f"{m}/{a}")

        total_general = sum(totales_por_mes.values())
        return JsonResponse({
            'success': True,
            'total_facturado': str(total_general),
            'num_clientes': sum(len(v) for v in datos_por_mes.values()),
            'meses': meses_guardados,
            'created': True,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error procesando archivo: {str(e)}'})


def calcular_facturado_por_vendedor(datos_json):
    """
    Dado el datos_json de ArchivoFacturacion {cliente_name: monto},
    retorna {user_id: monto_facturado} mapeando clientes a sus vendedores.
    """
    from collections import defaultdict
    facturado_por_vendedor = defaultdict(Decimal)

    for cliente_name, monto_str in datos_json.items():
        monto = Decimal(str(monto_str))
        # Buscar cliente en BD por nombre
        cliente_match = Cliente.objects.filter(nombre_empresa__iexact=cliente_name).first()
        if not cliente_match:
            cliente_match = Cliente.objects.filter(nombre_empresa__icontains=cliente_name).first()
        if not cliente_match:
            # Intentar match parcial inverso
            for cliente_obj in Cliente.objects.all():
                if cliente_obj.nombre_empresa and cliente_obj.nombre_empresa.upper() in cliente_name.upper():
                    cliente_match = cliente_obj
                    break

        if cliente_match:
            # El dueño es el usuario con más oportunidades de ese cliente
            owner = (
                TodoItem.objects
                .filter(cliente=cliente_match)
                .values('usuario_id')
                .annotate(count=Count('id'))
                .order_by('-count')
                .first()
            )
            if owner and owner['usuario_id']:
                facturado_por_vendedor[owner['usuario_id']] += monto

    return dict(facturado_por_vendedor)


@login_required
def api_cliente_oportunidades(request, cliente_id):
    """
    API que devuelve las oportunidades de un cliente específico en JSON.
    Formato compatible con buildCrmRow() del frontend.
    """
    from datetime import datetime
    user = request.user
    es_supervisor = is_supervisor(user)

    # Por defecto mostrar todo el historial del cliente, no solo el mes actual
    mes_filter = request.GET.get('mes', 'todos')
    anio_filter = request.GET.get('anio', 'todos')

    MES_CODE_TO_NAME = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre',
    }

    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cliente no encontrado'}, status=404)

    qs = TodoItem.objects.select_related('cliente', 'contacto', 'usuario').filter(cliente=cliente)
    
    # Aplicar filtros solo si no son 'todos'
    if anio_filter != 'todos':
        try:
            anio_int = int(anio_filter)
            qs = qs.filter(anio_cierre=anio_int)
        except (ValueError, TypeError):
            pass

    if mes_filter != 'todos':
        qs = qs.filter(mes_cierre=mes_filter)

    if not es_supervisor:
        qs = qs.filter(usuario=user)

    # Filtrar por tipo si se especifica
    tipo = request.GET.get('tipo', '')
    if tipo == 'cobrado':
        qs = qs.filter(probabilidad_cierre=100)

    qs = qs.order_by('-fecha_actualizacion')

    def format_money(val):
        if val is None:
            return '0'
        try:
            return '{:,.0f}'.format(val)
        except (ValueError, TypeError):
            return '0'

    rows = []
    for item in qs:
        rows.append({
            'id': item.id,
            'oportunidad': (item.oportunidad or '')[:35],
            'cliente': (item.cliente.nombre_empresa if item.cliente else '- Sin Cliente -')[:35],
            'cliente_id': item.cliente_id,
            'contacto': {'nombre': (item.contacto.nombre[:18] if item.contacto else '-')},
            'area': item.area or '-',
            'producto': item.producto or '',
            'producto_display': item.get_producto_display() if hasattr(item, 'get_producto_display') else (item.producto or ''),
            'usuario': item.usuario.get_full_name() or item.usuario.username if item.usuario else '—',
            'fecha': item.fecha_actualizacion.strftime('%d/%m/%y') if item.fecha_actualizacion else '—',
            'fecha_iso': item.fecha_actualizacion.isoformat() if item.fecha_actualizacion else '',
            'monto': format_money(item.monto),
            'monto_raw': float(item.monto or 0),
            'probabilidad_cierre': item.probabilidad_cierre,
        })

    # All contacts for this client (for filter dropdown)
    contactos_cliente = list(
        Contacto.objects.filter(cliente=cliente)
        .values_list('nombre', flat=True)
        .order_by('nombre')
    )

    return JsonResponse({
        'success': True,
        'cliente_nombre': cliente.nombre_empresa,
        'rows': rows,
        'contactos': contactos_cliente,
    })


@login_required
def api_cliente_cotizaciones(request, cliente_id):
    """API que devuelve las cotizaciones de un cliente específico."""
    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cliente no encontrado'}, status=404)

    qs = Cotizacion.objects.select_related('oportunidad', 'created_by').filter(cliente=cliente).order_by('-fecha_creacion')

    def fmt(val):
        try:
            return '{:,.0f}'.format(val or 0)
        except (ValueError, TypeError):
            return '0'

    rows = []
    for c in qs:
        rows.append({
            'id': c.id,
            'titulo': c.titulo or 'Cotización',
            'oportunidad': c.oportunidad.oportunidad[:35] if c.oportunidad else '—',
            'oportunidad_id': c.oportunidad_id,
            'usuario': c.created_by.get_full_name() or c.created_by.username if c.created_by else '—',
            'fecha': c.fecha_creacion.strftime('%d/%m/%y') if c.fecha_creacion else '—',
            'subtotal': fmt(c.subtotal),
            'total': fmt(c.total),
            'total_raw': float(c.total or 0),
            'moneda': c.moneda or 'MXN',
            'pdf_url': f'/app/cotizacion/view/{c.id}/',
        })

    return JsonResponse({
        'success': True,
        'cliente_nombre': cliente.nombre_empresa,
        'rows': rows,
    })


# ═══════════════════════════════════════════════════
# ══  ADMIN PANEL APIs (solo supervisores)  ═══════
# ═══════════════════════════════════════════════════

@login_required
def api_admin_usuarios(request):
    # El GET lo permitimos a cualquier autenticado para poder seleccionar usuarios en tareas
    if request.method != 'GET' and not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    
    if not request.user.is_authenticated:
         return JsonResponse({'error': 'No autenticado'}, status=401)

    if request.method == 'GET':
        usuarios = User.objects.select_related('userprofile').all().order_by('first_name', 'last_name')
        data = []
        for u in usuarios:
            grupos = [g.name for g in u.groups.all()]
            profile = getattr(u, 'userprofile', None)
            data.append({
                'id': u.id,
                'username': u.username,
                'first_name': u.first_name,
                'last_name': u.last_name,
                'email': u.email,
                'is_active': u.is_active,
                'is_supervisor': 'Supervisores' in grupos,
                'meta_mensual': str(profile.meta_mensual) if profile else '0',
                'meta_oportunidades': str(getattr(profile, 'meta_oportunidades', 0)) if profile else '0',
                'meta_cotizado': str(getattr(profile, 'meta_cotizado', 0)) if profile else '0',
                'meta_cobrado': str(getattr(profile, 'meta_cobrado', 0)) if profile else '0',
                'rol': getattr(profile, 'rol', 'vendedor') if profile else 'vendedor',
            })
        return JsonResponse({'usuarios': data})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        first_name = data.get('first_name', '').strip()
        last_name = data.get('last_name', '').strip()
        email = data.get('email', '').strip()

        if not username or not password:
            return JsonResponse({'error': 'Username y password son requeridos'}, status=400)

        if User.objects.filter(username=username).exists():
            return JsonResponse({'error': 'El username ya existe'}, status=400)

        from django.contrib.auth.models import Group
        user = User.objects.create_user(
            username=username, password=password,
            first_name=first_name, last_name=last_name, email=email
        )
        profile, _ = UserProfile.objects.get_or_create(user=user)
        rol = data.get('rol', 'vendedor').strip()
        if rol in ('vendedor', 'ingeniero'):
            profile.rol = rol
            profile.save(update_fields=['rol'])

        if data.get('is_supervisor'):
            grupo, _ = Group.objects.get_or_create(name='Supervisores')
            user.groups.add(grupo)

        return JsonResponse({'success': True, 'id': user.id})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_admin_usuario_detalle(request, user_id):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    try:
        usuario = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        if 'first_name' in data:
            usuario.first_name = data['first_name']
        if 'last_name' in data:
            usuario.last_name = data['last_name']
        if 'email' in data:
            usuario.email = data['email']
        if 'is_active' in data:
            usuario.is_active = data['is_active']
        if data.get('password', '').strip():
            usuario.set_password(data['password'].strip())
        usuario.save()

        # Actualizar rol en perfil
        if 'rol' in data:
            profile, _ = UserProfile.objects.get_or_create(user=usuario)
            profile.rol = data['rol']
            profile.save(update_fields=['rol'])

        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_ingeniero_actividades(request):
    """
    Devuelve las tareas asignadas al ingeniero ordenadas para su tablero personal.
    Combina TareaOportunidad y Tarea generales.
    """
    from app.models import IngenieroBoardItem, TareaOportunidad, Tarea
    user = request.user

    from django.db.models import Q
    _puede_ver_todo = is_supervisor(user) or is_ingeniero(user)

    # TareaOportunidad: admins/ingenieros ven todas, vendedores solo las suyas
    if _puede_ver_todo:
        tareas_opp = TareaOportunidad.objects.exclude(
            estado='completada'
        ).select_related('oportunidad', 'oportunidad__cliente', 'responsable')
    else:
        tareas_opp = TareaOportunidad.objects.filter(
            responsable=user
        ).exclude(estado='completada').select_related('oportunidad', 'oportunidad__cliente')

    # Tareas generales: admins/ingenieros ven todas las de proyectos de ingeniería + sus propias
    if _puede_ver_todo:
        tareas_gen = Tarea.objects.exclude(estado='completada').select_related(
            'oportunidad', 'oportunidad__cliente', 'proyecto', 'asignado_a'
        ).filter(Q(proyecto__es_ingenieria=True) | Q(asignado_a=user)).distinct()
    else:
        tareas_gen = Tarea.objects.exclude(estado='completada').select_related(
            'oportunidad', 'oportunidad__cliente', 'proyecto'
        ).filter(
            Q(asignado_a=user) |
            Q(proyecto__es_ingenieria=True, proyecto__miembros=user)
        ).distinct()

    # Obtener órdenes personales guardados
    board_map_opp = {
        bi.tarea_opp_id: bi
        for bi in IngenieroBoardItem.objects.filter(usuario=user, tarea_opp__isnull=False)
    }
    board_map_gen = {
        bi.tarea_id: bi
        for bi in IngenieroBoardItem.objects.filter(usuario=user, tarea__isnull=False)
    }

    items = []
    for t in tareas_opp:
        bi = board_map_opp.get(t.id)
        items.append({
            'key': f'opp_{t.id}',
            'tipo': 'tarea_opp',
            'id': t.id,
            'titulo': t.titulo,
            'descripcion': t.descripcion or '',
            'estado': t.estado,
            'prioridad': t.prioridad,
            'fecha_limite': t.fecha_limite.strftime('%Y-%m-%d') if t.fecha_limite else None,
            'fecha_limite_display': t.fecha_limite.strftime('%d/%m/%Y') if t.fecha_limite else 'Sin fecha',
            'oportunidad': t.oportunidad.oportunidad if t.oportunidad else '',
            'oportunidad_id': t.oportunidad_id,
            'cliente': (t.oportunidad.cliente.nombre_empresa if t.oportunidad and t.oportunidad.cliente else ''),
            'asignado': (f"{t.responsable.first_name} {t.responsable.last_name}".strip() if getattr(t, 'responsable', None) else ''),
            'orden': bi.orden if bi else 9999,
            'fecha_planeada': str(bi.fecha_planeada) if bi and bi.fecha_planeada else None,
        })

    for t in tareas_gen:
        bi = board_map_gen.get(t.id)
        nombre_contexto = ''
        if t.oportunidad:
            nombre_contexto = t.oportunidad.oportunidad or ''
        elif t.proyecto:
            nombre_contexto = t.proyecto.nombre or ''
        items.append({
            'key': f'gen_{t.id}',
            'tipo': 'tarea',
            'id': t.id,
            'titulo': t.titulo,
            'descripcion': t.descripcion or '',
            'estado': t.estado,
            'prioridad': t.prioridad,
            'fecha_limite': t.fecha_limite.strftime('%Y-%m-%d') if t.fecha_limite else None,
            'fecha_limite_display': t.fecha_limite.strftime('%d/%m/%Y') if t.fecha_limite else 'Sin fecha',
            'fecha_creacion': t.fecha_creacion.strftime('%Y-%m-%d') if getattr(t, 'fecha_creacion', None) else '1970-01-01',
            'oportunidad': nombre_contexto,
            'oportunidad_id': t.oportunidad_id,
            'cliente': (t.oportunidad.cliente.nombre_empresa if t.oportunidad and t.oportunidad.cliente else ''),
            'orden': bi.orden if bi else 9999,
            'fecha_planeada': str(bi.fecha_planeada) if bi and bi.fecha_planeada else None,
        })

    # Ordenar tareas por fecha_creacion (más reciente primero)
    items.sort(key=lambda x: x.get('fecha_creacion', '1970-01-01'), reverse=True)

    return JsonResponse({'items': items})


@login_required
@csrf_exempt
def api_ingeniero_board_reorder(request):
    """Guarda el orden personal del tablero del ingeniero."""
    from app.models import IngenieroBoardItem
    if request.method != 'POST':
        return JsonResponse({'error': 'POST requerido'}, status=405)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    # data = [{'key': 'opp_5', 'orden': 0, 'fecha_planeada': '2026-03-05'}, ...]
    for i, item in enumerate(data):
        key = item.get('key', '')
        fecha_planeada = item.get('fecha_planeada') or None
        if key.startswith('opp_'):
            tid = int(key[4:])
            IngenieroBoardItem.objects.update_or_create(
                usuario=request.user, tarea_opp_id=tid, tarea=None,
                defaults={'orden': i, 'fecha_planeada': fecha_planeada}
            )
        elif key.startswith('gen_'):
            tid = int(key[4:])
            IngenieroBoardItem.objects.update_or_create(
                usuario=request.user, tarea_id=tid, tarea_opp=None,
                defaults={'orden': i, 'fecha_planeada': fecha_planeada}
            )

    return JsonResponse({'ok': True})


@login_required
@login_required
def api_ingeniero_proyectos(request):
    """
    Devuelve proyectos de ingeniería (es_ingenieria=True) donde el usuario
    es miembro o los puede ver (supervisores ven todos).
    """
    from app.models import Proyecto
    user = request.user
    if is_supervisor(user) or is_ingeniero(user):
        qs = Proyecto.objects.filter(es_ingenieria=True)
    else:
        qs = Proyecto.objects.filter(es_ingenieria=True, miembros=user)

    _estado_map = {
        'pendiente': 'pendiente',
        'en_progreso': 'en_progreso',
        'completado': 'completada',
        'cancelado': 'cancelada',
    }
    items = []
    for p in qs.order_by('-fecha_creacion'):
        estado_raw = getattr(p, 'estado', 'en_progreso') or 'en_progreso'
        items.append({
            'id': p.id,
            'key': f'proy_{p.id}',
            'titulo': p.nombre,
            'prioridad': getattr(p, 'prioridad', 'media') or 'media',
            'estado': estado_raw,
            'fecha_limite': p.fecha_fin.strftime('%Y-%m-%d') if getattr(p, 'fecha_fin', None) else None,
            'fecha_limite_display': p.fecha_fin.strftime('%d/%m/%Y') if getattr(p, 'fecha_fin', None) else 'Sin fecha',
        })
    return JsonResponse({'items': items})


@login_required
def api_ingeniero_proyecto_detalle(request, proyecto_id):
    """Detalle de un proyecto de ingeniería: info, tareas, carpetas y archivos raíz."""
    from app.models import Proyecto, Tarea, CarpetaProyecto, ArchivoProyecto
    try:
        if is_supervisor(request.user) or is_ingeniero(request.user):
            proyecto = Proyecto.objects.get(pk=proyecto_id, es_ingenieria=True)
        else:
            proyecto = Proyecto.objects.get(pk=proyecto_id, es_ingenieria=True, miembros=request.user)
    except Proyecto.DoesNotExist:
        return JsonResponse({'error': 'Proyecto no encontrado'}, status=404)

    tareas = list(
        Tarea.objects.filter(proyecto=proyecto)
        .select_related('asignado_a')
        .order_by('estado', 'fecha_limite')
        .values('id', 'titulo', 'estado', 'prioridad', 'fecha_limite', 'asignado_a__first_name', 'asignado_a__last_name')
    )
    for t in tareas:
        fn = t.pop('asignado_a__first_name') or ''
        ln = t.pop('asignado_a__last_name') or ''
        t['asignado_a'] = (fn + ' ' + ln).strip() or 'Sin asignar'
        t['fecha_limite'] = t['fecha_limite'].strftime('%d/%m/%Y') if t['fecha_limite'] else None

    carpetas = list(
        CarpetaProyecto.objects.filter(proyecto=proyecto, carpeta_padre=None)
        .values('id', 'nombre')
    )
    for c in carpetas:
        c['archivos'] = list(
            ArchivoProyecto.objects.filter(carpeta_id=c['id'])
            .values('id', 'nombre_original', 'tipo_archivo', 'extension', 'bitrix_download_url', 'tamaño')
        )

    archivos_raiz = list(
        ArchivoProyecto.objects.filter(proyecto=proyecto, carpeta=None)
        .values('id', 'nombre_original', 'tipo_archivo', 'extension', 'bitrix_download_url', 'tamaño')
    )

    creado_por = None
    if proyecto.creado_por:
        fn = proyecto.creado_por.first_name or ''
        ln = proyecto.creado_por.last_name or ''
        creado_por = (fn + ' ' + ln).strip() or proyecto.creado_por.username

    miembros_data = []
    for m in proyecto.miembros.all():
        fn = m.first_name or ''
        ln = m.last_name or ''
        nombre = (fn + ' ' + ln).strip() or m.username
        iniciales = ((fn[:1] if fn else '') + (ln[:1] if ln else '')).upper() or m.username[:2].upper()
        miembros_data.append({
            'nombre': nombre,
            'iniciales': iniciales,
            'username': m.username
        })

    return JsonResponse({
        'id': proyecto.id,
        'titulo': proyecto.nombre,
        'descripcion': proyecto.descripcion or '',
        'estado': getattr(proyecto, 'estado', '') or '',
        'fecha_creacion': proyecto.fecha_creacion.strftime('%d/%m/%Y') if getattr(proyecto, 'fecha_creacion', None) else '',
        'creado_por': creado_por,
        'miembros': miembros_data,
        'tareas': tareas,
        'carpetas': carpetas,
        'archivos_raiz': archivos_raiz,
    })


@login_required
@csrf_exempt
def api_ingeniero_dashboard_stats(request):
    """
    Dashboard stats for the engineer: efficiency, hours, project status.
    Efficiency factors in: tareas, actividades (tareas_opp), and proyectos.
    Hours calculated using business hours only: Mon-Fri, 8am-6pm (10h/day).
    """
    from app.models import TareaOportunidad, Tarea, Proyecto, AsistenciaJornada, EficienciaMensual
    from datetime import timedelta, datetime as dt_class, time as time_class
    import math

    user = request.user
    ahora = timezone.now()
    hoy = timezone.localdate()

    # ── Tareas asignadas ──
    tareas = Tarea.objects.filter(asignado_a=user).exclude(estado='cancelada')
    total_tareas = tareas.count()
    tareas_completadas = tareas.filter(estado='completada').count()
    tareas_progreso = tareas.filter(estado__in=['en_progreso', 'iniciada']).count()
    tareas_pendientes = tareas.filter(estado='pendiente').count()
    tareas_vencidas = tareas.filter(fecha_limite__lt=ahora).exclude(estado='completada').count()

    # ── Actividades (Tareas de Oportunidad) ──
    acts = TareaOportunidad.objects.filter(responsable=user)
    total_acts = acts.count()
    acts_completadas = acts.filter(estado='completada').count()
    acts_pendientes = acts.filter(estado='pendiente').count()
    acts_vencidas = acts.filter(fecha_limite__lt=ahora, estado='pendiente').count()

    # ── Proyectos ── (where user has tasks assigned)
    proyecto_ids = set(
        tareas.exclude(proyecto__isnull=True).values_list('proyecto_id', flat=True)
    )
    proyectos = Proyecto.objects.filter(id__in=proyecto_ids) if proyecto_ids else Proyecto.objects.none()
    total_proyectos = proyectos.count()

    # ── EFICIENCIA DEL INGENIERO ──
    # Base: % of items completed on time vs total actionable items
    total_items = total_tareas + total_acts + total_proyectos
    completados = tareas_completadas + acts_completadas
    vencidos = tareas_vencidas + acts_vencidas

    if total_items > 0:
        # Positive: completed items add points
        puntos_posibles = total_items * 10
        puntos_obtenidos = completados * 10
        # Negative: overdue items penalize
        puntos_obtenidos -= vencidos * 5
        puntos_obtenidos = max(0, puntos_obtenidos)
        eficiencia = round((puntos_obtenidos / puntos_posibles) * 100, 1)
    else:
        eficiencia = 100.0

    # Also check EficienciaMensual for historical data
    em = EficienciaMensual.objects.filter(
        usuario=user, mes=hoy.month, anio=hoy.year
    ).first()
    eficiencia_mensual = float(em.promedio_eficiencia) if em else eficiencia

    # ── HORAS para item seleccionado ──
    # Business hours calculation: Mon-Fri, 8am-6pm (10h/day)
    selected_key = request.GET.get('selected', '')
    horas_plan = 0
    horas_actual = 0

    def _business_hours_between(start_dt, end_dt):
        """Calculate business hours between two datetimes (Mon-Fri, 8:00-18:00)."""
        if not start_dt or not end_dt:
            return 0
        # Ensure timezone aware
        if timezone.is_naive(start_dt):
            start_dt = timezone.make_aware(start_dt)
        if timezone.is_naive(end_dt):
            end_dt = timezone.make_aware(end_dt)
        if start_dt >= end_dt:
            return 0

        total_minutes = 0
        current = start_dt

        while current < end_dt:
            # Skip weekends
            if current.weekday() >= 5:
                current = current.replace(hour=8, minute=0, second=0) + timedelta(days=1)
                continue

            day_start = current.replace(hour=8, minute=0, second=0, microsecond=0)
            day_end = current.replace(hour=18, minute=0, second=0, microsecond=0)

            # Clamp to business hours
            effective_start = max(current, day_start)
            effective_end = min(end_dt, day_end)

            if effective_start < effective_end:
                total_minutes += (effective_end - effective_start).total_seconds() / 60

            # Move to next day
            current = day_start + timedelta(days=1)

        return round(total_minutes / 60, 1)

    if selected_key:
        if selected_key.startswith('opp_'):
            tid = int(selected_key[4:])
            try:
                t = TareaOportunidad.objects.get(id=tid)
                # Actual: business hours since creation until now (or completion)
                horas_actual = _business_hours_between(t.fecha_creacion, ahora)
                # Plan: business hours from creation to deadline
                if t.fecha_limite:
                    horas_plan = _business_hours_between(t.fecha_creacion, t.fecha_limite)
            except TareaOportunidad.DoesNotExist:
                pass
        elif selected_key.startswith('gen_'):
            tid = int(selected_key[4:])
            try:
                t = Tarea.objects.get(id=tid)
                horas_actual = _business_hours_between(t.fecha_creacion, ahora)
                if t.fecha_limite:
                    horas_plan = _business_hours_between(t.fecha_creacion, t.fecha_limite)
            except Tarea.DoesNotExist:
                pass

    return JsonResponse({
        'eficiencia': eficiencia,
        'eficiencia_mensual': eficiencia_mensual,
        'total_tareas': total_tareas,
        'tareas_completadas': tareas_completadas,
        'tareas_progreso': tareas_progreso,
        'tareas_pendientes': tareas_pendientes,
        'tareas_vencidas': tareas_vencidas,
        'total_acts': total_acts,
        'acts_completadas': acts_completadas,
        'acts_pendientes': acts_pendientes,
        'acts_vencidas': acts_vencidas,
        'total_proyectos': total_proyectos,
        'horas_plan': horas_plan,
        'horas_actual': horas_actual,
    })


# ═══ APIS PARA PROGRAMACIÓN DE ACTIVIDADES (CALENDARIO PROYECTO) ═══

@login_required
def api_programacion_actividades(request):
    """GET: lista actividades de un proyecto_key. POST: crear una nueva."""
    if request.method == 'GET':
        proyecto_key = request.GET.get('proyecto_key', '')
        if not proyecto_key:
            return JsonResponse({'error': 'proyecto_key requerido'}, status=400)

        acts = ProgramacionActividad.objects.filter(
            proyecto_key=proyecto_key
        ).prefetch_related('responsables')

        items = []
        for a in acts:
            responsables = []
            for u in a.responsables.all():
                profile = getattr(u, 'userprofile', None)
                iniciales = profile.iniciales() if profile else (u.first_name[:1] + u.last_name[:1]).upper() if u.first_name else u.username[:2].upper()
                responsables.append({
                    'id': u.id,
                    'nombre': u.get_full_name() or u.username,
                    'iniciales': iniciales,
                })
            items.append({
                'id': a.id,
                'titulo': a.titulo,
                'dia_semana': a.dia_semana,
                'fecha': str(a.fecha) if a.fecha else None,
                'hora_inicio': a.hora_inicio.strftime('%H:%M'),
                'hora_fin': a.hora_fin.strftime('%H:%M'),
                'responsables': responsables,
                'creado_por': a.creado_por.get_full_name() or a.creado_por.username,
            })

        return JsonResponse({'success': True, 'items': items})

    if request.method == 'POST':
        data = json.loads(request.body)
        proyecto_key = data.get('proyecto_key', '')
        titulo = data.get('titulo', '').strip()
        dia_semana = data.get('dia_semana', '')
        hora_inicio_str = data.get('hora_inicio', '')
        hora_fin_str = data.get('hora_fin', '')
        responsable_ids = data.get('responsables', [])
        fecha_str = data.get('fecha', '')

        if not proyecto_key or not dia_semana or not hora_inicio_str or not hora_fin_str:
            return JsonResponse({'error': 'Campos requeridos: proyecto_key, dia_semana, hora_inicio, hora_fin'}, status=400)

        from datetime import time as time_class, datetime as dt_class, timedelta
        try:
            h_ini = dt_class.strptime(hora_inicio_str, '%H:%M').time()
            h_fin = dt_class.strptime(hora_fin_str, '%H:%M').time()
        except ValueError:
            return JsonResponse({'error': 'Formato de hora inválido (usar HH:MM)'}, status=400)

        if h_ini >= h_fin:
            return JsonResponse({'error': 'La hora de inicio debe ser antes de la hora de fin'}, status=400)

        fecha = None
        if fecha_str:
            try:
                fecha = dt_class.strptime(fecha_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        # Verificar conflictos para cada responsable
        conflictos = []
        for uid in responsable_ids:
            user_conflicts = ProgramacionActividad.get_conflictos_usuario(
                uid, dia_semana, h_ini, h_fin, fecha=fecha
            )
            if user_conflicts.exists():
                try:
                    u = User.objects.get(id=uid)
                    nombre = u.get_full_name() or u.username
                except User.DoesNotExist:
                    nombre = f'ID {uid}'
                for c in user_conflicts:
                    conflictos.append({
                        'usuario': nombre,
                        'usuario_id': uid,
                        'actividad': c.titulo,
                        'proyecto_key': c.proyecto_key,
                        'hora': f"{c.hora_inicio.strftime('%H:%M')}-{c.hora_fin.strftime('%H:%M')}",
                    })

        if conflictos:
            return JsonResponse({
                'success': False,
                'error': 'conflicto',
                'conflictos': conflictos,
                'mensaje': 'Algunos responsables tienen conflictos de horario',
            }, status=409)

        act = ProgramacionActividad.objects.create(
            proyecto_key=proyecto_key,
            titulo=titulo or 'Actividad sin título',
            dia_semana=dia_semana,
            fecha=fecha,
            hora_inicio=h_ini,
            hora_fin=h_fin,
            creado_por=request.user,
        )
        usuarios_asignados = []
        if responsable_ids:
            usuarios_asignados = list(User.objects.filter(id__in=responsable_ids))
            act.responsables.set(usuarios_asignados)

        # ── Create calendar Actividad for each assigned person ──
        # Use next occurrence of dia_semana if no fecha given
        from datetime import datetime as dt_class
        if fecha:
            act_fecha = fecha
        else:
            dias_map = {'Lunes':0,'Martes':1,'Miércoles':2,'Jueves':3,'Viernes':4,'Sábado':5,'Domingo':6}
            target_day = dias_map.get(dia_semana, 0)
            today = timezone.localdate()
            days_ahead = target_day - today.weekday()
            if days_ahead <= 0:
                days_ahead += 7
            act_fecha = today + timedelta(days=days_ahead)

        fecha_inicio_cal = timezone.make_aware(dt_class.combine(act_fecha, h_ini))
        fecha_fin_cal = timezone.make_aware(dt_class.combine(act_fecha, h_fin))

        proyecto_titulo = data.get('proyecto_titulo', proyecto_key)

        cal_actividad = Actividad.objects.create(
            titulo=f"📋 {proyecto_titulo}: {titulo or 'Actividad programada'}",
            tipo_actividad='reunion',
            descripcion=f"Actividad programada para el proyecto. Día: {dia_semana}.",
            fecha_inicio=fecha_inicio_cal,
            fecha_fin=fecha_fin_cal,
            creado_por=request.user,
            color='#FF2D55',  # Rosa
        )
        if usuarios_asignados:
            cal_actividad.participantes.set(usuarios_asignados)

        # ── Send notifications to assigned users ──
        for u in usuarios_asignados:
            try:
                crear_notificacion(
                    usuario_destinatario=u,
                    tipo='programacion_proyecto',
                    titulo=f"Asignado a actividad de proyecto",
                    mensaje=f"Se te asignó a '{titulo or 'Actividad'}' el {dia_semana} de {h_ini.strftime('%H:%M')} a {h_fin.strftime('%H:%M')} en el proyecto '{proyecto_titulo}'.",
                    usuario_remitente=request.user,
                    proyecto_nombre=proyecto_titulo,
                )
            except Exception:
                pass

        return JsonResponse({'success': True, 'id': act.id})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_programacion_actividad_detail(request, actividad_id):
    """PUT actualizar / DELETE eliminar una actividad programada."""
    act = get_object_or_404(ProgramacionActividad, pk=actividad_id)

    if request.method == 'PUT':
        data = json.loads(request.body)
        titulo = data.get('titulo', '').strip()
        if titulo:
            act.titulo = titulo
        hora_inicio_str = data.get('hora_inicio', '')
        hora_fin_str = data.get('hora_fin', '')
        if hora_inicio_str and hora_fin_str:
            from datetime import datetime as dt_class
            try:
                act.hora_inicio = dt_class.strptime(hora_inicio_str, '%H:%M').time()
                act.hora_fin = dt_class.strptime(hora_fin_str, '%H:%M').time()
            except ValueError:
                pass
        responsable_ids = data.get('responsables')
        if responsable_ids is not None:
            conflictos = []
            for uid in responsable_ids:
                user_conflicts = ProgramacionActividad.get_conflictos_usuario(
                    uid, act.dia_semana, act.hora_inicio, act.hora_fin,
                    exclude_id=act.id, fecha=act.fecha
                )
                if user_conflicts.exists():
                    try:
                        u = User.objects.get(id=uid)
                        nombre = u.get_full_name() or u.username
                    except User.DoesNotExist:
                        nombre = f'ID {uid}'
                    for c in user_conflicts:
                        conflictos.append({
                            'usuario': nombre,
                            'actividad': c.titulo,
                            'hora': f"{c.hora_inicio.strftime('%H:%M')}-{c.hora_fin.strftime('%H:%M')}",
                        })
            if conflictos:
                return JsonResponse({
                    'success': False,
                    'error': 'conflicto',
                    'conflictos': conflictos,
                }, status=409)
            act.responsables.set(User.objects.filter(id__in=responsable_ids))
        act.save()
        return JsonResponse({'success': True})

    if request.method == 'DELETE':
        act.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_programacion_disponibilidad(request):
    """
    GET: Verifica disponibilidad de usuarios para un día y rango de hora.
    """
    dia_semana = request.GET.get('dia_semana', '')
    hora_inicio_str = request.GET.get('hora_inicio', '')
    hora_fin_str = request.GET.get('hora_fin', '')
    fecha_str = request.GET.get('fecha', '')
    exclude_id = request.GET.get('exclude_id')

    if not dia_semana or not hora_inicio_str or not hora_fin_str:
        return JsonResponse({'error': 'Parámetros requeridos: dia_semana, hora_inicio, hora_fin'}, status=400)

    from datetime import datetime as dt_class
    try:
        h_ini = dt_class.strptime(hora_inicio_str, '%H:%M').time()
        h_fin = dt_class.strptime(hora_fin_str, '%H:%M').time()
    except ValueError:
        return JsonResponse({'error': 'Formato de hora inválido'}, status=400)

    fecha = None
    if fecha_str:
        try:
            fecha = dt_class.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    usuarios = User.objects.filter(is_active=True).select_related('userprofile').order_by('first_name', 'last_name')

    resultado = []
    for u in usuarios:
        profile = getattr(u, 'userprofile', None)
        iniciales = profile.iniciales() if profile else (u.first_name[:1] + u.last_name[:1]).upper() if u.first_name else u.username[:2].upper()
        rol = profile.rol if profile else 'vendedor'

        conflictos = ProgramacionActividad.get_conflictos_usuario(
            u.id, dia_semana, h_ini, h_fin,
            exclude_id=int(exclude_id) if exclude_id else None,
            fecha=fecha
        )

        conflict_list = []
        for c in conflictos:
            conflict_list.append({
                'id': c.id,
                'titulo': c.titulo,
                'proyecto_key': c.proyecto_key,
                'hora': f"{c.hora_inicio.strftime('%H:%M')}-{c.hora_fin.strftime('%H:%M')}",
            })

        resultado.append({
            'id': u.id,
            'nombre': u.get_full_name() or u.username,
            'iniciales': iniciales,
            'rol': rol,
            'disponible': len(conflict_list) == 0,
            'conflictos': conflict_list,
        })

    return JsonResponse({'success': True, 'usuarios': resultado})


@login_required
def api_empleados_jornadas(request):
    """Retorna horas trabajadas y eficiencia por día de cada empleado en el mes dado."""
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    import calendar as _cal
    from datetime import date as _date
    try:
        mes = int(request.GET.get('mes', _date.today().month))
        anio = int(request.GET.get('anio', _date.today().year))
    except (ValueError, TypeError):
        mes, anio = _date.today().month, _date.today().year

    users = User.objects.filter(is_active=True).exclude(
        groups__name='Supervisores'
    ).order_by('first_name', 'last_name')

    jornadas = AsistenciaJornada.objects.filter(
        fecha__year=anio, fecha__month=mes
    ).select_related('usuario')

    data = {}
    for j in jornadas:
        uid = j.usuario_id
        day = j.fecha.day
        if uid not in data:
            data[uid] = {}
        data[uid][day] = {
            'horas': round(j.segundos_laborados / 3600, 1),
            'eficiencia': float(j.eficiencia_dia),
        }

    # Today's status (only relevant if viewing current month)
    today = _date.today()
    today_estados = {}
    if mes == today.month and anio == today.year:
        hoy_jornadas = AsistenciaJornada.objects.filter(fecha=today)
        for j in hoy_jornadas:
            if j.hora_fin:
                today_estados[j.usuario_id] = 'inactivo'
            elif j.pausado:
                today_estados[j.usuario_id] = 'pausa'
            else:
                today_estados[j.usuario_id] = 'activo'

    empleados = []
    for u in users:
        dias = data.get(u.id, {})
        empleados.append({
            'id': u.id,
            'nombre': ((u.first_name + ' ' + u.last_name).strip() or u.username),
            'dias': {str(d): v for d, v in dias.items()},
            'estado_hoy': today_estados.get(u.id),
        })

    _, num_dias = _cal.monthrange(anio, mes)
    return JsonResponse({'empleados': empleados, 'mes': mes, 'anio': anio, 'num_dias': num_dias})


@login_required
def api_clientes_rapido(request):
    """GET: lista de clientes. POST: crear cliente rápido."""
    if request.method == 'GET':
        if is_supervisor(request.user):
            clientes = Cliente.objects.all().order_by('nombre_empresa')
        else:
            clientes = Cliente.objects.filter(asignado_a=request.user).order_by('nombre_empresa')
        data = [{'id': c.id, 'nombre': c.nombre_empresa} for c in clientes]
        return JsonResponse({'clientes': data})
    elif request.method == 'POST':
        try:
            body = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'JSON inválido'}, status=400)
        nombre = (body.get('nombre') or '').strip()
        if not nombre:
            return JsonResponse({'error': 'Nombre requerido'}, status=400)
        cliente = Cliente.objects.create(
            nombre_empresa=nombre,
            asignado_a=request.user,
        )
        return JsonResponse({'id': cliente.id, 'nombre': cliente.nombre_empresa})
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_oportunidades_rapido(request):
    """GET: oportunidades de un cliente. POST: crea oportunidad rápida."""
    if request.method == 'GET':
        cliente_id = request.GET.get('cliente_id')
        if not cliente_id:
            return JsonResponse({'opps': []})
        qs = TodoItem.objects.filter(cliente_id=cliente_id).order_by('-id')
        if not is_supervisor(request.user):
            qs = qs.filter(usuario=request.user)
        data = [{'id': o.id, 'nombre': o.oportunidad} for o in qs[:50]]
        return JsonResponse({'opps': data})
    elif request.method == 'POST':
        try:
            body = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'JSON inválido'}, status=400)
        nombre = (body.get('nombre') or '').strip()
        cliente_id = body.get('cliente_id')
        if not nombre or not cliente_id:
            return JsonResponse({'error': 'Nombre y cliente requeridos'}, status=400)
        try:
            cliente = Cliente.objects.get(id=cliente_id)
        except Cliente.DoesNotExist:
            return JsonResponse({'error': 'Cliente no encontrado'}, status=404)
        today = timezone.localdate()
        opp = TodoItem.objects.create(
            usuario=request.user,
            oportunidad=nombre,
            cliente=cliente,
            producto='ZEBRA',
            monto=Decimal('0.00'),
            probabilidad_cierre=5,
            mes_cierre=str(today.month).zfill(2),
            anio_cierre=today.year,
            area='SISTEMAS',
            tipo_negociacion='runrate',
            estado_crm='nueva',
        )
        return JsonResponse({'id': opp.id, 'nombre': opp.oportunidad})
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_jornada_ayer(request):
    """Retorna la eficiencia del día anterior del usuario."""
    from datetime import timedelta
    ayer = timezone.localdate() - timedelta(days=1)
    jornada = AsistenciaJornada.objects.filter(usuario=request.user, fecha=ayer).first()
    if jornada:
        return JsonResponse({'eficiencia': float(jornada.eficiencia_dia), 'fecha': str(ayer)})
    return JsonResponse({'eficiencia': None, 'fecha': str(ayer)})


@login_required
def api_admin_clientes(request):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method == 'GET':
        clientes = Cliente.objects.select_related('asignado_a').all().order_by('nombre_empresa')
        data = []
        for c in clientes:
            data.append({
                'id': c.id,
                'nombre_empresa': c.nombre_empresa,
                'categoria': c.categoria,
                'asignado_a_id': c.asignado_a_id,
                'asignado_a_name': c.asignado_a.get_full_name() or c.asignado_a.username if c.asignado_a else '',
                'telefono': c.telefono or '',
                'email': c.email or '',
                'meta_mensual': str(c.meta_mensual or 0),
            })
        return JsonResponse({'clientes': data})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_admin_cliente_detalle(request, cliente_id):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)

    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        if 'asignado_a_id' in data:
            uid = data['asignado_a_id']
            if uid:
                try:
                    cliente.asignado_a = User.objects.get(id=uid)
                except User.DoesNotExist:
                    return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
            else:
                cliente.asignado_a = None
        if 'categoria' in data:
            cliente.categoria = data['categoria']
        if 'meta_mensual' in data:
            try:
                cliente.meta_mensual = Decimal(str(data['meta_mensual']))
            except Exception:
                pass
        cliente.save()

        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_admin_contactos(request):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method == 'GET':
        cliente_id = request.GET.get('cliente_id')
        qs = Contacto.objects.select_related('empresa').all().order_by('nombre')
        if cliente_id:
            qs = qs.filter(empresa_id=cliente_id)
        data = []
        for c in qs:
            data.append({
                'id': c.id,
                'nombre': c.nombre,
                'apellido': c.apellido or '',
                'email': c.email or '',
                'telefono': c.telefono or '',
                'empresa_id': c.empresa_id,
                'empresa_name': c.empresa.nombre_empresa if c.empresa else '',
            })
        return JsonResponse({'contactos': data})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        nombre = data.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'Nombre es requerido'}, status=400)

        empresa_id = data.get('empresa_id')
        if not empresa_id:
            return JsonResponse({'error': 'Cliente es requerido'}, status=400)

        try:
            empresa = Cliente.objects.get(id=empresa_id)
        except Cliente.DoesNotExist:
            return JsonResponse({'error': 'Cliente no encontrado'}, status=404)

        contacto = Contacto.objects.create(
            nombre=nombre,
            apellido=data.get('apellido', ''),
            email=data.get('email', ''),
            telefono=data.get('telefono', ''),
            empresa=empresa,
        )
        return JsonResponse({'success': True, 'id': contacto.id})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_admin_metas(request):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        metas = data.get('metas', {})  # {uid: {mensual: x, oportunidades: y...}} or {uid: val}
        updated = 0
        for uid_str, meta_data in metas.items():
            try:
                uid = int(uid_str)
                profile, _ = UserProfile.objects.get_or_create(user_id=uid)
                
                # Check if meta_data is dict or value (legacy)
                if isinstance(meta_data, dict):
                    if 'mensual' in meta_data: profile.meta_mensual = Decimal(str(meta_data['mensual']))
                    if 'oportunidades' in meta_data: profile.meta_oportunidades = Decimal(str(meta_data['oportunidades']))
                    if 'cotizado' in meta_data: profile.meta_cotizado = Decimal(str(meta_data['cotizado']))
                    if 'cobrado' in meta_data: profile.meta_cobrado = Decimal(str(meta_data['cobrado']))
                else:
                    # Legacy flat value = mensual
                    profile.meta_mensual = Decimal(str(meta_data))
                
                profile.save()
                updated += 1
            except (ValueError, User.DoesNotExist, decimal.InvalidOperation):
                continue

        return JsonResponse({'success': True, 'updated': updated})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_admin_permisos(request, user_id):
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        try:
            usuario = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return JsonResponse({'error': 'Usuario no encontrado'}, status=404)

        from django.contrib.auth.models import Group
        grupo, _ = Group.objects.get_or_create(name='Supervisores')

        if data.get('is_supervisor'):
            usuario.groups.add(grupo)
        else:
            usuario.groups.remove(grupo)

        return JsonResponse({'success': True, 'is_supervisor': data.get('is_supervisor', False)})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def bienvenida(request):
    """
    Vista de bienvenida que será la primera que vea el usuario al ingresar.
    """
    # Asegurar que el usuario tenga un perfil
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    # Determinar perfil
    perfil = "SUPERVISOR" if is_supervisor(request.user) else "VENDEDOR"
    # Fecha actual
    fecha_actual = timezone.localtime(timezone.now()).strftime('%A, %d de %B de %Y')

    # Usuario del mes: quien más oportunidades ha creado desde el 1 de septiembre 2025
    # Se calcula desde el 1 de septiembre hasta hoy
    today = date.today()
    inicio_mes_actual = date(2025, 9, 1)  # Fijado al 1 de septiembre 2025
    oportunidades_mes = (
        TodoItem.objects.filter(fecha_creacion__date__gte=inicio_mes_actual, fecha_creacion__date__lte=today)
        .exclude(usuario__groups__name='Supervisores')
        .values('usuario')
        .annotate(oportunidades_creadas=Count('id'))
        .order_by('-oportunidades_creadas')
    )
    usuario_mes = None
    
    # TEMPORAL: Forzar que Rivera sea empleado del mes para pruebas
    try:
        rivera_user = User.objects.get(username='Rivera')
        rivera_profile = UserProfile.objects.get(user=rivera_user)
        avatar_url = rivera_profile.get_avatar_url()
        if not avatar_url:
            avatar_url = f'https://ui-avatars.com/api/?name={rivera_user.get_full_name() or rivera_user.username}&background=38bdf8&color=fff'
        logger.info(f"DEBUG: URL del avatar del mes (Rivera): {avatar_url}")
        
        # Calcular oportunidades y monto para Rivera
        oportunidades_rivera = TodoItem.objects.filter(
            fecha_creacion__date__gte=inicio_mes_actual,
            fecha_creacion__date__lte=today,
            usuario=rivera_user
        ).count()
        
        monto_total_rivera = TodoItem.objects.filter(
            fecha_creacion__date__gte=inicio_mes_actual,
            fecha_creacion__date__lte=today,
            usuario=rivera_user
        ).aggregate(total=Sum('monto'))['total'] or 150000  # Valor ejemplo
        
        usuario_mes = {
            'nombre': rivera_user.get_full_name() or rivera_user.username,
            'avatar_url': avatar_url,
            'oportunidades_creadas': oportunidades_rivera if oportunidades_rivera > 0 else 25,  # Valor ejemplo
            'monto_total_mes': monto_total_rivera,
            'user': rivera_user  # Añadir el objeto user para el filtro
        }
    except User.DoesNotExist:
        # Si no existe Rivera, usar la lógica original
        if oportunidades_mes:
            user_id = oportunidades_mes[0]['usuario']
            user = User.objects.get(id=user_id)
            user_profile = UserProfile.objects.get(user=user)
            avatar_url = user_profile.get_avatar_url()
            if not avatar_url:
                avatar_url = f'https://ui-avatars.com/api/?name={user.get_full_name() or user.username}&background=38bdf8&color=fff'
            logger.info(f"DEBUG: URL del avatar del mes: {avatar_url}")

            # Calcular el monto total de oportunidades creadas por este usuario en el mes actual
            monto_total_mes = TodoItem.objects.filter(
                fecha_creacion__date__gte=inicio_mes_actual,
                fecha_creacion__date__lte=today,
                usuario=user
            ).aggregate(total=Sum('monto'))['total'] or 0
            usuario_mes = {
                'nombre': user.get_full_name() or user.username,
                'avatar_url': avatar_url,
                'oportunidades_creadas': oportunidades_mes[0]['oportunidades_creadas'],
                'monto_total_mes': monto_total_mes,
                'user': user  # Añadir el objeto user para el filtro
            }

    # Usuario del día: más oportunidades registradas hoy, solo vendedores
    hoy = date.today()
    usuario_dia = None
    oportunidades_hoy = (
        TodoItem.objects.filter(fecha_creacion__date=hoy)
        .exclude(usuario__groups__name='Supervisores')
        .values('usuario')
        .annotate(oportunidades_hoy=Count('id'))
        .order_by('-oportunidades_hoy')
    )
    if oportunidades_hoy and oportunidades_hoy[0]['oportunidades_hoy'] > 0:
        user_id = oportunidades_hoy[0]['usuario']
        user = User.objects.get(id=user_id)
        user_profile = UserProfile.objects.get(user=user)
        avatar_url = user_profile.get_avatar_url()
        if not avatar_url:
            avatar_url = f'https://ui-avatars.com/api/?name={user.get_full_name() or user.username}&background=f472b6&color=fff'
        logger.info(f"DEBUG: URL del avatar del día: {avatar_url}")

        usuario_dia = {
            'nombre': user.get_full_name() or user.username,
            'avatar_url': avatar_url,
            'oportunidades_hoy': oportunidades_hoy[0]['oportunidades_hoy'],
            'user': user  # Añadir el objeto user para el filtro
        }
    # Si no hay oportunidades hoy, usuario_dia queda en None

    # Últimas oportunidades (de todos)
    ultimas_oportunidades_qs = TodoItem.objects.select_related('cliente', 'usuario').order_by('-fecha_creacion')[:8]
    ultimas_oportunidades = [
        {
            'id': o.id,  # Add the ID here
            'nombre': o.oportunidad,
            'cliente': o.cliente.nombre_empresa if o.cliente else '',
            'monto': o.monto,
            'probabilidad': o.probabilidad_cierre,
            'usuario': o.usuario.get_full_name() or o.usuario.username if o.usuario else '',
        }
        for o in ultimas_oportunidades_qs
    ]

    # Clima: dejar None, preparado para integración futura
    clima = None

    context = {
        'perfil': perfil,
        'fecha_actual': fecha_actual,
        'usuario_mes': usuario_mes,
        'usuario_dia': usuario_dia,
        'ultimas_oportunidades': ultimas_oportunidades,
        'clima': clima,
    }
    return render(request, 'bienvenida.html', context)


@login_required
def feed(request):
    """
    Vista para el feed de actividad que muestra las actualizaciones recientes
    y publicaciones de los usuarios.
    """
    try:
        # Obtener el perfil del usuario actual
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        # Si no existe, crear uno con valores predeterminados
        user_profile = UserProfile.objects.create(
            user=request.user,
            language='es',
            avatar_tipo='dinosaur'
        )
    
    context = {
        'user_profile': user_profile,
    }
    
    return render(request, 'feed.html', context)


@login_required
def tareas_proyectos(request):
    """
    Vista para la sección de Tareas y Proyectos (solo superusuarios)
    """
    # Verificar que el usuario sea superusuario
    if not request.user.is_superuser:
        return redirect('home')
    
    # Crear datos de ejemplo si no existen (solo para desarrollo)
    if request.GET.get('crear_ejemplos') == 'si':
        crear_datos_ejemplo_tareas(request.user)
    
    context = {
        'user': request.user,
        'page_title': 'Tareas y Proyectos'
    }
    
    return render(request, 'tareas_proyectos.html', context)

def crear_datos_ejemplo_tareas(usuario):
    """
    Función temporal para crear datos de ejemplo
    """
    from datetime import datetime, timedelta
    
    # Crear proyecto de ejemplo si no existe
    proyecto, created = Proyecto.objects.get_or_create(
        nombre="Proyecto Demo",
        defaults={
            'descripcion': 'Proyecto de demostración para el sistema de tareas',
            'creado_por': usuario,
            'privacidad': 'publico'
        }
    )
    
    # Crear algunas tareas de ejemplo
    tareas_ejemplo = [
        {
            'titulo': 'Informacion de Pedidos recibidos',
            'descripcion': 'Revisar y procesar la información de pedidos recibidos del sistema',
            'estado': 'en_progreso',
            'prioridad': 'alta',
            'fecha_limite': datetime.now() + timedelta(days=2)
        },
        {
            'titulo': 'Documentos factura 10827',
            'descripcion': 'Procesar documentos de la factura número 10827',
            'estado': 'pendiente',
            'prioridad': 'media',
            'fecha_limite': datetime.now() + timedelta(days=5)
        },
        {
            'titulo': 'Actualizar base de datos',
            'descripcion': 'Realizar actualización de la base de datos del sistema',
            'estado': 'completada',
            'prioridad': 'baja',
            'fecha_limite': datetime.now() - timedelta(days=1)
        }
    ]
    
    for tarea_data in tareas_ejemplo:
        Tarea.objects.get_or_create(
            titulo=tarea_data['titulo'],
            proyecto=proyecto,
            defaults={
                'descripcion': tarea_data['descripcion'],
                'creado_por': usuario,
                'asignado_a': usuario,
                'estado': tarea_data['estado'],
                'prioridad': tarea_data['prioridad'],
                'fecha_limite': tarea_data['fecha_limite']
            }
        )


@login_required
def api_proyectos(request):
    """
    API para obtener proyectos con paginación
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    # Obtener parámetros de paginación
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 30))
    search = request.GET.get('search', '').strip()
    
    # Obtener proyectos reales de la base de datos
    try:
        # Obtener TODOS los proyectos primero para debug
        proyectos_query = Proyecto.objects.all()
        
        # Filtrar por búsqueda si hay término
        if search:
            proyectos_query = proyectos_query.filter(
                Q(nombre__icontains=search) | 
                Q(descripcion__icontains=search)
            )
        
        # Aplicar paginación
        from django.core.paginator import Paginator
        paginator = Paginator(proyectos_query, per_page)
        proyectos_pagina = paginator.get_page(page)
        
        # Convertir a formato JSON
        proyectos_data = []
        for proyecto in proyectos_pagina:
            proyectos_data.append({
                'id': proyecto.id,
                'nombre': proyecto.nombre,
                'descripcion': proyecto.descripcion or '',
                'avance': proyecto.get_avance_porcentaje(),
                'fecha_creacion': proyecto.fecha_creacion.strftime('%d de %b, %Y'),
                'creador': {
                    'id': proyecto.creado_por.id,
                    'nombre': proyecto.creado_por.get_full_name() or proyecto.creado_por.username,
                    'iniciales': ''.join([palabra[0].upper() for palabra in (proyecto.creado_por.get_full_name() or proyecto.creado_por.username).split()[:2]]),
                    'avatar_url': proyecto.creado_por.userprofile.get_avatar_url() if hasattr(proyecto.creado_por, 'userprofile') else None
                },
                'miembros': proyecto.get_miembros_display(),
                'privacidad': proyecto.privacidad,
                'tipo': proyecto.tipo,
                'mi_rol': proyecto.get_rol_usuario(request.user)
            })
            
            # Agregar oportunidades de forma segura
            try:
                oportunidades = []
                print(f"DEBUG: Proyecto {proyecto.id} - {proyecto.nombre}")
                print(f"DEBUG: Tiene oportunidades_ligadas attr: {hasattr(proyecto, 'oportunidades_ligadas')}")
                
                if hasattr(proyecto, 'oportunidades_ligadas'):
                    opp_count = proyecto.oportunidades_ligadas.count()
                    print(f"DEBUG: Número de oportunidades ligadas: {opp_count}")
                    
                    for oportunidad in proyecto.oportunidades_ligadas.all():
                        print(f"DEBUG: Oportunidad encontrada: {oportunidad.oportunidad}")
                        oportunidades.append({
                            'id': oportunidad.id,
                            'titulo': oportunidad.oportunidad
                        })
                proyectos_data[-1]['oportunidades_ligadas'] = oportunidades
                print(f"DEBUG: Total oportunidades procesadas: {len(oportunidades)}")
            except Exception as opp_error:
                print(f"ERROR obteniendo oportunidades para proyecto {proyecto.id}: {opp_error}")
                import traceback
                traceback.print_exc()
                proyectos_data[-1]['oportunidades_ligadas'] = []

        return JsonResponse({
            'success': True,
            'proyectos': proyectos_data,
            'total': paginator.count,
            'page': page,
            'per_page': per_page,
            'total_pages': paginator.num_pages
        })
        
    except Exception as e:
        print(f"Error obteniendo proyectos: {e}")
        # Fallback a datos de ejemplo si hay error
        usuario_actual_id = request.user.id
    
    proyectos_ejemplo = [
        {
            'id': 1380,
            'nombre': 'RFQ-20250428-11 CONTROL DE ACCESO TORNIQUETE',
            'descripcion': 'Sistema de control de acceso con torniquetes inteligentes',
            'avance': 63,
            'fecha_creacion': '15 de Ago, 2024',
            'creador': {
                'id': 1,
                'nombre': 'Diego Rivera',
                'iniciales': 'DR'
            },
            'miembros': [
                {'id': 1, 'nombre': 'Diego Rivera', 'iniciales': 'DR'},
                {'id': 2, 'nombre': 'Juan García', 'iniciales': 'JG'},
                {'id': 3, 'nombre': 'María Pérez', 'iniciales': 'MP'},
                {'id': 4, 'nombre': 'Ana López', 'iniciales': 'AL'}
            ],
            'privacidad': 'publico',
            'mi_rol': 'Jefe de proyecto' if usuario_actual_id == 1 else ('Miembro' if usuario_actual_id in [2, 3, 4] else 'No te has unido al proyecto')
        },
        {
            'id': 2056,
            'nombre': 'CRM NetHive - Sistema de Gestión',
            'descripcion': 'Desarrollo del sistema CRM personalizado',
            'avance': 100,
            'fecha_creacion': '10 de Jul, 2024',
            'creador': {
                'id': 2,
                'nombre': 'Juan García',
                'iniciales': 'JG'
            },
            'miembros': [
                {'id': 2, 'nombre': 'Juan García', 'iniciales': 'JG'},
                {'id': 5, 'nombre': 'Luis Martínez', 'iniciales': 'LM'}
            ],
            'privacidad': 'privado',
            'mi_rol': 'Jefe de proyecto' if usuario_actual_id == 2 else ('Miembro' if usuario_actual_id == 5 else 'No te has unido al proyecto')
        },
        {
            'id': 774,
            'nombre': 'Telvista 34630 CH - Sistema Avigilon',
            'descripcion': 'Implementación de sistema de videovigilancia',
            'avance': 45,
            'fecha_creacion': '5 de Sep, 2024',
            'creador': {
                'id': 3,
                'nombre': 'María Pérez',
                'iniciales': 'MP'
            },
            'miembros': [
                {'id': 3, 'nombre': 'María Pérez', 'iniciales': 'MP'},
                {'id': 1, 'nombre': 'Diego Rivera', 'iniciales': 'DR'},
                {'id': 4, 'nombre': 'Ana López', 'iniciales': 'AL'},
                {'id': 6, 'nombre': 'Carlos Ruiz', 'iniciales': 'CR'}
            ],
            'privacidad': 'privado',
            'mi_rol': 'Jefe de proyecto' if usuario_actual_id == 3 else ('Miembro' if usuario_actual_id in [1, 4, 6] else 'No te has unido al proyecto')
        },
        {
            'id': 1560,
            'nombre': '8700053503 - Equipos ZEBRA PIMS',
            'descripcion': 'Instalación y configuración de equipos ZEBRA',
            'avance': 85,
            'fecha_creacion': '20 de Ago, 2024',
            'creador': {
                'id': 1,
                'nombre': 'Diego Rivera',
                'iniciales': 'DR'
            },
            'miembros': [
                {'id': 1, 'nombre': 'Diego Rivera', 'iniciales': 'DR'},
                {'id': 7, 'nombre': 'Ana Beltrán', 'iniciales': 'AB'},
                {'id': 8, 'nombre': 'Carlos Fuentes', 'iniciales': 'CF'}
            ],
            'privacidad': 'publico',
            'mi_rol': 'Jefe de proyecto' if usuario_actual_id == 1 else ('Miembro' if usuario_actual_id in [7, 8] else 'No te has unido al proyecto')
        },
        {
            'id': 1130,
            'nombre': 'PO 4201027104 - Instalación Access Point',
            'descripcion': 'Instalación de puntos de acceso inalámbricos',
            'avance': 92,
            'fecha_creacion': '1 de Sep, 2024',
            'creador': {
                'id': 4,
                'nombre': 'Ana López',
                'iniciales': 'AL'
            },
            'miembros': [
                {'id': 4, 'nombre': 'Ana López', 'iniciales': 'AL'},
                {'id': 9, 'nombre': 'José Morales', 'iniciales': 'JM'},
                {'id': 10, 'nombre': 'Teresa Ramírez', 'iniciales': 'TR'},
                {'id': 11, 'nombre': 'David López', 'iniciales': 'DL'}
            ],
            'privacidad': 'privado',
            'mi_rol': 'Jefe de proyecto' if usuario_actual_id == 4 else ('Miembro' if usuario_actual_id in [9, 10, 11] else 'No te has unido al proyecto')
        }
    ]
    
    # Filtrar por búsqueda si se proporciona
    if search:
        proyectos_filtrados = [
            p for p in proyectos_ejemplo 
            if search.lower() in p['nombre'].lower() or search.lower() in p['descripcion'].lower()
        ]
    else:
        proyectos_filtrados = proyectos_ejemplo
    
    # Simular paginación
    total_proyectos = len(proyectos_filtrados)
    start_index = (page - 1) * per_page
    end_index = start_index + per_page
    proyectos_pagina = proyectos_filtrados[start_index:end_index]
    
    return JsonResponse({
        'success': True,
        'proyectos': proyectos_pagina,
        'total': total_proyectos,
        'page': page,
        'per_page': per_page,
        'total_pages': (total_proyectos + per_page - 1) // per_page
    })


# Variable temporal para almacenar tareas creadas durante la sesión
# En un entorno real, esto se almacenaría en la base de datos
tareas_temporales = []

@login_required
def api_tareas(request):
    """
    API para obtener y crear tareas
    """
    
    if request.method == 'GET':
        # Obtener tareas por proyecto si se especifica, sino mostrar TODAS las tareas
        proyecto_id = request.GET.get('proyecto_id')
        
        oportunidad_id = request.GET.get('oportunidad_id')
        try:
            if proyecto_id:
                # Filtrar por proyecto específico
                proyecto = Proyecto.objects.get(id=proyecto_id)
                tareas = Tarea.objects.filter(proyecto=proyecto).select_related(
                    'creado_por', 'asignado_a', 'proyecto'
                ).order_by('-fecha_creacion')
            elif oportunidad_id:
                # Filtrar por oportunidad (tareas del widget de oportunidad)
                tareas = Tarea.objects.filter(oportunidad_id=oportunidad_id).select_related(
                    'creado_por', 'asignado_a', 'proyecto', 'oportunidad'
                ).order_by('fecha_limite', '-fecha_creacion')
            else:
                # Mostrar TODAS las tareas
                if request.user.is_superuser:
                    # El admin ve todo el universo de tareas
                    tareas = Tarea.objects.all().distinct().select_related('creado_por', 'asignado_a', 'proyecto', 'oportunidad', 'oportunidad__cliente').order_by('-fecha_creacion')
                else:
                    # Usuario normal solo ve las suyas
                    tareas = Tarea.objects.filter(
                        Q(creado_por=request.user) |
                        Q(asignado_a=request.user) |
                        Q(participantes=request.user)
                    ).distinct().select_related('creado_por', 'asignado_a', 'proyecto', 'oportunidad', 'oportunidad__cliente').order_by('-fecha_creacion')
            
            tareas_data = []
            for tarea in tareas:
                # Formatear tiempo trabajado
                tiempo_total_str = "00:00:00"
                if hasattr(tarea, 'tiempo_trabajado') and tarea.tiempo_trabajado:
                    total_seconds = int(tarea.tiempo_trabajado.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    tiempo_total_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                
                tareas_data.append({
                    'id': tarea.id,
                    'titulo': tarea.titulo,
                    'descripcion': tarea.descripcion,
                    'estado': tarea.estado,
                    'prioridad': tarea.prioridad,
                    'fecha_creacion': tarea.fecha_creacion.isoformat(),
                    'fecha_limite': tarea.fecha_limite.isoformat() if tarea.fecha_limite else None,
                    'fecha_completada': tarea.fecha_completada.isoformat() if tarea.fecha_completada else None,
                    'creado_por': tarea.creado_por.get_full_name() or tarea.creado_por.username,
                    'responsable': tarea.asignado_a.get_full_name() or tarea.asignado_a.username if tarea.asignado_a else None,
                    'proyecto_nombre': tarea.proyecto.nombre if tarea.proyecto else 'Sin proyecto',
                    'proyecto_id': tarea.proyecto.id if tarea.proyecto else None,
                    'oportunidad_id': tarea.oportunidad.id if tarea.oportunidad else None,
                    'oportunidad_nombre': tarea.oportunidad.oportunidad if tarea.oportunidad else None,
                    'oportunidad_cliente': tarea.oportunidad.cliente.nombre_empresa if tarea.oportunidad and tarea.oportunidad.cliente else None,
                    # Datos del cronómetro
                    'trabajando_actualmente': getattr(tarea, 'trabajando_actualmente', False),
                    'pausado': getattr(tarea, 'pausado', False),
                    'tiempo_trabajado': tiempo_total_str,
                    'fecha_inicio_sesion': tarea.fecha_inicio_sesion.isoformat() if hasattr(tarea, 'fecha_inicio_sesion') and tarea.fecha_inicio_sesion else None,
                })
            
            return JsonResponse({
                'success': True,
                'tareas': tareas_data
            })
            
        except Proyecto.DoesNotExist:
            return JsonResponse({'error': 'Proyecto no encontrado'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Error obteniendo tareas: {str(e)}'}, status=500)
    
    elif request.method == 'POST':
        # Crear nueva tarea
        try:
            from datetime import datetime
            
            # Obtener datos del formulario
            titulo = request.POST.get('titulo', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            proyecto_id = request.POST.get('proyecto_id')
            
            # Manejar prioridad - puede venir como 'alta_prioridad' boolean o 'prioridad' string
            alta_prioridad = request.POST.get('alta_prioridad', 'false').lower() == 'true'
            prioridad = request.POST.get('prioridad', 'alta' if alta_prioridad else 'media')
            
            fecha_limite_str = request.POST.get('fecha_limite')
            responsable_id = request.POST.get('responsable_id')
            
            # Validaciones
            if not titulo:
                return JsonResponse({'error': 'El título es requerido'}, status=400)
            if not proyecto_id:
                return JsonResponse({'error': 'El proyecto es requerido'}, status=400)
            
            # Obtener proyecto
            try:
                proyecto = Proyecto.objects.get(id=proyecto_id)
            except Proyecto.DoesNotExist:
                return JsonResponse({'error': 'Proyecto no encontrado'}, status=404)
            
            # Procesar fecha límite
            fecha_limite = None
            if fecha_limite_str:
                try:
                    fecha_limite = datetime.fromisoformat(fecha_limite_str.replace('Z', '+00:00'))
                except ValueError:
                    return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)
            
            # Obtener responsable
            responsable = None
            if responsable_id:
                try:
                    responsable = User.objects.get(id=responsable_id)
                except User.DoesNotExist:
                    return JsonResponse({'error': 'Usuario responsable no encontrado'}, status=404)
            
            # Crear tarea
            tarea = Tarea.objects.create(
                titulo=titulo,
                descripcion=descripcion,
                proyecto=proyecto,
                creado_por=request.user,
                asignado_a=responsable,
                prioridad=prioridad,
                fecha_limite=fecha_limite
            )
            
            return JsonResponse({
                'success': True,
                'tarea_id': tarea.id,
                'mensaje': 'Tarea creada exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'error': f'Error al crear la tarea: {str(e)}'
            }, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_actualizar_estado_tarea(request):
    """
    API para actualizar el estado de una tarea
    """
    global tareas_temporales
    
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method == 'POST':
        try:
            tarea_id = request.POST.get('tarea_id')
            nuevo_estado = request.POST.get('nuevo_estado')
            
            # Validaciones básicas
            if not tarea_id:
                return JsonResponse({'error': 'ID de tarea requerido'}, status=400)
            
            if not nuevo_estado:
                return JsonResponse({'error': 'Nuevo estado requerido'}, status=400)
            
            estados_validos = ['pendiente', 'en_progreso', 'completada', 'cancelada']
            if nuevo_estado not in estados_validos:
                return JsonResponse({'error': 'Estado no válido'}, status=400)
            
            # Buscar la tarea en la lista temporal
            tarea_encontrada = None
            for i, tarea in enumerate(tareas_temporales):
                if str(tarea['id']) == str(tarea_id):
                    tarea_encontrada = tarea
                    break
            
            if not tarea_encontrada:
                return JsonResponse({'error': 'Tarea no encontrada'}, status=404)
            
            # Verificar permisos: solo el responsable puede cambiar estados
            if tarea_encontrada['responsable_id'] != request.user.id:
                return JsonResponse({'error': 'Solo el responsable puede cambiar el estado de la tarea'}, status=403)
            
            # Actualizar el estado
            tarea_encontrada['estado'] = nuevo_estado
            
            return JsonResponse({
                'success': True,
                'message': 'Estado actualizado exitosamente',
                'tarea': tarea_encontrada
            })
            
        except Exception as e:
            return JsonResponse({
                'error': f'Error al actualizar estado: {str(e)}'
            }, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_actualizar_tarea(request):
    """
    API para actualizar una tarea completa (solo creadores)
    """
    global tareas_temporales
    
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method == 'POST':
        try:
            import json
            from datetime import datetime
            from django.contrib.auth.models import User
            from .models import Proyecto
            
            # Obtener datos del request
            data = json.loads(request.body)
            tarea_id = data.get('tarea_id')
            titulo = data.get('titulo', data.get('nombre', '')).strip()
            descripcion = data.get('descripcion', '').strip()
            proyecto_id = data.get('proyecto_id')
            alta_prioridad = data.get('alta_prioridad', False)
            prioridad = data.get('prioridad', 'media')
            fecha_limite = data.get('fecha_limite')
            responsable_id = data.get('responsable_id')
            participantes_json = data.get('participantes', '[]')
            observadores_json = data.get('observadores', '[]')
            
            # Validaciones básicas
            if not tarea_id:
                return JsonResponse({'error': 'ID de tarea requerido'}, status=400)
            
            if not titulo:
                return JsonResponse({'error': 'El título es requerido'}, status=400)
            
            # Buscar la tarea en la lista temporal
            tarea_encontrada = None
            for i, tarea in enumerate(tareas_temporales):
                if str(tarea['id']) == str(tarea_id):
                    tarea_encontrada = tarea
                    break
            
            if not tarea_encontrada:
                return JsonResponse({'error': 'Tarea no encontrada'}, status=404)
            
            # Verificar permisos: solo el creador puede editar
            if tarea_encontrada['creado_por_id'] != request.user.id:
                return JsonResponse({'error': 'Solo el creador puede editar la tarea'}, status=403)
            
            # Obtener el nombre del responsable
            responsable_nombre = tarea_encontrada['responsable']  # Mantener actual por defecto
            if responsable_id:
                try:
                    responsable_user = User.objects.get(id=responsable_id)
                    responsable_nombre = responsable_user.username
                except User.DoesNotExist:
                    pass  # Mantener el valor actual
            
            # Actualizar los campos de la tarea
            tarea_encontrada.update({
                'titulo': titulo,
                'descripcion': descripcion,
                'prioridad': prioridad,
                'fecha_limite': fecha_limite,
                'responsable': responsable_nombre,
                'responsable_id': responsable_id or tarea_encontrada['responsable_id'],
            })
            
            return JsonResponse({
                'success': True,
                'message': 'Tarea actualizada exitosamente',
                'tarea': tarea_encontrada
            })
            
        except Exception as e:
            return JsonResponse({
                'error': f'Error al actualizar la tarea: {str(e)}'
            }, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required  
def api_estadisticas_tareas_proyectos(request):
    """
    API para obtener estadísticas de tareas y proyectos
    """
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    # Por ahora, devolver estadísticas de ejemplo
    estadisticas = {
        'proyectos_activos': 2,
        'tareas_pendientes': 3,
        'tareas_completadas': 12,
        'progreso_general': 68
    }
    
    return JsonResponse({
        'success': True,
        **estadisticas
    })

@login_required
def api_buscar_usuarios(request):
    """
    API para buscar usuarios para agregar como miembros del proyecto
    """
    print(f"🔍 api_buscar_usuarios - Usuario: {request.user.username}, Autenticado: {request.user.is_authenticated}")
    
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Usuario no autenticado'}, status=401)
    
    search_query = request.GET.get('q', '').strip()
    
    from django.contrib.auth.models import User
    
    if search_query and len(search_query) >= 2:
        # Buscar usuarios por nombre, apellido o username
        usuarios = User.objects.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(username__icontains=search_query)
        )[:100]  # Aumentamos el límite y permitimos mostrar al usuario actual
    else:
        # Si no hay query, devolver todos los usuarios (para mostrar lista completa)
        usuarios = User.objects.all().order_by('first_name', 'last_name', 'username')[:100]
    
    usuarios_data = []
    for usuario in usuarios:
        # Generar iniciales
        if usuario.first_name and usuario.last_name:
            iniciales = f"{usuario.first_name[0]}{usuario.last_name[0]}".upper()
            nombre_completo = f"{usuario.first_name} {usuario.last_name}"
        elif usuario.first_name:
            iniciales = usuario.first_name[0].upper()
            nombre_completo = usuario.first_name
        else:
            iniciales = usuario.username[0].upper() if usuario.username else "?"
            nombre_completo = usuario.username
        
        # Determinar rol/cargo
        if usuario.is_superuser:
            rol = "Administrador"
        elif hasattr(usuario, 'groups') and usuario.groups.filter(name='Supervisores').exists():
            rol = "Supervisor"
        else:
            rol = "Usuario"
        
        # Obtener avatar_url del usuario
        avatar_url = None
        if hasattr(usuario, 'userprofile'):
            try:
                avatar_url = usuario.userprofile.get_avatar_url()
                print(f"DEBUG: Avatar URL para {usuario.username}: {avatar_url}")
            except Exception as e:
                print(f"DEBUG: Error obteniendo avatar para {usuario.username}: {e}")
                avatar_url = None
        else:
            print(f"DEBUG: Usuario {usuario.username} no tiene UserProfile")
        
        usuarios_data.append({
            'id': usuario.id,
            'nombre': nombre_completo,
            'username': usuario.username,
            'iniciales': iniciales,
            'rol': rol,
            'email': usuario.email or '',
            'avatar_url': avatar_url
        })
    
    return JsonResponse({
        'success': True,
        'usuarios': usuarios_data
    })

@login_required
def api_crear_proyecto(request):
    """
    API para crear un nuevo proyecto con miembros y enviar notificaciones
    """
    print(f"🔍 Usuario: {request.user.username}, Autenticado: {request.user.is_authenticated}, Superusuario: {request.user.is_superuser}")
    
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Usuario no autenticado'}, status=401)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        
        # Validar datos requeridos
        nombre = data.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'El nombre del proyecto es requerido'}, status=400)
        
        descripcion = data.get('descripcion', '').strip()
        privacidad = data.get('privacidad', 'publico')
        tipo = data.get('tipo', 'runrate')
        miembros_ids = data.get('miembros', [])
        
        # Crear el proyecto real en la base de datos
        proyecto = Proyecto.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            tipo=tipo,
            privacidad=privacidad,
            creado_por=request.user
        )
        
        # Crear comentario inicial con la descripción del proyecto (si existe)
        if descripcion.strip():
            ProyectoComentario.objects.create(
                proyecto=proyecto,
                usuario=request.user,
                contenido=f"📝 Proyecto creado: {descripcion}"
            )
        
        # Agregar miembros al proyecto y enviar notificaciones
        from django.contrib.auth.models import User
        miembros_notificados = []
        
        for miembro_id in miembros_ids:
            try:
                usuario = User.objects.get(id=miembro_id)
                # Agregar como miembro del proyecto
                proyecto.miembros.add(usuario)
                
                # Enviar notificación
                notificacion = notificar_miembro_agregado_proyecto(
                    usuario_agregado=usuario,
                    proyecto_nombre=nombre,
                    proyecto_id=proyecto.id,
                    usuario_que_agrega=request.user
                )
                if notificacion:
                    miembros_notificados.append({
                        'id': usuario.id,
                        'nombre': usuario.get_full_name() or usuario.username,
                        'notificado': True
                    })
            except User.DoesNotExist:
                continue
        
        print(f"✅ Proyecto '{nombre}' creado. Notificaciones enviadas a {len(miembros_notificados)} miembros.")
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Proyecto creado exitosamente',
            'proyecto': {
                'id': proyecto.id,
                'nombre': proyecto.nombre,
                'descripcion': proyecto.descripcion,
                'privacidad': proyecto.privacidad,
                'tipo': proyecto.tipo,
                'miembros_notificados': miembros_notificados,
                'creado_por': request.user.get_full_name() or request.user.username,
                'fecha_creacion': proyecto.fecha_creacion.strftime('%Y-%m-%d')
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        print(f"❌ Error creando proyecto: {e}")
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)


@login_required
def api_crear_tarea(request):
    """
    API para crear una nueva tarea independiente (sin proyecto)
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Usuario no autenticado'}, status=401)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        
        # Validar datos requeridos
        nombre = data.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'El nombre de la tarea es requerido'}, status=400)
        
        descripcion = data.get('descripcion', '').strip()
        prioridad = data.get('prioridad', 'media')
        fecha_limite = data.get('fecha_limite')
        estimacion_horas = data.get('estimacion_horas')
        asignado_a_id = data.get('asignado_a')
        proyecto_id = data.get('proyecto_id')
        oportunidad_id = data.get('oportunidad_id')
        participantes_ids = data.get('participantes', [])
        observadores_ids = data.get('observadores', [])
        
        # Convertir fecha límite si está presente
        fecha_limite_obj = None
        
        if fecha_limite:
            from datetime import datetime
            try:
                # Intentar formato con hora primero (YYYY-MM-DDTHH:MM)
                fecha_limite_obj = datetime.strptime(fecha_limite, '%Y-%m-%dT%H:%M')
            except ValueError:
                try:
                    # Si falla, intentar solo fecha (YYYY-MM-DD) y agregar hora por defecto
                    fecha_limite_obj = datetime.strptime(fecha_limite + 'T23:59', '%Y-%m-%dT%H:%M')
                except ValueError:
                    print(f"⚠️ Formato de fecha_limite no válido: {fecha_limite}")
                    fecha_limite_obj = None
        
        # Obtener usuario asignado si se especificó
        asignado_a = None
        if asignado_a_id:
            try:
                from django.contrib.auth.models import User
                asignado_a = User.objects.get(id=asignado_a_id)
            except User.DoesNotExist:
                return JsonResponse({'error': 'Usuario asignado no encontrado'}, status=400)
        
        # Obtener proyecto si se especificó
        proyecto = None
        if proyecto_id:
            try:
                proyecto = Proyecto.objects.get(id=proyecto_id)
            except Proyecto.DoesNotExist:
                return JsonResponse({'error': 'Proyecto no encontrado'}, status=400)

        # Obtener oportunidad si se especificó
        oportunidad = None
        if oportunidad_id:
            try:
                oportunidad = TodoItem.objects.get(id=oportunidad_id)
            except TodoItem.DoesNotExist:
                return JsonResponse({'error': 'Oportunidad no encontrada'}, status=400)

        # Crear la tarea en la base de datos
        tarea = Tarea.objects.create(
            titulo=nombre,
            descripcion=descripcion,
            prioridad=prioridad,
            estado='pendiente',
            creado_por=request.user,
            asignado_a=asignado_a,
            fecha_limite=fecha_limite_obj,
            proyecto=proyecto,
            oportunidad=oportunidad,
        )
        
        # Agregar participantes y observadores
        from django.contrib.auth.models import User as AuthUser
        for pid in participantes_ids:
            try:
                tarea.participantes.add(AuthUser.objects.get(id=pid))
            except Exception:
                pass
        for oid in observadores_ids:
            try:
                tarea.observadores.add(AuthUser.objects.get(id=oid))
            except Exception:
                pass

        # Crear comentario inicial si hay descripción
        if descripcion.strip():
            TareaComentario.objects.create(
                tarea=tarea,
                usuario=request.user,
                contenido=f"📝 Tarea creada: {descripcion}"
            )
        
        # Enviar notificación al usuario asignado si es diferente al creador
        if asignado_a and asignado_a != request.user:
            try:
                crear_notificacion(
                    usuario_destinatario=asignado_a,
                    tipo='tarea_asignada',
                    titulo='Nueva tarea asignada',
                    mensaje=f'Se te ha asignado la tarea "{nombre}"',
                    usuario_remitente=request.user
                )
            except Exception as e:
                print(f"⚠️ Error enviando notificación: {e}")
        
        return JsonResponse({
            'success': True,
            'message': 'Tarea creada exitosamente',
            'tarea': {
                'id': tarea.id,
                'titulo': tarea.titulo,
                'descripcion': tarea.descripcion,
                'prioridad': tarea.prioridad,
                'estado': tarea.estado,
                'creado_por': request.user.get_full_name() or request.user.username,
                'asignado_a': asignado_a.get_full_name() or asignado_a.username if asignado_a else None,
                'fecha_creacion': tarea.fecha_creacion.strftime('%Y-%m-%d'),
                'fecha_limite': tarea.fecha_limite.strftime('%Y-%m-%d') if tarea.fecha_limite else None
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)


@login_required
@login_required
def proyectos_ingenieria(request):
    """
    Lista de proyectos de ingeniería importados de Bitrix24.
    Accesible para ingenieros y supervisores.
    """
    profile = getattr(request.user, 'userprofile', None)
    es_ing = profile and getattr(profile, 'rol', '') == 'ingeniero'
    if not (is_supervisor(request.user) or es_ing):
        return redirect('home')

    proyectos_qs = Proyecto.objects.filter(es_ingenieria=True).prefetch_related(
        'miembros', 'miembros__userprofile'
    ).order_by('-fecha_creacion')

    proyectos_data = []
    for p in proyectos_qs:
        miembros = []
        for m in p.miembros.all()[:5]:
            iniciales = ''.join(
                w[0].upper() for w in (m.get_full_name() or m.username).split()[:2]
            )
            avatar_url = None
            if hasattr(m, 'userprofile'):
                try:
                    avatar_url = m.userprofile.get_avatar_url()
                except Exception:
                    pass
            miembros.append({
                'nombre': m.get_full_name() or m.username,
                'iniciales': iniciales,
                'avatar_url': avatar_url,
            })
        proyectos_data.append({
            'id': p.id,
            'nombre': p.nombre,
            'fecha_creacion': p.fecha_creacion,
            'miembros': miembros,
            'total_miembros': p.miembros.count(),
        })

    return render(request, 'proyectos_ingenieria.html', {
        'proyectos': proyectos_data,
        'total': len(proyectos_data),
    })


def proyecto_detalle(request, proyecto_id):
    """
    Vista para el detalle individual de un proyecto
    """
    try:
        # Obtener el proyecto real de la base de datos
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        
        # Verificar permisos y establecer nivel de acceso
        tiene_acceso = True
        es_miembro = False
        puede_solicitar_acceso = False
        
        if proyecto.privacidad == 'privado':
            # Verificar si es miembro o creador
            es_miembro = (request.user == proyecto.creado_por or 
                         request.user in proyecto.miembros.all())
            
            if not es_miembro:
                tiene_acceso = False
                # Verificar si ya envió una solicitud pendiente (con manejo de errores)
                try:
                    from app.models import SolicitudAccesoProyecto
                    solicitud_pendiente = SolicitudAccesoProyecto.objects.filter(
                        proyecto=proyecto,
                        usuario_solicitante=request.user,
                        estado='pendiente'
                    ).exists()
                    puede_solicitar_acceso = not solicitud_pendiente
                except Exception:
                    # Si la tabla no existe, permitir solicitar acceso
                    puede_solicitar_acceso = True
        
        # Crear datos del proyecto para el template
        proyecto_data = {
            'id': proyecto.id,
            'nombre': proyecto.nombre,
            'descripcion': proyecto.descripcion,
            'tipo': proyecto.tipo,
            'privacidad': proyecto.privacidad,
            'fecha_creacion': proyecto.fecha_creacion.strftime('%d de %b, %Y'),
            'creado_por': proyecto.creado_por,
            'miembros': proyecto.get_miembros_display(),
            'mi_rol': proyecto.get_rol_usuario(request.user)
        }
        
        # Obtener usuarios disponibles para asignar como responsables y miembros
        from django.contrib.auth.models import User
        usuarios_disponibles = User.objects.filter(is_active=True).order_by('first_name', 'username')

        context = {
            'proyecto': proyecto,  # Pasar objeto completo para acceder a métodos
            'user': request.user,
            'page_title': f'Proyecto: {proyecto.nombre}',
            'usuarios_disponibles': usuarios_disponibles,
            'tiene_acceso': tiene_acceso,
            'es_miembro': es_miembro,
            'puede_solicitar_acceso': puede_solicitar_acceso
        }
        
        return render(request, 'proyecto_detalle.html', context)
        
    except Proyecto.DoesNotExist:
        return redirect('tareas_proyectos')


@login_required
def api_comentarios_proyecto(request, proyecto_id):
    """
    API para obtener comentarios de un proyecto
    """
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        
        # Verificar permisos
        if proyecto.privacidad == 'privado':
            if request.user != proyecto.creado_por and request.user not in proyecto.miembros.all():
                return JsonResponse({'error': 'Sin permisos'}, status=403)
        
        # Filtrar comentarios excluyendo los automáticos de configuración
        # PERO manteniendo comentarios de creación del proyecto y manuales
        comentarios = ProyectoComentario.objects.filter(
            proyecto=proyecto
        ).exclude(
            # Excluir comentarios que empiecen con patrones de configuración automática
            Q(contenido__startswith='📋 Configuración actualizada') |
            Q(contenido__startswith='⚙️ Configuración') |
            Q(contenido__startswith='🔧 Configuración') |
            Q(contenido__startswith='📊 Configuración') |
            Q(contenido__startswith='Configuración actualizada') |
            Q(contenido__startswith='Proyecto actualizado') |
            Q(contenido__startswith='Configuración modificada') |
            Q(contenido__contains='Miembro agregado al proyecto') |
            Q(contenido__contains='Miembro removido del proyecto') |
            Q(contenido__contains='Oportunidad ligada al proyecto') |
            Q(contenido__contains='Oportunidad removida del proyecto')
        ).order_by('-fecha_creacion')
        
        comentarios_data = []
        for comentario in comentarios:
            archivos_data = []
            for archivo in comentario.archivos.all():
                archivos_data.append({
                    'id': archivo.id,
                    'nombre': archivo.nombre_original,
                    'tamaño': archivo.get_tamaño_legible(),
                    'icono': archivo.get_icono(),
                    'url': archivo.archivo.url,
                    'tipo': archivo.tipo_contenido
                })
            
            comentarios_data.append({
                'id': comentario.id,
                'contenido': getattr(comentario, 'get_contenido_con_menciones', lambda: comentario.contenido)(),
                'contenido_raw': comentario.contenido,
                'usuario': {
                    'id': comentario.usuario.id,
                    'nombre': comentario.usuario.get_full_name() or comentario.usuario.username,
                    'username': comentario.usuario.username,
                    'iniciales': ''.join([palabra[0].upper() for palabra in (comentario.usuario.get_full_name() or comentario.usuario.username).split()[:2]]),
                    'avatar_url': getattr(comentario.usuario.userprofile, 'get_avatar_url', lambda: None)() if hasattr(comentario.usuario, 'userprofile') else None
                },
                'fecha': comentario.fecha_creacion.strftime('%d de %b, %Y - %H:%M'),
                'fecha_edicion': comentario.fecha_edicion.strftime('%d de %b, %Y - %H:%M') if comentario.fecha_edicion else None,
                'editado': comentario.editado,
                'archivos': archivos_data,
                'puede_editar': comentario.usuario == request.user,
                'puede_eliminar': comentario.usuario == request.user or proyecto.creado_por == request.user
            })
        
        return JsonResponse({
            'success': True,
            'comentarios': comentarios_data
        })
        
    except Exception as e:
        print(f"Error obteniendo comentarios: {e}")
        return JsonResponse({'error': 'Error interno'}, status=500)




def get_safe_file_info(archivo):
    """Helper function para obtener información segura de archivos"""
    try:
        tamaño = archivo.get_tamaño_legible() if hasattr(archivo, 'get_tamaño_legible') else f"{archivo.tamaño} bytes"
    except:
        tamaño = f"{archivo.tamaño} bytes"
    
    try:
        icono = archivo.get_icono() if hasattr(archivo, 'get_icono') else '📎'
    except:
        icono = '📎'
    
    return {
        'id': archivo.id,
        'nombre': archivo.nombre_original,
        'tamaño': tamaño,
        'icono': icono,
        'url': archivo.archivo.url,
        'tipo': archivo.tipo_contenido
    }



@login_required
def api_agregar_comentario_proyecto(request, proyecto_id):
    """
    API para agregar un comentario a un proyecto
    """
    print(f"🔍 api_agregar_comentario_proyecto - Usuario: {request.user.username}, Proyecto ID: {proyecto_id}")
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        print(f"✅ Proyecto encontrado: {proyecto.nombre}")
        
        # Verificar permisos
        if proyecto.privacidad == 'privado':
            if request.user != proyecto.creado_por and request.user not in proyecto.miembros.all():
                return JsonResponse({'error': 'Sin permisos'}, status=403)
        
        contenido = request.POST.get('contenido', '').strip()
        if not contenido:
            return JsonResponse({'error': 'El comentario no puede estar vacío'}, status=400)
        
        # Crear el comentario
        comentario = ProyectoComentario.objects.create(
            proyecto=proyecto,
            usuario=request.user,
            contenido=contenido
        )
        
        # Procesar archivos adjuntos
        archivos_data = []
        for key, file in request.FILES.items():
            if key.startswith('archivo_'):
                archivo = ProyectoArchivo.objects.create(
                    comentario=comentario,
                    archivo=file,
                    nombre_original=file.name,
                    tamaño=file.size,
                    tipo_contenido=file.content_type or ''
                )
                archivos_data.append({
                    'id': archivo.id,
                    'nombre': archivo.nombre_original,
                    'tamaño': archivo.get_tamaño_legible(),
                    'icono': archivo.get_icono(),
                    'url': archivo.archivo.url,
                    'tipo': archivo.tipo_contenido
                })
        
        # Enviar notificaciones a usuarios mencionados
        usuarios_mencionados = comentario.extraer_menciones()
        for usuario_mencionado in usuarios_mencionados:
            if usuario_mencionado != request.user:  # No notificar al autor
                crear_notificacion(
                    usuario_destinatario=usuario_mencionado,
                    tipo='proyecto_agregado',
                    titulo=f"Te mencionaron en {proyecto.nombre}",
                    mensaje=f"{request.user.get_full_name() or request.user.username} te mencionó en el proyecto '{proyecto.nombre}': {contenido[:100]}...",
                    usuario_remitente=request.user,
                    proyecto_id=proyecto.id,
                    proyecto_nombre=proyecto.nombre
                )
        
        return JsonResponse({
            'success': True,
            'comentario': {
                'id': comentario.id,
                'contenido': getattr(comentario, 'get_contenido_con_menciones', lambda: comentario.contenido)(),
                'contenido_raw': comentario.contenido,
                'usuario': {
                    'id': comentario.usuario.id,
                    'nombre': comentario.usuario.get_full_name() or comentario.usuario.username,
                    'username': comentario.usuario.username,
                    'iniciales': ''.join([palabra[0].upper() for palabra in (comentario.usuario.get_full_name() or comentario.usuario.username).split()[:2]])
                },
                'fecha': comentario.fecha_creacion.strftime('%d de %b, %Y - %H:%M'),
                'fecha_edicion': None,
                'editado': False,
                'archivos': archivos_data,
                'puede_editar': True,
                'puede_eliminar': True
            }
        })
        
    except Exception as e:
        print(f"Error agregando comentario: {e}")
        return JsonResponse({'error': 'Error interno'}, status=500)




@login_required
def api_editar_comentario_proyecto(request, comentario_id):
    """
    API para editar un comentario de proyecto
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        comentario = get_object_or_404(ProyectoComentario, id=comentario_id)
        
        # Verificar permisos
        if comentario.usuario != request.user:
            return JsonResponse({'error': 'Sin permisos'}, status=403)
        
        contenido = request.POST.get('contenido', '').strip()
        if not contenido:
            return JsonResponse({'error': 'El comentario no puede estar vacío'}, status=400)
        
        comentario.contenido = contenido
        comentario.save()
        
        return JsonResponse({
            'success': True,
            'contenido': comentario.get_contenido_con_menciones(),
            'fecha_edicion': comentario.fecha_edicion.strftime('%d de %b, %Y - %H:%M')
        })
        
    except Exception as e:
        print(f"Error editando comentario: {e}")
        return JsonResponse({'error': 'Error interno'}, status=500)




@login_required
def api_eliminar_comentario_proyecto(request, comentario_id):
    """
    API para eliminar un comentario de proyecto
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        comentario = get_object_or_404(ProyectoComentario, id=comentario_id)
        
        # Verificar permisos (autor del comentario o creador del proyecto)
        if comentario.usuario != request.user and comentario.proyecto.creado_por != request.user:
            return JsonResponse({'error': 'Sin permisos'}, status=403)
        
        comentario.delete()
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        print(f"Error eliminando comentario: {e}")
        return JsonResponse({'error': 'Error interno'}, status=500)


@login_required
def api_eliminar_comentario_proyecto(request, comentario_id):
    """
    API para eliminar un comentario de proyecto
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        comentario = get_object_or_404(ProyectoComentario, id=comentario_id)
        
        # Verificar permisos (autor del comentario o creador del proyecto)
        if comentario.usuario != request.user and comentario.proyecto.creado_por != request.user:
            return JsonResponse({'error': 'Sin permisos'}, status=403)
        
        comentario.delete()
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        print(f"Error eliminando comentario: {e}")
        return JsonResponse({'error': 'Error interno'}, status=500)


@login_required
def api_configuracion_proyecto(request, proyecto_id):
    """
    API para actualizar la configuración completa de un proyecto
    """
    print(f"🔍 api_configuracion_proyecto - Usuario: {request.user.username}, Proyecto ID: {proyecto_id}")
    print(f"🔍 Método de request: {request.method}")
    print(f"🔍 Content-Type: {request.content_type}")
    print(f"🔍 Request body: {request.body}")
    
    if request.method != 'PUT':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        print(f"🔍 Proyecto encontrado: {proyecto.nombre}")
        
        # Verificar permisos - solo el creador puede modificar la configuración
        if proyecto.creado_por != request.user:
            print(f"❌ Sin permisos: proyecto creado por {proyecto.creado_por}, usuario actual {request.user}")
            return JsonResponse({'error': 'Sin permisos para modificar este proyecto'}, status=403)
        
        import json
        data = json.loads(request.body)
        print(f"🔍 Datos recibidos: {data}")
        
        # Actualizar información básica
        if 'nombre' in data:
            nombre = data['nombre'].strip()
            if not nombre:
                return JsonResponse({'error': 'El nombre del proyecto es requerido'}, status=400)
            proyecto.nombre = nombre
        
        if 'descripcion' in data:
            proyecto.descripcion = data['descripcion'].strip()
        
        if 'tipo' in data:
            if data['tipo'] in ['runrate', 'ingenieria']:
                proyecto.tipo = data['tipo']
        
        if 'privacidad' in data:
            if data['privacidad'] in ['publico', 'privado']:
                proyecto.privacidad = data['privacidad']
        
        # El campo responsable no existe en el modelo Proyecto
        # Se usa creado_por como responsable del proyecto
        
        # Actualizar miembros
        if 'miembros_ids' in data:
            from django.contrib.auth.models import User
            miembros_ids = data['miembros_ids']
            
            # Limpiar miembros actuales
            proyecto.miembros.clear()
            
            # Agregar nuevos miembros
            for miembro_id in miembros_ids:
                try:
                    usuario = User.objects.get(id=miembro_id)
                    proyecto.miembros.add(usuario)
                    
                    # Enviar notificación si es un miembro nuevo (con try/catch para evitar errores)
                    try:
                        notificar_miembro_agregado_proyecto(
                            usuario_agregado=usuario,
                            proyecto_nombre=proyecto.nombre,
                            proyecto_id=proyecto.id,
                            usuario_que_agrega=request.user
                        )
                    except Exception as notif_error:
                        print(f"⚠️ Error enviando notificación: {notif_error}")
                        # Continuar sin fallar el guardado por error de notificación
                        
                except User.DoesNotExist:
                    continue
        
        # Actualizar oportunidad ligada
        # Manejar oportunidades ligadas
        if 'oportunidades_ids' in data:
            oportunidades_ids = data['oportunidades_ids']
            print(f"🔗 Oportunidades a ligar: {oportunidades_ids}")
            
            # Limpiar oportunidades actuales
            proyecto.oportunidades_ligadas.clear()
            
            # Agregar nuevas oportunidades
            if oportunidades_ids:
                for oportunidad_id in oportunidades_ids:
                    try:
                        oportunidad = TodoItem.objects.get(id=oportunidad_id)
                        proyecto.oportunidades_ligadas.add(oportunidad)
                        print(f"✅ Oportunidad ligada: {oportunidad.oportunidad}")
                    except TodoItem.DoesNotExist:
                        print(f"❌ Oportunidad con ID {oportunidad_id} no encontrada")
                        return JsonResponse({'error': f'Oportunidad con ID {oportunidad_id} no encontrada'}, status=400)
            else:
                print("✅ Todas las oportunidades desligadas")
        
        # Guardar cambios
        proyecto.save()
        
        # Comentarios del sistema temporalmente deshabilitados para evitar errores
        # TODO: Reactivar cuando se resuelvan los problemas de modelo ProyectoComentario
        print(f"📝 Cambios realizados en proyecto '{proyecto.nombre}': {list(data.keys())}")
        
        print(f"✅ Configuración del proyecto '{proyecto.nombre}' actualizada exitosamente")
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Configuración actualizada exitosamente',
            'proyecto': {
                'id': proyecto.id,
                'nombre': proyecto.nombre,
                'descripcion': proyecto.descripcion,
                'tipo': proyecto.tipo,
                'privacidad': proyecto.privacidad,
                'creado_por': proyecto.creado_por.get_full_name() or proyecto.creado_por.username,
                'miembros_count': proyecto.miembros.count(),
                'miembros': list(proyecto.miembros.values_list('id', flat=True)),
                'oportunidades_ligadas': [
                    {
                        'id': oportunidad.id,
                        'titulo': oportunidad.oportunidad,
                        'cliente': oportunidad.cliente.nombre_empresa if oportunidad.cliente else 'Sin cliente'
                    }
                    for oportunidad in proyecto.oportunidades_ligadas.all()
                ]
            }
        })
        
    except json.JSONDecodeError as e:
        print(f"❌ Error de JSON: {e}")
        return JsonResponse({'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        import traceback
        print(f"❌ Error actualizando configuración: {e}")
        print(f"❌ Traceback completo: {traceback.format_exc()}")
        return JsonResponse({'error': f'Error interno del servidor: {str(e)}'}, status=500)


@login_required
def api_buscar_oportunidades_proyecto(request):
    """
    API para buscar oportunidades para vincular con un proyecto
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Usuario no autenticado'}, status=401)
    
    query = request.GET.get('q', '').strip()
    
    try:
        if not query or len(query) < 2:
            # Si no hay query, devolver las últimas 5 oportunidades
            if is_supervisor(request.user):
                oportunidades = TodoItem.objects.select_related('cliente').order_by('-fecha_creacion')[:5]
            else:
                # Usuario regular solo ve sus oportunidades
                oportunidades = TodoItem.objects.filter(
                    usuario=request.user
                ).select_related('cliente').order_by('-fecha_creacion')[:5]
        else:
            # Buscar oportunidades por título, cliente o comentarios
            if is_supervisor(request.user):
                oportunidades = TodoItem.objects.filter(
                    Q(oportunidad__icontains=query) |
                    Q(cliente__nombre_empresa__icontains=query) |
                    Q(comentarios__icontains=query)
                ).select_related('cliente').order_by('-fecha_creacion')[:10]
            else:
                # Usuario regular solo ve sus oportunidades
                oportunidades = TodoItem.objects.filter(
                    usuario=request.user
                ).filter(
                    Q(oportunidad__icontains=query) |
                    Q(cliente__nombre_empresa__icontains=query) |
                    Q(comentarios__icontains=query)
                ).select_related('cliente').order_by('-fecha_creacion')[:10]
        
        oportunidades_data = []
        for oportunidad in oportunidades:
            oportunidades_data.append({
                'id': oportunidad.id,
                'titulo': oportunidad.oportunidad,
                'cliente__nombre_empresa': oportunidad.cliente.nombre_empresa if oportunidad.cliente else 'Cliente no asignado',
                'monto': float(oportunidad.monto) if oportunidad.monto else 0,
                'probabilidad': oportunidad.probabilidad_cierre,
                'mes_cierre': oportunidad.get_mes_cierre_display(),
                'usuario': oportunidad.usuario.get_full_name() or oportunidad.usuario.username
            })
        
        return JsonResponse({
            'success': True,
            'oportunidades': oportunidades_data
        })
        
    except Exception as e:
        print(f"❌ Error buscando oportunidades: {e}")
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)




@login_required
def dashboard(request):
    # Base queryset for user-specific or all opportunities
    if is_supervisor(request.user):
        base_opportunities = TodoItem.objects.all()
    else:
        base_opportunities = TodoItem.objects.filter(usuario=request.user)

    # --- Key Metrics ---

    # 1. Top Client by total opportunity amount
    top_cliente = base_opportunities.filter(cliente__isnull=False).values(
        'cliente__id', 'cliente__nombre_empresa'
    ).annotate(
        total_monto=Sum('monto')
    ).order_by('-total_monto').first()

    # 2. Top Brand (Product) by total opportunity amount
    top_marca_data = base_opportunities.values('producto').annotate(
        total_monto=Sum('monto')
    ).order_by('-total_monto').first()
    
    top_marca = None
    if top_marca_data and top_marca_data['producto']:
        top_marca = {
            'nombre_marca': _get_display_for_value(top_marca_data['producto'], TodoItem.PRODUCTO_CHOICES),
            'total_monto': top_marca_data['total_monto'],
            'marca': top_marca_data['producto']
        }

    # --- Current Month ---
    today = date.today()
    mes_actual_val = str(today.month).zfill(2)
    mes_actual_nombre = dict(TodoItem.MES_CHOICES).get(mes_actual_val, f"Mes {today.month}")
    
    # Filtrar por fecha de creación del mes actual
    oportunidades_mes_actual = base_opportunities.filter(
        fecha_creacion__year=today.year,
        fecha_creacion__month=today.month
    )
    monto_total_mes_actual = oportunidades_mes_actual.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    META_MENSUAL = Decimal('350000.00') if is_supervisor(request.user) else Decimal('130000.00')
    porcentaje_cobertura_mes_actual = min(100, int((monto_total_mes_actual / META_MENSUAL * 100) if META_MENSUAL > 0 else 0))

    # --- Next Month ---
    next_month_date = today + relativedelta(months=1)
    next_month_value = str(next_month_date.month).zfill(2)
    next_month_display = dict(TodoItem.MES_CHOICES).get(next_month_value, f"Mes {next_month_date.month}")
    
    # Filtrar por mes de cierre del próximo mes (búsqueda robusta)
    oportunidades_proximo_mes = base_opportunities.filter(
        Q(mes_cierre=next_month_value) |  # Formato con ceros (11)
        Q(mes_cierre=str(next_month_date.month)) |  # Formato sin ceros (11)
        Q(mes_cierre=next_month_display.upper()) |  # Nombre del mes
        Q(mes_cierre=next_month_display.lower())   # Nombre del mes en minúscula
    )
    total_oportunidades_proximo_mes = oportunidades_proximo_mes.count()
    total_monto_esperado_proximo_mes = oportunidades_proximo_mes.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')

    # --- Chart Data: Opportunities per month by creation date ---
    ano_actual = today.year
    ventas_por_mes = base_opportunities.filter(
        fecha_creacion__year=ano_actual
    ).extra(
        select={'mes_creacion': 'EXTRACT(month FROM fecha_creacion)'}
    ).values('mes_creacion').annotate(
        total=Sum('monto')
    ).order_by('mes_creacion')
    
    ventas_por_mes_dict = {int(v['mes_creacion']): float(v['total'] or 0) for v in ventas_por_mes}
    ventas_por_mes_list = []
    for i in range(1, 13):
        ventas_por_mes_list.append({'mes': f'{ano_actual}-{str(i).zfill(2)}', 'total': ventas_por_mes_dict.get(i, 0)})

    # --- Additional Statistics ---
    total_oportunidades_count = base_opportunities.count()
    total_monto_general = base_opportunities.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    promedio_oportunidad = total_monto_general / total_oportunidades_count if total_oportunidades_count > 0 else Decimal('0.00')
    oportunidades_ano_actual = base_opportunities.filter(fecha_creacion__year=ano_actual).count()

    context = {
        'is_supervisor': is_supervisor(request.user),
        
        # Key Widgets
        'top_cliente': top_cliente,
        'top_marca': top_marca,
        
        # Current Month
        'mes_actual_nombre': mes_actual_nombre,
        'monto_total_mes_actual': monto_total_mes_actual,
        'porcentaje_cobertura_mes_actual': porcentaje_cobertura_mes_actual,
        'mes_actual_val': mes_actual_val,

        # Next Month
        'next_month_display': next_month_display,
        'proximo_mes_val': next_month_value,
        'total_oportunidades_proximo_mes': total_oportunidades_proximo_mes,
        'total_monto_esperado_proximo_mes': total_monto_esperado_proximo_mes,
        
        # Chart
        'ventas_por_mes_list': ventas_por_mes_list,
        
        # Additional Statistics
        'total_oportunidades_count': total_oportunidades_count,
        'promedio_oportunidad': promedio_oportunidad,
        'oportunidades_ano_actual': oportunidades_ano_actual,
    }
    return render(request, "dashboard.html", context)


@login_required
def get_user_clients_api(request):
    """
    Vista API que devuelve los clientes asignados al usuario autenticado.
    Si el usuario es supervisor, devuelve todos los clientes.
    """
    try:
        if is_supervisor(request.user):
            # Si el usuario es supervisor, obtener todos los clientes
            clients_queryset = Cliente.objects.all()
            print("DEBUG: Usuario es supervisor. Obteniendo todos los clientes.")
        else:
            # Si no es supervisor, obtener solo los clientes asignados a este usuario
            # Usamos 'asignado_a' que es el campo correcto en tu modelo Cliente
            clients_queryset = Cliente.objects.filter(asignado_a=request.user)
            print(f"DEBUG: Usuario {request.user.username} es vendedor. Obteniendo sus clientes.")

        # Serializar los clientes a un formato que pueda ser convertido a JSON
        clients_data = []
        for client in clients_queryset:
            clients_data.append({
                'id': client.id, # Convertir ID a número para consistencia
                'name': client.nombre_empresa, # Mapear nombre_empresa a 'name'
                'nombre_empresa': client.nombre_empresa, # Mantener también el nombre original
                'address': client.direccion, # Mapear direccion a 'address'
                'taxId': client.email, # Mapear email a 'taxId' (o el campo que uses para ID Fiscal)
                'categoria': client.categoria, # Categoría del cliente (A, B, C)
                'porcentaje_utilidad': 15 if client.categoria == 'A' else 20 if client.categoria == 'B' else 25 # Porcentaje de utilidad según categoría
                # Puedes añadir más campos aquí si los necesitas en el frontend
            })
        return JsonResponse(clients_data, safe=False) # safe=False permite serializar listas directamente
    except Exception as e:
        print(f"ERROR en get_user_clients_api: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def view_cotizacion_pdf(request, cotizacion_id):
    """
    Vista para generar y mostrar el PDF de una cotización específica en el navegador.
    """
    print(f"DEBUG: Iniciando view_cotizacion_pdf para la cotización ID: {cotizacion_id}")
    cotizacion = get_object_or_404(Cotizacion, pk=cotizacion_id)
    
    if not is_supervisor(request.user) and cotizacion.created_by != request.user:
        return HttpResponse("Acceso denegado.", status=403)

    detalles_cotizacion = DetalleCotizacion.objects.filter(cotizacion=cotizacion).order_by('orden')
    iva_rate_percentage = (cotizacion.iva_rate * Decimal('100')).quantize(Decimal('1'))
    
    # Organizar productos en secciones basadas en títulos
    secciones = []
    seccion_actual = {'titulo': None, 'productos': []}
    
    for detalle in detalles_cotizacion:
        # Manejar productos existentes que no tienen el campo tipo definido
        tipo_detalle = getattr(detalle, 'tipo', 'producto') or 'producto'
        
        if tipo_detalle == 'titulo':
            # Si hay productos en la sección actual, guardarla ANTES de crear nueva sección
            if seccion_actual['productos']:
                secciones.append(seccion_actual)
            # Si hay una sección actual con título pero sin productos, también la guardamos
            elif seccion_actual['titulo']:
                secciones.append(seccion_actual)
                
            # Iniciar nueva sección
            seccion_actual = {'titulo': detalle.nombre_producto, 'productos': []}
        else:
            # Agregar producto a la sección actual
            seccion_actual['productos'].append(detalle)
    
    # Agregar la última sección (con o sin productos)
    if seccion_actual['titulo'] or seccion_actual['productos']:
        secciones.append(seccion_actual)
    
    # Si no hay títulos, crear una sección por defecto
    if not secciones:
        productos_sin_seccion = [d for d in detalles_cotizacion if getattr(d, 'tipo', 'producto') == 'producto']
        if productos_sin_seccion:
            secciones.append({'titulo': None, 'productos': productos_sin_seccion})
    
    print(f"DEBUG: Secciones organizadas en view_cotizacion_pdf: {len(secciones)} secciones encontradas")
    for i, seccion in enumerate(secciones):
        print(f"DEBUG: Sección {i+1}: titulo='{seccion['titulo']}', productos={len(seccion['productos'])}")
        for j, producto in enumerate(seccion['productos']):
            print(f"  Producto {j+1}: {producto.nombre_producto} (tipo: {getattr(producto, 'tipo', 'NO_DEFINIDO')})")

    pdf_name_raw = cotizacion.nombre_cotizacion or f"Cotizacion_{cotizacion.id}"
    pdf_name = "".join(c for c in pdf_name_raw if c.isalnum() or c in ('_', '-')).strip().replace(' ', '_')
    if not pdf_name:
        pdf_name = f"Cotizacion_{cotizacion.id}"

    tipo_cotizacion = cotizacion.tipo_cotizacion
    logo_base64 = ""
    company_name = ""
    company_address = ""
    company_phone = ""
    company_email = ""
    template_name = 'cotizacion_pdf_template.html' # Default template

    if tipo_cotizacion and tipo_cotizacion.lower() == 'iamet':
        template_name = 'iamet_cotizacion_pdf_template.html'
        company_name = 'IAMET S.A. de C.V.'
        company_address = 'Av. Principal #456, Col. Centro, Guadalajara, Jalisco'
        company_phone = '+52 33 9876 5432'
        company_email = 'contacto@iamet.com'
    elif tipo_cotizacion and tipo_cotizacion.lower() == 'bajanet':
        template_name = 'cotizacion_pdf_template.html'
        company_name = 'BAJANET S.A. de C.V.'
        company_address = 'Calle Ficticia #123, Colonia Ejemplo, Ciudad de México'
        company_phone = '+52 55 1234 5678'
        company_email = 'ventas@bajanet.com'
        try:
            logo_url = "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRV1xCutWCicl-yXCjMjH5P5jTZA0R993cG9g&s"
            response = requests.get(logo_url)
            response.raise_for_status() # Raise an exception for HTTP errors
            logo_base64 = base64.b64encode(response.content).decode('utf-8')
        except requests.exceptions.RequestException as e:
            print(f"ERROR: Error fetching Bajanet logo from URL: {e}")
            logo_base64 = "" # Fallback to empty string if fetching fails
    else: # Fallback to Bajanet if type is not recognized or None
        template_name = 'cotizacion_pdf_template.html'
        company_name = 'BAJANET S.A. de C.V.'
        company_address = 'Calle Ficticia #123, Colonia Ejemplo, Ciudad de México'
        company_phone = '+52 55 1234 5678'
        company_email = 'ventas@bajanet.com'
        try:
            logo_url = "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRV1xCutWCicl-yXCjMjH5P5jTZA0R993cG9g&s"
            response = requests.get(logo_url)
            response.raise_for_status() # Raise an exception for HTTP errors
            logo_base64 = base64.b64encode(response.content).decode('utf-8')
        except requests.exceptions.RequestException as e:
            print(f"ERROR: Error fetching Bajanet logo from URL: {e}")
            logo_base64 = "" # Fallback to empty string if fetching fails


    context = {
        'cotizacion': cotizacion,
        'detalles_cotizacion': detalles_cotizacion,
        'secciones': secciones,
        'request_user': request.user,
        'current_date': date.today(),
        'company_name': company_name,
        'company_address': company_address,
        'company_phone': company_phone,
        'company_email': company_email,
        'logo_base64': logo_base64,
        'iva_rate_percentage': iva_rate_percentage,
    }

    try:
        print(f"DEBUG: Attempting to render template: {template_name}")
        html_string = render_to_string(template_name, context)
        print("DEBUG: Template rendered to HTML string.")
    except Exception as e:
        print(f"ERROR: Error rendering template '{template_name}': {e}")
        return HttpResponse(f"Internal server error rendering PDF: {e}", status=500)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{pdf_name}.pdf"'

    try:
        print("DEBUG: Attempting to generate PDF with WeasyPrint.")
        HTML(string=html_string).write_pdf(response)
        print("DEBUG: PDF generated successfully.")
    except Exception as e:
        print(f"ERROR: Error generating PDF with WeasyPrint: {e}")
        return HttpResponse(f"Internal server error generating PDF: {e}", status=500)
        
    return response





@login_required
def todos (request):
    # Mostrar todas las oportunidades a todos los usuarios
    items = TodoItem.objects.select_related('usuario', 'cliente').all()

    filter_form = VentaFilterForm(request.GET)

    if filter_form.is_valid():
        area = filter_form.cleaned_data.get('area')
        producto = filter_form.cleaned_data.get('producto')
        orden_monto = filter_form.cleaned_data.get('orden_monto')
        empleado = filter_form.cleaned_data.get('empleado')
        mes_cierre = filter_form.cleaned_data.get('mes_cierre')
        tipo_negociacion = filter_form.cleaned_data.get('tipo_negociacion')
        etapa = filter_form.cleaned_data.get('etapa')
        cliente = filter_form.cleaned_data.get('cliente')

        if area:
            items = items.filter(area=area)
        if producto:
            # Modificación aquí: hacer la búsqueda del producto insensible a mayúsculas/minúsculas
            items = items.filter(producto__iexact=producto) # Usa icontains o iexact para insensibilidad
        if empleado:
            items = items.filter(usuario=empleado)
        if mes_cierre:
            items = items.filter(mes_cierre=mes_cierre)
        if tipo_negociacion:
            items = items.filter(tipo_negociacion=tipo_negociacion)
        if etapa:
            # Filtrar por etapa_corta (vigentes, ganadas, perdidas)
            items = items.filter(etapa_corta__iexact=etapa)
        if cliente:
            items = items.filter(cliente=cliente)

        if orden_monto:
            if orden_monto == 'monto_asc':
                items = items.order_by('monto')
            elif orden_monto == 'monto_desc':
                items = items.order_by('-monto')
        else:
            items = items.order_by('-fecha_creacion')
    else:
        # Ordenar por fecha_creación con índice optimizado
        items = items.order_by('-fecha_creacion')

    # Manejar búsqueda global ANTES de la paginación
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Búsqueda global en todas las oportunidades (por nombre de oportunidad o cliente)
        items = items.filter(
            Q(oportunidad__icontains=search_query) |
            Q(cliente__nombre_empresa__icontains=search_query)
        )

    reporte_activo = request.GET.get('reporte', 'false').lower() == 'true'

    if reporte_activo:
        # Modo reporte: No paginar, mostrar todos los items
        page_obj = None
        context_items = items
    else:
        # Modo normal: Aplicar paginación
        paginator = Paginator(items, 20)
        page_number = request.GET.get('page', 1)
        try:
            page_obj = paginator.get_page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        context_items = page_obj

    # Obtener lista de empleados para el filtro
    from django.contrib.auth.models import User
    empleados = User.objects.filter(is_active=True).order_by('first_name', 'last_name')

    # Obtener lista de clientes para el filtro
    clientes = Cliente.objects.all().order_by('nombre_empresa')

    context = {
        "items": context_items,
        "page_obj": page_obj,
        "filter_form": filter_form,
        "is_supervisor": is_supervisor(request.user),
        "search_query": search_query,
        "reporte_activo": reporte_activo,
        "empleados": empleados,
        "clientes": clientes,
    }
    return render (request, "todos.html", context)




from .bitrix_integration import get_or_create_bitrix_company, send_opportunity_to_bitrix, update_opportunity_in_bitrix, get_all_bitrix_companies, get_all_bitrix_contacts

@login_required
def ingresar_venta_todoitem(request):
    # Asegurarse de que messages y las funciones de bitrix_integration estén disponibles
    

    if request.method == 'POST':
        form = VentaForm(request.POST, user=request.user if not is_supervisor(request.user) else None)
        if form.is_valid():
            cliente_nombre = form.cleaned_data['cliente_nombre']
            bitrix_company_id_from_form = form.cleaned_data.get('bitrix_company_id')

            cliente = None
            # Escenario 1: Se seleccionó un cliente existente de Bitrix
            if bitrix_company_id_from_form:
                try:
                    cliente = Cliente.objects.get(bitrix_company_id=bitrix_company_id_from_form)
                except Cliente.DoesNotExist:
                    # If client doesn't exist locally, create it with the provided Bitrix ID
                    cliente = Cliente.objects.create(
                        nombre_empresa=cliente_nombre,
                        bitrix_company_id=bitrix_company_id_from_form
                    )
                except Exception as e:
                    messages.error(request, f"ERROR: No se pudo obtener o crear el cliente local: {e}")
                    form.add_error('cliente_nombre', 'Hubo un error al procesar el cliente.')
                    return render(request, 'ingresar_venta.html', {'form': form})
            # Escenario 2: Se escribió un nombre de cliente nuevo (sin ID de Bitrix)
            else:
                # Primero, crea la compañía en Bitrix para obtener un ID
                new_bitrix_company_id = get_or_create_bitrix_company(cliente_nombre, request=request)
                if new_bitrix_company_id:
                    cliente, created = Cliente.objects.get_or_create(
                        bitrix_company_id=new_bitrix_company_id,
                        defaults={'nombre_empresa': cliente_nombre}
                    )
                else:
                    messages.error(request, 'No se pudo crear la compañía en Bitrix. Verifique la configuración del webhook.')
                    form.add_error('cliente_nombre', 'No se pudo crear la compañía en Bitrix.')
                    return render(request, 'ingresar_venta.html', {'form': form})

            if cliente:
                venta = form.save(commit=False)
                venta.cliente = cliente
                if is_supervisor(request.user):
                    venta.usuario = form.cleaned_data['usuario']
                else:
                    venta.usuario = request.user
                
                venta.save() # Guardar la instancia de venta para obtener un PK
                
                # Crown Jewel Feature: Trigger immediate opportunity detection
                trigger_immediate_opportunity_notification(venta, request)

                try:
                    opportunity_data = {
                        'oportunidad': venta.oportunidad,
                        'monto': float(venta.monto),
                        'cliente': cliente.nombre_empresa,
                        'bitrix_company_id': cliente.bitrix_company_id,
                        'producto': venta.producto,
                        'area': venta.area,
                        'mes_cierre': venta.mes_cierre,
                        'probabilidad_cierre': venta.probabilidad_cierre,
                        'comentarios': venta.comentarios,
                        'bitrix_stage_id': venta.bitrix_stage_id,
                        'bitrix_contact_id': venta.contacto.bitrix_contact_id if venta.contacto else None,
                    }
                    # Obtener el bitrix_user_id del usuario asignado
                    bitrix_assigned_by_id = None
                    if venta.usuario and hasattr(venta.usuario, 'userprofile') and venta.usuario.userprofile.bitrix_user_id:
                        bitrix_assigned_by_id = venta.usuario.userprofile.bitrix_user_id
                    opportunity_data['bitrix_assigned_by_id'] = bitrix_assigned_by_id
                    
                    bitrix_response = send_opportunity_to_bitrix(opportunity_data, request=request)
                    
                    if bitrix_response and bitrix_response.get('result'):
                        venta.bitrix_deal_id = bitrix_response.get('result')
                        venta.save(update_fields=['bitrix_deal_id']) # Actualizar solo el campo bitrix_deal_id
                        messages.success(request, "Oportunidad creada y sincronizada con Bitrix24.")
                    else:
                        messages.warning(request, "Oportunidad creada, pero no se pudo sincronizar con Bitrix24. Verifique los logs.")

                except Exception as e:
                    messages.error(request, f"Oportunidad creada localmente, pero falló la sincronización con Bitrix24: {e}")
                    print(f"ERROR: Falló la sincronización con Bitrix24 para la oportunidad {venta.id}: {e}")

                return redirect('ingresar_venta_todoitem_exitosa')

    else:
        form = VentaForm(user=request.user if not is_supervisor(request.user) else None)
    
    return render(request, 'ingresar_venta.html', {'form': form})

def ingresar_venta_todoitem_exitosa(request):
    return render(request, 'ingresar_venta_exitosa.html')

# Vistas de autenticación
def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home') # Redirigir a 'home' después de registrar e iniciar sesión
    else:
        form = UserCreationForm()
    return render(request, 'register.html', {'form': form})

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def user_login(request):
    if request.method == 'POST':
        username_input = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        # Si el usuario escribió un correo, buscar el username asociado
        if '@' in username_input:
            from django.contrib.auth.models import User as AuthUser
            found_user = AuthUser.objects.filter(email__iexact=username_input).first()
            if found_user:
                username_input = found_user.username

        # Crear datos modificados con el username correcto
        post_data = request.POST.copy()
        post_data['username'] = username_input

        form = AuthenticationForm(request, data=post_data)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            if request.POST.get('remember'):
                request.session.set_expiry(1209600)  # 2 semanas
            else:
                request.session.set_expiry(0)  # Expira al cerrar el navegador
            return redirect(settings.LOGIN_REDIRECT_URL)
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form, 'hide_dock': True})


@require_POST
def solicitar_reset_password(request):
    try:
        import json
        data = json.loads(request.body)
        username = data.get('username', '').strip()
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Solicitud inválida.'}, status=400)

    if not username:
        return JsonResponse({'status': 'error', 'message': 'El nombre de usuario es requerido.'}, status=400)

    from django.contrib.auth.models import User as AuthUser
    from .models import Notificacion

    try:
        usuario = AuthUser.objects.get(username=username)
    except AuthUser.DoesNotExist:
        # Respuesta genérica para no revelar si el usuario existe
        return JsonResponse({'status': 'ok', 'message': 'Si el usuario existe, la solicitud fue enviada a los supervisores.'})

    supervisores = AuthUser.objects.filter(is_superuser=True)
    if not supervisores.exists():
        return JsonResponse({'status': 'error', 'message': 'No hay supervisores disponibles. Contacta al administrador.'}, status=500)

    titulo = f"Solicitud de restablecimiento de contraseña"
    mensaje = (
        f"El usuario '{usuario.username}' ha solicitado un restablecimiento de contraseña. "
        f"Por favor, autorice y gestione el cambio."
    )

    for supervisor in supervisores:
        Notificacion.objects.create(
            usuario_destinatario=supervisor,
            usuario_remitente=None,
            tipo='sistema',
            titulo=titulo,
            mensaje=mensaje,
        )

    return JsonResponse({'status': 'ok', 'message': 'Solicitud enviada. Un supervisor se pondrá en contacto contigo pronto.'})

@login_required
def user_logout(request):
    logout(request)
    return redirect('user_login')


def set_language(request):
    """
    Vista para cambiar el idioma de la aplicación
    Soporta tanto peticiones AJAX como formularios normales
    """
    logger = logging.getLogger(__name__)
    
    try:
        # Verificar método HTTP
        if request.method != 'POST':
            logger.warning('Intento de acceso con método no permitido: %s', request.method)
            return JsonResponse(
                {'status': 'error', 'message': 'Método no permitido'}, 
                status=405
            )
        
        # Obtener el idioma solicitado
        language = request.POST.get('language', 'es')
        logger.info('Solicitud de cambio de idioma a: %s', language)
        
        # Validar que el idioma esté en los idiomas configurados
        available_languages = [lang[0] for lang in settings.LANGUAGES]
        if language not in available_languages:
            logger.warning('Idioma no válido: %s. Idiomas disponibles: %s', language, available_languages)
            language = 'es'  # Valor por defecto
        
        # Activar el idioma para esta sesión
        translation.activate(language)
        logger.debug('Idioma activado: %s', language)
        
        # Configurar el idioma en la sesión
        if hasattr(request, 'session'):
            request.session[LANGUAGE_SESSION_KEY] = language
            request.session.modified = True
            logger.debug('Idioma guardado en la sesión')
        
        # Guardar en el perfil del usuario si está autenticado
        if request.user.is_authenticated:
            try:
                if hasattr(request.user, 'userprofile'):
                    request.user.userprofile.language = language
                    request.user.userprofile.save()
                    logger.debug('Idioma guardado en el perfil del usuario')
            except Exception as e:
                logger.error('Error al guardar el idioma en el perfil del usuario: %s', str(e), exc_info=True)
        
        # Determinar si es una petición AJAX
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                 request.content_type == 'application/json' or \
                 request.META.get('HTTP_ACCEPT') == 'application/json'
        
        # Configurar la respuesta
        if is_ajax:
            response_data = {
                'status': 'success', 
                'language': language,
                'message': 'Idioma cambiado correctamente',
                'redirect': request.META.get('HTTP_REFERER', '/')
            }
            response = JsonResponse(response_data)
        else:
            # Para formularios normales, redirigir a la página anterior o a la raíz
            next_url = request.POST.get('next', request.META.get('HTTP_REFERER', '/'))
            response = redirect(next_url)
        
        # Configurar la cookie de idioma
        response.set_cookie(
            settings.LANGUAGE_COOKIE_NAME,
            language,
            max_age=365 * 24 * 60 * 60,  # 1 año
            path=settings.LANGUAGE_COOKIE_PATH or '/',
            domain=settings.LANGUAGE_COOKIE_DOMAIN or None,
            secure=request.is_secure(),
            httponly=True,
            samesite='Lax'
        )
        
        logger.info('Cambio de idioma exitoso a: %s', language)
        return response
    
    except Exception as e:
        # Registrar el error completo para depuración
        logger.error('Error al cambiar el idioma', exc_info=True)
        
        # Devolver un mensaje de error genérico al cliente
        return JsonResponse(
            {'status': 'error', 'message': 'Error interno del servidor'},
            status=500
        )

@login_required
def calendario_view(request):
    """
    Vista para mostrar el calendario de actividades.
    """
    return render(request, 'calendario.html', {})

@login_required
@csrf_exempt
def actividad_list_create(request):
    """
    API para listar y crear actividades del calendario.
    """
    try:
        if request.method == 'GET':
            if is_supervisor(request.user):
                actividades = Actividad.objects.all()
            else:
                    actividades = Actividad.objects.filter(Q(creado_por=request.user) | Q(participantes=request.user)).distinct()

            # Filtro por vendedores (IDs separados por coma)
            vendedores_param = request.GET.get('vendedores', '').strip()
            if vendedores_param and is_supervisor(request.user):
                try:
                    ids = [int(x) for x in vendedores_param.split(',') if x.strip().isdigit()]
                    if ids:
                        actividades = actividades.filter(creado_por_id__in=ids)
                except Exception:
                    pass

            # Filtro por mes (YYYY-MM)
            mes_param = request.GET.get('mes', '').strip()
            if mes_param:
                try:
                    from datetime import datetime as _dt
                    year, month = int(mes_param[:4]), int(mes_param[5:7])
                    from django.utils import timezone as _tz
                    import calendar as _cal
                    last_day = _cal.monthrange(year, month)[1]
                    desde = _tz.make_aware(_dt(year, month, 1, 0, 0, 0))
                    hasta = _tz.make_aware(_dt(year, month, last_day, 23, 59, 59))
                    actividades = actividades.filter(fecha_inicio__lte=hasta, fecha_fin__gte=desde)
                except Exception:
                    pass

            # Filtro por oportunidad
            oportunidad_id = request.GET.get('oportunidad_id', '').strip()
            if oportunidad_id:
                actividades = actividades.filter(oportunidad_id=oportunidad_id)

            actividades = actividades.select_related('creado_por', 'oportunidad').prefetch_related('participantes')

            events = []
            _1h = timedelta(hours=1)
            for actividad in actividades:
                participants_data = [{'id': p.id, 'text': p.get_full_name() or p.username} for p in actividad.participantes.all()]
                opportunity_data = None
                if actividad.oportunidad:
                    opportunity_data = {'id': actividad.oportunidad.id, 'text': actividad.oportunidad.oportunidad}
                # Clampar eventos multi-día a 1h para que no crucen semanas en el calendario
                fin_display = actividad.fecha_fin
                if (fin_display - actividad.fecha_inicio).days >= 1:
                    fin_display = actividad.fecha_inicio + _1h
                events.append({
                    'id': actividad.id,
                    'title': actividad.titulo,
                    'tipo': actividad.tipo_actividad,
                    'start': actividad.fecha_inicio.isoformat(),
                    'end': fin_display.isoformat(),
                    'description': actividad.descripcion or '',
                    'color': actividad.color,
                    'participants': participants_data,
                    'opportunity': opportunity_data,
                    'creado_por': {'id': actividad.creado_por.id, 'text': actividad.creado_por.get_full_name() or actividad.creado_por.username},
                    'es_mio': actividad.creado_por_id == request.user.pk,
                })
            return JsonResponse(events, safe=False)

        elif request.method == 'POST':
            data = json.loads(request.body)

            # Validar fechas
            try:
                start_date = datetime.fromisoformat(data['start'])
                end_date = datetime.fromisoformat(data['end'])
                if timezone.is_naive(start_date):
                    start_date = timezone.make_aware(start_date)
                if timezone.is_naive(end_date):
                    end_date = timezone.make_aware(end_date)
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha inválido.'}, status=400)

            actividad = Actividad.objects.create(
                titulo=data['title'],
                tipo_actividad=data.get('tipo', 'otro'),
                descripcion=data.get('description', ''),
                fecha_inicio=start_date,
                fecha_fin=end_date,
                creado_por=request.user,
                color=data.get('color', '#1D1D1F'),
                oportunidad_id=data.get('opportunity')
            )
        
            if 'participants' in data and data['participants']:
                actividad.participantes.set(data['participants'])
        
            participants_data = []
            for p in actividad.participantes.all():
                participants_data.append({'id': p.id, 'text': p.get_full_name() or p.username})
        
            opportunity_data = None
            if actividad.oportunidad:
                opportunity_data = {'id': actividad.oportunidad.id, 'text': actividad.oportunidad.oportunidad}

            return JsonResponse({
                'id': actividad.id,
                'title': actividad.titulo,
                'tipo': actividad.tipo_actividad,
                'start': actividad.fecha_inicio.isoformat(),
                'end': actividad.fecha_fin.isoformat(),
                'description': actividad.descripcion,
                'color': actividad.color,
                'participants': participants_data,
                'opportunity': opportunity_data,
                'creado_por': {'id': actividad.creado_por.id, 'text': actividad.creado_por.get_full_name() or actividad.creado_por.username},
                'es_mio': True,
            }, status=201)
    
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)
@login_required
@csrf_exempt
def actividad_detail(request, pk):
    """
    API para obtener, actualizar o eliminar una actividad específica.
    """
    actividad = get_object_or_404(Actividad, pk=pk)

    # Verificar permisos: solo el creador o un participante puede ver/editar/eliminar
    if actividad.creado_por != request.user and request.user not in actividad.participantes.all() and not is_supervisor(request.user):
        return JsonResponse({'error': 'No tienes permiso para acceder a esta actividad.'}, status=403)

    if request.method == 'GET':
        participants_data = []
        for p in actividad.participantes.all():
            participants_data.append({'id': p.id, 'text': p.get_full_name() or p.username})
        
        opportunity_data = None
        if actividad.oportunidad:
            opportunity_data = {'id': actividad.oportunidad.id, 'text': actividad.oportunidad.oportunidad}

        return JsonResponse({
            'id': actividad.id,
            'title': actividad.titulo,
            'tipo': actividad.tipo_actividad,
            'start': actividad.fecha_inicio.isoformat(),
            'end': actividad.fecha_fin.isoformat(),
            'description': actividad.descripcion,
            'color': actividad.color,
            'participants': participants_data,
            'opportunity': opportunity_data,
            'creado_por': {'id': actividad.creado_por.id, 'text': actividad.creado_por.get_full_name() or actividad.creado_por.username},
            'es_mio': actividad.creado_por_id == request.user.pk,
        })

    elif request.method == 'PUT':
        data = json.loads(request.body)

        # Verificar permisos: solo el creador o un supervisor puede editar
        if actividad.creado_por != request.user and not is_supervisor(request.user):
            return JsonResponse({'error': 'No tienes permiso para editar esta actividad.'}, status=403)

        # Validar fechas
        try:
            start_date = datetime.fromisoformat(data['start'])
            end_date = datetime.fromisoformat(data['end'])
            if timezone.is_naive(start_date):
                start_date = timezone.make_aware(start_date)
            if timezone.is_naive(end_date):
                end_date = timezone.make_aware(end_date)
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha inválido.'}, status=400)

        actividad.titulo = data['title']
        actividad.tipo_actividad = data.get('tipo', 'otro')
        actividad.descripcion = data.get('description', '')
        actividad.fecha_inicio = start_date
        actividad.fecha_fin = end_date
        actividad.color = data.get('color', '#007AFF')
        actividad.oportunidad_id = data.get('opportunity')
        actividad.save()

        if 'participants' in data and data['participants'] is not None:
            actividad.participantes.set(data['participants'])
        else:
            actividad.participantes.clear() # Clear if no participants are sent

        participants_data = []
        for p in actividad.participantes.all():
            participants_data.append({'id': p.id, 'text': p.get_full_name() or p.username})

        opportunity_data = None
        if actividad.oportunidad:
            opportunity_data = {'id': actividad.oportunidad.id, 'text': actividad.oportunidad.oportunidad}

        return JsonResponse({
            'id': actividad.id,
            'title': actividad.titulo,
            'tipo': actividad.tipo_actividad,
            'start': actividad.fecha_inicio.isoformat(),
            'end': actividad.fecha_fin.isoformat(),
            'description': actividad.descripcion,
            'color': actividad.color,
            'participants': participants_data,
            'opportunity': opportunity_data,
            'creado_por': {'id': actividad.creado_por.id, 'text': actividad.creado_por.get_full_name() or actividad.creado_por.username},
            'es_mio': actividad.creado_por_id == request.user.pk,
        })

    elif request.method == 'DELETE':
        # Verificar permisos: solo el creador o un supervisor puede eliminar
        if actividad.creado_por != request.user and not is_supervisor(request.user):
            return JsonResponse({'error': 'No tienes permiso para eliminar esta actividad.'}, status=403)

        actividad.delete()
        return JsonResponse({'message': 'Actividad eliminada exitosamente'}, status=204)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required
def user_list_api(request):
    """
    API para obtener una lista de usuarios para Select2.
    """
    search_query = request.GET.get('q', '')
    users = User.objects.filter(
        Q(username__icontains=search_query) | 
        Q(first_name__icontains=search_query) | 
        Q(last_name__icontains=search_query)
    ).order_by('username')[:20] # Limitar resultados
    
    results = []
    for user in users:
        avatar_url = None
        if hasattr(user, 'userprofile'):
            try:
                avatar_url = user.userprofile.get_avatar_url()
            except:
                avatar_url = None
        results.append({
            'id': user.id, 
            'text': user.get_full_name() or user.username,
            'avatar_url': avatar_url
        })
    return JsonResponse(results, safe=False)
@login_required
def oportunidad_list_api(request):
    """
    API para obtener una lista de oportunidades para Select2.
    """
    search_query = request.GET.get('q', '')
    
    if is_supervisor(request.user):
        oportunidades = TodoItem.objects.filter(
            Q(oportunidad__icontains=search_query) |
            Q(cliente__nombre_empresa__icontains=search_query)
        ).order_by('-fecha_creacion')[:20]
    else:
        oportunidades = TodoItem.objects.filter(
            Q(usuario=request.user) &
            (Q(oportunidad__icontains=search_query) |
            Q(cliente__nombre_empresa__icontains=search_query))
        ).order_by('-fecha_creacion')[:20]
    
    results = []
    for op in oportunidades:
        results.append({'id': op.id, 'text': f"{op.oportunidad} ({op.cliente.nombre_empresa if op.cliente else 'Sin Cliente'})", 'title': op.oportunidad})
    return JsonResponse(results, safe=False)



@login_required
def editar_venta_todoitem(request, pk):
    # Supervisor puede editar cualquier venta, vendedor solo las suyas
    if is_supervisor(request.user):
        todo_item = get_object_or_404(TodoItem, pk=pk)
    else:
        todo_item = get_object_or_404(TodoItem, pk=pk, usuario=request.user)

    if request.method == 'POST':
        if 'delete' in request.POST:
            # Solo permite borrar si el usuario es supervisor o dueño
            if is_supervisor(request.user) or todo_item.usuario == request.user:
                todo_item.delete()
                messages.success(request, "Oportunidad eliminada con éxito.")
            else:
                messages.error(request, "No tienes permiso para eliminar esta oportunidad.")
            return redirect('todos')
        
        # Si no es delete, entonces es edición:
        form = VentaForm(request.POST, instance=todo_item, user=request.user if not is_supervisor(request.user) else None)
        if form.is_valid():
            cliente_nombre = form.cleaned_data['cliente_nombre']
            bitrix_company_id_from_form = form.cleaned_data.get('bitrix_company_id')

            cliente = None
            # Try to get client by bitrix_company_id if provided
            if bitrix_company_id_from_form:
                try:
                    cliente = Cliente.objects.get(bitrix_company_id=bitrix_company_id_from_form)
                except Cliente.DoesNotExist:
                    # If client doesn't exist locally, create it with the provided Bitrix ID
                    cliente = Cliente.objects.create(
                        nombre_empresa=cliente_nombre,
                        bitrix_company_id=bitrix_company_id_from_form
                    )
                except Exception as e:
                    messages.error(request, f"ERROR: No se pudo obtener o crear el cliente local por Bitrix ID: {e}")
                    form.add_error('cliente_nombre', 'Hubo un error al procesar el cliente.')
                    return render(request, 'editar_venta.html', {'form': form, 'todo_item': todo_item})
            else:
                # If no bitrix_company_id from form, try to find by name or create a new one without Bitrix ID
                try:
                    cliente, created = Cliente.objects.get_or_create(
                        nombre_empresa=cliente_nombre,
                        defaults={'bitrix_company_id': None}
                    )
                except Exception as e:
                    messages.error(request, f"ERROR: No se pudo obtener o crear el cliente local por nombre: {e}")
                    form.add_error('cliente_nombre', 'Hubo un error al procesar el cliente.')
                    return render(request, 'editar_venta.html', {'form': form, 'todo_item': todo_item})

            venta = form.save(commit=False)
            venta.cliente = cliente
            if is_supervisor(request.user):
                venta.usuario = form.cleaned_data['usuario']
            else:
                venta.usuario = request.user
            venta.save()

            # Actualizar en Bitrix si existe un bitrix_deal_id
            if venta.bitrix_deal_id:
                opportunity_data = {
                    'oportunidad': venta.oportunidad,
                    'monto': float(venta.monto),
                    'cliente': cliente.nombre_empresa,
                    'bitrix_company_id': cliente.bitrix_company_id,
                    'producto': venta.producto,
                    'area': venta.area,
                    'mes_cierre': venta.mes_cierre,
                    'probabilidad_cierre': venta.probabilidad_cierre,
                    'comentarios': venta.comentarios,
                    'bitrix_stage_id': venta.bitrix_stage_id,
                }
                # Obtener el bitrix_user_id del usuario asignado
                bitrix_assigned_by_id = None
                if venta.usuario and hasattr(venta.usuario, 'userprofile') and venta.usuario.userprofile.bitrix_user_id:
                    bitrix_assigned_by_id = venta.usuario.userprofile.bitrix_user_id
                opportunity_data['bitrix_assigned_by_id'] = bitrix_assigned_by_id
                bitrix_updated = update_opportunity_in_bitrix(venta.bitrix_deal_id, opportunity_data, request=request)
                if bitrix_updated:
                    messages.success(request, "Oportunidad actualizada en Bitrix24 con éxito.")
                else:
                    messages.error(request, "Error al actualizar la oportunidad en Bitrix24.")
            else:
                messages.warning(request, "La oportunidad no tiene un ID de Bitrix24 asociado. No se pudo actualizar en Bitrix24.")

            return redirect('todos')
    else:
        form = VentaForm(instance=todo_item, user=request.user if not is_supervisor(request.user) else None)

    return render(request, 'editar_venta.html', {'form': form, 'todo_item': todo_item})


@login_required
def reporte_ventas_por_cliente(request):
    from django.contrib.auth.models import User
    if is_supervisor(request.user):
        reporte_data = Cliente.objects.annotate(
            total_monto=Coalesce(
                Sum('oportunidades__monto', filter=Q(oportunidades__probabilidad_cierre=100)),
                Value(Decimal('0.00'))
            )
        ).values(
            'id',
            'nombre_empresa',
            'total_monto'
        ).order_by('nombre_empresa')
        total_general = TodoItem.objects.filter(probabilidad_cierre=100).aggregate(
            sum_monto=Sum('monto')
        )['sum_monto'] or Decimal('0.00')
        usuarios = User.objects.filter(is_active=True)
    else:
        reporte_data = Cliente.objects.filter(asignado_a=request.user).annotate(
            total_monto=Coalesce(
                Sum('oportunidades__monto', filter=Q(oportunidades__probabilidad_cierre=100, oportunidades__usuario=request.user)),
                Value(Decimal('0.00'))
            )
        ).values(
            'id',
            'nombre_empresa',
            'total_monto'
        ).order_by('nombre_empresa')
        total_general = TodoItem.objects.filter(
            usuario=request.user, probabilidad_cierre=100
        ).aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')
        usuarios = None
    context = {
        'reporte_data': reporte_data,
        'total_general': total_general,
        'is_supervisor': is_supervisor(request.user),
        'usuarios': usuarios,
    }
    return render(request, 'reporte_ventas_por_cliente.html', context)


@login_required
def oportunidades_por_cliente(request, cliente_id):
    # Determinar qué clientes pueden ser vistos por el usuario
    if is_supervisor(request.user):
        cliente_seleccionado = get_object_or_404(Cliente, pk=cliente_id) # No filtrar por usuario
        oportunidades = TodoItem.objects.filter(cliente=cliente_seleccionado) # Todas las oportunidades del cliente
        print("DEBUG: Supervisor viendo oportunidades de cliente.")
    else:
        cliente_seleccionado = get_object_or_404(Cliente, pk=cliente_id, asignado_a=request.user) # Usar asignado_a
        oportunidades = TodoItem.objects.filter(cliente=cliente_seleccionado, usuario=request.user)
        print(f"DEBUG: Vendedor {request.user.username} viendo sus propias oportunidades de cliente.")

    # El formulario de filtro no necesita el usuario para sus querysets de clientes en este contexto
    # ya que los clientes ya vienen filtrados por la vista o se obtienen todos.
    filter_form = VentaFilterForm(request.GET)

    if filter_form.is_valid():
        area = filter_form.cleaned_data.get('area')
        producto = filter_form.cleaned_data.get('producto')
        orden_monto = filter_form.cleaned_data.get('orden_monto')
        probabilidad_min = filter_form.cleaned_data.get('probabilidad_min')
        probabilidad_max = filter_form.cleaned_data.get('probabilidad_max')
        mes_cierre = filter_form.cleaned_data.get('mes_cierre')

        if area:
            oportunidades = oportunidades.filter(area=area)
        if producto:
            oportunidades = oportunidades.filter(producto=producto)
        if probabilidad_min is not None:
            oportunidades = oportunidades.filter(probabilidad_cierre__gte=probabilidad_min)
        if probabilidad_max is not None:
            oportunidades = oportunidades.filter(probabilidad_cierre__lte=probabilidad_max)
        if mes_cierre:
            oportunidades = oportunidades.filter(mes_cierre=mes_cierre)

        if orden_monto:
            if orden_monto == 'monto_asc':
                oportunidades = oportunidades.order_by('monto')
            elif orden_monto == 'monto_desc':
                oportunidades = oportunidades.order_by('-monto')
        else:
            oportunidades = oportunidades.order_by('-fecha_creacion')
    else:
        oportunidades = oportunidades.order_by('-fecha_creacion')


    context = {
        'cliente': cliente_seleccionado,
        'oportunidades': oportunidades,
        'filter_form': filter_form,
        'is_supervisor': is_supervisor(request.user),
    }
    return render(request, 'oportunidades_por_cliente.html', context)


@login_required
def producto_dashboard_detail(request, producto_val):
    print(f"DEBUG: producto_dashboard_detail - producto_val recibido RAW: {producto_val}")

    # Convertir a mayúsculas para asegurar que la comparación con PRODUCTO_CHOICES sea consistente
    producto_val_upper = producto_val.upper()
    print(f"DEBUG: producto_dashboard_detail - producto_val_upper: {producto_val_upper}")
    print(f"DEBUG: Keys de PRODUCTO_CHOICES: {list(dict(TodoItem.PRODUCTO_CHOICES).keys())}")

    # Verificar si el producto_val_upper es una clave válida en PRODUCTO_CHOICES
    if producto_val_upper not in dict(TodoItem.PRODUCTO_CHOICES):
        return redirect('dashboard')

    if is_supervisor(request.user):
        oportunidades = TodoItem.objects.filter(producto=producto_val_upper)
    else:
        oportunidades = TodoItem.objects.filter(producto=producto_val_upper, usuario=request.user)

    print(f"DEBUG: Oportunidades encontradas para {producto_val_upper} (antes de desglosar): {oportunidades.count()}")
    for op in oportunidades:
        print(f"DEBUG:   - ID: {op.id}, Oportunidad: {op.oportunidad}, Producto: {op.producto}, Usuario ID: {op.usuario.id}")

    # --- Ventas Cerradas (probabilidad 100%) para este producto ---
    ventas_cerradas = oportunidades.filter(probabilidad_cierre=100)
    total_vendido_cerrado = ventas_cerradas.aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')
    total_vendido_cerrado_count = ventas_cerradas.count() # Conteo de oportunidades cerradas
    print(f"DEBUG: Ventas Cerradas (100%) para '{producto_val_upper}': {total_vendido_cerrado_count} oportunidades, Monto: {total_vendido_cerrado}")
    for venta in ventas_cerradas:
        print(f"DEBUG:   - Oportunidad: {venta.oportunidad}, Monto: {venta.monto}, Probabilidad: {venta.probabilidad_cierre}%")

    # --- Oportunidades Vigentes (probabilidad del 1% al 99%) para este producto ---
    oportunidades_vigentes = oportunidades.filter(
        probabilidad_cierre__gt=0, # Mayor que 0%
        probabilidad_cierre__lt=100 # Menor que 100%
    )
    total_monto_vigente = oportunidades_vigentes.aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')
    total_monto_vigente_count = oportunidades_vigentes.count() # Conteo de oportunidades vigentes
    print(f"DEBUG: Oportunidades Vigentes (0% < prob < 100%) para '{producto_val_upper}': {total_monto_vigente_count} oportunidades, Monto: {total_monto_vigente}")
    for op_vigente in oportunidades_vigentes:
        print(f"DEBUG:   - Oportunidad: {op_vigente.oportunidad}, Monto: {op_vigente.monto}, Probabilidad: {op_vigente.probabilidad_cierre}%")

    # --- Oportunidades Perdidas (probabilidad 0%) para este producto ---
    oportunidades_perdidas = oportunidades.filter(probabilidad_cierre=0)
    total_monto_perdido = oportunidades_perdidas.aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')
    total_monto_perdido_count = oportunidades_perdidas.count() # Conteo de oportunidades perdidas
    print(f"DEBUG: Oportunidades Perdidas (0%) para '{producto_val_upper}': {total_monto_perdido_count} oportunidades, Monto: {total_monto_perdido}")
    for op_perdida in oportunidades_perdidas:
        print(f"DEBUG:   - Oportunidad: {op_perdida.oportunidad}, Monto: {op_perdida.monto}, Probabilidad: {op_perdida.probabilidad_cierre}%")


    # Clientes involucrados en este producto
    clientes_involucrados = oportunidades.filter(cliente__isnull=False).values('cliente__id', 'cliente__nombre_empresa').distinct()

    # Meses involucrados en este producto (mes de cierre esperado)
    meses_involucrados = oportunidades.values('mes_cierre').distinct()

    # Mapear valores crudos de mes a sus nombres de visualización
    meses_display = []
    for m in meses_involucrados:
        # Aseguramos que la clave sea un string de dos dígitos para la búsqueda
        mes_key = str(m['mes_cierre']).zfill(2)
        meses_display.append(dict(TodoItem.MES_CHOICES).get(mes_key, mes_key))
    context = {
        'producto_val': producto_val_upper, # Aseguramos que la clave pasada sea la que usará el template
        'producto_display': dict(TodoItem.PRODUCTO_CHOICES).get(producto_val_upper, producto_val_upper),
        'total_vendido_cerrado': total_vendido_cerrado,
        'total_vendido_cerrado_count': total_vendido_cerrado_count, # AÑADIDO
        'total_monto_vigente': total_monto_vigente, # Nuevo: Monto oportunidades vigentes
        'total_monto_vigente_count': total_monto_vigente_count, # AÑADIDO
        'total_monto_perdido': total_monto_perdido, # Nuevo: Monto oportunidades perdidas
        'total_monto_perdido_count': total_monto_perdido_count, # AÑADIDO
        'clientes_involucrados': clientes_involucrados,
        'meses_involucrados_display': meses_display,
        'oportunidades': oportunidades, # Pasar todas las oportunidades para listarlas
        'is_supervisor': is_supervisor(request.user), # Pasamos si el usuario es supervisor al contexto
    }
    return render(request, 'producto_dashboard_detail.html', context)


@login_required
def mes_dashboard_detail(request, mes_val):
    # Asegúrate de que el mes_val recibido es uno de los choices válidos
    mes_val_padded = str(mes_val).zfill(2) # Asegurar que mes_val sea de dos dígitos para la validación
    if mes_val_padded not in dict(TodoItem.MES_CHOICES).keys():
        return redirect('home') # Redirige a home si el mes no es válido

    # Base queryset de oportunidades según el rol
    if is_supervisor(request.user):
        oportunidades_mes = TodoItem.objects.filter(mes_cierre=mes_val_padded)
    else:
        oportunidades_mes = TodoItem.objects.filter(usuario=request.user, mes_cierre=mes_val_padded)


    # Monto total esperado para este mes
    total_monto_esperado = oportunidades_mes.aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')

    # Monto POR COBRAR: oportunidades con probabilidad entre 1 and 99%
    por_cobrar_monto = oportunidades_mes.filter(probabilidad_cierre__gte=1, probabilidad_cierre__lte=99).aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')

    # Clientes involucrados en oportunidades para este mes
    clientes_involucrados = oportunidades_mes.filter(cliente__isnull=False).values('cliente__id', 'cliente__nombre_empresa').distinct()

    # Datos para la gráfica: Probabilidad de cierre vs. Monto
    graph_data_raw = oportunidades_mes.values('id', 'oportunidad', 'producto', 'monto', 'probabilidad_cierre', 'cliente__nombre_empresa')

    # Añadir 'get_producto_display' a cada item en graph_data
    graph_data_with_display = []
    for item in graph_data_raw:
        item_copy = item.copy()
        item_copy['get_producto_display'] = dict(TodoItem.PRODUCTO_CHOICES).get(item_copy['producto'], item_copy['producto'])
        graph_data_with_display.append(item_copy)

    context = {
        'mes_val': mes_val_padded, # Aseguramos que la clave pasada sea la que usará el template
        'mes_display': dict(TodoItem.MES_CHOICES).get(mes_val_padded, mes_val_padded),
        'total_monto_esperado': total_monto_esperado,
        'por_cobrar_monto': por_cobrar_monto,
        'clientes_involucrados': clientes_involucrados,
        'oportunidades': oportunidades_mes, # Pasar todas las oportunidades para listarlas
        'graph_data_json': graph_data_with_display, # Pasa los datos procesados con display_value
        'is_supervisor': is_supervisor(request.user),
    }
    return render(request, 'mes_dashboard_detail.html', context)


@login_required
def oportunidades_perdidas_detail(request):
    """
    Vista para mostrar todas las oportunidades con 0% de probabilidad de cierre.
    Considera el rol de supervisor.
    """
    if is_supervisor(request.user):
        oportunidades_perdidas = TodoItem.objects.filter(probabilidad_cierre=0).order_by('-fecha_creacion')
    else:
        oportunidades_perdidas = TodoItem.objects.filter(usuario=request.user, probabilidad_cierre=0).order_by('-fecha_creacion')

    total_perdido_monto = oportunidades_perdidas.aggregate(
        sum_monto=Sum('monto')
    )['sum_monto'] or Decimal('0.00')

    context = {
        'oportunidades': oportunidades_perdidas,
        'titulo': "Oportunidades Perdidas (0% Probabilidad)",
        'total_perdido_monto': total_perdido_monto,
        'is_supervisor': is_supervisor(request.user),
    }
    return render(request, 'oportunidades_perdidas_detail.html', context)


# VISTA DEPRECADA - MANTENIDA POR REFERENCIA
@login_required
def generate_quote_pdf(request, pk):
    """
    DEPRECATED: Esta vista generaba PDF para TodoItem.
    Now it will use generate_cotizacion_pdf for the Cotizacion model.
    """
    # Ensure that only the owner or a supervisor can generate the quote
    if is_supervisor(request.user):
        opportunity = get_object_or_404(TodoItem, pk=pk)
    else:
        opportunity = get_object_or_404(TodoItem, pk=pk, usuario=request.user)

    # Context for the PDF template
    context = {
        'opportunity': opportunity,
        'request_user': request.user, # To show the user generating the quote
        'current_date': date.today(),
        # You can add more company data here if you have it in the model or settings
        'company_name': 'Tu Empresa de Ventas S.A. de C.V.',
        'company_address': 'Calle Ficticia #123, Colonia Ejemplo, Ciudad de México',
        'company_phone': '+52 55 1234 5678',
        'company_email': 'ventas@tuempresa.com',
    }

    # Render the HTML template to a string
    html_string = render_to_string('quote_pdf.html', context)

    # Create the PDF from the HTML string
    # You can add a base_url if you have external images or CSS that WeasyPrint needs to load
    # Example: HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    pdf_file = HTML(string=html_string).write_pdf()

    # Create the HTTP response with the PDF
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="cotizacion_{opportunity.oportunidad}.pdf"'
    return response



# --- NEW/CORRECTED VIEWS ADDED ---

@login_required
def exportar_oportunidades_csv(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import date
        OPENPYXL_AVAILABLE = True
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        OPENPYXL_AVAILABLE = False
        from datetime import date
    
    # 1. Get the base queryset with cotizations
    if is_supervisor(request.user):
        items = TodoItem.objects.select_related('usuario', 'cliente').prefetch_related('cotizaciones__detalles').all()
    else:
        items = TodoItem.objects.select_related('usuario', 'cliente').prefetch_related('cotizaciones__detalles').filter(usuario=request.user)

    # 2. Apply filters (supporting multiple values)
    oportunidad_filter = request.GET.get('filterOportunidad', '').strip()
    if oportunidad_filter:
        items = items.filter(oportunidad__icontains=oportunidad_filter)
        
    cliente_filter = request.GET.get('filterCliente', '').strip()
    if cliente_filter:
        # Si es un número, buscar por ID; si no, buscar por nombre
        if cliente_filter.isdigit():
            items = items.filter(cliente__id=int(cliente_filter))
        else:
            items = items.filter(cliente__nombre_empresa__icontains=cliente_filter)
        
    # Filtro de tipo
    tipo_filter = request.GET.get('filterTipo', '').strip()
    if tipo_filter:
        items = items.filter(tipo_negociacion=tipo_filter)
        
    # Filtro de empleado múltiple
    empleado_filter = request.GET.get('empleado', '').strip()
    if empleado_filter:
        empleado_ids = [e.strip() for e in empleado_filter.split(',') if e.strip()]
        if empleado_ids:
            items = items.filter(usuario__id__in=empleado_ids)
            
    # Filtro de mes de cierre múltiple
    mes_cierre_filter = request.GET.get('filterMesCierre', '').strip()
    if mes_cierre_filter:
        meses = [m.strip() for m in mes_cierre_filter.split(',') if m.strip()]
        if meses:
            items = items.filter(mes_cierre__in=meses)
            
    # Filtro de etapa múltiple (estado)
    etapa_filter = request.GET.get('filterEtapa', '').strip()
    if etapa_filter:
        from django.db.models import Q
        etapas = [e.strip() for e in etapa_filter.split(',') if e.strip()]
        if etapas:
            etapa_conditions = Q()
            for etapa in etapas:
                if etapa == 'vigentes':
                    # Excluir cerradas (ganadas, perdidas y pagadas)
                    etapa_conditions |= ~Q(etapa_completa__icontains='ganado') & ~Q(etapa_completa__icontains='perdido') & ~Q(etapa_completa__icontains='pagado')
                elif etapa == 'ganadas':
                    # Solo cerradas ganadas, incluir pagado
                    etapa_conditions |= Q(etapa_completa__icontains='ganado') | Q(etapa_completa__icontains='pagado')
                elif etapa == 'perdidas':
                    # Solo cerradas perdidas
                    etapa_conditions |= Q(etapa_completa__icontains='perdido')
            items = items.filter(etapa_conditions)

    area_filter = request.GET.get('filterArea', '').strip()
    if area_filter:
        items = items.filter(area=area_filter)

    # 3. Apply sorting
    orden_monto = request.GET.get('orden_monto')
    if orden_monto:
        if orden_monto == 'monto_asc':
            items = items.order_by('monto')
        elif orden_monto == 'monto_desc':
            items = items.order_by('-monto')

    orden_probabilidad = request.GET.get('orden_probabilidad')
    if orden_probabilidad:
        if orden_probabilidad == 'prob_asc':
            items = items.order_by('probabilidad_cierre')
        elif orden_probabilidad == 'prob_desc':
            items = items.order_by('-probabilidad_cierre')

    # 4. Create report based on available libraries
    if OPENPYXL_AVAILABLE:
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Oportunidades"
        
        # --- START: Add Report Metadata ---
        bold_font = Font(bold=True)
        
        # Row 1: Report Author
        ws['A1'] = "Reporte generado por:"
        ws['A1'].font = bold_font
        ws['B1'] = request.user.get_full_name() or request.user.username
        
        # Row 2: Generation Date
        ws['A2'] = "Fecha de generación:"
        ws['A2'].font = bold_font
        ws['B2'] = date.today().strftime("%d/%m/%Y")
        
        # Row 3: Applied Filters
        ws['A3'] = "Filtros aplicados:"
        ws['A3'].font = bold_font
        
        filters_applied = []
        
        # Filtro de oportunidad
        if oportunidad_filter:
            filters_applied.append(f"Oportunidad: {oportunidad_filter}")
            
        # Filtro de cliente
        if cliente_filter:
            filters_applied.append(f"Cliente: {cliente_filter}")
            
        # Filtro de tipo
        if tipo_filter:
            tipo_display = "Runrate" if tipo_filter == "runrate" else "Proyecto"
            filters_applied.append(f"Tipo: {tipo_display}")
            
        # Filtro de empleado múltiple
        if empleado_filter:
            empleado_ids = [e.strip() for e in empleado_filter.split(',') if e.strip()]
            empleado_nombres = []
            for emp_id in empleado_ids:
                try:
                    from django.contrib.auth.models import User
                    empleado = User.objects.get(id=emp_id)
                    empleado_nombres.append(empleado.get_full_name() or empleado.username)
                except User.DoesNotExist:
                    empleado_nombres.append(f"ID:{emp_id}")
            if empleado_nombres:
                filters_applied.append(f"Empleado(s): {', '.join(empleado_nombres)}")
                
        # Filtro de mes de cierre múltiple
        if mes_cierre_filter:
            meses = [m.strip() for m in mes_cierre_filter.split(',') if m.strip()]
            mes_nombres = []
            mes_map = {
                '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
                '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
                '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
            }
            for mes in meses:
                mes_nombres.append(mes_map.get(mes, mes))
            if mes_nombres:
                filters_applied.append(f"Mes(es) de Cierre: {', '.join(mes_nombres)}")
                
        # Filtro de etapa múltiple
        if etapa_filter:
            etapas = [e.strip() for e in etapa_filter.split(',') if e.strip()]
            etapa_nombres = []
            etapa_map = {
                'vigentes': 'Solo vigentes',
                'ganadas': 'Solo cerradas ganadas', 
                'perdidas': 'Solo cerradas perdidas'
            }
            for etapa in etapas:
                etapa_nombres.append(etapa_map.get(etapa, etapa))
            if etapa_nombres:
                filters_applied.append(f"Estado(s): {', '.join(etapa_nombres)}")
                
        # Filtro de área
        if area_filter:
            filters_applied.append(f"Área: {area_filter}")
        
        if filters_applied:
            ws['B3'] = "; ".join(filters_applied)
        else:
            ws['B3'] = "Ninguno"
            
        # --- END: Add Report Metadata ---

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define all headers (brands + months)
        all_headers = [
            'OPORTUNIDAD', 'CLIENTE', 'AREA', 'CONTACTO', 'ZEBRA', 'PANDUIT', 'APC', 'AVIGILON', 
            'GENETEC', 'AXIS', 'SOFTWARE', 'RUNRATE', 'PÓLIZA', 'CISCO',
            'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEPT', 'OCT', 'NOV', 'DIC', 'ESTATUS', 'EMPLEADO'
        ]
        
        # Write all headers in a single row (starting at row 5)
        header_row_num = 5
        for col, header in enumerate(all_headers, 1):
            cell = ws.cell(row=header_row_num, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
    else:
        # Fallback to CSV - simplified structure
        headers = [
            'OPORTUNIDAD', 'CLIENTE', 'AREA', 'CONTACTO', 'ZEBRA', 'PANDUIT', 'APC', 'AVIGILON', 
            'GENETEC', 'AXIS', 'SOFTWARE', 'RUNRATE', 'PÓLIZA', 'CISCO',
            'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEPT', 'OCT', 'NOV', 'DIC', 'ESTATUS', 'EMPLEADO'
        ]
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="reporte_cotizaciones_oportunidades.csv"'
        
        import csv
        writer = csv.writer(response)
        writer.writerow(headers)
        
    # Brand columns mapping - usar las marcas reales de PRODUCTO_CHOICES
    brand_columns = {
        'ZEBRA': 5, 'PANDUIT': 6, 'APC': 7, 'AVIGILON': 8,
        'GENETEC': 9, 'AXIS': 10, 'SOFTWARE': 11, 'RUNRATE': 12, 
        'PÓLIZA': 13, 'CISCO': 14
    }
        
    # Debug: Print item count and sample data
    items_count = items.count()
    print(f"DEBUG: Total items found: {items_count}")
    
    if items_count > 0:
        # Show first item fields for debugging
        first_item = items.first()
        print(f"DEBUG: First item fields:")
        print(f"  - oportunidad: '{first_item.oportunidad}'")
        print(f"  - mes_cierre: '{first_item.mes_cierre}' (type: {type(first_item.mes_cierre)})")
        print(f"  - probabilidad_cierre: {first_item.probabilidad_cierre} (type: {type(first_item.probabilidad_cierre)})")
        print(f"  - monto: {first_item.monto}")
        print(f"  - area: '{first_item.area}'")
        print(f"  - producto: '{first_item.producto}'")
    else:
        print("DEBUG: No items found - check your filters!")
    
    # Write data rows (start at row 6 for Excel since we now have metadata headers)
    row = 6
    for item in items:
        print(f"DEBUG: Processing item: {item.oportunidad}")
        # Get cotization details for this opportunity
        cotizaciones = item.cotizaciones.all()
        
        # Para cada oportunidad, el monto va en la columna del producto/área de la oportunidad
        # No en las cotizaciones, sino en el producto/área de la oportunidad misma
        oportunidad_producto = item.get_producto_display() if hasattr(item, 'get_producto_display') else ''
        oportunidad_monto = float(item.monto) if item.monto else 0
        
        # Mapear el producto de la oportunidad a las marcas disponibles
        brand_totals = {}
        if oportunidad_producto and oportunidad_monto > 0:
            # Detectar marca por el producto de la oportunidad
            producto_upper = oportunidad_producto.upper()
            marca_detectada = None
            
            for brand in brand_columns.keys():
                if brand.upper() in producto_upper:
                    marca_detectada = brand
                    break
            
            # Si no se detecta por nombre, usar mapeo específico
            if not marca_detectada:
                producto_mappings = {
                    'ZEBRA': ['ZEBRA'],
                    'PANDUIT': ['PANDUIT'],
                    'APC': ['APC', 'UPS'],
                    'AVIGILION': ['AVIGILION', 'CCTV', 'CAMARA'],
                    'GENETEC': ['GENETEC', 'SEGURIDAD'],
                    'AXIS': ['AXIS'],
                    'SOFTWARE': ['SOFTWARE', 'APP', 'DESARROLLO'],
                    'RUNRATE': ['RUNRATE', 'RUN RATE'],
                    'PÓLIZA': ['PÓLIZA', 'POLIZA', 'SEGURO'],
                    'CISCO': ['CISCO', 'NETWORKING', 'RED']
                }
                
                for brand, keywords in producto_mappings.items():
                    if any(keyword in producto_upper for keyword in keywords):
                        marca_detectada = brand
                        break
            
            # Si se detectó una marca, asignar el monto
            if marca_detectada:
                brand_totals[marca_detectada] = oportunidad_monto
        
        # Prepare row data
        row_data = [
            item.oportunidad,
            item.cliente.nombre_empresa if item.cliente else '',
            item.get_area_display(),
            str(item.contacto) if item.contacto else ''
        ]
        
        # Add brand amounts - usar las marcas correctas
        for brand in ['ZEBRA', 'PANDUIT', 'APC', 'AVIGILION', 'GENETEC', 'AXIS', 'SOFTWARE', 'RUNRATE', 'PÓLIZA', 'CISCO']:
            amount = brand_totals.get(brand, 0)
            row_data.append(f"${amount:,.2f}" if amount > 0 else '')
        
        # Add monthly data - LA PROBABILIDAD VA EN EL MES DE CIERRE
        months = [''] * 12
        
        # Extraer mes_cierre y probabilidad_cierre correctamente
        mes_cierre_valor = item.mes_cierre  # CharField con valores como '01', '02', etc.
        probabilidad_valor = item.probabilidad_cierre  # IntegerField
        
        print(f"DEBUG: Raw mes_cierre: '{mes_cierre_valor}' (type: {type(mes_cierre_valor)})")
        print(f"DEBUG: Raw probabilidad_cierre: {probabilidad_valor} (type: {type(probabilidad_valor)})")
        
        # Convertir mes_cierre a índice (0-11) para el array de meses
        if mes_cierre_valor and mes_cierre_valor.strip():
            # Crear mapeo de nombres de meses a números
            mes_nombres_a_numeros = {
                'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
                'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
                'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
            }
            
            mes_str = mes_cierre_valor.strip()
            
            # Intentar convertir usando el mapeo de nombres
            if mes_str in mes_nombres_a_numeros:
                mes = mes_nombres_a_numeros[mes_str]
                print(f"DEBUG: mes_cierre '{mes_str}' mapped to int: {mes}")
                
                # Formatear la probabilidad
                if probabilidad_valor is not None:
                    probabilidad_str = f"{probabilidad_valor}%"
                else:
                    probabilidad_str = "0%"  # Default si no hay probabilidad
                
                # Asignar al mes correspondiente (mes-1 porque el array es 0-indexed)
                month_names = ['ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEPT','OCT','NOV','DIC']
                months[mes - 1] = probabilidad_str
                print(f"DEBUG: Setting month {mes} ({month_names[mes-1]}) to {probabilidad_str}")
            else:
                # Intentar como número directo (fallback)
                try:
                    mes = int(mes_str)
                    if 1 <= mes <= 12:
                        if probabilidad_valor is not None:
                            probabilidad_str = f"{probabilidad_valor}%"
                        else:
                            probabilidad_str = "0%"
                        
                        month_names = ['ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEPT','OCT','NOV','DIC']
                        months[mes - 1] = probabilidad_str
                        print(f"DEBUG: Setting month {mes} ({month_names[mes-1]}) to {probabilidad_str}")
                    else:
                        print(f"DEBUG: mes_cierre {mes} is out of range (1-12)")
                except (ValueError, TypeError):
                    print(f"DEBUG: Unable to convert mes_cierre '{mes_str}' - not a recognized month name or number")
        else:
            print(f"DEBUG: mes_cierre is None, empty or whitespace: '{mes_cierre_valor}'")
        
        print(f"DEBUG: Final months array: {months}")
        row_data.extend(months)
        
        # Add estatus (etapa_corta) before empleado
        estatus = getattr(item, 'etapa_corta', '') or ''
        row_data.append(estatus)
        
        row_data.append(item.usuario.get_full_name() or item.usuario.username if item.usuario else '')
        
        if OPENPYXL_AVAILABLE:
            # Excel version
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                # Format currency columns (brand columns are 4-13)
                if col >= 4 and col <= 13 and value and value != '':
                    try:
                        numeric_value = float(value.replace('$', '').replace(',', ''))
                        cell.value = numeric_value
                        cell.number_format = '$#,##0.00'
                    except:
                        pass
        else:
            # CSV version
            writer.writerow(row_data)
        
        row += 1
    
    if OPENPYXL_AVAILABLE:
        # Auto-adjust column widths (14 main headers + 12 month columns + 1 estatus + 1 empleado = 28 total)
        total_columns = 14 + 12 + 1 + 1  # main headers + monthly columns + estatus + empleado
        for col in range(1, total_columns + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 12
            
        # Make first column wider for opportunity names
        ws.column_dimensions['A'].width = 30

        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_cotizaciones_oportunidades.xlsx"'
        
        wb.save(response)
    
    return response

@login_required
def oportunidades_por_cliente_view(request, cliente_id):
    """
    View to display sales opportunities for a specific client.
    """
    cliente = get_object_or_404(Cliente, pk=cliente_id)

    if is_supervisor(request.user):
        oportunidades = TodoItem.objects.filter(cliente=cliente).order_by('-fecha_creacion')
    else:
        oportunidades = TodoItem.objects.filter(cliente=cliente, usuario=request.user).order_by('-fecha_creacion')

    context = {
        'cliente_id': cliente_id,
        'cliente_nombre': cliente.nombre_empresa, # Use nombre_empresa
        'oportunidades': oportunidades,
    }
    return render(request, 'oportunidades_por_cliente.html', context)


from django.views.decorators.clickjacking import xframe_options_exempt, xframe_options_sameorigin

@xframe_options_exempt
@login_required
def bitrix_widget_launcher(request):
    return render(request, 'bitrix_widget_launcher.html')

from django.http import HttpResponseRedirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def bitrix_cotizador_redirect(request):
    """
    Renderiza una plantilla con JavaScript para redirigir al cotizador en una nueva pestaña.
    """
    cotizador_url = "https://nethive.mx/app/crear-cotizacion/"
    print(f"DEBUG: Renderizando plantilla de redirección para: {cotizador_url}", flush=True)
    return render(request, 'bitrix_redirect_cotizador.html', {'cotizador_url': cotizador_url})


@login_required
@csrf_exempt
@xframe_options_sameorigin
def crear_cotizacion_view(request, cliente_id=None, oportunidad_id=None):
    cliente_seleccionado = None
    oportunidad_seleccionada = None
    
    # Detectar si viene de la detección automática de oportunidades (Crown Jewel Feature)
    is_auto_filled = request.GET.get('auto_filled') == 'true'
    auto_fill_data = {}
    
    if is_auto_filled:
        # Obtener datos de auto-llenado desde los parámetros GET
        auto_fill_data = {
            'titulo': request.GET.get('titulo', ''),
            'monto_estimado': request.GET.get('monto_estimado', ''),
            'oportunidad_id': request.GET.get('oportunidad_id', ''),
            'cliente_id': request.GET.get('cliente_id', '')
        }
        print(f"DEBUG: Auto-fill detectado desde notificación de oportunidad: {auto_fill_data}")
        
        # Si viene de auto-fill, usar los IDs de los parámetros
        if not oportunidad_id and auto_fill_data['oportunidad_id']:
            oportunidad_id = auto_fill_data['oportunidad_id']
        if not cliente_id and auto_fill_data['cliente_id']:
            cliente_id = auto_fill_data['cliente_id']

    print(f"DEBUG: crear_cotizacion_view - Request method: {request.method}")
    print(f"DEBUG: crear_cotizacion_view - GET parameters: {dict(request.GET)}")
    print(f"DEBUG: crear_cotizacion_view - oportunidad_id inicial: {oportunidad_id}")
    
    # Si no viene oportunidad_id como parámetro de URL, verificar en GET parameters
    if not oportunidad_id and request.GET.get('oportunidad_id'):
        oportunidad_id = request.GET.get('oportunidad_id')
        print(f"DEBUG: crear_cotizacion_view - oportunidad_id obtenido de GET params: {oportunidad_id}")
    
    if oportunidad_id:
        try:
            oportunidad_seleccionada = TodoItem.objects.get(id=oportunidad_id)
            # Si se selecciona una oportunidad, el cliente de la oportunidad es el cliente de la cotización
            if oportunidad_seleccionada.cliente:
                cliente_seleccionado = oportunidad_seleccionada.cliente
                cliente_id = str(oportunidad_seleccionada.cliente.id) # Asegurarse de que cliente_id se actualice
        except TodoItem.DoesNotExist:
            messages.error(request, f"La oportunidad con ID {oportunidad_id} no fue encontrada.")
            oportunidad_id = None # Reset para evitar errores posteriores

    # Si cliente_id fue pasado directamente (no a través de oportunidad_id) y no hay cliente_seleccionado aún
    if cliente_id and not cliente_seleccionado:
        try:
            # CORRECTO: La URL pasa el ID de Django, no el bitrix_company_id
            cliente_seleccionado = Cliente.objects.get(id=cliente_id)
            print(f"DEBUG: crear_cotizacion_view - Cliente encontrado por ID Django: {cliente_seleccionado.nombre_empresa}")
        except Cliente.DoesNotExist:
            print(f"DEBUG: crear_cotizacion_view - Cliente con ID Django {cliente_id} no encontrado")
            messages.error(request, f"El cliente con ID {cliente_id} no fue encontrado.")
            return redirect('crear_cotizacion')

    if cliente_seleccionado:
        print(f"DEBUG: crear_cotizacion_view - Cliente seleccionado: {cliente_seleccionado.nombre_empresa}")
    else:
        print("DEBUG: crear_cotizacion_view - No hay cliente seleccionado.")

    if request.method == 'POST':
        try:
            print(f"DEBUG: Procesando POST request para crear cotización")
            print(f"DEBUG: Usuario: {request.user}")
            print(f"DEBUG: POST data keys: {list(request.POST.keys())}")
            
            # Instantiate the form with POST data, and provide the user-specific queryset for the cliente field
            form = CotizacionForm(request.POST, user=request.user)
            print(f"DEBUG: Formulario creado exitosamente")
        except Exception as e:
            print(f"ERROR: Error general en POST request: {e}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'errors': {'__all__': [{'message': f'Error al procesar solicitud: {str(e)}'}]}}, status=500)

        if form.is_valid():
            print(f"DEBUG: Formulario es válido")
            cotizacion = form.save(commit=False)
            cotizacion.created_by = request.user  # Asignar el usuario creador
            cotizacion.save()
            # Actualizar fecha_actualizacion de la oportunidad vinculada para que suba en la tabla CRM
            if cotizacion.oportunidad:
                cotizacion.oportunidad.save(update_fields=['fecha_actualizacion'])
            print(f"DEBUG: Quote saved with ID: {cotizacion.id}")

            # Collect product data sent from JavaScript
            productos_data = {}
            for key, value in request.POST.items():
                if key.startswith('productos['):
                    parts = key.split('[')
                    index = parts[1].split(']')[0]
                    field = parts[2].split(']')[0]

                    if index not in productos_data:
                        productos_data[index] = {}
                    productos_data[index][field] = value

            # También recopilar títulos para crear orden combinado
            titulos_data = {}
            for key, value in request.POST.items():
                if key.startswith('titulos['):
                    parts = key.split('[')
                    index = parts[1].split(']')[0]
                    field = parts[2].split(']')[0]
                    if index not in titulos_data:
                        titulos_data[index] = {}
                    titulos_data[index][field] = value
            
            # Para nuevas cotizaciones: crear lista combinada en orden del DOM
            # PROBLEMA: títulos y productos tienen sistemas de posición diferentes
            # SOLUCIÓN: usar un mapa de posición unificado
            elementos_combinados = []
            
            # Crear lista de todos los elementos con sus posiciones
            elementos_con_posicion = []
            
            # SOLUCIÓN SIMPLE: Intercalar títulos y productos según aparecen en el DOM
            # Los títulos ya tienen sus posiciones DOM reales, los productos deben ir donde corresponden
            
            # Obtener títulos con sus posiciones
            titulos_con_posicion = []
            for titulo_data in titulos_data.values():
                if titulo_data.get('type') == 'title' and titulo_data.get('texto'):
                    pos = int(titulo_data.get('position', 999))
                    titulos_con_posicion.append((pos, titulo_data))
            titulos_con_posicion.sort()  # Ordenar por posición
            
            # Productos en orden
            productos_lista = [(int(k), v) for k, v in productos_data.items()]
            productos_lista.sort()
            
            print(f"DEBUG INTERCALAR: {len(titulos_con_posicion)} títulos, {len(productos_lista)} productos")
            for pos, titulo in titulos_con_posicion:
                print(f"DEBUG INTERCALAR: Título '{titulo.get('texto')}' en posición DOM {pos}")
            
            # Nueva lógica: crear lista completa con títulos y productos en sus posiciones reales
            elementos_temporales = []
            
            # Agregar títulos con sus posiciones reales
            for pos_titulo, titulo_data in titulos_con_posicion:
                elementos_temporales.append({
                    'tipo': 'titulo',
                    'posicion_dom': pos_titulo,
                    'datos': titulo_data,
                    'nombre': titulo_data.get('texto', 'SIN_TITULO')
                })
                print(f"DEBUG INTERCALAR: Título '{titulo_data.get('texto')}' en posición DOM {pos_titulo}")
            
            # Agregar productos con posiciones secuenciales donde NO hay títulos
            producto_index = 0
            
            # Obtener todas las posiciones ocupadas por títulos
            posiciones_titulos = set(pos for pos, _ in titulos_con_posicion)
            
            # Recorrer todas las posiciones posibles y llenar con productos donde no hay títulos
            max_posicion = max(len(productos_lista) + len(titulos_con_posicion), max(posiciones_titulos, default=0) + 1)
            for pos in range(max_posicion):
                if pos not in posiciones_titulos and producto_index < len(productos_lista):
                    _, producto_data = productos_lista[producto_index]
                    elementos_temporales.append({
                        'tipo': 'producto',
                        'posicion_dom': pos,
                        'datos': producto_data,
                        'nombre': producto_data.get('nombre_producto', 'SIN_NOMBRE')
                    })
                    print(f"DEBUG INTERCALAR: Producto '{producto_data.get('nombre_producto')}' en posición DOM {pos}")
                    producto_index += 1
            
            # Ordenar por posición DOM y asignar posiciones finales secuenciales
            elementos_temporales.sort(key=lambda x: x['posicion_dom'])
            
            posicion_final = 0
            for elemento in elementos_temporales:
                elemento['posicion'] = posicion_final
                elementos_con_posicion.append({
                    'tipo': elemento['tipo'],
                    'posicion': posicion_final,
                    'datos': elemento['datos'],
                    'nombre': elemento['nombre']
                })
                print(f"DEBUG INTERCALAR: Agregado {elemento['tipo']} '{elemento['nombre']}' en posición final {posicion_final}")
                posicion_final += 1
            
            # Assign a new, sequential 'orden' value based on the sorted list
            # This loop modifies the dictionaries in elementos_con_posicion in place
            for i, elemento in enumerate(elementos_con_posicion):
                elemento['posicion_final'] = i 

            print(f"DEBUG POSITION: Total elementos encontrados: {len(elementos_con_posicion)}")
            print(f"DEBUG POSITION: Elementos ordenados y con posicion_final:")
            for i, elem in enumerate(elementos_con_posicion):
                print(f"  {i+1}. {elem['tipo'].upper()}: '{elem['nombre']}' (posición: {elem['posicion']}), posicion_final: {elem['posicion_final']}")
            
            # Crear lista final (now just copy the modified dictionaries)
            elementos_combinados = elementos_con_posicion # No need to create new dictionaries, just use the modified ones
            
            print(f"DEBUG ORDER: Elementos en orden: {len(elementos_combinados)} elementos")
            for i, elemento in enumerate(elementos_combinados):
                if elemento['tipo'] == 'titulo':
                    print(f"DEBUG ORDER: {i+1}. TÍTULO: '{elemento['datos'].get('texto', 'SIN_TEXTO')}' (posición: {elemento['posicion']}), posicion_final: {elemento['posicion_final']}")
                else:
                    print(f"DEBUG ORDER: {i+1}. PRODUCTO: '{elemento['datos'].get('nombre_producto', 'SIN_NOMBRE')}' (posición: {elemento['posicion']}), posicion_final: {elemento['posicion_final']}")
            
            # Mantener productos_list para compatibilidad con el cálculo de totales
            productos_list = [productos_data[key] for key in sorted(productos_data.keys(), key=int)]
            print(f"DEBUG: productos_list before saving details: {productos_list}")

            calculated_subtotal = Decimal('0.00')
            for item_data in productos_list:
                try:
                    cantidad = int(item_data.get('cantidad', 0))
                    
                    # Manejo seguro del precio
                    precio_str = str(item_data.get('precio', '0.00')).strip()
                    if not precio_str or precio_str == '':
                        precio_str = '0.00'
                    precio = Decimal(precio_str)
                    
                    # Manejo seguro del descuento
                    descuento_str = str(item_data.get('descuento', '0.00')).strip()
                    if not descuento_str or descuento_str == '':
                        descuento_str = '0.00'
                    descuento = Decimal(descuento_str)
                    
                    item_total = cantidad * precio
                    item_total -= item_total * (descuento / Decimal('100.00'))
                    calculated_subtotal += item_total.quantize(Decimal('0.01'))
                    print(f"DEBUG_CALC: Item: {item_data.get('nombre')}, Quantity: {cantidad}, Price: {precio}, Discount: {descuento}, Item Total (rounded): {item_total.quantize(Decimal('0.01'))}")
                except (ValueError, TypeError, decimal.InvalidOperation) as e:
                    print(f"DEBUG_ERROR: Error processing item {item_data}: {e}")
                    cotizacion.delete()
                    return JsonResponse({'success': False, 'errors': {'__all__': [{'message': f'Invalid product data in row. Error: {e}'}]}}, status=400)

            cotizacion.subtotal = calculated_subtotal.quantize(Decimal('0.01'))
            
            try:
                iva_rate_str = str(request.POST.get('iva_rate', '0.00')).strip()
                if not iva_rate_str or iva_rate_str == '':
                    iva_rate_str = '0.00'
                cotizacion.iva_rate = Decimal(iva_rate_str)
            except (ValueError, TypeError, decimal.InvalidOperation):
                cotizacion.iva_rate = Decimal('0.00')

            cotizacion.iva_amount = (cotizacion.subtotal * cotizacion.iva_rate).quantize(Decimal('0.01'))
            cotizacion.total = (cotizacion.subtotal + cotizacion.iva_amount).quantize(Decimal('0.01'))

            descuento_visible_str = request.POST.get('descuento_visible', 'true')
            cotizacion.descuento_visible = descuento_visible_str.lower() == 'true'
            cotizacion.tipo_cotizacion = request.POST.get('tipo_cotizacion')
            
            cotizacion.save(update_fields=['subtotal', 'iva_rate', 'iva_amount', 'total', 'descuento_visible', 'tipo_cotizacion', 'oportunidad'])
            print(f"DEBUG: Quote totals updated. Subtotal: {cotizacion.subtotal}, IVA: {cotizacion.iva_amount}, Total: {cotizacion.total}, Quote Type: {cotizacion.tipo_cotizacion}")
            
            # Guardar elementos en orden correcto (títulos Y productos)
            for elemento in elementos_combinados:
                if elemento['tipo'] == 'titulo':
                    titulo_data = elemento['datos']
                    try:
                        if titulo_data.get('type') == 'title' and titulo_data.get('texto'):
                            DetalleCotizacion.objects.create(
                                cotizacion=cotizacion,
                                nombre_producto=titulo_data.get('texto', ''),
                                descripcion=titulo_data.get('texto', ''),
                                cantidad=0,
                                precio_unitario=Decimal('0.00'),
                                descuento_porcentaje=Decimal('0.00'),
                                marca='',
                                no_parte='',
                                tipo='titulo',
                                orden=elemento['posicion_final']
                            )
                            print(f"DEBUG ORDER: Title created: {titulo_data.get('texto')} with orden {elemento['posicion_final']}")
                    except Exception as e:
                        print(f"WARNING: Error creating title {titulo_data}: {e}")
                        
                else:  # producto
                    item_data = elemento['datos']
                    try:
                        # Manejo seguro del precio unitario
                        precio_str = str(item_data.get('precio', '0.00')).strip()
                        if not precio_str or precio_str == '':
                            precio_str = '0.00'
                        precio_unitario = Decimal(precio_str)
                        
                        # Manejo seguro del descuento
                        descuento_str = str(item_data.get('descuento', '0.00')).strip()
                        if not descuento_str or descuento_str == '':
                            descuento_str = '0.00'
                        descuento_porcentaje = Decimal(descuento_str)
                        
                        DetalleCotizacion.objects.create(
                            cotizacion=cotizacion,
                            nombre_producto=item_data.get('nombre_producto', ''),
                            descripcion=item_data.get('descripcion', ''),
                            cantidad=int(item_data.get('cantidad', 1)),
                            precio_unitario=precio_unitario,
                            descuento_porcentaje=descuento_porcentaje,
                            marca=item_data.get('marca', ''),
                            no_parte=item_data.get('no_parte', ''),
                            tipo='producto',
                            orden=elemento['posicion_final']
                        )
                        print(f"DEBUG ORDER: Product created: {item_data.get('nombre_producto')} with orden {elemento['posicion_final']}")
                    except (ValueError, TypeError, decimal.InvalidOperation) as e:
                        cotizacion.delete()
                        return JsonResponse({'success': False, 'errors': {'__all__': [{'message': f'Invalid product data in row. Error: {e}'}]}}, status=400)

            # Los títulos ya se procesaron en orden combinado arriba
            print(f"DEBUG: Todos los elementos (productos y títulos) fueron guardados en orden correcto")
            
            # Check if this is a widget_mode request (AJAX from iframe)
            is_widget_mode = request.POST.get('widget_mode') == '1' or request.GET.get('widget_mode') == '1'

            if is_widget_mode:
                # Return JSON response for the widget iframe
                from django.urls import reverse
                pdf_url = reverse('generate_cotizacion_pdf', args=[cotizacion.id])
                return JsonResponse({
                    'success': True,
                    'cotizacion_id': cotizacion.id,
                    'pdf_url': pdf_url,
                })

            # Si hay una oportunidad seleccionada, redirigir a página de descarga y luego a cotizaciones por oportunidad
            # sino, al PDF directo como antes
            oportunidad_para_redirect = form.cleaned_data.get('oportunidad')
            if oportunidad_para_redirect:
                # Crear URL que descarga PDF y luego redirige a cotizaciones por oportunidad
                from django.urls import reverse
                redirect_url = reverse('download_and_redirect_cotizacion',
                                     args=[cotizacion.id, oportunidad_para_redirect.id])
                return redirect(redirect_url)
            else:
                # Flujo original para cotizaciones sin oportunidad
                return redirect('generate_cotizacion_pdf', cotizacion_id=cotizacion.id)

        else:
            errors_dict = {}
            for field, field_errors in form.errors.items():
                errors_dict[field] = [{'message': str(e), 'code': e.code if hasattr(e, 'code') else 'invalid'} for e in field_errors]
            
            # Enhanced debugging information
            print(f"DEBUG: Form validation failed for crear_cotizacion_view")
            print(f"DEBUG: Form errors: {errors_dict}")
            print(f"DEBUG: Form data received: {dict(request.POST.items())}")
            print(f"DEBUG: Form cleaned_data (if available): {getattr(form, 'cleaned_data', 'N/A')}")
            print(f"DEBUG: Form non_field_errors: {form.non_field_errors()}")
            print(f"DEBUG: User: {request.user.username if request.user.is_authenticated else 'Anonymous'}")
            print(f"DEBUG: Request method: {request.method}")
            print(f"DEBUG: Content type: {request.content_type}")
            
            return JsonResponse({'success': False, 'errors': errors_dict}, status=400)

    else: # If the request is GET
        initial_data = {}
        if oportunidad_seleccionada:
            initial_data['titulo'] = oportunidad_seleccionada.oportunidad
            initial_data['cliente'] = oportunidad_seleccionada.cliente.id if oportunidad_seleccionada.cliente else None
            initial_data['oportunidad'] = oportunidad_seleccionada.id
            
        # Si viene de auto-fill (Crown Jewel Feature), usar ese título preferentemente
        if is_auto_filled and auto_fill_data.get('titulo'):
            initial_data['titulo'] = f"Cotización - {auto_fill_data['titulo']}"
            print(f"DEBUG: Título auto-llenado: {initial_data['titulo']}")

        form = CotizacionForm(initial=initial_data, user=request.user)
        from django.urls import reverse
        form_action_url = reverse('crear_cotizacion_with_id', args=[cliente_id]) if cliente_id else reverse('crear_cotizacion')
        print(f"DEBUG: crear_cotizacion_view - Generated form_action_url: {form_action_url}")

        # Todos los usuarios pueden ver todos los clientes para la creación de cotizaciones
        clientes_django = Cliente.objects.all().order_by('nombre_empresa')

        print(f"DEBUG: crear_cotizacion_view - Clientes obtenidos de Django: {len(clientes_django)}")

        clientes_data_json = []
        for c in clientes_django:
            clientes_data_json.append({
                'id': str(c.id),
                'name': c.nombre_empresa,
            })
        print(f"DEBUG: crear_cotizacion_view - Clientes serializados para JSON: {clientes_data_json}")

        context = {
            'form': form,
            'cliente_seleccionado': cliente_seleccionado,
            'clientes_data_json': json.dumps(clientes_data_json),
            'cliente_id_inicial': cliente_id,
            'oportunidad_id_inicial': oportunidad_id,
            'form_action_url': form_action_url,
            'producto_choices': TodoItem.PRODUCTO_CHOICES,
            'area_choices': TodoItem.AREA_CHOICES,
            'probabilidad_choices_list': [i for i in range(0, 101, 10)],
            # Crown Jewel Feature: Auto-fill data from opportunity detection
            'is_auto_filled': is_auto_filled,
            'auto_fill_data': auto_fill_data,
            'oportunidad_seleccionada': oportunidad_seleccionada,
            'widget_mode': request.GET.get('widget_mode') == '1',
        }
        return render(request, 'crear_cotizacion.html', context)


@login_required
@xframe_options_sameorigin
def editar_cotizacion_view(request, cotizacion_id):
    cotizacion_original = get_object_or_404(Cotizacion, pk=cotizacion_id)
    detalles_originales = DetalleCotizacion.objects.filter(cotizacion=cotizacion_original).order_by('id')

    # Formatear detalles para el JavaScript del template
    detalles_list = [{
        'nombre_producto': d.nombre_producto,
        'marca': d.marca,
        'no_parte': d.no_parte,
        'descripcion': d.descripcion,
        'cantidad': d.cantidad,
        'precio': str(d.precio_unitario),
        'descuento': str(d.descuento_porcentaje),
        'tipo': getattr(d, 'tipo', 'producto') or 'producto',  # Incluir el tipo (producto o titulo)
    } for d in detalles_originales]

    # Obtener todos los clientes para el dropdown
    if is_supervisor(request.user):
        clientes_django = Cliente.objects.all().order_by('nombre_empresa')
    else:
        clientes_django = Cliente.objects.filter(Q(asignado_a=request.user) | Q(asignado_a__isnull=True)).order_by('nombre_empresa')

    clientes_data_json = []
    for c in clientes_django:
        clientes_data_json.append({
            'id': str(c.id),
            'name': c.nombre_empresa,
        })

    # Obtener oportunidades del cliente para el dropdown
    oportunidades_data_json = []
    if cotizacion_original.cliente:
        if is_supervisor(request.user):
            oportunidades = TodoItem.objects.filter(cliente=cotizacion_original.cliente).order_by('-fecha_creacion')
        else:
            oportunidades = TodoItem.objects.filter(cliente=cotizacion_original.cliente, usuario=request.user).order_by('-fecha_creacion')
        
        for o in oportunidades:
            oportunidades_data_json.append({
                'id': str(o.id),
                'title': o.oportunidad,
                'monto': str(o.monto),
            })

    # URL del formulario apunta siempre a crear_cotizacion estándar para reutilizar la lógica
    from django.urls import reverse
    form_action_url = reverse('crear_cotizacion')

    # Datos iniciales para el formulario
    initial_data = {
        'titulo': cotizacion_original.titulo,
        'cliente': cotizacion_original.cliente.id,
        'usuario_final': cotizacion_original.usuario_final,
        'oportunidad': cotizacion_original.oportunidad.id if cotizacion_original.oportunidad else None,
        'descripcion': cotizacion_original.descripcion,
        'comentarios': cotizacion_original.comentarios,
        'nombre_cotizacion': cotizacion_original.nombre_cotizacion,
        'iva_rate': cotizacion_original.iva_rate,
        'moneda': cotizacion_original.moneda,
        'tipo_cotizacion': cotizacion_original.tipo_cotizacion,
        'descuento_visible': cotizacion_original.descuento_visible,
    }

    is_widget = request.GET.get('widget_mode') == '1'
    context = {
        'form': CotizacionForm(initial=initial_data, user=request.user),
        'cliente_seleccionado': cotizacion_original.cliente,
        'clientes_data_json': json.dumps(clientes_data_json),
        'oportunidades_data_json': json.dumps(oportunidades_data_json),
        'cliente_id_inicial': cotizacion_original.cliente.id,
        'oportunidad_id_inicial': cotizacion_original.oportunidad.id if cotizacion_original.oportunidad else None,
        'detalles_cotizacion': json.dumps(detalles_list),  # Convertir a JSON para el template
        'form_action_url': form_action_url,
        'editing_mode': True,  # Indicador para el template
        'cotizacion_original': cotizacion_original,  # Datos de la cotización original
        'probabilidad_choices_list': [i for i in range(0, 101, 10)],
        'widget_mode': is_widget,
    }

    return render(request, 'crear_cotizacion.html', context)



@login_required
def download_and_redirect_cotizacion(request, cotizacion_id, oportunidad_id):
    """
    Vista que muestra una página que descarga el PDF automáticamente y luego redirige 
    a la página de cotizaciones por oportunidad.
    """
    print(f"DEBUG: download_and_redirect_cotizacion - cotizacion_id: {cotizacion_id}, oportunidad_id: {oportunidad_id}")
    
    # Verificar que la cotización existe y el usuario tiene permisos
    cotizacion = get_object_or_404(Cotizacion, pk=cotizacion_id)
    if not is_supervisor(request.user) and cotizacion.created_by != request.user:
        messages.error(request, "No tienes permisos para descargar esta cotización.")
        return redirect('cotizaciones')
    
    # Verificar que la oportunidad existe
    oportunidad = get_object_or_404(TodoItem, pk=oportunidad_id)
    if not is_supervisor(request.user) and oportunidad.usuario != request.user:
        messages.error(request, "No tienes permisos para acceder a esta oportunidad.")
        return redirect('todos')
    
    # URLs necesarias para la página de descarga y redirección
    pdf_download_url = reverse('generate_cotizacion_pdf', args=[cotizacion_id])
    redirect_url = reverse('cotizaciones_por_oportunidad', args=[oportunidad_id])
    
    context = {
        'cotizacion': cotizacion,
        'oportunidad': oportunidad,
        'pdf_download_url': pdf_download_url,
        'redirect_url': redirect_url
    }
    
    return render(request, 'download_and_redirect.html', context)

def generate_cotizacion_pdf(request, cotizacion_id):
    """
    View to generate the PDF of a specific quote.
    """
    print(f"DEBUG: Starting generate_cotizacion_pdf for quote ID: {cotizacion_id}")
    cotizacion = get_object_or_404(Cotizacion, pk=cotizacion_id)
    print(f"DEBUG: Quote found: {cotizacion.id} - Quote Type: {cotizacion.tipo_cotizacion}")
    
    # Ensure that the user has permission to view this quote
    if not is_supervisor(request.user) and cotizacion.created_by != request.user:
        print(f"DEBUG: Access denied for user {request.user.username} to quote {cotizacion.id}")
        return HttpResponse("Acceso denegado.", status=403)

    detalles_cotizacion = DetalleCotizacion.objects.filter(cotizacion=cotizacion).order_by('orden')
    iva_rate_percentage = (cotizacion.iva_rate * Decimal('100')).quantize(Decimal('1'))
    
    # DEBUG: Mostrar todos los detalles antes de procesar
    print(f"DEBUG: Total detalles encontrados: {len(detalles_cotizacion)}")
    for detalle in detalles_cotizacion:
        tipo_actual = getattr(detalle, 'tipo', 'NO_DEFINIDO')
        print(f"DEBUG: Detalle ID={detalle.id}, nombre='{detalle.nombre_producto}', tipo='{tipo_actual}'")
    
    # Organizar productos en secciones basadas en títulos
    secciones = []
    seccion_actual = None

    for detalle in detalles_cotizacion:
        tipo_detalle = getattr(detalle, 'tipo', 'producto') or 'producto'

        if tipo_detalle == 'titulo':
            if seccion_actual:
                secciones.append(seccion_actual)
            seccion_actual = {'titulo': detalle.nombre_producto, 'productos': []}
        else:  # Es un producto
            if not seccion_actual:
                # Si hay productos antes del primer título, se agrupan en una sección sin título.
                seccion_actual = {'titulo': None, 'productos': []}
            seccion_actual['productos'].append(detalle)

    # Agregar la última sección si existe
    if seccion_actual:
        secciones.append(seccion_actual)

    # Si después de todo no hay secciones pero sí detalles, se crea una sección por defecto.
    if not secciones and detalles_cotizacion:
        secciones.append({'titulo': None, 'productos': [d for d in detalles_cotizacion if getattr(d, 'tipo', 'producto') == 'producto']})
    
    print(f"DEBUG: ===== RESUMEN FINAL =====")
    print(f"DEBUG: Secciones organizadas: {len(secciones)} secciones encontradas")
    for i, seccion in enumerate(secciones):
        print(f"DEBUG: 📋 Sección {i+1}: titulo='{seccion['titulo']}', productos={len(seccion['productos'])}")
        for j, producto in enumerate(seccion['productos']):
            print(f"     📦 Producto {j+1}: {producto.nombre_producto} (tipo: {getattr(producto, 'tipo', 'NO_DEFINIDO')})")
    print(f"DEBUG: =========================")

    pdf_name_raw = cotizacion.nombre_cotizacion or f"Cotizacion_{cotizacion.id}"
    pdf_name = "".join(c for c in pdf_name_raw if c.isalnum() or c in ('_', '-')).strip().replace(' ', '_')
    if not pdf_name:
        pdf_name = f"Cotizacion_{cotizacion.id}"

    tipo_cotizacion = cotizacion.tipo_cotizacion
    logo_base64 = ""
    company_name = ""
    company_address = ""
    company_phone = ""
    company_email = ""
    template_name = 'cotizacion_pdf_template.html' # Default template

    if tipo_cotizacion and tipo_cotizacion.lower() == 'iamet':
        template_name = 'iamet_cotizacion_pdf_template.html'
        company_name = 'IAMET S.A. de C.V.'
        company_address = 'Av. Principal #456, Col. Centro, Guadalajara, Jalisco'
        company_phone = '+52 33 9876 5432'
        company_email = 'contacto@iamet.com'
    elif tipo_cotizacion and tipo_cotizacion.lower() == 'bajanet':
        template_name = 'cotizacion_pdf_template.html'
        company_name = 'BAJANET S.A. de C.V.'
        company_address = 'Calle Ficticia #123, Colonia Ejemplo, Ciudad de México'
        company_phone = '+52 55 1234 5678'
        company_email = 'ventas@bajanet.com'
        try:
            logo_url = "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRV1xCutWCicl-yXCjMjH5P5jTZA0R993cG9g&s"
            response = requests.get(logo_url)
            response.raise_for_status() # Raise an exception for HTTP errors
            logo_base64 = base64.b64encode(response.content).decode('utf-8')
        except requests.exceptions.RequestException as e:
            print(f"ERROR: Error fetching Bajanet logo from URL: {e}")
            logo_base64 = "" # Fallback to empty string if fetching fails
    else: # Fallback to Bajanet if type is not recognized or None
        template_name = 'cotizacion_pdf_template.html'
        company_name = 'BAJANET S.A. de C.V.'
        company_address = 'Calle Ficticia #123, Colonia Ejemplo, Ciudad de México'
        company_phone = '+52 55 1234 5678'
        company_email = 'ventas@bajanet.com'
        try:
            logo_url = "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRV1xCutWCicl-yXCjMjH5P5jTZA0R993cG9g&s"
            response = requests.get(logo_url)
            response.raise_for_status() # Raise an exception for HTTP errors
            logo_base64 = base64.b64encode(response.content).decode('utf-8')
        except requests.exceptions.RequestException as e:
            print(f"ERROR: Error fetching Bajanet logo from URL: {e}")
            logo_base64 = "" # Fallback to empty string if fetching fails


    context = {
        'cotizacion': cotizacion,
        'detalles_cotizacion': detalles_cotizacion,
        'secciones': secciones,
        'request_user': request.user,
        'current_date': date.today(),
        'company_name': company_name,
        'company_address': company_address,
        'company_phone': company_phone,
        'company_email': company_email,
        'logo_base64': logo_base64,
        'iva_rate_percentage': iva_rate_percentage,
    }

    try:
        print(f"DEBUG: Attempting to render template: {template_name}")
        html_string = render_to_string(template_name, context)
        print("DEBUG: Template rendered to HTML string.")
    except Exception as e:
        print(f"ERROR: Error rendering template '{template_name}': {e}")
        return HttpResponse(f"Internal server error rendering PDF: {e}", status=500)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{pdf_name}.pdf"'

    try:
        print("DEBUG: Attempting to generate PDF with WeasyPrint.")
        HTML(string=html_string).write_pdf(response)
        print("DEBUG: PDF generated successfully.")
    except Exception as e:
        print(f"ERROR: Error generating PDF with WeasyPrint: {e}")
        return HttpResponse(f"Internal server error generating PDF: {e}", status=500)
        
    return response


from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test

def supervisor_required(view_func):
    decorated_view_func = login_required(user_passes_test(is_supervisor)(view_func))
    return decorated_view_func


@login_required
def usuario_redirect_view(request):
    """
    Muestra la página de configuración avanzada y maneja el envío del formulario de avatar.
    """
    if request.method == 'POST':
        # Obtener o crear el perfil del usuario
        profile, created = UserProfile.objects.get_or_create(user=request.user)
        
        # Manejar la subida de imagen
        if 'avatar' in request.FILES and request.FILES['avatar']:
            profile.avatar = request.FILES['avatar']
            profile.usar_animado = False
            profile.avatar_tipo = '1'  # Reset a humano por defecto
        # Si se está seleccionando un avatar animado
        elif 'avatar_choice' in request.POST and request.POST['avatar_choice']:
            profile.usar_animado = True
            profile.avatar_tipo = request.POST['avatar_choice']  # Guardar el tipo seleccionado
            # Limpiar avatar anterior si existía
            profile.avatar = None
        
        profile.save()
        
        # Agregar mensaje de éxito
        messages.success(request, 'Avatar actualizado correctamente')
        
        # Redirigir a la configuración avanzada
        return redirect('configuracion_avanzada')
    
    # Si es GET, redirigir a configuración avanzada
    return redirect('configuracion_avanzada')


@supervisor_required
def reporte_usuarios(request):
    usuarios = User.objects.all().order_by('username')
    return render(request, 'reporte_usuarios.html', {'usuarios': usuarios})

@login_required
def perfil_usuario(request, usuario_id):
    from datetime import date
    from django.db.models import Sum, Count
    usuario = get_object_or_404(User, id=usuario_id)

    today = date.today()

    # --- Filtros de fecha (igual que estadisticas_usuarios) ---
    filtro = request.GET.get('filtro', 'todas')
    fecha_especifica = request.GET.get('fecha', None)
    mes_especifico = request.GET.get('mes', None)
    anio_especifico = request.GET.get('anio', None)

    oportunidades = TodoItem.objects.filter(usuario=usuario).select_related('cliente')
    cotizaciones = Cotizacion.objects.filter(created_by=usuario)

    titulo_filtro = "Todas las oportunidades"
    if filtro == 'dia':
        if fecha_especifica:
            try:
                fecha = date.fromisoformat(fecha_especifica)
            except ValueError:
                fecha = today
        else:
            fecha = today
        oportunidades = oportunidades.filter(fecha_creacion__date=fecha)
        cotizaciones = cotizaciones.filter(fecha_creacion__date=fecha)
        titulo_filtro = f"Día: {fecha.strftime('%d/%m/%Y')}"
    elif filtro == 'mes':
        if mes_especifico:
            try:
                año, mes = mes_especifico.split('-')
                año = int(año)
                mes = int(mes)
            except (ValueError, AttributeError):
                año = today.year
                mes = today.month
        else:
            año = today.year
            mes = today.month
        oportunidades = oportunidades.filter(fecha_creacion__year=año, fecha_creacion__month=mes)
        cotizaciones = cotizaciones.filter(fecha_creacion__year=año, fecha_creacion__month=mes)
        from calendar import month_name
        import locale
        try:
            locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        except:
            pass
        titulo_filtro = f"Mes: {today.replace(year=año, month=mes, day=1).strftime('%B %Y').capitalize()}"
    elif filtro == 'anio':
        if anio_especifico:
            try:
                año = int(anio_especifico)
            except ValueError:
                año = today.year
        else:
            año = today.year
        oportunidades = oportunidades.filter(fecha_creacion__year=año)
        cotizaciones = cotizaciones.filter(fecha_creacion__year=año)
        titulo_filtro = f"Año: {año}"

    # --- Estadísticas principales (mismas que la vista general) ---
    total_vendido = oportunidades.filter(probabilidad_cierre=100).aggregate(
        total=Sum('monto')
    )['total'] or Decimal('0.00')

    total_pipeline = oportunidades.aggregate(
        total=Sum('monto')
    )['total'] or Decimal('0.00')

    num_oportunidades = oportunidades.count()
    num_cotizaciones = cotizaciones.count()

    # Cliente top por monto
    cliente_top = oportunidades.values(
        'cliente__id', 'cliente__nombre_empresa'
    ).annotate(
        total_cliente=Sum('monto'),
        num_ops=Count('id')
    ).order_by('-total_cliente').first()

    # Oportunidad más grande (1-99% probabilidad)
    oportunidad_mayor = oportunidades.filter(
        probabilidad_cierre__gte=1, probabilidad_cierre__lte=99
    ).order_by('-monto').first()

    context = {
        'usuario': usuario,
        'oportunidades': oportunidades.order_by('-monto'),
        'oportunidad_mayor': oportunidad_mayor,
        'total_vendido': total_vendido,
        'total_pipeline': total_pipeline,
        'num_oportunidades': num_oportunidades,
        'num_cotizaciones': num_cotizaciones,
        'cliente_top': cliente_top,
        'filtro_actual': filtro,
        'titulo_filtro': titulo_filtro,
        'fecha_actual': today.isoformat(),
        'mes_actual': today.strftime('%Y-%m'),
        'anio_actual': str(today.year),
    }
    return render(request, 'perfil_usuario.html', context)


@login_required
def estadisticas_usuarios(request):
    """
    Vista para mostrar estadísticas de usuarios:
    - Número de oportunidades/cotizaciones por usuario
    - Total vendido por usuario
    - Cliente al que más venden
    Filtros: día, mes, período, todas
    """
    from datetime import date, timedelta
    from django.db.models import Sum, Count
    from django.db.models.functions import TruncMonth, TruncDay

    # Obtener parámetro de filtro
    filtro = request.GET.get('filtro', 'todas')
    fecha_especifica = request.GET.get('fecha', None)
    mes_especifico = request.GET.get('mes', None)
    fecha_desde = request.GET.get('fecha_desde', None)
    fecha_hasta = request.GET.get('fecha_hasta', None)

    today = date.today()

    # Base queryset de oportunidades
    oportunidades = TodoItem.objects.all()

    # Base queryset de cotizaciones (PDFs generados)
    cotizaciones = Cotizacion.objects.all()

    # Aplicar filtros
    titulo_filtro = "Todas las oportunidades"
    if filtro == 'dia':
        if fecha_especifica:
            try:
                fecha = date.fromisoformat(fecha_especifica)
            except ValueError:
                fecha = today
        else:
            fecha = today
        oportunidades = oportunidades.filter(fecha_creacion__date=fecha)
        cotizaciones = cotizaciones.filter(fecha_creacion__date=fecha)
        titulo_filtro = f"Día: {fecha.strftime('%d/%m/%Y')}"
    elif filtro == 'mes':
        if mes_especifico:
            try:
                año, mes = mes_especifico.split('-')
                año = int(año)
                mes = int(mes)
            except (ValueError, AttributeError):
                año = today.year
                mes = today.month
        else:
            año = today.year
            mes = today.month
        oportunidades = oportunidades.filter(fecha_creacion__year=año, fecha_creacion__month=mes)
        cotizaciones = cotizaciones.filter(fecha_creacion__year=año, fecha_creacion__month=mes)
        from calendar import month_name
        import locale
        try:
            locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        except:
            pass
        titulo_filtro = f"Mes: {today.replace(year=año, month=mes, day=1).strftime('%B %Y').capitalize()}"
    elif filtro == 'periodo':
        if fecha_desde and fecha_hasta:
            try:
                f_desde = date.fromisoformat(fecha_desde)
                f_hasta = date.fromisoformat(fecha_hasta)
            except ValueError:
                f_desde = today.replace(day=1)
                f_hasta = today
            oportunidades = oportunidades.filter(fecha_creacion__date__gte=f_desde, fecha_creacion__date__lte=f_hasta)
            cotizaciones = cotizaciones.filter(fecha_creacion__date__gte=f_desde, fecha_creacion__date__lte=f_hasta)
            titulo_filtro = f"Período: {f_desde.strftime('%d/%m/%Y')} — {f_hasta.strftime('%d/%m/%Y')}"
        else:
            titulo_filtro = "Todas las oportunidades"

    # Exportar a Excel
    if request.GET.get('exportar') == 'excel':
        return _exportar_estadisticas_excel(request, oportunidades, cotizaciones, titulo_filtro)

    # Obtener todos los usuarios que tienen oportunidades
    usuarios_con_datos = []
    usuarios = User.objects.filter(oportunidades__isnull=False).distinct()

    for usuario in usuarios:
        oportunidades_usuario = oportunidades.filter(usuario=usuario)

        # Número de oportunidades
        num_oportunidades = oportunidades_usuario.count()

        if num_oportunidades == 0:
            continue

        # Total vendido (oportunidades con probabilidad 100%)
        total_vendido = oportunidades_usuario.filter(probabilidad_cierre=100).aggregate(
            total=Sum('monto')
        )['total'] or Decimal('0.00')

        # Total en pipeline (todas las oportunidades)
        total_pipeline = oportunidades_usuario.aggregate(
            total=Sum('monto')
        )['total'] or Decimal('0.00')

        # Cliente al que más vende (por monto total)
        cliente_top = oportunidades_usuario.values(
            'cliente__id', 'cliente__nombre_empresa'
        ).annotate(
            total_cliente=Sum('monto'),
            num_ops=Count('id')
        ).order_by('-total_cliente').first()

        # Número de cotizaciones PDF generadas por este usuario
        num_cotizaciones = cotizaciones.filter(created_by=usuario).count()

        usuarios_con_datos.append({
            'usuario': usuario,
            'num_oportunidades': num_oportunidades,
            'num_cotizaciones': num_cotizaciones,
            'total_vendido': total_vendido,
            'total_pipeline': total_pipeline,
            'cliente_top': cliente_top,
        })

    # Ordenar por total vendido descendente
    usuarios_con_datos.sort(key=lambda x: x['total_vendido'], reverse=True)

    # Totales generales
    total_general_vendido = sum(u['total_vendido'] for u in usuarios_con_datos)
    total_general_pipeline = sum(u['total_pipeline'] for u in usuarios_con_datos)
    total_oportunidades = sum(u['num_oportunidades'] for u in usuarios_con_datos)
    total_cotizaciones = sum(u['num_cotizaciones'] for u in usuarios_con_datos)

    context = {
        'usuarios_datos': usuarios_con_datos,
        'filtro_actual': filtro,
        'titulo_filtro': titulo_filtro,
        'fecha_actual': today.isoformat(),
        'mes_actual': today.strftime('%Y-%m'),
        'fecha_desde': fecha_desde or '',
        'fecha_hasta': fecha_hasta or '',
        'total_general_vendido': total_general_vendido,
        'total_general_pipeline': total_general_pipeline,
        'total_oportunidades': total_oportunidades,
        'total_cotizaciones': total_cotizaciones,
    }

    return render(request, 'estadisticas_usuarios.html', context)


def _exportar_estadisticas_excel(request, oportunidades, cotizaciones, titulo_filtro):
    """Genera un archivo Excel con las estadísticas de usuarios."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Sum, Count

    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas Usuarios"

    # Estilos
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='0A84FF', end_color='0A84FF', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D1D6'),
        right=Side(style='thin', color='D1D1D6'),
        top=Side(style='thin', color='D1D1D6'),
        bottom=Side(style='thin', color='D1D1D6'),
    )
    money_format = '#,##0.00'

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = f'Reporte de Estadísticas de Usuarios — {titulo_filtro}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='0A84FF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = ['Usuario', 'Oportunidades', 'Cotizaciones PDF', 'Total Vendido (100%)', 'Pipeline', 'Cliente Top']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos
    usuarios = User.objects.filter(oportunidades__isnull=False).distinct()
    row = 4
    for usuario in usuarios:
        oportunidades_usuario = oportunidades.filter(usuario=usuario)
        num_oportunidades = oportunidades_usuario.count()
        if num_oportunidades == 0:
            continue
        total_vendido = oportunidades_usuario.filter(probabilidad_cierre=100).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        total_pipeline = oportunidades_usuario.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        num_cotizaciones = cotizaciones.filter(created_by=usuario).count()
        cliente_top = oportunidades_usuario.values('cliente__nombre_empresa').annotate(
            total_cliente=Sum('monto'), num_ops=Count('id')
        ).order_by('-total_cliente').first()
        cliente_nombre = cliente_top['cliente__nombre_empresa'] if cliente_top else '-'

        nombre = usuario.get_full_name() or usuario.username

        ws.cell(row=row, column=1, value=nombre).border = thin_border
        ws.cell(row=row, column=2, value=num_oportunidades).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=num_cotizaciones).border = thin_border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        cell_vendido = ws.cell(row=row, column=4, value=float(total_vendido))
        cell_vendido.number_format = money_format
        cell_vendido.border = thin_border
        cell_pipeline = ws.cell(row=row, column=5, value=float(total_pipeline))
        cell_pipeline.number_format = money_format
        cell_pipeline.border = thin_border
        ws.cell(row=row, column=6, value=cliente_nombre).border = thin_border
        row += 1

    # Ajustar anchos
    col_widths = [25, 18, 18, 22, 22, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="estadisticas_usuarios.xlsx"'
    wb.save(response)
    return response


from django.views.decorators.csrf import csrf_exempt
from .bitrix_integration import get_bitrix_company_details, get_bitrix_deal_details, BITRIX_WEBHOOK_URL, map_bitrix_category_to_tipo_negociacion, get_producto_from_bitrix_deal, get_etapa_from_bitrix_stage
from django.contrib.auth.models import User
import requests
import traceback

@csrf_exempt
def bitrix_webhook_receiver(request):
    """
    Receives webhook notifications from Bitrix24.
    This view handles the 'ONCRMDEALADD' event to create a new opportunity (TodoItem)
    in the local database when a new deal is created in Bitrix24.
    """
    # Log EVERY request to debug
    print(f"BITRIX WEBHOOK: Received {request.method} request from {request.META.get('REMOTE_ADDR', 'unknown')}", flush=True)
    print(f"BITRIX WEBHOOK: Request headers: {dict(request.headers)}", flush=True)
    print(f"BITRIX WEBHOOK: Request body: {request.body}", flush=True)
    
    if request.method != 'POST':
        print("BITRIX WEBHOOK: Received non-POST request. Ignoring.", flush=True)
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

    try:
        data = request.POST
        event = data.get('event')

        print(f"BITRIX WEBHOOK: POST data: {dict(data)}", flush=True)
        print(f"BITRIX WEBHOOK: Received event '{event}'", flush=True)

        if event in ['ONCRMDEALADD', 'ONCRMDEALUPDATE', 'ONCRMDEALDEL']:
            deal_id = data.get('data[FIELDS][ID]')
            if not deal_id:
                print(f"BITRIX WEBHOOK: '{event}' event received but no deal ID found.", flush=True)
                return JsonResponse({'status': 'error', 'message': 'Deal ID missing'}, status=400)

            print(f"BITRIX WEBHOOK: Processing {event} for Deal ID: {deal_id}", flush=True)
            
            # Check if this is a deletion event
            if event == 'ONCRMDEALDEL':
                print(f"BITRIX WEBHOOK: Processing DELETE event for Deal ID: {deal_id}", flush=True)
                try:
                    existing_opportunity = TodoItem.objects.get(bitrix_deal_id=deal_id)
                    opportunity_name = existing_opportunity.oportunidad
                    existing_opportunity.delete()
                    print(f"BITRIX WEBHOOK: Successfully deleted opportunity '{opportunity_name}' (Deal ID: {deal_id})", flush=True)
                    return JsonResponse({'status': 'success', 'message': f'Opportunity deleted: {opportunity_name}'})
                except TodoItem.DoesNotExist:
                    print(f"BITRIX WEBHOOK: Deal ID {deal_id} not found locally for deletion. Nothing to delete.", flush=True)
                    return JsonResponse({'status': 'success', 'message': 'Opportunity not found locally, nothing to delete'})
                except Exception as delete_error:
                    print(f"BITRIX WEBHOOK: Error deleting opportunity for Deal ID {deal_id}: {delete_error}", flush=True)
                    return JsonResponse({'status': 'error', 'message': f'Error deleting opportunity: {str(delete_error)}'}, status=500)

            # Check if this is an update to an existing deal
            is_update = event == 'ONCRMDEALUPDATE'
            existing_opportunity = None
            
            if is_update:
                try:
                    existing_opportunity = TodoItem.objects.get(bitrix_deal_id=deal_id)
                    print(f"BITRIX WEBHOOK: Found existing opportunity '{existing_opportunity.oportunidad}' for update", flush=True)
                except TodoItem.DoesNotExist:
                    print(f"BITRIX WEBHOOK: Deal ID {deal_id} not found locally. Will create new opportunity.", flush=True)
                    is_update = False
            else:
                # For ONCRMDEALADD, check for duplicates
                if TodoItem.objects.filter(bitrix_deal_id=deal_id).exists():
                    print(f"BITRIX WEBHOOK: An opportunity for Deal ID {deal_id} already exists. Skipping creation.", flush=True)
                    return JsonResponse({'status': 'success', 'message': 'Duplicate ignored'})

            # Intentar obtener detalles de la oportunidad con una llamada simple
            deal_details = None
            if BITRIX_WEBHOOK_URL:
                try:
                    get_url = BITRIX_WEBHOOK_URL.replace("crm.deal.add.json", "crm.deal.get.json")
                    response = requests.post(get_url, json={
                        'id': deal_id,
                        'select': [
                            'ID', 'TITLE', 'OPPORTUNITY', 'CURRENCY_ID', 'COMMENTS',
                            'COMPANY_ID', 'CONTACT_ID', 'ASSIGNED_BY_ID', 'STAGE_ID', 'CATEGORY_ID',
                            'UF_CRM_1752859685662',  # Producto (Runrate)
                            'UF_CRM_1750723256972',  # Solución (Proyectos)
                            'UF_CRM_1752859525038',  # Área
                            'UF_CRM_1752859877756',  # Mes de Cobro
                            'UF_CRM_1752855787179',  # Probabilidad de cierre
                            'UF_CRM_1755615484859',  # Utilidad
                        ]
                    }, timeout=5)
                    if response.status_code == 200:
                        deal_data = response.json()
                        if 'result' in deal_data and deal_data['result']:
                            deal_details = deal_data['result']
                            print(f"BITRIX WEBHOOK: Successfully fetched deal details for ID: {deal_id}", flush=True)
                        else:
                            print(f"BITRIX WEBHOOK: No result in API response for Deal ID: {deal_id}", flush=True)
                    else:
                        print(f"BITRIX WEBHOOK: API returned status {response.status_code} for Deal ID: {deal_id}", flush=True)
                except requests.exceptions.RequestException as e:
                    print(f"BITRIX WEBHOOK: Request failed for Deal ID {deal_id}: {e}", flush=True)
                except Exception as e:
                    print(f"BITRIX WEBHOOK: Unexpected error fetching deal details for ID {deal_id}: {e}", flush=True)
            
            # Si no se pudieron obtener los detalles, usar valores por defecto
            if not deal_details:
                deal_details = {
                    'ID': deal_id,
                    'TITLE': f'Oportunidad #{deal_id}',  # Título por defecto
                    'OPPORTUNITY': 0.0,  # Monto por defecto
                    'COMPANY_ID': None,  # Se buscará si hay datos disponibles
                    'ASSIGNED_BY_ID': None,  # Se buscará si hay datos disponibles
                    'CONTACT_ID': None,  # Se buscará si hay datos disponibles
                }
                print(f"BITRIX WEBHOOK: Using default deal details for Deal ID: {deal_id}", flush=True)
            else:
                print(f"BITRIX WEBHOOK: Using fetched deal details: Title='{deal_details.get('TITLE', 'N/A')}', Amount={deal_details.get('OPPORTUNITY', 'N/A')}", flush=True)

            # --- Find or Create the Associated Client (Company) ---
            company_id = deal_details.get('COMPANY_ID')
            cliente = None
            if company_id:
                try:
                    cliente = Cliente.objects.get(bitrix_company_id=company_id)
                    print(f"BITRIX WEBHOOK: Found existing client '{cliente.nombre_empresa}' for Company ID: {company_id}", flush=True)
                except Cliente.DoesNotExist:
                    print(f"BITRIX WEBHOOK: Client with Company ID {company_id} not found. Creating new client.", flush=True)
                    company_details = get_bitrix_company_details(company_id, request=request)
                    if company_details and company_details.get('TITLE'):
                        cliente = Cliente.objects.create(
                            bitrix_company_id=company_details.get('ID'),
                            nombre_empresa=company_details.get('TITLE')
                        )
                        print(f"BITRIX WEBHOOK: Created new client '{cliente.nombre_empresa}'", flush=True)
                    else:
                        print(f"BITRIX WEBHOOK: Could not fetch details for Company ID: {company_id}", flush=True)
            else:
                print("BITRIX WEBHOOK: Deal has no associated Company ID.", flush=True)

            # --- Find the Assigned User ---
            assigned_by_id = deal_details.get('ASSIGNED_BY_ID')
            usuario = None
            
            if assigned_by_id:
                try:
                    usuario = User.objects.filter(userprofile__bitrix_user_id=assigned_by_id).first()
                    if usuario:
                        print(f"BITRIX WEBHOOK: Found user '{usuario.username}' for Assigned ID: {assigned_by_id}", flush=True)
                    else:
                        print(f"BITRIX WEBHOOK: User with Bitrix User ID {assigned_by_id} not found in local DB.", flush=True)
                except Exception as e:
                    print(f"BITRIX WEBHOOK: Error finding user for Bitrix ID {assigned_by_id}: {e}", flush=True)
                    usuario = None
            else:
                print("BITRIX WEBHOOK: Deal has no assigned user.", flush=True)
            
            # --- Find or Create the Contact ---
            contact_id = deal_details.get('CONTACT_ID')
            contacto_obj = None
            
            if contact_id:
                try:
                    # Buscar contacto existente por bitrix_contact_id
                    contacto_obj = Contacto.objects.filter(bitrix_contact_id=contact_id).first()
                    if contacto_obj:
                        print(f"BITRIX WEBHOOK: Found existing contact '{contacto_obj.nombre} {contacto_obj.apellido}' for Contact ID: {contact_id}", flush=True)
                    else:
                        print(f"BITRIX WEBHOOK: Contact with ID {contact_id} not found locally. Fetching from Bitrix24...", flush=True)
                        # Obtener detalles del contacto desde Bitrix24
                        from .bitrix_integration import get_bitrix_contact_details
                        contact_details = get_bitrix_contact_details(contact_id, request=request)
                        if contact_details:
                            try:
                                contacto_obj = Contacto.objects.create(
                                    nombre=contact_details.get('NAME', 'Sin nombre'),
                                    apellido=contact_details.get('LAST_NAME', ''),
                                    bitrix_contact_id=contact_id,
                                    company_id=contact_details.get('COMPANY_ID'),
                                    cliente=cliente  # Asociar con el cliente
                                )
                                print(f"BITRIX WEBHOOK: Created new contact '{contacto_obj.nombre} {contacto_obj.apellido}'", flush=True)
                            except Exception as e:
                                print(f"BITRIX WEBHOOK: Error creating contact: {e}", flush=True)
                                contacto_obj = None
                        else:
                            print(f"BITRIX WEBHOOK: Could not fetch contact details from Bitrix24 for ID {contact_id}", flush=True)
                except Exception as e:
                    print(f"BITRIX WEBHOOK: Error processing contact ID {contact_id}: {e}", flush=True)
                    contacto_obj = None
            else:
                print("BITRIX WEBHOOK: Deal has no associated Contact ID.", flush=True)

            print(f"BITRIX WEBHOOK: DEBUG - After user assignment. is_update={locals().get('is_update', 'NOT_DEFINED')}, existing_opportunity={locals().get('existing_opportunity', 'NOT_DEFINED')}", flush=True)

            # --- Create the Opportunity (TodoItem) ---
            # Always create the opportunity, using defaults if needed
            if not usuario:
                print(f"BITRIX WEBHOOK: User not found, using default user", flush=True)
                usuario, _ = User.objects.get_or_create(
                    username='default_user',
                    defaults={
                        'first_name': 'Usuario',
                        'last_name': 'Sin Asignar',
                        'is_active': True
                    }
                )
            
            if not cliente:
                print(f"BITRIX WEBHOOK: Client not found, using default client", flush=True)
                cliente, _ = Cliente.objects.get_or_create(
                    nombre_empresa='Cliente Desconocido',
                    defaults={'bitrix_company_id': None}
                )

            # Now create the opportunity
            # Mapeos de Bitrix ID a Django Value - igual que en import_bitrix_opportunities.py
            PRODUCTO_BITRIX_ID_TO_DJANGO_VALUE = {
                "176": "ZEBRA", "178": "PANDUIT", "180": "APC", "182": "AVIGILION",
                "184": "GENETEC", "186": "AXIS", "188": "SOFTWARE", "190": "RUNRATE",
                "192": "PÓLIZA", "194": "CISCO", "374": "RFID", "376": "CONSUMIBLE",
                "378": "IMPRESORA INDUSTRIAL", "380": "SCANNER", "382": "TABLETA",
                "582": "SERVICIO",
            }

            AREA_BITRIX_ID_TO_DJANGO_VALUE = {
                "164": "SISTEMAS", "166": "Recursos Humanos", "168": "Compras",
                "170": "Seguridad", "172": "Mantenimiento", "174": "Almacén",
            }

            MES_COBRO_BITRIX_ID_TO_DJANGO_VALUE = {
                "196": "Enero", "198": "Febrero", "200": "Marzo", "202": "Abril",
                "204": "Mayo", "206": "Junio", "208": "Julio", "210": "Agosto",
                "212": "Septiembre", "214": "Octubre", "216": "Noviembre", "218": "Diciembre",
            }

            PROBABILIDAD_BITRIX_ID_TO_VALUE_STRING = {
                "220": "0%", "124": "10%", "126": "20%", "128": "30%",
                "130": "40%", "132": "50%", "134": "60%", "136": "70%",
                "138": "80%", "140": "90%", "142": "100%",
            }

            UTILIDAD_BITRIX_ID_TO_VALUE_STRING = {
                "718": "10%", "720": "15%", "722": "20%", 
                "724": "25%", "726": "30%", "728": "35%",
            }

            # Obtener valores raw de Bitrix
            area_bitrix_id = deal_details.get('UF_CRM_1752859525038')
            mes_cierre_bitrix_id = deal_details.get('UF_CRM_1752859877756')
            probabilidad_bitrix_id = deal_details.get('UF_CRM_1752855787179')
            utilidad_bitrix_id = deal_details.get('UF_CRM_1755615484859')
            category_id = deal_details.get('CATEGORY_ID')

            print(f"BITRIX WEBHOOK: Raw area ID: {area_bitrix_id}", flush=True)
            print(f"BITRIX WEBHOOK: Raw mes_cierre ID: {mes_cierre_bitrix_id}", flush=True)
            print(f"BITRIX WEBHOOK: Raw probabilidad ID: {probabilidad_bitrix_id}", flush=True)
            print(f"BITRIX WEBHOOK: Raw utilidad ID: {utilidad_bitrix_id}", flush=True)
            print(f"BITRIX WEBHOOK: Raw category ID: {category_id}", flush=True)

            # Mapear tipo de negociación desde CATEGORY_ID
            tipo_negociacion = map_bitrix_category_to_tipo_negociacion(category_id)
            
            # Aplicar conversiones de mapeo
            area = AREA_BITRIX_ID_TO_DJANGO_VALUE.get(str(area_bitrix_id), 'SISTEMAS') # Default
            mes_cierre = MES_COBRO_BITRIX_ID_TO_DJANGO_VALUE.get(str(mes_cierre_bitrix_id), 'Enero') # Default
            
            # Mapear producto/solución según el tipo de negociación
            producto = get_producto_from_bitrix_deal(deal_details, tipo_negociacion)
            
            # Mapear etapa según el tipo de negociación
            stage_id = deal_details.get('STAGE_ID')
            etapa_corta, etapa_completa, etapa_color = get_etapa_from_bitrix_stage(stage_id, tipo_negociacion)

            probabilidad_cierre = 0 # Default value
            if probabilidad_bitrix_id is not None:
                prob_str = PROBABILIDAD_BITRIX_ID_TO_VALUE_STRING.get(str(probabilidad_bitrix_id))
                if prob_str:
                    try:
                        parsed_prob = int(prob_str.replace('%', ''))
                        probabilidad_cierre = min(parsed_prob, 100) # Cap at 100
                    except ValueError:
                        print(f"BITRIX WEBHOOK: Invalid probability string from Bitrix: {prob_str}. Setting to default (0).", flush=True)
                        probabilidad_cierre = 0
                else:
                    print(f"BITRIX WEBHOOK: Unknown probability ID from Bitrix: {probabilidad_bitrix_id}. Setting to default (0).", flush=True)
                    probabilidad_cierre = 0

            # Procesar porcentaje de utilidad
            porcentaje_utilidad = 0 # Default value
            if utilidad_bitrix_id is not None:
                utilidad_str = UTILIDAD_BITRIX_ID_TO_VALUE_STRING.get(str(utilidad_bitrix_id))
                if utilidad_str:
                    try:
                        parsed_utilidad = int(utilidad_str.replace('%', ''))
                        porcentaje_utilidad = min(parsed_utilidad, 100) # Cap at 100
                        print(f"BITRIX WEBHOOK: Porcentaje de utilidad configurado: {porcentaje_utilidad}%", flush=True)
                    except ValueError:
                        print(f"BITRIX WEBHOOK: Invalid utilidad string from Bitrix: {utilidad_str}. Setting to default (0).", flush=True)
                        porcentaje_utilidad = 0
                else:
                    print(f"BITRIX WEBHOOK: Unknown utilidad ID from Bitrix: {utilidad_bitrix_id}. Setting to default (0).", flush=True)
                    porcentaje_utilidad = 0

                print(f"BITRIX WEBHOOK: Mapped product: {producto}", flush=True)
                print(f"BITRIX WEBHOOK: Mapped area: {area}", flush=True)
                print(f"BITRIX WEBHOOK: Mapped mes_cierre: {mes_cierre}", flush=True)
                print(f"BITRIX WEBHOOK: Mapped probabilidad_cierre: {probabilidad_cierre}", flush=True)
                print(f"BITRIX WEBHOOK: About to create opportunity with usuario={usuario}, cliente={cliente}", flush=True)

                if is_update and existing_opportunity:
                    # Verificar si cambió el producto (para detectar cuando se agrega Zebra)
                    producto_anterior = existing_opportunity.producto
                    print(f"BITRIX WEBHOOK: PRODUCTO DEBUG - Anterior: '{producto_anterior}' | Nuevo: '{producto}' | Son diferentes: {producto_anterior != producto}", flush=True)
                    
                    # Obtener el estado de Bitrix24
                    bitrix_stage_id = deal_details.get('STAGE_ID')
                    print(f"BITRIX WEBHOOK: Raw STAGE_ID: {bitrix_stage_id}", flush=True)
                    
                    # Actualizar oportunidad existente
                    existing_opportunity.oportunidad = deal_details.get('TITLE', existing_opportunity.oportunidad)
                    existing_opportunity.monto = deal_details.get('OPPORTUNITY', existing_opportunity.monto) or existing_opportunity.monto
                    existing_opportunity.cliente = cliente or existing_opportunity.cliente
                    existing_opportunity.usuario = usuario or existing_opportunity.usuario
                    existing_opportunity.producto = producto
                    existing_opportunity.area = area
                    existing_opportunity.mes_cierre = mes_cierre
                    existing_opportunity.tipo_negociacion = tipo_negociacion  # Nuevo campo
                    existing_opportunity.etapa_corta = etapa_corta
                    existing_opportunity.etapa_completa = etapa_completa
                    existing_opportunity.etapa_color = etapa_color
                    existing_opportunity.probabilidad_cierre = probabilidad_cierre
                    existing_opportunity.bitrix_stage_id = bitrix_stage_id  # Guardar el estado
                    existing_opportunity.contacto = contacto_obj  # Asignar contacto
                    existing_opportunity.save()
                    print(f"BITRIX WEBHOOK: Successfully updated opportunity '{existing_opportunity.oportunidad}' with ID {existing_opportunity.id}", flush=True)
                    
                    # ========================================
                    # VERIFICAR SI CAMBIÓ EL PRODUCTO A UNA MARCA AUTOMÁTICA
                    # ========================================
                    if producto_anterior != producto:
                        print(f"BITRIX WEBHOOK: Producto cambió de '{producto_anterior}' a '{producto}' - Verificando cotización automática", flush=True)
                        
                        # Verificar si debe generar cotización automática
                        if es_cotizacion_automatica(deal_details):
                            print(f"BITRIX WEBHOOK: Oportunidad actualizada califica para cotización automática - Marca: {producto}", flush=True)
                            
                            # Verificar si ya tiene cotización para evitar duplicados
                            from app.models import Cotizacion
                            cotizacion_existente = Cotizacion.objects.filter(oportunidad=existing_opportunity).first()
                            
                            if cotizacion_existente:
                                print(f"BITRIX WEBHOOK: Ya existe cotización ID {cotizacion_existente.id} para esta oportunidad - No se creará duplicado", flush=True)
                            else:
                                # Obtener usuario final del contacto
                                contact_id = deal_details.get('CONTACT_ID')
                                usuario_final = ''
                                if contact_id:
                                    try:
                                        from .bitrix_integration import get_bitrix_contact_details
                                        contact_details = get_bitrix_contact_details(contact_id, request=request)
                                        if contact_details:
                                            usuario_final = f"{contact_details.get('NAME', '')} {contact_details.get('LAST_NAME', '')}".strip()
                                            print(f"BITRIX WEBHOOK: Usuario final obtenido: {usuario_final}", flush=True)
                                    except Exception as e:
                                        print(f"BITRIX WEBHOOK: Error obteniendo contacto: {e}", flush=True)
                                
                                # Crear cotización automática
                                cotizacion_automatica = crear_cotizacion_automatica_bitrix(deal_details, cliente, usuario, porcentaje_utilidad)
                                
                                if cotizacion_automatica:
                                    # Agregar usuario final si se obtuvo
                                    if usuario_final:
                                        cotizacion_automatica.usuario_final = usuario_final
                                        cotizacion_automatica.save()
                                    
                                    # Subir PDF a Bitrix24
                                    resultado_subida = subir_cotizacion_a_bitrix(cotizacion_automatica, deal_id, request)
                                    
                                    if resultado_subida:
                                        print(f"BITRIX WEBHOOK: ✅ Cotización automática completada exitosamente para deal actualizado {deal_id}")
                                    else:
                                        print(f"BITRIX WEBHOOK: ⚠️ Cotización creada pero falló la subida a Bitrix24 para deal actualizado {deal_id}")
                                else:
                                    print(f"BITRIX WEBHOOK: ❌ No se pudo crear la cotización automática para deal actualizado {deal_id}")
                        else:
                            print(f"BITRIX WEBHOOK: Oportunidad actualizada NO califica para cotización automática - Producto: {producto}", flush=True)
                else:
                    # Crear nueva oportunidad
                    print(f"BITRIX WEBHOOK: Creating new opportunity with data:", flush=True)
                    print(f"  - Title: {deal_details.get('TITLE', 'Oportunidad sin título')}", flush=True)
                    print(f"  - Amount: {deal_details.get('OPPORTUNITY', 0.0)}", flush=True)
                    print(f"  - Cliente: {cliente}", flush=True)
                    print(f"  - Usuario: {usuario}", flush=True)
                    print(f"  - Deal ID: {deal_id}", flush=True)
                    
                    # Obtener el estado de Bitrix24
                    bitrix_stage_id = deal_details.get('STAGE_ID')
                    print(f"BITRIX WEBHOOK: Raw STAGE_ID for new opportunity: {bitrix_stage_id}", flush=True)
                    
                    try:
                        new_opportunity = TodoItem.objects.create(
                            oportunidad=deal_details.get('TITLE', 'Oportunidad sin título'),
                            monto=deal_details.get('OPPORTUNITY', 0.0) or 0.0,
                            cliente=cliente,
                            usuario=usuario,
                            bitrix_deal_id=deal_id,
                            producto=producto,
                            area=area,
                            mes_cierre=mes_cierre,
                            tipo_negociacion=tipo_negociacion,  # Nuevo campo
                            etapa_corta=etapa_corta,
                            etapa_completa=etapa_completa,
                            etapa_color=etapa_color,
                            probabilidad_cierre=probabilidad_cierre,
                            bitrix_stage_id=bitrix_stage_id,  # Guardar el estado
                            contacto=contacto_obj,  # Asignar contacto
                        )
                        print(f"BITRIX WEBHOOK: Successfully created new opportunity '{new_opportunity.oportunidad}' with ID {new_opportunity.id}", flush=True)
                        
                        # ========================================
                        # LÓGICA DE COTIZACIÓN AUTOMÁTICA
                        # ========================================
                        
                        # Verificar si debe generar cotización automática
                        if es_cotizacion_automatica(deal_details):
                            print(f"BITRIX WEBHOOK: Oportunidad califica para cotización automática - Marca: {producto}", flush=True)
                            
                            # Obtener usuario final del contacto
                            contact_id = deal_details.get('CONTACT_ID')
                            usuario_final = ''
                            if contact_id:
                                try:
                                    from .bitrix_integration import get_bitrix_contact_details
                                    contact_details = get_bitrix_contact_details(contact_id, request=request)
                                    if contact_details:
                                        usuario_final = f"{contact_details.get('NAME', '')} {contact_details.get('LAST_NAME', '')}".strip()
                                        print(f"BITRIX WEBHOOK: Usuario final obtenido: {usuario_final}", flush=True)
                                except Exception as e:
                                    print(f"BITRIX WEBHOOK: Error obteniendo contacto: {e}", flush=True)
                            
                            # Crear cotización automática
                            cotizacion_automatica = crear_cotizacion_automatica_bitrix(deal_details, cliente, usuario, porcentaje_utilidad)
                            
                            if cotizacion_automatica:
                                # Agregar usuario final si se obtuvo
                                if usuario_final:
                                    cotizacion_automatica.usuario_final = usuario_final
                                    cotizacion_automatica.save()
                                
                                # Subir PDF a Bitrix24
                                resultado_subida = subir_cotizacion_a_bitrix(cotizacion_automatica, deal_id, request)
                                
                                if resultado_subida:
                                    print(f"BITRIX WEBHOOK: ✅ Cotización automática completada exitosamente para deal {deal_id}")
                                else:
                                    print(f"BITRIX WEBHOOK: ⚠️ Cotización creada pero falló la subida a Bitrix24 para deal {deal_id}")
                            else:
                                print(f"BITRIX WEBHOOK: ❌ No se pudo crear la cotización automática para deal {deal_id}")
                        else:
                            print(f"BITRIX WEBHOOK: Oportunidad NO califica para cotización automática - Producto: {producto}", flush=True)
                        
                    except Exception as create_error:
                        print(f"BITRIX WEBHOOK: Error creating opportunity: {create_error}", flush=True)
                        print(f"BITRIX WEBHOOK: Traceback: {traceback.format_exc()}", flush=True)

        return JsonResponse({'status': 'success'})

    except Exception as e:
        print(f"BITRIX WEBHOOK: An unexpected error occurred: {e}", flush=True)
        print(traceback.format_exc(), flush=True)
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

from django.template.context_processors import request as request_context

def add_is_supervisor_to_context(request):
    return {'is_supervisor': is_supervisor(request.user) if request.user.is_authenticated else False}

from .bitrix_integration import get_all_bitrix_companies

@login_required
def cotizaciones_view(request):
    user = request.user
    is_supervisor_flag = is_supervisor(request.user)
    print(f"DEBUG: cotizaciones_view - Usuario: {user.username}, Es supervisor: {is_supervisor_flag}")

    # Obtener clientes de la base de datos local
    clientes_locales = Cliente.objects.all().order_by('nombre_empresa')
    print(f"DEBUG: cotizaciones_view - Clientes locales obtenidos: {clientes_locales.count()}")

    # Obtener clientes de Bitrix
    clientes_bitrix = get_all_bitrix_companies(request=request)
    print(f"DEBUG: cotizaciones_view - Clientes de Bitrix obtenidos: {len(clientes_bitrix)}")

    # Combinar y desduplicar clientes
    clientes_combinados = {cliente.bitrix_company_id: cliente for cliente in clientes_locales if cliente.bitrix_company_id}
    for cliente_b in clientes_bitrix:
        bitrix_id = cliente_b['ID']
        if bitrix_id not in clientes_combinados:
            # Si el cliente de Bitrix no existe localmente, créalo
            cliente_nuevo, created = Cliente.objects.get_or_create(
                bitrix_company_id=bitrix_id,
                defaults={'nombre_empresa': cliente_b['TITLE']}
            )
            if created:
                print(f"DEBUG: cotizaciones_view - Cliente de Bitrix creado localmente: {cliente_nuevo.nombre_empresa}")
    
    # Obtener todos los clientes de nuevo para incluir los recién creados
    clientes = Cliente.objects.all().order_by('nombre_empresa')

    clientes_data = []
    for cliente in clientes:
        if is_supervisor_flag:
            cotizaciones_qs = Cotizacion.objects.filter(cliente=cliente)
        else:
            cotizaciones_qs = Cotizacion.objects.filter(cliente=cliente, created_by=request.user)
        
        from django.urls import reverse
        cotizaciones_list = [
            {
                'id': c.id,
                'nombre': c.nombre_cotizacion or f'Cotización #{c.id}',
                'descargar_url': reverse('generate_cotizacion_pdf', args=[c.id]),
                'view_url': reverse('view_cotizacion_pdf', args=[c.id])
            }
            for c in cotizaciones_qs
        ]

        clientes_data.append({
            'id': cliente.id,
            'nombre': cliente.nombre_empresa,
            'cotizaciones': cotizaciones_list
        })
    print(f"DEBUG: cotizaciones_view - Clientes finales en clientes_asignados: {len(clientes_data)}")

    context = {
        'clientes_asignados': clientes_data,
        'is_supervisor': is_supervisor_flag
    }
    return render(request, 'cotizaciones.html', context)

def cotizaciones_por_oportunidad_view(request, oportunidad_id):
    """
    Vista para mostrar todas las cotizaciones y volumetrías de una oportunidad específica
    """
    oportunidad = get_object_or_404(TodoItem, pk=oportunidad_id)
    
    # Determinar qué tipo de contenido mostrar basado en el parámetro 'tipo'
    tipo_contenido = request.GET.get('tipo', 'cotizaciones')  # Por defecto mostrar cotizaciones
    
    cotizaciones = Cotizacion.objects.filter(oportunidad=oportunidad).order_by('-fecha_creacion')
    volumetrias = Volumetria.objects.filter(oportunidad=oportunidad).order_by('-fecha_creacion')
    
    # Detectar cotizaciones nuevas y crear comentarios automáticos
    if cotizaciones.exists():
        try:
            from app.models import OportunidadComentario, OportunidadActividad
            import re
            
            # LIMPIEZA ÚNICA: Eliminar comentarios duplicados de cotizaciones (ejecutar solo una vez)
            # Buscar comentarios de cotización sin el patrón [COT_ID:X] (comentarios antiguos)
            comentarios_sin_patron = OportunidadComentario.objects.filter(
                oportunidad=oportunidad,
                contenido__contains='📋 Nueva cotización creada:'
            ).exclude(contenido__contains='[COT_ID:')
            
            if comentarios_sin_patron.exists():
                print(f"🧹 Limpiando {comentarios_sin_patron.count()} comentarios duplicados antiguos...")
                
                # También eliminar las actividades asociadas
                for comentario in comentarios_sin_patron:
                    OportunidadActividad.objects.filter(
                        oportunidad=oportunidad,
                        descripcion__contains=f'[COMENTARIO_ID:{comentario.id}]'
                    ).delete()
                
                comentarios_sin_patron.delete()
                print("✅ Comentarios duplicados eliminados")
            
            # Obtener todas las actividades de comentarios de esta oportunidad
            actividades_comentarios = OportunidadActividad.objects.filter(
                oportunidad=oportunidad,
                tipo='comentario'
            )
            
            # Obtener IDs de cotizaciones que ya tienen comentarios automáticos
            # Usamos un patrón único en el contenido del comentario para identificarlos
            cotizaciones_con_comentario = set()
            
            # Buscar comentarios que contengan el patrón específico de cotización automática
            comentarios_auto = OportunidadComentario.objects.filter(
                oportunidad=oportunidad,
                contenido__contains='📋 Nueva cotización creada:'
            )
            
            for comentario in comentarios_auto:
                # Extraer ID de cotización del patrón [COT_ID:X] que agregaremos
                match = re.search(r'\[COT_ID:(\d+)\]', comentario.contenido)
                if match:
                    cotizaciones_con_comentario.add(int(match.group(1)))
                
            print(f"🔍 Cotizaciones con comentarios automáticos: {cotizaciones_con_comentario}")
            print(f"📊 Cotizaciones actuales en BD: {[c.id for c in cotizaciones]}")
            
            # Crear comentarios para cotizaciones sin comentario
            for cotizacion in cotizaciones:
                if cotizacion.id not in cotizaciones_con_comentario:
                    try:
                        # Crear comentario automático con identificador único de la cotización
                        cot_title = cotizacion.titulo or cotizacion.nombre_cotizacion or f'Cotización #{cotizacion.id}'
                        contenido_comentario = f"📋 Nueva cotización creada: {cot_title} [COT_ID:{cotizacion.id}]"
                        
                        comentario_auto = OportunidadComentario.objects.create(
                            oportunidad=oportunidad,
                            usuario=cotizacion.created_by,  # Usuario que creó la cotización
                            contenido=contenido_comentario
                        )
                        
                        # Crear actividad en el timeline (sin el [COT_ID:X] para que no se vea en el frontend)
                        descripcion_actividad_limpia = f"📋 Nueva cotización creada: {cot_title} [COMENTARIO_ID:{comentario_auto.id}]"
                        
                        OportunidadActividad.objects.create(
                            oportunidad=oportunidad,
                            tipo='comentario',
                            titulo='Nueva Cotización',
                            descripcion=descripcion_actividad_limpia,
                            usuario=cotizacion.created_by
                        )
                        
                        print(f"✅ Comentario automático creado para cotización {cotizacion.id} - {cot_title}")
                        
                    except Exception as e:
                        print(f"❌ Error creando comentario automático para cotización {cotizacion.id}: {e}")
                        
        except Exception as e:
            print(f"❌ Error en detección de cotizaciones nuevas: {e}")
    
        # Verificar si la oportunidad está ligada a un proyecto
        try:
            from app.models import Proyecto
            proyecto_ligado = Proyecto.objects.filter(oportunidades_ligadas=oportunidad).first()
        except:
            proyecto_ligado = None
        
        context = {
    
            'oportunidad': oportunidad,
    
            'cotizaciones': cotizaciones,
    
            'volumetrias': volumetrias,
    
            'tiene_cotizaciones': cotizaciones.exists(),
    
            'tiene_volumetrias': volumetrias.exists(),
    
            'tipo_contenido': tipo_contenido,
    
            'is_engineer': True,
            
            'proyecto_ligado': proyecto_ligado,
    
        }
    
    # Asegurarse de que context esté definido antes de usarlo
    if 'context' not in locals():
        context = {}
    
    # Verificar que la oportunidad esté en el contexto
    if 'oportunidad' not in context:
        context['oportunidad'] = oportunidad
    
    return render(request, 'cotizaciones_por_oportunidad.html', context)

@login_required
@require_http_methods(["POST"])
def editar_oportunidad_api(request, oportunidad_id):
    """
    API para editar información de una oportunidad
    """
    try:
        oportunidad = get_object_or_404(TodoItem, pk=oportunidad_id)
        
        # Verificar permisos - supervisores pueden editar todo, usuarios solo sus oportunidades
        if not is_supervisor(request.user) and oportunidad.usuario != request.user:
            return JsonResponse({'success': False, 'error': 'No tienes permisos para editar esta oportunidad'})
        
        updated_values = {}
        
        # Actualizar nombre (titulo) de la oportunidad
        if 'oportunidad' in request.POST and request.POST['oportunidad']:
            oportunidad.oportunidad = request.POST['oportunidad']
            updated_values['oportunidad'] = oportunidad.oportunidad

        # Actualizar cliente
        if 'cliente' in request.POST and request.POST['cliente']:
            try:
                cliente = Cliente.objects.get(id=request.POST['cliente'])
                oportunidad.cliente = cliente
                updated_values['cliente'] = cliente.nombre_empresa
            except Cliente.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Cliente no encontrado'})
        
        # Actualizar contacto (acepta ID de contacto)
        if 'contacto' in request.POST and request.POST['contacto']:
            try:
                contacto_obj = Contacto.objects.get(id=int(request.POST['contacto']))
                oportunidad.contacto = contacto_obj
                updated_values['contacto'] = f"{contacto_obj.nombre} {contacto_obj.apellido or ''}".strip()
            except (Contacto.DoesNotExist, ValueError):
                pass
        
        # Actualizar área
        if 'area' in request.POST and request.POST['area']:
            oportunidad.area = request.POST['area']
            updated_values['area'] = oportunidad.get_area_display()
        
        # Actualizar tipo de negociación
        if 'tipo_negociacion' in request.POST and request.POST['tipo_negociacion']:
            if request.POST['tipo_negociacion'] in ['runrate', 'proyecto']:
                oportunidad.tipo_negociacion = request.POST['tipo_negociacion']
                updated_values['tipo_negociacion'] = oportunidad.get_tipo_negociacion_display()
            else:
                return JsonResponse({'success': False, 'error': 'Tipo de negociación inválido'})
        
        # Actualizar producto
        if 'producto' in request.POST and request.POST['producto']:
            oportunidad.producto = request.POST['producto']
            updated_values['producto'] = oportunidad.get_producto_display()
        
        # Actualizar monto
        if 'monto' in request.POST:
            try:
                monto = float(request.POST['monto'])
                oportunidad.monto = monto
                updated_values['monto'] = f"${monto:,.2f}"
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'error': 'Monto inválido'})
        
        # Actualizar probabilidad
        if 'probabilidad' in request.POST:
            try:
                probabilidad = int(request.POST['probabilidad'])
                if 0 <= probabilidad <= 100:
                    oportunidad.probabilidad_cierre = probabilidad
                    updated_values['probabilidad'] = f"{probabilidad}%"
                else:
                    return JsonResponse({'success': False, 'error': 'Probabilidad debe estar entre 0 y 100'})
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'error': 'Probabilidad inválida'})
        
        # Actualizar mes de cierre
        if 'mes_cierre' in request.POST and request.POST['mes_cierre']:
            oportunidad.mes_cierre = request.POST['mes_cierre']
            updated_values['mes_cierre'] = oportunidad.get_mes_cierre_display()

        # Actualizar etapa (desde widget CRM)
        if 'etapa_corta' in request.POST and request.POST['etapa_corta']:
            from .bitrix_integration import get_etapa_from_bitrix_stage
            nueva_etapa = request.POST['etapa_corta']
            oportunidad.etapa_corta = nueva_etapa
            oportunidad.etapa_completa = nueva_etapa
            updated_values['etapa_corta'] = nueva_etapa

        # Actualizar usuario/vendedor
        if 'usuario' in request.POST and request.POST['usuario']:
            try:
                nuevo_usuario = User.objects.get(id=request.POST['usuario'])
                oportunidad.usuario = nuevo_usuario
                updated_values['usuario'] = nuevo_usuario.get_full_name() or nuevo_usuario.username
            except User.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Usuario no encontrado'})

        oportunidad.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Oportunidad actualizada correctamente',
            'updated_values': updated_values
        })
        
    except Exception as e:
        print(f"Error editando oportunidad: {e}")
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def clientes_api(request):
    """
    API para obtener lista de clientes
    """
    try:
        clientes = Cliente.objects.all().order_by('nombre_empresa')
        clientes_data = [
            {
                'id': cliente.id,
                'nombre_empresa': cliente.nombre_empresa
            }
            for cliente in clientes
        ]
        return JsonResponse(clientes_data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def cotizaciones_por_cliente_view(request, cliente_id):
    user = request.user
    is_supervisor = user.groups.filter(name='Supervisores').exists()
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if is_supervisor:
        cotizaciones = Cotizacion.objects.filter(cliente=cliente)
    else:
        cotizaciones = Cotizacion.objects.filter(cliente=cliente, created_by=user)
    cotizaciones_list = [
        {
            'id': c.id,
            'nombre_cotizacion': c.nombre_cotizacion or c.titulo or f'Cotización #{c.id}',
            'descargar_url': f"/cotizacion/pdf/{c.id}/"
        }
        for c in cotizaciones
    ]
    return render(request, 'cotizaciones_cliente.html', {
        'cliente': cliente,
        'cotizaciones': cotizaciones_list,
        'is_supervisor': is_supervisor
    })




import traceback



from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import TodoItem



from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import Cliente
from .bitrix_integration import get_or_create_bitrix_company

@csrf_exempt
@login_required
def crear_cliente_api(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            # Create the company in Bitrix24 first
            company_name = form.cleaned_data['nombre_empresa']
            contact_name = form.cleaned_data['contacto_principal']
            email = form.cleaned_data['email']
            bitrix_company_id = get_or_create_bitrix_company(company_name, email, contact_name, request=request)

            if bitrix_company_id:
                # Check if a client with this bitrix_company_id already exists
                cliente, created = Cliente.objects.get_or_create(
                    bitrix_company_id=bitrix_company_id,
                    defaults={
                        'nombre_empresa': form.cleaned_data['nombre_empresa'],
                        'contacto_principal': form.cleaned_data['contacto_principal'],
                        'email': form.cleaned_data['email'],
                        'asignado_a': request.user
                    }
                )
                if not created:
                    # If the client already existed, update its fields
                    cliente.nombre_empresa = form.cleaned_data['nombre_empresa']
                    cliente.contacto_principal = form.cleaned_data['contacto_principal']
                    cliente.email = form.cleaned_data['email']
                    cliente.save()

                return JsonResponse({
                    'success': True,
                    'cliente': {
                        'id': cliente.id, # Return the local DB ID
                        'nombre_empresa': cliente.nombre_empresa,
                    }
                })
            else:
                return JsonResponse({'success': False, 'errors': {'__all__': ['Could not create company in Bitrix24.']}}, status=400)
        else:
            print(f"DEBUG: OportunidadModalForm errors: {form.errors}", flush=True)
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required
def oportunidad_detalle_api(request, id):
    """
    Devuelve los datos de una oportunidad en JSON para actualizar la fila tras edición.
    """
    try:
        todo = TodoItem.objects.get(pk=id)
        return JsonResponse({
            'id': todo.id,
            'oportunidad': todo.oportunidad,
            'monto': float(todo.monto),
            'probabilidad_cierre': todo.probabilidad_cierre,
            'cliente': str(todo.cliente) if todo.cliente else '',
            'mes_cierre': str(todo.get_mes_cierre_display()),
            'producto': str(todo.get_producto_display()),
            'area': str(todo.get_area_display()),
            'contacto': str(todo.contacto) if todo.contacto else '',
        })
    except TodoItem.DoesNotExist:
        return JsonResponse({'error': 'Oportunidad no encontrada'}, status=404)

@login_required
def view_incrementa_pdf(request, quote_id):
    # This is a dummy PDF content. In a real scenario, you would fetch the PDF
    # from Incrementa API or a storage.
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Cotización Incrementa #{quote_id}</title>
        <style>
            body {{ font-family: sans-serif; margin: 40px; }}
            h1 {{ color: #333; }}
            p {{ color: #666; }}
        </style>
    </head>
    <body>
        <h1>Cotización Incrementa #{quote_id}</h1>
        <p>Este es un PDF de ejemplo para la cotización de Incrementa.</p>
        <p>Contenido detallado de la cotización #{quote_id} iría aquí.</p>
        <p>Fecha: 2023-XX-XX</p>
        <p>Monto: $XXXX.XX</p>
        <p>Cliente: Cliente de Ejemplo</p>
        <p>Este documento simula el PDF de una cotización de Incrementa.</p>
    </body>
    </html>
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="cotizacion_incrementa_{quote_id}.pdf"'
    HTML(string=html_content).write_pdf(response)
    return response

@login_required
def download_incrementa_pdf(request, quote_id):
    # This is a dummy PDF content. In a real scenario, you would fetch the PDF
    # from Incrementa API or a storage.
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Cotización Incrementa #{quote_id}</title>
        <style>
            body {{ font-family: sans-serif; margin: 40px; }}
            h1 {{ color: #333; }}
            p {{ color: #666; }}
        </style>
    </head>
    <body>
        <h1>Cotización Incrementa #{quote_id}</h1>
        <p>Este es un PDF de ejemplo para la cotización de Incrementa.</p>
        <p>Contenido detallado de la cotización #{quote_id} iría aquí.</p>
        <p>Fecha: 2023-XX-XX</p>
        <p>Monto: $XXXX.XX</p>
        <p>Cliente: Cliente de Ejemplo</p>
        <p>Este documento simula el PDF de una cotización de Incrementa.</p>
    </body>
    </html>
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="cotizacion_incrementa_{quote_id}.pdf"'
    HTML(string=html_content).write_pdf(response)
    return response

@csrf_exempt
@login_required
def incrementa_view(request):
    if not is_supervisor(request.user) and not request.user.groups.filter(name='Supervisores').exists():
        return render(request, 'incrementa.html', {'coming_soon': True})

    # Dummy data for Incrementa quotes (replace with API integration later)
    all_incrementa_quotes = [
        {'id': 1, 'nombre': 'Cotización Incrementa #001 - Cliente A', 'fecha': '2023-01-15', 'monto': 1500.00},
        {'id': 2, 'nombre': 'Cotización Incrementa #002 - Cliente B', 'fecha': '2023-02-20', 'monto': 2500.50},
        {'id': 3, 'nombre': 'Cotización Incrementa #003 - Cliente C', 'fecha': '2023-03-10', 'monto': 1200.75},
        {'id': 4, 'nombre': 'Cotización Incrementa #004 - Cliente A', 'fecha': '2023-04-01', 'monto': 3000.00},
        {'id': 5, 'nombre': 'Cotización Incrementa #005 - Cliente D', 'fecha': '2023-05-22', 'monto': 800.00},
        {'id': 6, 'nombre': 'Cotización Incrementa #006 - Cliente E', 'fecha': '2023-06-05', 'monto': 4500.00},
    ]

    # Add view and download URLs to dummy data
    from django.urls import reverse
    for quote in all_incrementa_quotes:
        quote['view_url'] = reverse('view_incrementa_pdf', args=[quote['id']])
        quote['download_url'] = reverse('download_incrementa_pdf', args=[quote['id']])

    search_query = request.GET.get('q', '')
    incrementa_quotes = all_incrementa_quotes

    if search_query:
        incrementa_quotes = [
            q for q in all_incrementa_quotes if search_query.lower() in q['nombre'].lower()
        ]

    context = {
        'coming_soon': False,
        'incrementa_quotes': incrementa_quotes,
        'search_query': search_query,
    }
    return render(request, 'incrementa.html', context)

@login_required
def actualizar_probabilidad(request, id):
    """
    API para actualizar la probabilidad_cierre de una oportunidad (TodoItem).
    URL: /api/oportunidad/<id>/probabilidad/
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    try:
        prob = int(request.POST.get('probabilidad', -1))
        if prob < 0 or prob > 100:
            return JsonResponse({'error': 'Probabilidad fuera de rango'}, status=400)
        todo = TodoItem.objects.get(pk=id)
        todo.probabilidad_cierre = prob
        todo.save(update_fields=['probabilidad_cierre'])
        return JsonResponse({'ok': True, 'probabilidad': prob})
    except TodoItem.DoesNotExist:
        return JsonResponse({'error': 'Oportunidad no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def actualizar_po(request, id):
    """API para actualizar el campo PO de una oportunidad."""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    try:
        po = request.POST.get('po_number', '').strip()
        todo = TodoItem.objects.get(pk=id)
        todo.po_number = po
        todo.save(update_fields=['po_number'])
        return JsonResponse({'ok': True, 'po_number': po})
    except TodoItem.DoesNotExist:
        return JsonResponse({'error': 'Oportunidad no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@login_required

@login_required
def importar_oportunidades(request):
    clientes = Cliente.objects.all().order_by('nombre_empresa')

    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        tabla_json = request.POST.get('tabla_json')

        if not cliente_id or not tabla_json:
            return JsonResponse({'error': 'Faltan datos obligatorios (cliente o tabla_json)'}, status=400)

        try:
            cliente = Cliente.objects.get(pk=cliente_id)
            oportunidades_data = json.loads(tabla_json)
        except Cliente.DoesNotExist:
            return JsonResponse({'error': 'Cliente no encontrado'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Formato JSON inválido para la tabla'}, status=400)

        created_count = 0
        errors = []

        # Mapeo de nombres de meses a números de dos dígitos
        MONTH_MAPPING = {
            "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
            "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
            "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12"
        }

        # Columnas de producto (deben coincidir con los choices de TodoItem.PRODUCTO_CHOICES)
        PRODUCT_COLUMNS = [
            "zebra", "panduit", "apc", "avigilon", "genetec", "axis",
            "desarrollo_app", "runrate", "poliza", "cisco"
        ]

        for row_data in oportunidades_data:
            try:
                oportunidad_nombre = row_data.get('oportunidad', '')
                area = row_data.get('area', '')
                contacto = row_data.get('contacto', '')

                # Encontrar el producto y el monto/mes de cierre
                producto = None
                monto = Decimal('0.00')
                mes_cierre = None

                # Buscar el producto en las columnas de producto
                for prod_col in PRODUCT_COLUMNS:
                    if row_data.get(prod_col) and row_data.get(prod_col).strip() != '':
                        producto = row_data[prod_col].strip().upper() # Convertir a mayúsculas para coincidir con choices
                        break
                
                # Buscar el monto y mes de cierre en las columnas de meses
                for month_name, month_num in MONTH_MAPPING.items():
                    if row_data.get(month_name) and row_data.get(month_name).strip() != '':
                        try:
                            monto = Decimal(row_data[month_name].strip())
                            mes_cierre = month_num
                            break
                        except (ValueError, TypeError):
                            pass # Ignorar si el monto no es un número válido

                # Validaciones básicas
                if not oportunidad_nombre:
                    errors.append(f"Fila con oportunidad vacía: {row_data}")
                    continue
                if not producto:
                    errors.append(f"Fila '{oportunidad_nombre}': Producto no especificado o inválido.")
                    continue
                if not mes_cierre or monto == Decimal('0.00'):
                    errors.append(f"Fila '{oportunidad_nombre}': Mes de cierre o monto no especificado/inválido.")
                    continue

                # Crear el TodoItem
                TodoItem.objects.create(
                    oportunidad=oportunidad_nombre,
                    cliente=cliente,
                    area=area,
                    contacto=contacto,
                    producto=producto,
                    monto=monto,
                    mes_cierre=mes_cierre,
                    usuario=request.user # Asignar al usuario que importa
                )
                created_count += 1
            except Exception as e:
                errors.append(f"Error procesando fila {row_data}: {e}")

        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        else:
            return JsonResponse({'success': True, 'message': f'{created_count} oportunidades importadas con éxito.'})

    context = {
        'clientes': clientes,
    }
    return render(request, 'importar_oportunidades.html', context)


@login_required
def importar_oportunidades(request):
    clientes = Cliente.objects.all().order_by('nombre_empresa')

    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        tabla_json = request.POST.get('tabla_json')

        if not cliente_id or not tabla_json:
            return JsonResponse({'error': 'Faltan datos obligatorios (cliente o tabla_json)'}, status=400)

        try:
            cliente = Cliente.objects.get(pk=cliente_id)
            oportunidades_data = json.loads(tabla_json)
        except Cliente.DoesNotExist:
            return JsonResponse({'error': 'Cliente no encontrado'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Formato JSON inválido para la tabla'}, status=400)

        created_count = 0
        errors = []

        # Mapeo de nombres de meses a números de dos dígitos
        MONTH_MAPPING = {
            "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
            "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
            "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12"
        }

        # Columnas de producto (deben coincidir con los choices de TodoItem.PRODUCTO_CHOICES)
        PRODUCT_COLUMNS = [
            "zebra", "panduit", "apc", "avigilon", "genetec", "axis",
            "desarrollo_app", "runrate", "poliza", "cisco"
        ]

        for row_data in oportunidades_data:
            try:
                oportunidad_nombre = row_data.get('oportunidad', '')
                area = row_data.get('area', '')
                contacto = row_data.get('contacto', '')

                # Encontrar el producto y el monto/mes de cierre
                producto = None
                monto = Decimal('0.00')
                mes_cierre = None

                # Buscar el producto en las columnas de producto
                for prod_col in PRODUCT_COLUMNS:
                    if row_data.get(prod_col) and row_data.get(prod_col).strip() != '':
                        producto = row_data[prod_col].strip().upper() # Convertir a mayúsculas para coincidir con choices
                        break
                
                # Buscar el monto y mes de cierre en las columnas de meses
                for month_name, month_num in MONTH_MAPPING.items():
                    if row_data.get(month_name) and row_data.get(month_name).strip() != '':
                        try:
                            monto = Decimal(row_data[month_name].strip())
                            mes_cierre = month_num
                            break
                        except (ValueError, TypeError):
                            pass # Ignorar si el monto no es un número válido

                # Validaciones básicas
                if not oportunidad_nombre:
                    errors.append(f"Fila con oportunidad vacía: {row_data}")
                    continue
                if not producto:
                    errors.append(f"Fila '{oportunidad_nombre}': Producto no especificado o inválido.")
                    continue
                if not mes_cierre or monto == Decimal('0.00'):
                    errors.append(f"Fila '{oportunidad_nombre}': Mes de cierre o monto no especificado/inválido.")
                    continue

                # Crear el TodoItem
                TodoItem.objects.create(
                    oportunidad=oportunidad_nombre,
                    cliente=cliente,
                    area=area,
                    contacto=contacto,
                    producto=producto,
                    monto=monto,
                    mes_cierre=mes_cierre,
                    usuario=request.user # Asignar al usuario que importa
                )
                created_count += 1
            except Exception as e:
                errors.append(f"Error procesando fila {row_data}: {e}")

        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        else:
            return JsonResponse({'success': True, 'message': f'{created_count} oportunidades importadas con éxito.'})

    context = {
        'clientes': clientes,
    }
    return render(request, 'importar_oportunidades.html', context) 

# --- VISTA DE PRUEBA TEMPORAL ---
from django.http import HttpResponse

@csrf_exempt
@login_required
def crear_oportunidad_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'errors': 'Invalid request method'}, status=405)

    cliente_id = request.POST.get('cliente')
    oportunidad_nombre = request.POST.get('oportunidad') # Get opportunity name from POST
    monto = request.POST.get('monto')
    area = request.POST.get('area')
    mes_cierre = request.POST.get('mes_cierre')
    probabilidad_cierre = request.POST.get('probabilidad_cierre')
    
    # Set default values for fields not present in the simplified modal
    producto = request.POST.get('producto', '') # Default to empty string
    comentarios = request.POST.get('comentarios', '') # Default to empty string
    bitrix_stage_id = request.POST.get('bitrix_stage_id', 'UC_YUQKW6') # Default to 'Cotizando'

    if not cliente_id:
        return JsonResponse({'success': False, 'errors': {'cliente': 'ID de cliente es requerido.'}}, status=400)
    if not oportunidad_nombre:
        return JsonResponse({'success': False, 'errors': {'oportunidad': 'Nombre de oportunidad es requerido.'}}, status=400)
    if not monto:
        return JsonResponse({'success': False, 'errors': {'monto': 'Monto es requerido.'}}, status=400)
    if not area:
        return JsonResponse({'success': False, 'errors': {'area': 'Área es requerida.'}}, status=400)
    if not mes_cierre:
        return JsonResponse({'success': False, 'errors': {'mes_cierre': 'Mes de cierre es requerido.'}}, status=400)
    if not probabilidad_cierre:
        return JsonResponse({'success': False, 'errors': {'probabilidad_cierre': 'Probabilidad de cierre es requerida.'}}, status=400)

    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({'success': False, 'errors': {'cliente': 'Cliente seleccionado no encontrado.'}}, status=404)

    # Ensure the client has a bitrix_company_id
    if not cliente.bitrix_company_id:
        bitrix_company_id = get_or_create_bitrix_company(cliente.nombre_empresa, request=request)
        if bitrix_company_id:
            cliente.bitrix_company_id = bitrix_company_id
            cliente.save()
        else:
            return JsonResponse({'success': False, 'errors': {'cliente': 'No se pudo obtener o crear el ID de compañía de Bitrix para el cliente.'}}, status=400)
    
    # Create a dictionary for the TodoItem data
    todo_item_data = {
        'oportunidad': oportunidad_nombre,
        'monto': float(monto),
        'cliente': cliente,
        'usuario': request.user, # Assign the current user
        'area': area,
        'mes_cierre': mes_cierre,
        'anio_cierre': timezone.localdate().year,
        'probabilidad_cierre': int(probabilidad_cierre),
        'producto': producto, # Use default or provided
        'comentarios': comentarios, # Use default or provided
        'bitrix_stage_id': bitrix_stage_id, # Use default or provided
        'po_number': '', # Ensure PO is empty on creation
    }

    try:
        venta = TodoItem.objects.create(**todo_item_data)
        
        opportunity_data = {
            'oportunidad': venta.oportunidad,
            'monto': float(venta.monto),
            'cliente': cliente.nombre_empresa,
            'bitrix_company_id': cliente.bitrix_company_id,
            'producto': venta.producto,
            'area': venta.area,
            'mes_cierre': venta.mes_cierre,
            'probabilidad_cierre': venta.probabilidad_cierre,
            'comentarios': venta.comentarios,
            'bitrix_stage_id': venta.bitrix_stage_id,
            'bitrix_contact_id': None, # No contact from simplified form
        }
        
        bitrix_assigned_by_id = None
        if venta.usuario and hasattr(venta.usuario, 'userprofile') and venta.usuario.userprofile.bitrix_user_id:
            bitrix_assigned_by_id = venta.usuario.userprofile.bitrix_user_id
        opportunity_data['bitrix_assigned_by_id'] = bitrix_assigned_by_id
        
        bitrix_response = send_opportunity_to_bitrix(opportunity_data, request=request)
        
        if bitrix_response and bitrix_response.get('result'):
            venta.bitrix_deal_id = bitrix_response.get('result')
            venta.save(update_fields=['bitrix_deal_id'])
            
        return JsonResponse({
            'success': True,
            'oportunidad': {
                'id': venta.id,
                'nombre': venta.oportunidad,
            }
        })

    except Exception as e:
        print(f"ERROR: Falló la sincronización con Bitrix24 para la oportunidad {oportunidad_nombre}: {e}")
        return JsonResponse({
            'success': True, # Still return success for local creation
            'oportunidad': {
                'id': venta.id if 'venta' in locals() else None, # Return ID if created locally
                'nombre': oportunidad_nombre,
            },
            'warning': f'Oportunidad creada localmente, pero falló la sincronización con Bitrix24: {e}'
        })

@login_required
def check_new_local_opportunities(request):
    """
    API endpoint para detectar nuevas oportunidades creadas en nuestro sistema.
    Esta función verifica si hay oportunidades creadas después del último timestamp de verificación.
    """
    try:
        # Obtener timestamp de la última verificación desde el parámetro GET
        last_check_timestamp = request.GET.get('last_check')
        
        if not last_check_timestamp:
            return JsonResponse({
                'success': False,
                'error': 'Falta el parámetro last_check'
            })
        
        # Convertir timestamp a datetime
        from datetime import datetime
        last_check = datetime.fromtimestamp(int(last_check_timestamp) / 1000, tz=timezone.utc)
        
        # Primero verificar si hay una alerta inmediata en la sesión (Crown Jewel Feature)
        opportunities_data = []
        session_key = f'new_opportunity_alert_{request.user.id}'
        
        if session_key in request.session:
            # Hay una oportunidad que se acaba de crear - procesarla inmediatamente
            alert_data = request.session.pop(session_key)  # Remover después de leer
            opportunities_data.append(alert_data)
            print(f"DEBUG: Detectada oportunidad inmediata desde sesión: {alert_data}")
        
        # También buscar oportunidades nuevas creadas después del último check
        # Solo buscar las del usuario actual o todas si es supervisor
        if is_supervisor(request.user):
            new_opportunities = TodoItem.objects.filter(
                created_at__gt=last_check
            ).select_related('cliente').order_by('-created_at')[:5]
        else:
            new_opportunities = TodoItem.objects.filter(
                user=request.user,
                created_at__gt=last_check
            ).select_related('cliente').order_by('-created_at')[:5]
        
        # Agregar oportunidades de base de datos a la lista existente
        for opp in new_opportunities:
            opportunities_data.append({
                'id': opp.id,
                'titulo': opp.oportunidad,
                'cliente_id': opp.cliente.id if opp.cliente else None,
                'cliente_nombre': opp.cliente.nombre_empresa if opp.cliente else 'Sin cliente',
                'monto_estimado': str(opp.precio_estimado) if opp.precio_estimado else 'N/A',
                'probabilidad': opp.probabilidad_exito,
                'created_at': opp.created_at.isoformat(),
                'user': opp.usuario.username if opp.usuario else 'Sin usuario'
            })
        
        return JsonResponse({
            'success': True,
            'new_opportunities': opportunities_data,
            'count': len(opportunities_data),
            'last_check': last_check.isoformat()
        })
        
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en el formato del timestamp: {e}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error interno del servidor: {e}'
        })

def trigger_immediate_opportunity_notification(opportunity, request):
    """
    Crown Jewel Feature: Notifica inmediatamente a la sesión del usuario 
    sobre una nueva oportunidad creada, marcándola para detección inmediata
    """
    try:
        # Crear una entrada en la caché/session para que el frontend la detecte inmediatamente
        if hasattr(request, 'session'):
            session_key = f'new_opportunity_alert_{request.user.id}'
            request.session[session_key] = {
                'opportunity_id': opportunity.id,
                'titulo': opportunity.oportunidad,
                'cliente_id': opportunity.cliente.id if opportunity.cliente else None,
                'cliente_nombre': opportunity.cliente.nombre_empresa if opportunity.cliente else 'Sin cliente',
                'monto_estimado': str(opportunity.precio_estimado) if opportunity.precio_estimado else str(opportunity.monto) if opportunity.monto else 'N/A',
                'created_at': opportunity.created_at.isoformat() if opportunity.created_at else timezone.now().isoformat(),
                'user_id': opportunity.usuario.id if opportunity.usuario else request.user.id,
                'timestamp': timezone.now().timestamp()
            }
            # Marcar para que expire en 5 minutos
            request.session.set_expiry(300)
            print(f"DEBUG: Nueva oportunidad marcada para notificación inmediata: {opportunity.oportunidad}")
        
    except Exception as e:
        print(f"ERROR: No se pudo configurar notificación de oportunidad: {e}")


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@login_required
def sync_bitrix_manual(request):
    """
    Vista para que los supervisores ejecuten sincronización manual con Bitrix
    """
    # Verificar que el usuario sea supervisor
    if not is_supervisor(request.user):
        return JsonResponse({
            'success': False,
            'error': 'Solo los supervisores pueden ejecutar sincronización'
        })
    
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        })
    
    try:
        from django.core.management import call_command
        import io
        import sys
        from contextlib import redirect_stdout, redirect_stderr
        
        # Capturar la salida de los comandos
        stdout_buffer = io.StringIO()
        stderr_buffer = io.StringIO()
        
        results = {
            'success': [],
            'errors': [],
            'output': []
        }
        
        sync_commands = [
            ('sync_bitrix_users', 'Usuarios de Bitrix'),
            ('sync_bitrix', 'Empresas/Clientes'),
            ('sync_bitrix_contacts', 'Contactos'),
            ('import_bitrix_opportunities', 'Oportunidades')
        ]
        
        for command, description in sync_commands:
            try:
                stdout_buffer.seek(0)
                stdout_buffer.truncate(0)
                stderr_buffer.seek(0) 
                stderr_buffer.truncate(0)
                
                with redirect_stdout(stdout_buffer), redirect_stderr(stderr_buffer):
                    call_command(command)
                
                output = stdout_buffer.getvalue()
                results['success'].append({
                    'command': command,
                    'description': description,
                    'output': output[:500]  # Limitar output
                })
                
            except Exception as e:
                error_output = stderr_buffer.getvalue()
                results['errors'].append({
                    'command': command,
                    'description': description,
                    'error': str(e),
                    'output': error_output[:500]
                })
        
        # Estadísticas finales
        from app.models import Cliente, TodoItem, UserProfile
        from django.contrib.auth.models import User
        
        stats = {
            'usuarios': User.objects.count(),
            'profiles_bitrix': UserProfile.objects.filter(bitrix_user_id__isnull=False).count(),
            'clientes': Cliente.objects.count(),
            'oportunidades': TodoItem.objects.count()
        }
        
        return JsonResponse({
            'success': True,
            'message': f'Sincronización completada: {len(results["success"])} exitosos, {len(results["errors"])} errores',
            'results': results,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error general en sincronización: {str(e)}'
        })

@login_required
def search_clientes_api(request):
    """
    API endpoint para buscar clientes por nombre de empresa.
    Permite búsqueda parcial insensible a mayúsculas y minúsculas.
    """
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({'clientes': []})
    
    try:
        # Filtrar clientes basado en permisos del usuario
        if is_supervisor(request.user):
            # Supervisores pueden ver todos los clientes
            clientes_queryset = Cliente.objects.filter(
                nombre_empresa__icontains=query
            ).order_by('nombre_empresa')[:20]  # Limitar a 20 resultados
        else:
            # Usuarios normales solo ven sus clientes asignados
            clientes_queryset = Cliente.objects.filter(
                asignado_a=request.user,
                nombre_empresa__icontains=query
            ).order_by('nombre_empresa')[:20]
        
        # Serializar los resultados
        clientes_data = []
        for cliente in clientes_queryset:
            clientes_data.append({
                'id': cliente.id,
                'nombre_empresa': cliente.nombre_empresa,
                'contacto_principal': cliente.contacto_principal,
                'email': cliente.email,
                'telefono': cliente.telefono,
            })
        
        return JsonResponse({
            'clientes': clientes_data,
            'total': len(clientes_data)
        })
        
    except Exception as e:
        print(f"ERROR: Error en búsqueda de clientes: {e}")
        return JsonResponse({
            'clientes': [],
            'error': 'Error interno del servidor'
        }, status=500)

@login_required
def check_new_bitrix_opportunities(request):
    """
    API endpoint para detectar nuevas oportunidades directamente desde Bitrix24.
    Esta función consulta la API de Bitrix para detectar oportunidades creadas recientemente.
    """
    from .bitrix_integration import get_all_bitrix_deals
    from datetime import datetime, timedelta
    from django.utils import timezone, translation
    from django.utils.translation import gettext_lazy as _
    from django.http import JsonResponse
    from django.utils.translation import activate, get_language

    try:
        # Obtener timestamp de la última verificación desde el parámetro GET
        last_check_timestamp = request.GET.get('last_check')
        if not last_check_timestamp:
            return JsonResponse({
                'success': False,
                'error': 'Falta el parámetro last_check'
            })
        
        # Convertir timestamp a datetime
        last_check = datetime.fromtimestamp(int(last_check_timestamp) / 1000, tz=django_timezone.utc)
        
        # Obtener el ID de usuario de Bitrix24 para este usuario de Django
        try:
            user_profile = request.user.userprofile
            user_bitrix_id = str(user_profile.bitrix_user_id) if user_profile.bitrix_user_id else None
        except:
            user_bitrix_id = None
            
        if not user_bitrix_id:
            return JsonResponse({
                'success': True,
                'new_opportunities': [],
                'count': 0,
                'message': 'Usuario no tiene ID de Bitrix24 configurado'
            })
        
        # Consultar todas las oportunidades desde Bitrix24
        bitrix_deals = get_all_bitrix_deals(request)
        
        if not bitrix_deals:
            return JsonResponse({
                'success': True,
                'new_opportunities': [],
                'count': 0,
                'message': 'No se pudieron obtener oportunidades de Bitrix24'
            })
        
        # Filtrar solo oportunidades asignadas a este usuario en Bitrix24
        user_deals = [deal for deal in bitrix_deals if deal.get('ASSIGNED_BY_ID') == user_bitrix_id]
        
        print(f"DEBUG Bot: Usuario Django {request.user.username} → Bitrix ID {user_bitrix_id}")
        print(f"DEBUG Bot: Encontradas {len(user_deals)} oportunidades para este usuario de {len(bitrix_deals)} totales")
        
        # Filtrar oportunidades nuevas (que no existen en nuestro sistema)
        recent_deals = []
        
        for deal in user_deals[:10]:  # Solo las 10 más recientes del usuario
            # Verificar si esta oportunidad ya existe en nuestro sistema
            deal_id = deal.get('ID')
            if deal_id:
                existing_opportunity = TodoItem.objects.filter(
                    bitrix_deal_id=deal_id
                ).first()
                
                if not existing_opportunity:
                    # Esta es una nueva oportunidad que no tenemos en nuestro sistema
                    # Obtener datos de la compañía si existe
                    company_name = 'Cliente por definir'
                    if deal.get('COMPANY_ID'):
                        try:
                            from .bitrix_integration import get_bitrix_company_details
                            company_data = get_bitrix_company_details(deal.get('COMPANY_ID'), request)
                            if company_data and company_data.get('TITLE'):
                                company_name = company_data.get('TITLE')
                        except Exception as e:
                            print(f"Error obteniendo datos de compañía: {e}")
                    
                    recent_deals.append({
                        'id': deal_id,
                        'bitrix_id': deal_id,
                        'titulo': deal.get('TITLE', 'Sin título'),
                        'monto_estimado': deal.get('OPPORTUNITY', '0'),
                        'company_id': deal.get('COMPANY_ID'),
                        'company_name': company_name,
                        'contact_id': deal.get('CONTACT_ID'),
                        'stage_id': deal.get('STAGE_ID'),
                        'comentarios': deal.get('COMMENTS', ''),
                        'assigned_by_id': deal.get('ASSIGNED_BY_ID'),
                        # Mapear campos personalizados
                        'producto_bitrix_id': deal.get('UF_CRM_1752859685662'),
                        'area_bitrix_id': deal.get('UF_CRM_1752859525038'),
                        'mes_cierre_bitrix_id': deal.get('UF_CRM_1752859877756'),
                        'probabilidad_bitrix_id': deal.get('UF_CRM_1752855787179'),
                        'is_from_bitrix': True,
                        'detected_at': django_timezone.now().isoformat()
                    })
        
        print(f"DEBUG Bot: Encontradas {len(recent_deals)} nuevas oportunidades desde Bitrix24")
        
        return JsonResponse({
            'success': True,
            'new_opportunities': recent_deals,
            'count': len(recent_deals),
            'last_check': django_timezone.now().timestamp() * 1000  # Nuevo timestamp
        })
        
    except Exception as e:
        print(f"ERROR Bot: Error al verificar oportunidades desde Bitrix24: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Error al verificar oportunidades desde Bitrix24: {str(e)}'
        }, status=500)


@login_required
def volumetria(request):
    """
    Vista para la sección de Volumetría - Solo para ingenieros y superusuarios
    """
    # Verificar permisos: solo ingenieros y superusuarios
    if not (is_engineer(request.user) or request.user.is_superuser):
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return redirect('home')
    
    # Obtener datos de oportunidad de la sesión si existen
    oportunidad_data = request.session.get('volumetria_oportunidad_data', None)
    
    context = {
        'page_title': 'Volumetría',
        'user_role': 'Ingeniero' if is_engineer(request.user) else 'Superusuario',
        'oportunidad_data': oportunidad_data
    }
    
    return render(request, 'volumetria.html', context)

@login_required
def crear_volumetria_with_opportunity(request, oportunidad_id):
    """
    Vista para crear volumetría desde una oportunidad específica
    """
    # Verificar permisos: solo ingenieros y superusuarios
    if not (is_engineer(request.user) or request.user.is_superuser):
        messages.error(request, "No tienes permisos para crear volumetrías.")
        return redirect('home')
    
    try:
        # Obtener la oportunidad
        oportunidad = get_object_or_404(TodoItem, pk=oportunidad_id)
        
        # Verificar que el usuario tenga acceso a esta oportunidad
        if not is_supervisor(request.user) and oportunidad.usuario != request.user:
            messages.error(request, "No tienes permisos para acceder a esta oportunidad.")
            return redirect('todos')
        
        print(f"DEBUG: Creando volumetría para oportunidad {oportunidad.id}: {oportunidad.oportunidad}")
        
        # Guardar datos de la oportunidad en la sesión
        request.session['volumetria_oportunidad_data'] = {
            'oportunidad_id': oportunidad.id,
            'oportunidad_nombre': oportunidad.oportunidad,
            'cliente_id': oportunidad.cliente.id if oportunidad.cliente else None,
            'cliente_nombre': oportunidad.cliente.nombre_empresa if oportunidad.cliente else None,
            'monto': float(oportunidad.monto) if oportunidad.monto else 0,
            'timestamp': timezone.now().timestamp()
        }
        
        # Marcar para autostart de volumetría general
        request.session['auto_start_volumetria_general'] = True
        
        # Configurar expiración de la sesión para 30 minutos
        request.session.set_expiry(1800)  # 30 minutos
        
        # messages.success(request, f'Generando volumetría integral para la oportunidad "{oportunidad.oportunidad}".')
        
        # Redirigir a la sección de volumetrías (ahora irá directo a volumetría general)
        return redirect('volumetria')
        
    except Exception as e:
        print(f"ERROR: Error al procesar oportunidad para volumetría: {str(e)}")
        messages.error(request, "Error al cargar los datos de la oportunidad.")
        return redirect('todos')

@csrf_exempt
@login_required
def generar_pdf_volumetria(request):
    """
    Vista para generar el PDF de una volumetría
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Verificar permisos
        if not (is_engineer(request.user) or request.user.is_superuser):
            return JsonResponse({'error': 'No tienes permisos para generar PDFs de volumetría'}, status=403)
        
        # Parsear datos JSON
        data = json.loads(request.body)
        print(f"DEBUG: Datos recibidos para PDF volumetría: {data}")
        
        # Obtener información del cliente y oportunidad si están disponibles
        cliente_nombre = "No especificado"
        oportunidad_nombre = "Sin oportunidad asignada"
        vendedor_responsable = None  # Para el responsable del proyecto
        
        if data.get('cliente_id'):
            try:
                cliente = Cliente.objects.get(id=data['cliente_id'])
                cliente_nombre = cliente.nombre_empresa
            except Cliente.DoesNotExist:
                pass
        
        if data.get('oportunidad_id'):
            try:
                oportunidad = TodoItem.objects.get(id=data['oportunidad_id'])
                oportunidad_nombre = oportunidad.oportunidad
                vendedor_responsable = oportunidad.usuario  # Obtener el vendedor responsable
                print(f"DEBUG: Vendedor responsable de la oportunidad: {vendedor_responsable.get_full_name() if vendedor_responsable else 'No definido'}")
            except TodoItem.DoesNotExist:
                pass
        
        # 1. VERIFICAR SI YA EXISTE UN PROYECTO PARA ESTA OPORTUNIDAD
        project_id = None
        carpeta_volumetrias_id = None
        oportunidad_proyecto = None
        oportunidad = None
        
        # Si hay una oportunidad asociada, verificar si ya existe un proyecto
        if data.get('oportunidad_id'):
            try:
                from .models import OportunidadProyecto
                oportunidad = TodoItem.objects.get(id=data['oportunidad_id'])
                
                # Buscar si ya existe un proyecto para esta oportunidad
                try:
                    oportunidad_proyecto = OportunidadProyecto.objects.get(oportunidad=oportunidad)
                    project_id = oportunidad_proyecto.bitrix_project_id
                    carpeta_volumetrias_id = oportunidad_proyecto.carpeta_volumetrias_id
                    
                    # Incrementar contador de volumetrías
                    oportunidad_proyecto.volumetrias_generadas += 1
                    oportunidad_proyecto.save()
                    
                    print(f"DEBUG: Proyecto existente encontrado: ID {project_id}")
                    print(f"DEBUG: Esta será la volumetría #{oportunidad_proyecto.volumetrias_generadas} para este proyecto")
                    
                except OportunidadProyecto.DoesNotExist:
                    print(f"DEBUG: No existe proyecto para oportunidad {data['oportunidad_id']}, se creará uno nuevo")
                    
            except TodoItem.DoesNotExist:
                print(f"WARNING: Oportunidad {data['oportunidad_id']} no encontrada")
                oportunidad = None
        
        # 2. Si NO existe proyecto, crear uno nuevo en Bitrix24
        if not project_id:
            project_name = data.get('nombre_volumetria', 'Análisis Volumétrico')
            project_description = f"""
Proyecto automatizado desde Nethive para volumetría: {project_name}

Cliente: {cliente_nombre}
Oportunidad: {oportunidad_nombre}
Elaborado por: {data.get('elaborado_por', request.user.get_full_name() or request.user.username)}
Fecha: {data.get('fecha', convert_to_tijuana_time(datetime.now()).strftime('%d/%m/%Y'))}

Categoría: {data.get('categoria', 'CAT6')}
Cantidad de nodos: {data.get('cantidad_nodos', 1)}
Total: ${float(data.get('total', 0)):,.2f} USD

Este proyecto contiene la documentación técnica y volumetría del proyecto.
                """.strip()

            try:
                from .bitrix_integration import create_bitrix_project
                project_id = create_bitrix_project(
                    project_name=project_name,
                    description=project_description,
                    vendedor_responsable=vendedor_responsable,
                    request=request
                )
                
                if project_id:
                    print(f"DEBUG: Nuevo proyecto Bitrix24 creado exitosamente: ID {project_id}")
                    
                    # Crear registro en OportunidadProyecto si hay oportunidad
                    if data.get('oportunidad_id') and oportunidad:
                        try:
                            from .models import OportunidadProyecto
                            oportunidad_proyecto = OportunidadProyecto.objects.create(
                                oportunidad=oportunidad,
                                bitrix_project_id=project_id,
                                bitrix_deal_id=str(oportunidad.bitrix_deal_id) if oportunidad.bitrix_deal_id else None,
                                proyecto_nombre=project_name,
                                created_by=request.user,
                                volumetrias_generadas=1
                            )
                            print(f"DEBUG: Registro OportunidadProyecto creado: ID {oportunidad_proyecto.id}")
                        except Exception as e:
                            print(f"WARNING: Error creando registro OportunidadProyecto: {e}")
                else:
                    print("WARNING: No se pudo crear el proyecto en Bitrix24. La volumetría no se adjuntará al proyecto.")

            except Exception as e:
                print(f"WARNING: Error al intentar crear el proyecto en Bitrix24: {e}")
                project_id = None # Asegurar que project_id sea None si la creación falla

        # Calcular valores financieros (estos dependen de los datos de la volumetría, no del proyecto)
        subtotal = float(data.get('subtotal', 0))
        iva_rate = 0.16  # 16% IVA por defecto
        iva = subtotal * iva_rate
        total = subtotal + iva
        
        # Calcular métricas de rentabilidad basadas en los items
        items = data.get('items', [])
        total_costo_proveedor = sum(float(item.get('total_proveedor', 0)) for item in items)
        ganancia_total = subtotal - total_costo_proveedor
        margen_utilidad = (ganancia_total / total_costo_proveedor * 100) if total_costo_proveedor > 0 else 0
        
        cantidad_nodos = int(data.get('cantidad_nodos', 1))
        precio_por_nodo = total / cantidad_nodos if cantidad_nodos > 0 else 0
        costo_por_nodo = total_costo_proveedor / cantidad_nodos if cantidad_nodos > 0 else 0
        
        # Preparar contexto para el template
        context = {
            'volumetria': {
                'nombre_volumetria': data.get('nombre_volumetria', 'Análisis Volumétrico'),
                'cliente_nombre': cliente_nombre,
                'usuario_final': data.get('usuario_final', ''),
                'oportunidad_nombre': oportunidad_nombre,
                'elaborado_por': data.get('elaborado_por', request.user.get_full_name() or request.user.username),
                'fecha': data.get('fecha', convert_to_tijuana_time(datetime.now()).strftime('%d/%m/%Y')),
                'categoria': data.get('categoria', 'CAT6'),
                'color': data.get('color', 'Azul'),
                'cantidad_nodos': cantidad_nodos,
                'distancia': data.get('distancia', 0),
                'subtotal': subtotal,
                'iva': iva,
                'total': total,
                'precio_por_nodo': precio_por_nodo,
                'costo_por_nodo': costo_por_nodo,
                'ganancia_total': ganancia_total,
                'margen_utilidad': margen_utilidad,
                'items': data.get('items', [])
            },
            'logo_base64': get_logo_base64(),
            'current_date': convert_to_tijuana_time(datetime.now()).strftime('%d/%m/%Y')
        }
        
        # Renderizar template HTML
        html_string = render_to_string('volumetria_pdf_template.html', context)
        
        # Generar PDF usando WeasyPrint
        pdf_file = HTML(string=html_string).write_pdf()
        
        # 3. GUARDAR VOLUMETRÍA EN LA BASE DE DATOS
        volumetria_record = None
        if oportunidad:  # Solo si tenemos una oportunidad
            try:
                # Crear registro de volumetría
                volumetria_record = Volumetria.objects.create(
                    titulo=data.get('nombre_volumetria', 'Análisis Volumétrico'),
                    cliente_id=data.get('cliente_id') if data.get('cliente_id') else oportunidad.cliente.id,
                    usuario_final=data.get('usuario_final', ''),
                    oportunidad=oportunidad,
                    categoria=data.get('categoria', 'CAT6'),
                    color=data.get('color', 'Azul'),
                    cantidad_nodos=cantidad_nodos,
                    distancia=Decimal(str(data.get('distancia', 0))),
                    subtotal=Decimal(str(subtotal)),
                    iva_rate=Decimal(str(iva_rate)),
                    iva_amount=Decimal(str(iva)),
                    total=Decimal(str(total)),
                    moneda='USD',
                    total_costo_proveedor=Decimal(str(total_costo_proveedor)),
                    ganancia_total=Decimal(str(ganancia_total)),
                    margen_utilidad=Decimal(str(margen_utilidad)),
                    precio_por_nodo=Decimal(str(precio_por_nodo)),
                    costo_por_nodo=Decimal(str(costo_por_nodo)),
                    pdf_content=pdf_file,
                    bitrix_project_id=project_id,
                    elaborado_por=data.get('elaborado_por', request.user.get_full_name() or request.user.username),
                    created_by=request.user if request.user.is_authenticated else None
                )
                
                # Crear detalles de volumetría
                for item in items:
                    DetalleVolumetria.objects.create(
                        volumetria=volumetria_record,
                        nombre_producto=item.get('nombre', ''),
                        descripcion=item.get('descripcion', ''),
                        cantidad=int(item.get('cantidad', 1)),
                        precio_unitario=Decimal(str(item.get('precio_unitario', 0))),
                        precio_proveedor=Decimal(str(item.get('precio_proveedor', 0))),
                        total=Decimal(str(item.get('total', 0))),
                        total_proveedor=Decimal(str(item.get('total_proveedor', 0)))
                    )
                
                print(f"DEBUG: Volumetría guardada en BD con ID: {volumetria_record.id}")
                
            except Exception as e:
                print(f"WARNING: Error guardando volumetría en BD: {e}")
        
        # 4. Si el proyecto fue creado, subir el PDF de la volumetría a su drive
        # Definir filename con numeración única para proyectos reutilizados
        if project_id:
            if oportunidad_proyecto and oportunidad_proyecto.volumetrias_generadas > 1:
                # Proyecto reutilizado - usar numeración
                filename = f"Volumetria_Proyecto_{project_id}_V{oportunidad_proyecto.volumetrias_generadas}.pdf"
            else:
                # Proyecto nuevo o primera volumetría
                filename = f"Volumetria_Proyecto_{project_id}.pdf"
        else:
            filename = "Volumetria_SinProyecto.pdf"
        
        # NUEVO SISTEMA: Subida en background + opción manual
        if project_id:
            print(f"DEBUG: Proyecto {project_id} creado exitosamente.")
            print(f"DEBUG: Guardando archivo para subida en background...")
            
            # Guardar el archivo en la base de datos para subida posterior
            try:
                from .models import PendingFileUpload
                
                pending_upload = PendingFileUpload.objects.create(
                    project_id=project_id,
                    filename=filename,
                    file_content=pdf_file,
                    file_size=len(pdf_file),
                    created_by=request.user if request.user.is_authenticated else None,
                    oportunidad_id=data.get('oportunidad_id')
                )
                
                print(f"DEBUG: Archivo guardado para subida (ID: {pending_upload.id})")
                
                # Iniciar tarea en background
                def upload_file_in_background(volumetrias_folder_id_to_use=None):
                    import time
                    print(f"DEBUG BACKGROUND: Iniciando tarea de subida para proyecto {project_id} con folder_id {volumetrias_folder_id_to_use}")
                    
                    # Esperar 10 segundos para que el proyecto esté completamente creado en Bitrix24
                    time.sleep(10)
                    
                    try:
                        from .models import PendingFileUpload, OportunidadProyecto
                        from .bitrix_integration import upload_file_to_project_drive
                        import base64
                        
                        # Obtener el registro pendiente
                        upload_record = PendingFileUpload.objects.get(id=pending_upload.id)
                        
                        if upload_record.status != 'pending':
                            print(f"DEBUG BACKGROUND: Archivo ya procesado ({upload_record.status})")
                            return
                        
                        # Marcar como en progreso
                        upload_record.status = 'in_progress'
                        upload_record.attempts += 1
                        upload_record.save()
                        
                        print(f"DEBUG BACKGROUND: Iniciando subida (intento {upload_record.attempts})")
                        
                        # Codificar archivo
                        pdf_base64 = base64.b64encode(upload_record.file_content).decode('utf-8')
                        
                        # Intentar subir, pasando el ID de la carpeta si existe
                        success, returned_folder_id = upload_file_to_project_drive(
                            project_id=upload_record.project_id,
                            file_name=upload_record.filename,
                            file_content_base64=pdf_base64,
                            request=None,
                            volumetrias_folder_id=volumetrias_folder_id_to_use
                        )
                        
                        if success:
                            upload_record.status = 'success'
                            upload_record.completed_at = timezone.now()
                            print(f"SUCCESS BACKGROUND: Archivo subido exitosamente al proyecto {project_id}")

                            # Si la subida fue exitosa y se creó una nueva carpeta, guardamos su ID
                            if returned_folder_id and not volumetrias_folder_id_to_use:
                                try:
                                    op_proyecto = OportunidadProyecto.objects.get(bitrix_project_id=upload_record.project_id)
                                    op_proyecto.carpeta_volumetrias_id = returned_folder_id
                                    op_proyecto.save(update_fields=['carpeta_volumetrias_id'])
                                    print(f"SUCCESS BACKGROUND: ID de carpeta de volumetrías ({returned_folder_id}) guardado.")
                                except OportunidadProyecto.DoesNotExist:
                                    print(f"ERROR BACKGROUND: No se encontró OportunidadProyecto para guardar el ID de la carpeta.")
                                except Exception as e:
                                    print(f"ERROR BACKGROUND: No se pudo guardar el ID de la carpeta: {e}")
                        else:
                            upload_record.status = 'failed'
                            upload_record.error_message = "La función upload_file_to_project_drive retornó False o un error."
                            print(f"ERROR BACKGROUND: Falló la subida al proyecto {project_id}")
                        
                        upload_record.save()
                        
                    except Exception as e:
                        print(f"ERROR BACKGROUND: Excepción en subida: {e}")
                        try:
                            upload_record.status = 'failed'
                            upload_record.error_message = str(e)
                            upload_record.save()
                        except:
                            pass
                
                # Ejecutar en hilo separado, pasando el ID de la carpeta
                import threading
                upload_thread = threading.Thread(target=upload_file_in_background, kwargs={'volumetrias_folder_id_to_use': carpeta_volumetrias_id})
                upload_thread.daemon = True
                upload_thread.start()
                
                print(f"INFO: Tarea de subida en background iniciada para proyecto {project_id}")
                print(f"INFO: El PDF se descargará inmediatamente. La subida se procesará en segundo plano.")
                
            except Exception as e:
                print(f"WARNING: Error creando tarea de subida en background: {e}")
                print(f"INFO: El PDF se descargará normalmente.")
        
        # Crear respuesta HTTP con el PDF
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        print(f"DEBUG: PDF de volumetría generado exitosamente: {filename}")
        return response
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        print(f"ERROR: Error generando PDF de volumetría: {str(e)}")
        return JsonResponse({'error': f'Error interno del servidor: {str(e)}'}, status=500)


@login_required
def view_volumetria_pdf(request, volumetria_id):
    """
    Vista para mostrar el PDF de una volumetría en el navegador
    """
    volumetria = get_object_or_404(Volumetria, id=volumetria_id)
    
    # Verificar permisos
    if not (is_supervisor(request.user) or volumetria.created_by == request.user or volumetria.oportunidad.usuario == request.user):
        return HttpResponse("No tienes permisos para ver esta volumetría.", status=403)
    
    if not volumetria.pdf_content:
        return HttpResponse("PDF no disponible para esta volumetría.", status=404)
    
    response = HttpResponse(volumetria.pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{volumetria.get_filename()}"'
    return response

@login_required
def download_volumetria_pdf(request, volumetria_id):
    """
    Vista para descargar el PDF de una volumetría
    """
    volumetria = get_object_or_404(Volumetria, id=volumetria_id)
    
    # Verificar permisos
    if not (is_supervisor(request.user) or volumetria.created_by == request.user or volumetria.oportunidad.usuario == request.user):
        return HttpResponse("No tienes permisos para descargar esta volumetría.", status=403)
    
    if not volumetria.pdf_content:
        return HttpResponse("PDF no disponible para esta volumetría.", status=404)
    
    response = HttpResponse(volumetria.pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{volumetria.get_filename()}"'
    return response

@login_required
@require_http_methods(["POST"])
def eliminar_volumetria(request, volumetria_id):
    """
    Vista para eliminar una volumetría
    """
    try:
        volumetria = get_object_or_404(Volumetria, id=volumetria_id)
        
        # Verificar permisos - solo supervisores o el creador pueden eliminar
        if not (is_supervisor(request.user) or volumetria.created_by == request.user):
            return JsonResponse({'success': False, 'error': 'No tienes permisos para eliminar esta volumetría'}, status=403)
        
        # Guardar información antes de eliminar
        titulo = volumetria.titulo
        oportunidad_id = volumetria.oportunidad.id
        
        # Eliminar la volumetría (los detalles se eliminan automáticamente por cascade)
        volumetria.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Volumetría "{titulo}" eliminada exitosamente',
            'redirect_url': f'/app/cotizaciones/oportunidad/{oportunidad_id}/?tipo=volumetrias'
        })
        
    except Volumetria.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Volumetría no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'}, status=500)

def get_logo_base64():
    """
    Función auxiliar para obtener el logo en base64
    """
    try:
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'img', 'bajanet_logo.png')
        with open(logo_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except FileNotFoundError:
        return ""

@csrf_exempt
@login_required
def limpiar_datos_oportunidad_volumetria(request):
    """
    Vista para limpiar los datos de oportunidad de la sesión
    """
    if request.method == 'POST':
        try:
            # Limpiar datos de la sesión
            if 'volumetria_oportunidad_data' in request.session:
                del request.session['volumetria_oportunidad_data']
                print("DEBUG: Datos de oportunidad limpiados de la sesión")
            
            return JsonResponse({'success': True, 'message': 'Datos limpiados correctamente'})
        except Exception as e:
            print(f"ERROR: Error limpiando datos de oportunidad: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt
@login_required
def crear_cotizacion_desde_volumetria(request):
    """
    Vista para crear cotización automáticamente desde una volumetría
    """
    print(f"DEBUG: === INICIO crear_cotizacion_desde_volumetria ===")
    print(f"DEBUG: Método: {request.method}")
    
    if request.method != 'POST':
        print(f"ERROR: Método no permitido: {request.method}")
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    print(f"DEBUG: Llegamos después del check de método")
    
    try:
        print(f"DEBUG: Intentando acceder a request.body...")
        body = request.body
        print(f"DEBUG: Request body obtenido, tipo: {type(body)}, tamaño: {len(body) if body else 'None'}")
        try:
            print(f"DEBUG: Request body preview: {str(body)[:200]}...")  # Solo primeros 200 caracteres
        except:
            print(f"DEBUG: No se puede mostrar el body, pero existe")
        
        print(f"DEBUG: About to parse JSON")
        data = json.loads(body)
    except Exception as body_error:
        print(f"ERROR: Error accediendo al request.body: {body_error}")
        return JsonResponse({'success': False, 'error': f'Error en request body: {str(body_error)}'}, status=400)
    print(f"DEBUG: JSON parseado exitosamente")
    print(f"DEBUG: Tipo de data: {type(data)}")
    print(f"DEBUG: Keys en data: {list(data.keys()) if isinstance(data, dict) else 'No es dict'}")
    # Comentado temporalmente: print(f"DEBUG: Creando cotización desde volumetría: {data}")
    print(f"DEBUG: Iniciando validaciones...")
    print(f"DEBUG: cliente_id en data: {data.get('cliente_id')}")
    print(f"DEBUG: oportunidadId en data: {data.get('oportunidadId')}")
    
    try:
        
        if not data.get('cliente_id'):
            return JsonResponse({'success': False, 'error': 'Cliente requerido'})
        
        if not data.get('items') or len(data.get('items', [])) == 0:
            return JsonResponse({'success': False, 'error': 'Items requeridos'})
        
        cliente = get_object_or_404(Cliente, id=data['cliente_id'])
        
        oportunidad = None
        if data.get('oportunidadId'):
            oportunidad = TodoItem.objects.filter(id=data['oportunidadId']).first()
        
        vendedor_responsable = request.user
        if oportunidad and oportunidad.usuario:
            vendedor_responsable = oportunidad.usuario
        
        cotizacion = Cotizacion.objects.create(
            cliente=cliente,
            nombre_cotizacion=data.get('nombreVolumetria', 'Cotización desde Volumetría'),
            descripcion=f"Cotización generada automáticamente desde volumetría por {request.user.get_full_name() or request.user.username}",
            usuario_final=data.get('usuarioFinal', ''),
            iva_rate=Decimal('0.16'), # Asumir 16% por ahora
            moneda='USD',
            created_by=vendedor_responsable,
            oportunidad=oportunidad,
            bitrix_deal_id=oportunidad.bitrix_deal_id if oportunidad else None
        )
        
        total_cotizacion = Decimal('0')
        for item_data in data['items']:
            try:
                precio_unitario = Decimal(str(item_data.get('precio_cliente', 0)))
                cantidad = int(item_data.get('cantidad', 1))
                descuento = Decimal(str(item_data.get('descuento_cliente', 0)))
                
                precio_con_descuento = precio_unitario * (1 - descuento / 100)
                total_item = precio_con_descuento * cantidad
                
                DetalleCotizacion.objects.create(
                    cotizacion=cotizacion,
                    no_parte=item_data.get('numero_parte', ''),
                    descripcion=item_data.get('descripcion', ''),
                    cantidad=cantidad,
                    precio_unitario=precio_unitario,
                    descuento_porcentaje=descuento,
                    precio_con_descuento=precio_con_descuento,
                    total=total_item
                )
                
                total_cotizacion += total_item
                
            except (ValueError, TypeError, decimal.InvalidOperation) as e:
                print(f"ERROR: Error procesando item {item_data}: {str(e)}")
                continue
        
        cotizacion.subtotal = total_cotizacion
        cotizacion.iva_amount = (cotizacion.subtotal * cotizacion.iva_rate).quantize(Decimal('0.01'))
        cotizacion.total = (cotizacion.subtotal + cotizacion.iva_amount).quantize(Decimal('0.01'))
        cotizacion.save()
        
        # El comentario automático se creará cuando se visite la página de cotizaciones por oportunidad
        from django.urls import reverse
        pdf_url = request.build_absolute_uri(reverse('generate_cotizacion_pdf', args=[cotizacion.id]))
        bitrix_comentario = False
        
        print(f"DEBUG: Verificando condiciones para Bitrix24:")
        print(f"DEBUG: - Oportunidad existe: {oportunidad is not None}")
        if oportunidad:
            print(f"DEBUG: - Oportunidad ID: {oportunidad.id}")
            print(f"DEBUG: - Bitrix Deal ID: {oportunidad.bitrix_deal_id}")
        
        if oportunidad and oportunidad.bitrix_deal_id:
            # Generar PDF y adjuntarlo a Bitrix24
            print(f"DEBUG: Intentando adjuntar cotización a Bitrix24 deal {oportunidad.bitrix_deal_id}")
            try:
                from django.http import HttpRequest
                from django.test import RequestFactory
                
                # Crear una petición temporal para generar el PDF
                factory = RequestFactory()
                pdf_request = factory.get(f'/app/cotizacion/{cotizacion.id}/pdf/')
                pdf_request.user = request.user
                
                # Generar el PDF
                print(f"DEBUG: Generando PDF para cotización {cotizacion.id}")
                # from . import views_exportar # Removed as generate_cotizacion_pdf is in this file
                pdf_response = generate_cotizacion_pdf(pdf_request, cotizacion.id)
                print(f"DEBUG: PDF generado con status {pdf_response.status_code}")
                
                if pdf_response.status_code == 200:
                    # Convertir PDF a base64
                    import base64
                    pdf_content = pdf_response.content
                    pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
                    
                    # Nombre del archivo
                    pdf_filename = f"Cotizacion_{cotizacion.nombre_cotizacion}_{cotizacion.id}.pdf"
                    
                    # Comentario para Bitrix
                    comentario_texto = f"""
📋 Cotización Automática desde Volumetría

Cotización: {cotizacion.nombre_cotizacion}
Total: ${float(total_cotizacion):.2f} USD
Generada por: {request.user.get_full_name() or request.user.username}
Fecha: {convert_to_tijuana_time(timezone.now()).strftime('%d/%m/%Y %H:%M')}

Volumetría origen: {data.get('volumetria_nombre', 'N/A')}

🔗 Creada automáticamente desde el sistema Nethive
                    """.strip()
                    
                    # Subir a Bitrix24
                    print(f"DEBUG: Subiendo PDF '{pdf_filename}' a Bitrix24 deal {oportunidad.bitrix_deal_id}")
                    from .bitrix_integration import add_comment_with_attachment_to_deal
                    resultado_bitrix = add_comment_with_attachment_to_deal(
                        deal_id=oportunidad.bitrix_deal_id,
                        file_name=pdf_filename,
                        file_content_base64=pdf_base64,
                        comment_text=comentario_texto,
                        request=request
                    )
                    print(f"DEBUG: Resultado de subida a Bitrix24: {resultado_bitrix}")
                    
                    if resultado_bitrix:
                        bitrix_comentario = True
                        print(f"SUCCESS: PDF de cotización adjuntado a Bitrix24 deal {oportunidad.bitrix_deal_id}")
                    else:
                        print(f"ERROR: No se pudo adjuntar PDF a Bitrix24 deal {oportunidad.bitrix_deal_id}")
                        
                else:
                    print(f"ERROR: No se pudo generar PDF. Status: {pdf_response.status_code}")
                    
            except Exception as e:
                print(f"ERROR: Excepción adjuntando PDF a Bitrix24: {str(e)}")
                bitrix_comentario = False
        else:
            if not oportunidad:
                print("DEBUG: No hay oportunidad, no se puede adjuntar a Bitrix24")
            elif not oportunidad.bitrix_deal_id:
                print("DEBUG: La oportunidad no tiene bitrix_deal_id, no se puede adjuntar a Bitrix24")

        return JsonResponse({
            'success': True,
            'cotizacion_id': cotizacion.id,
            'cotizacion_nombre': cotizacion.nombre_cotizacion,
            'cotizacion_url': pdf_url,
            'total_cotizacion': float(total_cotizacion),
            'bitrix_comentario': bitrix_comentario,
            'pdf_generado': True,
            'message': 'Cotización creada exitosamente desde volumetría'
        })
        
    except Exception as e:
        import traceback
        print(f"ERROR: Error creando cotización desde volumetría: {str(e)}")
        print(f"ERROR: Traceback completo: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': f'Error interno: {str(e)}'}, status=500)

@login_required
def get_session_data(request):
    """
    Endpoint para obtener datos almacenados en la sesión (usado por volumetría)
    """
    try:
        session_data = {
            'volumetria_oportunidad_data': request.session.get('volumetria_oportunidad_data'),
            'auto_start_volumetria_general': request.session.get('auto_start_volumetria_general', False),
        }
        
        # Limpiar el flag de auto-start después de leerlo para evitar bucles
        if 'auto_start_volumetria_general' in request.session:
            del request.session['auto_start_volumetria_general']
        
        return JsonResponse(session_data)
    except Exception as e:
        print(f"ERROR: Error obteniendo datos de sesión: {str(e)}")
        return JsonResponse({'error': 'Error obteniendo datos de sesión'}, status=500)

@login_required
def get_opportunities_by_client_api(request, cliente_id):
    """
    Endpoint para obtener oportunidades de un cliente específico (usado por volumetría)
    """
    try:
        cliente = get_object_or_404(Cliente, pk=cliente_id)
        
        # Filtrar oportunidades por cliente
        oportunidades = TodoItem.objects.filter(cliente=cliente).order_by('-fecha_creacion')
        
        # Si no es supervisor, solo mostrar las propias oportunidades
        if not is_supervisor(request.user):
            oportunidades = oportunidades.filter(usuario=request.user)
        
        # Serializar las oportunidades
        oportunidades_data = []
        for oportunidad in oportunidades:
            oportunidades_data.append({
                'id': oportunidad.id,
                'oportunidad': oportunidad.oportunidad,
                'monto': float(oportunidad.monto) if oportunidad.monto else 0,
                'mes_cierre': oportunidad.get_mes_cierre_display(),
                'probabilidad_cierre': oportunidad.probabilidad_cierre,
            })
        
        return JsonResponse(oportunidades_data, safe=False)
        
    except Exception as e:
        print(f"ERROR: Error obteniendo oportunidades por cliente: {str(e)}")
        return JsonResponse({'error': 'Error obteniendo oportunidades'}, status=500)

@login_required
def get_client_by_opportunity_api(request, oportunidad_id):
    """
    API para obtener datos del cliente (incluyendo categoría) de una oportunidad específica
    Usado para auto-detectar cliente cuando se selecciona oportunidad en volumetría
    """
    try:
        oportunidad = get_object_or_404(TodoItem, pk=oportunidad_id)
        
        # Verificar que el usuario tenga acceso a esta oportunidad
        if not is_supervisor(request.user) and oportunidad.usuario != request.user:
            return JsonResponse({'error': 'Acceso denegado'}, status=403)
        
        if not oportunidad.cliente:
            return JsonResponse({'error': 'Esta oportunidad no tiene cliente asignado'}, status=400)
        
        cliente = oportunidad.cliente
        cliente_data = {
            'id': cliente.id,
            'nombre_empresa': cliente.nombre_empresa,
            'categoria': cliente.categoria,
            'porcentaje_utilidad': 15 if cliente.categoria == 'A' else 20 if cliente.categoria == 'B' else 25
        }
        
        return JsonResponse(cliente_data)
        
    except Exception as e:
        print(f"ERROR: Error obteniendo cliente por oportunidad: {str(e)}")
        return JsonResponse({'error': 'Error obteniendo datos del cliente'}, status=500)

# ===============================================
# NUEVAS FUNCIONES PARA SUBIDA MANUAL DE ARCHIVOS
# ===============================================

@login_required
def pending_file_uploads(request):
    """
    Vista para mostrar archivos pendientes de subir al usuario
    """
    if request.user.is_authenticated:
        # Si es supervisor, ver todos los archivos
        if is_supervisor(request.user):
            uploads = PendingFileUpload.objects.all()
        else:
            # Si no, solo los propios
            uploads = PendingFileUpload.objects.filter(created_by=request.user)
    else:
        uploads = PendingFileUpload.objects.none()
    
    return render(request, 'pending_file_uploads.html', {
        'uploads': uploads
    })

@login_required
def retry_file_upload(request, upload_id):
    """
    Endpoint para reintentar la subida manual de un archivo
    """
    if request.method == 'POST':
        try:
            upload = get_object_or_404(PendingFileUpload, id=upload_id)
            
            # Verificar permisos
            if not is_supervisor(request.user) and upload.created_by != request.user:
                return JsonResponse({'error': 'Sin permisos para este archivo'}, status=403)
            
            # Verificar que no se exceda el máximo de intentos
            if upload.attempts >= upload.max_attempts:
                return JsonResponse({'error': 'Máximo de intentos excedido'}, status=400)
            
            # Marcar como en progreso
            upload.status = 'in_progress'
            upload.attempts += 1
            upload.error_message = None
            upload.save()
            
            print(f"DEBUG MANUAL: Reintentando subida manual para archivo {upload.filename}")
            
            # Ejecutar subida en background para no bloquear la respuesta
            def manual_upload_background():
                import time
                time.sleep(2)  # Pequeña pausa
                
                try:
                    from .models import OportunidadProyecto, TodoItem
                    from .bitrix_integration import upload_file_to_project_drive
                    import base64

                    # Obtener el ID de la carpeta de volumetrías si existe
                    volumetrias_folder_id = None
                    if upload.oportunidad_id:
                        try:
                            oportunidad = TodoItem.objects.get(id=upload.oportunidad_id)
                            oportunidad_proyecto = OportunidadProyecto.objects.get(oportunidad=oportunidad)
                            volumetrias_folder_id = oportunidad_proyecto.carpeta_volumetrias_id
                            print(f"DEBUG MANUAL: Carpeta de volumetrías encontrada para reintento: {volumetrias_folder_id}")
                        except (TodoItem.DoesNotExist, OportunidadProyecto.DoesNotExist):
                            print(f"DEBUG MANUAL: No se encontró OportunidadProyecto para la oportunidad {upload.oportunidad_id}. Se subirá a la raíz del proyecto.")
                        except Exception as e:
                            print(f"ERROR MANUAL: Buscando carpeta de volumetrías: {e}")

                    # Codificar archivo
                    pdf_base64 = base64.b64encode(upload.file_content).decode('utf-8')
                    
                    print(f"DEBUG MANUAL: Iniciando subida manual para proyecto {upload.project_id}")
                    
                    # Intentar subir, pasando el ID de la carpeta
                    success, returned_folder_id = upload_file_to_project_drive(
                        project_id=upload.project_id,
                        file_name=upload.filename,
                        file_content_base64=pdf_base64,
                        request=None,
                        volumetrias_folder_id=volumetrias_folder_id
                    )
                    
                    if success:
                        upload.status = 'success'
                        upload.completed_at = timezone.now()
                        upload.error_message = None
                        print(f"SUCCESS MANUAL: Archivo {upload.filename} subido exitosamente")
                        # Si se retornó un ID de carpeta y no teníamos uno, lo guardamos.
                        if returned_folder_id and not volumetrias_folder_id:
                             if upload.oportunidad_id:
                                try:
                                    oportunidad = TodoItem.objects.get(id=upload.oportunidad_id)
                                    op_proyecto, created = OportunidadProyecto.objects.get_or_create(oportunidad=oportunidad, defaults={'bitrix_project_id': upload.project_id})
                                    op_proyecto.carpeta_volumetrias_id = returned_folder_id
                                    op_proyecto.save(update_fields=['carpeta_volumetrias_id'])
                                    print(f"SUCCESS MANUAL: ID de carpeta de volumetrías ({returned_folder_id}) guardado en reintento.")
                                except Exception as e:
                                    print(f"ERROR MANUAL: No se pudo guardar el ID de la carpeta en reintento: {e}")
                    else:
                        upload.status = 'failed'
                        upload.error_message = "La función upload_file_to_project_drive retornó False"
                        print(f"ERROR MANUAL: Falló la subida manual de {upload.filename}")
                    
                    upload.save()
                    
                except Exception as e:
                    print(f"ERROR MANUAL: Excepción en subida manual: {e}")
                    upload.status = 'failed'
                    upload.error_message = str(e)
                    upload.save()
            
            # Ejecutar en hilo separado
            import threading
            upload_thread = threading.Thread(target=manual_upload_background)
            upload_thread.daemon = True
            upload_thread.start()
            
            return JsonResponse({
                'success': True, 
                'message': f'Reintentando subida del archivo {upload.filename}',
                'attempts': upload.attempts
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required 
def delete_pending_upload(request, upload_id):
    """
    Endpoint para eliminar un archivo pendiente
    """
    if request.method == 'POST':
        try:
            upload = get_object_or_404(PendingFileUpload, id=upload_id)
            
            # Verificar permisos
            if not is_supervisor(request.user) and upload.created_by != request.user:
                return JsonResponse({'error': 'Sin permisos para este archivo'}, status=403)
            
            filename = upload.filename
            upload.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Archivo {filename} eliminado exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required
def upload_status_api(request):
    """
    API para obtener el estado de las subidas pendientes
    """
    if request.user.is_authenticated:
        if is_supervisor(request.user):
            uploads = PendingFileUpload.objects.all()
        else:
            uploads = PendingFileUpload.objects.filter(created_by=request.user)
    else:
        uploads = PendingFileUpload.objects.none()
    
    uploads_data = []
    for upload in uploads.order_by('-created_at')[:20]:  # Últimas 20
        uploads_data.append({
            'id': upload.id,
            'project_id': upload.project_id,
            'filename': upload.filename,
            'status': upload.status,
            'attempts': upload.attempts,
            'max_attempts': upload.max_attempts,
            'error_message': upload.error_message,
            'created_at': convert_to_tijuana_time(upload.created_at).strftime('%d/%m/%Y %H:%M'),
            'completed_at': convert_to_tijuana_time(upload.completed_at).strftime('%d/%m/%Y %H:%M') if upload.completed_at else None,
        })
    
    return JsonResponse({'uploads': uploads_data})


@login_required
def spotlight_search_api(request):
    """
    API Spotlight - Búsqueda universal de cotizaciones y oportunidades
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:  # Mínimo 2 caracteres para buscar
        return JsonResponse({'results': []})
    
    results = []
    
    # Búsqueda de Cotizaciones (incluyendo búsqueda por número con #)
    # Detectar si buscan por número (#123 o solo 123)
    numeric_query = query.replace('#', '').strip()
    is_numeric_search = numeric_query.isdigit()
    
    cotizaciones_query = Q(nombre_cotizacion__icontains=query) | Q(cliente__nombre_empresa__icontains=query) | Q(descripcion__icontains=query)
    
    # Si es una búsqueda numérica, darle prioridad
    if is_numeric_search:
        cotizaciones_query |= Q(id__exact=int(numeric_query))
    else:
        # Búsqueda regular por ID como string
        cotizaciones_query |= Q(id__icontains=query)
    
    cotizaciones = Cotizacion.objects.filter(cotizaciones_query).select_related('cliente', 'created_by', 'oportunidad')
    
    # Filtrar por permisos de usuario
    if not is_supervisor(request.user):
        cotizaciones = cotizaciones.filter(created_by=request.user)
    
    for cotizacion in cotizaciones[:10]:  # Limitar a 10 resultados
        # Marcar si es una coincidencia exacta de número para priorizar
        is_exact_match = is_numeric_search and str(cotizacion.id) == numeric_query
        
        # Si la cotización tiene una oportunidad asociada, ir a cotizaciones por oportunidad
        # Si no, ir a la página general de cotizaciones
        if cotizacion.oportunidad:
            cotizacion_url = f'/app/cotizaciones/oportunidad/{cotizacion.oportunidad.id}/'
        else:
            cotizacion_url = f'/app/cotizaciones/'
        
        results.append({
            'type': 'cotizacion',
            'id': cotizacion.id,
            'title': cotizacion.nombre_cotizacion or f'Cotización #{cotizacion.id}',
            'subtitle': f'{cotizacion.cliente.nombre_empresa} • ${cotizacion.total:.2f} {cotizacion.moneda}',
            'description': f'Creada por {cotizacion.created_by.get_full_name() or cotizacion.created_by.username if cotizacion.created_by else "Usuario desconocido"}',
            'date': convert_to_tijuana_time(cotizacion.fecha_creacion).strftime('%d/%m/%Y'),
            'icon': 'document',
            'url': cotizacion_url,
            'priority': 1 if is_exact_match else 2,  # Para ordenamiento
            'actions': [
                {'name': 'Ver', 'action': 'view', 'color': 'green'},
                {'name': 'Descargar', 'action': 'download', 'color': 'blue'}, 
                {'name': 'Editar', 'action': 'edit', 'color': 'yellow'}
            ]
        })
    
    # Búsqueda de Oportunidades (TodoItem)
    oportunidades = TodoItem.objects.filter(
        Q(oportunidad__icontains=query) |
        Q(cliente__nombre_empresa__icontains=query) |
        Q(comentarios__icontains=query)
    ).select_related('cliente', 'usuario')
    
    # Filtrar por permisos de usuario
    if not is_supervisor(request.user):
        oportunidades = oportunidades.filter(usuario=request.user)
    
    for oportunidad in oportunidades[:8]:  # Limitar a 8 resultados
        anio = oportunidad.anio_cierre or oportunidad.fecha_creacion.year
        mes = oportunidad.mes_cierre or str(oportunidad.fecha_creacion.month).zfill(2)
        crm_url = f'/app/todos/?tab=crm&anio={anio}&mes=todos&open_opp={oportunidad.id}'
        results.append({
            'type': 'oportunidad',
            'id': oportunidad.id,
            'title': oportunidad.oportunidad,
            'subtitle': f'{oportunidad.cliente.nombre_empresa if oportunidad.cliente else "Sin cliente"} • {oportunidad.etapa_corta or oportunidad.mes_cierre}',
            'description': f'Año {anio} • {oportunidad.get_area_display()} • ${oportunidad.monto:,.0f}',
            'date': convert_to_tijuana_time(oportunidad.fecha_creacion).strftime('%d/%m/%Y'),
            'icon': 'star',
            'url': crm_url,
            'priority': 3,
            'actions': [
                {'name': 'Ver en CRM', 'action': 'view', 'color': 'blue'}
            ]
        })
    
    # Ordenar resultados por prioridad (1=exacto, 2=cotización, 3=oportunidad), luego por título
    results.sort(key=lambda x: (x.get('priority', 3), x['title'].lower()))
    
    return JsonResponse({
        'results': results[:15],  # Máximo 15 resultados totales
        'query': query,
        'total': len(results)
    })

@login_required
def cotizaciones_automaticas_view(request):
    """
    Vista para cotizaciones automáticas - Solo para superusuarios
    """
    # Verificar que el usuario sea superusuario
    if not request.user.is_superuser:
        return redirect('home')  # Redirigir si no es superusuario
    
    if request.method == 'POST':
        # Procesar el formulario de cotización automática
        try:
            titulo = request.POST.get('titulo', '')
            cliente_id = request.POST.get('cliente_hidden', '')
            usuario_final = request.POST.get('usuario_final', '')
            oportunidad_id = request.POST.get('oportunidad', '')
            comentarios = request.POST.get('comentarios', '')
            moneda = request.POST.get('moneda', 'USD')
            iva_rate = float(request.POST.get('iva_rate', '0.16'))
            tipo_cotizacion = request.POST.get('tipo_cotizacion', '')
            descuento_visible = request.POST.get('descuento_visible', 'false') == 'true'
            
            # Verificar campos requeridos
            if not titulo or not cliente_id or not tipo_cotizacion:
                messages.error(request, 'Por favor, completa todos los campos obligatorios.')
                return redirect('cotizaciones_automaticas')
            
            # Obtener cliente
            try:
                cliente = Cliente.objects.get(id=cliente_id)
            except Cliente.DoesNotExist:
                messages.error(request, 'Cliente no encontrado.')
                return redirect('cotizaciones_automaticas')
            
            # Obtener oportunidad si se especificó
            oportunidad = None
            if oportunidad_id:
                try:
                    oportunidad = TodoItem.objects.get(id=oportunidad_id)
                except TodoItem.DoesNotExist:
                    pass
            
            # Crear cotización
            cotizacion = Cotizacion.objects.create(
                titulo=titulo,
                cliente=cliente,
                usuario_final=usuario_final,
                oportunidad=oportunidad,
                comentarios=comentarios,
                moneda=moneda,
                iva_rate=iva_rate,
                tipo_cotizacion=tipo_cotizacion,
                descuento_visible=descuento_visible,
                usuario=request.user
            )
            
            # Procesar productos automáticos
            productos_data = {}
            for key, value in request.POST.items():
                if key.startswith('productos[') and '][' in key:
                    # Extraer índice y campo del producto
                    index_start = key.find('[') + 1
                    index_end = key.find(']')
                    field_start = key.rfind('[') + 1
                    field_end = key.rfind(']')
                    
                    if index_start < index_end and field_start < field_end:
                        index = key[index_start:index_end]
                        field = key[field_start:field_end]
                        
                        if index not in productos_data:
                            productos_data[index] = {}
                        productos_data[index][field] = value
            
            # Procesar títulos/secciones
            titulos_data = {}
            for key, value in request.POST.items():
                if key.startswith('titulos[') and '][' in key:
                    index_start = key.find('[') + 1
                    index_end = key.find(']')
                    field_start = key.rfind('[') + 1
                    field_end = key.rfind(']')
                    
                    if index_start < index_end and field_start < field_end:
                        index = key[index_start:index_end]
                        field = key[field_start:field_end]
                        
                        if index not in titulos_data:
                            titulos_data[index] = {}
                        titulos_data[index][field] = value
            
            # Crear detalles de productos y títulos en el orden correcto
            all_items = {}
            
            # Agregar productos
            for index, producto in productos_data.items():
                if all(field in producto for field in ['marca', 'no_parte', 'descripcion', 'cantidad', 'precio']):
                    all_items[f'producto_{index}'] = {
                        'type': 'producto',
                        'order': int(index),
                        'data': producto
                    }
            
            # Agregar títulos
            for index, titulo in titulos_data.items():
                if 'texto' in titulo and 'position' in titulo:
                    all_items[f'titulo_{index}'] = {
                        'type': 'titulo',
                        'order': int(titulo.get('position', 0)),
                        'data': titulo
                    }
            
            # Ordenar por posición y crear detalles
            sorted_items = sorted(all_items.items(), key=lambda x: x[1]['order'])
            
            for item_key, item in sorted_items:
                if item['type'] == 'producto':
                    producto = item['data']
                    cantidad = int(producto.get('cantidad', 1))
                    precio = float(producto.get('precio', 0))
                    descuento = float(producto.get('descuento', 0))
                    
                    DetalleCotizacion.objects.create(
                        cotizacion=cotizacion,
                        marca=producto.get('marca', ''),
                        no_parte=producto.get('no_parte', ''),
                        descripcion=producto.get('descripcion', ''),
                        nombre_producto=producto.get('nombre_producto', ''),
                        cantidad=cantidad,
                        precio=precio,
                        descuento=descuento,
                        tipo='producto'
                    )
                elif item['type'] == 'titulo':
                    titulo_data = item['data']
                    DetalleCotizacion.objects.create(
                        cotizacion=cotizacion,
                        nombre_producto=titulo_data.get('texto', 'NUEVA SECCIÓN'),
                        tipo='titulo',
                        cantidad=0,
                        precio=0,
                        descuento=0
                    )
            
            messages.success(request, f'Cotización automática "{titulo}" creada exitosamente.')
            
            # Redirigir al PDF de la cotización
            return redirect('generate_cotizacion_pdf', cotizacion_id=cotizacion.id)
            
        except Exception as e:
            messages.error(request, f'Error al crear la cotización automática: {str(e)}')
            return redirect('cotizaciones_automaticas')
    
    # GET request - mostrar formulario
    clientes = Cliente.objects.all().order_by('nombre_empresa')
    clientes_data = [
        {'id': cliente.id, 'name': cliente.nombre_empresa}
        for cliente in clientes
    ]
    
    context = {
        'clientes_data_json': json.dumps(clientes_data),
        'is_cotizaciones_automaticas': True,
    }
    
    return render(request, 'cotizaciones_automaticas.html', context)

def parse_producto_inteligente(linea, marcas_validas):
    """
    Parser inteligente que detecta productos por patrón marca-precio
    """
    import re
    
    # Dividir la línea en tokens por espacios
    tokens = linea.split()
    if len(tokens) < 3:  # Mínimo: marca, no_parte, precio
        return None
    
    # Buscar marca válida al inicio
    marca = None
    marca_index = -1
    for i, token in enumerate(tokens):
        if token.upper() in marcas_validas:
            marca = token.upper()
            marca_index = i
            break
    
    if not marca or marca_index == -1:
        return None
    
    # Buscar precio (número decimal) desde el final
    precio = None
    precio_index = -1
    for i in range(len(tokens) - 1, marca_index, -1):
        token = tokens[i].replace(',', '').replace('$', '')
        try:
            precio_val = float(token)
            if precio_val >= 0:
                precio = precio_val
                precio_index = i
                break
        except ValueError:
            continue
    
    if precio is None or precio_index == -1:
        return None
    
    # Extraer no_parte (siguiente token después de marca)
    if marca_index + 1 >= precio_index:
        return None
    
    no_parte = tokens[marca_index + 1]
    
    # Extraer descripción (todos los tokens entre no_parte y precio)
    descripcion_tokens = tokens[marca_index + 2:precio_index]
    descripcion = ' '.join(descripcion_tokens) if descripcion_tokens else no_parte
    
    # Si no hay descripción separada, usar no_parte como descripción base
    if not descripcion.strip():
        descripcion = f"Producto {no_parte}"
    
    return {
        'marca': marca,
        'no_parte': no_parte,
        'descripcion': descripcion,
        'precio': precio
    }

@login_required
def avatar_generator(request):
    """
    Vista para el generador de avatares con IA
    """
    if request.method == 'POST':
        # Obtener el animal seleccionado
        data = json.loads(request.body)
        animal = data.get('animal', 'dinosaur')
        
        # Definir los prompts para cada animal
        AVATAR_PROMPTS = {
            'dinosaur': "Cute, friendly, cartoon dinosaur, T-Rex species, small round eyes, smiling, 3D render, highly detailed, vibrant green and yellow colors, clean white background, digital art, children's book style. Avatar circle crop.",
            'fox': "Cute, friendly, cartoon fox, sitting up, fluffy orange and white fur, big curious eyes, 3D render, highly detailed, vibrant colors, clean white background, digital art. Avatar circle crop.",
            'owl': "Cute, friendly, cartoon owl, wise, large golden eyes, elegant brown and white feathers, 3D render, highly detailed, vibrant colors, clean white background, digital art. Avatar circle crop.",
            'panda': "Cute, friendly, cartoon giant panda, holding bamboo, round face, black and white fur, 3D render, highly detailed, clean white background, digital art. Avatar circle crop.",
            'wolf': "Cute, friendly, cartoon wolf cub, blue-gray fur, sharp but kind eyes, howling playfully, 3D render, highly detailed, vibrant colors, clean white background, digital art. Avatar circle crop."
        }
        
        # Obtener el prompt según el animal seleccionado
        prompt = AVATAR_PROMPTS.get(animal, AVATAR_PROMPTS['dinosaur'])
        
        # Configurar la petición a la API de generación de imágenes
        # NOTA: Necesitarás configurar tu propia API key en settings.py
        api_key = getattr(settings, 'IMAGEN_API_KEY', '')
        if not api_key:
            return JsonResponse({'error': 'API key no configurada'}, status=500)
            
        url = f"https://generativelanguage.googleapis.com/v1beta/models/imagen-3.0-generate-002:predict?key={api_key}"
        
        payload = {
            "instances": [
                {
                    "prompt": prompt,
                    "aspectRatio": "1:1",
                    "sampleCount": 1
                }
            ],
            "parameters": {
                "sampleImageSize": "1024"
            }
        }
        
        try:
            # Realizar la petición a la API
            response = requests.post(url, json=payload)
            response.raise_for_status()
            result = response.json()
            
            # Extraer la imagen generada
            if 'predictions' in result and result['predictions']:
                image_data = result['predictions'][0].get('bytesBase64Encoded', '')
                if image_data:
                    # Guardar la imagen en el perfil del usuario si está autenticado
                    if hasattr(request.user, 'userprofile'):
                        request.user.userprofile.avatar = f"data:image/png;base64,{image_data}"
                        request.user.userprofile.save()
                    
                    return JsonResponse({
                        'success': True,
                        'image_url': f"data:image/png;base64,{image_data}"
                    })
            
            return JsonResponse({'error': 'No se pudo generar la imagen'}, status=500)
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    # Si es GET, mostrar la página del generador
    return render(request, 'avatar_generator.html')

@login_required
def configuracion_avanzada(request):
    """
    Vista para la configuración avanzada de usuario con diseño tipo macOS
    """
    # Asegurar que el usuario tenga un perfil
    user_profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        try:
            # Actualizar información del usuario
            user = request.user
            
            # Actualizar campos del modelo User
            username = request.POST.get('username', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            
            if username and username != user.username:
                # Verificar que el username no esté en uso
                if User.objects.filter(username=username).exclude(id=user.id).exists():
                    return JsonResponse({'success': False, 'error': 'Este nombre de usuario ya está en uso'})
                user.username = username
            
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.save()
            
            # Actualizar campos del perfil
            birthday = request.POST.get('birthday', '').strip()
            gender = request.POST.get('gender', '').strip()
            
            if birthday:
                try:
                    from datetime import datetime
                    user_profile.birthday = datetime.strptime(birthday, '%Y-%m-%d').date()
                except ValueError:
                    pass  # Ignorar fecha inválida
            
            if gender in ['M', 'F', 'O', 'N']:
                user_profile.gender = gender
            
            user_profile.save()
            
            return JsonResponse({'success': True, 'message': 'Información actualizada correctamente'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    context = {
        'user_profile': user_profile,
    }
    return render(request, 'personalizacion_avanzada.html', context)


def limpiar_y_corregir_texto(texto):
    """
    Limpia caracteres de encoding corrupto y corrige palabras comunes
    """
    import re
    
    # Diccionario de correcciones comunes de caracteres de encoding
    correcciones_encoding = {
        'Bimet√°lica': 'Bimetálica',
        'bimet√°lica': 'bimetálica',
        'met√°lica': 'metálica',
        'Met√°lica': 'Metálica',
        'el√©ctrica': 'eléctrica',
        'El√©ctrica': 'Eléctrica',
        'electr√≥nica': 'electrónica',
        'Electr√≥nica': 'Electrónica',
        'mec√°nica': 'mecánica',
        'Mec√°nica': 'Mecánica',
        'hidr√°ulica': 'hidráulica',
        'Hidr√°ulica': 'Hidráulica',
        'neum√°tica': 'neumática',
        'Neum√°tica': 'Neumática',
        'autom√°tica': 'automática',
        'Autom√°tica': 'Automática',
        'est√°ndar': 'estándar',
        'Est√°ndar': 'Estándar',
        'b√°sica': 'básica',
        'B√°sica': 'Básica',
        'avanzada': 'avanzada',
        'Avanzada': 'Avanzada',
        'pr√°ctica': 'práctica',
        'Pr√°ctica': 'Práctica',
        'gu√≠a': 'guía',
        'Gu√≠a': 'Guía',
        'energ√≠a': 'energía',
        'Energ√≠a': 'Energía',
        'bater√≠a': 'batería',
        'Bater√≠a': 'Batería',
        'H√∫medas': 'Húmedas',
        'h√∫medas': 'húmedas',
        'h√∫meda': 'húmeda',
        'H√∫meda': 'Húmeda',
        'C√©dula': 'Cédula',
        'c√©dula': 'cédula',
        'c√©lula': 'célula',
        'C√©lula': 'Célula',
        'f√°cil': 'fácil',
        'F√°cil': 'Fácil',
        'r√°pida': 'rápida',
        'R√°pida': 'Rápida',
        'r√°pido': 'rápido',
        'R√°pido': 'Rápido',
        'c√°mara': 'cámara',
        'C√°mara': 'Cámara',
        '√°rea': 'área',
        '√Área': 'Área',
        'energ√≠a': 'energía',
        'Energ√≠a': 'Energía',
        'tecnolog√≠a': 'tecnología',
        'Tecnolog√≠a': 'Tecnología',
        'categor√≠a': 'categoría',
        'Categor√≠a': 'Categoría',
        'garant√≠a': 'garantía',
        'Garant√≠a': 'Garantía',
        'compa√±√≠a': 'compañía',
        'Compa√±√≠a': 'Compañía',
        'dise√±o': 'diseño',
        'Dise√±o': 'Diseño',
        'peque√±o': 'pequeño',
        'Peque√±o': 'Pequeño',
        'peque√±a': 'pequeña',
        'Peque√±a': 'Pequeña',
        'a√±o': 'año',
        'A√±o': 'Año',
        'ni√±o': 'niño',
        'Ni√±o': 'Niño',
        'ni√±a': 'niña',
        'Ni√±a': 'Niña',
        'espa√±ol': 'español',
        'Espa√±ol': 'Español',
        'espa√±ola': 'española',
        'Espa√±ola': 'Española'
    }
    
    # Patrones de corrección automática
    patrones_correccion = [
        # √° -> á
        (r'(.*)√°(.*)', r'\1á\2'),
        # √© -> é  
        (r'(.*)√©(.*)', r'\1é\2'),
        # √≠ -> í
        (r'(.*)√≠(.*)', r'\1í\2'),
        # √≥ -> ó
        (r'(.*)√≥(.*)', r'\1ó\2'),
        # √∫ -> ú (común en "húmedo")
        (r'(.*)√∫(.*)', r'\1ú\2'),
        # √± -> ñ
        (r'(.*)√±(.*)', r'\1ñ\2'),
        # Otros caracteres problemáticos
        (r'(.*)‚Ä¢(.*)', r'\1"\2'),
        (r'(.*)‚Äù(.*)', r'\1"\2'),
        (r'(.*)‚Äî(.*)', r'\1-\2'),
        (r'(.*)√¢(.*)', r'\1â\2'),
        (r'(.*)√™(.*)', r'\1ê\2'),
        (r'(.*)√Æ(.*)', r'\1Æ\2'),
        # Limpiar caracteres raros restantes
        (r'√[^a-zA-Z0-9]', ''),
        # Múltiples espacios
        (r'\s+', ' ')
    ]
    
    # Aplicar correcciones específicas del diccionario primero
    texto_corregido = texto
    for incorrecto, correcto in correcciones_encoding.items():
        texto_corregido = texto_corregido.replace(incorrecto, correcto)
    
    # Aplicar patrones de corrección automática
    for patron, reemplazo in patrones_correccion:
        texto_corregido = re.sub(patron, reemplazo, texto_corregido)
    
    # Normalizar espacios y limpiar
    texto_corregido = re.sub(r'\s+', ' ', texto_corregido).strip()
    
    return texto_corregido

def parse_producto_volumetria_inteligente(linea):
    """
    Parser inteligente específico para productos de volumetría
    Detecta patrón: numero_parte marca descripcion precio_lista precio_proveedor
    Ejemplo: TC-CARRIER Bobina de cable de 305 metros... 384 226.56
    """
    import re
    
    # Limpiar la línea y corregir caracteres de encoding
    linea = limpiar_y_corregir_texto(linea.strip())
    if not linea:
        return None
    
    # Dividir en tokens por espacios
    tokens = linea.split()
    if len(tokens) < 3:  # Mínimo: numero_parte, marca, precio
        return None
    
    # Buscar el primer token que parezca un número de parte (con guiones, letras y números)
    numero_parte = None
    numero_parte_index = -1
    
    # Patrón para número de parte: letras, números, guiones
    patron_numero_parte = re.compile(r'^[A-Z0-9\-\/\.]+$', re.IGNORECASE)
    
    for i, token in enumerate(tokens):
        if patron_numero_parte.match(token) and len(token) >= 3:
            numero_parte = token.upper()
            numero_parte_index = i
            break
    
    if not numero_parte or numero_parte_index == -1:
        return None
    
    # Buscar precios desde el final (debe haber al menos 2: precio_lista y precio_proveedor)
    precios = []
    precio_indices = []
    
    # Buscar los últimos 2 números como precios
    for i in range(len(tokens) - 1, numero_parte_index, -1):
        token = tokens[i].replace(',', '').replace('$', '')
        try:
            precio_val = float(token)
            if precio_val >= 0:
                precios.append(precio_val)
                precio_indices.append(i)
                if len(precios) >= 2:  # Solo necesitamos 2 precios
                    break
        except ValueError:
            continue
    
    if len(precios) < 2:
        return None
    
    # Los precios están en orden inverso (el último es precio_lista, el penúltimo es precio_proveedor)
    precio_proveedor = precios[0]  # Último precio encontrado
    precio_lista = precios[1]      # Penúltimo precio encontrado
    
    # El índice donde termina la descripción es el penúltimo precio
    fin_descripcion_index = precio_indices[1]
    
    # Extraer marca (siguiente token después de numero_parte)
    if numero_parte_index + 1 >= fin_descripcion_index:
        return None
    
    marca = tokens[numero_parte_index + 1] if numero_parte_index + 1 < len(tokens) else 'UNKNOWN'
    
    # Extraer descripción (todos los tokens entre marca y primer precio)
    inicio_descripcion = numero_parte_index + 2
    descripcion_tokens = tokens[inicio_descripcion:fin_descripcion_index]
    descripcion = ' '.join(descripcion_tokens) if descripcion_tokens else f"Producto {numero_parte}"
    
    # Limpiar y corregir caracteres de encoding en la descripción
    descripcion = limpiar_y_corregir_texto(descripcion)
    
    # Si la descripción está muy corta, usar una más descriptiva
    if len(descripcion) < 10:
        descripcion = f"Producto {numero_parte} - {marca}"
    
    # Detectar categoría automáticamente desde número de parte o descripción (solo para referencia, no restrictiva)
    categoria = ''
    descripcion_lower = descripcion.lower()
    numero_parte_lower = numero_parte.lower()
    
    if 'cat6a' in descripcion_lower or 'cat6a' in numero_parte_lower:
        categoria = 'Cat6A'
    elif 'cat6' in descripcion_lower or 'cat6' in numero_parte_lower:
        categoria = 'Cat6'
    elif 'cat5e' in descripcion_lower or 'cat5e' in numero_parte_lower:
        categoria = 'Cat5e'
    elif 'jack' in descripcion_lower:
        categoria = 'Jack'
    elif 'patch' in descripcion_lower:
        categoria = 'Patchcord'
    elif 'faceplate' in descripcion_lower or 'face' in descripcion_lower:
        categoria = 'Faceplate'
    elif 'tubo' in descripcion_lower or 'conduit' in descripcion_lower:
        categoria = 'Conduit'
    elif 'conector' in descripcion_lower:
        categoria = 'Conector'
    elif 'cable' in descripcion_lower:
        categoria = 'Cable'
    elif 'charola' in descripcion_lower or 'malla' in descripcion_lower:
        categoria = 'Charola'
    # Si no se detecta ninguna categoría específica, dejar vacío para que el usuario decida
    # No forzar una categoría "Otro"
    
    # Detectar color automáticamente
    color = ''
    if 'negro' in descripcion_lower or 'black' in descripcion_lower:
        color = 'NEGRO'
    elif 'azul' in descripcion_lower or 'blue' in descripcion_lower:
        color = 'AZUL'
    elif 'blanco' in descripcion_lower or 'white' in descripcion_lower:
        color = 'BLANCO'
    elif 'gris' in descripcion_lower or 'gray' in descripcion_lower:
        color = 'GRIS'
    elif 'amarillo' in descripcion_lower or 'yellow' in descripcion_lower:
        color = 'AMARILLO'
    elif 'rojo' in descripcion_lower or 'red' in descripcion_lower:
        color = 'ROJO'
    elif 'verde' in descripcion_lower or 'green' in descripcion_lower:
        color = 'VERDE'
    
    return {
        'numero_parte': numero_parte,
        'marca': marca.upper(),
        'descripcion': descripcion,
        'precio_lista': precio_lista,
        'precio_proveedor': precio_proveedor,
        'categoria': categoria,
        'color': color
    }

@login_required
def gestion_productos_view(request):
    """
    Vista para gestión de productos - Solo para superusuarios
    """
    # Verificar que el usuario sea superusuario
    if not request.user.is_superuser:
        return redirect('home')  # Redirigir si no es superusuario
    
    if request.method == 'POST':
        # Procesar importación de productos desde Excel
        try:
            excel_data = request.POST.get('excel_data', '').strip()
            actualizar_existentes = request.POST.get('actualizar_existentes') == 'on'
            
            if not excel_data:
                messages.error(request, 'Por favor, pega los datos de Excel.')
                return redirect('gestion_productos')
            
            # Procesar datos del Excel
            lineas = excel_data.split('\n')
            productos_procesados = []
            productos_para_crear = []  # Para bulk_create
            productos_para_actualizar = []  # Para bulk_update
            productos_nuevos = 0
            productos_actualizados = 0
            productos_duplicados = 0
            errores = []
            
            # Determinar batch size basado en volumen de datos
            total_lineas = len(lineas)
            if total_lineas > 10000:
                batch_size = 1000  # Para datasets muy grandes
                print(f"IMPORTACIÓN: 🚀 Modo ULTRA-ESCALABLE activado para {total_lineas} líneas (batch: {batch_size})", flush=True)
            elif total_lineas > 5000:
                batch_size = 750   # Para datasets grandes
                print(f"IMPORTACIÓN: 🏃 Modo RÁPIDO activado para {total_lineas} líneas (batch: {batch_size})", flush=True)
            else:
                batch_size = 500   # Para datasets normales
                print(f"IMPORTACIÓN: ⚡ Modo ESTÁNDAR para {total_lineas} líneas (batch: {batch_size})", flush=True)
            
            from app.models import Marca, ProductoCatalogo, ImportacionProductos
            import re
            
            # Marcas válidas
            marcas_validas = {'ZEBRA', 'PANDUIT', 'APC', 'AVIGILON', 'AXIS', 'GENETEC', 'CISCO'}
            
            for i, linea in enumerate(lineas, 1):
                linea = linea.strip()
                if not linea:
                    continue
                
                # Indicador de progreso para datasets grandes
                if total_lineas > 5000 and i % 1000 == 0:
                    progreso = (i / total_lineas) * 100
                    print(f"IMPORTACIÓN: Progreso {progreso:.1f}% ({i}/{total_lineas} líneas procesadas)", flush=True)
                
                # Detectar si es línea de encabezado y saltarla
                if linea.lower().startswith('marca') and 'no' in linea.lower() and 'precio' in linea.lower():
                    continue
                
                try:
                    # Usar algoritmo inteligente de detección por patrones
                    producto_data = parse_producto_inteligente(linea, marcas_validas)
                    
                    if not producto_data:
                        errores.append(f'Línea {i}: No se pudo detectar el patrón marca-precio válido')
                        continue
                    
                    marca_nombre = producto_data['marca']
                    no_parte = producto_data['no_parte']
                    descripcion = producto_data['descripcion']
                    precio_str = str(producto_data['precio'])
                    
                    # Validar datos
                    if not marca_nombre or not no_parte or not descripcion:
                        errores.append(f'Línea {i}: Datos vacíos')
                        continue
                    
                    # Limpiar y convertir precio
                    precio = float(precio_str.replace(',', '').replace('$', ''))
                    if precio < 0:
                        errores.append(f'Línea {i}: Precio no puede ser negativo')
                        continue
                    
                    # Crear o obtener marca
                    marca, created = Marca.objects.get_or_create(
                        nombre=marca_nombre,
                        defaults={'activa': True}
                    )
                    
                    # Crear clave única para identificar duplicados
                    clave_producto = f"{marca.id}_{no_parte}"
                    
                    # Verificar si el producto ya existe en BD
                    producto_existente = ProductoCatalogo.objects.filter(
                        marca=marca,
                        no_parte=no_parte
                    ).first()
                    
                    if producto_existente:
                        if actualizar_existentes:
                            # Agregar a lista para bulk_update
                            producto_existente.descripcion = descripcion
                            producto_existente.precio = precio
                            productos_para_actualizar.append(producto_existente)
                            productos_actualizados += 1
                            # Solo mostrar logs detallados si dataset es pequeño
                            if total_lineas <= 5000:
                                print(f"IMPORTACIÓN: Producto actualizado - {marca_nombre} {no_parte}", flush=True)
                        else:
                            # Producto existe pero no actualizar
                            productos_duplicados += 1
                            # Solo mostrar logs detallados si dataset es pequeño
                            if total_lineas <= 5000:
                                print(f"IMPORTACIÓN: Producto duplicado ignorado - {marca_nombre} {no_parte}", flush=True)
                    else:
                        # Agregar a lista para bulk_create
                        nuevo_producto = ProductoCatalogo(
                            marca=marca,
                            no_parte=no_parte,
                            descripcion=descripcion,
                            precio=precio
                        )
                        productos_para_crear.append(nuevo_producto)
                        productos_nuevos += 1
                        # Solo mostrar logs detallados si dataset es pequeño
                        if total_lineas <= 5000:
                            print(f"IMPORTACIÓN: Producto nuevo preparado - {marca_nombre} {no_parte}", flush=True)
                    
                    productos_procesados.append({
                        'marca': marca_nombre,
                        'no_parte': no_parte,
                        'descripcion': descripcion,
                        'precio': precio
                    })
                    
                except ValueError as e:
                    errores.append(f'Línea {i}: Error en precio "{precio_str}" - {str(e)}')
                except Exception as e:
                    errores.append(f'Línea {i}: Error inesperado - {str(e)}')
            
            # ========================================
            # EJECUTAR OPERACIONES EN LOTE (BULK)
            # ========================================
            
            # Crear productos nuevos en lote
            if productos_para_crear:
                print(f"IMPORTACIÓN: Creando {len(productos_para_crear)} productos nuevos en lote (batch: {batch_size})...", flush=True)
                ProductoCatalogo.objects.bulk_create(productos_para_crear, batch_size=batch_size)
                print(f"IMPORTACIÓN: ✅ {len(productos_para_crear)} productos creados exitosamente", flush=True)
            
            # Actualizar productos existentes en lote
            if productos_para_actualizar:
                print(f"IMPORTACIÓN: Actualizando {len(productos_para_actualizar)} productos en lote (batch: {batch_size})...", flush=True)
                ProductoCatalogo.objects.bulk_update(
                    productos_para_actualizar, 
                    ['descripcion', 'precio', 'fecha_actualizacion'], 
                    batch_size=batch_size
                )
                print(f"IMPORTACIÓN: ✅ {len(productos_para_actualizar)} productos actualizados exitosamente", flush=True)
            
            # Registrar importación
            if productos_procesados:
                importacion = ImportacionProductos.objects.create(
                    usuario=request.user,
                    productos_importados=len(productos_procesados),
                    productos_actualizados=productos_actualizados,
                    productos_nuevos=productos_nuevos,
                    observaciones=f'Errores: {len(errores)}' if errores else 'Importación exitosa'
                )
            
            # Mostrar resultados detallados
            if productos_procesados:
                mensaje_exito = f'✅ Importación completada: {productos_nuevos} productos nuevos, {productos_actualizados} actualizados'
                if productos_duplicados > 0:
                    mensaje_exito += f', {productos_duplicados} duplicados ignorados'
                if errores:
                    mensaje_exito += f'. ⚠️ {len(errores)} errores encontrados.'
                messages.success(request, mensaje_exito)
                
                # Log detallado para debugging
                print(f"IMPORTACIÓN COMPLETADA:", flush=True)
                print(f"  - Líneas procesadas: {len(lineas)}", flush=True)
                print(f"  - Productos nuevos: {productos_nuevos}", flush=True)
                print(f"  - Productos actualizados: {productos_actualizados}", flush=True)
                print(f"  - Productos duplicados: {productos_duplicados}", flush=True)
                print(f"  - Errores: {len(errores)}", flush=True)
            
            if errores:
                for error in errores[:5]:  # Mostrar solo los primeros 5 errores
                    messages.warning(request, error)
                if len(errores) > 5:
                    messages.warning(request, f'... y {len(errores) - 5} errores más.')
            
            return redirect('gestion_productos')
            
        except Exception as e:
            messages.error(request, f'Error al procesar la importación: {str(e)}')
            return redirect('gestion_productos')
    
    # GET request - mostrar formulario
    from app.models import Marca, ProductoCatalogo, ImportacionProductos
    from django.utils import timezone
    from datetime import timedelta
    
    # Estadísticas
    total_productos = ProductoCatalogo.objects.filter(activo=True).count()
    total_marcas = Marca.objects.filter(activa=True).count()
    
    # Días desde última importación
    ultima_importacion = ImportacionProductos.objects.first()
    if ultima_importacion:
        dias_ultima_importacion = (timezone.now() - ultima_importacion.fecha_importacion).days
    else:
        dias_ultima_importacion = 'N/A'
    
    # Productos recientes (últimos 50)
    productos = ProductoCatalogo.objects.filter(activo=True).select_related('marca').order_by('-fecha_actualizacion')[:50]
    
    # Marcas activas
    marcas = Marca.objects.filter(activa=True).order_by('nombre')
    
    # Importaciones recientes
    importaciones_recientes = ImportacionProductos.objects.select_related('usuario', 'marca').order_by('-fecha_importacion')[:10]
    
    context = {
        'total_productos': total_productos,
        'total_marcas': total_marcas,
        'dias_ultima_importacion': dias_ultima_importacion,
        'productos': productos,
        'marcas': marcas,
        'importaciones_recientes': importaciones_recientes,
    }
    
    return render(request, 'gestion_productos.html', context)

@login_required
def get_productos_por_marca_api(request):
    """
    API para obtener productos por marca para cotizaciones automáticas
    """
    marca = request.GET.get('marca', '').upper()
    query = request.GET.get('query', '').lower()
    
    if not marca:
        return JsonResponse({'productos': []})
    
    try:
        from app.models import ProductoCatalogo, Marca
        
        # Obtener la marca
        marca_obj = Marca.objects.filter(nombre=marca, activa=True).first()
        if not marca_obj:
            return JsonResponse({'productos': []})
        
        # Obtener productos de la marca
        productos = ProductoCatalogo.objects.filter(
            marca=marca_obj,
            activo=True
        )
        
        # Filtrar por consulta si se proporciona
        if query:
            productos = productos.filter(
                Q(no_parte__icontains=query) |
                Q(descripcion__icontains=query)
            )
        
        # Limitar resultados
        productos = productos.order_by('no_parte')[:20]
        
        # Formatear respuesta
        productos_data = []
        for producto in productos:
            productos_data.append({
                'no_parte': producto.no_parte,
                'descripcion': producto.descripcion,
                'precio': float(producto.precio),
                'marca': producto.marca.nombre
            })
        
        return JsonResponse({
            'productos': productos_data,
            'marca': marca,
            'query': query
        })
        
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'productos': []
        })

# ========================================
# COTIZACIONES AUTOMÁTICAS DESDE BITRIX24
# ========================================

import re

def es_cotizacion_automatica(deal_details):
    """
    Detecta si una oportunidad debe generar cotización automática
    basado en si el campo producto es una de las 7 marcas principales
    """
    producto_bitrix_id = deal_details.get('UF_CRM_1752859685662')
    
    # IDs de las 7 marcas que tienen cotización automática
    marcas_automaticas = [
        "176",  # ZEBRA
        "178",  # PANDUIT
        "180",  # APC
        "182",  # AVIGILON
        "184",  # GENETEC
        "186",  # AXIS
        "194",  # CISCO
    ]
    
    return str(producto_bitrix_id) in marcas_automaticas

def extraer_productos_inteligente(texto_requisicion, marca_seleccionada):
    """
    Parser súper inteligente que extrae números de parte y cantidades
    de cualquier formato de texto libre
    """
    if not texto_requisicion or not marca_seleccionada:
        return []
    
    productos_encontrados = []
    
    # Limpiar texto y dividir en líneas
    lineas = texto_requisicion.strip().replace('\r\n', '\n').split('\n')
    
    for linea in lineas:
        linea = linea.strip()
        if not linea:
            continue
            
        # Buscar patrones flexibles: texto alfanumérico + número
        # Patrones que detecta:
        # - "ZT411-203DPI 2"
        # - "ZT411 cantidad: 3"
        # - "necesito 5 del ZT411-203DPI"
        # - "450145 x3"
        # - "ZT411 (2 piezas)"
        
        # Patrón principal: buscar código alfanumérico y número cercano
        patron_principal = r'([A-Z0-9\-_]+).*?(\d+)'
        matches = re.findall(patron_principal, linea, re.IGNORECASE)
        
        for posible_parte, cantidad_str in matches:
            # Limpiar el posible número de parte
            posible_parte = posible_parte.strip().upper()
            
            # Verificar que sea un código válido (al menos 3 caracteres)
            if len(posible_parte) < 3:
                continue
                
            # Verificar que existe en nuestro catálogo para esta marca
            from app.models import ProductoCatalogo, Marca
            try:
                marca_obj = Marca.objects.get(nombre=marca_seleccionada, activa=True)
                producto = ProductoCatalogo.objects.filter(
                    marca=marca_obj,
                    no_parte__iexact=posible_parte,
                    activo=True
                ).first()
                
                if producto:
                    cantidad = int(cantidad_str)
                    productos_encontrados.append({
                        'no_parte': producto.no_parte,
                        'descripcion': producto.descripcion,
                        'precio': float(producto.precio),
                        'cantidad': cantidad,
                        'total': float(producto.precio) * cantidad
                    })
                    print(f"COTIZACIÓN AUTOMÁTICA: Producto encontrado: {producto.no_parte} x{cantidad}")
                    
            except (Marca.DoesNotExist, ValueError) as e:
                print(f"COTIZACIÓN AUTOMÁTICA: Error procesando {posible_parte}: {e}")
                continue
    
    return productos_encontrados

def crear_cotizacion_automatica_bitrix(deal_details, cliente, usuario, porcentaje_utilidad=0):
    """
    Crea una cotización automática basada en los datos de Bitrix24
    """
    try:
        # Obtener la marca del campo producto
        PRODUCTO_BITRIX_ID_TO_DJANGO_VALUE = {
            "176": "ZEBRA", "178": "PANDUIT", "180": "APC", "182": "AVIGILON",
            "184": "GENETEC", "186": "AXIS", "194": "CISCO"
        }
        
        producto_bitrix_id = deal_details.get('UF_CRM_1752859685662')
        marca_seleccionada = PRODUCTO_BITRIX_ID_TO_DJANGO_VALUE.get(str(producto_bitrix_id))
        
        if not marca_seleccionada:
            print(f"COTIZACIÓN AUTOMÁTICA: Marca no válida para ID {producto_bitrix_id}")
            return None
            
        # Obtener el texto de requisición del cliente desde COMMENTS
        # Los ejecutivos pueden escribir números de parte aquí en cualquier formato
        requisicion_cliente = deal_details.get('COMMENTS', '')
        
        print(f"COTIZACIÓN AUTOMÁTICA: Texto de requisición obtenido: '{requisicion_cliente[:100]}...'", flush=True)
        
        # Extraer productos inteligentemente
        productos = extraer_productos_inteligente(requisicion_cliente, marca_seleccionada)
        
        if not productos:
            print(f"COTIZACIÓN AUTOMÁTICA: No se encontraron productos válidos en la requisición")
            return None
            
        # Crear la cotización
        from app.models import Cotizacion, DetalleCotizacion
        
        # Usar el título del deal directamente
        titulo_cotizacion = deal_details.get('TITLE', 'Sin título')

        # Encontrar la oportunidad local correspondiente
        oportunidad_local = None
        deal_id = deal_details.get('ID')
        if deal_id:
            oportunidad_local = TodoItem.objects.filter(bitrix_deal_id=deal_id).first()
        
        cotizacion = Cotizacion.objects.create(
            cliente=cliente,
            created_by=usuario,
            titulo=titulo_cotizacion,
            nombre_cotizacion=titulo_cotizacion,
            oportunidad=oportunidad_local, # Asociar la oportunidad local
            descripcion="", # Descripción vacía
            moneda='USD',
            iva_rate=0.16,  # 16% por defecto
            tipo_cotizacion='Iamet',  # IAMET por defecto
            subtotal=0,  # Se calculará después
            iva_amount=0,  # Se calculará después
            total=0,  # Se calculará después
            descuento_visible=False
        )
        
        # Agregar productos a la cotización
        subtotal = 0
        print(f"COTIZACIÓN AUTOMÁTICA: Aplicando {porcentaje_utilidad}% de utilidad a los precios", flush=True)
        
        for i, producto in enumerate(productos, 1):
            # Aplicar porcentaje de utilidad al precio base
            precio_base = producto['precio']
            if porcentaje_utilidad > 0:
                factor_utilidad = 1 + (porcentaje_utilidad / 100)
                precio_con_utilidad = precio_base * factor_utilidad
                total_con_utilidad = precio_con_utilidad * producto['cantidad']
                print(f"COTIZACIÓN AUTOMÁTICA: Producto {producto['no_parte']} - Precio base: ${precio_base:.2f} → Con {porcentaje_utilidad}% utilidad: ${precio_con_utilidad:.2f}", flush=True)
            else:
                precio_con_utilidad = precio_base
                total_con_utilidad = precio_base * producto['cantidad']
            
            DetalleCotizacion.objects.create(
                cotizacion=cotizacion,
                orden=i,
                marca=marca_seleccionada,
                nombre_producto=producto['no_parte'],
                no_parte=producto['no_parte'],
                descripcion=producto['descripcion'],
                cantidad=producto['cantidad'],
                precio_unitario=precio_con_utilidad,
                descuento_porcentaje=0,
                total=total_con_utilidad
            )
            subtotal += total_con_utilidad
        
        # Calcular totales
        iva_amount = Decimal(str(subtotal)) * Decimal('0.16')
        total = Decimal(str(subtotal)) + iva_amount
        
        # Actualizar cotización con totales
        cotizacion.subtotal = Decimal(str(subtotal))
        cotizacion.iva_amount = iva_amount
        cotizacion.total = total
        cotizacion.save()
        
        print(f"COTIZACIÓN AUTOMÁTICA: Creada cotización {cotizacion.id} con {len(productos)} productos. Total: ${total:.2f}")
        return cotizacion
        
    except Exception as e:
        print(f"COTIZACIÓN AUTOMÁTICA: Error creando cotización: {e}")
        return None

def subir_cotizacion_a_bitrix(cotizacion, deal_id, request=None):
    """
    Genera el PDF de la cotización y lo sube como comentario a Bitrix24
    """
    try:
        from django.test import RequestFactory
        import base64
        
        # Crear una petición temporal para generar el PDF
        factory = RequestFactory()
        pdf_request = factory.get(f'/app/cotizacion/{cotizacion.id}/pdf/')
        pdf_request.user = cotizacion.created_by
        
        # Generar el PDF
        pdf_response = generate_cotizacion_pdf(pdf_request, cotizacion.id)
        
        if pdf_response.status_code == 200:
            # Convertir PDF a base64
            pdf_content = pdf_response.content
            pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
            
            # Nombre del archivo
            pdf_name_raw = cotizacion.nombre_cotizacion or f"Cotizacion_{cotizacion.id}"
            sanitized_name = "".join(c for c in pdf_name_raw if c.isalnum() or c in ('_', '-')).strip().replace(' ', '_')
            pdf_filename = f"{sanitized_name}.pdf"
            
            # Comentario para Bitrix
            comentario_texto = f"""
🤖 Cotización Automática Generada

Cotización: {cotizacion.nombre_cotizacion}
Total: ${float(cotizacion.total):.2f} {cotizacion.moneda}
Productos: {cotizacion.detalles.count()} artículos
Generada por: {cotizacion.created_by.get_full_name() or cotizacion.created_by.username}
Fecha: {convert_to_tijuana_time(cotizacion.fecha_creacion).strftime('%d/%m/%Y %H:%M')}

⚡ Generada automáticamente desde el sistema Nethive
            """.strip()
            
            # Subir a Bitrix24
            from .bitrix_integration import add_comment_with_attachment_to_deal
            resultado_bitrix = add_comment_with_attachment_to_deal(
                deal_id=deal_id,
                file_name=pdf_filename,
                file_content_base64=pdf_base64,
                comment_text=comentario_texto,
                request=request
            )
            
            if resultado_bitrix:
                print(f"COTIZACIÓN AUTOMÁTICA: PDF subido exitosamente a Bitrix24 deal {deal_id}")
                return True
            else:
                print(f"COTIZACIÓN AUTOMÁTICA: Error subiendo PDF a Bitrix24 deal {deal_id}")
                return False
                
        else:
            print(f"COTIZACIÓN AUTOMÁTICA: Error generando PDF (status {pdf_response.status_code})")
            return False
            
    except Exception as e:
        print(f"COTIZACIÓN AUTOMÁTICA: Error subiendo a Bitrix24: {e}")
        return False

# ========================================
# FUNCIÓN DE TEST PARA COTIZACIONES AUTOMÁTICAS
# ========================================

@login_required
def test_cotizacion_automatica(request):
    """
    Función de test para probar las cotizaciones automáticas sin Bitrix24
    Solo accesible para superusuarios
    """
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Acceso denegado'}, status=403)
    
    try:
        # Simular datos de Bitrix24
        deal_details_test = {
            'ID': '999999',  # ID de prueba
            'TITLE': 'Oportunidad de Prueba - Cotización Automática',
            'OPPORTUNITY': '5000.00',
            'COMPANY_ID': '123',
            'CONTACT_ID': '456',
            'UF_CRM_1752859685662': '176',  # ZEBRA
            'COMMENTS': 'ZT411-203DPI 2\n450145 x3\nNecesito 1 unidad del 20-61019-04R',
        }
        
        # Obtener o crear cliente de prueba
        from app.models import Cliente, User
        cliente_test, _ = Cliente.objects.get_or_create(
            nombre_empresa='Cliente Test Automático',
            defaults={'bitrix_company_id': '123'}
        )
        
        # Usar el usuario actual
        usuario_test = request.user
        
        print("=== INICIANDO TEST DE COTIZACIÓN AUTOMÁTICA ===")
        
        # Test 1: Verificar detección de cotización automática
        es_automatica = es_cotizacion_automatica(deal_details_test)
        print(f"✓ Test detección automática: {es_automatica}")
        
        # Test 2: Probar el parser inteligente
        productos = extraer_productos_inteligente(deal_details_test['COMMENTS'], 'ZEBRA')
        print(f"✓ Test parser inteligente: {len(productos)} productos encontrados")
        for p in productos:
            print(f"  - {p['no_parte']} x{p['cantidad']} = ${p['total']:.2f}")
        
        # Test 3: Crear cotización automática
        if es_automatica and productos:
            cotizacion_test = crear_cotizacion_automatica_bitrix(deal_details_test, cliente_test, usuario_test)
            if cotizacion_test:
                print(f"✓ Test creación cotización: ID {cotizacion_test.id}, Total: ${cotizacion_test.total:.2f}")
                
                return JsonResponse({
                    'success': True,
                    'message': 'Test de cotización automática completado exitosamente',
                    'cotizacion_id': cotizacion_test.id,
                    'total': float(cotizacion_test.total),
                    'productos_count': len(productos),
                    'productos': productos
                })
            else:
                return JsonResponse({'error': 'Falló la creación de la cotización'}, status=400)
        else:
            return JsonResponse({'error': 'Test falló en la detección o parsing'}, status=400)
            
    except Exception as e:
        print(f"ERROR EN TEST: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def buscar_producto_catalogo(request):
    """
    API endpoint para buscar productos en el catálogo por número de parte
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    numero_parte = request.GET.get('numero_parte', '').strip()
    tipo_producto = request.GET.get('tipo', '').strip()
    
    if not numero_parte:
        return JsonResponse({'error': 'Número de parte es requerido'}, status=400)
    
    try:
        # Buscar producto por número de parte
        query = CatalogoCableado.objects.filter(
            numero_parte__icontains=numero_parte,
            activo=True
        )
        
        # Filtrar por tipo si se especifica
        if tipo_producto:
            query = query.filter(tipo_producto=tipo_producto.upper())
        
        producto = query.first()
        
        if producto:
            return JsonResponse({
                'success': True,
                'producto': {
                    'numero_parte': producto.numero_parte,
                    'descripcion': producto.descripcion,
                    'tipo_producto': producto.tipo_producto,
                    'marca': producto.marca,
                    'precio_unitario': float(producto.precio_unitario),
                    'precio_proveedor': float(producto.precio_proveedor),
                    'categoria': producto.categoria,
                    'color': producto.color,
                    'activo': producto.activo
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Producto no encontrado en el catálogo'
            })
            
    except Exception as e:
        print(f"ERROR buscando producto: {e}")
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)

@login_required  
def agregar_producto_catalogo(request):
    """
    API endpoint para agregar productos al catálogo
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Validar campos requeridos
        required_fields = ['numero_parte', 'descripcion', 'tipo_producto', 'precio_unitario']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({'error': f'Campo {field} es requerido'}, status=400)
        
        # Verificar si el producto ya existe
        if CatalogoCableado.objects.filter(numero_parte=data['numero_parte']).exists():
            return JsonResponse({'error': 'Ya existe un producto con este número de parte'}, status=400)
        
        # Crear nuevo producto
        producto = CatalogoCableado.objects.create(
            numero_parte=data['numero_parte'],
            descripcion=data['descripcion'],
            tipo_producto=data['tipo_producto'].upper(),
            marca=data.get('marca', 'PANDUIT'),
            precio_unitario=Decimal(str(data['precio_unitario'])),
            precio_proveedor=Decimal(str(data.get('precio_proveedor', 0))),
            categoria=data.get('categoria', ''),
            color=data.get('color', ''),
            activo=data.get('activo', True)
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Producto agregado exitosamente',
            'producto_id': producto.id,
            'numero_parte': producto.numero_parte
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)
    except Exception as e:
        print(f"ERROR agregando producto: {e}")
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)

@login_required
def gestion_catalogo_volumetria(request):
    """
    Vista para gestión del catálogo de productos de volumetría - Solo para superusuarios
    """
    # Verificar que el usuario sea superusuario
    if not request.user.is_superuser:
        messages.error(request, 'No tiene permisos para acceder a esta sección.')
        return redirect('home')
    
    if request.method == 'POST':
        try:
            action = request.POST.get('action')
            
            if action == 'agregar_producto':
                # Agregar un nuevo producto al catálogo
                numero_parte = request.POST.get('numero_parte', '').strip()
                descripcion = request.POST.get('descripcion', '').strip()
                categoria = request.POST.get('categoria', '').strip()
                color = request.POST.get('color', '').strip()
                precio_unitario = request.POST.get('precio_unitario', '0')
                precio_proveedor = request.POST.get('precio_proveedor', '0')
                marca = request.POST.get('marca', 'PANDUIT').strip()
                
                # Validaciones
                if not numero_parte or not descripcion:
                    messages.error(request, 'Número de parte y descripción son obligatorios.')
                    return redirect('gestion_catalogo_volumetria')
                
                # Verificar si ya existe
                if CatalogoCableado.objects.filter(numero_parte=numero_parte).exists():
                    messages.error(request, f'Ya existe un producto con el número de parte: {numero_parte}')
                    return redirect('gestion_catalogo_volumetria')
                
                # Crear producto
                CatalogoCableado.objects.create(
                    numero_parte=numero_parte,
                    descripcion=descripcion,
                    categoria=categoria,
                    color=color.upper(),
                    precio_unitario=Decimal(precio_unitario),
                    precio_proveedor=Decimal(precio_proveedor),
                    marca=marca.upper(),
                    activo=True
                )
                
                messages.success(request, f'Producto {numero_parte} agregado exitosamente.')
                return redirect('gestion_catalogo_volumetria')
                
            elif action == 'importar_masivo':
                # Importación masiva desde texto usando parser inteligente
                productos_texto = request.POST.get('productos_texto', '').strip()
                actualizar_existentes = request.POST.get('actualizar_existentes') == 'on'
                
                if not productos_texto:
                    messages.error(request, 'Por favor, ingrese los datos de productos.')
                    return redirect('gestion_catalogo_volumetria')
                
                productos_procesados = 0
                productos_actualizados = 0
                errores = []
                
                lineas = productos_texto.split('\n')
                for i, linea in enumerate(lineas, 1):
                    linea = linea.strip()
                    if not linea:
                        continue
                    
                    try:
                        # Usar parser inteligente para detectar productos automáticamente
                        producto_data = parse_producto_volumetria_inteligente(linea)
                        
                        if not producto_data:
                            errores.append(f'Línea {i}: No se pudo detectar el patrón número_parte-marca-descripción-precios')
                            continue
                        
                        numero_parte = producto_data['numero_parte']
                        descripcion = producto_data['descripcion']
                        categoria = producto_data['categoria']
                        color = producto_data['color']
                        precio_unitario = Decimal(str(producto_data['precio_lista']))
                        precio_proveedor = Decimal(str(producto_data['precio_proveedor']))
                        marca = producto_data['marca']
                        
                        if not numero_parte or not descripcion:
                            errores.append(f'Línea {i}: Número de parte y descripción son obligatorios')
                            continue
                        
                        # Verificar si existe
                        producto_existente = CatalogoCableado.objects.filter(numero_parte=numero_parte).first()
                        
                        if producto_existente:
                            if actualizar_existentes:
                                producto_existente.descripcion = descripcion
                                producto_existente.categoria = categoria
                                producto_existente.color = color
                                producto_existente.precio_unitario = precio_unitario
                                producto_existente.precio_proveedor = precio_proveedor
                                producto_existente.marca = marca
                                producto_existente.save()
                                productos_actualizados += 1
                            else:
                                errores.append(f'Línea {i}: Producto {numero_parte} ya existe (no se actualiza)')
                        else:
                            CatalogoCableado.objects.create(
                                numero_parte=numero_parte,
                                descripcion=descripcion,
                                categoria=categoria,
                                color=color,
                                precio_unitario=precio_unitario,
                                precio_proveedor=precio_proveedor,
                                marca=marca,
                                activo=True
                            )
                            productos_procesados += 1
                            
                    except Exception as e:
                        errores.append(f'Línea {i}: Error procesando - {str(e)}')
                
                # Mostrar resultados
                if productos_procesados > 0:
                    messages.success(request, f'Se agregaron {productos_procesados} productos nuevos.')
                if productos_actualizados > 0:
                    messages.success(request, f'Se actualizaron {productos_actualizados} productos existentes.')
                if errores:
                    for error in errores[:10]:  # Mostrar solo los primeros 10 errores
                        messages.warning(request, error)
                    if len(errores) > 10:
                        messages.warning(request, f'Y {len(errores) - 10} errores más...')
                
                return redirect('gestion_catalogo_volumetria')
                
        except Exception as e:
            messages.error(request, f'Error procesando solicitud: {str(e)}')
            return redirect('gestion_catalogo_volumetria')
    
    # GET request - Mostrar la página
    productos = CatalogoCableado.objects.all().order_by('categoria', 'numero_parte')
    
    # Estadísticas
    total_productos = productos.count()
    total_categorias = productos.values('categoria').distinct().count()
    productos_sin_precio = productos.filter(precio_unitario=0).count()
    
    # Categorías disponibles para filtro
    categorias = productos.values_list('categoria', flat=True).distinct().order_by('categoria')
    
    context = {
        'productos': productos,
        'total_productos': total_productos,
        'total_categorias': total_categorias,
        'productos_sin_precio': productos_sin_precio,
        'categorias': categorias,
    }
    
    return render(request, 'gestion_catalogo_volumetria.html', context)

@login_required
def eliminar_producto_catalogo(request, producto_id):
    """
    Eliminar un producto del catálogo de volumetría
    """
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        producto = CatalogoCableado.objects.get(id=producto_id)
        numero_parte = producto.numero_parte
        producto.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Producto {numero_parte} eliminado exitosamente'
        })
    except CatalogoCableado.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def editar_producto_catalogo(request, producto_id):
    """
    Editar un producto del catálogo de volumetría
    """
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        producto = CatalogoCableado.objects.get(id=producto_id)
        data = json.loads(request.body)
        
        # Actualizar campos
        producto.numero_parte = data.get('numero_parte', producto.numero_parte)
        producto.descripcion = data.get('descripcion', producto.descripcion)
        producto.categoria = data.get('categoria', producto.categoria)
        producto.color = data.get('color', producto.color).upper()
        producto.precio_unitario = Decimal(str(data.get('precio_unitario', producto.precio_unitario)))
        producto.precio_proveedor = Decimal(str(data.get('precio_proveedor', producto.precio_proveedor)))
        producto.marca = data.get('marca', producto.marca).upper()
        
        producto.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Producto {producto.numero_parte} actualizado exitosamente'
        })
        
    except CatalogoCableado.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def vista_previa_volumetria_api(request):
    """
    API para vista previa de importación masiva de productos de volumetría
    """
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        data = json.loads(request.body)
        productos_texto = data.get('productos_texto', '').strip()
        
        if not productos_texto:
            return JsonResponse({'error': 'No hay datos para procesar'}, status=400)
        
        productos_detectados = []
        errores = []
        
        lineas = productos_texto.split('\n')
        for i, linea in enumerate(lineas, 1):
            linea = linea.strip()
            if not linea:
                continue
            
            try:
                producto_data = parse_producto_volumetria_inteligente(linea)
                
                if producto_data:
                    # Verificar si ya existe en la base de datos
                    existe = CatalogoCableado.objects.filter(numero_parte=producto_data['numero_parte']).exists()
                    
                    productos_detectados.append({
                        'linea': i,
                        'numero_parte': producto_data['numero_parte'],
                        'marca': producto_data['marca'],
                        'descripcion': producto_data['descripcion'],
                        'categoria': producto_data['categoria'],
                        'color': producto_data['color'],
                        'precio_lista': float(producto_data['precio_lista']),
                        'precio_proveedor': float(producto_data['precio_proveedor']),
                        'existe': existe,
                        'status': 'actualizar' if existe else 'nuevo'
                    })
                else:
                    errores.append({
                        'linea': i,
                        'texto': linea[:100] + '...' if len(linea) > 100 else linea,
                        'error': 'No se pudo detectar el patrón número_parte-marca-descripción-precios'
                    })
            except Exception as e:
                errores.append({
                    'linea': i,
                    'texto': linea[:100] + '...' if len(linea) > 100 else linea,
                    'error': str(e)
                })
        
        # Agrupar por marca para mejor visualización
        productos_por_marca = {}
        for producto in productos_detectados:
            marca = producto['marca']
            if marca not in productos_por_marca:
                productos_por_marca[marca] = []
            productos_por_marca[marca].append(producto)
        
        return JsonResponse({
            'success': True,
            'productos_detectados': productos_detectados,
            'productos_por_marca': productos_por_marca,
            'errores': errores,
            'total_productos': len(productos_detectados),
            'total_nuevos': len([p for p in productos_detectados if not p['existe']]),
            'total_existentes': len([p for p in productos_detectados if p['existe']]),
            'total_errores': len(errores)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def sugerencias_productos_api(request):
    """
    API para obtener sugerencias de productos mientras se escribe
    """
    print(f"DEBUG: API sugerencias llamada con método: {request.method}")
    
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    query = request.GET.get('q', '').strip()
    limit = min(int(request.GET.get('limit', 10)), 20)  # Máximo 20 sugerencias
    
    print(f"DEBUG: Query recibida: '{query}', Limit: {limit}")
    
    if len(query) < 2:
        return JsonResponse({'success': True, 'sugerencias': [], 'total': 0})
    
    try:
        print(f"DEBUG: Buscando en CatalogoCableado...")
        
        # Verificar si hay productos en el catálogo
        total_productos = CatalogoCableado.objects.count()
        productos_activos = CatalogoCableado.objects.filter(activo=True).count()
        
        print(f"DEBUG: Total productos en catálogo: {total_productos}, Activos: {productos_activos}")
        
        # Buscar productos que contengan el texto en número de parte o descripción
        productos = CatalogoCableado.objects.filter(
            activo=True
        ).filter(
            Q(numero_parte__icontains=query) | 
            Q(descripcion__icontains=query)
        ).order_by(
            # Priorizar coincidencias exactas en número de parte
            Case(
                When(numero_parte__istartswith=query, then=Value(1)),
                When(numero_parte__icontains=query, then=Value(2)),
                When(descripcion__icontains=query, then=Value(3)),
                default=Value(4)
            ),
            'numero_parte'
        )[:limit]
        
        print(f"DEBUG: Productos encontrados: {productos.count()}")
        
        sugerencias = []
        for producto in productos:
            print(f"DEBUG: Procesando producto: {producto.numero_parte}")
            sugerencia = {
                'numero_parte': producto.numero_parte,
                'descripcion': producto.descripcion[:80] + '...' if len(producto.descripcion) > 80 else producto.descripcion,
                'marca': producto.marca if producto.marca else 'N/A',
                'categoria': producto.categoria if producto.categoria else 'N/A',
                'precio_lista': float(producto.precio_unitario) if producto.precio_unitario else 0.0,
                'precio_proveedor': float(producto.precio_proveedor) if producto.precio_proveedor else 0.0,
                'color': producto.color if producto.color else 'N/A'
            }
            sugerencias.append(sugerencia)
        
        response_data = {
            'success': True,
            'sugerencias': sugerencias,
            'total': len(sugerencias)
        }
        
        print(f"DEBUG: Respuesta exitosa con {len(sugerencias)} sugerencias")
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"ERROR obteniendo sugerencias: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': 'Error interno del servidor',
            'sugerencias': []
        }, status=500)

@login_required
def buscar_producto_catalogo_api(request):
    """
    API para buscar un producto específico en el catálogo de cableado
    por número de parte y devolver todos sus datos
    """
    numero_parte = request.GET.get('numero_parte', '').strip()
    
    if not numero_parte:
        return JsonResponse({
            'success': False,
            'error': 'Número de parte requerido'
        }, status=400)
    
    try:
        # Buscar producto por número de parte (case insensitive)
        producto = CatalogoCableado.objects.filter(
            numero_parte__iexact=numero_parte,
            activo=True
        ).first()
        
        if producto:
            data = {
                'success': True,
                'found': True,
                'producto': {
                    'numero_parte': producto.numero_parte,
                    'descripcion': producto.descripcion,
                    'precio_unitario': float(producto.precio_unitario),
                    'precio_proveedor': float(producto.precio_proveedor),
                    'marca': producto.marca,
                    'categoria': producto.categoria,
                    'color': producto.color if hasattr(producto, 'color') else '',
                    'tipo_producto': producto.tipo_producto,
                    # Calcular margen automáticamente
                    'margen_porcentaje': float(
                        ((producto.precio_unitario - producto.precio_proveedor) / producto.precio_unitario * 100) 
                        if producto.precio_unitario > 0 else 0
                    )
                }
            }
            print(f"✅ Producto encontrado: {producto.numero_parte} - ${producto.precio_unitario}")
        else:
            data = {
                'success': True,
                'found': False,
                'message': f'Producto {numero_parte} no encontrado en catálogo'
            }
            print(f"❌ Producto no encontrado: {numero_parte}")
        
        return JsonResponse(data)
        
    except Exception as e:
        print(f"ERROR buscando producto {numero_parte}: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': 'Error interno del servidor'
        }, status=500)

@login_required
@require_http_methods(["GET"])
def buscar_productos_catalogo(request):
    """
    Endpoint para buscar productos en el catálogo de volumetría con motor inteligente.
    Filtra por tipo, categoría, marca, color, uso y texto, y prioriza palabras clave en la descripción.
    """
    try:
        query = request.GET.get('q', '').strip()
        filtros_json = request.GET.get('filtros', '{}')
        tipo_producto = request.GET.get('tipo', '').strip()
        
        print(f"🔍 Búsqueda en catálogo: query='{query}', filtros='{filtros_json}', tipo='{tipo_producto}'")

        try:
            filtros = json.loads(filtros_json)
        except json.JSONDecodeError:
            filtros = {}

        # Si no hay ni query ni filtros, no devolver nada.
        if not query and not filtros:
             return JsonResponse({'success': True, 'productos': [], 'total': 0})

        # Consulta base sobre el modelo correcto: CatalogoCableado
        productos_query = CatalogoCableado.objects.filter(activo=True)

        # 1. Priorizar y/o filtrar por palabra clave en la descripción según el tipo de producto.
        if tipo_producto:
            keyword = ''
            if tipo_producto == 'cable':
                keyword = 'bobina'
                # Filtro restrictivo para cable
                productos_query = productos_query.filter(
                    Q(descripcion__icontains='bobina') | Q(descripcion__icontains='cable')
                )
                print(f"🔌 Filtro RESTRICTIVO para 'cable'/'bobina' aplicado.")
            elif tipo_producto == 'jack':
                keyword = 'jack'
            elif tipo_producto == 'patchcord':
                keyword = 'patchcord'
            elif tipo_producto == 'faceplate':
                keyword = 'faceplate'
            
            if keyword:
                productos_query = productos_query.annotate(
                    priority=Case(
                        When(descripcion__icontains=keyword, then=Value(1)),
                        default=Value(2)
                    )
                )
                print(f"🔌 Prioridad aplicada para '{keyword}' en tipo '{tipo_producto}'.")

        # 2. Filtrar por categoría si se proporciona en los filtros.
        categoria_filtro = filtros.get('categoria')
        if categoria_filtro:
            exact_category = ''
            if categoria_filtro == '6A':
                exact_category = 'Cat6A'
            elif categoria_filtro == '6':
                exact_category = 'Cat6'
            elif categoria_filtro == '5e':
                exact_category = 'Cat5e'
            
            if exact_category:
                productos_query = productos_query.filter(categoria__iexact=exact_category)
                print(f"🎯 Filtro de categoría exacto aplicado: {exact_category}")
            else:
                productos_query = productos_query.filter(categoria__icontains=categoria_filtro)
                print(f"🎯 Filtro de categoría icontains aplicado: {categoria_filtro}")

        # 3. Filtrar por marca, color y uso.
        if 'marca' in filtros and filtros['marca']:
            productos_query = productos_query.filter(marca__iexact=filtros['marca'])
            print(f"🏷️ Filtro de marca aplicado: {filtros['marca']}")
        
        if 'color' in filtros and filtros['color']:
            productos_query = productos_query.filter(color__iexact=filtros['color'])
            print(f"🎨 Filtro de color aplicado: {filtros['color']}")

        if 'uso' in filtros and filtros['uso']:
            productos_query = productos_query.filter(descripcion__icontains=filtros['uso'])
            print(f"🏠 Filtro de uso aplicado: {filtros['uso']}")

        # 4. Filtrar por el query de texto en número de parte y descripción.
        if query:
            productos_query = productos_query.filter(
                Q(numero_parte__icontains=query) | 
                Q(descripcion__icontains=query)
            )
            print(f"🎯 Filtro de texto aplicado: '{query}'")

        # Ordenar y limitar resultados
        if tipo_producto and 'priority' in productos_query.query.annotations:
            productos = productos_query.distinct().order_by('priority', 'numero_parte')[:50]
        else:
            productos = productos_query.distinct().order_by('numero_parte')[:50]
        
        print(f"📦 Productos encontrados después del filtrado: {productos.count()}")

        # Convertir a formato JSON para la respuesta.
        productos_data = []
        for producto in productos:
            productos_data.append({
                'numero_parte': producto.numero_parte,
                'descripcion': producto.descripcion,
                'precio_lista': float(producto.precio_unitario) if producto.precio_unitario else 0.0,
                'costo_unitario': float(producto.precio_proveedor) if producto.precio_proveedor else 0.0,
                'marca': producto.marca,
                'categoria': producto.categoria or '',
            })
        
        print(f"✅ Enviando {len(productos_data)} productos al frontend")
        
        return JsonResponse({
            'success': True,
            'productos': productos_data,
            'total': len(productos_data)
        })
        
    except Exception as e:
        print(f"❌ Error en búsqueda de catálogo: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["POST"])
def buscar_productos_por_numeros_parte(request):
    try:
        data = json.loads(request.body)
        numeros_parte = data.get('numeros_parte', [])

        if not numeros_parte:
            return JsonResponse({'success': False, 'error': 'No se proporcionaron números de parte'})

        productos = CatalogoCableado.objects.filter(numero_parte__in=numeros_parte, activo=True)
        
        productos_data = []
        for producto in productos:
            productos_data.append({
                'numero_parte': producto.numero_parte,
                'descripcion': producto.descripcion,
                'precio_lista': float(producto.precio_unitario) if producto.precio_unitario else 0.0,
                'costo_unitario': float(producto.precio_proveedor) if producto.precio_proveedor else 0.0,
                'marca': producto.marca,
                'categoria': producto.categoria or '',
            })

        return JsonResponse({'success': True, 'productos': productos_data})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def get_marcas_catalogo(request):
    marcas = CatalogoCableado.objects.values_list('marca', flat=True).distinct().order_by('marca')
    return JsonResponse(list(marcas), safe=False)


@login_required
def nueva_oportunidad(request):
    """
    Vista optimizada para crear nuevas oportunidades con mejor UX y automatización.
    """
    if request.method == 'POST':
        form = NuevaOportunidadForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                oportunidad = form.save()
                messages.success(request, f'Oportunidad "{oportunidad.oportunidad}" creada exitosamente.')
                return redirect('todos')  # Redirigir a la lista de oportunidades
            except Exception as e:
                messages.error(request, f'Error al crear la oportunidad: {str(e)}')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = NuevaOportunidadForm(user=request.user)
    
    # Obtener lista de clientes para autocompletado
    clientes = Cliente.objects.all().order_by('nombre_empresa')
    
    context = {
        'form': form,
        'clientes': clientes,
        'title': 'Nueva Oportunidad'
    }
    
    return render(request, 'nueva_oportunidad.html', context)


@login_required
def api_crear_oportunidad(request):
    """
    API AJAX para crear oportunidad desde el widget CRM.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)

    try:
        import json
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST

        cliente_nombre = data.get('cliente_nombre', '').strip()
        if not cliente_nombre or len(cliente_nombre) < 2:
            return JsonResponse({'ok': False, 'error': 'El nombre del cliente es requerido (mín. 2 caracteres).'})

        oportunidad_nombre = data.get('oportunidad', '').strip()
        if not oportunidad_nombre:
            return JsonResponse({'ok': False, 'error': 'El nombre de la oportunidad es requerido.'})

        monto = data.get('monto', 0)
        try:
            monto = Decimal(str(monto))
        except Exception:
            monto = Decimal('0')

        # Buscar o crear cliente
        cliente, _ = Cliente.objects.get_or_create(
            nombre_empresa__iexact=cliente_nombre,
            defaults={'nombre_empresa': cliente_nombre, 'asignado_a': request.user}
        )

        # Buscar o crear contacto
        contacto = None
        contacto_nombre = data.get('contacto_nombre', '').strip()
        if contacto_nombre:
            nombre_parts = contacto_nombre.split(' ', 1)
            contacto, _ = Contacto.objects.get_or_create(
                nombre__iexact=nombre_parts[0],
                cliente=cliente,
                defaults={
                    'nombre': nombre_parts[0],
                    'apellido': nombre_parts[1] if len(nombre_parts) > 1 else '',
                    'cliente': cliente
                }
            )

        tipo_neg = data.get('tipo_negociacion', 'runrate')
        # Asignar etapa inicial según tipo de negociación
        if tipo_neg == 'proyecto':
            etapa_corta_init = 'Oportunidad'
            etapa_completa_init = 'Oportunidad'
            etapa_color_init = '#FFFFFF'
        else:
            etapa_corta_init = 'En Solicitud'
            etapa_completa_init = 'Solicitud de Cotizacion'
            etapa_color_init = '#FFFFFF'

        from datetime import datetime as dt_create
        now_dt = dt_create.now()
        mes_actual = str(now_dt.month).zfill(2)

        todo = TodoItem(
            usuario=request.user,
            oportunidad=oportunidad_nombre,
            cliente=cliente,
            contacto=contacto,
            monto=monto,
            probabilidad_cierre=int(data.get('probabilidad_cierre', 25)),
            mes_cierre=data.get('mes_cierre', mes_actual),
            anio_cierre=now_dt.year,
            area=data.get('area', 'SISTEMAS'),
            producto=data.get('producto', 'SOFTWARE'),
            tipo_negociacion=tipo_neg,
            comentarios=data.get('comentarios', ''),
            etapa_corta=etapa_corta_init,
            etapa_completa=etapa_completa_init,
            etapa_color=etapa_color_init,
            po_number='', # Ensure PO is empty on creation
        )
        todo.save()

        # If there are comments, add them as a chat message
        if todo.comentarios:
            from .models import MensajeOportunidad
            MensajeOportunidad.objects.create(
                oportunidad=todo,
                usuario=request.user,
                texto=todo.comentarios
            )

        return JsonResponse({
            'ok': True,
            'message': f'Oportunidad "{oportunidad_nombre}" creada exitosamente.',
            'id': todo.id
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required
def api_oportunidad_detalle_crm(request, oportunidad_id):
    """
    API para obtener detalle completo de una oportunidad para el widget CRM.
    """
    try:
        todo = get_object_or_404(TodoItem, pk=oportunidad_id)

        # Obtener cotizaciones de esta oportunidad
        cotizaciones = Cotizacion.objects.filter(oportunidad=todo).order_by('-fecha_creacion')
        cots_list = []
        for cot in cotizaciones:
            cots_list.append({
                'id': cot.id,
                'titulo': cot.titulo or f'COT-{cot.id}',
                'fecha': cot.fecha_creacion.strftime('%d %b %Y') if cot.fecha_creacion else '',
                'total': float(cot.total) if cot.total else 0,
                'moneda': cot.moneda or 'MXN',
            })

        data = {
            'id': todo.id,
            'oportunidad': todo.oportunidad or '',
            'monto': float(todo.monto) if todo.monto else 0,
            'cliente': {
                'id': todo.cliente_id,
                'nombre': todo.cliente.nombre_empresa if todo.cliente else '',
            } if todo.cliente else None,
            'contacto': '',
            'contacto_id': todo.contacto_id,
            'producto': todo.producto or '',
            'area': todo.area or '',
            'probabilidad_cierre': todo.probabilidad_cierre or 0,
            'po_number': todo.po_number or '',
            'mes_cierre': todo.mes_cierre or '',
            'tipo_negociacion': todo.tipo_negociacion or 'runrate',
            'etapa_corta': todo.etapa_corta or '',
            'etapa_completa': todo.etapa_completa or '',
            'etapa_color': todo.etapa_color or '#FFFFFF',
            'usuario': todo.usuario.get_full_name() or todo.usuario.username if todo.usuario else '',
            'usuario_id': todo.usuario_id,
            'comentarios': todo.comentarios or '',
            'fecha_creacion': todo.fecha_creacion.strftime('%d/%m/%Y') if todo.fecha_creacion else '',
            'cotizaciones': cots_list,
            'productos_adicionales': [
                {'id': p.id, 'producto': p.producto, 'notas': p.notas}
                for p in todo.productos_adicionales.all()
            ],
        }

        # Contacto
        if todo.contacto:
            if hasattr(todo.contacto, 'nombre'):
                data['contacto'] = f"{todo.contacto.nombre} {todo.contacto.apellido or ''}".strip()
            else:
                data['contacto'] = str(todo.contacto)

        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
def api_oportunidad_productos(request, oportunidad_id):
    """Lista y agrega productos adicionales a una oportunidad."""
    todo = get_object_or_404(TodoItem, pk=oportunidad_id)
    if request.method == 'GET':
        prods = [{'id': p.id, 'producto': p.producto, 'notas': p.notas}
                 for p in todo.productos_adicionales.all()]
        return JsonResponse({'productos': prods})
    elif request.method == 'POST':
        data = json.loads(request.body)
        producto = data.get('producto', '').strip()
        if not producto:
            return JsonResponse({'error': 'Producto requerido.'}, status=400)
        p = ProductoOportunidad.objects.create(
            oportunidad=todo,
            producto=producto,
            notas=data.get('notas', ''),
        )
        return JsonResponse({'id': p.id, 'producto': p.producto, 'notas': p.notas}, status=201)
    return JsonResponse({'error': 'Método no permitido.'}, status=405)


@login_required
@csrf_exempt
def api_oportunidad_producto_delete(request, oportunidad_id, producto_id):
    """Elimina un producto adicional de una oportunidad."""
    p = get_object_or_404(ProductoOportunidad, pk=producto_id, oportunidad_id=oportunidad_id)
    if request.method == 'DELETE':
        p.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Método no permitido.'}, status=405)


@login_required
def api_buscar_clientes(request):
    """
    API para autocompletado de clientes en el formulario de nueva oportunidad.
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'clientes': []})
    
    # Buscar clientes que coincidan con el query
    if is_supervisor(request.user):
        clientes = Cliente.objects.filter(
            nombre_empresa__icontains=query
        ).order_by('nombre_empresa')[:10]
    else:
        clientes = Cliente.objects.filter(
            Q(nombre_empresa__icontains=query) & 
            (Q(asignado_a=request.user) | Q(asignado_a__isnull=True))
        ).order_by('nombre_empresa')[:10]
    
    clientes_data = []
    for cliente in clientes:
        clientes_data.append({
            'id': cliente.id,
            'nombre': cliente.nombre_empresa,
            'contacto_principal': cliente.contacto_principal or '',
            'email': cliente.email or '',
            'telefono': cliente.telefono or ''
        })
    
    return JsonResponse({'clientes': clientes_data})


@login_required  
def api_buscar_contactos(request):
    """
    API para autocompletado de contactos basado en el cliente seleccionado.
    """
    cliente_id = request.GET.get('cliente_id')
    query = request.GET.get('q', '').strip()
    
    if not cliente_id:
        return JsonResponse({'contactos': []})
    
    try:
        contactos = Contacto.objects.filter(cliente_id=cliente_id)
        
        if query:
            contactos = contactos.filter(
                Q(nombre__icontains=query) | Q(apellido__icontains=query)
            )
        
        contactos = contactos.order_by('nombre')[:10]
        
        contactos_data = []
        for contacto in contactos:
            contactos_data.append({
                'id': contacto.id,
                'nombre_completo': f"{contacto.nombre} {contacto.apellido or ''}".strip(),
                'nombre': contacto.nombre,
                'apellido': contacto.apellido or ''
            })
        
        return JsonResponse({'contactos': contactos_data})
        
    except Exception as e:
        return JsonResponse({'contactos': [], 'error': str(e)}, status=500)


# ===============================================
# API PARA CRM AVANZADO DE OPORTUNIDADES
# ===============================================

@login_required
@require_http_methods(["POST"])
def cambiar_estado_oportunidad(request, oportunidad_id):
    """
    API para cambiar el estado CRM de una oportunidad
    """
    
    try:
        import json
        data = json.loads(request.body)
        estado = data.get('estado')
        
        if not estado:
            return JsonResponse({'error': 'Estado requerido'}, status=400)
        
        # Obtener la oportunidad
        oportunidad = get_object_or_404(TodoItem, id=oportunidad_id)
        
        # Validar que el usuario puede modificar esta oportunidad
        if not is_supervisor(request.user) and oportunidad.usuario != request.user:
            return JsonResponse({'error': 'No tienes permisos para modificar esta oportunidad'}, status=403)
        
        # Guardar estado anterior para actividad
        estado_anterior = oportunidad.estado_crm
        
        # Actualizar estado
        oportunidad.estado_crm = estado
        oportunidad.save()
        
        # Crear actividad en el timeline
        actividad = OportunidadActividad.objects.create(
            oportunidad=oportunidad,
            tipo='cambio_estado',
            titulo='Cambio de Estado',
            descripcion=f'Estado cambiado de "{estado_anterior}" a "{estado}"',
            usuario=request.user,
            estado_anterior=estado_anterior,
            estado_nuevo=estado
        )
        
        print(f"🔄 Actividad creada: {actividad.id}, estado_anterior: {actividad.estado_anterior}, estado_nuevo: {actividad.estado_nuevo}")
        
        # Preparar datos del timeline item para el frontend
        usuario_nombre = request.user.get_full_name() or request.user.username
        timeline_item = {
            'id': actividad.id,
            'tipo': 'cambio_estado',
            'titulo': 'Cambio de Estado',
            'descripcion': f'Estado cambiado de "{estado_anterior}" a "{estado}"',
            'usuario': usuario_nombre,
            'fecha': convert_to_tijuana_time(actividad.fecha_creacion).strftime('%d/%m/%Y %H:%M'),
            'icono': '🔄 Cambio de Estado',
            'estado_anterior': estado_anterior,
            'estado_nuevo': estado
        }
        
        return JsonResponse({
            'success': True,
            'nuevo_estado': estado,
            'message': f'Estado actualizado a {estado}',
            'timeline_item': timeline_item
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def agregar_comentario_oportunidad(request, oportunidad_id):
    """
    API para agregar comentarios con archivos a una oportunidad
    """
    try:
        # Obtener contenido del comentario
        contenido = request.POST.get('contenido', '').strip()
        
        # Obtener la oportunidad
        oportunidad = get_object_or_404(TodoItem, id=oportunidad_id)
        
        # Verificar permisos: supervisores ven todo, usuarios solo sus propias oportunidades
        if not is_supervisor(request.user) and oportunidad.usuario != request.user:
            return JsonResponse({'error': 'No tienes permisos para comentar en esta oportunidad'}, status=403)
        
        # Verificar que hay contenido o archivos
        archivos_subidos = []
        archivos_keys = [key for key in request.FILES.keys() if key.startswith('archivo_')]
        
        if not contenido and not archivos_keys:
            return JsonResponse({'error': 'Debe proporcionar contenido o archivos'}, status=400)
        
        # Crear comentario (puede estar vacío si solo hay archivos)
        comentario = OportunidadComentario.objects.create(
            oportunidad=oportunidad,
            usuario=request.user,
            contenido=contenido or "Archivo adjunto"
        )
        
        # Procesar archivos adjuntos
        print(f"📁 Procesando {len(archivos_keys)} archivos: {archivos_keys}")
        for key in archivos_keys:
            archivo = request.FILES[key]
            print(f"📄 Procesando archivo: {archivo.name}, tamaño: {archivo.size}, tipo: {archivo.content_type}")
            
            # Determinar tipo de archivo
            content_type = archivo.content_type.lower()
            if content_type.startswith('image/'):
                tipo_archivo = 'imagen'
            elif content_type in ['application/pdf']:
                tipo_archivo = 'documento'
            elif content_type in ['application/msword', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document']:
                tipo_archivo = 'documento'
            elif content_type in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'text/csv']:
                tipo_archivo = 'documento'
            else:
                tipo_archivo = 'otro'
            
            print(f"📋 Tipo determinado: {tipo_archivo}")
            
            try:
                # Crear registro de archivo
                archivo_obj = OportunidadArchivo.objects.create(
                    oportunidad=oportunidad,
                    usuario=request.user,
                    archivo=archivo,
                    nombre_original=archivo.name,
                    tipo=tipo_archivo,
                    tamaño=archivo.size,
                    descripcion=f"Adjuntado en comentario #{comentario.id}"
                )
                
                print(f"✅ Archivo guardado exitosamente: ID={archivo_obj.id}, URL={archivo_obj.archivo.url}")
                
                archivos_subidos.append({
                    'id': archivo_obj.id,
                    'nombre': archivo_obj.nombre_original,
                    'tipo': archivo_obj.tipo,
                    'tamaño': archivo_obj.tamaño,
                    'url': archivo_obj.archivo.url if archivo_obj.archivo else None
                })
                
            except Exception as e:
                print(f"❌ Error guardando archivo {archivo.name}: {e}")
                import traceback
                print(traceback.format_exc())
        
        # Crear actividad en el timeline con referencia al comentario
        descripcion_actividad = contenido[:200] + ('...' if len(contenido) > 200 else '')
        if archivos_subidos:
            if contenido:
                descripcion_actividad += f" ({len(archivos_subidos)} archivo{'s' if len(archivos_subidos) > 1 else ''} adjunto{'s' if len(archivos_subidos) > 1 else ''})"
            else:
                descripcion_actividad = f"Subió {len(archivos_subidos)} archivo{'s' if len(archivos_subidos) > 1 else ''}"
        
        # Agregar referencia al comentario en la descripción para linking directo
        descripcion_actividad += f" [COMENTARIO_ID:{comentario.id}]"
        
        print(f"🔥 Creando actividad - Usuario: {request.user}, Usuario ID: {request.user.id}, Nombre: {request.user.get_full_name()}")
        actividad_creada = OportunidadActividad.objects.create(
            oportunidad=oportunidad,
            tipo='comentario',
            titulo='Nuevo Comentario' + (' con archivos' if archivos_subidos else ''),
            descripcion=descripcion_actividad,
            usuario=request.user
        )
        print(f"💬 Actividad creada: ID={actividad_creada.id}, Usuario={actividad_creada.usuario}, Descripcion='{actividad_creada.descripcion}'")
        
        # ======================================
        # CREAR NOTIFICACIONES AUTOMÁTICAMENTE
        # ======================================
        
        # 1. Detectar menciones @usuario en el comentario
        if contenido:
            detectar_menciones_en_comentario(contenido, request.user, oportunidad, comentario)
        
        # 2. Notificar al dueño de la oportunidad (si no es el mismo que comenta)
        if oportunidad.usuario != request.user:
            mensaje_notif = f'{request.user.get_full_name() or request.user.username} comentó en tu oportunidad "{oportunidad.oportunidad}"'
            if contenido:
                mensaje_notif += f': {contenido[:100]}...' if len(contenido) > 100 else f': {contenido}'
            else:
                mensaje_notif += ' y adjuntó archivos'
                
            crear_notificacion(
                usuario_destinatario=oportunidad.usuario,
                tipo='comentario_oportunidad',
                titulo='Nuevo comentario en tu oportunidad',
                mensaje=mensaje_notif,
                oportunidad=oportunidad,
                comentario=comentario,
                usuario_remitente=request.user
            )
        
        # 3. Notificar a otros usuarios que han comentado en esta oportunidad (excepto el autor actual y el dueño)
        otros_comentaristas = User.objects.filter(
            oportunidadcomentario__oportunidad=oportunidad
        ).exclude(
            id__in=[request.user.id, oportunidad.usuario.id]
        ).distinct()
        
        for usuario in otros_comentaristas:
            mensaje_notif = f'{request.user.get_full_name() or request.user.username} también comentó en la oportunidad "{oportunidad.oportunidad}"'
            if contenido:
                mensaje_notif += f': {contenido[:100]}...' if len(contenido) > 100 else f': {contenido}'
            
            crear_notificacion(
                usuario_destinatario=usuario,
                tipo='comentario_oportunidad',
                titulo='Nuevo comentario en oportunidad que sigues',
                mensaje=mensaje_notif,
                oportunidad=oportunidad,
                comentario=comentario,
                usuario_remitente=request.user
            )
        
        return JsonResponse({
            'success': True,
            'comentario': {
                'id': comentario.id,
                'contenido': comentario.contenido,
                'usuario': request.user.get_full_name() or request.user.username,
                'fecha': convert_to_tijuana_time(comentario.fecha_creacion).strftime('%d/%m/%Y %H:%M'),
                'archivos': archivos_subidos
            }
        })
        
    except Exception as e:
        import traceback
        print(f"Error en agregar_comentario_oportunidad: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def timeline_oportunidad(request, oportunidad_id):
    """
    API para obtener el timeline completo de una oportunidad
    """
    from datetime import timedelta
    
    oportunidad = get_object_or_404(TodoItem, id=oportunidad_id)
    
    # Verificar permisos: supervisores ven todo, usuarios solo sus propias oportunidades
    if not is_supervisor(request.user) and oportunidad.usuario != request.user:
        return JsonResponse({'error': 'No tienes permisos para ver este timeline'}, status=403)
    
    try:
        # Limpiar actividades huérfanas antes de generar el timeline
        try:
            limpiar_actividades_huerfanas(oportunidad)
        except Exception as e:
            print(f"Error limpiando actividades huérfanas: {e}")
            # Continuar sin limpiar si hay error
        
        # Obtener todas las actividades
        actividades = oportunidad.actividades_crm.all().order_by('-fecha_creacion')
        
        timeline_data = []
        for actividad in actividades:
            # Obtener información del usuario
            usuario_nombre = 'Sistema'
            if actividad.usuario:
                usuario_nombre = actividad.usuario.get_full_name() or actividad.usuario.username
            else:
                usuario_nombre = 'Sistema'
            
            # Convertir fecha a zona horaria de Tijuana (Pacific Time)
            fecha_tijuana = convert_to_tijuana_time(actividad.fecha_creacion)
            
            item_data = {
                'id': actividad.id,
                'tipo': actividad.tipo,
                'titulo': actividad.titulo,
                'descripcion': actividad.descripcion,
                'usuario': usuario_nombre,
                'fecha': fecha_tijuana.strftime('%d/%m/%Y %H:%M'),
                'icono': dict(OportunidadActividad.TIPO_ACTIVIDAD_CHOICES).get(actividad.tipo, '⚙️')
            }
            
            # Agregar campos específicos según el tipo de actividad
            if actividad.tipo == 'cambio_estado':
                item_data['estado_anterior'] = actividad.estado_anterior
                item_data['estado_nuevo'] = actividad.estado_nuevo
            elif actividad.tipo == 'comentario':
                # Para comentarios, buscar el contenido real del comentario
                try:
                    # Nueva estrategia: buscar por ID directo en la descripción
                    comentario = None
                    
                    # Estrategia 1: Buscar por ID directo en la descripción
                    import re
                    match = re.search(r'\[COMENTARIO_ID:(\d+)\]', actividad.descripcion or '')
                    if match:
                        comentario_id = int(match.group(1))
                        try:
                            comentario = OportunidadComentario.objects.get(id=comentario_id)
                        except OportunidadComentario.DoesNotExist:
                            comentario = None
                    
                    # Estrategia 2 (fallback): Buscar por rango de tiempo
                    if not comentario:
                        comentarios_candidatos = OportunidadComentario.objects.filter(
                            oportunidad=oportunidad,
                            fecha_creacion__gte=actividad.fecha_creacion - timedelta(minutes=1),
                            fecha_creacion__lte=actividad.fecha_creacion + timedelta(minutes=1)
                        ).order_by('-fecha_creacion')
                        
                        if comentarios_candidatos.exists():
                            comentario = comentarios_candidatos.first()
                    
                    # Estrategia 3 (último recurso): Buscar por usuario y fecha cercana
                    if not comentario and actividad.usuario:
                        comentarios_por_usuario = OportunidadComentario.objects.filter(
                            oportunidad=oportunidad,
                            usuario=actividad.usuario
                        ).order_by('-fecha_creacion')
                        
                        for c in comentarios_por_usuario:
                            diff = abs((c.fecha_creacion - actividad.fecha_creacion).total_seconds())
                            if diff <= 300:  # 5 minutos
                                comentario = c
                                break
                    
                    if comentario:
                        item_data['contenido'] = comentario.contenido
                        item_data['comentario_id'] = comentario.id
                        
                        # Limpiar la descripción para mostrar solo el contenido real (sin el ID)
                        descripcion_limpia = re.sub(r' \[COMENTARIO_ID:\d+\]', '', actividad.descripcion or '')
                        item_data['descripcion'] = descripcion_limpia
                        item_data['puede_editar'] = comentario.usuario == request.user or is_supervisor(request.user)
                        
                        # Buscar archivos asociados a este comentario específico
                        # Usar la descripción del archivo que contiene el ID del comentario
                        archivos_asociados = OportunidadArchivo.objects.filter(
                            oportunidad=oportunidad,
                            descripcion__contains=f"Adjuntado en comentario #{comentario.id}"
                        )
                        
                        # Buscar archivos asociados específicamente por descripción
                        
                        # Si no encuentra por descripción exacta, NO usar fallback para evitar contaminación cruzada
                        # (El fallback por tiempo era lo que causaba que todos los archivos aparecieran en todos los comentarios)
                        
                        if archivos_asociados.exists():
                            item_data['archivos'] = []
                            for archivo in archivos_asociados:
                                item_data['archivos'].append({
                                    'id': archivo.id,
                                    'nombre': archivo.nombre_original,
                                    'tipo': archivo.tipo,
                                    'tamaño': archivo.tamaño_legible,
                                    'fecha': convert_to_tijuana_time(archivo.fecha_subida).strftime('%d/%m/%Y %H:%M'),
                                    'url': archivo.archivo.url if archivo.archivo else None
                                })
                        
                        # Actualizar usuario con el del comentario si está disponible
                        if comentario.usuario:
                            usuario_comentario = comentario.usuario.get_full_name() or comentario.usuario.username
                            item_data['usuario'] = usuario_comentario
                            item_data['usuario_id'] = comentario.usuario.id
                    else:
                        # Si no se encuentra comentario, saltar esta actividad para evitar huérfanas
                        continue
                        
                except Exception as e:
                    # Si hay error, saltar esta actividad
                    continue
            
            timeline_data.append(item_data)
        
        return JsonResponse({
            'success': True,
            'timeline': timeline_data
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
def editar_comentario_oportunidad(request, comentario_id):
    """
    API para editar un comentario específico de una oportunidad
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Obtener el comentario
        comentario = get_object_or_404(OportunidadComentario, id=comentario_id)
        
        # Verificar permisos: solo el autor del comentario o supervisores pueden editarlo
        if comentario.usuario != request.user and not is_supervisor(request.user):
            return JsonResponse({'error': 'No tienes permisos para editar este comentario'}, status=403)
        
        # Obtener el nuevo contenido
        nuevo_contenido = request.POST.get('contenido', '').strip()
        
        if not nuevo_contenido:
            return JsonResponse({'error': 'El contenido del comentario no puede estar vacío'}, status=400)
        
        # Actualizar el comentario
        comentario.contenido = nuevo_contenido
        comentario.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Comentario actualizado exitosamente',
            'nuevo_contenido': nuevo_contenido,
            'fecha_actualizacion': convert_to_tijuana_time(comentario.fecha_actualizacion).strftime('%d/%m/%Y %H:%M')
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt  
def eliminar_comentario_oportunidad(request, comentario_id):
    """
    API para eliminar un comentario específico de una oportunidad
    """
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Obtener el comentario
        comentario = get_object_or_404(OportunidadComentario, id=comentario_id)
        
        # Verificar permisos: solo el autor del comentario o supervisores pueden eliminarlo
        if comentario.usuario != request.user and not is_supervisor(request.user):
            return JsonResponse({'error': 'No tienes permisos para eliminar este comentario'}, status=403)
        
        # Guardar información antes de eliminar
        oportunidad_id = comentario.oportunidad.id
        oportunidad = comentario.oportunidad
        usuario_comentario = comentario.usuario
        fecha_comentario = comentario.fecha_creacion
        
        print(f"🗑️ Eliminando comentario ID={comentario_id}, usuario={usuario_comentario}, fecha={fecha_comentario}")
        
        # Buscar y eliminar TODAS las actividades que podrían estar apuntando a este comentario
        try:
            # Estrategia más amplia: buscar todas las actividades de comentario que podrían estar relacionadas
            actividades_candidatas = OportunidadActividad.objects.filter(
                oportunidad=oportunidad,
                tipo='comentario'
            )
            
            actividades_eliminadas = 0
            for actividad in actividades_candidatas:
                # Verificar si esta actividad apunta al comentario que vamos a eliminar
                # usando el nuevo sistema de IDs
                deberia_eliminar = False
                
                # Estrategia 1: Buscar por ID directo en la descripción (nuevo sistema)
                import re
                match = re.search(r'\[COMENTARIO_ID:(\d+)\]', actividad.descripcion or '')
                if match:
                    comentario_referenciado = int(match.group(1))
                    if comentario_referenciado == comentario_id:
                        deberia_eliminar = True
                        print(f"🎯 Actividad {actividad.id} apunta al comentario que se va a eliminar: {comentario_id}")
                else:
                    # Estrategia 2: Fallback para actividades del sistema viejo
                    # Verificar por rango de tiempo
                    diff_tiempo = abs((actividad.fecha_creacion - fecha_comentario).total_seconds())
                    if diff_tiempo <= 300:  # 5 minutos
                        deberia_eliminar = True
                        print(f"⏱️ Actividad {actividad.id} encontrada por tiempo: diff={diff_tiempo}s")
                    
                    # Verificar por usuario y descripción similar
                    if (actividad.usuario == usuario_comentario and 
                        comentario.contenido in actividad.descripcion):
                        deberia_eliminar = True
                        print(f"📝 Actividad {actividad.id} encontrada por contenido")
                
                if deberia_eliminar:
                    print(f"🗑️ Eliminando actividad relacionada ID={actividad.id}")
                    actividad.delete()
                    actividades_eliminadas += 1
            
            print(f"✅ Eliminadas {actividades_eliminadas} actividades relacionadas")
            
        except Exception as e:
            print(f"⚠️ Error eliminando actividades relacionadas: {e}")
            # No fallar si no se pueden eliminar las actividades, el comentario sí se debe eliminar
        
        # Eliminar el comentario
        comentario.delete()
        print(f"✅ Comentario ID={comentario_id} eliminado exitosamente")
        
        return JsonResponse({
            'success': True,
            'message': 'Comentario eliminado exitosamente',
            'oportunidad_id': oportunidad_id
        })
        
    except Exception as e:
        print(f"❌ Error eliminando comentario: {e}")
        return JsonResponse({'error': str(e)}, status=500)


def limpiar_actividades_huerfanas(oportunidad):
    """
    Función para limpiar actividades de comentario que ya no tienen comentario asociado
    """
    try:
        actividades_comentario = OportunidadActividad.objects.filter(
            oportunidad=oportunidad,
            tipo='comentario'
        )
        
        comentarios_existentes = OportunidadComentario.objects.filter(
            oportunidad=oportunidad
        )
        
        actividades_huerfanas = []
        
        for actividad in actividades_comentario:
            # Buscar si existe un comentario para esta actividad usando el nuevo sistema de IDs
            comentario_encontrado = False
            
            # Estrategia 1: Buscar por ID directo en la descripción (nuevo sistema)
            import re
            match = re.search(r'\[COMENTARIO_ID:(\d+)\]', actividad.descripcion or '')
            if match:
                comentario_id = int(match.group(1))
                if comentarios_existentes.filter(id=comentario_id).exists():
                    comentario_encontrado = True
                    print(f"✅ Actividad {actividad.id} tiene comentario válido: {comentario_id}")
                else:
                    print(f"❌ Actividad {actividad.id} referencia comentario inexistente: {comentario_id}")
            else:
                # Estrategia 2: Fallback para actividades del sistema viejo (por tiempo y usuario)
                for comentario in comentarios_existentes:
                    # Verificar por rango de tiempo
                    diff_tiempo = abs((actividad.fecha_creacion - comentario.fecha_creacion).total_seconds())
                    if (diff_tiempo <= 300 and  # 5 minutos
                        actividad.usuario == comentario.usuario):
                        comentario_encontrado = True
                        print(f"✅ Actividad {actividad.id} encontrada por tiempo: comentario {comentario.id}")
                        break
                    
                    # Verificar por contenido en descripción
                    if (actividad.usuario == comentario.usuario and 
                        comentario.contenido in actividad.descripcion):
                        comentario_encontrado = True
                        print(f"✅ Actividad {actividad.id} encontrada por contenido: comentario {comentario.id}")
                        break
            
            if not comentario_encontrado:
                actividades_huerfanas.append(actividad)
        
        # Eliminar actividades huérfanas
        if actividades_huerfanas:
            print(f"🧹 Limpiando {len(actividades_huerfanas)} actividades huérfanas")
            for actividad in actividades_huerfanas:
                print(f"🗑️ Eliminando actividad huérfana ID={actividad.id}")
                actividad.delete()
        
    except Exception as e:
        print(f"⚠️ Error limpiando actividades huérfanas: {e}")


@login_required
def descargar_archivo_oportunidad(request, archivo_id):
    """
    Vista para descargar archivos adjuntos de oportunidades
    """
    try:
        archivo = get_object_or_404(OportunidadArchivo, id=archivo_id)
        
        # Verificar permisos: supervisores ven todo, usuarios solo archivos de sus oportunidades
        if not is_supervisor(request.user) and archivo.oportunidad.usuario != request.user:
            return JsonResponse({'error': 'No tienes permisos para descargar este archivo'}, status=403)
        
        # Verificar que el archivo existe
        if not archivo.archivo:
            return JsonResponse({'error': 'Archivo no encontrado'}, status=404)
        
        # Importar las clases necesarias para la respuesta
        from django.http import HttpResponse, Http404
        from django.utils.encoding import smart_str
        import os
        import mimetypes
        
        # Obtener la ruta del archivo
        file_path = archivo.archivo.path
        
        if not os.path.exists(file_path):
            raise Http404("El archivo no existe en el servidor")
        
        # Determinar el tipo MIME
        content_type, _ = mimetypes.guess_type(file_path)
        if content_type is None:
            content_type = 'application/octet-stream'
        
        # Leer el archivo
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=content_type)
        
        # Configurar headers para descarga
        filename = smart_str(archivo.nombre_original)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = os.path.getsize(file_path)
        
        return response
        
    except Exception as e:
        print(f"❌ Error descargando archivo: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def vista_previa_archivo_oportunidad(request, archivo_id):
    """
    Vista para mostrar archivos en vista previa (inline)
    """
    try:
        archivo = get_object_or_404(OportunidadArchivo, id=archivo_id)
        
        # Verificar permisos: supervisores ven todo, usuarios solo archivos de sus oportunidades
        if not is_supervisor(request.user) and archivo.oportunidad.usuario != request.user:
            return HttpResponse('No tienes permisos para ver este archivo', status=403)
        
        # Verificar que el archivo existe
        if not archivo.archivo:
            return HttpResponse('Archivo no encontrado', status=404)
        
        # Importar las clases necesarias para la respuesta
        from django.http import HttpResponse, Http404
        from django.utils.encoding import smart_str
        import os
        import mimetypes
        
        # Obtener la ruta del archivo
        file_path = archivo.archivo.path
        
        if not os.path.exists(file_path):
            raise Http404("El archivo no existe en el servidor")
        
        # Determinar el tipo MIME
        content_type, _ = mimetypes.guess_type(file_path)
        if content_type is None:
            content_type = 'application/octet-stream'
        
        # Servir el archivo inline siempre (para vista previa en nueva pestaña)
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=content_type)
        
        # Configurar headers para vista inline (no descarga)
        filename = smart_str(archivo.nombre_original)
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        response['Content-Length'] = os.path.getsize(file_path)
        
        return response
        
    except Exception as e:
        print(f"❌ Error en vista previa de archivo: {e}")
        return HttpResponse(f'Error al abrir archivo: {str(e)}', status=500)


# ==========================================
# APIs DE NOTIFICACIONES
# ==========================================

@login_required
def obtener_notificaciones_api(request):
    """
    API para obtener las notificaciones del usuario actual
    """
    try:
        from .models import Notificacion
        user = request.user
        
        # --- Lógica para verificar tareas a punto de vencer y vencidas ---
        try:
            from django.utils import timezone
            from django.db.models import Q
            from datetime import timedelta
            from .models import Tarea, TareaOportunidad
            now = timezone.now()
            umbral_por_vencer = now + timedelta(days=1)
            
            # Buscar tareas generales del usuario
            mis_tareas = Tarea.objects.filter(
                Q(asignado_a=user) | Q(participantes=user),
                estado__in=['pendiente', 'iniciada', 'en_progreso'],
                fecha_limite__isnull=False
            ).distinct()
            
            for t in mis_tareas:
                if t.fecha_limite < now:
                    # Vencida
                    if not Notificacion.objects.filter(usuario_destinatario=user, tipo='tarea_vencida', tarea_id=t.id).exists():
                        crear_notificacion(user, 'tarea_vencida', 'Tarea Vencida', f'La tarea "{t.titulo}" ha vencido.', tarea_id=t.id)
                elif t.fecha_limite <= umbral_por_vencer:
                    # Por vencer
                    if not Notificacion.objects.filter(usuario_destinatario=user, tipo='tarea_por_vencer', tarea_id=t.id).exists():
                        crear_notificacion(user, 'tarea_por_vencer', 'Tarea por Vencer', f'La tarea "{t.titulo}" vencerá pronto.', tarea_id=t.id)

            # Idem para tareas de oportunidad
            mis_tareas_opp = TareaOportunidad.objects.filter(
                responsable=user,
                estado__in=['pendiente', 'en_progreso'],
                fecha_limite__isnull=False
            )
            for t in mis_tareas_opp:
                if t.fecha_limite < now:
                    if not Notificacion.objects.filter(usuario_destinatario=user, tipo='actividad_vencida', tarea_opp=t).exists():
                        crear_notificacion(user, 'actividad_vencida', 'Actividad Vencida', f'La actividad "{t.titulo}" ha vencido.', tarea_opp=t, oportunidad=t.oportunidad)
                elif t.fecha_limite <= umbral_por_vencer:
                    if not Notificacion.objects.filter(usuario_destinatario=user, tipo='actividad_por_vencer', tarea_opp=t).exists():
                        crear_notificacion(user, 'actividad_por_vencer', 'Actividad por Vencer', f'La actividad "{t.titulo}" vencerá pronto.', tarea_opp=t, oportunidad=t.oportunidad)
                        
        except Exception as ex_exp:
            print(f"Error verificando vencimientos: {ex_exp}")
        
        # Obtener notificaciones del usuario (últimas 50)
        notificaciones = Notificacion.objects.filter(
            usuario_destinatario=user
        ).select_related(
            'usuario_remitente', 'oportunidad', 'comentario', 'tarea_opp'
        ).order_by('-fecha_creacion')[:50]
        
        # Contar notificaciones no leídas
        unread_count = Notificacion.objects.filter(
            usuario_destinatario=user,
            leida=False
        ).count()
        
        # Serializar notificaciones
        notifications_data = []
        for notif in notificaciones:
            # Determinar URL basada en el tipo de notificación
            url = ''
            if notif.tarea_id:
                url = f'/app/tareas-proyectos/?task_id={notif.tarea_id}'
            elif notif.tarea_opp_id:
                if notif.oportunidad_id:
                    url = f'/app/cotizaciones/oportunidad/{notif.oportunidad_id}/?tab=actividades'
                else:
                    url = '/app/todos/'
            elif notif.oportunidad:
                url = f'/app/cotizaciones/oportunidad/{notif.oportunidad.id}/'
            elif notif.proyecto_id:
                url = f'/app/proyecto/{notif.proyecto_id}/'
            elif notif.tipo in ['muro_post', 'muro_mencion', 'mencion', 'respuesta']:
                url = '/app/home/?open_muro=1'
            elif notif.tipo == 'rendimiento_bajo' and notif.usuario_remitente:
                url = f'/app/perfil-usuario/{notif.usuario_remitente.id}/'
            notifications_data.append({
                'id': notif.id,
                'titulo': notif.titulo,
                'mensaje': notif.mensaje,
                'tipo': notif.tipo,
                'leida': notif.leida,
                'fecha': notif.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                'remitente': notif.usuario_remitente.get_full_name() if notif.usuario_remitente else 'Sistema',
                'url': url,
                'remitente_id': notif.usuario_remitente.id if notif.usuario_remitente else None,
                'tarea_id': notif.tarea_id,
                'proyecto_id': notif.proyecto_id,
                'tarea_opp_id': notif.tarea_opp_id,
                'oportunidad_id': notif.oportunidad_id,
                'solicitud_perfil_data': {
                    'id': notif.solicitud_perfil.id,
                    'first_name': notif.solicitud_perfil.first_name,
                    'last_name': notif.solicitud_perfil.last_name,
                    'email': notif.solicitud_perfil.email,
                    'language': notif.solicitud_perfil.language,
                    'avatar_url': notif.solicitud_perfil.avatar.url if notif.solicitud_perfil.avatar else None,
                    'old_first_name': notif.solicitud_perfil.old_first_name,
                    'old_last_name': notif.solicitud_perfil.old_last_name,
                    'old_email': notif.solicitud_perfil.old_email,
                    'old_language': notif.solicitud_perfil.old_language,
                } if notif.solicitud_perfil else None,
            })
        
        return JsonResponse({
            'success': True,
            'notifications': notifications_data,
            'unread_count': unread_count
        })
        
    except Exception as e:
        print(f"❌ Error obteniendo notificaciones: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'notifications': [],
            'unread_count': 0
        })


@login_required
@require_http_methods(["POST"])
def marcar_notificacion_leida_api(request, notificacion_id):
    """
    API para marcar una notificación específica como leída
    """
    try:
        from .models import Notificacion
        notificacion = get_object_or_404(
            Notificacion, 
            id=notificacion_id, 
            usuario_destinatario=request.user
        )
        
        notificacion.marcar_como_leida()
        
        return JsonResponse({
            'success': True,
            'message': 'Notificación marcada como leída'
        })
        
    except Exception as e:
        print(f"❌ Error marcando notificación como leída: {e}")
        return JsonResponse({
            'success': False,
            'error': 'Error al marcar notificación como leída'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def marcar_todas_notificaciones_leidas_api(request):
    """
    API para marcar todas las notificaciones del usuario como leídas
    """
    try:
        from .models import Notificacion
        from django.utils import timezone
        import json
        
        filters = {'usuario_destinatario': request.user, 'leida': False}
        
        # Intentar leer filtros del body
        try:
            data = json.loads(request.body)
            if 'tarea_id' in data and data['tarea_id']:
                filters['tarea_id'] = data['tarea_id']
            if 'oportunidad_id' in data and data['oportunidad_id']:
                filters['oportunidad_id'] = data['oportunidad_id']
            if 'tarea_opp_id' in data and data['tarea_opp_id']:
                filters['tarea_opp_id'] = data['tarea_opp_id']
        except:
            pass

        # Marcar notificaciones filtradas
        notificaciones_no_leidas = Notificacion.objects.filter(**filters)
        
        count = notificaciones_no_leidas.update(
            leida=True,
            fecha_lectura=timezone.now()
        )
        
        return JsonResponse({
            'success': True,
            'message': f'{count} notificaciones marcadas como leídas',
            'count_marked': count
        })
        
    except Exception as e:
        print(f"❌ Error marcando todas las notificaciones como leídas: {e}")
        return JsonResponse({
            'success': False,
            'error': 'Error al marcar notificaciones como leídas'
        }, status=500)


def crear_notificacion(usuario_destinatario, tipo, titulo, mensaje, oportunidad=None, comentario=None, usuario_remitente=None, proyecto_id=None, proyecto_nombre=None, tarea_opp=None, solicitud_perfil=None, tarea_id=None, tarea_titulo=None):
    """
    Función auxiliar para crear notificaciones
    """
    try:
        # No crear notificación si el remitente y destinatario son el mismo
        if usuario_remitente and usuario_destinatario == usuario_remitente:
            return None

        notificacion = Notificacion.objects.create(
            usuario_destinatario=usuario_destinatario,
            usuario_remitente=usuario_remitente,
            tipo=tipo,
            titulo=titulo,
            mensaje=mensaje,
            oportunidad=oportunidad,
            comentario=comentario,
            proyecto_id=proyecto_id,
            proyecto_nombre=proyecto_nombre,
            tarea_opp=tarea_opp,
            solicitud_perfil=solicitud_perfil,
            tarea_id=tarea_id,
            tarea_titulo=tarea_titulo,
        )
        
        print(f"✅ Notificación creada: {titulo} para {usuario_destinatario.username}")
        return notificacion
        
    except Exception as e:
        print(f"❌ Error creando notificación: {e}")
        return None

def notificar_miembro_agregado_proyecto(usuario_agregado, proyecto_nombre, proyecto_id, usuario_que_agrega):
    """
    Función específica para notificar cuando se agrega un usuario a un proyecto
    """
    titulo = f"Te han agregado al proyecto: {proyecto_nombre}"
    mensaje = f"{usuario_que_agrega.get_full_name() or usuario_que_agrega.username} te ha agregado como miembro del proyecto '{proyecto_nombre}'. Ahora puedes colaborar en este proyecto."
    
    return crear_notificacion(
        usuario_destinatario=usuario_agregado,
        tipo='proyecto_agregado',
        titulo=titulo,
        mensaje=mensaje,
        usuario_remitente=usuario_que_agrega,
        proyecto_id=proyecto_id,
        proyecto_nombre=proyecto_nombre
    )


def detectar_menciones_en_comentario(contenido, usuario_remitente, oportunidad, comentario=None):
    """
    Detecta menciones @usuario en un comentario y crea notificaciones
    """
    import re
    
    # Buscar menciones en el formato @usuario
    menciones = re.findall(r'@(\w+)', contenido)
    
    for username in menciones:
        try:
            usuario_mencionado = User.objects.get(username=username)
            
            # Crear notificación de mención
            crear_notificacion(
                usuario_destinatario=usuario_mencionado,
                tipo='mencion',
                titulo='Te han mencionado en un comentario',
                mensaje=f'{usuario_remitente.get_full_name() or usuario_remitente.username} te mencionó en la oportunidad "{oportunidad.oportunidad}": {contenido[:100]}...',
                oportunidad=oportunidad,
                comentario=comentario,
                usuario_remitente=usuario_remitente
            )
            
        except User.DoesNotExist:
            # Usuario mencionado no existe, ignorar
            pass




@login_required
@require_http_methods(["POST"])
def actualizar_avatar(request):
    """
    API para actualizar el avatar del usuario
    """
    try:
        # Obtener o crear el perfil del usuario
        user_profile, created = UserProfile.objects.get_or_create(user=request.user)
        
        # Verificar si se envió un archivo de avatar
        if 'avatar' in request.FILES:
            avatar_file = request.FILES['avatar']
            
            # Validar tipo de archivo
            allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp']
            if avatar_file.content_type not in allowed_types:
                return JsonResponse({
                    'success': False, 
                    'error': 'Tipo de archivo no permitido. Solo se permiten imágenes (JPG, PNG, GIF, WebP)'
                })
            
            # Validar tamaño de archivo (máximo 5MB)
            if avatar_file.size > 5 * 1024 * 1024:
                return JsonResponse({
                    'success': False, 
                    'error': 'El archivo es demasiado grande. Máximo 5MB permitido'
                })
            
            # Eliminar avatar anterior si existe
            if user_profile.avatar:
                try:
                    user_profile.avatar.delete(save=False)
                except:
                    pass  # Ignorar errores al eliminar archivo anterior
            
            # Guardar nuevo avatar
            user_profile.avatar = avatar_file
            user_profile.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Avatar actualizado correctamente',
                'avatar_url': user_profile.get_avatar_url()
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'No se envió ningún archivo de avatar'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar avatar: {str(e)}'
        })

@login_required
@user_passes_test(lambda u: u.is_superuser)
def bitrix_sync_admin(request):
    """
    Vista de administración para sincronización con Bitrix24
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'sync_deletions':
            # Sincronizar eliminaciones
            from .bitrix_integration import get_all_bitrix_deals
            
            try:
                # Obtener todas las oportunidades locales con bitrix_deal_id
                local_opportunities = TodoItem.objects.filter(bitrix_deal_id__isnull=False)
                
                # Obtener todos los deals de Bitrix24
                bitrix_deals = get_all_bitrix_deals()
                if not bitrix_deals:
                    messages.error(request, 'No se pudieron obtener deals de Bitrix24')
                    return redirect('bitrix-sync-admin')
                
                # Crear set de IDs que existen en Bitrix24
                bitrix_deal_ids = {str(deal['ID']) for deal in bitrix_deals}
                
                # Encontrar oportunidades locales que ya no existen en Bitrix24
                opportunities_to_delete = []
                for opportunity in local_opportunities:
                    if str(opportunity.bitrix_deal_id) not in bitrix_deal_ids:
                        opportunities_to_delete.append(opportunity)
                
                # Eliminar las oportunidades
                deleted_count = 0
                for opportunity in opportunities_to_delete:
                    try:
                        opportunity.delete()
                        deleted_count += 1
                    except Exception as e:
                        print(f"Error eliminando oportunidad {opportunity.id}: {e}")
                
                if deleted_count > 0:
                    messages.success(request, f'Se eliminaron {deleted_count} oportunidades que ya no existen en Bitrix24')
                else:
                    messages.info(request, 'No hay oportunidades para eliminar. Todo está sincronizado.')
                    
            except Exception as e:
                messages.error(request, f'Error durante la sincronización: {str(e)}')
        
        elif action == 'import_all':
            # Importar todas las oportunidades
            from django.core.management import call_command
            from io import StringIO
            
            try:
                out = StringIO()
                call_command('import_bitrix_opportunities', stdout=out)
                output = out.getvalue()
                messages.success(request, f'Importación completada: {output}')
            except Exception as e:
                messages.error(request, f'Error durante la importación: {str(e)}')
        
        elif action == 'sync_lost_opportunities':
            # Ejecutar sincronización de oportunidades perdidas
            try:
                from django.core.management import call_command
                from io import StringIO
                
                # Capturar la salida del comando
                out = StringIO()
                call_command('sync_lost_opportunities', stdout=out)
                output = out.getvalue()
                
                # Extraer estadísticas del output
                lines = output.split('\n')
                stats_found = False
                for line in lines:
                    if 'Oportunidades perdidas encontradas:' in line:
                        found_count = line.split(':')[-1].strip()
                        messages.success(
                            request, 
                            f'Sincronización completada. Se encontraron {found_count} oportunidades perdidas que fueron actualizadas.'
                        )
                        stats_found = True
                        break
                
                if not stats_found:
                    messages.success(request, 'Sincronización de oportunidades perdidas completada')
                    
            except Exception as e:
                messages.error(request, f'Error durante la sincronización: {str(e)}')
        
        elif action == 'full_bitrix_sync':
            # Ejecutar sincronización completa bidireccional
            try:
                from django.core.management import call_command
                from io import StringIO
                
                # Capturar la salida del comando
                out = StringIO()
                call_command('full_bitrix_sync', stdout=out)
                output = out.getvalue()
                
                # Extraer estadísticas del output
                lines = output.split('\n')
                for line in lines:
                    if 'Nuevas importadas:' in line:
                        new_count = line.split(':')[-1].strip()
                    elif 'Actualizadas:' in line:
                        updated_count = line.split(':')[-1].strip()
                    elif 'Perdidas detectadas:' in line:
                        lost_count = line.split(':')[-1].strip()
                
                messages.success(
                    request, 
                    f'Sincronización completa terminada. Nuevas: {new_count if "new_count" in locals() else "0"}, '
                    f'Actualizadas: {updated_count if "updated_count" in locals() else "0"}, '
                    f'Perdidas: {lost_count if "lost_count" in locals() else "0"}'
                )
                    
            except Exception as e:
                messages.error(request, f'Error durante la sincronización completa: {str(e)}')
        
        return redirect('bitrix-sync-admin')
    
    # GET request - mostrar estadísticas
    context = {
        'total_opportunities': TodoItem.objects.count(),
        'active_opportunities': TodoItem.objects.count() - len([item for item in TodoItem.objects.all() if is_lost_opportunity(item.bitrix_stage_id)]),
        'lost_opportunities': len([item for item in TodoItem.objects.all() if is_lost_opportunity(item.bitrix_stage_id)]),
        'bitrix_opportunities': TodoItem.objects.filter(bitrix_deal_id__isnull=False).count(),
        'local_opportunities': TodoItem.objects.filter(bitrix_deal_id__isnull=True).count(),
    }
    
    return render(request, 'bitrix_sync_admin.html', context)

def handle_lost_opportunity(opportunity, action='log_only'):
    """
    Maneja oportunidades marcadas como "Cerrado Perdido" en Bitrix24
    
    Actions:
    - 'log_only': Solo registra en logs (default)
    - 'delete': Elimina la oportunidad
    - 'mark_inactive': Marca como inactiva (requiere campo adicional)
    """
    opportunity_name = opportunity.oportunidad if hasattr(opportunity, 'oportunidad') else 'Oportunidad'
    
    if action == 'delete':
        opportunity.delete()
        print(f"BITRIX WEBHOOK: 🗑️ Oportunidad '{opportunity_name}' eliminada por estar marcada como PERDIDA", flush=True)
        return {'deleted': True}
    elif action == 'mark_inactive':
        # Si implementas un campo 'activa' en el modelo
        # opportunity.activa = False
        # opportunity.save()
        print(f"BITRIX WEBHOOK: 📴 Oportunidad '{opportunity_name}' marcada como inactiva por estar PERDIDA", flush=True)
        return {'marked_inactive': True}
    else:  # log_only
        print(f"BITRIX WEBHOOK: ⚠️ Oportunidad '{opportunity_name}' está marcada como PERDIDA en Bitrix24 (solo registro)", flush=True)
        return {'logged': True}

@login_required
@user_passes_test(lambda u: u.is_superuser)
@login_required  
def bitrix_lost_opportunities(request):
    """
    Vista para mostrar y gestionar oportunidades marcadas como "Cerrado Perdido"
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        opportunity_id = request.POST.get('opportunity_id')
        
        if opportunity_id and action in ['delete', 'reactivate']:
            try:
                opportunity = TodoItem.objects.get(id=opportunity_id)
                
                if action == 'delete':
                    opportunity_name = opportunity.oportunidad
                    opportunity.delete()
                    messages.success(request, f'Oportunidad "{opportunity_name}" eliminada exitosamente')
                elif action == 'reactivate':
                    # Cambiar el estado a uno activo (ej: Cotizando)
                    opportunity.bitrix_stage_id = 'UC_YUQKW6'  # Cotizando
                    opportunity.save()
                    messages.success(request, f'Oportunidad "{opportunity.oportunidad}" reactivada')
                    
            except TodoItem.DoesNotExist:
                messages.error(request, 'Oportunidad no encontrada')
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
        
        return redirect('bitrix-lost-opportunities')
    
    # Obtener oportunidades perdidas usando detección precisa
    all_opportunities = TodoItem.objects.select_related('cliente', 'usuario').all()
    lost_ids = [item.id for item in all_opportunities if is_lost_opportunity(item.bitrix_stage_id)]
    lost_opportunities = TodoItem.objects.filter(id__in=lost_ids).select_related('cliente', 'usuario').order_by('-fecha_actualizacion')
    
    context = {
        'lost_opportunities': lost_opportunities,
        'total_lost': lost_opportunities.count(),
    }
    
    return render(request, 'bitrix_lost_opportunities.html', context)


@login_required
def api_tarea_detalle(request, tarea_id):
    """
    API para obtener detalles de una tarea específica
    """
    if request.method == 'GET':
        try:
            # Buscar la tarea en la base de datos
            tarea = Tarea.objects.get(id=tarea_id)
            
            # Verificar permisos - el usuario debe ser participante, creador o asignado
            user_can_view = (
                tarea.creado_por == request.user or
                tarea.asignado_a == request.user or
                tarea.participantes.filter(id=request.user.id).exists() or
                request.user.is_superuser
            )
            
            if not user_can_view:
                return JsonResponse({'error': 'Sin permisos para ver esta tarea'}, status=403)
            
            # Función auxiliar para obtener datos del usuario con avatar
            def get_user_data(user):
                if not user:
                    return None
                
                user_data = {
                    'id': user.id,
                    'nombre': user.get_full_name() or user.username,
                    'username': user.username,
                    'avatar_url': None
                }
                
                # Intentar obtener avatar del UserProfile
                try:
                    if hasattr(user, 'userprofile') and user.userprofile.avatar:
                        user_data['avatar_url'] = user.userprofile.avatar.url
                except:
                    pass
                
                return user_data
            
            # Formatear tiempo trabajado
            tiempo_total_str = "00:00:00"
            tiempo_total_segundos = 0
            if hasattr(tarea, 'tiempo_trabajado') and tarea.tiempo_trabajado:
                tiempo_total_segundos = int(tarea.tiempo_trabajado.total_seconds())
                hours = tiempo_total_segundos // 3600
                minutes = (tiempo_total_segundos % 3600) // 60
                seconds = tiempo_total_segundos % 60
                tiempo_total_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"

            # Preparar datos de la tarea
            tarea_data = {
                'id': tarea.id,
                'titulo': tarea.titulo,
                'descripcion': tarea.descripcion,
                'descripcion_html': tarea.get_descripcion_html(),
                'estado': tarea.estado,
                'prioridad': tarea.prioridad,
                'proyecto_nombre': tarea.proyecto.nombre if tarea.proyecto else 'Sin proyecto',
                'proyecto_id': tarea.proyecto.id if tarea.proyecto else None,
                'creado_por': tarea.creado_por.get_full_name() or tarea.creado_por.username,
                'creado_por_data': get_user_data(tarea.creado_por),
                'responsable': tarea.asignado_a.get_full_name() or tarea.asignado_a.username if tarea.asignado_a else None,
                'responsable_data': get_user_data(tarea.asignado_a) if tarea.asignado_a else None,
                'fecha_limite': tarea.fecha_limite.isoformat() if tarea.fecha_limite else None,
                'fecha_creacion': tarea.fecha_creacion.isoformat(),
                'fecha_completada': tarea.fecha_completada.isoformat() if tarea.fecha_completada else None,
                'participantes': [get_user_data(p) for p in tarea.participantes.all()] if hasattr(tarea, 'participantes') else [],
                'observadores': [get_user_data(o) for o in tarea.observadores.all()] if hasattr(tarea, 'observadores') else [],
                # Datos del cronómetro
                'trabajando_actualmente': getattr(tarea, 'trabajando_actualmente', False),
                'pausado': getattr(tarea, 'pausado', False),
                'tiempo_trabajado': tiempo_total_str,
                'tiempo_total_trabajado': tiempo_total_segundos,
                'fecha_inicio_sesion': tarea.fecha_inicio_sesion.isoformat() if hasattr(tarea, 'fecha_inicio_sesion') and tarea.fecha_inicio_sesion else None,
                # Oportunidad y cliente
                'oportunidad_id': tarea.oportunidad.id if tarea.oportunidad else None,
                'oportunidad_nombre': tarea.oportunidad.oportunidad if tarea.oportunidad else None,
                # Cliente: primero el directo, luego el de la oportunidad
                'cliente_id': (tarea.cliente.id if tarea.cliente else (tarea.oportunidad.cliente.id if tarea.oportunidad and tarea.oportunidad.cliente else None)),
                'cliente_nombre': (tarea.cliente.nombre_empresa if tarea.cliente else (tarea.oportunidad.cliente.nombre_empresa if tarea.oportunidad and tarea.oportunidad.cliente else None)),
            }
            
            return JsonResponse(tarea_data)
            
        except Tarea.DoesNotExist:
            return JsonResponse({'error': 'Tarea no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    elif request.method == 'PUT':
        try:
            import json
            from django.db import transaction
            data = json.loads(request.body)
            
            print(f"🔍 PUT request data: {data}")
            print(f"🔍 User making request: {request.user.username}")
            
            # Usar transacción atómica para asegurar consistencia
            with transaction.atomic():
                # Buscar la tarea
                tarea = Tarea.objects.get(id=tarea_id)
                print(f"🔍 Tarea encontrada: {tarea.titulo} (ID: {tarea.id})")
                
                # Verificar permisos - solo el creador o admin puede modificar participantes
                user_can_edit = (
                    tarea.creado_por == request.user or
                    request.user.is_superuser
                )
                
                print(f"🔍 Permisos - Creador: {tarea.creado_por.username}, Current: {request.user.username}, Can edit: {user_can_edit}")
                
                if not user_can_edit:
                    return JsonResponse({'error': 'Sin permisos para modificar esta tarea'}, status=403)
                
                # Obtener datos de la petición
                user_id = data.get('user_id')
                action = data.get('action')  # 'add' o 'remove'
                tipo = data.get('tipo')      # 'participantes' o 'observadores'
                
                print(f"🔍 Datos: user_id={user_id}, action={action}, tipo={tipo}")
                
                if not all([user_id, action, tipo]):
                    return JsonResponse({'error': 'Datos faltantes: user_id, action y tipo son requeridos'}, status=400)
                
                # Obtener usuario
                try:
                    usuario = User.objects.get(id=user_id)
                    print(f"🔍 Usuario encontrado: {usuario.username} ({usuario.get_full_name()})")
                except User.DoesNotExist:
                    return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
                
                # Verificar estado actual antes del cambio
                if tipo == 'participantes':
                    current_participantes = list(tarea.participantes.all())
                    print(f"🔍 Participantes actuales ANTES: {[p.username for p in current_participantes]}")
                elif tipo == 'observadores':
                    current_observadores = list(tarea.observadores.all())
                    print(f"🔍 Observadores actuales ANTES: {[o.username for o in current_observadores]}")
                
                # Aplicar cambios según el tipo
                if tipo == 'participantes':
                    if action == 'add':
                        tarea.participantes.add(usuario)
                        print(f"✅ AGREGADO como participante: {usuario.username}")
                        mensaje = f"{usuario.get_full_name() or usuario.username} ha sido agregado como participante a la tarea '{tarea.titulo}'"
                    elif action == 'remove':
                        tarea.participantes.remove(usuario)
                        print(f"❌ REMOVIDO como participante: {usuario.username}")
                        mensaje = f"{usuario.get_full_name() or usuario.username} ha sido removido como participante de la tarea '{tarea.titulo}'"
                elif tipo == 'observadores':
                    if action == 'add':
                        tarea.observadores.add(usuario)
                        print(f"✅ AGREGADO como observador: {usuario.username}")
                        mensaje = f"{usuario.get_full_name() or usuario.username} ha sido agregado como observador a la tarea '{tarea.titulo}'"
                    elif action == 'remove':
                        tarea.observadores.remove(usuario)
                        print(f"❌ REMOVIDO como observador: {usuario.username}")
                        mensaje = f"{usuario.get_full_name() or usuario.username} ha sido removido como observador de la tarea '{tarea.titulo}'"
                else:
                    return JsonResponse({'error': 'Tipo inválido. Use "participantes" o "observadores"'}, status=400)
                
                # Guardar cambios
                tarea.save()
                print(f"💾 Tarea guardada")
                
                # Verificar estado después del cambio
                if tipo == 'participantes':
                    new_participantes = list(tarea.participantes.all())
                    print(f"🔍 Participantes actuales DESPUÉS: {[p.username for p in new_participantes]}")
                elif tipo == 'observadores':
                    new_observadores = list(tarea.observadores.all())
                    print(f"🔍 Observadores actuales DESPUÉS: {[o.username for o in new_observadores]}")
                
                # Enviar notificación al usuario agregado
                if action == 'add':
                    try:
                        from .models import Notificacion
                        
                        # Determinar el tipo de notificación
                        tipo_notificacion = 'tarea_participante' if tipo == 'participantes' else 'tarea_observador'
                        
                        # Crear la notificación
                        notificacion = Notificacion.objects.create(
                            usuario_destinatario=usuario,
                            usuario_remitente=request.user,
                            tipo=tipo_notificacion,
                            titulo=f"Agregado a tarea: {tarea.titulo}",
                            mensaje=mensaje,
                            tarea_id=tarea.id,
                            tarea_titulo=tarea.titulo,
                            proyecto_id=tarea.proyecto.id if tarea.proyecto else None,
                            proyecto_nombre=tarea.proyecto.nombre if tarea.proyecto else None
                        )
                        
                        print(f"🔔 Notificación creada para {usuario.username}: {notificacion.titulo}")
                        
                    except Exception as e:
                        print(f"❌ Error creando notificación: {e}")
                        # No fallar si hay error en notificación, solo loggearlo
                
                return JsonResponse({
                    'success': True,
                    'message': mensaje,
                    'user_data': {
                        'id': usuario.id,
                        'nombre': usuario.get_full_name() or usuario.username,
                        'username': usuario.username,
                        'avatar_url': usuario.userprofile.get_avatar_url() if hasattr(usuario, 'userprofile') else None
                    }
                })
            
        except Tarea.DoesNotExist:
            return JsonResponse({'error': 'Tarea no encontrada'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


# Funciones de utilidad para notificaciones
# Función duplicada eliminada - usar la función de la línea 9891

def notificar_nueva_tarea(tarea, creado_por):
    """
    Envía notificación cuando se crea una nueva tarea
    """
    try:
        # Notificar al responsable de la tarea
        if tarea.asignado_a and tarea.asignado_a != creado_por:
            crear_notificacion(
                usuario_destinatario=tarea.asignado_a,
                tipo='tarea_nueva',
                titulo=f'Nueva tarea asignada: {tarea.titulo}',
                mensaje=f'Se te ha asignado la tarea "{tarea.titulo}" en el proyecto {tarea.proyecto.nombre if tarea.proyecto else "Sin proyecto"}',
                usuario_remitente=creado_por
            )
        
        # Notificar a todos los participantes
        for participante in tarea.participantes.all():
            if participante != creado_por and participante != tarea.asignado_a:
                crear_notificacion(
                    usuario_destinatario=participante,
                    tipo='tarea_nueva',
                    titulo=f'Nueva tarea: {tarea.titulo}',
                    mensaje=f'Has sido agregado como participante en la tarea "{tarea.titulo}"',
                    usuario_remitente=creado_por
                )
        
        # Notificar a todos los observadores
        for observador in tarea.observadores.all():
            if observador != creado_por and observador != tarea.asignado_a:
                crear_notificacion(
                    usuario_destinatario=observador,
                    tipo='tarea_nueva',
                    titulo=f'Nueva tarea: {tarea.titulo}',
                    mensaje=f'Has sido agregado como observador en la tarea "{tarea.titulo}"',
                    usuario_remitente=creado_por
                )
                
    except Exception as e:
        print(f"Error enviando notificaciones de nueva tarea: {e}")


@login_required 
def api_notificaciones(request):
    """
    API para obtener notificaciones del usuario actual
    """
    if request.method == 'GET':
        try:
            # Obtener notificaciones no leídas del usuario
            notificaciones = request.user.notificaciones.filter(leida=False).order_by('-fecha_creacion')[:10]
            
            notificaciones_data = []
            for notif in notificaciones:
                notificaciones_data.append({
                    'id': notif.id,
                    'tipo': notif.tipo,
                    'titulo': notif.titulo,
                    'mensaje': notif.mensaje,
                    'url': notif.url,
                    'fecha_creacion': notif.fecha_creacion.isoformat(),
                    'creado_por': notif.creado_por.get_full_name() or notif.creado_por.username if notif.creado_por else None
                })
            
            return JsonResponse({
                'success': True,
                'notificaciones': notificaciones_data,
                'total_no_leidas': request.user.notificaciones.filter(leida=False).count()
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    elif request.method == 'PUT':
        try:
            import json
            data = json.loads(request.body)
            
            # Marcar notificación como leída
            notif_id = data.get('id')
            if notif_id:
                from .models import Notificacion
                notificacion = Notificacion.objects.get(id=notif_id, usuario=request.user)
                notificacion.leida = True
                notificacion.save()
                
                return JsonResponse({'success': True, 'message': 'Notificación marcada como leída'})
            else:
                return JsonResponse({'error': 'ID de notificación requerido'}, status=400)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_toggle_task_timer(request, tarea_id):
    """
    API para iniciar/pausar el cronómetro de una tarea
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Importaciones necesarias
        from datetime import datetime, timezone, timedelta
        from .models import Tarea
        import json
        
        # Obtener la tarea
        tarea = get_object_or_404(Tarea, id=tarea_id)
        
        # Verificar permisos - solo el asignado puede iniciar/pausar
        if request.user != tarea.asignado_a and not request.user.is_superuser:
            return JsonResponse({'error': 'Solo el responsable puede iniciar/pausar esta tarea'}, status=403)
        
        # Obtener acción del request
        data = json.loads(request.body) if request.body else {}
        action = data.get('action', 'toggle')  # 'start', 'pause', 'toggle'
        
        ahora = datetime.now(timezone.utc)
        
        if not tarea.trabajando_actualmente:
            # INICIAR TAREA
            tarea.trabajando_actualmente = True
            tarea.fecha_inicio_sesion = ahora
            tarea.estado = 'en_progreso'
            tarea.pausado = False
            
            mensaje = f"Tarea iniciada. Cronómetro en marcha."
            
        else:
            # PAUSAR TAREA
            if tarea.fecha_inicio_sesion:
                # Calcular tiempo trabajado en esta sesión
                tiempo_sesion = ahora - tarea.fecha_inicio_sesion
                # Sumar al tiempo total trabajado
                if tarea.tiempo_trabajado:
                    tarea.tiempo_trabajado += tiempo_sesion
                else:
                    tarea.tiempo_trabajado = tiempo_sesion
            
            tarea.trabajando_actualmente = False
            tarea.fecha_inicio_sesion = None
            tarea.estado = 'iniciada'  # Usamos 'iniciada' para representar pausado
            tarea.pausado = True
            
            mensaje = f"Tarea pausada. Tiempo registrado."
        
        tarea.save()
        
        # Formatear tiempo trabajado para respuesta
        tiempo_total_str = "00:00:00"
        if tarea.tiempo_trabajado:
            total_seconds = int(tarea.tiempo_trabajado.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            tiempo_total_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'tarea': {
                'id': tarea.id,
                'trabajando_actualmente': tarea.trabajando_actualmente,
                'estado': tarea.estado,
                'tiempo_trabajado': tiempo_total_str,
                'fecha_inicio_sesion': tarea.fecha_inicio_sesion.isoformat() if tarea.fecha_inicio_sesion else None
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_completar_tarea(request, tarea_id):
    """
    API para marcar una tarea como completada
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Importaciones necesarias
        from datetime import datetime, timezone
        from .models import Tarea
        
        # Obtener la tarea
        tarea = get_object_or_404(Tarea, id=tarea_id)
        
        # Verificar permisos - solo el asignado, creador o superusuario pueden completar
        if (request.user != tarea.asignado_a and 
            request.user != tarea.creado_por and 
            not request.user.is_superuser):
            return JsonResponse({'error': 'Sin permisos para completar esta tarea'}, status=403)
        
        # Verificar que no esté ya completada
        if tarea.estado == 'completada':
            return JsonResponse({'error': 'La tarea ya está completada'}, status=400)
        
        ahora = datetime.now(timezone.utc)
        
        # Si está corriendo el cronómetro, detenerlo y guardar tiempo
        if getattr(tarea, 'trabajando_actualmente', False):
            if hasattr(tarea, 'fecha_inicio_sesion') and tarea.fecha_inicio_sesion:
                # Calcular tiempo trabajado en esta sesión
                tiempo_sesion = ahora - tarea.fecha_inicio_sesion
                # Sumar al tiempo total trabajado
                if hasattr(tarea, 'tiempo_trabajado') and tarea.tiempo_trabajado:
                    tarea.tiempo_trabajado += tiempo_sesion
                else:
                    tarea.tiempo_trabajado = tiempo_sesion
            
            # Detener cronómetro
            tarea.trabajando_actualmente = False
            tarea.fecha_inicio_sesion = None
            tarea.pausado = False
        
        # Marcar como completada
        tarea.estado = 'completada'
        tarea.fecha_completada = ahora
        tarea.save()

        # Notificar al creador si es distinto al que completó
        if tarea.creado_por and tarea.creado_por != request.user:
            completador = request.user.get_full_name() or request.user.username
            crear_notificacion(
                usuario_destinatario=tarea.creado_por,
                tipo='tarea_vencida',
                titulo=f'Tarea completada: {tarea.titulo}',
                mensaje=f'{completador} marcó como completada la tarea "{tarea.titulo}".',
                usuario_remitente=request.user,
                tarea_id=tarea.id,
                tarea_titulo=tarea.titulo,
            )

        # Formatear tiempo trabajado para respuesta
        tiempo_total_str = "00:00:00"
        if hasattr(tarea, 'tiempo_trabajado') and tarea.tiempo_trabajado:
            total_seconds = int(tarea.tiempo_trabajado.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            tiempo_total_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        
        return JsonResponse({
            'success': True,
            'message': 'Tarea completada exitosamente',
            'tarea': {
                'id': tarea.id,
                'estado': tarea.estado,
                'fecha_completada': tarea.fecha_completada.isoformat(),
                'tiempo_trabajado_total': tiempo_total_str,
                'trabajando_actualmente': False
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# =============================================
# API ENDPOINTS PARA SISTEMA DE CARPETAS Y ARCHIVOS
# =============================================

@csrf_exempt
@login_required
def api_carpetas_proyecto(request, proyecto_id):
    """API para listar carpetas de un proyecto y crear nuevas carpetas"""
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        
        if request.method == 'GET':
            parent_id = request.GET.get('parent')
            if parent_id:
                carpetas = CarpetaProyecto.objects.filter(
                    proyecto=proyecto,
                    carpeta_padre_id=parent_id
                ).order_by('nombre')
            else:
                # Carpetas raíz (sin padre)
                carpetas = CarpetaProyecto.objects.filter(
                    proyecto=proyecto,
                    carpeta_padre__isnull=True
                ).order_by('nombre')
            
            carpetas_data = []
            for carpeta in carpetas:
                carpetas_data.append({
                    'id': carpeta.id,
                    'nombre': carpeta.nombre,
                    'fecha_creacion': carpeta.fecha_creacion.isoformat(),
                    'fecha_modificacion': carpeta.fecha_modificacion.isoformat(),
                    'creado_por': carpeta.creado_por.get_full_name() or carpeta.creado_por.username,
                    'carpeta_padre_id': carpeta.carpeta_padre_id,
                    'tipo': 'carpeta'
                })
            
            # También incluir archivos en la carpeta actual
            if parent_id:
                archivos = ArchivoProyecto.objects.filter(
                    proyecto=proyecto,
                    carpeta_id=parent_id
                ).order_by('nombre_original')
            else:
                # Archivos en raíz (sin carpeta)
                archivos = ArchivoProyecto.objects.filter(
                    proyecto=proyecto,
                    carpeta__isnull=True
                ).order_by('nombre_original')
            
            archivos_data = []
            for archivo in archivos:
                archivos_data.append({
                    'id': archivo.id,
                    'nombre': archivo.nombre_original,
                    'tipo_archivo': archivo.tipo_archivo,
                    'tamaño': archivo.tamaño,
                    'extension': archivo.extension,
                    'fecha_subida': archivo.fecha_subida.isoformat(),
                    'subido_por': archivo.subido_por.get_full_name() or archivo.subido_por.username,
                    'es_publico': archivo.es_publico,
                    'descripcion': archivo.descripcion,
                    'url': f'/app/api/proyecto/{proyecto_id}/archivo/{archivo.id}/stream/',
                    'tipo': 'archivo'
                })
            
            # Obtener información de la carpeta actual si estamos en una subcarpeta
            carpeta_actual = None
            if parent_id:
                try:
                    carpeta_obj = CarpetaProyecto.objects.get(id=parent_id)
                    carpeta_actual = {
                        'id': carpeta_obj.id,
                        'nombre': carpeta_obj.nombre
                    }
                except CarpetaProyecto.DoesNotExist:
                    pass
            
            return JsonResponse({
                'success': True,
                'carpetas': carpetas_data,
                'archivos': archivos_data,
                'carpeta_actual': carpeta_actual
            })
        
        elif request.method == 'POST':
            data = json.loads(request.body)
            nombre = data.get('nombre')
            parent_id = data.get('parent_id')
            
            print(f"📂 Creando carpeta: {nombre} en proyecto {proyecto_id}, parent: {parent_id}")
            
            if not nombre:
                return JsonResponse({'error': 'Nombre de carpeta requerido'}, status=400)
            
            # Verificar que no exista una carpeta con el mismo nombre en el mismo nivel
            carpeta_padre = None
            if parent_id:
                carpeta_padre = get_object_or_404(CarpetaProyecto, id=parent_id)
            
            if CarpetaProyecto.objects.filter(
                proyecto=proyecto,
                carpeta_padre=carpeta_padre,
                nombre=nombre
            ).exists():
                return JsonResponse({'error': 'Ya existe una carpeta con ese nombre'}, status=400)
            
            carpeta = CarpetaProyecto.objects.create(
                nombre=nombre,
                proyecto=proyecto,
                carpeta_padre=carpeta_padre,
                creado_por=request.user
            )
            
            print(f"✅ Carpeta creada exitosamente: ID={carpeta.id}, Nombre={carpeta.nombre}")
            
            return JsonResponse({
                'success': True,
                'carpeta': {
                    'id': carpeta.id,
                    'nombre': carpeta.nombre,
                    'fecha_creacion': carpeta.fecha_creacion.isoformat(),
                    'creado_por': carpeta.creado_por.get_full_name() or carpeta.creado_por.username,
                    'carpeta_padre_id': carpeta.carpeta_padre_id,
                    'tipo': 'carpeta'
                }
            })
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@login_required
def api_carpeta_detalle(request, proyecto_id, carpeta_id):
    """API para obtener, actualizar y eliminar una carpeta específica"""
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        carpeta = get_object_or_404(CarpetaProyecto, id=carpeta_id, proyecto=proyecto)
        
        if request.method == 'GET':
            return JsonResponse({
                'success': True,
                'carpeta': {
                    'id': carpeta.id,
                    'nombre': carpeta.nombre,
                    'fecha_creacion': carpeta.fecha_creacion.isoformat(),
                    'fecha_modificacion': carpeta.fecha_modificacion.isoformat(),
                    'creado_por': carpeta.creado_por.get_full_name() or carpeta.creado_por.username,
                    'carpeta_padre_id': carpeta.carpeta_padre_id
                }
            })
        
        elif request.method == 'PUT':
            data = json.loads(request.body)
            nuevo_nombre = data.get('nombre')
            
            if not nuevo_nombre:
                return JsonResponse({'error': 'Nombre requerido'}, status=400)
            
            # Verificar que no exista otra carpeta con el mismo nombre
            if CarpetaProyecto.objects.filter(
                proyecto=proyecto,
                carpeta_padre=carpeta.carpeta_padre,
                nombre=nuevo_nombre
            ).exclude(id=carpeta_id).exists():
                return JsonResponse({'error': 'Ya existe una carpeta con ese nombre'}, status=400)
            
            carpeta.nombre = nuevo_nombre
            carpeta.save()
            
            return JsonResponse({
                'success': True,
                'carpeta': {
                    'id': carpeta.id,
                    'nombre': carpeta.nombre,
                    'fecha_modificacion': carpeta.fecha_modificacion.isoformat()
                }
            })
        
        elif request.method == 'DELETE':
            force_delete = request.GET.get('force', 'false').lower() == 'true'
            
            # Si no es forzado, verificar que la carpeta esté vacía
            if not force_delete and (carpeta.subcarpetas.exists() or carpeta.archivos.exists()):
                return JsonResponse({'error': 'No se puede eliminar una carpeta que contiene archivos o subcarpetas'}, status=400)
            
            # Si es eliminación forzada, eliminar recursivamente todo el contenido
            if force_delete:
                # Eliminar todos los archivos de la carpeta
                for archivo in carpeta.archivos.all():
                    # Eliminar archivo físico
                    if archivo.archivo:
                        try:
                            archivo.archivo.delete(save=False)
                        except:
                            pass  # Ignorar errores al eliminar archivo físico
                    archivo.delete()
                
                # Eliminar subcarpetas recursivamente
                def eliminar_carpeta_recursiva(carpeta_obj):
                    # Eliminar archivos de esta carpeta
                    for archivo in carpeta_obj.archivos.all():
                        if archivo.archivo:
                            try:
                                archivo.archivo.delete(save=False)
                            except:
                                pass
                        archivo.delete()
                    
                    # Eliminar subcarpetas recursivamente
                    for subcarpeta in carpeta_obj.subcarpetas.all():
                        eliminar_carpeta_recursiva(subcarpeta)
                        subcarpeta.delete()
                
                eliminar_carpeta_recursiva(carpeta)
            
            carpeta.delete()
            return JsonResponse({'success': True, 'message': 'Carpeta eliminada' + (' forzadamente' if force_delete else '')})
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@login_required
def api_archivos_proyecto(request, proyecto_id):
    """API para subir archivos a un proyecto - Usando lógica exitosa de oportunidades"""
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        print(f"🏗️ Procesando archivos para proyecto: {proyecto.nombre} (ID: {proyecto_id})")
        
        if request.method == 'POST':
            carpeta_id = request.POST.get('carpeta_id')
            descripcion = request.POST.get('descripcion', '')
            es_publico = request.POST.get('es_publico', 'true').lower() == 'true'
            archivo_file = request.FILES.get('archivo')
            
            if not archivo_file:
                return JsonResponse({'error': 'Archivo requerido'}, status=400)
            
            carpeta = None
            if carpeta_id:
                carpeta = get_object_or_404(CarpetaProyecto, id=carpeta_id, proyecto=proyecto)
            
            # Usar lógica simplificada similar a OportunidadArchivo que funciona
            print(f"📁 Subiendo archivo: {archivo_file.name} (Tamaño: {archivo_file.size} bytes)")
            
            # Detectar tipo de archivo simplificado
            content_type = archivo_file.content_type or ''
            extension = archivo_file.name.split('.')[-1].lower() if '.' in archivo_file.name else ''
            
            # Determinar tipo basado en extensión (similar a oportunidades)
            if extension in ['pdf']:
                tipo_archivo = 'pdf'
            elif extension in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'svg']:
                tipo_archivo = 'imagen'
            elif extension in ['doc', 'docx', 'txt', 'rtf', 'odt']:
                tipo_archivo = 'documento'
            elif extension in ['xls', 'xlsx', 'csv', 'ods']:
                tipo_archivo = 'hoja_calculo'
            elif extension in ['ppt', 'pptx', 'odp']:
                tipo_archivo = 'presentacion'
            elif extension in ['mp4', 'avi', 'mov', 'wmv']:
                tipo_archivo = 'video'
            elif extension in ['mp3', 'wav', 'aac', 'flac']:
                tipo_archivo = 'audio'
            elif extension in ['zip', 'rar', '7z', 'tar', 'gz']:
                tipo_archivo = 'archivo_comprimido'
            else:
                tipo_archivo = 'otro'
            
            print(f"📋 Tipo determinado: {tipo_archivo}")
            
            try:
                # Crear registro de archivo siguiendo el patrón exitoso de oportunidades
                archivo = ArchivoProyecto.objects.create(
                    nombre_original=archivo_file.name,
                    archivo=archivo_file,
                    tipo_archivo=tipo_archivo,
                    tamaño=archivo_file.size,
                    proyecto=proyecto,
                    carpeta=carpeta,
                    subido_por=request.user,
                    descripcion=descripcion or f"Archivo subido a {'carpeta' if carpeta else 'raíz'}",
                    es_publico=es_publico,
                    extension=extension,
                    mime_type=content_type
                )
                
                print(f"✅ Archivo guardado exitosamente: ID={archivo.id}, URL={archivo.archivo.url}")
                
            except Exception as e:
                print(f"❌ Error guardando archivo {archivo_file.name}: {e}")
                import traceback
                print(traceback.format_exc())
                return JsonResponse({'error': f'Error guardando archivo: {str(e)}'}, status=500)
            
            return JsonResponse({
                'success': True,
                'archivo': {
                    'id': archivo.id,
                    'nombre': archivo.nombre_original,
                    'tipo_archivo': archivo.tipo_archivo,
                    'tamaño': archivo.tamaño,
                    'extension': archivo.extension,
                    'fecha_subida': archivo.fecha_subida.isoformat(),
                    'subido_por': archivo.subido_por.get_full_name() or archivo.subido_por.username,
                    'es_publico': archivo.es_publico,
                    'descripcion': archivo.descripcion,
                    'url': f'/app/api/proyecto/{proyecto_id}/archivo/{archivo.id}/stream/',
                    'tipo': 'archivo'
                }
            })
            
    except Exception as e:
        print(f"❌ Error general en api_archivos_proyecto: {e}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': f'Error en el servidor: {str(e)}'}, status=500)


@csrf_exempt
@login_required
def api_archivo_detalle(request, proyecto_id, archivo_id):
    """API para obtener, actualizar y eliminar un archivo específico"""
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        archivo = get_object_or_404(ArchivoProyecto, id=archivo_id, proyecto=proyecto)
        
        if request.method == 'GET':
            return JsonResponse({
                'success': True,
                'archivo': {
                    'id': archivo.id,
                    'nombre': archivo.nombre_original,
                    'tipo_archivo': archivo.tipo_archivo,
                    'tamaño': archivo.tamaño,
                    'extension': archivo.extension,
                    'fecha_subida': archivo.fecha_subida.isoformat(),
                    'subido_por': archivo.subido_por.get_full_name() or archivo.subido_por.username,
                    'es_publico': archivo.es_publico,
                    'descripcion': archivo.descripcion,
                    'url': archivo.archivo.url,
                    'mime_type': archivo.mime_type
                }
            })
        
        elif request.method == 'PUT':
            data = json.loads(request.body)
            nuevo_nombre = data.get('nombre')
            nueva_descripcion = data.get('descripcion', archivo.descripcion)
            nuevo_es_publico = data.get('es_publico', archivo.es_publico)
            
            # Si se proporciona un nuevo nombre, actualizarlo
            if nuevo_nombre:
                archivo.nombre_original = nuevo_nombre
            
            archivo.descripcion = nueva_descripcion
            archivo.es_publico = nuevo_es_publico
            archivo.save()
            
            return JsonResponse({
                'success': True,
                'archivo': {
                    'id': archivo.id,
                    'nombre': archivo.nombre_original,
                    'descripcion': archivo.descripcion,
                    'es_publico': archivo.es_publico
                }
            })
        
        elif request.method == 'DELETE':
            # Eliminar archivo físico
            if archivo.archivo:
                try:
                    archivo.archivo.delete(save=False)
                except:
                    pass  # Ignorar errores al eliminar archivo físico
            
            archivo.delete()
            return JsonResponse({'success': True, 'message': 'Archivo eliminado'})
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_archivo_proyecto_stream(request, proyecto_id, archivo_id):
    """
    Descarga/previsualización en tiempo real para archivos de proyecto.
    """
    from django.http import FileResponse
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    archivo = get_object_or_404(ArchivoProyecto, id=archivo_id, proyecto=proyecto)

    if archivo.archivo:
        try:
            content_type, _ = mimetypes.guess_type(archivo.nombre_original)
            if not content_type:
                content_type = 'application/octet-stream'
            
            response = FileResponse(
                archivo.archivo.open('rb'),
                content_type=content_type
            )
            
            nombre_encoded = urllib.parse.quote(archivo.nombre_original)
            if request.GET.get('dl') == '1':
                response['Content-Disposition'] = f"attachment; filename*=UTF-8''{nombre_encoded}"
            else:
                response['Content-Disposition'] = f"inline; filename*=UTF-8''{nombre_encoded}"
                
            return response
        except Exception as e:
            print(f"Error streaming local project file: {e}")
            pass

    return JsonResponse({'error': 'Archivo no disponible'}, status=404)


# =============================================
# DRIVE DE OPORTUNIDADES
# =============================================

@csrf_exempt
@login_required
def api_drive_oportunidad(request, opp_id):
    """Lista carpetas + archivos en raíz (o en subcarpeta) de una oportunidad. POST crea carpeta."""
    opp = get_object_or_404(TodoItem, id=opp_id)

    if request.method == 'GET':
        parent_id = request.GET.get('parent')
        carpetas = CarpetaOportunidad.objects.filter(
            oportunidad=opp,
            carpeta_padre_id=parent_id  # None = raíz
        ).order_by('nombre')

        archivos = ArchivoOportunidad.objects.filter(
            oportunidad=opp,
            carpeta_id=parent_id
        ).order_by('nombre_original')

        carpeta_actual = None
        if parent_id:
            try:
                c = CarpetaOportunidad.objects.get(id=parent_id, oportunidad=opp)
                carpeta_actual = {'id': c.id, 'nombre': c.nombre, 'padre_id': c.carpeta_padre_id}
            except CarpetaOportunidad.DoesNotExist:
                pass

        carpetas_data = [{'id': c.id, 'nombre': c.nombre,
                          'carpeta_padre_id': c.carpeta_padre_id,
                          'fecha_creacion': c.fecha_creacion.isoformat(),
                          'tipo': 'carpeta'} for c in carpetas]
        archivos_data = [{'id': a.id, 'nombre': a.nombre_original,
                          'tipo_archivo': a.tipo_archivo,
                          'extension': a.extension,
                          'tamaño': a.tamaño,
                          'url': f'/app/api/oportunidad/{opp.id}/drive/archivo/{a.id}/stream/',
                          'es_bitrix': bool(not a.archivo and getattr(a, 'bitrix_file_id', None)),
                          'fecha_subida': a.fecha_subida.isoformat(),
                          'tipo': 'archivo'} for a in archivos]

        # Mezclar contenido del proyecto vinculado en la raíz
        if not parent_id:
            opp_proyecto = OportunidadProyecto.objects.filter(oportunidad=opp).first()
            if opp_proyecto:
                try:
                    proyecto_vinculado = Proyecto.objects.get(bitrix_group_id=int(opp_proyecto.bitrix_project_id))
                    for c in CarpetaProyecto.objects.filter(proyecto=proyecto_vinculado, carpeta_padre__isnull=True).order_by('nombre'):
                        carpetas_data.append({'id': c.id, 'nombre': c.nombre,
                                              'fecha_creacion': c.fecha_creacion.isoformat(),
                                              'tipo': 'carpeta_proyecto'})
                    for a in ArchivoProyecto.objects.filter(proyecto=proyecto_vinculado, carpeta__isnull=True).order_by('nombre_original'):
                        archivos_data.append({'id': a.id, 'nombre': a.nombre_original,
                                              'tipo_archivo': a.tipo_archivo,
                                              'extension': a.extension,
                                              'tamaño': a.tamaño,
                                              'url': a.bitrix_download_url or f'/app/api/proyecto/{proyecto_vinculado.id}/archivo/{a.id}/stream/',
                                              'tipo': 'archivo_proyecto'})
                except (Proyecto.DoesNotExist, ValueError):
                    pass

        return JsonResponse({
            'success': True,
            'carpeta_actual': carpeta_actual,
            'carpetas': carpetas_data,
            'archivos': archivos_data,
        })

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'JSON inválido'}, status=400)
        nombre = (data.get('nombre') or '').strip()
        parent_id = data.get('parent_id')
        if not nombre:
            return JsonResponse({'error': 'Nombre requerido'}, status=400)
        carpeta_padre = None
        if parent_id:
            carpeta_padre = get_object_or_404(CarpetaOportunidad, id=parent_id, oportunidad=opp)
        if CarpetaOportunidad.objects.filter(oportunidad=opp, carpeta_padre=carpeta_padre, nombre=nombre).exists():
            return JsonResponse({'error': 'Ya existe una carpeta con ese nombre'}, status=400)
        c = CarpetaOportunidad.objects.create(
            nombre=nombre, oportunidad=opp, carpeta_padre=carpeta_padre, creado_por=request.user
        )
        return JsonResponse({'success': True, 'carpeta': {
            'id': c.id, 'nombre': c.nombre,
            'carpeta_padre_id': c.carpeta_padre_id,
            'fecha_creacion': c.fecha_creacion.isoformat(), 'tipo': 'carpeta'
        }})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@csrf_exempt
@login_required
def api_drive_oportunidad_carpeta(request, opp_id, carpeta_id):
    """Renombra o elimina una carpeta del drive de una oportunidad."""
    opp = get_object_or_404(TodoItem, id=opp_id)
    carpeta = get_object_or_404(CarpetaOportunidad, id=carpeta_id, oportunidad=opp)

    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'JSON inválido'}, status=400)
        nuevo = (data.get('nombre') or '').strip()
        if not nuevo:
            return JsonResponse({'error': 'Nombre requerido'}, status=400)
        if CarpetaOportunidad.objects.filter(
            oportunidad=opp, carpeta_padre=carpeta.carpeta_padre, nombre=nuevo
        ).exclude(id=carpeta_id).exists():
            return JsonResponse({'error': 'Ya existe una carpeta con ese nombre'}, status=400)
        carpeta.nombre = nuevo
        carpeta.save()
        return JsonResponse({'success': True, 'carpeta': {'id': carpeta.id, 'nombre': carpeta.nombre}})

    elif request.method == 'DELETE':
        force = request.GET.get('force', 'false').lower() == 'true'
        if not force and not carpeta.puede_eliminar():
            return JsonResponse({'error': 'La carpeta no está vacía'}, status=400)

        def borrar_recursivo(c):
            for sub in c.subcarpetas.all():
                borrar_recursivo(sub)
            for a in c.archivos.all():
                if a.archivo:
                    try:
                        a.archivo.delete(save=False)
                    except Exception:
                        pass
                a.delete()
            c.delete()

        borrar_recursivo(carpeta)
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@csrf_exempt
@login_required
def api_drive_oportunidad_archivo(request, opp_id):
    """Sube un archivo al drive de una oportunidad."""
    opp = get_object_or_404(TodoItem, id=opp_id)

    if request.method == 'POST':
        archivo_file = request.FILES.get('archivo')
        if not archivo_file:
            return JsonResponse({'error': 'Archivo requerido'}, status=400)
        carpeta_id = request.POST.get('carpeta_id')
        carpeta = None
        if carpeta_id:
            carpeta = get_object_or_404(CarpetaOportunidad, id=carpeta_id, oportunidad=opp)

        ext = archivo_file.name.rsplit('.', 1)[-1].lower() if '.' in archivo_file.name else ''
        tipo_map = {
            'pdf': 'pdf',
            **{e: 'imagen' for e in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'svg', 'webp']},
            **{e: 'documento' for e in ['doc', 'docx', 'txt', 'rtf', 'odt']},
            **{e: 'hoja_calculo' for e in ['xls', 'xlsx', 'csv', 'ods']},
            **{e: 'presentacion' for e in ['ppt', 'pptx', 'odp']},
            **{e: 'video' for e in ['mp4', 'avi', 'mov', 'wmv', 'mkv']},
            **{e: 'audio' for e in ['mp3', 'wav', 'aac', 'flac']},
            **{e: 'archivo_comprimido' for e in ['zip', 'rar', '7z', 'tar', 'gz']},
        }
        tipo = tipo_map.get(ext, 'otro')

        a = ArchivoOportunidad.objects.create(
            nombre_original=archivo_file.name,
            archivo=archivo_file,
            tipo_archivo=tipo,
            tamaño=archivo_file.size,
            oportunidad=opp,
            carpeta=carpeta,
            subido_por=request.user,
            extension=ext,
            mime_type=archivo_file.content_type or '',
        )
        return JsonResponse({'success': True, 'archivo': {
            'id': a.id, 'nombre': a.nombre_original,
            'tipo_archivo': a.tipo_archivo, 'extension': a.extension,
            'tamaño': a.tamaño, 'url': f'/app/api/oportunidad/{opp.id}/drive/archivo/{a.id}/stream/',
            'fecha_subida': a.fecha_subida.isoformat(), 'tipo': 'archivo'
        }})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@csrf_exempt
@login_required
def api_drive_oportunidad_archivo_detalle(request, opp_id, archivo_id):
    """Renombra o elimina un archivo del drive de una oportunidad."""
    opp = get_object_or_404(TodoItem, id=opp_id)
    archivo = get_object_or_404(ArchivoOportunidad, id=archivo_id, oportunidad=opp)

    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'JSON inválido'}, status=400)
        nuevo = (data.get('nombre') or '').strip()
        if nuevo:
            archivo.nombre_original = nuevo
            archivo.save()
        return JsonResponse({'success': True, 'archivo': {'id': archivo.id, 'nombre': archivo.nombre_original}})

    elif request.method == 'DELETE':
        if archivo.archivo:
            try:
                archivo.archivo.delete(save=False)
            except Exception:
                pass
        archivo.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_drive_archivo_stream(request, opp_id, archivo_id):
    """
    Descarga/previsualización en tiempo real (proxy streaming).
    - Archivo local → FileResponse directo.
    - Solo bitrix_file_id → pide URL fresca a Bitrix y hace stream al cliente sin guardar nada.
    Parámetro opcional: ?dl=1 para forzar descarga, sin él intenta inline (previsualización).
    """
    import urllib.parse
    from django.http import StreamingHttpResponse, FileResponse

    opp = get_object_or_404(TodoItem, id=opp_id)
    archivo = get_object_or_404(ArchivoOportunidad, id=archivo_id, oportunidad=opp)

    # Caso 1: archivo físico local
    if archivo.archivo:
        try:
            content_type, _ = mimetypes.guess_type(archivo.nombre_original)
            if not content_type:
                content_type = 'application/octet-stream'

            response = FileResponse(
                archivo.archivo.open('rb'),
                content_type=content_type
            )
            
            nombre_encoded = urllib.parse.quote(archivo.nombre_original)
            if request.GET.get('dl') == '1':
                response['Content-Disposition'] = f"attachment; filename*=UTF-8''{nombre_encoded}"
            else:
                # IMPORTANTE: disposition='inline' para que el navegador lo muestre
                response['Content-Disposition'] = f"inline; filename*=UTF-8''{nombre_encoded}"
            
            return response
        except Exception as e:
            print(f"Error streaming local file: {e}")
            pass  # fallback a Bitrix si el físico no existe

    # Caso 2: archivo en Bitrix (streaming sin guardar)
    if not archivo.bitrix_file_id:
        return JsonResponse({'error': 'Archivo no disponible'}, status=404)

    bitrix_webhook = os.getenv(
        "BITRIX_PROJECTS_WEBHOOK_URL",
        "https://bajanet.bitrix24.mx/rest/86/hwpxu5dr31b6wve3/sonet_group.create.json"
    )
    base = bitrix_webhook.rsplit("/", 1)[0] + "/"

    # Obtener URL de descarga fresca (expiran, por eso se pide cada vez)
    download_url = ""
    try:
        r = requests.post(base + "disk.file.get.json",
                          json={"id": archivo.bitrix_file_id}, timeout=15)
        r.raise_for_status()
        result = r.json().get("result") or {}
        download_url = result.get("DOWNLOAD_URL") or result.get("downloadUrl") or ""
    except Exception:
        download_url = archivo.bitrix_download_url  # fallback a URL guardada

    if not download_url:
        return JsonResponse({'error': 'No se pudo obtener URL del archivo en Bitrix'}, status=502)

    # Stream desde Bitrix → cliente (chunk de 32 KB, sin tocar disco)
    try:
        bitrix_resp = requests.get(download_url, stream=True, timeout=60)
        bitrix_resp.raise_for_status()

        content_type = bitrix_resp.headers.get("Content-Type", "")
        if not content_type or content_type == 'application/octet-stream':
             guessed_type, _ = mimetypes.guess_type(archivo.nombre_original)
             if guessed_type:
                 content_type = guessed_type
             else:
                 content_type = 'application/octet-stream'

        nombre_encoded = urllib.parse.quote(archivo.nombre_original)
        disposition = "attachment" if request.GET.get('dl') == '1' else "inline"

        streaming = StreamingHttpResponse(
            bitrix_resp.iter_content(chunk_size=32768),
            content_type=content_type,
        )
        streaming["Content-Disposition"] = f"{disposition}; filename*=UTF-8''{nombre_encoded}"
        size = bitrix_resp.headers.get("Content-Length") or (archivo.tamaño if archivo.tamaño else None)
        if size:
            streaming["Content-Length"] = str(size)
        return streaming

    except Exception as e:
        return JsonResponse({'error': f'Error al obtener archivo: {str(e)}'}, status=502)


# =============================================
# CHAT / BITÁCORA DE OPORTUNIDAD
# =============================================

@csrf_exempt
@login_required
def api_chat_oportunidad(request, opp_id):
    from .models import MensajeOportunidad
    opp = get_object_or_404(TodoItem, id=opp_id)

    import re as _re

    def serializar(m):
        u = m.usuario
        nombre = u.get_full_name() or u.username if u else 'Usuario'
        iniciales = ''.join([p[0].upper() for p in nombre.split()[:2]]) if nombre else '?'
        reply = None
        if m.reply_to:
            ru = m.reply_to.usuario
            rnombre = ru.get_full_name() or ru.username if ru else 'Usuario'
            reply = {
                'id': m.reply_to.id,
                'texto': m.reply_to.texto[:80],
                'nombre': rnombre,
                'tiene_imagen': bool(m.reply_to.imagen),
            }
        # Detectar y limpiar prefijo técnico de Bitrix
        texto = m.texto
        es_bitrix = False
        bitrix_tipo = None
        bitrix_match = _re.match(r'^\[BITRIX_(ACT|CMT):(\d+)\]\s*', texto)
        if bitrix_match:
            es_bitrix = True
            bitrix_tipo = 'actividad' if bitrix_match.group(1) == 'ACT' else 'comentario'
            texto = texto[bitrix_match.end():]
            # Comentarios de Bitrix: mostrar como mensaje normal del autor
        return {
            'id': m.id,
            'texto': texto,
            'imagen_url': m.imagen.url if m.imagen else None,
            'editado': m.editado,
            'reply_to': reply,
            'fecha': m.fecha.strftime('%d/%m/%Y %H:%M'),
            'usuario_id': u.id if u else None,
            'nombre': nombre,
            'iniciales': iniciales,
            'es_mio': (u.id == request.user.id) if u else False,
            'es_bitrix': es_bitrix,
            'bitrix_tipo': bitrix_tipo,
        }

    if request.method == 'GET':
        msgs = MensajeOportunidad.objects.filter(oportunidad=opp).select_related('usuario', 'reply_to', 'reply_to__usuario')
        return JsonResponse({'mensajes': [serializar(m) for m in msgs]})

    if request.method == 'POST':
        texto = (request.POST.get('texto') or '').strip()
        imagen = request.FILES.get('imagen')
        reply_to_id = request.POST.get('reply_to_id')
        if not texto and not imagen:
            return JsonResponse({'error': 'El mensaje no puede estar vacío'}, status=400)
        reply_obj = None
        if reply_to_id:
            try:
                reply_obj = MensajeOportunidad.objects.get(id=int(reply_to_id), oportunidad=opp)
            except MensajeOportunidad.DoesNotExist:
                pass
        m = MensajeOportunidad.objects.create(
            oportunidad=opp,
            usuario=request.user,
            texto=texto,
            imagen=imagen,
            reply_to=reply_obj,
        )

        # Notificar al dueño de la oportunidad si es distinto al que envió
        if opp.usuario and opp.usuario != request.user:
            remitente_nombre = request.user.get_full_name() or request.user.username
            preview = texto[:100] + ('…' if len(texto) > 100 else '') if texto else '📷 Imagen'
            crear_notificacion(
                usuario_destinatario=opp.usuario,
                tipo='oportunidad_mensaje',
                titulo=f'Mensaje en: {opp.oportunidad}',
                mensaje=f'{remitente_nombre}: {preview}',
                oportunidad=opp,
                usuario_remitente=request.user,
            )

        return JsonResponse({'success': True, 'mensaje': serializar(m)})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@csrf_exempt
@login_required
def api_chat_mensaje(request, opp_id, msg_id):
    from .models import MensajeOportunidad
    opp = get_object_or_404(TodoItem, id=opp_id)
    msg = get_object_or_404(MensajeOportunidad, id=msg_id, oportunidad=opp)
    es_supervisor = request.user.groups.filter(name='Supervisores').exists() or request.user.is_superuser

    if request.method == 'PUT':
        if msg.usuario != request.user:
            return JsonResponse({'error': 'Sin permiso'}, status=403)
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'JSON inválido'}, status=400)
        nuevo = (data.get('texto') or '').strip()
        if not nuevo:
            return JsonResponse({'error': 'El texto no puede estar vacío'}, status=400)
        msg.texto = nuevo
        msg.editado = True
        msg.save()
        return JsonResponse({'success': True})

    if request.method == 'DELETE':
        if msg.usuario != request.user and not es_supervisor:
            return JsonResponse({'error': 'Sin permiso'}, status=403)
        if msg.imagen:
            msg.imagen.delete(save=False)
        msg.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


# =============================================
# VIEWS PARA INTERCAMBIO NAVIDEÑO 🎄
# =============================================

@login_required
def intercambio_navidad(request):
    """Vista principal del intercambio navideño"""
    from datetime import datetime
    from .models import IntercambioNavidad, ParticipanteIntercambio
    
    # Obtener el intercambio del año actual
    año_actual = datetime.now().year
    intercambio = IntercambioNavidad.objects.filter(año=año_actual).first()
    
    # Variables para el template
    context = {
        'employees': User.objects.filter(is_active=True).order_by('first_name', 'last_name'),
        'sorteo_realizado': False,
        'persona_asignada': None,
        'fecha_intercambio': None,
    }
    
    # Si existe intercambio, obtener información
    if intercambio:
        context['intercambio'] = intercambio
        context['sorteo_realizado'] = intercambio.sorteo_realizado
        context['fecha_intercambio'] = intercambio.fecha_intercambio
        
        # Si el usuario no es superusuario y ya se realizó el sorteo
        if not request.user.is_superuser and intercambio.sorteo_realizado:
            # Buscar la asignación del usuario actual
            participante = ParticipanteIntercambio.objects.filter(
                intercambio=intercambio,
                usuario=request.user
            ).first()
            
            if participante and participante.regalo_para:
                context['persona_asignada'] = participante.regalo_para.get_full_name() or participante.regalo_para.username
    
    return render(request, 'intercambio_navidad.html', context)


@csrf_exempt
@login_required
def realizar_sorteo_navidad(request):
    """API para realizar el sorteo navideño (solo superusuarios)"""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'No tienes permisos para realizar el sorteo'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        import json
        import random
        from datetime import datetime
        from .models import IntercambioNavidad, ParticipanteIntercambio, HistorialIntercambio
        
        # Crear o obtener intercambio del año actual para verificar participantes
        año_actual = datetime.now().year
        intercambio = IntercambioNavidad.objects.filter(año=año_actual).first()
        
        if not intercambio:
            return JsonResponse({'error': 'No hay intercambio configurado para este año'}, status=400)
        
        # Obtener participantes existentes de la base de datos
        participantes_db = list(ParticipanteIntercambio.objects.filter(intercambio=intercambio))
        
        if len(participantes_db) < 2:
            return JsonResponse({'error': 'Se necesitan al menos 2 participantes'}, status=400)
        
        # Permitir repetir el sorteo (útil cuando se agregan más participantes)
        if intercambio.sorteo_realizado:
            # Solo mostrar una advertencia en consola, pero permitir continuar
            print(f"Repitiendo sorteo para intercambio {intercambio.año}")
        
        # Limpiar asignaciones anteriores (por si se vuelve a ejecutar)
        for participante in participantes_db:
            participante.regalo_para = None
            participante.save()
        
        # Los participantes ya están en la base de datos
        participantes = participantes_db
        
        # Realizar el algoritmo de sorteo (amigo invisible)
        usuarios = [p.usuario for p in participantes]
        usuarios_disponibles = usuarios.copy()
        
        # Mezclar para randomizar
        random.shuffle(usuarios_disponibles)
        
        asignaciones = []
        for i, participante in enumerate(participantes):
            # Filtrar usuarios disponibles (no puede regalarse a sí mismo)
            opciones = [u for u in usuarios_disponibles if u != participante.usuario]
            
            if not opciones:
                # Si no hay opciones, reiniciar el proceso
                return JsonResponse({'error': 'Error en el sorteo, intenta de nuevo'}, status=500)
            
            # Seleccionar usuario aleatorio
            usuario_asignado = random.choice(opciones)
            usuarios_disponibles.remove(usuario_asignado)
            
            # Guardar asignación
            participante.regalo_para = usuario_asignado
            participante.save()
            
            asignaciones.append({
                'participante': participante.usuario.get_full_name(),
                'regalo_para': usuario_asignado.get_full_name()
            })
        
        # Actualizar estado del intercambio
        intercambio.estado = 'sorteo_realizado'
        intercambio.fecha_sorteo = datetime.now()
        intercambio.save()
        
        # Registrar en historial
        HistorialIntercambio.objects.create(
            intercambio=intercambio,
            accion='sorteo_realizado',
            usuario=request.user,
            detalles={
                'total_participantes': len(participantes),
                'asignaciones': asignaciones
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Sorteo realizado exitosamente',
            'total_participantes': len(participantes),
            'intercambio_id': intercambio.id
        })
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        print(f"Error en sorteo navideño: {e}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': f'Error interno del servidor: {str(e)}'}, status=500)


@csrf_exempt
def agregar_participante_navidad(request):
    """Agregar participante al intercambio navideño"""
    import json
    from datetime import datetime
    from .models import IntercambioNavidad, ParticipanteIntercambio
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
        
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        empleado_id = data.get('empleado_id')
        
        if not empleado_id:
            return JsonResponse({'error': 'ID de empleado requerido'}, status=400)
        
        # Obtener usuario
        empleado = User.objects.get(id=empleado_id)
        
        # Crear o obtener intercambio del año actual
        año_actual = datetime.now().year
        intercambio, created = IntercambioNavidad.objects.get_or_create(
            año=año_actual,
            defaults={
                'fecha_intercambio': datetime(año_actual, 12, 15).date(),
                'creado_por': request.user
            }
        )
        
        # Verificar si ya está participando
        if ParticipanteIntercambio.objects.filter(intercambio=intercambio, usuario=empleado).exists():
            return JsonResponse({'error': 'El empleado ya está participando'}, status=400)
        
        # Agregar participante
        participante = ParticipanteIntercambio.objects.create(
            intercambio=intercambio,
            usuario=empleado
        )
        
        return JsonResponse({
            'success': True,
            'participante': {
                'id': participante.id,
                'usuario': {
                    'id': empleado.id,
                    'first_name': empleado.first_name,
                    'last_name': empleado.last_name,
                    'username': empleado.username
                }
            }
        })
        
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def eliminar_participante_navidad(request):
    """Eliminar participante del intercambio navideño"""
    import json
    from .models import ParticipanteIntercambio
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
        
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        participante_id = data.get('participante_id')
        
        if not participante_id:
            return JsonResponse({'error': 'ID de participante requerido'}, status=400)
        
        # Eliminar participante
        participante = ParticipanteIntercambio.objects.get(id=participante_id)
        participante.delete()
        
        return JsonResponse({'success': True})
        
    except ParticipanteIntercambio.DoesNotExist:
        return JsonResponse({'error': 'Participante no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def listar_participantes_navidad(request):
    """Listar participantes del intercambio navideño"""
    from datetime import datetime
    from .models import IntercambioNavidad, ParticipanteIntercambio
    
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        año_actual = datetime.now().year
        intercambio = IntercambioNavidad.objects.filter(año=año_actual).first()
        
        if not intercambio:
            return JsonResponse({'participantes': []})
        
        participantes = ParticipanteIntercambio.objects.filter(intercambio=intercambio)
        
        data = []
        for participante in participantes:
            regalo_para = None
            if participante.regalo_para:
                regalo_para = {
                    'id': participante.regalo_para.id,
                    'first_name': participante.regalo_para.first_name,
                    'last_name': participante.regalo_para.last_name,
                    'username': participante.regalo_para.username
                }
            
            data.append({
                'id': participante.id,
                'usuario': {
                    'id': participante.usuario.id,
                    'first_name': participante.usuario.first_name,
                    'last_name': participante.usuario.last_name,
                    'username': participante.usuario.username
                },
                'regalo_para': regalo_para
            })
        
        return JsonResponse({'participantes': data})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def actualizar_evento_navidad(request):
    """Actualizar configuración del evento navideño"""
    import json
    from datetime import datetime
    from .models import IntercambioNavidad
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
        
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        data = json.loads(request.body)
        fecha_intercambio_str = data.get('fecha_intercambio')
        monto_sugerido = data.get('monto_sugerido', 500)
        descripcion = data.get('descripcion', '')
        
        if not fecha_intercambio_str:
            return JsonResponse({'error': 'Fecha del intercambio es requerida'}, status=400)
        
        # Convertir la fecha string a objeto date
        try:
            from datetime import datetime as dt
            fecha_intercambio = dt.strptime(fecha_intercambio_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)
        
        # Crear o obtener intercambio del año actual
        año_actual = datetime.now().year
        intercambio, created = IntercambioNavidad.objects.get_or_create(
            año=año_actual,
            defaults={
                'fecha_intercambio': fecha_intercambio,
                'monto_sugerido': monto_sugerido,
                'descripcion': descripcion,
                'creado_por': request.user
            }
        )
        
        # Si ya existe, actualizar los valores
        if not created:
            intercambio.fecha_intercambio = fecha_intercambio
            intercambio.monto_sugerido = monto_sugerido
            intercambio.descripcion = descripcion
            intercambio.save()
        
        return JsonResponse({
            'success': True,
            'intercambio': {
                'id': intercambio.id,
                'fecha_intercambio': intercambio.fecha_intercambio.strftime('%Y-%m-%d'),
                'monto_sugerido': intercambio.monto_sugerido,
                'descripcion': intercambio.descripcion
            }
        })
        
    except Exception as e:
        print(f"Error en actualizar_evento_navidad: {e}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': f'Error interno del servidor: {str(e)}'}, status=500)


@csrf_exempt
def estado_usuario_navidad(request):
    """Obtener estado del usuario en el intercambio navideño"""
    from datetime import datetime
    from .models import IntercambioNavidad, ParticipanteIntercambio
    
    if request.method != 'GET':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        año_actual = datetime.now().year
        intercambio = IntercambioNavidad.objects.filter(año=año_actual).first()
        
        if not intercambio:
            return JsonResponse({
                'participa': False,
                'sorteo_realizado': False,
                'regalo_para': None,
                'monto_sugerido': 500,
                'fecha_intercambio': None
            })
        
        # Buscar si el usuario participa
        participante = ParticipanteIntercambio.objects.filter(
            intercambio=intercambio,
            usuario=request.user
        ).first()
        
        if not participante:
            return JsonResponse({
                'participa': False,
                'sorteo_realizado': intercambio.sorteo_realizado,
                'regalo_para': None,
                'monto_sugerido': float(intercambio.monto_sugerido) if intercambio.monto_sugerido else 500,
                'fecha_intercambio': intercambio.fecha_intercambio.strftime('%d de %B') if intercambio.fecha_intercambio else None
            })
        
        # El usuario participa
        regalo_para_data = None
        if participante.regalo_para:
            regalo_para_data = {
                'id': participante.regalo_para.id,
                'first_name': participante.regalo_para.first_name,
                'last_name': participante.regalo_para.last_name,
                'username': participante.regalo_para.username
            }
        
        return JsonResponse({
            'participa': True,
            'sorteo_realizado': intercambio.sorteo_realizado,
            'regalo_para': regalo_para_data,
            'monto_sugerido': float(intercambio.monto_sugerido) if intercambio.monto_sugerido else 500,
            'fecha_intercambio': intercambio.fecha_intercambio.strftime('%d de %B') if intercambio.fecha_intercambio else None
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def solicitar_acceso_proyecto(request, proyecto_id):
    """
    API para solicitar acceso a un proyecto privado
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Verificar si el modelo existe antes de usar
        try:
            from app.models import SolicitudAccesoProyecto
            from django.db import connection
            
            # Verificar si la tabla existe en la base de datos
            with connection.cursor() as cursor:
                # Consulta compatible con MySQL y SQLite
                if 'mysql' in connection.vendor:
                    cursor.execute("""
                        SELECT table_name FROM information_schema.tables 
                        WHERE table_schema = DATABASE() AND table_name = 'app_solicitudaccesoproyecto';
                    """)
                else:
                    cursor.execute("""
                        SELECT name FROM sqlite_master WHERE type='table' AND name='app_solicitudaccesoproyecto';
                    """)
                table_exists = cursor.fetchone()
                
            if not table_exists:
                return JsonResponse({'error': 'Sistema de solicitudes no disponible. Tabla no existe en la base de datos.'}, status=503)
                
        except ImportError:
            return JsonResponse({'error': 'Sistema de solicitudes no disponible. Aplique migraciones.'}, status=503)
            
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        
        # Verificar que sea un proyecto privado
        if proyecto.privacidad != 'privado':
            return JsonResponse({'error': 'Este proyecto es público'}, status=400)
        
        # Verificar que no sea ya miembro
        if (request.user == proyecto.creado_por or 
            request.user in proyecto.miembros.all()):
            return JsonResponse({'error': 'Ya eres miembro de este proyecto'}, status=400)
        
        # Verificar si ya tiene una solicitud pendiente
        from app.models import SolicitudAccesoProyecto
        if SolicitudAccesoProyecto.objects.filter(
            proyecto=proyecto,
            usuario_solicitante=request.user,
            estado='pendiente'
        ).exists():
            return JsonResponse({'error': 'Ya tienes una solicitud pendiente'}, status=400)
        
        # Crear la solicitud
        import json
        data = json.loads(request.body) if request.body else {}
        mensaje = data.get('mensaje', '')
        
        solicitud = SolicitudAccesoProyecto.objects.create(
            proyecto=proyecto,
            usuario_solicitante=request.user,
            mensaje=mensaje
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Solicitud enviada exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def responder_solicitud_proyecto(request, solicitud_id):
    """
    API para aceptar o rechazar una solicitud de acceso a proyecto
    """
    print(f"🔍 DEBUG responder_solicitud_proyecto - Usuario: {request.user.username if request.user.is_authenticated else 'No autenticado'}")
    print(f"🔍 DEBUG - Solicitud ID: {solicitud_id}")
    print(f"🔍 DEBUG - Método: {request.method}")
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Verificar si el modelo existe antes de usar
        try:
            from app.models import SolicitudAccesoProyecto
            from django.db import connection
            
            # Verificar si la tabla existe en la base de datos
            with connection.cursor() as cursor:
                # Consulta compatible con MySQL y SQLite
                if 'mysql' in connection.vendor:
                    cursor.execute("""
                        SELECT table_name FROM information_schema.tables 
                        WHERE table_schema = DATABASE() AND table_name = 'app_solicitudaccesoproyecto';
                    """)
                else:
                    cursor.execute("""
                        SELECT name FROM sqlite_master WHERE type='table' AND name='app_solicitudaccesoproyecto';
                    """)
                table_exists = cursor.fetchone()
                
            if not table_exists:
                return JsonResponse({'error': 'Sistema de solicitudes no disponible. Tabla no existe en la base de datos.'}, status=503)
                
        except ImportError:
            return JsonResponse({'error': 'Sistema de solicitudes no disponible. Aplique migraciones.'}, status=503)
            
        solicitud = get_object_or_404(SolicitudAccesoProyecto, id=solicitud_id)
        print(f"🔍 DEBUG - Solicitud encontrada: {solicitud.proyecto.nombre}")
        print(f"🔍 DEBUG - Creador proyecto: {solicitud.proyecto.creado_por.username}")
        print(f"🔍 DEBUG - Usuario actual: {request.user.username}")
        
        # Verificar que sea el creador del proyecto o un miembro
        es_creador = request.user == solicitud.proyecto.creado_por
        es_miembro = request.user in solicitud.proyecto.miembros.all()
        print(f"🔍 DEBUG - ¿Es creador?: {es_creador}")
        print(f"🔍 DEBUG - ¿Es miembro?: {es_miembro}")
        
        if not (es_creador or es_miembro):
            print(f"🔍 DEBUG - NO TIENE PERMISOS - 403")
            return JsonResponse({'error': 'No tienes permisos para responder esta solicitud'}, status=403)
        
        # Verificar que la solicitud esté pendiente
        if solicitud.estado != 'pendiente':
            return JsonResponse({'error': 'Esta solicitud ya fue respondida'}, status=400)
        
        import json
        data = json.loads(request.body)
        accion = data.get('accion')  # 'aceptar' o 'rechazar'
        
        if accion == 'aceptar':
            # Agregar usuario como miembro del proyecto
            solicitud.proyecto.miembros.add(solicitud.usuario_solicitante)
            solicitud.estado = 'aprobada'
            mensaje_respuesta = 'Solicitud aceptada. El usuario ahora es miembro del proyecto.'
        elif accion == 'rechazar':
            solicitud.estado = 'rechazada'
            mensaje_respuesta = 'Solicitud rechazada.'
        else:
            return JsonResponse({'error': 'Acción no válida'}, status=400)
        
        # Actualizar la solicitud
        from django.utils import timezone
        solicitud.fecha_respuesta = timezone.now()
        solicitud.usuario_respuesta = request.user
        solicitud.save()
        
        return JsonResponse({
            'success': True,
            'message': mensaje_respuesta
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def obtener_solicitudes_proyecto(request):
    """
    API para obtener solicitudes pendientes de proyectos donde el usuario puede responder
    """
    try:
        # Verificar si el modelo existe antes de usar
        try:
            from app.models import SolicitudAccesoProyecto, Proyecto
            from django.db import connection
            
            # Verificar si la tabla existe en la base de datos
            with connection.cursor() as cursor:
                # Consulta compatible con MySQL y SQLite
                if 'mysql' in connection.vendor:
                    cursor.execute("""
                        SELECT table_name FROM information_schema.tables 
                        WHERE table_schema = DATABASE() AND table_name = 'app_solicitudaccesoproyecto';
                    """)
                else:
                    cursor.execute("""
                        SELECT name FROM sqlite_master WHERE type='table' AND name='app_solicitudaccesoproyecto';
                    """)
                table_exists = cursor.fetchone()
                
            if not table_exists:
                return JsonResponse({
                    'solicitudes': [], 
                    'total': 0,
                    'message': 'Sistema de solicitudes no disponible'
                })
                
        except (ImportError, Exception) as e:
            return JsonResponse({
                'solicitudes': [], 
                'total': 0,
                'message': f'Sistema de solicitudes no disponible: {str(e)}'
            })
        
        # Obtener proyectos donde el usuario es creador o miembro
        proyectos_con_permisos = Proyecto.objects.filter(
            models.Q(creado_por=request.user) | 
            models.Q(miembros=request.user),
            privacidad='privado'
        ).distinct()
        
        # Obtener solicitudes pendientes para esos proyectos
        solicitudes = SolicitudAccesoProyecto.objects.filter(
            proyecto__in=proyectos_con_permisos,
            estado='pendiente'
        ).select_related('proyecto', 'usuario_solicitante').order_by('-fecha_solicitud')
        
        solicitudes_data = []
        for solicitud in solicitudes:
            solicitudes_data.append({
                'id': solicitud.id,
                'proyecto_nombre': solicitud.proyecto.nombre,
                'proyecto_id': solicitud.proyecto.id,
                'usuario_nombre': solicitud.usuario_solicitante.get_full_name(),
                'usuario_username': solicitud.usuario_solicitante.username,
                'mensaje': solicitud.mensaje,
                'fecha_solicitud': solicitud.fecha_solicitud.strftime('%d de %b, %Y a las %H:%M'),
            })
        
        return JsonResponse({
            'solicitudes': solicitudes_data,
            'total': len(solicitudes_data)
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_eliminar_proyecto_completo(request, proyecto_id):
    """
    API para eliminar un proyecto completo con todas sus dependencias
    Elimina: proyecto, tareas, carpetas, archivos, comentarios, etc.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        print(f"🔍 DEBUG: Buscando proyecto con ID: {proyecto_id}")
        
        # Obtener el proyecto
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        print(f"✅ DEBUG: Proyecto encontrado: {proyecto.nombre}")
        
        # Verificar que el usuario es el creador del proyecto
        print(f"🔍 DEBUG: Creado por: {proyecto.creado_por}, Usuario actual: {request.user}")
        if proyecto.creado_por != request.user:
            return JsonResponse({'error': 'No tienes permisos para eliminar este proyecto'}, status=403)
        
        # Log para debugging
        print(f"🗑️ Iniciando eliminación del proyecto '{proyecto.nombre}' (ID: {proyecto_id})")
        
        import os
        from django.conf import settings
        
        # Contador para logging
        elementos_eliminados = {
            'tareas': 0,
            'archivos': 0,
            'carpetas': 0,
            'comentarios': 0,
            'solicitudes': 0,
            'archivos_fisicos': 0
        }
        
        # 1. Eliminar todas las tareas del proyecto
        tareas = Tarea.objects.filter(proyecto=proyecto)
        elementos_eliminados['tareas'] = tareas.count()
        tareas.delete()
        print(f"   ✅ {elementos_eliminados['tareas']} tareas eliminadas")
        
        # 2. Eliminar archivos físicos y registros
        archivos = ArchivoProyecto.objects.filter(proyecto=proyecto)
        elementos_eliminados['archivos'] = archivos.count()
        
        for archivo in archivos:
            # Eliminar archivo físico del servidor
            try:
                if archivo.archivo and hasattr(archivo.archivo, 'path'):
                    archivo_path = archivo.archivo.path
                    if os.path.exists(archivo_path):
                        os.remove(archivo_path)
                        elementos_eliminados['archivos_fisicos'] += 1
                        print(f"   🗂️ Archivo físico eliminado: {archivo_path}")
            except Exception as e:
                print(f"   ⚠️ Error eliminando archivo físico {archivo.nombre}: {e}")
        
        archivos.delete()
        print(f"   ✅ {elementos_eliminados['archivos']} archivos eliminados")
        
        # 3. Eliminar carpetas
        carpetas = CarpetaProyecto.objects.filter(proyecto=proyecto)
        elementos_eliminados['carpetas'] = carpetas.count()
        carpetas.delete()
        print(f"   ✅ {elementos_eliminados['carpetas']} carpetas eliminadas")
        
        # 4. Eliminar comentarios
        comentarios = ProyectoComentario.objects.filter(proyecto=proyecto)
        elementos_eliminados['comentarios'] = comentarios.count()
        comentarios.delete()
        print(f"   ✅ {elementos_eliminados['comentarios']} comentarios eliminados")
        
        # 5. Eliminar solicitudes de acceso
        solicitudes = SolicitudAccesoProyecto.objects.filter(proyecto=proyecto)
        elementos_eliminados['solicitudes'] = solicitudes.count()
        solicitudes.delete()
        print(f"   ✅ {elementos_eliminados['solicitudes']} solicitudes de acceso eliminadas")
        
        # 6. Finalmente, eliminar el proyecto
        proyecto_nombre = proyecto.nombre
        proyecto.delete()
        
        print(f"🎯 Proyecto '{proyecto_nombre}' eliminado completamente")
        print(f"📊 Resumen: {elementos_eliminados}")
        
        return JsonResponse({
            'success': True, 
            'message': f'Proyecto "{proyecto_nombre}" eliminado exitosamente',
            'elementos_eliminados': elementos_eliminados
        })
        
    except Proyecto.DoesNotExist:
        print(f"❌ DEBUG: Proyecto con ID {proyecto_id} no encontrado")
        return JsonResponse({'error': 'Proyecto no encontrado'}, status=404)
    except Exception as e:
        print(f"❌ DEBUG: Error eliminando proyecto: {e}")
        print(f"❌ DEBUG: Tipo de error: {type(e)}")
        import traceback
        print(f"❌ DEBUG: Traceback completo:")
        traceback.print_exc()
        return JsonResponse({'error': f'Error interno del servidor: {str(e)}'}, status=500)


@login_required
def api_actualizar_tarea_real(request, tarea_id):
    """
    API para actualizar una tarea real en la base de datos
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        import json
        from datetime import datetime
        
        # Obtener la tarea de la base de datos
        tarea = get_object_or_404(Tarea, id=tarea_id)
        
        # Verificar permisos - solo el creador o superusuario puede editar campos principales
        if tarea.creado_por != request.user and not request.user.is_superuser:
            return JsonResponse({'error': 'Sin permisos para editar esta tarea'}, status=403)
        
        # Obtener datos del request
        data = json.loads(request.body)
        
        # Actualizar campos si están presentes en la petición
        if 'nombre' in data or 'titulo' in data:
            nuevo_titulo = data.get('nombre') or data.get('titulo')
            if nuevo_titulo and nuevo_titulo.strip():
                tarea.titulo = nuevo_titulo.strip()
        
        if 'descripcion' in data:
            tarea.descripcion = data.get('descripcion', '') or ''
        
        if 'prioridad' in data:
            prioridad = data.get('prioridad')
            if prioridad in ['baja', 'media', 'alta']:
                tarea.prioridad = prioridad
        
        if 'fecha_limite' in data:
            fecha_limite_str = data.get('fecha_limite')
            if fecha_limite_str:
                try:
                    # Intentar parsear la fecha
                    from django.utils import timezone
                    from datetime import datetime
                    fecha_obj = datetime.fromisoformat(fecha_limite_str.replace('Z', '+00:00'))
                    tarea.fecha_limite = timezone.make_aware(fecha_obj) if timezone.is_naive(fecha_obj) else fecha_obj
                except (ValueError, TypeError):
                    pass  # Ignorar fechas inválidas
            else:
                tarea.fecha_limite = None

        if 'asignado_a' in data:
            asignado_id = data.get('asignado_a')
            if asignado_id:
                try:
                    from django.contrib.auth.models import User
                    tarea.asignado_a = User.objects.get(id=asignado_id)
                except User.DoesNotExist:
                    pass
            else:
                tarea.asignado_a = None

        if 'cliente_id' in data:
            cliente_id_val = data.get('cliente_id')
            if cliente_id_val:
                try:
                    tarea.cliente = Cliente.objects.get(id=cliente_id_val)
                except Cliente.DoesNotExist:
                    pass
            else:
                tarea.cliente = None

        if 'oportunidad_id' in data:
            opp_id_val = data.get('oportunidad_id')
            if opp_id_val:
                try:
                    tarea.oportunidad = TodoItem.objects.get(id=opp_id_val)
                except TodoItem.DoesNotExist:
                    pass
            else:
                tarea.oportunidad = None

        # Guardar cambios
        tarea.save()
        
        # Log para debugging
        print(f"✅ Tarea {tarea_id} actualizada: {tarea.titulo}")
        
        # Devolver datos actualizados
        return JsonResponse({
            'success': True,
            'message': 'Tarea actualizada exitosamente',
            'tarea': {
                'id': tarea.id,
                'titulo': tarea.titulo,
                'descripcion': tarea.descripcion,
                'prioridad': tarea.prioridad,
                'fecha_limite': tarea.fecha_limite.isoformat() if tarea.fecha_limite else None,
                'estado': tarea.estado
            }
        })
        
    except Tarea.DoesNotExist:
        return JsonResponse({'error': 'Tarea no encontrada'}, status=404)
    except Exception as e:
        print(f"❌ Error actualizando tarea {tarea_id}: {e}")
        return JsonResponse({'error': f'Error interno del servidor: {str(e)}'}, status=500)


# ═══════════════════════════════════════════════════════════
# API: EXPORTAR A EXCEL
# ═══════════════════════════════════════════════════════════

@login_required
def api_export_excel(request):
    """
    Exporta la tabla activa a Excel con metadata
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from datetime import datetime
        
        # Obtener parámetros
        tab = request.GET.get('tab', 'crm')
        mes = request.GET.get('mes', '')
        anio = request.GET.get('anio', '')
        usuario = request.GET.get('usuario', request.user.get_full_name() or request.user.username)
        filtros = request.GET.get('filtros', 'Sin filtros')
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Reporte {tab.upper()}"
        
        # Metadata
        ws['A1'] = 'Reporte CRM'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'Descargado por: {usuario}'
        ws['A3'] = f'Fecha: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A4'] = f'Periodo: {mes}/{anio}'
        ws['A5'] = f'Filtros aplicados: {filtros}'
        
        # Espacio
        row_start = 7
        
        # Obtener datos según el tab
        if tab == 'crm':
            headers = ['Oportunidad', 'Cliente', 'Contacto', 'Área', 'Zebra', 'Panduit', 'APC', 'Avigilon', 'Genetec', 'Axis', 'Software', 'Runrate', 'Póliza', 'Otros', 'Total']
            data = get_oportunidades_data(request, mes, anio)
        elif tab == 'facturado':
            headers = ['Cliente', 'Zebra', 'Panduit', 'APC', 'Avigilon', 'Genetec', 'Axis', 'Software', 'Runrate', 'Póliza', 'Otros', 'Total', 'Meta', 'Fact.']
            data = get_facturado_data(request, mes, anio)
        else:
            headers = ['Datos']
            data = []
        
        # Escribir headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_start, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0052D4", end_color="0052D4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Escribir datos
        for row_num, row_data in enumerate(data, row_start + 1):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=reporte_crm_{tab}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response
        
    except ImportError:
        return JsonResponse({'error': 'openpyxl no está instalado. Ejecute: pip install openpyxl'}, status=500)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def get_oportunidades_data(request, mes, anio):
    """Helper para obtener datos de oportunidades"""
    # Implementar lógica similar a crm_home
    return []


def get_facturado_data(request, mes, anio):
    """Helper para obtener datos de facturado"""
    # Implementar lógica similar a crm_home
    return []


@login_required
def api_clientes(request):
    """
    Retorna lista de clientes para filtros
    """
    try:
        clientes = Cliente.objects.all().values('id', 'nombre_empresa')
        return JsonResponse({'clientes': list(clientes)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ── Tareas de Oportunidad ──────────────────────────────────────────────────

def _serialize_tarea_opp(t):
    return {
        'id': t.id,
        'titulo': t.titulo,
        'descripcion': t.descripcion,
        'prioridad': t.prioridad,
        'estado': t.estado,
        'fecha_limite': t.fecha_limite.isoformat() if t.fecha_limite else None,
        'fecha_creacion': t.fecha_creacion.isoformat(),
        'creado_por': (t.creado_por.get_full_name() or t.creado_por.username) if t.creado_por else '',
        'creado_por_id': t.creado_por_id,
        'responsable': (t.responsable.get_full_name() or t.responsable.username) if t.responsable else None,
        'responsable_id': t.responsable_id,
        'oportunidad_id': t.oportunidad_id,
        'oportunidad_nombre': t.oportunidad.oportunidad if t.oportunidad else '',
    }


@login_required
def api_tareas_oportunidad(request, opp_id):
    """GET lista / POST crear — tareas de una oportunidad específica."""
    try:
        opp = get_object_or_404(TodoItem, pk=opp_id)

        if request.method == 'GET':
            from django.db.models import IntegerField
            tareas = opp.tareas_oportunidad.select_related('creado_por', 'responsable').annotate(
                prio_order=Case(When(prioridad='alta', then=Value(0)), default=Value(1), output_field=IntegerField())
            ).order_by('prio_order', F('fecha_limite').asc(nulls_last=True), 'fecha_creacion')
            return JsonResponse({'success': True, 'tareas': [_serialize_tarea_opp(t) for t in tareas]})

        if request.method == 'POST':
            data = json.loads(request.body)
            titulo = data.get('titulo', '').strip()
            if not titulo:
                return JsonResponse({'success': False, 'error': 'Título requerido'}, status=400)

            fecha_limite = None
            raw_fecha = data.get('fecha_limite')
            if raw_fecha:
                from django.utils.dateparse import parse_datetime
                fecha_limite = parse_datetime(raw_fecha)

            tarea = TareaOportunidad.objects.create(
                oportunidad=opp,
                titulo=titulo,
                descripcion=data.get('descripcion', ''),
                prioridad='alta' if data.get('alta_prioridad') else 'normal',
                fecha_limite=fecha_limite,
                creado_por=request.user,
                responsable_id=data.get('responsable_id') or None,
            )
            if data.get('participantes'):
                tarea.participantes.set(data['participantes'])
            if data.get('observadores'):
                tarea.observadores.set(data['observadores'])

            # Crear actividad en el calendario (roja por defecto)
            from django.utils.dateparse import parse_datetime as _pdt
            cal_inicio_raw = data.get('cal_inicio')
            cal_fin_raw = data.get('cal_fin')
            if cal_inicio_raw and cal_fin_raw:
                cal_ini = _pdt(cal_inicio_raw)
                cal_fin = _pdt(cal_fin_raw)
                if cal_ini and cal_fin:
                    actividad = Actividad.objects.create(
                        titulo=tarea.titulo,
                        tipo_actividad='tarea',
                        descripcion=tarea.descripcion or '',
                        fecha_inicio=cal_ini,
                        fecha_fin=cal_fin,
                        creado_por=request.user,
                        color='#0052D4',  # azul
                        oportunidad=opp,
                    )
                    tarea.actividad_calendario = actividad
                    tarea.save(update_fields=['actividad_calendario'])

            # Notificar al responsable si es distinto al creador
            if tarea.responsable and tarea.responsable != request.user:
                remitente_nombre = request.user.get_full_name() or request.user.username
                crear_notificacion(
                    usuario_destinatario=tarea.responsable,
                    tipo='tarea_opp_asignada',
                    titulo=f'Tarea asignada: {tarea.titulo}',
                    mensaje=f'{remitente_nombre} te asignó una tarea en la oportunidad "{opp.oportunidad}".',
                    oportunidad=opp,
                    usuario_remitente=request.user,
                    tarea_opp=tarea,
                )

            return JsonResponse({'success': True, 'tarea': _serialize_tarea_opp(tarea)})

    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)

    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def api_tarea_oportunidad_detail(request, tarea_id):
    """PUT actualizar / DELETE eliminar — una tarea de oportunidad."""
    tarea = get_object_or_404(TareaOportunidad, pk=tarea_id)
    user = request.user

    if tarea.creado_por != user and tarea.responsable != user and not is_supervisor(user):
        return JsonResponse({'error': 'Sin permiso'}, status=403)

    if request.method == 'PUT':
        data = json.loads(request.body)
        if 'titulo' in data:
            tarea.titulo = data['titulo']
        if 'descripcion' in data:
            tarea.descripcion = data['descripcion']
        if 'estado' in data:
            tarea.estado = data['estado']
        if 'prioridad' in data:
            tarea.prioridad = data['prioridad']
        if 'fecha_limite' in data:
            from django.utils.dateparse import parse_datetime
            tarea.fecha_limite = parse_datetime(data['fecha_limite']) if data['fecha_limite'] else None
        if 'responsable_id' in data:
            new_resp_id = data['responsable_id'] or None
            old_resp_id = tarea.responsable_id
            tarea.responsable_id = new_resp_id
            tarea.save()
            # Notificar al nuevo responsable si cambió y es distinto al editor
            if new_resp_id and new_resp_id != old_resp_id:
                try:
                    nuevo_resp = User.objects.get(pk=new_resp_id)
                    if nuevo_resp != request.user:
                        remitente_nombre = request.user.get_full_name() or request.user.username
                        crear_notificacion(
                            usuario_destinatario=nuevo_resp,
                            tipo='tarea_opp_asignada',
                            titulo=f'Tarea asignada: {tarea.titulo}',
                            mensaje=f'{remitente_nombre} te asignó como responsable de una tarea en "{tarea.oportunidad.oportunidad if tarea.oportunidad else ""}".',
                            oportunidad=tarea.oportunidad,
                            usuario_remitente=request.user,
                            tarea_opp=tarea,
                        )
                except User.DoesNotExist:
                    pass
            return JsonResponse({'success': True})
        # M2M: añadir/quitar participantes y observadores
        remitente_nombre = request.user.get_full_name() or request.user.username
        opp_nombre = tarea.oportunidad.oportunidad if tarea.oportunidad else ''
        if 'participante_add' in data:
            try:
                u = User.objects.get(pk=data['participante_add'])
                tarea.participantes.add(u)
                if u != request.user:
                    crear_notificacion(
                        usuario_destinatario=u,
                        tipo='tarea_opp_asignada',
                        titulo=f'Te agregaron como participante: {tarea.titulo}',
                        mensaje=f'{remitente_nombre} te agregó como participante en la tarea "{tarea.titulo}" de "{opp_nombre}".',
                        oportunidad=tarea.oportunidad,
                        usuario_remitente=request.user,
                        tarea_opp=tarea,
                    )
                return JsonResponse({'success': True})
            except User.DoesNotExist:
                return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
        if 'participante_remove' in data:
            try:
                u = User.objects.get(pk=data['participante_remove'])
                tarea.participantes.remove(u)
                return JsonResponse({'success': True})
            except User.DoesNotExist:
                return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
        if 'observador_add' in data:
            try:
                u = User.objects.get(pk=data['observador_add'])
                tarea.observadores.add(u)
                if u != request.user:
                    crear_notificacion(
                        usuario_destinatario=u,
                        tipo='tarea_opp_asignada',
                        titulo=f'Te agregaron como observador: {tarea.titulo}',
                        mensaje=f'{remitente_nombre} te agregó como observador en la tarea "{tarea.titulo}" de "{opp_nombre}".',
                        oportunidad=tarea.oportunidad,
                        usuario_remitente=request.user,
                        tarea_opp=tarea,
                    )
                return JsonResponse({'success': True})
            except User.DoesNotExist:
                return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
        if 'observador_remove' in data:
            try:
                u = User.objects.get(pk=data['observador_remove'])
                tarea.observadores.remove(u)
                return JsonResponse({'success': True})
            except User.DoesNotExist:
                return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
        tarea.save()
        # Si se marcó como completada, actualizar color de la actividad a verde
        if data.get('estado') == 'completada' and tarea.actividad_calendario_id:
            Actividad.objects.filter(pk=tarea.actividad_calendario_id).update(color='#34C759')
        # Si se cambió la fecha_limite, actualizar la actividad del calendario
        if 'fecha_limite' in data and tarea.actividad_calendario_id and tarea.fecha_limite:
            from datetime import timedelta
            new_start = tarea.fecha_limite
            new_end = tarea.fecha_limite + timedelta(hours=1)
            Actividad.objects.filter(pk=tarea.actividad_calendario_id).update(
                fecha_inicio=new_start, fecha_fin=new_end
            )
        return JsonResponse({'success': True})

    if request.method == 'DELETE':
        tarea.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def api_todas_tareas_opp(request):
    """GET todas las tareas de oportunidad accesibles para el usuario."""
    from django.db.models import IntegerField
    import calendar
    user = request.user

    if is_supervisor(user):
        qs = TareaOportunidad.objects.select_related('oportunidad', 'creado_por', 'responsable').all()
    else:
        qs = TareaOportunidad.objects.select_related('oportunidad', 'creado_por', 'responsable').filter(
            Q(creado_por=user) | Q(responsable=user) |
            Q(participantes=user) | Q(observadores=user)
        ).distinct()

    # Filter by month/year (based on fecha_creacion)
    mes = request.GET.get('mes', '')
    anio = request.GET.get('anio', '')
    if mes and anio:
        try:
            mes_int = int(mes)
            anio_int = int(anio)
            last_day = calendar.monthrange(anio_int, mes_int)[1]
            from datetime import date
            qs = qs.filter(
                fecha_creacion__date__gte=date(anio_int, mes_int, 1),
                fecha_creacion__date__lte=date(anio_int, mes_int, last_day),
            )
        except (ValueError, TypeError):
            pass

    # Filter by vendedor (user IDs)
    vendedores_param = request.GET.get('vendedores', '')
    if vendedores_param:
        try:
            vids = [int(v) for v in vendedores_param.split(',') if v.strip()]
            if vids:
                qs = qs.filter(
                    Q(creado_por_id__in=vids) | Q(responsable_id__in=vids)
                ).distinct()
        except (ValueError, TypeError):
            pass

    qs = qs.annotate(
        prio_order=Case(When(prioridad='alta', then=Value(0)), default=Value(1), output_field=IntegerField())
    ).order_by('prio_order', F('fecha_limite').asc(nulls_last=True), 'fecha_creacion')

    return JsonResponse({'success': True, 'tareas': [_serialize_tarea_opp(t) for t in qs]})


@login_required
def api_tarea_opp_detalle(request, tarea_id):
    """GET detalle completo de una TareaOportunidad (incluye M2M)."""
    tarea = get_object_or_404(TareaOportunidad, pk=tarea_id)

    def user_data(u):
        if not u:
            return None
        return {'id': u.id, 'nombre': u.get_full_name() or u.username}

    data = {
        'id': tarea.id,
        'titulo': tarea.titulo,
        'descripcion': tarea.descripcion,
        'prioridad': tarea.prioridad,
        'estado': tarea.estado,
        'fecha_limite': tarea.fecha_limite.isoformat() if tarea.fecha_limite else None,
        'fecha_creacion': tarea.fecha_creacion.isoformat(),
        'oportunidad_id': tarea.oportunidad_id,
        'oportunidad_nombre': tarea.oportunidad.oportunidad if tarea.oportunidad else '',
        'creado_por_data': user_data(tarea.creado_por),
        'responsable_data': user_data(tarea.responsable),
        'participantes': [user_data(u) for u in tarea.participantes.all()],
        'observadores': [user_data(u) for u in tarea.observadores.all()],
    }
    return JsonResponse(data)


@login_required
def api_tarea_opp_comentarios(request, tarea_id):
    """GET lista de comentarios / POST agregar comentario a una TareaOportunidad."""
    tarea = get_object_or_404(TareaOportunidad, pk=tarea_id)

    if request.method == 'GET':
        comentarios = tarea.comentarios.select_related('autor').all()
        return JsonResponse({
            'success': True,
            'comentarios': [
                {
                    'id': c.id,
                    'usuario': c.autor.get_full_name() or c.autor.username if c.autor else 'Desconocido',
                    'usuario_id': c.autor_id,
                    'contenido': c.contenido,
                    'fecha': c.fecha_creacion.isoformat(),
                }
                for c in comentarios
            ]
        })

    if request.method == 'POST':
        contenido = request.POST.get('contenido', '').strip()
        if not contenido:
            return JsonResponse({'success': False, 'error': 'Comentario vacío'}, status=400)
        ComentarioTareaOpp.objects.create(tarea=tarea, autor=request.user, contenido=contenido)

        # Notificar a creador y responsable (si son distintos al comentarista)
        remitente_nombre = request.user.get_full_name() or request.user.username
        msg_corto = contenido[:100] + ('…' if len(contenido) > 100 else '')
        notif_titulo = f'Comentario en: {tarea.titulo}'
        notif_msg = f'{remitente_nombre}: {msg_corto}'
        destinatarios = set()
        if tarea.creado_por and tarea.creado_por != request.user:
            destinatarios.add(tarea.creado_por)
        if tarea.responsable and tarea.responsable != request.user:
            destinatarios.add(tarea.responsable)
        for dest in destinatarios:
            crear_notificacion(
                usuario_destinatario=dest,
                tipo='tarea_opp_comentario',
                titulo=notif_titulo,
                mensaje=notif_msg,
                oportunidad=tarea.oportunidad,
                usuario_remitente=request.user,
                tarea_opp=tarea,
            )

        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_tarea_opp_comentario_detail(request, tarea_id, comentario_id):
    """PUT editar / DELETE eliminar un comentario de TareaOportunidad."""
    tarea = get_object_or_404(TareaOportunidad, pk=tarea_id)
    comentario = get_object_or_404(ComentarioTareaOpp, pk=comentario_id, tarea=tarea)

    if comentario.autor != request.user and not is_supervisor(request.user):
        return JsonResponse({'error': 'Sin permiso'}, status=403)

    if request.method == 'PUT':
        data = json.loads(request.body)
        contenido = data.get('contenido', '').strip()
        if not contenido:
            return JsonResponse({'success': False, 'error': 'Contenido vacío'}, status=400)
        comentario.contenido = contenido
        comentario.save()
        return JsonResponse({'success': True})

    if request.method == 'DELETE':
        comentario.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


# ═══════════════════════════════════════════════════════
#  MURO EMPRESARIAL
# ═══════════════════════════════════════════════════════

def _muro_post_dict(post, request_user):
    """Serializa un PostMuro a dict para la API."""
    autor = post.autor
    profile = getattr(autor, 'userprofile', None)
    avatar_url = profile.get_avatar_url() if profile else None
    iniciales = profile.iniciales() if profile else (autor.first_name[:1] + autor.last_name[:1]).upper() if autor.first_name else autor.username[:2].upper()

    etiquetados = [
        {'id': u.id, 'nombre': u.get_full_name() or u.username}
        for u in post.etiquetados.all()
    ]
    num_likes = post.likes.count()
    yo_like = post.likes.filter(pk=request_user.pk).exists()
    num_comentarios = post.comentarios.count()

    imagen_url = post.imagen.url if post.imagen else None
    programado_str = post.programado_para.strftime('%Y-%m-%dT%H:%M') if post.programado_para else None

    res = {
        'id': post.id,
        'autor_id': autor.id,
        'autor_nombre': autor.get_full_name() or autor.username,
        'autor_avatar': avatar_url,
        'autor_iniciales': iniciales,
        'contenido': post.contenido,
        'imagen': imagen_url,
        'etiquetados': etiquetados,
        'es_anuncio': post.es_anuncio,
        'programado_para': programado_str,
        'fecha': post.fecha_creacion.strftime('%d %b %Y %H:%M'),
        'editado': post.editado,
        'num_likes': num_likes,
        'yo_like': yo_like,
        'num_comentarios': num_comentarios,
        'es_autor': post.autor_id == request_user.pk,
    }

    # Sobrescribir autor si es un anuncio del sistema (como Empleado del Mes)
    if post.es_anuncio and ("EMPLEADO DEL MES" in post.contenido or "IAMET" in post.contenido):
        res['autor_nombre'] = "IAMET"
        res['autor_avatar'] = "/static/images/apple-touch-icon.png"  # Logo corporativo
        res['autor_iniciales'] = "IA"

    return res


@login_required
def api_muro_posts(request):
    """GET lista de posts / POST crear post."""
    if request.method == 'GET':
        filtro = request.GET.get('filtro', 'todos')
        ahora = timezone.now()
        qs = PostMuro.objects.prefetch_related('likes', 'etiquetados', 'comentarios').select_related('autor').filter(
            Q(programado_para__isnull=True) | Q(programado_para__lte=ahora)
        )
        if filtro == 'anuncios':
            qs = qs.filter(es_anuncio=True)
        elif filtro == 'mios':
            qs = qs.filter(autor=request.user)
        posts = [_muro_post_dict(p, request.user) for p in qs]
        return JsonResponse({'success': True, 'posts': posts})

    if request.method == 'POST':
        contenido = ''
        imagen = None
        etiquetados_ids = []
        es_anuncio = False
        programado_para = None

        if request.content_type and 'multipart' in request.content_type:
            contenido = request.POST.get('contenido', '').strip()
            etiquetados_raw = request.POST.get('etiquetados', '[]')
            try:
                etiquetados_ids = json.loads(etiquetados_raw)
            except Exception:
                etiquetados_ids = []
            es_anuncio = request.POST.get('es_anuncio', 'false') == 'true'
            programado_str = request.POST.get('programado_para', '').strip()
            if programado_str:
                try:
                    from dateutil.parser import parse as parse_dt
                    programado_para = timezone.make_aware(parse_dt(programado_str).replace(tzinfo=None))
                except Exception:
                    programado_para = None
            imagen = request.FILES.get('imagen')
        else:
            data = json.loads(request.body)
            contenido = data.get('contenido', '').strip()
            etiquetados_ids = data.get('etiquetados', [])
            es_anuncio = data.get('es_anuncio', False)
            programado_str = data.get('programado_para', '').strip() if data.get('programado_para') else ''
            if programado_str:
                try:
                    from dateutil.parser import parse as parse_dt
                    programado_para = timezone.make_aware(parse_dt(programado_str).replace(tzinfo=None))
                except Exception:
                    programado_para = None

        if not contenido and not imagen:
            return JsonResponse({'success': False, 'error': 'El contenido no puede estar vacío'}, status=400)

        # Solo supervisores pueden publicar anuncios o programar posts
        if es_anuncio and not is_supervisor(request.user):
            es_anuncio = False
        if programado_para and not is_supervisor(request.user):
            programado_para = None

        post = PostMuro.objects.create(
            autor=request.user,
            contenido=contenido,
            imagen=imagen,
            es_anuncio=es_anuncio,
            programado_para=programado_para,
        )
        if etiquetados_ids:
            usuarios_etiq = User.objects.filter(pk__in=etiquetados_ids)
            post.etiquetados.set(usuarios_etiq)
            # Notificar a etiquetados
            remitente_nombre = request.user.get_full_name() or request.user.username
            for u in usuarios_etiq:
                crear_notificacion(
                    usuario_destinatario=u,
                    tipo='muro_mencion',
                    titulo=f'{remitente_nombre} te etiquetó en el muro',
                    mensaje=contenido[:120],
                    usuario_remitente=request.user,
                )

        if es_anuncio:
            # Notificar a todos los usuarios del sistema
            todos = User.objects.filter(is_active=True).exclude(pk=request.user.pk)
            remitente_nombre = request.user.get_full_name() or request.user.username
            for u in todos:
                crear_notificacion(
                    usuario_destinatario=u,
                    tipo='muro_post',
                    titulo=f'Anuncio de {remitente_nombre}',
                    mensaje=contenido[:120],
                    usuario_remitente=request.user,
                )

        return JsonResponse({'success': True, 'post': _muro_post_dict(post, request.user)})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_muro_post_detail(request, post_id):
    """PUT editar / DELETE eliminar un post del muro."""
    post = get_object_or_404(PostMuro, pk=post_id)

    if post.autor != request.user and not is_supervisor(request.user):
        return JsonResponse({'error': 'Sin permiso'}, status=403)

    if request.method == 'PUT':
        data = json.loads(request.body)
        contenido = data.get('contenido', '').strip()
        if not contenido:
            return JsonResponse({'success': False, 'error': 'Contenido vacío'}, status=400)
        post.contenido = contenido
        post.editado = True
        post.save()
        return JsonResponse({'success': True, 'post': _muro_post_dict(post, request.user)})

    if request.method == 'DELETE':
        post.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_muro_like(request, post_id):
    """POST toggle like en un post."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    post = get_object_or_404(PostMuro, pk=post_id)
    if post.likes.filter(pk=request.user.pk).exists():
        post.likes.remove(request.user)
        yo_like = False
    else:
        post.likes.add(request.user)
        yo_like = True
    return JsonResponse({'success': True, 'num_likes': post.likes.count(), 'yo_like': yo_like})


@login_required
def api_muro_comentarios(request, post_id):
    """GET lista de comentarios / POST crear comentario."""
    post = get_object_or_404(PostMuro, pk=post_id)

    if request.method == 'GET':
        comentarios = []
        for c in post.comentarios.select_related('autor').all():
            autor = c.autor
            profile = getattr(autor, 'userprofile', None)
            avatar_url = profile.get_avatar_url() if profile else None
            iniciales = profile.iniciales() if profile else (autor.first_name[:1] + autor.last_name[:1]).upper() if autor.first_name else autor.username[:2].upper()
            comentarios.append({
                'id': c.id,
                'autor_id': autor.id,
                'autor_nombre': autor.get_full_name() or autor.username,
                'autor_avatar': avatar_url,
                'autor_iniciales': iniciales,
                'contenido': c.contenido,
                'fecha': c.fecha_creacion.strftime('%d %b %Y %H:%M'),
                'editado': c.editado,
                'es_autor': c.autor_id == request.user.pk,
            })
        return JsonResponse({'success': True, 'comentarios': comentarios})

    if request.method == 'POST':
        data = json.loads(request.body)
        contenido = data.get('contenido', '').strip()
        if not contenido:
            return JsonResponse({'success': False, 'error': 'Contenido vacío'}, status=400)
        comentario = ComentarioMuro.objects.create(
            post=post,
            autor=request.user,
            contenido=contenido,
        )
        # Notificar al autor del post si es distinto
        if post.autor != request.user:
            remitente_nombre = request.user.get_full_name() or request.user.username
            crear_notificacion(
                usuario_destinatario=post.autor,
                tipo='muro_mencion',
                titulo=f'{remitente_nombre} comentó tu post',
                mensaje=contenido[:120],
                usuario_remitente=request.user,
            )
        autor = comentario.autor
        profile = getattr(autor, 'userprofile', None)
        avatar_url = profile.get_avatar_url() if profile else None
        iniciales = profile.iniciales() if profile else (autor.first_name[:1] + autor.last_name[:1]).upper() if autor.first_name else autor.username[:2].upper()
        return JsonResponse({
            'success': True,
            'comentario': {
                'id': comentario.id,
                'autor_id': autor.id,
                'autor_nombre': autor.get_full_name() or autor.username,
                'autor_avatar': avatar_url,
                'autor_iniciales': iniciales,
                'contenido': comentario.contenido,
                'fecha': comentario.fecha_creacion.strftime('%d %b %Y %H:%M'),
                'editado': False,
                'es_autor': True,
            }
        })

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def api_muro_comentario_detail(request, comentario_id):
    """PUT editar / DELETE eliminar un comentario del muro."""
    comentario = get_object_or_404(ComentarioMuro, pk=comentario_id)

    if comentario.autor != request.user and not is_supervisor(request.user):
        return JsonResponse({'error': 'Sin permiso'}, status=403)

    if request.method == 'PUT':
        data = json.loads(request.body)
        contenido = data.get('contenido', '').strip()
        if not contenido:
            return JsonResponse({'success': False, 'error': 'Contenido vacío'}, status=400)
        comentario.contenido = contenido
        comentario.editado = True
        comentario.save()
        return JsonResponse({'success': True})

    if request.method == 'DELETE':
        comentario.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


# ═══ APIS PARA CONTROL DE ASISTENCIA Y EFICIENCIA ═══


@login_required
@csrf_exempt
def api_jornada_estado(request):
    """Retorna el estado actual de la jornada del usuario."""
    hoy = timezone.localdate()
    # Buscar jornada de hoy que no haya terminado o la más reciente terminada hoy
    jornada = AsistenciaJornada.objects.filter(usuario=request.user, fecha=hoy).first()
    
    if not jornada:
        return JsonResponse({'activo': False})
    
    # Calcular segundos laborados acumulados
    segundos = jornada.segundos_laborados
    if not jornada.pausado and not jornada.hora_fin:
        # Si está activo, sumar el tramo actual
        delta = timezone.now() - jornada.hora_inicio
        segundos += int(delta.total_seconds())

    return JsonResponse({
        'activo': True,
        'inicio': jornada.hora_inicio.isoformat() if jornada.hora_inicio else None,
        'fin': jornada.hora_fin.isoformat() if jornada.hora_fin else None,
        'pausado': jornada.pausado,
        'segundos': segundos,
        'eficiencia': float(jornada.eficiencia_dia)
    })

@login_required
@csrf_exempt
def api_jornada_iniciar(request):
    """Inicia la jornada laboral si no existe una hoy."""
    hoy = timezone.localdate()
    jornada, created = AsistenciaJornada.objects.get_or_create(
        usuario=request.user, 
        fecha=hoy,
        defaults={'hora_inicio': timezone.now()}
    )
    # Si ya existía y estaba terminada, podemos decidir si reabrirla o no.
    # Por ahora permitimos re-iniciar si estaba terminada hoy.
    if not created and jornada.hora_fin:
        jornada.hora_fin = None
        jornada.hora_inicio = timezone.now()
        jornada.save()

    return JsonResponse({'success': True, 'jornada_id': jornada.id})

@login_required
@csrf_exempt
def api_jornada_pausar(request):
    """Alterna el estado de pausa de la jornada laboral."""
    hoy = timezone.localdate()
    jornada = AsistenciaJornada.objects.filter(usuario=request.user, fecha=hoy, hora_fin__isnull=True).first()
    if not jornada:
        return JsonResponse({'error': 'No hay jornada activa'}, status=400)
    
    ahora = timezone.now()
    if not jornada.pausado:
        # Pausar: guardar segundos del tramo que termina ahora
        delta = ahora - jornada.hora_inicio
        jornada.segundos_laborados += int(delta.total_seconds())
        jornada.pausado = True
        jornada.ultima_pausa = ahora
    else:
        # Reanudar: actualizar hora_inicio para el nuevo tramo activo
        jornada.pausado = False
        jornada.hora_inicio = ahora
        if jornada.ultima_pausa:
            delta_p = ahora - jornada.ultima_pausa
            jornada.segundos_pausa += int(delta_p.total_seconds())
    
    jornada.save()
    return JsonResponse({'success': True, 'pausado': jornada.pausado})

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def api_jornada_terminar(request):
    """Termina la jornada laboral y calcula la eficiencia del día."""
    hoy = timezone.localdate()
    jornada = AsistenciaJornada.objects.filter(usuario=request.user, fecha=hoy, hora_fin__isnull=True).first()
    if not jornada:
        return JsonResponse({'error': 'No hay jornada activa que terminar'}, status=400)
    
    ahora = timezone.now()
    if not jornada.pausado:
        delta = ahora - jornada.hora_inicio
        jornada.segundos_laborados += int(delta.total_seconds())
    
    jornada.hora_fin = ahora
    
    # --- CÁLCULO DE EFICIENCIA (Algoritmo solicitado) ---
    # Prioridades: 1. Tareas, 2. Actividades, 3. Ventas cobradas
    
    # Tareas (Highest Priority)
    tareas = Tarea.objects.filter(asignado_a=request.user).exclude(estado='cancelada')
    # Tareas que vencen hoy o fueron completadas hoy
    tareas_totales = tareas.filter(Q(fecha_limite__date=hoy) | Q(estado='completada', fecha_completada__date=hoy)).count()
    tareas_completadas_hoy = tareas.filter(estado='completada', fecha_completada__date=hoy).count()
    tareas_vencidas = tareas.filter(fecha_limite__lt=ahora).exclude(estado='completada').count()
    
    # Actividades / Tareas de Oportunidad (Medium Priority)
    acts = TareaOportunidad.objects.filter(responsable=request.user)
    # Como TareaOportunidad no tiene fecha_completada, usamos la lógica de estado y fecha_limite
    act_totales = acts.filter(Q(fecha_limite__date=hoy) | Q(estado='completada', fecha_limite__date=hoy)).count()
    act_completadas_hoy = acts.filter(estado='completada', fecha_limite__date=hoy).count()
    act_vencidas = acts.filter(fecha_limite__lt=ahora, estado='pendiente').count()
    
    # Oportunidades Cobradas (High Impact / Bonus)
    # Nota: TodoItem SI tiene fecha_actualizacion
    opps_cobradas_hoy = TodoItem.objects.filter(usuario=request.user, estado_crm='pagada', fecha_actualizacion__date=hoy).count()

    # Cálculo por puntos:
    puntos_obtenidos = 0
    puntos_posibles = 0
    
    if tareas_totales > 0:
        puntos_posibles += (tareas_totales * 10)
        puntos_obtenidos += (tareas_completadas_hoy * 10)
        puntos_obtenidos -= (tareas_vencidas * 5) # Penalización
        
    if act_totales > 0:
        puntos_posibles += (act_totales * 6)
        puntos_obtenidos += (act_completadas_hoy * 6)
        puntos_obtenidos -= (act_vencidas * 3) # Penalización
    
    # Bonus por cada venta cerrada hoy
    puntos_obtenidos += (opps_cobradas_hoy * 20)
    
    eficiencia = 0
    if puntos_posibles > 0:
        eficiencia = (puntos_obtenidos / puntos_posibles) * 100
    elif opps_cobradas_hoy > 0:
        eficiencia = 100
    
    # Limitar entre 0 y 100
    eficiencia = max(0, min(100, eficiencia))
    
    jornada.eficiencia_dia = Decimal(str(round(eficiencia, 2)))
    jornada.save()
    
    # Actualizar registro mensual
    em, created = EficienciaMensual.objects.get_or_create(
        usuario=request.user, 
        mes=hoy.month, 
        anio=hoy.year,
        defaults={'promedio_eficiencia': eficiencia}
    )
    
    # Recalcular promedio mensual basado en todas las jornadas del mes
    jornadas_mes = AsistenciaJornada.objects.filter(
        usuario=request.user, 
        fecha__month=hoy.month, 
        fecha__year=hoy.year, 
        hora_fin__isnull=False
    )
    stats_mes = jornadas_mes.aggregate(avg_ef=models.Avg('eficiencia_dia'))
    em.promedio_eficiencia = stats_mes['avg_ef'] or eficiencia
    
    # Acumular hitos
    em.tareas_completadas += tareas_completadas_hoy
    em.actividades_completadas += act_completadas_hoy
    em.oportunidades_cobradas += opps_cobradas_hoy
    em.save()

    # --- NOTIFICACIÓN PARA ADMINISTRADORES (Nivel de Alerta) ---
    horas_totales = jornada.segundos_laborados / 3600
    if eficiencia < 80 or horas_totales < 9:
        admins = User.objects.filter(is_superuser=True)
        nombre_usuario = request.user.get_full_name() or request.user.username
        
        razones = []
        if eficiencia < 80:
            razones.append(f"Eficiencia baja ({round(eficiencia, 1)}%)")
        if horas_totales < 9:
            razones.append(f"Jornada incompleta ({round(horas_totales, 1)} horas)")
        
        # Detalle estructurado para el administrador
        mensaje_html = f"<div class='notif-alerta-rendimiento'>"
        mensaje_html += f"<p style='margin:0 0 8px;'><b>Alertas:</b> {', '.join(razones)}</p>"
        
        tareas_pendientes = tareas.exclude(estado='completada').order_by('fecha_limite')[:4]
        if tareas_pendientes:
            mensaje_html += "<div style='background:rgba(0,0,0,0.03); border-radius:8px; padding:8px; margin-bottom:8px;'>"
            mensaje_html += "<b style='font-size:0.75rem; color:#4B5563; text-transform:uppercase;'>Tareas no completadas:</b>"
            mensaje_html += "<ul style='margin:5px 0 0; padding-left:15px; font-size:0.8rem;'>"
            for t in tareas_pendientes:
                fecha_venc = t.fecha_limite.strftime('%d/%m/%Y') if t.fecha_limite else "Sin fecha"
                mensaje_html += f"<li style='margin-bottom:4px;'><a href='#' onclick='event.stopPropagation(); window.crmTaskVerDetalle({t.id})' style='color:#007AFF; text-decoration:none; font-weight:600;'>{t.titulo}</a> <span style='color:#EF4444; font-size:0.7rem;'>(Venció: {fecha_venc})</span></li>"
            mensaje_html += "</ul></div>"
            
        acts_pendientes = acts.exclude(estado='completada').order_by('fecha_limite')[:4]
        if acts_pendientes:
            mensaje_html += "<div style='background:rgba(0,122,255,0.03); border-radius:8px; padding:8px;'>"
            mensaje_html += "<b style='font-size:0.75rem; color:#4B5563; text-transform:uppercase;'>Actividades en Oportunidades:</b>"
            mensaje_html += "<ul style='margin:5px 0 0; padding-left:15px; font-size:0.8rem;'>"
            for a in acts_pendientes:
                fecha_venc = a.fecha_limite.strftime('%d/%m/%Y') if a.fecha_limite else "Sin fecha"
                opp_id = a.oportunidad.id
                mensaje_html += f"<li style='margin-bottom:4px;'><a href='#' onclick='event.stopPropagation(); window.openDetalle({opp_id})' style='color:#007AFF; text-decoration:none; font-weight:600;'>{a.titulo}</a> <span style='color:#EF4444; font-size:0.7rem;'>(Venció: {fecha_venc})</span></li>"
            mensaje_html += "</ul></div>"
        
        mensaje_html += "</div>"

        for admin in admins:
            Notificacion.objects.create(
                usuario_destinatario=admin,
                usuario_remitente=request.user,
                tipo='rendimiento_bajo',
                titulo=f"Alerta Rendimiento: {nombre_usuario}",
                mensaje=mensaje_html
            )

    return JsonResponse({
        'success': True, 
        'eficiencia': float(jornada.eficiencia_dia),
        'segundos': jornada.segundos_laborados
    })

@login_required
def api_verificar_empleado_mes(request):
    """
    Verifica si ha concluido el mes previo y publica al ganador en el muro.
    Configurado para iniciar el 1 de Abril de 2026 a las 8:00 AM (Tijuana).
    """
    ahora = timezone.now()
    ahora_tj = convert_to_tijuana_time(ahora)
    
    # 1. Bloqueo hasta el inicio oficial: 1 de Abril 2026, 8 AM Tijuana
    inicio_oficial = datetime(2026, 4, 1, 8, 0, 0, tzinfo=ZoneInfo("America/Tijuana"))
    if ahora_tj < inicio_oficial:
        return JsonResponse({
            'status': 'scheduled', 
            'message': 'El primer anuncio oficial está programado para el 1 de Abril a las 8:00 AM (Tijuana).'
        })

    # 2. Regla de publicación mensual: Solo el día 1, a partir de las 8:00 AM
    if ahora_tj.day != 1:
        return JsonResponse({'status': 'not_ready', 'message': 'Las publicaciones automáticas ocurren el día 1 de cada mes.'})
        
    if ahora_tj.hour < 8:
        return JsonResponse({'status': 'too_morning', 'message': 'El anuncio se publicará hoy a las 8:00 AM.'})

    # Revisar mes anterior (ej: si es 1 de Abril, revisa Marzo)
    mes_target_date = ahora - relativedelta(months=1)
    m = mes_target_date.month
    y = mes_target_date.year
    
    # Si ya se publicó este mes, salir
    if EficienciaMensual.objects.filter(mes=m, anio=y, anuncio_publicado=True).exists():
        return JsonResponse({'status': 'period_already_announced'})
    
    # Obtener el mejor del mes pasado
    ganador_em = EficienciaMensual.objects.filter(mes=m, anio=y).order_by('-promedio_eficiencia').first()
    
    if ganador_em and ganador_em.promedio_eficiencia > 0:
        ganador_em.empleado_del_mes = True
        ganador_em.anuncio_publicado = True
        ganador_em.save()
        
        # Muro Post (Anuncio automático)
        nombre_ganador = ganador_em.usuario.get_full_name() or ganador_em.usuario.username
        mes_nombre = mes_target_date.strftime('%B %Y')
        
        mensaje = f"¡Es un honor anunciar que <b>@{ganador_em.usuario.username} ({nombre_ganador})</b> ha sido seleccionado como el empleado del mes de <b>{mes_nombre}</b>! 🎉\n\n"
        mensaje += f"Su compromiso ha sido pieza clave en nuestro éxito:\n\n"
        mensaje += f"✨ <b>Eficiencia General:</b> {ganador_em.promedio_eficiencia}%\n"
        mensaje += f"✅ <b>Tareas resueltas:</b> {ganador_em.tareas_completadas}\n"
        mensaje += f"🛠 <b>Actividades finalizadas:</b> {ganador_em.actividades_completadas}\n"
        mensaje += f"💰 <b>Ventas cobradas:</b> {ganador_em.oportunidades_cobradas}\n\n"
        mensaje += f"¡Gracias por dar siempre el 100%! 🚀🔥"
        
        # Intentar que el autor sea un Admin, pero se mostrará como IAMET en el muro
        admin_user = User.objects.filter(is_superuser=True).first() or request.user
        
        # Obtener la foto del ganador para ponerla como imagen principal del post
        ganador_profile = getattr(ganador_em.usuario, 'userprofile', None)
        foto_ganador = ganador_profile.avatar if ganador_profile and ganador_profile.avatar else None

        PostMuro.objects.create(
            autor=admin_user,
            contenido=mensaje,
            es_anuncio=True,
            imagen=foto_ganador
        )
        return JsonResponse({'status': 'announced', 'winner': ganador_em.usuario.username})
    
    return JsonResponse({'status': 'no_eligible_data'})

@login_required
@require_http_methods(["POST"])
def api_solicitar_cambio_perfil(request):
    """
    Recibe los datos propuestos por el usuario. 
    Idioma y Tema se aplican de inmediato. 
    Si es Admin, todo se aplica de inmediato.
    Si es usuario normal y cambia Nombre/Email/Foto, se crea una solicitud de aprobación.
    """
    try:
        user = request.user
        first_name = request.POST.get('first_name', user.first_name)
        last_name = request.POST.get('last_name', user.last_name)
        email = request.POST.get('email', user.email)
        language = request.POST.get('language', user.userprofile.language)
        theme = request.POST.get('theme', user.userprofile.theme)
        avatar_file = request.FILES.get('avatar')

        # 1. Aplicar cambios inmediatos (Idioma y Tema)
        profile = user.userprofile
        profile.language = language
        profile.theme = theme
        profile.save()

        # 2. Si es superusuario, aplicar todo de inmediato
        if user.is_superuser:
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.save()
            if avatar_file:
                profile.avatar = avatar_file
                profile.save()
            return JsonResponse({'success': True, 'message': 'Información actualizada correctamente.', 'immediate': True})

        # 3. Si es usuario normal, verificar cambios sensibles
        cambio_sensible = (
            first_name != user.first_name or 
            last_name != user.last_name or 
            email != user.email or 
            avatar_file
        )

        if cambio_sensible:
            # Crear solicitud
            solicitud = SolicitudCambioPerfil.objects.create(
                solicitante=user,
                first_name=first_name,
                last_name=last_name,
                email=email,
                language=language or 'es',
                avatar=avatar_file,
                old_first_name=user.first_name,
                old_last_name=user.last_name,
                old_email=user.email,
                old_language=getattr(user.userprofile, 'language', 'es')
            )

            # Notificar a administradores (supervisores)
            admins = User.objects.filter(is_superuser=True)
            if not admins.exists():
                 admins = User.objects.all()[:1] 

            for admin in admins:
                crear_notificacion(
                    usuario_destinatario=admin,
                    tipo='solicitud_cambio_perfil',
                    titulo='Solicitud de Cambio de Perfil',
                    mensaje=f'El usuario {user.username} ha solicitado cambiar su información sensible. Requiere autorización.',
                    usuario_remitente=user,
                    solicitud_perfil=solicitud
                )

            return JsonResponse({
                'success': True, 
                'message': 'Preferencias guardadas. Los cambios en nombre, email o foto han sido enviados para aprobación administrativa.',
                'immediate': True # Recargar para aplicar idioma/tema si cambiaron
            })

        return JsonResponse({'success': True, 'message': 'Preferencias actualizadas correctamente.', 'immediate': True})

    except Exception as e:
        logger.error(f"Error en api_solicitar_cambio_perfil: {e}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
@user_passes_test(is_supervisor)
@require_http_methods(["POST"])
def api_procesar_solicitud_perfil(request, solicitud_id):
    """
    Permite que un administrador apruebe o rechace una solicitud de cambio de perfil.
    """
    try:
        solicitud = get_object_or_404(SolicitudCambioPerfil, id=solicitud_id)
        if solicitud.procesada:
            return JsonResponse({'success': False, 'message': 'Esta solicitud ya ha sido procesada.'}, status=400)

        data = json.loads(request.body)
        aprobado = data.get('aprobado', False)
        comentario = data.get('comentario', '')

        solicitud.procesada = True
        solicitud.aprobada = aprobado
        solicitud.procesado_por = request.user
        solicitud.fecha_procesamiento = timezone.now()
        solicitud.comentario_admin = comentario
        solicitud.save()

        user = solicitud.solicitante
        if aprobado:
            # Aplicar cambios al usuario
            user.first_name = solicitud.first_name
            user.last_name = solicitud.last_name
            user.email = solicitud.email
            user.save()

            # Aplicar cambios al perfil
            profile = user.userprofile
            profile.language = solicitud.language
            if solicitud.avatar:
                profile.avatar = solicitud.avatar
            profile.save()

            # Notificar al usuario que su solicitud fue aprobada
            crear_notificacion(
                usuario_destinatario=user,
                tipo='sistema',
                titulo='Perfil Actualizado',
                mensaje=f'Tu solicitud de cambio de perfil ha sido aprobada por {request.user.username}.'
            )
        else:
            # Notificar al usuario que su solicitud fue rechazada
            crear_notificacion(
                usuario_destinatario=user,
                tipo='sistema',
                titulo='Solicitud de Perfil Rechazada',
                mensaje=f'Tu solicitud de cambio de perfil ha sido rechazada. Motivo: {comentario}'
            )

        return JsonResponse({'success': True, 'message': 'Solicitud procesada correctamente.'})
    except Exception as e:
        logger.error(f"Error en api_procesar_solicitud_perfil: {e}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ─────────────────────────────────────────────────────────────────────────────
# API: Vinculo Proyecto <-> Oportunidad
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def api_oportunidad_proyectos(request, opp_id):
    """GET: devuelve confirmados y sugerencias de proyectos para una oportunidad."""
    from app.models import ProyectoOportunidadLink, TodoItem
    try:
        opp = TodoItem.objects.get(pk=opp_id)
    except TodoItem.DoesNotExist:
        return JsonResponse({'error': 'Oportunidad no encontrada'}, status=404)

    confirmados_qs = ProyectoOportunidadLink.objects.filter(
        oportunidad=opp, confirmado=True, rechazado=False
    ).select_related('proyecto')
    sugerencias_qs = ProyectoOportunidadLink.objects.filter(
        oportunidad=opp, confirmado=False, rechazado=False
    ).select_related('proyecto').order_by('-score')

    confirmados = [
        {
            'link_id': lnk.id,
            'id': lnk.proyecto.id,
            'nombre': lnk.proyecto.nombre,
            'score': lnk.score,
            'bitrix_group_id': lnk.proyecto.bitrix_group_id,
        }
        for lnk in confirmados_qs
    ]
    sugerencias = [
        {
            'link_id': lnk.id,
            'id': lnk.proyecto.id,
            'nombre': lnk.proyecto.nombre,
            'score': lnk.score,
            'bitrix_group_id': lnk.proyecto.bitrix_group_id,
        }
        for lnk in sugerencias_qs
    ]
    return JsonResponse({'confirmados': confirmados, 'sugerencias': sugerencias})


@login_required
def api_oportunidad_proyectos_accion(request, opp_id, link_id):
    """POST {"accion": "confirmar"|"rechazar"}: actualiza el link."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Metodo no permitido'}, status=405)

    from app.models import ProyectoOportunidadLink
    try:
        lnk = ProyectoOportunidadLink.objects.get(pk=link_id, oportunidad_id=opp_id)
    except ProyectoOportunidadLink.DoesNotExist:
        return JsonResponse({'error': 'Vinculo no encontrado'}, status=404)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'JSON invalido'}, status=400)

    accion = data.get('accion', '')
    if accion == 'confirmar':
        lnk.confirmado = True
        lnk.rechazado = False
        lnk.vinculado_por = request.user
        lnk.save()
    elif accion == 'rechazar':
        lnk.rechazado = True
        lnk.confirmado = False
        lnk.save()
    else:
        return JsonResponse({'error': 'Accion invalida'}, status=400)

    return JsonResponse({'success': True})


@login_required
def api_oportunidad_proyectos_buscar(request, opp_id):
    """
    GET ?q=texto : busca proyectos por nombre, devuelve top 10 no ligados.
    POST {"proyecto_id": X} : crea un vinculo confirmado manualmente (score=100).
    """
    from app.models import ProyectoOportunidadLink, TodoItem, Proyecto

    try:
        opp = TodoItem.objects.get(pk=opp_id)
    except TodoItem.DoesNotExist:
        return JsonResponse({'error': 'Oportunidad no encontrada'}, status=404)

    if request.method == 'GET':
        q = request.GET.get('q', '').strip()
        if not q:
            return JsonResponse({'results': []})

        ya_ligados = set(
            ProyectoOportunidadLink.objects.filter(oportunidad=opp).values_list('proyecto_id', flat=True)
        )
        proyectos = Proyecto.objects.filter(nombre__icontains=q).exclude(id__in=ya_ligados)[:10]

        try:
            from rapidfuzz import fuzz
            opp_nombre = opp.oportunidad or ''
        except ImportError:
            fuzz = None
            opp_nombre = ''

        results = []
        for p in proyectos:
            score = fuzz.token_set_ratio(opp_nombre, p.nombre) if fuzz else 0
            results.append({
                'id': p.id,
                'nombre': p.nombre,
                'bitrix_group_id': p.bitrix_group_id,
                'score': score,
            })
        results.sort(key=lambda x: x['score'], reverse=True)
        return JsonResponse({'results': results})

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({'error': 'JSON invalido'}, status=400)

        proyecto_id = data.get('proyecto_id')
        if not proyecto_id:
            return JsonResponse({'error': 'proyecto_id requerido'}, status=400)

        try:
            proyecto = Proyecto.objects.get(pk=proyecto_id)
        except Proyecto.DoesNotExist:
            return JsonResponse({'error': 'Proyecto no encontrado'}, status=404)

        lnk, _ = ProyectoOportunidadLink.objects.update_or_create(
            proyecto=proyecto,
            oportunidad=opp,
            defaults={
                'score': 100.0,
                'confirmado': True,
                'rechazado': False,
                'vinculado_por': request.user,
            },
        )
        return JsonResponse({'success': True, 'link_id': lnk.id})

    return JsonResponse({'error': 'Metodo no permitido'}, status=405)
