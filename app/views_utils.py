# ----------------------------------------------------------------------
# views_utils.py — Shared utilities and helpers (no URL routing).
# ----------------------------------------------------------------------

import json
import logging
import requests
import mimetypes
import os
from django.conf import settings
import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.db import models
from .models import TodoItem, Cliente, Cotizacion, DetalleCotizacion, UserProfile, Contacto, PendingFileUpload, OportunidadProyecto, Volumetria, DetalleVolumetria, CatalogoCableado, OportunidadActividad, OportunidadComentario, OportunidadArchivo, OportunidadEstado, Notificacion, Proyecto, ProyectoComentario, ProyectoArchivo, Tarea, TareaComentario, TareaArchivo, Actividad, CarpetaProyecto, ArchivoProyecto, CompartirArchivo, IntercambioNavidad, ParticipanteIntercambio, HistorialIntercambio, SolicitudAccesoProyecto, ArchivoFacturacion, CarpetaOportunidad, ArchivoOportunidad, MensajeOportunidad, TareaOportunidad, ComentarioTareaOpp, PostMuro, ComentarioMuro, ProductoOportunidad, AsistenciaJornada, EficienciaMensual, SolicitudCambioPerfil, ProgramacionActividad
from . import views_exportar
from .views_tarea_comentarios import api_comentarios_tarea, api_agregar_comentario_tarea, api_editar_comentario_tarea, api_eliminar_comentario_tarea
from .forms import VentaForm, VentaFilterForm, CotizacionForm, ClienteForm, OportunidadModalForm, NuevaOportunidadForm
from django.db.models import Sum, Count, F, Q, Case, When, Value
from django.db.models.functions import Upper, Coalesce
from django.db.models import Value
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from django.utils import timezone
from decimal import Decimal
import decimal
from django.utils.html import json_script

# Helper function to detect lost opportunities

def is_lost_opportunity(stage_id):
    """
    Detecta si una oportunidad está marcada como "Cerrado Perdido" en Bitrix24
    Maneja múltiples formatos de stage_id:
    - '2': Formato simple
    - 'C2:11': Formato con prefijo numérico (perdido)
    - 'C2:LOSE': Formato "Cerrar negociación" -> "Cerrado perdido"
    - 'LOSE': Formato texto
    
    NOTA: C2:NEW no es perdido, es nuevo
    """
    if not stage_id:
        return False
    
    stage_id = str(stage_id)
    
    # Casos específicos de estado perdido
    if stage_id == '2':  # Formato simple
        return True
    elif stage_id == 'LOSE':  # Formato texto
        return True
    elif stage_id == 'C2:LOSE':  # Formato "Cerrar negociación" -> "Cerrado perdido"
        return True
    elif stage_id.startswith('C2:'):
        # Solo considerar perdido si después de C2: hay un número
        # C2:NEW = nuevo, C2:11 = perdido, C2:LOSE = perdido (ya manejado arriba)
        try:
            suffix = stage_id[3:]  # Obtener la parte después de 'C2:'
            int(suffix)  # Si es un número, es perdido
            return True
        except ValueError:
            # Si no es un número (ej: C2:NEW), no es perdido
            return False
    
    return False


def convert_to_tijuana_time(utc_datetime):
    """Convierte una fecha UTC a tiempo de Tijuana"""
    try:
        # Usar zoneinfo que es parte de la biblioteca estándar de Python
        tijuana_tz = ZoneInfo('America/Tijuana')
        if utc_datetime.tzinfo is None:
            # Si no tiene timezone info, asumir UTC
            utc_datetime = timezone.make_aware(utc_datetime, timezone.utc)
        return utc_datetime.astimezone(tijuana_tz)
    except Exception:
        return utc_datetime


def is_supervisor(user):
    return user.is_superuser or user.groups.filter(name='Supervisores').exists()


def is_ingeniero(user):
    try:
        from app.models import UserProfile
        profile = UserProfile.objects.get(user=user)
        return getattr(profile, 'rol', 'vendedor') == 'ingeniero'
    except Exception:
        return False


def is_administrador(user):
    """Devuelve True si el usuario tiene rol 'administrador' en UserProfile."""
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    try:
        from app.models import UserProfile
        profile = UserProfile.objects.get(user=user)
        return getattr(profile, 'rol', '') == 'administrador'
    except Exception:
        return False


def is_engineer(user):
    return user.groups.filter(name='Ingenieros').exists()


def _get_display_for_value(value, choices_list):
    return dict(choices_list).get(value, value)


def calcular_facturado_por_vendedor(datos_json):
    """
    Dado el datos_json de ArchivoFacturacion {cliente_name: monto},
    retorna {user_id: monto_facturado} mapeando clientes a sus vendedores.
    """
    from collections import defaultdict
    facturado_por_vendedor = defaultdict(Decimal)

    for cliente_name, monto_str in datos_json.items():
        monto = Decimal(str(monto_str))
        # Buscar cliente en BD por nombre
        cliente_match = Cliente.objects.filter(nombre_empresa__iexact=cliente_name).first()
        if not cliente_match:
            cliente_match = Cliente.objects.filter(nombre_empresa__icontains=cliente_name).first()
        if not cliente_match:
            # Intentar match parcial inverso
            for cliente_obj in Cliente.objects.all():
                if cliente_obj.nombre_empresa and cliente_obj.nombre_empresa.upper() in cliente_name.upper():
                    cliente_match = cliente_obj
                    break

        if cliente_match:
            # El dueño es el usuario con más oportunidades de ese cliente
            owner = (
                TodoItem.objects
                .filter(cliente=cliente_match)
                .values('usuario_id')
                .annotate(count=Count('id'))
                .order_by('-count')
                .first()
            )
            if owner and owner['usuario_id']:
                facturado_por_vendedor[owner['usuario_id']] += monto

    return dict(facturado_por_vendedor)


def get_safe_file_info(archivo):
    """Helper function para obtener información segura de archivos"""
    try:
        tamaño = archivo.get_tamaño_legible() if hasattr(archivo, 'get_tamaño_legible') else f"{archivo.tamaño} bytes"
    except:
        tamaño = f"{archivo.tamaño} bytes"
    
    try:
        icono = archivo.get_icono() if hasattr(archivo, 'get_icono') else '📎'
    except:
        icono = '📎'
    
    return {
        'id': archivo.id,
        'nombre': archivo.nombre_original,
        'tamaño': tamaño,
        'icono': icono,
        'url': archivo.archivo.url,
        'tipo': archivo.tipo_contenido
    }


def supervisor_required(view_func):
    decorated_view_func = login_required(user_passes_test(is_supervisor)(view_func))
    return decorated_view_func


def _exportar_estadisticas_excel(request, oportunidades, cotizaciones, titulo_filtro):
    """Genera un archivo Excel con las estadísticas de usuarios."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Sum, Count

    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas Usuarios"

    # Estilos
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='0A84FF', end_color='0A84FF', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D1D6'),
        right=Side(style='thin', color='D1D1D6'),
        top=Side(style='thin', color='D1D1D6'),
        bottom=Side(style='thin', color='D1D1D6'),
    )
    money_format = '#,##0.00'

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = f'Reporte de Estadísticas de Usuarios — {titulo_filtro}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='0A84FF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = ['Usuario', 'Oportunidades', 'Cotizaciones PDF', 'Total Vendido (100%)', 'Pipeline', 'Cliente Top']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos
    usuarios = User.objects.filter(oportunidades__isnull=False).distinct()
    row = 4
    for usuario in usuarios:
        oportunidades_usuario = oportunidades.filter(usuario=usuario)
        num_oportunidades = oportunidades_usuario.count()
        if num_oportunidades == 0:
            continue
        total_vendido = oportunidades_usuario.filter(probabilidad_cierre=100).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        total_pipeline = oportunidades_usuario.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
        num_cotizaciones = cotizaciones.filter(created_by=usuario).count()
        cliente_top = oportunidades_usuario.values('cliente__nombre_empresa').annotate(
            total_cliente=Sum('monto'), num_ops=Count('id')
        ).order_by('-total_cliente').first()
        cliente_nombre = cliente_top['cliente__nombre_empresa'] if cliente_top else '-'

        nombre = usuario.get_full_name() or usuario.username

        ws.cell(row=row, column=1, value=nombre).border = thin_border
        ws.cell(row=row, column=2, value=num_oportunidades).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=num_cotizaciones).border = thin_border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        cell_vendido = ws.cell(row=row, column=4, value=float(total_vendido))
        cell_vendido.number_format = money_format
        cell_vendido.border = thin_border
        cell_pipeline = ws.cell(row=row, column=5, value=float(total_pipeline))
        cell_pipeline.number_format = money_format
        cell_pipeline.border = thin_border
        ws.cell(row=row, column=6, value=cliente_nombre).border = thin_border
        row += 1

    # Ajustar anchos
    col_widths = [25, 18, 18, 22, 22, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="estadisticas_usuarios.xlsx"'
    wb.save(response)
    return response


def trigger_immediate_opportunity_notification(opportunity, request):
    """
    Crown Jewel Feature: Notifica inmediatamente a la sesión del usuario 
    sobre una nueva oportunidad creada, marcándola para detección inmediata
    """
    try:
        # Crear una entrada en la caché/session para que el frontend la detecte inmediatamente
        if hasattr(request, 'session'):
            session_key = f'new_opportunity_alert_{request.user.id}'
            request.session[session_key] = {
                'opportunity_id': opportunity.id,
                'titulo': opportunity.oportunidad,
                'cliente_id': opportunity.cliente.id if opportunity.cliente else None,
                'cliente_nombre': opportunity.cliente.nombre_empresa if opportunity.cliente else 'Sin cliente',
                'monto_estimado': str(opportunity.precio_estimado) if opportunity.precio_estimado else str(opportunity.monto) if opportunity.monto else 'N/A',
                'created_at': opportunity.created_at.isoformat() if opportunity.created_at else timezone.now().isoformat(),
                'user_id': opportunity.usuario.id if opportunity.usuario else request.user.id,
                'timestamp': timezone.now().timestamp()
            }
            # Marcar para que expire en 5 minutos
            request.session.set_expiry(300)
            print(f"DEBUG: Nueva oportunidad marcada para notificación inmediata: {opportunity.oportunidad}")
        
    except Exception as e:
        print(f"ERROR: No se pudo configurar notificación de oportunidad: {e}")


def get_logo_base64():
    """
    Función auxiliar para obtener el logo en base64
    """
    try:
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'img', 'bajanet_logo.png')
        with open(logo_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except FileNotFoundError:
        return ""


def parse_producto_inteligente(linea, marcas_validas):
    """
    Parser inteligente que detecta productos por patrón marca-precio
    """
    import re
    
    # Dividir la línea en tokens por espacios
    tokens = linea.split()
    if len(tokens) < 3:  # Mínimo: marca, no_parte, precio
        return None
    
    # Buscar marca válida al inicio
    marca = None
    marca_index = -1
    for i, token in enumerate(tokens):
        if token.upper() in marcas_validas:
            marca = token.upper()
            marca_index = i
            break
    
    if not marca or marca_index == -1:
        return None
    
    # Buscar precio (número decimal) desde el final
    precio = None
    precio_index = -1
    for i in range(len(tokens) - 1, marca_index, -1):
        token = tokens[i].replace(',', '').replace('$', '')
        try:
            precio_val = float(token)
            if precio_val >= 0:
                precio = precio_val
                precio_index = i
                break
        except ValueError:
            continue
    
    if precio is None or precio_index == -1:
        return None
    
    # Extraer no_parte (siguiente token después de marca)
    if marca_index + 1 >= precio_index:
        return None
    
    no_parte = tokens[marca_index + 1]
    
    # Extraer descripción (todos los tokens entre no_parte y precio)
    descripcion_tokens = tokens[marca_index + 2:precio_index]
    descripcion = ' '.join(descripcion_tokens) if descripcion_tokens else no_parte
    
    # Si no hay descripción separada, usar no_parte como descripción base
    if not descripcion.strip():
        descripcion = f"Producto {no_parte}"
    
    return {
        'marca': marca,
        'no_parte': no_parte,
        'descripcion': descripcion,
        'precio': precio
    }


def limpiar_y_corregir_texto(texto):
    """
    Limpia caracteres de encoding corrupto y corrige palabras comunes
    """
    import re
    
    # Diccionario de correcciones comunes de caracteres de encoding
    correcciones_encoding = {
        'Bimet√°lica': 'Bimetálica',
        'bimet√°lica': 'bimetálica',
        'met√°lica': 'metálica',
        'Met√°lica': 'Metálica',
        'el√©ctrica': 'eléctrica',
        'El√©ctrica': 'Eléctrica',
        'electr√≥nica': 'electrónica',
        'Electr√≥nica': 'Electrónica',
        'mec√°nica': 'mecánica',
        'Mec√°nica': 'Mecánica',
        'hidr√°ulica': 'hidráulica',
        'Hidr√°ulica': 'Hidráulica',
        'neum√°tica': 'neumática',
        'Neum√°tica': 'Neumática',
        'autom√°tica': 'automática',
        'Autom√°tica': 'Automática',
        'est√°ndar': 'estándar',
        'Est√°ndar': 'Estándar',
        'b√°sica': 'básica',
        'B√°sica': 'Básica',
        'avanzada': 'avanzada',
        'Avanzada': 'Avanzada',
        'pr√°ctica': 'práctica',
        'Pr√°ctica': 'Práctica',
        'gu√≠a': 'guía',
        'Gu√≠a': 'Guía',
        'energ√≠a': 'energía',
        'Energ√≠a': 'Energía',
        'bater√≠a': 'batería',
        'Bater√≠a': 'Batería',
        'H√∫medas': 'Húmedas',
        'h√∫medas': 'húmedas',
        'h√∫meda': 'húmeda',
        'H√∫meda': 'Húmeda',
        'C√©dula': 'Cédula',
        'c√©dula': 'cédula',
        'c√©lula': 'célula',
        'C√©lula': 'Célula',
        'f√°cil': 'fácil',
        'F√°cil': 'Fácil',
        'r√°pida': 'rápida',
        'R√°pida': 'Rápida',
        'r√°pido': 'rápido',
        'R√°pido': 'Rápido',
        'c√°mara': 'cámara',
        'C√°mara': 'Cámara',
        '√°rea': 'área',
        '√Área': 'Área',
        'energ√≠a': 'energía',
        'Energ√≠a': 'Energía',
        'tecnolog√≠a': 'tecnología',
        'Tecnolog√≠a': 'Tecnología',
        'categor√≠a': 'categoría',
        'Categor√≠a': 'Categoría',
        'garant√≠a': 'garantía',
        'Garant√≠a': 'Garantía',
        'compa√±√≠a': 'compañía',
        'Compa√±√≠a': 'Compañía',
        'dise√±o': 'diseño',
        'Dise√±o': 'Diseño',
        'peque√±o': 'pequeño',
        'Peque√±o': 'Pequeño',
        'peque√±a': 'pequeña',
        'Peque√±a': 'Pequeña',
        'a√±o': 'año',
        'A√±o': 'Año',
        'ni√±o': 'niño',
        'Ni√±o': 'Niño',
        'ni√±a': 'niña',
        'Ni√±a': 'Niña',
        'espa√±ol': 'español',
        'Espa√±ol': 'Español',
        'espa√±ola': 'española',
        'Espa√±ola': 'Española'
    }
    
    # Patrones de corrección automática
    patrones_correccion = [
        # √° -> á
        (r'(.*)√°(.*)', r'\1á\2'),
        # √© -> é  
        (r'(.*)√©(.*)', r'\1é\2'),
        # √≠ -> í
        (r'(.*)√≠(.*)', r'\1í\2'),
        # √≥ -> ó
        (r'(.*)√≥(.*)', r'\1ó\2'),
        # √∫ -> ú (común en "húmedo")
        (r'(.*)√∫(.*)', r'\1ú\2'),
        # √± -> ñ
        (r'(.*)√±(.*)', r'\1ñ\2'),
        # Otros caracteres problemáticos
        (r'(.*)‚Ä¢(.*)', r'\1"\2'),
        (r'(.*)‚Äù(.*)', r'\1"\2'),
        (r'(.*)‚Äî(.*)', r'\1-\2'),
        (r'(.*)√¢(.*)', r'\1â\2'),
        (r'(.*)√™(.*)', r'\1ê\2'),
        (r'(.*)√Æ(.*)', r'\1Æ\2'),
        # Limpiar caracteres raros restantes
        (r'√[^a-zA-Z0-9]', ''),
        # Múltiples espacios
        (r'\s+', ' ')
    ]
    
    # Aplicar correcciones específicas del diccionario primero
    texto_corregido = texto
    for incorrecto, correcto in correcciones_encoding.items():
        texto_corregido = texto_corregido.replace(incorrecto, correcto)
    
    # Aplicar patrones de corrección automática
    for patron, reemplazo in patrones_correccion:
        texto_corregido = re.sub(patron, reemplazo, texto_corregido)
    
    # Normalizar espacios y limpiar
    texto_corregido = re.sub(r'\s+', ' ', texto_corregido).strip()
    
    return texto_corregido


def es_cotizacion_automatica(deal_details):
    """
    Detecta si una oportunidad debe generar cotización automática
    basado en si el campo producto es una de las 7 marcas principales
    """
    producto_bitrix_id = deal_details.get('UF_CRM_1752859685662')
    
    # IDs de las 7 marcas que tienen cotización automática
    marcas_automaticas = [
        "176",  # ZEBRA
        "178",  # PANDUIT
        "180",  # APC
        "182",  # AVIGILON
        "184",  # GENETEC
        "186",  # AXIS
        "194",  # CISCO
    ]
    
    return str(producto_bitrix_id) in marcas_automaticas


def extraer_productos_inteligente(texto_requisicion, marca_seleccionada):
    """
    Parser súper inteligente que extrae números de parte y cantidades
    de cualquier formato de texto libre
    """
    if not texto_requisicion or not marca_seleccionada:
        return []
    
    productos_encontrados = []
    
    # Limpiar texto y dividir en líneas
    lineas = texto_requisicion.strip().replace('\r\n', '\n').split('\n')
    
    for linea in lineas:
        linea = linea.strip()
        if not linea:
            continue
            
        # Buscar patrones flexibles: texto alfanumérico + número
        # Patrones que detecta:
        # - "ZT411-203DPI 2"
        # - "ZT411 cantidad: 3"
        # - "necesito 5 del ZT411-203DPI"
        # - "450145 x3"
        # - "ZT411 (2 piezas)"
        
        # Patrón principal: buscar código alfanumérico y número cercano
        patron_principal = r'([A-Z0-9\-_]+).*?(\d+)'
        matches = re.findall(patron_principal, linea, re.IGNORECASE)
        
        for posible_parte, cantidad_str in matches:
            # Limpiar el posible número de parte
            posible_parte = posible_parte.strip().upper()
            
            # Verificar que sea un código válido (al menos 3 caracteres)
            if len(posible_parte) < 3:
                continue
                
            # Verificar que existe en nuestro catálogo para esta marca
            from app.models import ProductoCatalogo, Marca
            try:
                marca_obj = Marca.objects.get(nombre=marca_seleccionada, activa=True)
                producto = ProductoCatalogo.objects.filter(
                    marca=marca_obj,
                    no_parte__iexact=posible_parte,
                    activo=True
                ).first()
                
                if producto:
                    cantidad = int(cantidad_str)
                    productos_encontrados.append({
                        'no_parte': producto.no_parte,
                        'descripcion': producto.descripcion,
                        'precio': float(producto.precio),
                        'cantidad': cantidad,
                        'total': float(producto.precio) * cantidad
                    })
                    print(f"COTIZACIÓN AUTOMÁTICA: Producto encontrado: {producto.no_parte} x{cantidad}")
                    
            except (Marca.DoesNotExist, ValueError) as e:
                print(f"COTIZACIÓN AUTOMÁTICA: Error procesando {posible_parte}: {e}")
                continue
    
    return productos_encontrados


def limpiar_actividades_huerfanas(oportunidad):
    """
    Función para limpiar actividades de comentario que ya no tienen comentario asociado
    """
    try:
        actividades_comentario = OportunidadActividad.objects.filter(
            oportunidad=oportunidad,
            tipo='comentario'
        )
        
        comentarios_existentes = OportunidadComentario.objects.filter(
            oportunidad=oportunidad
        )
        
        actividades_huerfanas = []
        
        for actividad in actividades_comentario:
            # Buscar si existe un comentario para esta actividad usando el nuevo sistema de IDs
            comentario_encontrado = False
            
            # Estrategia 1: Buscar por ID directo en la descripción (nuevo sistema)
            import re
            match = re.search(r'\[COMENTARIO_ID:(\d+)\]', actividad.descripcion or '')
            if match:
                comentario_id = int(match.group(1))
                if comentarios_existentes.filter(id=comentario_id).exists():
                    comentario_encontrado = True
                    print(f"✅ Actividad {actividad.id} tiene comentario válido: {comentario_id}")
                else:
                    print(f"❌ Actividad {actividad.id} referencia comentario inexistente: {comentario_id}")
            else:
                # Estrategia 2: Fallback para actividades del sistema viejo (por tiempo y usuario)
                for comentario in comentarios_existentes:
                    # Verificar por rango de tiempo
                    diff_tiempo = abs((actividad.fecha_creacion - comentario.fecha_creacion).total_seconds())
                    if (diff_tiempo <= 300 and  # 5 minutos
                        actividad.usuario == comentario.usuario):
                        comentario_encontrado = True
                        print(f"✅ Actividad {actividad.id} encontrada por tiempo: comentario {comentario.id}")
                        break
                    
                    # Verificar por contenido en descripción
                    if (actividad.usuario == comentario.usuario and 
                        comentario.contenido in actividad.descripcion):
                        comentario_encontrado = True
                        print(f"✅ Actividad {actividad.id} encontrada por contenido: comentario {comentario.id}")
                        break
            
            if not comentario_encontrado:
                actividades_huerfanas.append(actividad)
        
        # Eliminar actividades huérfanas
        if actividades_huerfanas:
            print(f"🧹 Limpiando {len(actividades_huerfanas)} actividades huérfanas")
            for actividad in actividades_huerfanas:
                print(f"🗑️ Eliminando actividad huérfana ID={actividad.id}")
                actividad.delete()
        
    except Exception as e:
        print(f"⚠️ Error limpiando actividades huérfanas: {e}")


def crear_notificacion(usuario_destinatario, tipo, titulo, mensaje, oportunidad=None, comentario=None, usuario_remitente=None, proyecto_id=None, proyecto_nombre=None, tarea_opp=None, solicitud_perfil=None, tarea_id=None, tarea_titulo=None):
    """
    Función auxiliar para crear notificaciones
    """
    try:
        # No crear notificación si el remitente y destinatario son el mismo
        if usuario_remitente and usuario_destinatario == usuario_remitente:
            return None

        notificacion = Notificacion.objects.create(
            usuario_destinatario=usuario_destinatario,
            usuario_remitente=usuario_remitente,
            tipo=tipo,
            titulo=titulo,
            mensaje=mensaje,
            oportunidad=oportunidad,
            comentario=comentario,
            proyecto_id=proyecto_id,
            proyecto_nombre=proyecto_nombre,
            tarea_opp=tarea_opp,
            solicitud_perfil=solicitud_perfil,
            tarea_id=tarea_id,
            tarea_titulo=tarea_titulo,
        )
        
        print(f"✅ Notificación creada: {titulo} para {usuario_destinatario.username}")
        return notificacion
        
    except Exception as e:
        print(f"❌ Error creando notificación: {e}")
        return None


def notificar_miembro_agregado_proyecto(usuario_agregado, proyecto_nombre, proyecto_id, usuario_que_agrega):
    """
    Función específica para notificar cuando se agrega un usuario a un proyecto
    """
    titulo = f"Te han agregado al proyecto: {proyecto_nombre}"
    mensaje = f"{usuario_que_agrega.get_full_name() or usuario_que_agrega.username} te ha agregado como miembro del proyecto '{proyecto_nombre}'. Ahora puedes colaborar en este proyecto."
    
    return crear_notificacion(
        usuario_destinatario=usuario_agregado,
        tipo='proyecto_agregado',
        titulo=titulo,
        mensaje=mensaje,
        usuario_remitente=usuario_que_agrega,
        proyecto_id=proyecto_id,
        proyecto_nombre=proyecto_nombre
    )


def detectar_menciones_en_comentario(contenido, usuario_remitente, oportunidad, comentario=None):
    """
    Detecta menciones @usuario en un comentario y crea notificaciones
    """
    import re
    
    # Buscar menciones en el formato @usuario
    menciones = re.findall(r'@(\w+)', contenido)
    
    for username in menciones:
        try:
            usuario_mencionado = User.objects.get(username=username)
            
            # Crear notificación de mención
            crear_notificacion(
                usuario_destinatario=usuario_mencionado,
                tipo='mencion',
                titulo='Te han mencionado en un comentario',
                mensaje=f'{usuario_remitente.get_full_name() or usuario_remitente.username} te mencionó en la oportunidad "{oportunidad.oportunidad}": {contenido[:100]}...',
                oportunidad=oportunidad,
                comentario=comentario,
                usuario_remitente=usuario_remitente
            )
            
        except User.DoesNotExist:
            # Usuario mencionado no existe, ignorar
            pass


def handle_lost_opportunity(opportunity, action='log_only'):
    """
    Maneja oportunidades marcadas como "Cerrado Perdido" en Bitrix24
    
    Actions:
    - 'log_only': Solo registra en logs (default)
    - 'delete': Elimina la oportunidad
    - 'mark_inactive': Marca como inactiva (requiere campo adicional)
    """
    opportunity_name = opportunity.oportunidad if hasattr(opportunity, 'oportunidad') else 'Oportunidad'
    
    if action == 'delete':
        opportunity.delete()
        print(f"BITRIX WEBHOOK: 🗑️ Oportunidad '{opportunity_name}' eliminada por estar marcada como PERDIDA", flush=True)
        return {'deleted': True}
    elif action == 'mark_inactive':
        # Si implementas un campo 'activa' en el modelo
        # opportunity.activa = False
        # opportunity.save()
        print(f"BITRIX WEBHOOK: 📴 Oportunidad '{opportunity_name}' marcada como inactiva por estar PERDIDA", flush=True)
        return {'marked_inactive': True}
    else:  # log_only
        print(f"BITRIX WEBHOOK: ⚠️ Oportunidad '{opportunity_name}' está marcada como PERDIDA en Bitrix24 (solo registro)", flush=True)
        return {'logged': True}


def notificar_nueva_tarea(tarea, creado_por):
    """
    Envía notificación cuando se crea una nueva tarea
    """
    try:
        # Notificar al responsable de la tarea
        if tarea.asignado_a and tarea.asignado_a != creado_por:
            crear_notificacion(
                usuario_destinatario=tarea.asignado_a,
                tipo='tarea_nueva',
                titulo=f'Nueva tarea asignada: {tarea.titulo}',
                mensaje=f'Se te ha asignado la tarea "{tarea.titulo}" en el proyecto {tarea.proyecto.nombre if tarea.proyecto else "Sin proyecto"}',
                usuario_remitente=creado_por
            )
        
        # Notificar a todos los participantes
        for participante in tarea.participantes.all():
            if participante != creado_por and participante != tarea.asignado_a:
                crear_notificacion(
                    usuario_destinatario=participante,
                    tipo='tarea_nueva',
                    titulo=f'Nueva tarea: {tarea.titulo}',
                    mensaje=f'Has sido agregado como participante en la tarea "{tarea.titulo}"',
                    usuario_remitente=creado_por
                )
        
        # Notificar a todos los observadores
        for observador in tarea.observadores.all():
            if observador != creado_por and observador != tarea.asignado_a:
                crear_notificacion(
                    usuario_destinatario=observador,
                    tipo='tarea_nueva',
                    titulo=f'Nueva tarea: {tarea.titulo}',
                    mensaje=f'Has sido agregado como observador en la tarea "{tarea.titulo}"',
                    usuario_remitente=creado_por
                )
                
    except Exception as e:
        print(f"Error enviando notificaciones de nueva tarea: {e}")


def get_oportunidades_data(request, mes, anio):
    """Helper para obtener datos de oportunidades"""
    # Implementar lógica similar a crm_home
    return []


def get_facturado_data(request, mes, anio):
    """Helper para obtener datos de facturado"""
    # Implementar lógica similar a crm_home
    return []


def _serialize_tarea_opp(t):
    return {
        'id': t.id,
        'titulo': t.titulo,
        'descripcion': t.descripcion,
        'prioridad': t.prioridad,
        'estado': t.estado,
        'fecha_limite': t.fecha_limite.isoformat() if t.fecha_limite else None,
        'fecha_creacion': t.fecha_creacion.isoformat(),
        'creado_por': (t.creado_por.get_full_name() or t.creado_por.username) if t.creado_por else '',
        'creado_por_id': t.creado_por_id,
        'responsable': (t.responsable.get_full_name() or t.responsable.username) if t.responsable else None,
        'responsable_id': t.responsable_id,
        'oportunidad_id': t.oportunidad_id,
        'oportunidad_nombre': t.oportunidad.oportunidad if t.oportunidad else '',
    }


def _muro_post_dict(post, request_user):
    """Serializa un PostMuro a dict para la API."""
    autor = post.autor
    profile = getattr(autor, 'userprofile', None)
    avatar_url = profile.get_avatar_url() if profile else None
    iniciales = profile.iniciales() if profile else (autor.first_name[:1] + autor.last_name[:1]).upper() if autor.first_name else autor.username[:2].upper()

    etiquetados = [
        {'id': u.id, 'nombre': u.get_full_name() or u.username}
        for u in post.etiquetados.all()
    ]
    num_likes = post.likes.count()
    yo_like = post.likes.filter(pk=request_user.pk).exists()
    num_comentarios = post.comentarios.count()

    imagen_url = post.imagen.url if post.imagen else None
    programado_str = post.programado_para.strftime('%Y-%m-%dT%H:%M') if post.programado_para else None

    res = {
        'id': post.id,
        'autor_id': autor.id,
        'autor_nombre': autor.get_full_name() or autor.username,
        'autor_avatar': avatar_url,
        'autor_iniciales': iniciales,
        'contenido': post.contenido,
        'imagen': imagen_url,
        'etiquetados': etiquetados,
        'es_anuncio': post.es_anuncio,
        'programado_para': programado_str,
        'fecha': convert_to_tijuana_time(post.fecha_creacion).strftime('%d %b %Y %H:%M'),
        'editado': post.editado,
        'num_likes': num_likes,
        'yo_like': yo_like,
        'num_comentarios': num_comentarios,
        'es_autor': post.autor_id == request_user.pk,
    }

    # Sobrescribir autor si es un anuncio del sistema (como Empleado del Mes)
    contenido_upper = (post.contenido or '').upper()
    if post.es_anuncio and ("EMPLEADO DEL MES" in contenido_upper or "empleado del mes" in post.contenido or "IAMET" in contenido_upper):
        res['autor_nombre'] = "IAMET"
        res['autor_iniciales'] = "IA"
        res['autor_avatar'] = None
        res['autor_avatar'] = "/static/images/apple-touch-icon.png"  # Logo corporativo
        res['autor_iniciales'] = "IA"

    return res
