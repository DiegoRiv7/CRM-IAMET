# ═══════════════════════════════════════════════════════════════
#  views_iamet.py — APIs del modulo de Proyectos IAMET
#  Gestion de proyectos de telecomunicaciones
# ═══════════════════════════════════════════════════════════════

import json
import re
from decimal import Decimal, InvalidOperation
from datetime import date
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.db.models import Sum, Count, Q, DecimalField
from django.utils import timezone
from .models import (
    ProyectoIAMET as Proyecto, ProyectoPartida, ProyectoOrdenCompra,
    ProyectoFacturaProveedor, ProyectoFacturaIngreso,
    ProyectoGasto, ProyectoTarea, ProyectoAlerta,
    ProyectoConfiguracion, ProyectoVolumetriaVersion,
    ProyectoLevantamiento, LevantamientoEvidencia,
    ProyectoVolumetria,
)
from .views_utils import is_supervisor
from .views_grupos import get_usuarios_visibles_ids


# ─── Helpers ──────────────────────────────────────────────────

def _get_proyectos_qs(user):
    # Todos los usuarios ven todos los proyectos (vendedores, ingenieros, supervisores)
    return Proyecto.objects.all()


def _check_access(user, proyecto):
    # Todos los usuarios tienen acceso a todos los proyectos
    return True
    # --- código original deshabilitado ---
    if is_supervisor(user):
        return True
    vis_ids = get_usuarios_visibles_ids(user)
    if vis_ids is None:
        return True
    return proyecto.usuario_id in vis_ids


def _dec(val, default=Decimal('0')):
    if val is None or val == '':
        return default
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return default


def _parse_date(val):
    if not val:
        return None
    try:
        parts = str(val).split('-')
        return date(int(parts[0]), int(parts[1]), int(parts[2]))
    except Exception:
        return None


def _fmt(val):
    if val is None:
        return None
    if hasattr(val, 'isoformat'):
        return val.isoformat()
    return str(val)


def _calcular_avance(proyecto):
    """Calcula % de avance del proyecto basado en partidas + etapa oportunidad."""
    from .models import ProgramacionActividad

    # 40% weight: partidas completadas (status closed or received)
    total_partidas = proyecto.partidas.count()
    if total_partidas > 0:
        partidas_completas = proyecto.partidas.filter(status__in=['closed', 'received']).count()
        pct_partidas = (partidas_completas / total_partidas) * 100
    else:
        pct_partidas = 0

    # 30% weight: etapa de oportunidad
    pct_etapa = 0
    if proyecto.oportunidad and proyecto.oportunidad.etapa_corta:
        etapa = proyecto.oportunidad.etapa_corta
        # Map etapas to progress %
        etapa_map = {
            'En Solicitud': 5, 'Cotizando': 15, 'Enviada': 25, 'Seguimiento': 35,
            'Vendido s/PO': 45, 'Vendido c/PO': 55, 'En Tránsito': 65,
            'Facturado': 75, 'Programado': 80, 'Entregado': 85,
            'Esperando Pago': 90, 'Ganado': 95, 'Pagado': 100,
        }
        pct_etapa = etapa_map.get(etapa, 50)

    # 30% weight: programa de obra (actividades completadas)
    pct_programa = 0
    actividades = ProgramacionActividad.objects.filter(proyecto_key=f'proy_{proyecto.id}')
    total_acts = actividades.count()
    if total_acts > 0:
        # Activities with fecha in the past are considered "done"
        acts_pasadas = actividades.filter(fecha__lt=timezone.localdate()).count()
        pct_programa = (acts_pasadas / total_acts) * 100

    avance = (pct_partidas * 0.4) + (pct_etapa * 0.3) + (pct_programa * 0.3)

    # Cap: if oportunidad is not 'Pagado', max 95%
    if proyecto.oportunidad and proyecto.oportunidad.etapa_corta != 'Pagado':
        avance = min(avance, 95)

    return round(min(avance, 100), 1)


def _calcular_efectividad(proyecto):
    """Calcula % de efectividad. Empieza en 100% y baja por penalizaciones."""
    from .models import ProgramacionActividad

    efectividad = 100.0
    hoy = timezone.localdate()

    # -5% por cada tarea vencida no completada
    tareas_vencidas = proyecto.tareas_proyecto.filter(
        fecha_limite__lt=timezone.now()
    ).exclude(status__in=['completed', 'cancelled']).count()
    efectividad -= tareas_vencidas * 5

    # -5% si gastado > costo total presupuestado
    gastado = float(proyecto.ordenes_compra.exclude(status='cancelled').aggregate(t=Sum('monto_total'))['t'] or 0)
    costo_total = float(proyecto.partidas.aggregate(t=Sum('costo_total'))['t'] or 0)
    if costo_total > 0 and gastado > costo_total:
        efectividad -= 5

    # -3% por cada OC con precio_unitario > partida.costo_unitario
    for oc in proyecto.ordenes_compra.exclude(status='cancelled').select_related('partida'):
        if oc.partida and oc.precio_unitario and oc.partida.costo_unitario:
            if oc.precio_unitario > oc.partida.costo_unitario:
                efectividad -= 3

    # -10% si fecha actual > fecha_fin y oportunidad no está en Pagado
    if proyecto.fecha_fin and hoy > proyecto.fecha_fin:
        if not proyecto.oportunidad or proyecto.oportunidad.etapa_corta != 'Pagado':
            efectividad -= 10

    # -5% si programa de obra tiene actividades atrasadas
    acts_atrasadas = ProgramacionActividad.objects.filter(
        proyecto_key=f'proy_{proyecto.id}',
        fecha__lt=hoy
    ).count()
    # Only penalize if there are activities that should have happened
    if acts_atrasadas > 0:
        total_acts = ProgramacionActividad.objects.filter(proyecto_key=f'proy_{proyecto.id}').count()
        if total_acts > 0:
            # Check if there are future activities too (meaning project is ongoing)
            acts_futuras = ProgramacionActividad.objects.filter(
                proyecto_key=f'proy_{proyecto.id}',
                fecha__gte=hoy
            ).count()
            if acts_futuras == 0 and proyecto.status != 'completed':
                efectividad -= 5

    return round(max(0, min(100, efectividad)), 1)


# ─── Helpers para overview enriquecido (widget detalle proyecto) ──

# Paleta determinística para color de cliente (basada en hash del id)
_CLIENT_COLOR_PALETTE = [
    '#0052D4', '#7B61FF', '#10B981', '#F59E0B', '#EF4444',
    '#06B6D4', '#EC4899', '#8B5CF6', '#22C55E', '#F97316',
    '#3B82F6', '#14B8A6',
]


def _iniciales(nombre):
    """Devuelve las iniciales (máx 2) de un nombre. '' si vacío."""
    if not nombre:
        return ''
    partes = [p for p in str(nombre).strip().split() if p]
    if not partes:
        return ''
    if len(partes) == 1:
        return partes[0][:2].upper()
    return (partes[0][0] + partes[-1][0]).upper()


def _get_user_card(user, rol_label=''):
    """Construye un dict ligero con datos del usuario para el frontend.

    Devuelve None si user es falsy.
    """
    if not user:
        return None
    nombre = (user.first_name + ' ' + user.last_name).strip() or user.username
    avatar_url = None
    rol = ''
    try:
        prof = user.userprofile
        if prof:
            try:
                avatar_url = prof.get_avatar_url()
            except Exception:
                avatar_url = None
            rol = (prof.rol or '').strip()
    except Exception:
        # Sin profile asociado — seguimos sin avatar/rol
        pass
    card = {
        'id': user.id,
        'nombre': nombre,
        'iniciales': _iniciales(nombre),
        'avatar_url': avatar_url,
        'rol': rol,
    }
    if rol_label:
        card['rol_label'] = rol_label
    return card


def _extraer_codigo_proyecto(nombre, opp):
    """Intenta extraer un código 'PO 1234567' del nombre del proyecto o de la oportunidad.

    Patrones soportados (case-insensitive):
    - "PO 1234567" / "PO-1234567" / "PO1234567"
    - po_number directo de la oportunidad

    Asunción: si no se encuentra patrón, regresa cadena vacía.
    """
    # 1) Si hay PO directo en la oportunidad ligada, úsalo
    if opp and getattr(opp, 'po_number', None):
        po = str(opp.po_number).strip()
        if po:
            # Normalizar a "PO <num>" si parece ser sólo el número
            if re.fullmatch(r'\d+', po):
                return f'PO {po}'
            return po
    # 2) Buscar en el nombre del proyecto
    if nombre:
        m = re.search(r'\bPO[\s\-]*([A-Z0-9]+)\b', str(nombre), flags=re.IGNORECASE)
        if m:
            return f'PO {m.group(1)}'
    return ''


def _color_cliente(cliente_id):
    """Color determinístico basado en id de cliente (hash mod paleta)."""
    if not cliente_id:
        return _CLIENT_COLOR_PALETTE[0]
    return _CLIENT_COLOR_PALETTE[int(cliente_id) % len(_CLIENT_COLOR_PALETTE)]


def _derivar_status(proyecto, avance_pct):
    """Devuelve (status, label, color).

    Reutiliza proyecto.status si es 'paused' (detenido) o 'completed'.
    En otro caso deriva por fechas + avance.
    Asunción: el modelo IAMET no tiene un status 'riesgo' nativo, lo derivamos.
    """
    hoy = timezone.localdate()
    raw_status = (proyecto.status or '').lower()

    # Mapeos directos cuando el modelo lo dice claro
    if raw_status == 'paused':
        return ('detenido', 'Detenido', '#ef4444')
    if raw_status == 'completed' or (avance_pct is not None and avance_pct >= 100):
        return ('completado', 'Completado', '#6b7280')

    # Vencido y no completado → riesgo (ámbar)
    if proyecto.fecha_fin and hoy > proyecto.fecha_fin and (avance_pct or 0) < 100:
        return ('riesgo', 'En riesgo', '#f59e0b')

    # Sin fecha_inicio o futura → planificación
    if not proyecto.fecha_inicio or proyecto.fecha_inicio > hoy:
        return ('planificacion', 'Planificación', '#3b82f6')

    # Hoy entre fecha_inicio y fecha_fin (o sin fecha_fin)
    if proyecto.fecha_inicio <= hoy and (not proyecto.fecha_fin or hoy <= proyecto.fecha_fin):
        return ('ejecucion', 'En ejecución', '#10b981')

    # Fallback conservador
    return ('ejecucion', 'En ejecución', '#10b981')


def _calcular_salud(margen_pct, avance_vs_tiempo_delta, dias_restantes, avance_pct):
    """Devuelve {rag, label, razones} aplicando reglas en orden (rojo > ámbar > verde)."""
    razones_rojo = []
    razones_ambar = []

    # Reglas ROJO
    if margen_pct is not None and margen_pct < 0:
        razones_rojo.append('En pérdida')
    if avance_vs_tiempo_delta is not None and avance_vs_tiempo_delta < -25:
        razones_rojo.append(f'Atrasado {abs(int(round(avance_vs_tiempo_delta)))}% vs cronograma')
    if dias_restantes is not None and dias_restantes < 0 and (avance_pct or 0) < 100:
        razones_rojo.append(f'Vencido hace {abs(int(dias_restantes))} días')

    if razones_rojo:
        return {
            'rag': 'rojo',
            'label': razones_rojo[0],
            'razones': razones_rojo,
        }

    # Reglas ÁMBAR
    if margen_pct is not None and margen_pct < 15:
        razones_ambar.append(f'Margen apretado ({int(round(margen_pct))}%)')
    if avance_vs_tiempo_delta is not None and avance_vs_tiempo_delta < -10:
        razones_ambar.append(f'Atrasado {abs(int(round(avance_vs_tiempo_delta)))}% vs cronograma')
    if (
        dias_restantes is not None and 0 <= dias_restantes <= 3
        and (avance_pct or 0) < 90
    ):
        razones_ambar.append(
            'Vence hoy' if dias_restantes == 0 else f'Vence en {int(dias_restantes)} días'
        )

    if razones_ambar:
        return {
            'rag': 'ambar',
            'label': razones_ambar[0],
            'razones': razones_ambar,
        }

    return {
        'rag': 'verde',
        'label': 'En tiempo · sin riesgos',
        'razones': [],
    }


def _proyecto_overview(proyecto):
    """Construye el sub-objeto `overview` con datos enriquecidos para el widget de detalle.

    NOTA: el modelo `ProyectoIAMET` (alias `Proyecto`) sólo tiene un único
    `usuario` (sin M2M de miembros ni `creado_por`); por eso `equipo` se
    arma con el `usuario` del proyecto y el `usuario` (vendedor) de la
    oportunidad, deduplicado. `ingeniero_responsable` toma el `usuario`
    del proyecto si su userprofile.rol == 'ingeniero', o cae al usuario
    igualmente si no hay otro candidato.
    """
    hoy = timezone.localdate()
    opp = proyecto.oportunidad if proyecto.oportunidad_id else None

    # ── Fechas / tiempo ──
    fecha_inicio = proyecto.fecha_inicio
    fecha_fin = proyecto.fecha_fin
    dias_totales = None
    dias_restantes = None
    tiempo_transcurrido_pct = 0
    if fecha_inicio and fecha_fin and fecha_fin >= fecha_inicio:
        dias_totales = (fecha_fin - fecha_inicio).days
        dias_restantes = (fecha_fin - hoy).days
        if dias_totales > 0:
            transcurridos = (hoy - fecha_inicio).days
            pct = (transcurridos / dias_totales) * 100
            tiempo_transcurrido_pct = int(max(0, min(100, round(pct))))
        else:
            # mismo día inicio/fin: si ya pasó o es hoy, considerarlo 100%
            tiempo_transcurrido_pct = 100 if hoy >= fecha_inicio else 0
    elif fecha_fin:
        dias_restantes = (fecha_fin - hoy).days

    # ── Avance / status ──
    avance_pct = _calcular_avance(proyecto)
    avance_vs_tiempo_delta = int(round((avance_pct or 0) - tiempo_transcurrido_pct))
    status_code, status_label, status_color = _derivar_status(proyecto, avance_pct)

    # ── Cliente ──
    cliente_dict = None
    if opp and opp.cliente_id:
        cli = opp.cliente
        nombre_cli = cli.nombre_empresa or ''
        cliente_dict = {
            'id': cli.id,
            'nombre': nombre_cli,
            # 'ubicacion' = direccion del cliente (no hay city/state separado)
            'ubicacion': (cli.direccion or '').strip(),
            'iniciales': _iniciales(nombre_cli),
            'color': _color_cliente(cli.id),
            'logo_url': None,  # No hay campo de logo en Cliente (asunción)
        }
    elif proyecto.cliente_nombre:
        # Fallback: el proyecto tiene cliente_nombre como texto plano
        cliente_dict = {
            'id': None,
            'nombre': proyecto.cliente_nombre,
            'ubicacion': '',
            'iniciales': _iniciales(proyecto.cliente_nombre),
            'color': _color_cliente(proyecto.id),  # estable por proyecto
            'logo_url': None,
        }

    # ── Oportunidad ──
    oportunidad_dict = None
    if opp:
        codigo_opp = ''
        if getattr(opp, 'po_number', None):
            po = str(opp.po_number).strip()
            if po:
                codigo_opp = f'PO {po}' if re.fullmatch(r'\d+', po) else po
        oportunidad_dict = {
            'id': opp.id,
            'codigo': codigo_opp,
            'nombre': opp.oportunidad or '',
            'monto': float(opp.monto) if opp.monto is not None else 0.0,
        }

    # ── Vendedor (usuario de la oportunidad) ──
    vendedor_dict = _get_user_card(opp.usuario) if opp and opp.usuario_id else None
    if vendedor_dict:
        # No exponemos rol_label para vendedor (frontend ya lo etiqueta)
        vendedor_dict.pop('rol_label', None)

    # ── Ingeniero responsable ──
    # Asunción: el modelo IAMET no tiene `miembros` M2M; usamos el `usuario`
    # del proyecto como ingeniero si su rol coincide, si no hacemos fallback
    # al mismo usuario (que actúa como creado_por de facto).
    ingeniero_user = proyecto.usuario
    ingeniero_dict = _get_user_card(ingeniero_user, rol_label='Ingeniero responsable')

    # ── Equipo (dedupe por id) ──
    # Orden de precedencia: ingeniero responsable (proyecto.usuario), vendedor
    # (opp.usuario), miembros agregados manualmente vía el botón "+" del topbar.
    equipo_users = []
    seen_ids = set()
    candidatos = [proyecto.usuario, opp.usuario if opp else None]
    try:
        candidatos.extend(list(proyecto.miembros.all()))
    except Exception:
        # M2M aún no migrado — degradamos silenciosamente.
        pass
    for u in candidatos:
        if u and u.id not in seen_ids:
            equipo_users.append(u)
            seen_ids.add(u.id)
    equipo_total = len(equipo_users)
    equipo_cards = []
    for u in equipo_users[:8]:
        card = _get_user_card(u)
        if card:
            equipo_cards.append({
                'id': card['id'],
                'nombre': card['nombre'],
                'iniciales': card['iniciales'],
                'avatar_url': card['avatar_url'],
                'rol': card['rol'],
            })

    # ── Financiero ──
    contratado = 0.0
    if opp and opp.monto is not None:
        contratado = float(opp.monto)
    else:
        contratado = float(proyecto.partidas.aggregate(t=Sum('precio_venta_total'))['t'] or 0)

    cobrado = float(proyecto.facturas_ingreso.aggregate(t=Sum('monto'))['t'] or 0)
    gastado_oc = proyecto.ordenes_compra.exclude(status='cancelled').aggregate(t=Sum('monto_total'))['t'] or Decimal('0')
    gastos_no_rechazados = proyecto.gastos.exclude(estado_aprobacion='rejected').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    gastado = float(gastado_oc) + float(gastos_no_rechazados)

    cobrado_pct = 0
    if contratado > 0:
        cobrado_pct = int(max(0, min(100, round((cobrado / contratado) * 100))))

    if cobrado > 0:
        margen_pct_raw = ((cobrado - gastado) / cobrado) * 100
    else:
        # Sin ingresos cobrados aún — margen "neutro" (asunción conservadora: 0)
        margen_pct_raw = 0
    margen_pct = int(round(margen_pct_raw))

    if margen_pct < 0:
        margen_label = 'Pérdida'
    elif margen_pct < 15:
        margen_label = 'En riesgo'
    elif margen_pct < 30:
        margen_label = 'Apretado'
    else:
        margen_label = 'Saludable'

    # ── Salud (RAG) ──
    salud = _calcular_salud(
        margen_pct=margen_pct if cobrado > 0 else None,
        avance_vs_tiempo_delta=avance_vs_tiempo_delta if (fecha_inicio and fecha_fin) else None,
        dias_restantes=dias_restantes,
        avance_pct=avance_pct,
    )

    # ── Counts (mini-stats del Dashboard) ──
    def _safe_count(rel):
        try:
            return rel.count() if rel is not None else 0
        except Exception:
            return 0

    counts = {
        'levantamientos': _safe_count(getattr(proyecto, 'levantamientos', None)),
        'tareas_pendientes': _safe_count(proyecto.tareas_proyecto.exclude(status__in=['completed', 'cancelled'])),
        'alertas': _safe_count(proyecto.alertas.filter(resuelta=False)),
        'ordenes_compra_activas': _safe_count(proyecto.ordenes_compra.exclude(status='cancelled')),
    }

    # ── Breakdown del presupuesto por categoría (de partidas) ──
    # Si no hay partidas con monto, devolvemos lista vacía y el frontend
    # muestra un mensaje informativo en su lugar.
    breakdown_presupuesto = []
    try:
        partidas_by_cat = list(
            proyecto.partidas.values('categoria')
            .annotate(monto=Sum('precio_venta_total'))
            .order_by('-monto')
        )
        for p in partidas_by_cat:
            monto = float(p.get('monto') or 0)
            if monto > 0:
                breakdown_presupuesto.append({
                    'categoria': p.get('categoria') or 'otros',
                    'monto': monto,
                })
    except Exception:
        breakdown_presupuesto = []

    return {
        'codigo': _extraer_codigo_proyecto(proyecto.nombre, opp),
        'status': status_code,
        'status_label': status_label,
        'status_color': status_color,
        'fecha_inicio': _fmt(fecha_inicio),
        'fecha_fin': _fmt(fecha_fin),
        'dias_totales': dias_totales,
        'dias_restantes': dias_restantes,
        'tiempo_transcurrido_pct': tiempo_transcurrido_pct,
        'avance_pct': avance_pct,
        'avance_vs_tiempo_delta': avance_vs_tiempo_delta,
        'cliente': cliente_dict,
        'oportunidad': oportunidad_dict,
        'vendedor': vendedor_dict,
        'ingeniero_responsable': ingeniero_dict,
        'equipo': equipo_cards,
        'equipo_total': equipo_total,
        'salud': salud,
        'financiero': {
            'contratado': contratado,
            'cobrado': cobrado,
            'gastado': gastado,
            'cobrado_pct': cobrado_pct,
            'margen_pct': margen_pct,
            'margen_label': margen_label,
        },
        'counts': counts,
        'breakdown_presupuesto': breakdown_presupuesto,
    }


def _proyecto_to_dict(p, include_alerts=False):
    # Calcular utilidad real desde facturas y gastos (no del campo del modelo)
    ingresos = float(p.facturas_ingreso.aggregate(t=Sum('monto'))['t'] or 0)
    costos_prov = float(p.facturas_proveedor.aggregate(t=Sum('monto'))['t'] or 0)
    gastos_apr = float(p.gastos.filter(estado_aprobacion='approved').aggregate(t=Sum('monto'))['t'] or 0)
    utilidad_real_calc = ingresos - costos_prov - gastos_apr

    opp = p.oportunidad if hasattr(p, 'oportunidad_id') and p.oportunidad_id else None

    # Agregados de levantamientos (para la lista tipo tabla)
    lev_qs = getattr(p, 'levantamientos', None)
    lev_count = 0
    lev_fase_max = 0
    if lev_qs is not None:
        try:
            lev_count = lev_qs.count()
            if lev_count:
                from django.db.models import Max
                lev_fase_max = lev_qs.aggregate(m=Max('fase_actual'))['m'] or 0
        except Exception:
            pass

    d = {
        'id': p.id,
        'usuario_id': p.usuario_id,
        'usuario_nombre': (p.usuario.first_name + ' ' + p.usuario.last_name).strip() or p.usuario.username,
        'nombre': p.nombre,
        'descripcion': p.descripcion,
        'cliente_nombre': p.cliente_nombre,
        'status': p.status,
        'utilidad_presupuestada': float(p.utilidad_presupuestada),
        'utilidad_real': utilidad_real_calc,
        'fecha_inicio': _fmt(p.fecha_inicio),
        'fecha_fin': _fmt(p.fecha_fin),
        'created_at': _fmt(p.created_at),
        'updated_at': _fmt(p.updated_at),
        'oportunidad_id': opp.id if opp else None,
        'oportunidad_nombre': opp.oportunidad if opp else None,
        'oportunidad_monto': float(opp.monto) if opp and opp.monto is not None else 0.0,
        'oportunidad_producto': opp.producto if opp else None,
        'oportunidad_etapa': opp.etapa_corta if opp else None,
        'oportunidad_etapa_color': opp.etapa_color if opp else None,
        'levantamientos_count': lev_count,
        'levantamiento_fase_max': lev_fase_max,
    }
    if include_alerts:
        d['alertas_pendientes'] = p.alertas.filter(resuelta=False).count()
    return d


def _partida_to_dict(p):
    return {
        'id': p.id,
        'proyecto_id': p.proyecto_id,
        'categoria': p.categoria,
        'descripcion': p.descripcion,
        'marca': p.marca,
        'numero_parte': p.numero_parte,
        'cantidad': float(p.cantidad),
        'cantidad_pendiente': float(p.cantidad_pendiente),
        'precio_lista': float(p.precio_lista),
        'descuento': float(p.descuento),
        'costo_unitario': float(p.costo_unitario),
        'costo_total': float(p.costo_total),
        'precio_venta_unitario': float(p.precio_venta_unitario),
        'precio_venta_total': float(p.precio_venta_total),
        'ganancia': float(p.ganancia),
        'proveedor': p.proveedor,
        'status': p.status,
        'created_at': _fmt(p.created_at),
        'updated_at': _fmt(p.updated_at),
    }


def _oc_to_dict(oc):
    # URL del archivo vinculado (si vino del drive de la oportunidad)
    archivo_url = None
    if hasattr(oc, 'archivo_drive') and oc.archivo_drive_id:
        try:
            a = oc.archivo_drive
            if a and a.oportunidad_id:
                archivo_url = f'/app/api/oportunidad/{a.oportunidad_id}/drive/archivo/{a.id}/stream/'
        except Exception:
            pass
    return {
        'id': oc.id,
        'proyecto_id': oc.proyecto_id,
        'partida_id': oc.partida_id,
        'partida_descripcion': oc.partida.descripcion if oc.partida else '',
        'numero_oc': oc.numero_oc,
        'proveedor': oc.proveedor,
        'cantidad': float(oc.cantidad),
        'precio_unitario': float(oc.precio_unitario),
        'monto_total': float(oc.monto_total),
        'status': oc.status,
        'fecha_emision': _fmt(oc.fecha_emision),
        'fecha_entrega_esperada': _fmt(oc.fecha_entrega_esperada),
        'fecha_entrega_real': _fmt(oc.fecha_entrega_real),
        'notas': oc.notas,
        'archivo_url': archivo_url,
        'created_at': _fmt(oc.created_at),
        'updated_at': _fmt(oc.updated_at),
    }


def _factura_prov_to_dict(f):
    return {
        'id': f.id,
        'proyecto_id': f.proyecto_id,
        'orden_compra_id': f.orden_compra_id,
        'numero_factura': f.numero_factura,
        'proveedor': f.proveedor,
        'monto': float(f.monto),
        'monto_presupuestado': float(f.monto_presupuestado) if f.monto_presupuestado else None,
        'varianza': float(f.varianza) if f.varianza is not None else None,
        'varianza_porcentaje': float(f.varianza_porcentaje) if f.varianza_porcentaje is not None else None,
        'fecha_factura': _fmt(f.fecha_factura),
        'fecha_vencimiento': _fmt(f.fecha_vencimiento),
        'fecha_pago': _fmt(f.fecha_pago),
        'status': f.status,
        'notas': f.notas,
        'created_at': _fmt(f.created_at),
        'updated_at': _fmt(f.updated_at),
    }


def _factura_ingreso_to_dict(f):
    return {
        'id': f.id,
        'proyecto_id': f.proyecto_id,
        'numero_factura': f.numero_factura,
        'monto': float(f.monto),
        'fecha_factura': _fmt(f.fecha_factura),
        'fecha_vencimiento': _fmt(f.fecha_vencimiento),
        'fecha_pago': _fmt(f.fecha_pago),
        'status': f.status,
        'metodo_pago': f.metodo_pago,
        'notas': f.notas,
        'created_at': _fmt(f.created_at),
        'updated_at': _fmt(f.updated_at),
    }


def _gasto_to_dict(g):
    return {
        'id': g.id,
        'proyecto_id': g.proyecto_id,
        'categoria': g.categoria,
        'descripcion': g.descripcion,
        'monto': float(g.monto),
        'monto_presupuestado': float(g.monto_presupuestado) if g.monto_presupuestado else None,
        'varianza': float(g.varianza) if g.varianza is not None else None,
        'fecha_gasto': _fmt(g.fecha_gasto),
        'estado_aprobacion': g.estado_aprobacion,
        'aprobado_por_id': g.aprobado_por_id,
        'aprobado_por_nombre': (
            (g.aprobado_por.first_name + ' ' + g.aprobado_por.last_name).strip() or g.aprobado_por.username
        ) if g.aprobado_por else None,
        'fecha_aprobacion': _fmt(g.fecha_aprobacion),
        'notas': g.notas,
        'created_at': _fmt(g.created_at),
        'updated_at': _fmt(g.updated_at),
    }


def _tarea_to_dict(t):
    return {
        'id': t.id,
        'proyecto_id': t.proyecto_id,
        'titulo': t.titulo,
        'descripcion': t.descripcion,
        'status': t.status,
        'prioridad': t.prioridad,
        'asignado_a_id': t.asignado_a_id,
        'asignado_a_nombre': (
            (t.asignado_a.first_name + ' ' + t.asignado_a.last_name).strip() or t.asignado_a.username
        ) if t.asignado_a else None,
        'fecha_limite': _fmt(t.fecha_limite),
        'fecha_completada': _fmt(t.fecha_completada),
        'horas_estimadas': float(t.horas_estimadas) if t.horas_estimadas else None,
        'horas_reales': float(t.horas_reales) if t.horas_reales else None,
        'notas': t.notas,
        'created_at': _fmt(t.created_at),
        'updated_at': _fmt(t.updated_at),
    }


def _alerta_to_dict(a):
    return {
        'id': a.id,
        'proyecto_id': a.proyecto_id,
        'tipo_alerta': a.tipo_alerta,
        'severidad': a.severidad,
        'titulo': a.titulo,
        'mensaje': a.mensaje,
        'entidad_tipo': a.entidad_tipo,
        'entidad_id': a.entidad_id,
        'resuelta': a.resuelta,
        'fecha_resolucion': _fmt(a.fecha_resolucion),
        'created_at': _fmt(a.created_at),
    }


# ═══════════════════════════════════════════════════════════════
#  PROYECTOS DASHBOARD / FINANCIERO
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def api_proyectos_dashboard(request):
    """KPIs agregados de todos los proyectos visibles, con filtro opcional por mes/año."""
    try:
        qs = _get_proyectos_qs(request.user)

        # Filtrar por mes/año de creación si se proporcionan
        mes = request.GET.get('mes', '')
        anio = request.GET.get('anio', '')
        if anio:
            try:
                qs = qs.filter(created_at__year=int(anio))
            except (ValueError, TypeError):
                pass
        if mes and mes != 'todos':
            try:
                qs = qs.filter(created_at__month=int(mes))
            except (ValueError, TypeError):
                pass

        # No filtrar solo activos — contar todos los del periodo
        activos = qs.filter(status='active')
        data = {
            'total_proyectos': qs.count(),
            'proyectos_ejecucion': activos.count(),
            'proyectos_programados': qs.filter(status='planning').count(),
            'proyectos_completados': qs.filter(status='completed').count(),
            'utilidad_total': float(
                qs.aggregate(t=Sum('partidas__ganancia'))['t'] or 0
            ),
            'costo_total': float(
                qs.aggregate(t=Sum('partidas__costo_total'))['t'] or 0
            ),
            'venta_total': float(
                qs.aggregate(t=Sum('partidas__precio_venta_total'))['t'] or 0
            ),
        }
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def api_proyectos_financiero(request):
    """Lista de proyectos con datos financieros agregados."""
    try:
        qs = _get_proyectos_qs(request.user)
        status_filter = request.GET.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        qs = qs.select_related('usuario').annotate(
            partidas_costo_total=Sum('partidas__costo_total'),
            partidas_venta_total=Sum('partidas__precio_venta_total'),
            partidas_ganancia=Sum('partidas__ganancia'),
        )

        proyectos = []
        for p in qs:
            costo = float(p.partidas_costo_total or 0)
            venta = float(p.partidas_venta_total or 0)
            utilidad = float(p.partidas_ganancia or 0)
            margen = (utilidad / venta * 100) if venta > 0 else 0.0
            proyectos.append({
                'id': p.id,
                'nombre': p.nombre,
                'cliente_nombre': p.cliente_nombre,
                'status': p.status,
                'utilidad_presupuestada': utilidad,
                'costo_total': costo,
                'venta_total': venta,
                'margen': round(margen, 2),
            })

        return JsonResponse({'success': True, 'data': proyectos})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ═══════════════════════════════════════════════════════════════
#  PROYECTOS CRUD
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def api_proyectos_lista(request):
    try:
        qs = _get_proyectos_qs(request.user)
        status_filter = request.GET.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        qs = qs.select_related('usuario')
        proyectos = [_proyecto_to_dict(p, include_alerts=True) for p in qs]
        return JsonResponse({'ok': True, 'data': proyectos})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_proyecto_crear(request):
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    nombre = (data.get('nombre') or '').strip()
    if not nombre:
        return JsonResponse({'success': False, 'error': 'El nombre es obligatorio'}, status=400)

    proyecto = Proyecto.objects.create(
        usuario=request.user,
        nombre=nombre,
        descripcion=data.get('descripcion', ''),
        cliente_nombre=data.get('cliente_nombre', ''),
        status=data.get('status', 'planning'),
        utilidad_presupuestada=_dec(data.get('utilidad_presupuestada')),
        fecha_inicio=_parse_date(data.get('fecha_inicio')),
        fecha_fin=_parse_date(data.get('fecha_fin')),
    )
    ProyectoConfiguracion.objects.create(proyecto=proyecto)
    return JsonResponse({'success': True, 'data': _proyecto_to_dict(proyecto, include_alerts=True)})


@login_required
@require_http_methods(["GET"])
def api_proyecto_detalle(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.select_related('usuario', 'configuracion', 'oportunidad').get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    d = _proyecto_to_dict(proyecto, include_alerts=True)

    # KPIs calculados
    partidas_agg = proyecto.partidas.aggregate(
        total_costo=Sum('costo_total'),
        total_venta=Sum('precio_venta_total'),
        total_ganancia=Sum('ganancia'),
    )
    ingresos = proyecto.facturas_ingreso.aggregate(total=Sum('monto'))['total'] or Decimal('0')
    costos_prov = proyecto.facturas_proveedor.aggregate(total=Sum('monto'))['total'] or Decimal('0')
    gastos_aprobados = proyecto.gastos.filter(estado_aprobacion='approved').aggregate(total=Sum('monto'))['total'] or Decimal('0')
    utilidad_real = ingresos - costos_prov - gastos_aprobados
    margen = (utilidad_real / ingresos * 100) if ingresos > 0 else Decimal('0')

    # Gastado: OCs (no canceladas) + Gastos operativos (no rechazados, es decir pendientes + aprobados)
    gastado_oc = proyecto.ordenes_compra.exclude(status='cancelled').aggregate(t=Sum('monto_total'))['t'] or Decimal('0')
    gastos_no_rechazados = proyecto.gastos.exclude(estado_aprobacion='rejected').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    gastado_total = float(gastado_oc) + float(gastos_no_rechazados)

    d['current_user_is_supervisor'] = is_supervisor(request.user)
    d['kpis'] = {
        'total_costo_partidas': float(partidas_agg['total_costo'] or 0),
        'total_venta_partidas': float(partidas_agg['total_venta'] or 0),
        'utilidad_presupuestada': float(partidas_agg['total_ganancia'] or 0),
        'ingresos': float(ingresos),
        'costos_proveedor': float(costos_prov),
        'gastos_aprobados': float(gastos_aprobados),
        'utilidad_real': float(utilidad_real),
        'margen': float(margen),
        'alertas_pendientes': proyecto.alertas.filter(resuelta=False).count(),
        'tareas_pendientes': proyecto.tareas_proyecto.exclude(status__in=['completed', 'cancelled']).count(),
        'ordenes_compra_activas': proyecto.ordenes_compra.exclude(status='cancelled').count(),
        # Gastado: OCs (no canceladas) + Gastos operativos (no rechazados)
        'gastado': gastado_total,
        # Cobrado: sum of facturas ingreso
        'cobrado': float(proyecto.facturas_ingreso.aggregate(t=Sum('monto'))['t'] or 0),
        # KPIs operativos
        'avance_pct': _calcular_avance(proyecto),
        'efectividad_pct': _calcular_efectividad(proyecto),
    }

    # Configuracion
    try:
        config = proyecto.configuracion
        d['configuracion'] = {
            'umbral_utilidad_minima': float(config.umbral_utilidad_minima),
            'umbral_alerta_varianza': float(config.umbral_alerta_varianza),
            'umbral_cantidad_critica': float(config.umbral_cantidad_critica),
            'requiere_aprobacion_gastos': config.requiere_aprobacion_gastos,
        }
    except ProyectoConfiguracion.DoesNotExist:
        d['configuracion'] = None

    # Overview enriquecido para el nuevo widget de detalle (no rompe legacy)
    try:
        d['overview'] = _proyecto_overview(proyecto)
    except Exception as e:
        # Conservador: si algo truena al construir el overview, lo dejamos vacío
        # en vez de tirar todo el endpoint. El frontend legacy sigue funcionando.
        d['overview'] = None
        d['overview_error'] = str(e)

    return JsonResponse({'success': True, 'data': d})


# ─── Miembros del proyecto (M2M ProyectoIAMET.miembros) ──────────────────
# Habilita el botón "+" del topbar del detalle. Cualquier usuario autenticado
# con acceso al proyecto puede agregar/quitar miembros (consistente con la
# política existente "no sensible en este CRM").

@login_required
@require_http_methods(["POST"])
def api_proyecto_miembro_agregar(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    try:
        data = json.loads(request.body or '{}')
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    user_id = data.get('user_id')
    if not user_id:
        return JsonResponse({'success': False, 'error': 'user_id requerido'}, status=400)

    try:
        user = User.objects.get(id=int(user_id))
    except (User.DoesNotExist, ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Usuario no encontrado'}, status=404)

    # Si el usuario ya es el `usuario` principal del proyecto, no agregamos
    # duplicado al M2M (queda implícito como ingeniero responsable).
    if proyecto.usuario_id == user.id:
        return JsonResponse({'success': True, 'already_member': True, 'user': _get_user_card(user)})

    proyecto.miembros.add(user)
    return JsonResponse({'success': True, 'user': _get_user_card(user)})


@login_required
@require_http_methods(["POST", "DELETE"])
def api_proyecto_miembro_quitar(request, proyecto_id, user_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    try:
        user = User.objects.get(id=int(user_id))
    except (User.DoesNotExist, ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Usuario no encontrado'}, status=404)

    proyecto.miembros.remove(user)
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["POST"])
def api_proyecto_actualizar(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    updatable = ['nombre', 'descripcion', 'cliente_nombre', 'status']
    for field in updatable:
        if field in data:
            setattr(proyecto, field, data[field])

    if 'utilidad_presupuestada' in data:
        proyecto.utilidad_presupuestada = _dec(data['utilidad_presupuestada'])
    if 'fecha_inicio' in data:
        proyecto.fecha_inicio = _parse_date(data['fecha_inicio'])
    if 'fecha_fin' in data:
        proyecto.fecha_fin = _parse_date(data['fecha_fin'])

    proyecto.save()

    # Actualizar configuracion si viene
    config_data = data.get('configuracion')
    if config_data and isinstance(config_data, dict):
        config, _ = ProyectoConfiguracion.objects.get_or_create(proyecto=proyecto)
        if 'umbral_utilidad_minima' in config_data:
            config.umbral_utilidad_minima = _dec(config_data['umbral_utilidad_minima'], Decimal('15'))
        if 'umbral_alerta_varianza' in config_data:
            config.umbral_alerta_varianza = _dec(config_data['umbral_alerta_varianza'], Decimal('5'))
        if 'umbral_cantidad_critica' in config_data:
            config.umbral_cantidad_critica = _dec(config_data['umbral_cantidad_critica'], Decimal('5'))
        if 'requiere_aprobacion_gastos' in config_data:
            config.requiere_aprobacion_gastos = bool(config_data['requiere_aprobacion_gastos'])
        config.save()

    return JsonResponse({'success': True, 'data': _proyecto_to_dict(proyecto, include_alerts=True)})


@login_required
@require_http_methods(["POST"])
def api_proyecto_eliminar(request, proyecto_id):
    try:
        try:
            proyecto = Proyecto.objects.select_related('usuario', 'oportunidad').get(id=proyecto_id)
        except Exception:
            proyecto = Proyecto.objects.select_related('usuario').get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Proyecto no encontrado'}, status=404)

    # Solo el creador del proyecto o el responsable de la oportunidad puede eliminar
    puede_eliminar = False
    if proyecto.usuario_id == request.user.id:
        puede_eliminar = True
    try:
        if proyecto.oportunidad and proyecto.oportunidad.usuario_id == request.user.id:
            puede_eliminar = True
    except Exception:
        pass
    if is_supervisor(request.user):
        puede_eliminar = True

    if not puede_eliminar:
        return JsonResponse({'ok': False, 'error': 'Solo el creador o supervisor puede eliminarlo'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'ok': False, 'error': 'JSON invalido'}, status=400)

    motivo = (data.get('motivo') or '').strip()
    if not motivo:
        return JsonResponse({'ok': False, 'error': 'Debes documentar el motivo de eliminacion'}, status=400)

    proyecto_nombre = proyecto.nombre
    cliente_nombre = proyecto.cliente_nombre or ''
    usuario_nombre = (request.user.first_name + ' ' + request.user.last_name).strip() or request.user.username

    try:
        from .models import Notificacion
        from django.contrib.auth.models import User as AuthUser
        admins = AuthUser.objects.filter(
            Q(is_superuser=True) | Q(groups__name='Supervisores')
        ).distinct().exclude(id=request.user.id)
        for admin in admins:
            Notificacion.objects.create(
                usuario_destinatario=admin,
                usuario_remitente=request.user,
                tipo='sistema',
                titulo=f'Proyecto eliminado: {proyecto_nombre}',
                mensaje=f'{usuario_nombre} elimino el proyecto "{proyecto_nombre}" (Cliente: {cliente_nombre}).\n\nMotivo: {motivo}',
            )
    except Exception:
        pass  # No fallar si la notificacion no se puede crear

    proyecto.delete()
    return JsonResponse({'ok': True, 'data': {'deleted': proyecto_id}})


# ═══════════════════════════════════════════════════════════════
#  PARTIDAS
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def api_partidas_lista(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    partidas = list(proyecto.partidas.all())
    items = [_partida_to_dict(p) for p in partidas]

    # Calcular totales en Python (evita problemas con Sum/Decimal en MySQL)
    t_costo = Decimal('0')
    t_venta = Decimal('0')
    for p in partidas:
        cu = p.costo_unitario if p.costo_unitario else Decimal('0')
        vu = p.precio_venta_unitario if p.precio_venta_unitario else Decimal('0')
        q = p.cantidad if p.cantidad else Decimal('0')
        t_costo += cu * q
        t_venta += vu * q
        print(f'[PARTIDA] {p.descripcion[:30]} | cu={cu} vu={vu} q={q} | costo={cu*q} venta={vu*q} ganancia={(vu-cu)*q}')
    t_ganancia = t_venta - t_costo
    print(f'[TOTALES] costo={t_costo} venta={t_venta} ganancia={t_ganancia}')

    return JsonResponse({
        'success': True,
        'data': {
            'partidas': items,
            'totales': {
                'costo_total': float(t_costo),
                'precio_venta_total': float(t_venta),
                'ganancia': float(t_ganancia),
            },
        },
    })


@login_required
@require_http_methods(["POST"])
def api_partida_crear(request):
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    proyecto_id = data.get('proyecto_id')
    if not proyecto_id:
        return JsonResponse({'success': False, 'error': 'proyecto_id es obligatorio'}, status=400)

    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    cantidad = _dec(data.get('cantidad'), Decimal('1'))
    cantidad_pendiente = _dec(data.get('cantidad_pendiente'), cantidad)

    partida = ProyectoPartida(
        proyecto=proyecto,
        categoria=data.get('categoria', 'equipamiento'),
        descripcion=data.get('descripcion', ''),
        marca=data.get('marca', ''),
        numero_parte=data.get('numero_parte', ''),
        cantidad=cantidad,
        cantidad_pendiente=cantidad_pendiente,
        precio_lista=_dec(data.get('precio_lista')),
        descuento=_dec(data.get('descuento')),
        costo_unitario=_dec(data.get('costo_unitario')),
        precio_venta_unitario=_dec(data.get('precio_venta_unitario')),
        proveedor=data.get('proveedor', ''),
        status=data.get('status', 'pending'),
    )
    partida.save()  # calcular_totales via model save()

    return JsonResponse({'success': True, 'data': _partida_to_dict(partida)})


@login_required
@require_http_methods(["POST"])
def api_partida_actualizar(request, partida_id):
    try:
        partida = ProyectoPartida.objects.select_related('proyecto').get(id=partida_id)
    except ProyectoPartida.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Partida no encontrada'}, status=404)

    if not _check_access(request.user, partida.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    str_fields = ['categoria', 'descripcion', 'marca', 'numero_parte', 'proveedor', 'status']
    for field in str_fields:
        if field in data:
            setattr(partida, field, data[field])

    dec_fields = ['cantidad', 'cantidad_pendiente', 'precio_lista', 'descuento',
                  'costo_unitario', 'precio_venta_unitario']
    for field in dec_fields:
        if field in data:
            setattr(partida, field, _dec(data[field]))

    partida.save()  # auto-recalculates totals

    return JsonResponse({'success': True, 'data': _partida_to_dict(partida)})


@login_required
@require_http_methods(["POST"])
def api_partida_eliminar(request, partida_id):
    try:
        partida = ProyectoPartida.objects.select_related('proyecto').get(id=partida_id)
    except ProyectoPartida.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Partida no encontrada'}, status=404)

    if not _check_access(request.user, partida.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    partida.delete()
    return JsonResponse({'success': True, 'data': {'deleted': partida_id}})


# ═══════════════════════════════════════════════════════════════
#  ORDENES DE COMPRA
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def api_oc_lista(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    ordenes = proyecto.ordenes_compra.select_related('partida').all()
    items = [_oc_to_dict(oc) for oc in ordenes]
    return JsonResponse({'success': True, 'data': items})


@login_required
@require_http_methods(["POST"])
def api_oc_crear(request):
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    proyecto_id = data.get('proyecto_id')
    partida_id = data.get('partida_id')
    if not proyecto_id or not partida_id:
        return JsonResponse({'success': False, 'error': 'proyecto_id y partida_id son obligatorios'}, status=400)

    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    try:
        partida = ProyectoPartida.objects.get(id=partida_id, proyecto=proyecto)
    except ProyectoPartida.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Partida no encontrada en este proyecto'}, status=404)

    cantidad_oc = _dec(data.get('cantidad'), Decimal('0'))
    if cantidad_oc <= 0:
        return JsonResponse({'success': False, 'error': 'La cantidad debe ser mayor a 0'}, status=400)

    if cantidad_oc > partida.cantidad_pendiente:
        return JsonResponse({
            'success': False,
            'error': f'Cantidad excede pendiente ({float(partida.cantidad_pendiente)} disponibles)'
        }, status=400)

    oc = ProyectoOrdenCompra(
        proyecto=proyecto,
        partida=partida,
        numero_oc=data.get('numero_oc', ''),
        proveedor=data.get('proveedor', ''),
        cantidad=cantidad_oc,
        precio_unitario=_dec(data.get('precio_unitario')),
        status=data.get('status', 'draft'),
        fecha_emision=_parse_date(data.get('fecha_emision')),
        fecha_entrega_esperada=_parse_date(data.get('fecha_entrega_esperada')),
        fecha_entrega_real=_parse_date(data.get('fecha_entrega_real')),
        notas=data.get('notas', ''),
    )
    oc.save()  # auto-calc monto_total + auto-generate numero_oc

    # Deduct from partida cantidad_pendiente
    partida.cantidad_pendiente -= cantidad_oc
    if partida.cantidad_pendiente <= 0:
        partida.cantidad_pendiente = Decimal('0')
        partida.status = 'closed'
    partida.save()

    return JsonResponse({'success': True, 'data': _oc_to_dict(oc)})


@login_required
@require_http_methods(["POST"])
def api_oc_actualizar(request, oc_id):
    try:
        oc = ProyectoOrdenCompra.objects.select_related('proyecto', 'partida').get(id=oc_id)
    except ProyectoOrdenCompra.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Orden de compra no encontrada'}, status=404)

    if not _check_access(request.user, oc.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    str_fields = ['proveedor', 'status', 'notas']
    for field in str_fields:
        if field in data:
            setattr(oc, field, data[field])

    if 'precio_unitario' in data:
        oc.precio_unitario = _dec(data['precio_unitario'])
    if 'cantidad' in data:
        oc.cantidad = _dec(data['cantidad'])

    date_fields = ['fecha_emision', 'fecha_entrega_esperada', 'fecha_entrega_real']
    for field in date_fields:
        if field in data:
            setattr(oc, field, _parse_date(data[field]))

    oc.save()  # auto-recalc monto_total

    return JsonResponse({'success': True, 'data': _oc_to_dict(oc)})


@login_required
@require_http_methods(["DELETE"])
def api_oc_eliminar(request, oc_id):
    try:
        oc = ProyectoOrdenCompra.objects.select_related('proyecto').get(id=oc_id)
    except ProyectoOrdenCompra.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'OC no encontrada'}, status=404)
    if not _check_access(request.user, oc.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    oc.delete()
    return JsonResponse({'success': True})


# ═══════════════════════════════════════════════════════════════
#  FACTURAS PROVEEDOR
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def api_facturas_proveedor_lista(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    facturas = proyecto.facturas_proveedor.all()
    items = [_factura_prov_to_dict(f) for f in facturas]
    return JsonResponse({'success': True, 'data': items})


@login_required
@require_http_methods(["POST"])
def api_factura_proveedor_crear(request):
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    proyecto_id = data.get('proyecto_id')
    if not proyecto_id:
        return JsonResponse({'success': False, 'error': 'proyecto_id es obligatorio'}, status=400)

    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    fecha_factura = _parse_date(data.get('fecha_factura'))
    if not fecha_factura:
        return JsonResponse({'success': False, 'error': 'fecha_factura es obligatoria'}, status=400)

    orden_compra = None
    oc_id = data.get('orden_compra_id')
    if oc_id:
        try:
            orden_compra = ProyectoOrdenCompra.objects.get(id=oc_id, proyecto=proyecto)
        except ProyectoOrdenCompra.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Orden de compra no encontrada'}, status=404)

    factura = ProyectoFacturaProveedor(
        proyecto=proyecto,
        orden_compra=orden_compra,
        numero_factura=data.get('numero_factura', ''),
        proveedor=data.get('proveedor', ''),
        monto=_dec(data.get('monto')),
        monto_presupuestado=_dec(data.get('monto_presupuestado')) if data.get('monto_presupuestado') else None,
        fecha_factura=fecha_factura,
        fecha_vencimiento=_parse_date(data.get('fecha_vencimiento')),
        fecha_pago=_parse_date(data.get('fecha_pago')),
        status=data.get('status', 'pending'),
        notas=data.get('notas', ''),
    )
    factura.save()  # auto-calc varianza via model save()

    # Si la varianza_porcentaje > 5, crear alerta
    if factura.varianza_porcentaje is not None and factura.varianza_porcentaje > 5:
        ProyectoAlerta.objects.create(
            proyecto=proyecto,
            tipo_alerta='budget_variance',
            severidad='warning' if factura.varianza_porcentaje <= 15 else 'critical',
            titulo=f'Varianza de {float(factura.varianza_porcentaje):.1f}% en factura {factura.numero_factura}',
            mensaje=(
                f'La factura {factura.numero_factura} de {factura.proveedor} tiene una varianza '
                f'de {float(factura.varianza_porcentaje):.1f}% respecto al presupuesto. '
                f'Monto: ${float(factura.monto):,.2f}, '
                f'Presupuestado: ${float(factura.monto_presupuestado):,.2f}'
            ),
            entidad_tipo='factura_proveedor',
            entidad_id=factura.id,
        )

    return JsonResponse({'success': True, 'data': _factura_prov_to_dict(factura)})


@login_required
@require_http_methods(["POST"])
def api_factura_proveedor_actualizar(request, factura_id):
    try:
        factura = ProyectoFacturaProveedor.objects.select_related('proyecto').get(id=factura_id)
    except ProyectoFacturaProveedor.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Factura no encontrada'}, status=404)

    if not _check_access(request.user, factura.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    str_fields = ['numero_factura', 'proveedor', 'status', 'notas']
    for field in str_fields:
        if field in data:
            setattr(factura, field, data[field])

    if 'monto' in data:
        factura.monto = _dec(data['monto'])
    if 'monto_presupuestado' in data:
        factura.monto_presupuestado = _dec(data['monto_presupuestado']) if data['monto_presupuestado'] else None

    date_fields = ['fecha_factura', 'fecha_vencimiento', 'fecha_pago']
    for field in date_fields:
        if field in data:
            setattr(factura, field, _parse_date(data[field]))

    if 'orden_compra_id' in data:
        oc_id = data['orden_compra_id']
        if oc_id:
            try:
                factura.orden_compra = ProyectoOrdenCompra.objects.get(id=oc_id, proyecto=factura.proyecto)
            except ProyectoOrdenCompra.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Orden de compra no encontrada'}, status=404)
        else:
            factura.orden_compra = None

    factura.save()  # auto-recalc varianza

    return JsonResponse({'success': True, 'data': _factura_prov_to_dict(factura)})


# ═══════════════════════════════════════════════════════════════
#  FACTURAS INGRESO
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def api_facturas_ingreso_lista(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    facturas = proyecto.facturas_ingreso.all()
    items = [_factura_ingreso_to_dict(f) for f in facturas]
    return JsonResponse({'success': True, 'data': items})


@login_required
@require_http_methods(["POST"])
def api_factura_ingreso_crear(request):
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    proyecto_id = data.get('proyecto_id')
    if not proyecto_id:
        return JsonResponse({'success': False, 'error': 'proyecto_id es obligatorio'}, status=400)

    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    fecha_factura = _parse_date(data.get('fecha_factura'))
    if not fecha_factura:
        return JsonResponse({'success': False, 'error': 'fecha_factura es obligatoria'}, status=400)

    factura = ProyectoFacturaIngreso.objects.create(
        proyecto=proyecto,
        numero_factura=data.get('numero_factura', ''),
        monto=_dec(data.get('monto')),
        fecha_factura=fecha_factura,
        fecha_vencimiento=_parse_date(data.get('fecha_vencimiento')),
        fecha_pago=_parse_date(data.get('fecha_pago')),
        status=data.get('status', 'emitted'),
        metodo_pago=data.get('metodo_pago', ''),
        notas=data.get('notas', ''),
    )

    return JsonResponse({'success': True, 'data': _factura_ingreso_to_dict(factura)})


@login_required
@require_http_methods(["POST"])
def api_factura_ingreso_actualizar(request, factura_id):
    try:
        factura = ProyectoFacturaIngreso.objects.select_related('proyecto').get(id=factura_id)
    except ProyectoFacturaIngreso.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Factura no encontrada'}, status=404)

    if not _check_access(request.user, factura.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    str_fields = ['numero_factura', 'status', 'metodo_pago', 'notas']
    for field in str_fields:
        if field in data:
            setattr(factura, field, data[field])

    if 'monto' in data:
        factura.monto = _dec(data['monto'])

    date_fields = ['fecha_factura', 'fecha_vencimiento', 'fecha_pago']
    for field in date_fields:
        if field in data:
            setattr(factura, field, _parse_date(data[field]))

    factura.save()

    return JsonResponse({'success': True, 'data': _factura_ingreso_to_dict(factura)})


# ═══════════════════════════════════════════════════════════════
#  GASTOS
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def api_gastos_lista(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    gastos = proyecto.gastos.select_related('aprobado_por').all()
    items = [_gasto_to_dict(g) for g in gastos]
    return JsonResponse({'success': True, 'data': items})


@login_required
@require_http_methods(["POST"])
def api_gasto_crear(request):
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    proyecto_id = data.get('proyecto_id')
    if not proyecto_id:
        return JsonResponse({'success': False, 'error': 'proyecto_id es obligatorio'}, status=400)

    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    fecha_gasto = _parse_date(data.get('fecha_gasto'))
    if not fecha_gasto:
        return JsonResponse({'success': False, 'error': 'fecha_gasto es obligatoria'}, status=400)

    gasto = ProyectoGasto(
        proyecto=proyecto,
        categoria=data.get('categoria', 'other'),
        descripcion=data.get('descripcion', ''),
        monto=_dec(data.get('monto')),
        monto_presupuestado=_dec(data.get('monto_presupuestado')) if data.get('monto_presupuestado') else None,
        fecha_gasto=fecha_gasto,
        estado_aprobacion=data.get('estado_aprobacion', 'pending'),
        notas=data.get('notas', ''),
    )
    gasto.save()  # auto-calc varianza

    return JsonResponse({'success': True, 'data': _gasto_to_dict(gasto)})


@login_required
@require_http_methods(["POST"])
def api_gasto_actualizar(request, gasto_id):
    try:
        gasto = ProyectoGasto.objects.select_related('proyecto', 'aprobado_por').get(id=gasto_id)
    except ProyectoGasto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Gasto no encontrado'}, status=404)

    if not _check_access(request.user, gasto.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    str_fields = ['categoria', 'descripcion', 'notas']
    for field in str_fields:
        if field in data:
            setattr(gasto, field, data[field])

    if 'monto' in data:
        gasto.monto = _dec(data['monto'])
    if 'monto_presupuestado' in data:
        gasto.monto_presupuestado = _dec(data['monto_presupuestado']) if data['monto_presupuestado'] else None

    if 'fecha_gasto' in data:
        gasto.fecha_gasto = _parse_date(data['fecha_gasto'])

    gasto.save()  # auto-recalc varianza

    return JsonResponse({'success': True, 'data': _gasto_to_dict(gasto)})


@login_required
@require_http_methods(["POST"])
def api_gasto_aprobar(request, gasto_id):
    if not is_supervisor(request.user):
        return JsonResponse({'success': False, 'error': 'Solo supervisores pueden aprobar gastos'}, status=403)

    try:
        gasto = ProyectoGasto.objects.select_related('proyecto', 'aprobado_por').get(id=gasto_id)
    except ProyectoGasto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Gasto no encontrado'}, status=404)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        data = {}

    accion = data.get('accion', 'approved')  # 'approved' or 'rejected'
    if accion not in ('approved', 'rejected'):
        return JsonResponse({'success': False, 'error': 'accion debe ser approved o rejected'}, status=400)

    gasto.estado_aprobacion = accion
    gasto.aprobado_por = request.user
    gasto.fecha_aprobacion = timezone.now()
    gasto.save()

    return JsonResponse({'success': True, 'data': _gasto_to_dict(gasto)})


# ═══════════════════════════════════════════════════════════════
#  FINANCIERO: SYNC DRIVE + UPLOAD MANUAL
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["POST"])
def api_financiero_sync_drive(request, proyecto_id):
    """Sincroniza archivos del drive de la oportunidad vinculada
    que aún no se hayan procesado (OCC y Facturas)."""
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not proyecto.oportunidad_id:
        return JsonResponse({'success': False, 'error': 'El proyecto no tiene oportunidad vinculada'}, status=400)

    from .services_financiero import procesar_archivos_pendientes_oportunidad
    resultado = procesar_archivos_pendientes_oportunidad(proyecto.oportunidad_id)

    return JsonResponse({
        'success': True,
        'total': resultado['total'],
        'procesados': resultado['procesados'],
        'errores': resultado['errores'],
    })


@login_required
@require_http_methods(["POST"])
def api_financiero_upload_oc(request, proyecto_id):
    """Subir un PDF de OC manualmente al financiero del proyecto."""
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'success': False, 'error': 'Archivo requerido'}, status=400)

    from decimal import Decimal as D
    nombre = archivo.name
    ext = nombre.rsplit('.', 1)[-1].lower() if '.' in nombre else ''

    # Extraer datos completos del PDF usando el parser central
    pdf_data = {}
    if ext == 'pdf':
        try:
            import pdfplumber, io
            archivo.seek(0)
            content = archivo.read()
            archivo.seek(0)
            # Crear un file-like para el parser
            from .services_financiero import _extraer_datos_pdf_from_bytes
            pdf_data = _extraer_datos_pdf_from_bytes(content)
        except Exception as exc:
            import logging
            logging.warning(f"[Financiero] Error parseando PDF upload: {exc}")

    from .services_financiero import _extraer_numero_oc, _extraer_proveedor_de_nombre, _acortar_nombre
    numero_oc = pdf_data.get('numero_oc') or _extraer_numero_oc(nombre)
    proveedor = pdf_data.get('proveedor') or _extraer_proveedor_de_nombre(nombre)
    monto = pdf_data.get('monto') or D('0')
    fecha = pdf_data.get('fecha')

    # Verificar duplicado
    if ProyectoOrdenCompra.objects.filter(proyecto=proyecto, numero_oc=numero_oc).exists():
        return JsonResponse({'success': False, 'error': f'Ya existe una OC con número {numero_oc} en este proyecto'}, status=409)

    oc = ProyectoOrdenCompra.objects.create(
        proyecto=proyecto,
        partida=None,
        numero_oc=numero_oc,
        proveedor=_acortar_nombre(proveedor),
        cantidad=D('1'),
        precio_unitario=monto,
        monto_total=monto,
        status='emitted',
        fecha_emision=fecha,
        notas=f'Subido manualmente. Archivo: {nombre}',
    )

    return JsonResponse({
        'success': True,
        'data': _oc_to_dict(oc),
        'monto_extraido': float(monto),
    })


@login_required
@require_http_methods(["POST"])
def api_financiero_upload_factura_ingreso(request, proyecto_id):
    """Subir un PDF de Factura de Ingreso manualmente."""
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'success': False, 'error': 'Archivo requerido'}, status=400)

    from decimal import Decimal as D
    nombre = archivo.name
    ext = nombre.rsplit('.', 1)[-1].lower() if '.' in nombre else ''

    pdf_data = {}
    if ext == 'pdf':
        try:
            archivo.seek(0)
            content = archivo.read()
            archivo.seek(0)
            from .services_financiero import _extraer_datos_pdf_from_bytes
            pdf_data = _extraer_datos_pdf_from_bytes(content)
        except Exception as exc:
            import logging
            logging.warning(f"[Financiero] Error parseando Factura PDF: {exc}")

    from .services_financiero import _extraer_numero_factura
    monto = pdf_data.get('monto') or D('0')
    fecha = pdf_data.get('fecha')
    numero = pdf_data.get('numero_factura') or _extraer_numero_factura(nombre)

    # Verificar duplicado
    if ProyectoFacturaIngreso.objects.filter(proyecto=proyecto, numero_factura=numero).exists():
        return JsonResponse({'success': False, 'error': f'Ya existe una factura con número {numero} en este proyecto'}, status=409)

    factura = ProyectoFacturaIngreso.objects.create(
        proyecto=proyecto,
        numero_factura=numero,
        monto=monto,
        fecha_factura=fecha or timezone.localdate(),
        status='emitted',
        notas=f'Subida manualmente. Archivo: {nombre}',
    )

    return JsonResponse({
        'success': True,
        'data': {
            'id': factura.id,
            'numero_factura': factura.numero_factura,
            'monto': float(factura.monto),
            'fecha_factura': str(factura.fecha_factura),
            'status': factura.status,
        },
        'monto_extraido': float(monto),
    })


@login_required
@require_http_methods(["POST"])
def api_financiero_upload_factura_proveedor(request, proyecto_id):
    """Subir un PDF de Factura de Proveedor manualmente."""
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'success': False, 'error': 'Archivo requerido'}, status=400)

    from decimal import Decimal as D
    nombre = archivo.name
    ext = nombre.rsplit('.', 1)[-1].lower() if '.' in nombre else ''

    pdf_data = {}
    if ext == 'pdf':
        try:
            archivo.seek(0)
            content = archivo.read()
            archivo.seek(0)
            from .services_financiero import _extraer_datos_pdf_from_bytes
            pdf_data = _extraer_datos_pdf_from_bytes(content)
        except Exception as exc:
            import logging
            logging.warning(f"[Financiero] Error parseando Factura Proveedor PDF: {exc}")

    from .services_financiero import _extraer_numero_factura, _extraer_proveedor_de_nombre, _acortar_nombre
    monto = pdf_data.get('monto') or D('0')
    fecha = pdf_data.get('fecha')
    numero = pdf_data.get('numero_factura') or _extraer_numero_factura(nombre)
    proveedor = pdf_data.get('proveedor') or _extraer_proveedor_de_nombre(nombre)

    # Verificar duplicado
    if ProyectoFacturaProveedor.objects.filter(proyecto=proyecto, numero_factura=numero).exists():
        return JsonResponse({'success': False, 'error': f'Ya existe una factura de proveedor con número {numero} en este proyecto'}, status=409)

    factura = ProyectoFacturaProveedor.objects.create(
        proyecto=proyecto,
        numero_factura=numero,
        proveedor=_acortar_nombre(proveedor),
        monto=monto,
        fecha_factura=fecha or timezone.localdate(),
        status='received',
        notas=f'Subida manualmente. Archivo: {nombre}',
    )

    return JsonResponse({
        'success': True,
        'data': _factura_prov_to_dict(factura),
        'monto_extraido': float(monto),
    })


@login_required
@require_http_methods(["DELETE"])
def api_factura_proveedor_eliminar(request, factura_id):
    """Elimina una factura de proveedor."""
    try:
        factura = ProyectoFacturaProveedor.objects.select_related('proyecto').get(id=factura_id)
    except ProyectoFacturaProveedor.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Factura no encontrada'}, status=404)

    if not _check_access(request.user, factura.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    factura.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["DELETE"])
def api_factura_ingreso_eliminar(request, factura_id):
    """Elimina una factura de ingreso."""
    try:
        factura = ProyectoFacturaIngreso.objects.select_related('proyecto').get(id=factura_id)
    except ProyectoFacturaIngreso.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Factura no encontrada'}, status=404)

    if not _check_access(request.user, factura.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    factura.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["DELETE"])
def api_gasto_eliminar(request, gasto_id):
    """Elimina un gasto operativo."""
    try:
        gasto = ProyectoGasto.objects.select_related('proyecto').get(id=gasto_id)
    except ProyectoGasto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Gasto no encontrado'}, status=404)

    if not _check_access(request.user, gasto.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    gasto.delete()
    return JsonResponse({'success': True})


# ═══════════════════════════════════════════════════════════════
#  TAREAS PROYECTO
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def api_tareas_proyecto_lista(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    # 1) Tareas propias del proyecto
    tareas = proyecto.tareas_proyecto.select_related('asignado_a').all()
    items = []
    for t in tareas:
        d = _tarea_to_dict(t)
        d['source'] = 'proyecto'
        items.append(d)

    # 2) Tareas de la oportunidad vinculada (modelo Tarea del CRM)
    if proyecto.oportunidad_id:
        from .models import Tarea
        tareas_crm = Tarea.objects.filter(
            oportunidad_id=proyecto.oportunidad_id
        ).select_related('asignado_a', 'creado_por')
        prioridad_map = {'baja': 'low', 'media': 'medium', 'alta': 'high'}
        estado_map = {'pendiente': 'pending', 'iniciada': 'in_progress', 'en_progreso': 'in_progress', 'completada': 'completed', 'cancelada': 'cancelled'}
        for t in tareas_crm:
            resp_name = None
            if t.asignado_a:
                resp_name = (t.asignado_a.first_name + ' ' + t.asignado_a.last_name).strip() or t.asignado_a.username
            items.append({
                'id': t.id,
                'source': 'oportunidad',
                'titulo': t.titulo,
                'descripcion': getattr(t, 'descripcion', ''),
                'prioridad': prioridad_map.get(t.prioridad, 'medium'),
                'status': estado_map.get(t.estado, 'pending'),
                'asignado_a_nombre': resp_name,
                'fecha_limite': _fmt(t.fecha_limite),
            })

    return JsonResponse({'success': True, 'data': items})


@login_required
@require_http_methods(["POST"])
def api_tarea_proyecto_crear(request):
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    proyecto_id = data.get('proyecto_id')
    if not proyecto_id:
        return JsonResponse({'success': False, 'error': 'proyecto_id es obligatorio'}, status=400)

    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    asignado_a = None
    asignado_a_id = data.get('asignado_a_id')
    if asignado_a_id:
        from django.contrib.auth.models import User
        try:
            asignado_a = User.objects.get(id=asignado_a_id)
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Usuario asignado no encontrado'}, status=404)

    tarea = ProyectoTarea.objects.create(
        proyecto=proyecto,
        titulo=data.get('titulo', ''),
        descripcion=data.get('descripcion', ''),
        status=data.get('status', 'pending'),
        prioridad=data.get('prioridad', 'medium'),
        asignado_a=asignado_a,
        fecha_limite=_parse_date(data.get('fecha_limite')),
        horas_estimadas=_dec(data.get('horas_estimadas')) if data.get('horas_estimadas') else None,
        notas=data.get('notas', ''),
    )

    return JsonResponse({'success': True, 'data': _tarea_to_dict(tarea)})


@login_required
@require_http_methods(["POST"])
def api_tarea_proyecto_actualizar(request, tarea_id):
    try:
        tarea = ProyectoTarea.objects.select_related('proyecto', 'asignado_a').get(id=tarea_id)
    except ProyectoTarea.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tarea no encontrada'}, status=404)

    if not _check_access(request.user, tarea.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON invalido'}, status=400)

    str_fields = ['titulo', 'descripcion', 'prioridad', 'notas']
    for field in str_fields:
        if field in data:
            setattr(tarea, field, data[field])

    if 'status' in data:
        new_status = data['status']
        if new_status == 'completed' and tarea.status != 'completed':
            tarea.fecha_completada = timezone.now()
        elif new_status != 'completed':
            tarea.fecha_completada = None
        tarea.status = new_status

    if 'fecha_limite' in data:
        tarea.fecha_limite = _parse_date(data['fecha_limite'])

    if 'horas_estimadas' in data:
        tarea.horas_estimadas = _dec(data['horas_estimadas']) if data['horas_estimadas'] else None
    if 'horas_reales' in data:
        tarea.horas_reales = _dec(data['horas_reales']) if data['horas_reales'] else None

    if 'asignado_a_id' in data:
        asignado_a_id = data['asignado_a_id']
        if asignado_a_id:
            from django.contrib.auth.models import User
            try:
                tarea.asignado_a = User.objects.get(id=asignado_a_id)
            except User.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Usuario asignado no encontrado'}, status=404)
        else:
            tarea.asignado_a = None

    tarea.save()

    return JsonResponse({'success': True, 'data': _tarea_to_dict(tarea)})


# ═══════════════════════════════════════════════════════════════
#  ALERTAS
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def api_alertas_lista(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    qs = proyecto.alertas.all()
    show_all = request.GET.get('all')
    if not show_all:
        qs = qs.filter(resuelta=False)

    items = [_alerta_to_dict(a) for a in qs]
    return JsonResponse({'success': True, 'data': items})


@login_required
@require_http_methods(["POST"])
def api_alerta_resolver(request, alerta_id):
    try:
        alerta = ProyectoAlerta.objects.select_related('proyecto').get(id=alerta_id)
    except ProyectoAlerta.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Alerta no encontrada'}, status=404)

    if not _check_access(request.user, alerta.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    alerta.resuelta = True
    alerta.fecha_resolucion = timezone.now()
    alerta.save()

    return JsonResponse({'success': True, 'data': _alerta_to_dict(alerta)})


# ═══════════════════════════════════════════════════════════════
#  FINANCIEROS (KPIs consolidados)
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def api_proyecto_financieros(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)

    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    # Usar utilidad_presupuestada del proyecto (viene del resumen del Excel)
    utilidad_presupuestada = proyecto.utilidad_presupuestada or Decimal('0')
    ingresos = proyecto.facturas_ingreso.aggregate(total=Sum('monto'))['total'] or Decimal('0')
    costos = proyecto.facturas_proveedor.aggregate(total=Sum('monto'))['total'] or Decimal('0')
    gastos = proyecto.gastos.filter(estado_aprobacion='approved').aggregate(total=Sum('monto'))['total'] or Decimal('0')

    utilidad_real = ingresos - costos - gastos
    costos_gastos = costos + gastos
    margen = (utilidad_real / ingresos * 100) if ingresos > 0 else Decimal('0')
    cobertura = (ingresos / costos_gastos * 100) if costos_gastos > 0 else Decimal('0')
    alertas_pendientes = proyecto.alertas.filter(resuelta=False).count()

    # KPIs operativos
    gastado = float(proyecto.ordenes_compra.exclude(status='cancelled').aggregate(t=Sum('monto_total'))['t'] or 0)
    cobrado = float(ingresos)

    # Costo presupuestado (de partidas)
    costo_presupuestado = float(proyecto.partidas.aggregate(t=Sum('costo_total'))['t'] or 0)

    return JsonResponse({
        'success': True,
        'data': {
            'utilidad_presupuestada': float(utilidad_presupuestada),
            'costo_presupuestado': costo_presupuestado,
            'ingresos': float(ingresos),
            'costos': float(costos),
            'gastos': float(gastos),
            'utilidad_real': float(utilidad_real),
            'margen': float(margen),
            'cobertura': float(cobertura),
            'alertas_pendientes': alertas_pendientes,
            'gastado': gastado,
            'cobrado': cobrado,
            'avance_pct': _calcular_avance(proyecto),
            'efectividad_pct': _calcular_efectividad(proyecto),
        },
    })


# ═══════════════════════════════════════════════════════════════
#  IMPORTAR EXCEL (Volumetria)
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(['POST'])
def api_importar_excel(request):
    """Import volumetria Excel — specific format used by IAMET engineers.

    2-pass approach:
      Pass 1: Pre-scan to find MO cost/venta totals and summary ganancia.
      Pass 2: Import rows — material items get normal costo, MO VENTA items
              get proportionally-distributed costo, MO COST items are skipped.
    """
    try:
        proyecto_id = request.POST.get('proyecto_id')
        if not proyecto_id:
            return JsonResponse({'ok': False, 'error': 'proyecto_id requerido'}, status=400)

        proyecto = Proyecto.objects.get(id=proyecto_id)
        if not _check_access(request.user, proyecto):
            return JsonResponse({'ok': False, 'error': 'Sin permiso'}, status=403)

        archivo = request.FILES.get('archivo')
        if not archivo:
            return JsonResponse({'ok': False, 'error': 'Archivo requerido'}, status=400)

        import openpyxl
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active

        # Get exchange rate USD->MXN
        exchange_rate = Decimal('19.50')  # fallback
        try:
            import urllib.request
            resp = urllib.request.urlopen('https://api.exchangerate-api.com/v4/latest/USD', timeout=5)
            import json as _json
            rates = _json.loads(resp.read())
            exchange_rate = Decimal(str(rates['rates'].get('MXN', 19.50)))
        except Exception:
            pass

        # Also try to get exchange rate from cell I4 which says "Moneda : Dolares 19.50"
        try:
            cell_i4 = str(ws.cell(4, 9).value or '')
            if 'dolar' in cell_i4.lower():
                import re
                match = re.search(r'[\d.]+', cell_i4.replace('Dolares', '').replace('dolares', ''))
                if match:
                    tc_from_excel = Decimal(match.group())
                    if tc_from_excel > 10:  # sanity check
                        exchange_rate = tc_from_excel
        except Exception:
            pass

        # Section category mapping
        SECTION_MAP = {
            'material': 'equipamiento',
            'equipamiento': 'equipamiento',
            'canalizacion': 'accesorios',
            'canalización': 'accesorios',
            'equipos de elevacion': 'equipamiento',
            'equipos de elevación': 'equipamiento',
            'accesorios': 'accesorios',
            'mano de obra': 'mano_obra',
        }

        # ─── PASS 1: Pre-scan for MO totals and summary ─────────────
        total_mo_venta = Decimal('0')
        total_mo_costo = Decimal('0')
        total_ganancia_resumen = Decimal('0')
        mo_cost_header_row = None  # Row of the cost section header

        for r in range(1, ws.max_row + 1):
            cell_a = str(ws.cell(r, 1).value or '').strip().lower()
            cell_d = str(ws.cell(r, 4).value or '').strip().lower()
            cell_e = str(ws.cell(r, 5).value or '').strip().lower()
            cell_h = ws.cell(r, 8).value
            cell_c3 = ws.cell(r, 3).value

            # "TOTAL MANO DE OBRA:" row — col H has total MO venta
            if 'total mano de obra' in cell_a or 'total mano de obra' in cell_d:
                total_mo_venta = _dec(cell_h)

            # "COSTO MANO DE OBRA:" row — col H has total MO cost
            if 'costo mano de obra' in cell_a or 'costo mano de obra' in cell_d:
                total_mo_costo = _dec(cell_h)

            # "Total de Ganancia" in the summary section
            if 'total de ganancia' in cell_a:
                total_ganancia_resumen = _dec(cell_c3)

            # Detect cost header: row with "Descripcion" in col D and "Costo unit"/"Costo unitario" in col E
            if cell_d in ('descripcion', 'descripción') and cell_e in ('costo unit', 'costo unitario'):
                mo_cost_header_row = r

        print(f'[IMPORT PRE-SCAN] MO venta={total_mo_venta}, MO costo={total_mo_costo}, '
              f'ganancia_resumen={total_ganancia_resumen}, cost_header_row={mo_cost_header_row}')

        # MO cost ratio: what fraction of MO venta is cost
        mo_cost_ratio = Decimal('0')
        if total_mo_venta > 0 and total_mo_costo > 0:
            mo_cost_ratio = (total_mo_costo / total_mo_venta).quantize(Decimal('0.000001'))

        # ─── Archive current partidas as a new version before importing ──
        existing_partidas = list(proyecto.partidas.all())
        if existing_partidas:
            last_version = ProyectoVolumetriaVersion.objects.filter(
                proyecto=proyecto
            ).order_by('-version').first()
            next_version = (last_version.version + 1) if last_version else 1

            snapshot = []
            snap_costo = Decimal('0')
            snap_venta = Decimal('0')
            for p in existing_partidas:
                cu = p.costo_unitario or Decimal('0')
                vu = p.precio_venta_unitario or Decimal('0')
                q = p.cantidad or Decimal('0')
                snap_costo += cu * q
                snap_venta += vu * q
                snapshot.append({
                    'categoria': p.categoria,
                    'descripcion': p.descripcion,
                    'marca': p.marca,
                    'numero_parte': p.numero_parte,
                    'cantidad': float(q),
                    'precio_lista': float(p.precio_lista or 0),
                    'descuento': float(p.descuento or 0),
                    'costo_unitario': float(cu),
                    'precio_venta_unitario': float(vu),
                    'ganancia': float((vu - cu) * q),
                    'proveedor': p.proveedor,
                    'status': p.status,
                })
            snap_ganancia = snap_venta - snap_costo
            snap_margen = (snap_ganancia / snap_venta * 100) if snap_venta > 0 else Decimal('0')

            ProyectoVolumetriaVersion.objects.create(
                proyecto=proyecto,
                version=next_version,
                archivo_nombre=archivo.name if hasattr(archivo, 'name') else '',
                subido_por=request.user,
                total_costo=snap_costo,
                total_venta=snap_venta,
                ganancia=snap_ganancia,
                margen=snap_margen,
                num_partidas=len(existing_partidas),
                partidas_json=snapshot,
            )

            # Delete old partidas
            proyecto.partidas.all().delete()

        # ─── PASS 2: Import rows with intelligence ──────────────────
        current_category = 'equipamiento'  # default
        items_created = 0
        errors = []
        total_venta = Decimal('0')
        total_costo = Decimal('0')
        in_mo_cost_zone = False  # True after "TOTAL MANO DE OBRA:" row

        # Detect summary section start
        summary_keywords = ['analisis de costos', 'análisis de costos', 'precio de lista',
                            'precio bajanet', 'ganancia de material', 'total de ganancia']

        for row_idx in range(6, ws.max_row + 1):
            try:
                # Read all cells
                col_a = ws.cell(row_idx, 1).value   # Marca
                col_b = ws.cell(row_idx, 2).value   # No. Parte
                col_c = ws.cell(row_idx, 3).value   # Cantidad
                col_d = ws.cell(row_idx, 4).value   # Descripcion
                col_e = ws.cell(row_idx, 5).value   # Precio Lista
                col_f = ws.cell(row_idx, 6).value   # Descuento
                col_g = ws.cell(row_idx, 7).value   # Unitario (venta)
                col_h = ws.cell(row_idx, 8).value   # Total (venta)
                col_i = ws.cell(row_idx, 9).value   # Desc costo
                col_j = ws.cell(row_idx, 10).value  # Unitario (costo)
                col_k = ws.cell(row_idx, 11).value  # Total (costo)

                col_a_str = str(col_a or '').strip().lower()
                col_d_str = str(col_d or '').strip().lower()

                # Check for summary section — STOP processing
                if any(kw in col_a_str for kw in summary_keywords) or \
                   any(kw in col_d_str for kw in summary_keywords):
                    break

                # Detect "TOTAL MANO DE OBRA:" row — everything after is MO cost zone
                if 'total mano de obra' in col_a_str or 'total mano de obra' in col_d_str:
                    in_mo_cost_zone = True
                    continue

                # Skip all rows in MO cost zone (Tecnico, combustible, casetas, etc.)
                if in_mo_cost_zone:
                    continue

                # Also skip rows that are clearly in the cost header zone
                if mo_cost_header_row and row_idx >= mo_cost_header_row:
                    continue

                # Check for section header
                if col_a and isinstance(col_a, str) and col_a.strip():
                    is_section = False
                    for section_name in SECTION_MAP:
                        if section_name in col_a_str:
                            current_category = SECTION_MAP[section_name]
                            is_section = True
                            break
                    if is_section:
                        continue

                if col_d and isinstance(col_d, str) and col_d.strip():
                    for section_name in SECTION_MAP:
                        if section_name in col_d_str:
                            current_category = SECTION_MAP[section_name]

                # Check for header rows (like "Marca | No. Parte | Cantidad | Descripcion...")
                if col_a_str in ('marca', '') and col_d_str in ('descripcion', 'descripción', ''):
                    if col_a_str == 'marca':
                        continue  # skip header row

                # Validate: must have quantity > 0 and description
                cantidad = None
                try:
                    if col_c is not None:
                        cantidad = Decimal(str(col_c))
                except (ValueError, InvalidOperation):
                    pass

                if not cantidad or cantidad <= 0:
                    continue

                descripcion = str(col_d or '').strip()
                if not descripcion:
                    continue

                # Skip if description is just a number (subtotal row)
                try:
                    float(descripcion.replace(',', ''))
                    continue  # it's a number, skip
                except ValueError:
                    pass

                # Skip "Total de..." rows
                if descripcion.lower().startswith('total de') or \
                   descripcion.lower().startswith('total '):
                    continue

                # Parse prices
                marca = str(col_a or '').strip()
                numero_parte = str(col_b or '').strip()

                precio_lista = _dec(col_e)
                descuento_dec = _dec(col_f)  # decimal like 0.3
                precio_venta_unit = _dec(col_g)
                costo_unit = _dec(col_j)
                total_venta_excel = _dec(col_h)
                total_costo_excel = _dec(col_k)

                # Determine if this is a MO VENTA item:
                # category is mano_obra AND has venta but no costo columns
                is_mo_venta = (
                    current_category == 'mano_obra'
                    and total_venta_excel > 0
                    and total_costo_excel == 0
                    and costo_unit == 0
                )

                if is_mo_venta:
                    # Distribute MO cost proportionally across MO venta items
                    # item_costo_total = item_venta_total * (total_mo_costo / total_mo_venta)
                    total_costo_excel = (total_venta_excel * mo_cost_ratio).quantize(Decimal('0.01'))
                    if cantidad > 0:
                        costo_unit = (total_costo_excel / cantidad).quantize(Decimal('0.01'))

                # Fallbacks (for material items)
                if precio_venta_unit == 0 and precio_lista > 0:
                    precio_venta_unit = precio_lista
                if total_venta_excel == 0 and precio_venta_unit > 0:
                    total_venta_excel = precio_venta_unit * cantidad
                if total_costo_excel == 0 and costo_unit > 0:
                    total_costo_excel = costo_unit * cantidad

                # Convert to MXN
                precio_lista_mxn = (precio_lista * exchange_rate).quantize(Decimal('0.01'))
                precio_venta_mxn = (precio_venta_unit * exchange_rate).quantize(Decimal('0.01'))
                costo_mxn = (costo_unit * exchange_rate).quantize(Decimal('0.01'))
                descuento_pct = (descuento_dec * Decimal('100')).quantize(Decimal('0.01'))

                venta_total = (total_venta_excel * exchange_rate).quantize(Decimal('0.01'))
                costo_total = (total_costo_excel * exchange_rate).quantize(Decimal('0.01'))
                ganancia = (venta_total - costo_total).quantize(Decimal('0.01'))

                ProyectoPartida.objects.create(
                    proyecto=proyecto,
                    categoria=current_category,
                    descripcion=descripcion,
                    marca=marca,
                    numero_parte=numero_parte,
                    cantidad=cantidad,
                    cantidad_pendiente=cantidad,
                    precio_lista=precio_lista_mxn,
                    descuento=descuento_pct,
                    costo_unitario=costo_mxn,
                    precio_venta_unitario=precio_venta_mxn,
                    costo_total=costo_total,
                    precio_venta_total=venta_total,
                    ganancia=ganancia,
                    proveedor=marca if marca else '',
                )
                items_created += 1
                total_venta += venta_total
                total_costo += costo_total

            except Exception as e:
                errors.append(f'Fila {row_idx}: {str(e)}')

        # ─── Read "Total de Ganancia" from summary ──────────────────
        ganancia_excel = Decimal('0')
        if total_ganancia_resumen > 0:
            ganancia_excel = total_ganancia_resumen * exchange_rate
        else:
            # Fallback: scan bottom of sheet
            for r in range(max(ws.max_row - 30, 1), ws.max_row + 1):
                cell_a = str(ws.cell(r, 1).value or '').strip().lower()
                if 'total de ganancia' in cell_a:
                    val = ws.cell(r, 3).value
                    if val:
                        try:
                            ganancia_excel = _dec(val) * exchange_rate
                        except Exception:
                            pass
                    break

        # Fallback: sum partidas ganancia if no summary found
        if ganancia_excel == 0:
            ganancia_excel = proyecto.partidas.aggregate(total=Sum('ganancia'))['total'] or Decimal('0')

        proyecto.utilidad_presupuestada = ganancia_excel.quantize(Decimal('0.01'))
        proyecto.save(update_fields=['utilidad_presupuestada'])
        print(f'[IMPORT] Utilidad guardada: {proyecto.utilidad_presupuestada} '
              f'(ganancia_excel={ganancia_excel}, mo_ratio={mo_cost_ratio})')

        ganancia_mxn = total_venta - total_costo

        return JsonResponse({
            'ok': True,
            'items_created': items_created,
            'total_venta_mxn': float(total_venta.quantize(Decimal('0.01'))),
            'total_costo_mxn': float(total_costo.quantize(Decimal('0.01'))),
            'ganancia_mxn': float(ganancia_mxn.quantize(Decimal('0.01'))),
            'exchange_rate': float(exchange_rate),
            'mo_venta': float(total_mo_venta),
            'mo_costo': float(total_mo_costo),
            'mo_cost_ratio': float(mo_cost_ratio),
            'errors': errors,
        })
    except Proyecto.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Proyecto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


# ─── Historial de versiones de volumetria ─────────────────────

@login_required
@require_http_methods(["GET"])
def api_volumetria_versiones(request, proyecto_id):
    """GET: Historial de versiones de volumetria"""
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Proyecto no encontrado'}, status=404)
    if not _check_access(request.user, proyecto):
        return JsonResponse({'ok': False, 'error': 'Sin acceso'}, status=403)

    versiones = ProyectoVolumetriaVersion.objects.filter(proyecto=proyecto).select_related('subido_por')
    data = []
    for v in versiones:
        data.append({
            'version': v.version,
            'archivo': v.archivo_nombre,
            'subido_por': (v.subido_por.first_name + ' ' + v.subido_por.last_name).strip() if v.subido_por else '—',
            'fecha': v.fecha.isoformat() if v.fecha else None,
            'total_costo': float(v.total_costo),
            'total_venta': float(v.total_venta),
            'ganancia': float(v.ganancia),
            'margen': float(v.margen),
            'num_partidas': v.num_partidas,
            'partidas_json': v.partidas_json or [],
        })

    # Add current version info
    current = list(proyecto.partidas.all())
    if current:
        c_costo = sum((p.costo_unitario or Decimal('0')) * (p.cantidad or Decimal('0')) for p in current)
        c_venta = sum((p.precio_venta_unitario or Decimal('0')) * (p.cantidad or Decimal('0')) for p in current)
        c_gan = c_venta - c_costo
        # Get last import info
        last_ver = versiones.first()  # ordered by -version
        last_user = ''
        last_date = None
        if last_ver:
            # The current is the one AFTER the last version
            last_user = (last_ver.subido_por.first_name + ' ' + last_ver.subido_por.last_name).strip() if last_ver.subido_por else '—'
            last_date = last_ver.fecha.isoformat() if last_ver.fecha else None
        data.insert(0, {
            'version': len(data) + 1,
            'archivo': 'Actual',
            'subido_por': last_user,
            'fecha': last_date,
            'total_costo': float(c_costo),
            'total_venta': float(c_venta),
            'ganancia': float(c_gan),
            'margen': float((c_gan / c_venta * 100) if c_venta > 0 else 0),
            'num_partidas': len(current),
            'is_current': True,
        })

    return JsonResponse({'ok': True, 'data': data})


@login_required
@require_http_methods(["POST"])
def api_restaurar_version(request, proyecto_id):
    """POST: Restaurar volumetria a una version anterior"""
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Proyecto no encontrado'}, status=404)
    if not _check_access(request.user, proyecto):
        return JsonResponse({'ok': False, 'error': 'Sin acceso'}, status=403)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'ok': False, 'error': 'JSON invalido'}, status=400)

    version_num = data.get('version')
    if not version_num:
        return JsonResponse({'ok': False, 'error': 'Version requerida'}, status=400)

    try:
        version = ProyectoVolumetriaVersion.objects.get(proyecto=proyecto, version=version_num)
    except ProyectoVolumetriaVersion.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Version no encontrada'}, status=404)

    # Archive current partidas first
    existing = list(proyecto.partidas.all())
    if existing:
        last_ver = ProyectoVolumetriaVersion.objects.filter(proyecto=proyecto).order_by('-version').first()
        next_ver = (last_ver.version + 1) if last_ver else 1
        snapshot = []
        sc = Decimal('0')
        sv = Decimal('0')
        for p in existing:
            cu = p.costo_unitario or Decimal('0')
            vu = p.precio_venta_unitario or Decimal('0')
            q = p.cantidad or Decimal('0')
            sc += cu * q
            sv += vu * q
            snapshot.append({
                'categoria': p.categoria, 'descripcion': p.descripcion, 'marca': p.marca,
                'numero_parte': p.numero_parte, 'cantidad': float(q),
                'precio_lista': float(p.precio_lista or 0), 'descuento': float(p.descuento or 0),
                'costo_unitario': float(cu), 'precio_venta_unitario': float(vu),
                'ganancia': float((vu - cu) * q), 'proveedor': p.proveedor, 'status': p.status,
            })
        sg = sv - sc
        ProyectoVolumetriaVersion.objects.create(
            proyecto=proyecto, version=next_ver, archivo_nombre='Antes de restaurar v' + str(version_num),
            subido_por=request.user, total_costo=sc, total_venta=sv, ganancia=sg,
            margen=(sg / sv * 100) if sv > 0 else Decimal('0'),
            num_partidas=len(existing), partidas_json=snapshot,
        )
        proyecto.partidas.all().delete()

    # Restore partidas from version snapshot
    restored = 0
    for item in (version.partidas_json or []):
        ProyectoPartida.objects.create(
            proyecto=proyecto,
            categoria=item.get('categoria', 'otros'),
            descripcion=item.get('descripcion', ''),
            marca=item.get('marca', ''),
            numero_parte=item.get('numero_parte', ''),
            cantidad=Decimal(str(item.get('cantidad', 0))),
            cantidad_pendiente=Decimal(str(item.get('cantidad', 0))),
            precio_lista=Decimal(str(item.get('precio_lista', 0))),
            descuento=Decimal(str(item.get('descuento', 0))),
            costo_unitario=Decimal(str(item.get('costo_unitario', 0))),
            precio_venta_unitario=Decimal(str(item.get('precio_venta_unitario', 0))),
            proveedor=item.get('proveedor', ''),
        )
        restored += 1

    # Update project utilidad
    total_g = sum(
        ((p.precio_venta_unitario or Decimal('0')) - (p.costo_unitario or Decimal('0'))) * (p.cantidad or Decimal('0'))
        for p in proyecto.partidas.all()
    )
    proyecto.utilidad_presupuestada = total_g
    proyecto.save(update_fields=['utilidad_presupuestada'])

    return JsonResponse({'ok': True, 'restored': restored})


# ═══════════════════════════════════════════════════════════════
#  CFDI XML → PDF (modulo independiente de prueba)
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def cfdi_convertidor_page(request):
    """Pagina standalone con form para subir XML CFDI y descargar PDF."""
    from django.shortcuts import render
    return render(request, 'crm/cfdi_convertidor.html')


@login_required
@require_http_methods(["POST"])
def api_cfdi_convertir(request):
    """Recibe un XML CFDI, retorna PDF inline."""
    from django.http import HttpResponse
    archivo = request.FILES.get('xml')
    if not archivo:
        return JsonResponse({'success': False, 'error': 'Archivo XML requerido'}, status=400)

    try:
        xml_bytes = archivo.read()
        from .services_cfdi import generate_cfdi_pdf
        pdf_bytes, data = generate_cfdi_pdf(xml_bytes)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Error generando PDF: {str(e)}'}, status=500)

    # Nombre de archivo: Factura_<RFC_EMISOR>_<SERIE><FOLIO>_<RFC_RECEPTOR>.pdf
    rfc_e = data.get('emisor', {}).get('rfc', 'EMISOR')
    rfc_r = data.get('receptor', {}).get('rfc', 'RECEPTOR')
    serie_folio = f"{data.get('serie', '')}{data.get('folio', '')}"
    filename = f"Factura_{rfc_e}_{serie_folio}_{rfc_r}.pdf"

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    # Inline para verlo en el navegador; descarga con el nombre correcto si el usuario usa "Guardar como"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


# ══════════════════════════════════════════════════════════════
#  LEVANTAMIENTOS (wizard de 5 fases)
# ══════════════════════════════════════════════════════════════

def _user_es_solo_lectura_levantamiento(user):
    """Vendedores tienen acceso de SOLO LECTURA a los levantamientos:
    pueden ver fases completadas y descargar PDFs, pero no editar ni borrar.
    Los ingenieros, supervisores, administradores y superusers tienen
    acceso completo.
    """
    if not user or not user.is_authenticated:
        return True
    if user.is_superuser:
        return False
    if is_supervisor(user):
        return False
    try:
        if hasattr(user, 'userprofile') and user.userprofile and getattr(user.userprofile, 'rol', None):
            return user.userprofile.rol == 'vendedor'
    except Exception:
        pass
    return False


def _lev_to_dict(lev, include_evidencias=False):
    # Datos del proyecto padre (útil para Fase 4 — programa de obra — y PDFs)
    proy = lev.proyecto
    d = {
        'id': lev.id,
        'proyecto_id': lev.proyecto_id,
        'proyecto_nombre': getattr(proy, 'nombre', '') if proy else '',
        'proyecto_fecha_inicio': _fmt(getattr(proy, 'fecha_inicio', None)) if proy else None,
        'proyecto_fecha_fin': _fmt(getattr(proy, 'fecha_fin', None)) if proy else None,
        'nombre': lev.nombre,
        'status': lev.status,
        'status_label': lev.get_status_display(),
        'fase_actual': lev.fase_actual,
        'fase1_data': lev.fase1_data or {},
        'fase2_data': lev.fase2_data or {},
        'fase3_data': lev.fase3_data or {},
        'fase4_data': lev.fase4_data or {},
        'fase5_data': lev.fase5_data or {},
        'creado_por_id': lev.creado_por_id,
        'creado_por_nombre': (
            (lev.creado_por.get_full_name() or lev.creado_por.username)
            if lev.creado_por else None
        ),
        'actualizado_por_id': getattr(lev, 'actualizado_por_id', None),
        'actualizado_por_nombre': (
            (lev.actualizado_por.get_full_name() or lev.actualizado_por.username)
            if getattr(lev, 'actualizado_por_id', None) else None
        ),
        'fecha_creacion': _fmt(lev.fecha_creacion),
        'fecha_actualizacion': _fmt(lev.fecha_actualizacion),
    }
    if include_evidencias:
        d['evidencias'] = [_evidencia_to_dict(e) for e in lev.evidencias.all()]
    return d


def _evidencia_to_dict(e):
    return {
        'id': e.id,
        'url': e.archivo.url if e.archivo else None,
        'nombre_original': e.nombre_original,
        'comentario': e.comentario,
        'producto_idx': e.producto_idx,
        'subido_por_id': e.subido_por_id,
        'subido_por_nombre': (
            (e.subido_por.get_full_name() or e.subido_por.username)
            if e.subido_por else None
        ),
        'fecha_subida': _fmt(e.fecha_subida),
    }


@login_required
@require_http_methods(["GET"])
def api_levantamientos_lista(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)
    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    qs = proyecto.levantamientos.select_related('creado_por', 'actualizado_por').all()
    return JsonResponse({
        'ok': True,
        'data': [_lev_to_dict(l) for l in qs],
        # Flag de UI: el frontend usa esto para decidir si renderiza el
        # wizard en modo editor o en modo lectura (vendedores).
        'puede_editar': not _user_es_solo_lectura_levantamiento(request.user),
    })


@login_required
@require_http_methods(["POST"])
def api_levantamiento_crear(request, proyecto_id):
    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)
    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    if _user_es_solo_lectura_levantamiento(request.user):
        return JsonResponse({'success': False, 'error': 'Los vendedores no pueden crear levantamientos'}, status=403)

    try:
        data = json.loads(request.body or '{}')
    except (json.JSONDecodeError, ValueError):
        data = {}

    nombre = (data.get('nombre') or '').strip()
    if not nombre:
        # Autonombrar secuencialmente si no viene nombre
        existentes = proyecto.levantamientos.count()
        nombre = f'Levantamiento {existentes + 1}'

    # Pre-poblar Fase 1 AGRESIVAMENTE desde proyecto + oportunidad +
    # cliente ligado. El ingeniero abre el wizard y encuentra hasta
    # No prellenamos ningun dato: el ingeniero debe capturar cada
    # campo manualmente para asegurar que los datos son reales y no
    # asumidos. El unico valor que respetamos es el override explicito
    # que venga en la peticion.
    fase1_default = {
        'cliente':     '',
        'cliente_id':  None,
        'contacto':    '',
        'area':        '',
        'fecha':       '',
        'email':       '',
        'telefono':    '',
        'descripcion': '',
        'servicios':   [],
        'componentes': [],
        'productos':   [],
    }
    fase1_override = data.get('fase1_data') or {}
    fase1 = {**fase1_default, **fase1_override}

    lev = ProyectoLevantamiento.objects.create(
        proyecto=proyecto,
        nombre=nombre,
        status=data.get('status', 'borrador'),
        fase_actual=int(data.get('fase_actual') or 1),
        fase1_data=fase1,
        fase2_data=data.get('fase2_data') or {},
        fase3_data=data.get('fase3_data') or {},
        fase4_data=data.get('fase4_data') or {},
        fase5_data=data.get('fase5_data') or {},
        creado_por=request.user,
    )
    return JsonResponse({'success': True, 'data': _lev_to_dict(lev, include_evidencias=True)})


@login_required
@require_http_methods(["GET"])
def api_levantamiento_detalle(request, levantamiento_id):
    try:
        lev = ProyectoLevantamiento.objects.select_related(
            'creado_por', 'actualizado_por', 'proyecto'
        ).prefetch_related('evidencias').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    return JsonResponse({
        'ok': True,
        'data': _lev_to_dict(lev, include_evidencias=True),
        'puede_editar': not _user_es_solo_lectura_levantamiento(request.user),
    })


@login_required
@require_http_methods(["POST"])
def api_levantamiento_actualizar(request, levantamiento_id):
    """Actualiza metadata (nombre, status, fase_actual) y/o data completa de fases."""
    try:
        lev = ProyectoLevantamiento.objects.select_related('proyecto').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    if _user_es_solo_lectura_levantamiento(request.user):
        return JsonResponse({'success': False, 'error': 'Los vendedores no pueden editar levantamientos'}, status=403)
    try:
        data = json.loads(request.body or '{}')
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON inválido'}, status=400)

    if 'nombre' in data:
        n = (data['nombre'] or '').strip()
        if n:
            lev.nombre = n
    if 'status' in data and data['status'] in dict(ProyectoLevantamiento.STATUS_CHOICES):
        lev.status = data['status']
    if 'fase_actual' in data:
        try:
            f = int(data['fase_actual'])
            if 1 <= f <= 5:
                lev.fase_actual = f
        except (ValueError, TypeError):
            pass
    for k in ('fase1_data', 'fase2_data', 'fase3_data', 'fase4_data', 'fase5_data'):
        if k in data and isinstance(data[k], dict):
            setattr(lev, k, data[k])

    lev.actualizado_por = request.user
    lev.save()
    return JsonResponse({'success': True, 'data': _lev_to_dict(lev, include_evidencias=True)})


@login_required
@require_http_methods(["POST"])
def api_levantamiento_fase(request, levantamiento_id):
    """Parche sólo de una fase. Body: {fase: 1..5, data: {...}, fase_actual?: int}"""
    try:
        lev = ProyectoLevantamiento.objects.select_related('proyecto').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    if _user_es_solo_lectura_levantamiento(request.user):
        return JsonResponse({'success': False, 'error': 'Los vendedores no pueden editar levantamientos'}, status=403)
    try:
        data = json.loads(request.body or '{}')
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON inválido'}, status=400)

    try:
        fase = int(data.get('fase') or 0)
    except (ValueError, TypeError):
        fase = 0
    if fase < 1 or fase > 5:
        return JsonResponse({'success': False, 'error': 'Fase inválida (1-5)'}, status=400)
    fase_data = data.get('data')
    if not isinstance(fase_data, dict):
        return JsonResponse({'success': False, 'error': 'data debe ser objeto'}, status=400)

    setattr(lev, f'fase{fase}_data', fase_data)
    # Avanzar fase_actual si el cliente envía una señal
    if 'fase_actual' in data:
        try:
            f = int(data['fase_actual'])
            if 1 <= f <= 5:
                lev.fase_actual = f
        except (ValueError, TypeError):
            pass
    lev.actualizado_por = request.user
    lev.save(update_fields=[f'fase{fase}_data', 'fase_actual', 'fecha_actualizacion', 'actualizado_por'])

    # Auto-sync de fechas Fase 2 → ProyectoIAMET.
    # El Programa de Implementación de la propuesta técnica define
    # la duración del proyecto. Reglas:
    #   - Si el proyecto NO tiene fecha → se asigna automáticamente.
    #   - Si el proyecto YA tiene fecha y coincide → no se toca.
    #   - Si difiere → solo se sobrescribe cuando el cliente envía
    #     force_override_proyecto_fechas=True (botón explícito en UI).
    proyecto_fechas_aplicadas = False
    proyecto_fechas_difieren = False
    if fase == 2 and lev.proyecto:
        prog = (fase_data or {}).get('programa') or {}
        force = bool(data.get('force_override_proyecto_fechas'))
        proy = lev.proyecto
        cambios = []
        for key in ('fecha_inicio', 'fecha_fin'):
            nueva = _parse_date(prog.get(key))
            if not nueva:
                continue
            actual = getattr(proy, key, None)
            if actual is None:
                setattr(proy, key, nueva)
                cambios.append(key)
            elif actual != nueva:
                if force:
                    setattr(proy, key, nueva)
                    cambios.append(key)
                else:
                    proyecto_fechas_difieren = True
        if cambios:
            proy.save(update_fields=cambios + ['updated_at'])
            proyecto_fechas_aplicadas = True

    payload = _lev_to_dict(lev)
    payload['proyecto_fechas_aplicadas'] = proyecto_fechas_aplicadas
    payload['proyecto_fechas_difieren'] = proyecto_fechas_difieren
    return JsonResponse({'success': True, 'data': payload})


@login_required
@require_http_methods(["POST"])
def api_levantamiento_eliminar(request, levantamiento_id):
    try:
        lev = ProyectoLevantamiento.objects.select_related('proyecto').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    if _user_es_solo_lectura_levantamiento(request.user):
        return JsonResponse({'success': False, 'error': 'Los vendedores no pueden eliminar levantamientos'}, status=403)
    lev.delete()
    return JsonResponse({'success': True})


# ══════════════════════════════════════════════════════════════
#  VOLUMETRÍAS — Borradores múltiples por levantamiento
#  ─────────────────────────────────────────────────────────────
#  Cada volumetría es un escenario/diseño editable con su propio
#  status (borrador/completada). Vendedores sólo ven completadas;
#  ingenieros y admins ven todas.
# ══════════════════════════════════════════════════════════════

def _vol_to_dict(vol):
    return {
        'id': vol.id,
        'levantamiento_id': vol.levantamiento_id,
        'nombre': vol.nombre,
        'status': vol.status,
        'status_label': vol.get_status_display(),
        'creado_por_id': vol.creado_por_id,
        'creado_por_nombre': (vol.creado_por.get_full_name() or vol.creado_por.username) if vol.creado_por else '',
        'actualizado_por_id': vol.actualizado_por_id,
        'actualizado_por_nombre': (vol.actualizado_por.get_full_name() or vol.actualizado_por.username) if vol.actualizado_por else '',
        'fecha_creacion': _fmt(vol.fecha_creacion),
        'fecha_actualizacion': _fmt(vol.fecha_actualizacion),
    }


@login_required
@require_http_methods(["GET"])
def api_volumetrias_lista(request, levantamiento_id):
    """Lista volumetrías de un levantamiento. Vendedores reciben sólo
    las marcadas como `completada`; ingenieros ven todas."""
    try:
        lev = ProyectoLevantamiento.objects.select_related('proyecto').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    qs = lev.volumetrias.select_related('creado_por', 'actualizado_por').all()
    if _user_es_solo_lectura_levantamiento(request.user):
        qs = qs.filter(status='completada')
    return JsonResponse({
        'ok': True,
        'data': [_vol_to_dict(v) for v in qs],
        'puede_editar': not _user_es_solo_lectura_levantamiento(request.user),
    })


@login_required
@require_http_methods(["POST"])
def api_volumetria_crear(request, levantamiento_id):
    """Crea una volumetría borrador nueva en el levantamiento.
    Auto-numera el nombre si no se provee uno."""
    try:
        lev = ProyectoLevantamiento.objects.select_related('proyecto').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    if _user_es_solo_lectura_levantamiento(request.user):
        return JsonResponse({'success': False, 'error': 'Los vendedores no pueden crear volumetrías'}, status=403)
    try:
        data = json.loads(request.body or '{}')
    except (json.JSONDecodeError, ValueError):
        data = {}
    nombre = (data.get('nombre') or '').strip()
    if not nombre:
        existentes = lev.volumetrias.count()
        nombre = f'Volumetría {existentes + 1}'
    vol = ProyectoVolumetria.objects.create(
        levantamiento=lev,
        nombre=nombre,
        status='borrador',
        data=data.get('data') or {},
        creado_por=request.user,
    )
    return JsonResponse({'success': True, 'data': _vol_to_dict(vol)})


@login_required
@require_http_methods(["GET"])
def api_volumetria_detalle(request, volumetria_id):
    """Devuelve la volumetría completa, incluyendo `data` (JSON con las
    partidas/materiales/etc. — puede ser pesado)."""
    try:
        vol = ProyectoVolumetria.objects.select_related(
            'levantamiento__proyecto', 'creado_por', 'actualizado_por',
        ).get(id=volumetria_id)
    except ProyectoVolumetria.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Volumetría no encontrada'}, status=404)
    if not _check_access(request.user, vol.levantamiento.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    # Vendedor sólo puede ver completadas.
    if _user_es_solo_lectura_levantamiento(request.user) and vol.status != 'completada':
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    payload = _vol_to_dict(vol)
    payload['data'] = vol.data or {}
    return JsonResponse({'ok': True, 'data': payload})


@login_required
@require_http_methods(["POST"])
def api_volumetria_actualizar(request, volumetria_id):
    """Actualiza meta de la volumetría: nombre y/o status. La data se
    guarda en otro endpoint (`api_volumetria_data`) por separado para
    no mezclar autosave pesado con cambios de meta livianos."""
    try:
        vol = ProyectoVolumetria.objects.select_related('levantamiento__proyecto').get(id=volumetria_id)
    except ProyectoVolumetria.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Volumetría no encontrada'}, status=404)
    if not _check_access(request.user, vol.levantamiento.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    if _user_es_solo_lectura_levantamiento(request.user):
        return JsonResponse({'success': False, 'error': 'Los vendedores no pueden editar volumetrías'}, status=403)
    try:
        body = json.loads(request.body or '{}')
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON inválido'}, status=400)

    update_fields = ['fecha_actualizacion', 'actualizado_por']
    if 'nombre' in body:
        nuevo = (body.get('nombre') or '').strip()
        if not nuevo:
            return JsonResponse({'success': False, 'error': 'Nombre vacío'}, status=400)
        vol.nombre = nuevo[:200]
        update_fields.append('nombre')
    if 'status' in body:
        nuevo_status = (body.get('status') or '').strip()
        if nuevo_status not in dict(ProyectoVolumetria.STATUS_CHOICES):
            return JsonResponse({'success': False, 'error': 'Status inválido'}, status=400)
        vol.status = nuevo_status
        update_fields.append('status')
    vol.actualizado_por = request.user
    vol.save(update_fields=update_fields)
    return JsonResponse({'success': True, 'data': _vol_to_dict(vol)})


@login_required
@require_http_methods(["POST"])
def api_volumetria_data(request, volumetria_id):
    """Autosave de la data (partidas, materiales, mano de obra, gastos).
    Reemplaza completa la `data` con lo que venga en el body."""
    try:
        vol = ProyectoVolumetria.objects.select_related('levantamiento__proyecto').get(id=volumetria_id)
    except ProyectoVolumetria.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Volumetría no encontrada'}, status=404)
    if not _check_access(request.user, vol.levantamiento.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    if _user_es_solo_lectura_levantamiento(request.user):
        return JsonResponse({'success': False, 'error': 'Los vendedores no pueden editar volumetrías'}, status=403)
    try:
        body = json.loads(request.body or '{}')
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'JSON inválido'}, status=400)
    nueva_data = body.get('data')
    if not isinstance(nueva_data, dict):
        return JsonResponse({'success': False, 'error': 'data debe ser objeto'}, status=400)
    vol.data = nueva_data
    vol.actualizado_por = request.user
    vol.save(update_fields=['data', 'fecha_actualizacion', 'actualizado_por'])
    payload = _vol_to_dict(vol)
    return JsonResponse({'success': True, 'data': payload})


@login_required
@require_http_methods(["POST"])
def api_volumetria_eliminar(request, volumetria_id):
    """Elimina una volumetría borrador. Las completadas no se pueden
    eliminar — primero hay que bajarlas a borrador (regla anti-pérdida
    de trabajo aprobado)."""
    try:
        vol = ProyectoVolumetria.objects.select_related('levantamiento__proyecto').get(id=volumetria_id)
    except ProyectoVolumetria.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Volumetría no encontrada'}, status=404)
    if not _check_access(request.user, vol.levantamiento.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    if _user_es_solo_lectura_levantamiento(request.user):
        return JsonResponse({'success': False, 'error': 'Los vendedores no pueden eliminar volumetrías'}, status=403)
    if vol.status == 'completada':
        return JsonResponse({
            'success': False,
            'error': 'No se puede eliminar una volumetría completada. Bájala a borrador primero.',
        }, status=400)
    vol.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["POST"])
def api_levantamiento_evidencia_subir(request, levantamiento_id):
    """Sube una foto de evidencia. Multipart: archivo + opcionales producto_idx, comentario."""
    try:
        lev = ProyectoLevantamiento.objects.select_related('proyecto').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'success': False, 'error': 'Archivo requerido (campo "archivo")'}, status=400)

    producto_idx = request.POST.get('producto_idx')
    try:
        producto_idx = int(producto_idx) if producto_idx not in (None, '', 'null') else None
    except (ValueError, TypeError):
        producto_idx = None

    ev = LevantamientoEvidencia.objects.create(
        levantamiento=lev,
        archivo=archivo,
        nombre_original=archivo.name[:255],
        comentario=(request.POST.get('comentario') or '')[:255],
        producto_idx=producto_idx,
        subido_por=request.user,
    )
    return JsonResponse({'success': True, 'data': _evidencia_to_dict(ev)})


@login_required
@require_http_methods(["POST"])
def api_levantamiento_evidencia_eliminar(request, evidencia_id):
    try:
        ev = LevantamientoEvidencia.objects.select_related('levantamiento__proyecto').get(id=evidencia_id)
    except LevantamientoEvidencia.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Evidencia no encontrada'}, status=404)
    if not _check_access(request.user, ev.levantamiento.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)
    ev.archivo.delete(save=False)
    ev.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["GET"])
def api_levantamiento_sitio_pdf(request, levantamiento_id):
    """Genera el PDF del Levantamiento en Sitio (Fase 1).

    Replica el formato del docx "FORMATO DE LEVANTAMIENTO EN SITIO
    INSTALACIONES" pero con styling moderno (IAMET branding).
    """
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.conf import settings
    import os
    try:
        lev = ProyectoLevantamiento.objects.select_related(
            'proyecto', 'creado_por'
        ).prefetch_related('evidencias').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    f1 = lev.fase1_data or {}

    def _static_file_url(rel_path):
        candidates = []
        if getattr(settings, 'STATIC_ROOT', None):
            candidates.append(os.path.join(settings.STATIC_ROOT, rel_path))
        candidates.append(os.path.join(settings.BASE_DIR, 'app', 'static', rel_path))
        for p in candidates:
            if os.path.exists(p):
                return 'file://' + p
        return ''

    # Formato de fecha: "04 / Nov / 2025"
    import datetime as _dt
    def _fmt_fecha(iso):
        if not iso:
            return ''
        try:
            dt = _dt.datetime.strptime(str(iso)[:10], '%Y-%m-%d')
        except Exception:
            return iso
        meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        return f'{dt.day:02d} / {meses[dt.month - 1]} / {dt.year}'

    # Listas canónicas de Tipo de Servicio y Componentes (igual al docx)
    TIPOS_SERVICIO = ['CCTV', 'Alarma Intrusión', 'Control de Acceso', 'Cableado Estructurado', 'Voceo', 'Telefonía', 'Otros']
    COMPONENTES = [
        'Cámara IP', 'Cámara Analógica', 'NVR', 'DVR', 'Video Balum',
        'Tubería', 'Canalización', 'Escalerilla',
        'Cable Cat5e', 'Cable Cat6', 'Cable Cat6A', 'Fibra Óptica',
        'Gabinete/Rack', 'Serpentines', 'Controladora',
        'PBX', 'Teléfonos', 'Work Station', 'Servidor', 'Switches',
        'Antenas', 'Enlaces', 'Bocinas', 'Cableado',
        'Fuentes de Poder', 'Amplificadores', 'Access Point', 'UPS',
    ]
    selected_svc = set(f1.get('servicios') or [])
    selected_comp = set(f1.get('componentes') or [])

    productos = [{
        'qty':       p.get('qty') or 1,
        'unidad':    p.get('unidad') or 'PZA',
        'desc':      p.get('desc') or '',
        'marca':     p.get('marca') or '',
        'modelo':    p.get('modelo') or '',
        'ubicacion': p.get('ubicacion') or '',
    } for p in (f1.get('productos') or [])]

    # Programa de implementación — ahora se captura en Fase 2
    # (fase2_data.programa). Para datos viejos, caemos a fase1_data.
    f2 = lev.fase2_data or {}
    prog = f2.get('programa') or {}
    def _prog(key):
        v = prog.get(key)
        if v is not None and v != '':
            return v
        return f1.get(key) or ''

    # Notas libres de campo (nueva Isla 4)
    notas_materiales = f1.get('notas_materiales') or ''

    # Evidencias fotográficas — convertir a rutas file:// para WeasyPrint
    def _file_url_for_field(field):
        if not field:
            return ''
        try:
            return 'file://' + field.path
        except Exception:
            try:
                return request.build_absolute_uri(field.url)
            except Exception:
                return ''
    evidencias = [{
        'abs_url': _file_url_for_field(e.archivo),
        'comentario': e.comentario or '',
    } for e in lev.evidencias.all()]

    ctx = {
        'lev':          lev,
        'cliente':      f1.get('cliente')  or (lev.proyecto.cliente_nombre if lev.proyecto else ''),
        'contacto':     f1.get('contacto') or '',
        'area':         f1.get('area')     or '',
        'email':        f1.get('email')    or '',
        'telefono':     f1.get('telefono') or '',
        'fecha_fmt':    _fmt_fecha(f1.get('fecha') or timezone.localdate().isoformat()),
        'descripcion':  f1.get('descripcion') or '',

        'tipos_servicio':   [{'name': s, 'on': s in selected_svc} for s in TIPOS_SERVICIO],
        'componentes':      [{'name': c, 'on': c in selected_comp} for c in COMPONENTES],

        # Programa (leer de fase2.programa, fallback a fase1)
        'fecha_inicio_fmt': _fmt_fecha(_prog('fecha_inicio')),
        'fecha_fin_fmt':    _fmt_fecha(_prog('fecha_fin')),
        'duracion':         _prog('duracion'),
        'turno':            _prog('turno'),
        'apoyo_cliente':    _prog('apoyo_cliente'),
        'personal_req':     _prog('personal_req'),
        'realizo':          lev.creado_por.get_full_name() if lev.creado_por else '',

        # Equipo de elevación
        'elev_alto':   _prog('elev_alto'),
        'elev_ancho':  _prog('elev_ancho'),
        'elev_modelo': _prog('elev_modelo'),

        'productos':        productos,
        'notas_materiales': notas_materiales,
        'evidencias':       evidencias,

        # Flag: ¿hay datos de programa para renderizar esa sección?
        'has_programa':     any([_prog(k) for k in ('fecha_inicio', 'fecha_fin', 'duracion', 'turno', 'apoyo_cliente', 'personal_req', 'elev_alto', 'elev_ancho', 'elev_modelo')]),

        # Assets
        'logo_url':         _static_file_url('images/iamet-logo.png'),
        'bajanet_hero_url': _static_file_url('images/propuesta/bajanet_hero.jpeg'),
        'footer_logos_url': _static_file_url('images/propuesta/footer_logos.png'),

        # Empresa (hardcoded por ahora)
        'empresa_tel':       '664 000 0000',
        'empresa_web':       'www.iamet.mx',
        'empresa_direccion': 'Tijuana, B.C.',
    }

    html = render_to_string('crm/levantamiento_sitio_pdf.html', ctx, request=request)
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Error generando PDF: {e}'}, status=500)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '_' for c in (lev.nombre or 'levantamiento')).strip()[:80] or 'levantamiento'
    filename = f'Levantamiento_{safe_name}.pdf'
    disp = 'attachment' if request.GET.get('download') else 'inline'
    response['Content-Disposition'] = f'{disp}; filename="{filename}"'
    return response


@login_required
@require_http_methods(["GET"])
def api_levantamiento_propuesta_pdf(request, levantamiento_id):
    """Genera el PDF de la Propuesta Técnica de un levantamiento.

    Respeta el formato del docx original (tablas azules con secciones).
    Params:
      ?download=1  → fuerza Content-Disposition: attachment
      por defecto  → inline (se abre en la pestaña para preview)
    """
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.conf import settings
    import os
    try:
        lev = ProyectoLevantamiento.objects.select_related(
            'proyecto', 'creado_por'
        ).prefetch_related('evidencias').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    f1 = lev.fase1_data or {}
    f2 = lev.fase2_data or {}

    # Para WeasyPrint: usar rutas filesystem (file://) en lugar de HTTP URLs.
    # Así evitamos que WeasyPrint haga peticiones HTTP (que pueden fallar por
    # SSL/auth/red interna del contenedor) y las imágenes cargan directo del disco.
    def _file_url_for_field(field):
        if not field:
            return ''
        try:
            return 'file://' + field.path
        except Exception:
            try:
                return request.build_absolute_uri(field.url)
            except Exception:
                return ''

    def _static_file_url(rel_path):
        """Busca un archivo estático en STATIC_ROOT (prod tras collectstatic)
        o en los directorios STATICFILES_DIRS/BASE_DIR (dev)."""
        candidates = []
        if getattr(settings, 'STATIC_ROOT', None):
            candidates.append(os.path.join(settings.STATIC_ROOT, rel_path))
        # Dev: el directorio app/static/
        candidates.append(os.path.join(settings.BASE_DIR, 'app', 'static', rel_path))
        for p in candidates:
            if os.path.exists(p):
                return 'file://' + p
        return ''

    evidencias = [{
        'abs_url': _file_url_for_field(e.archivo),
        'comentario': e.comentario or '',
    } for e in lev.evidencias.all()]

    # Formato de fecha: "04 / Nov / 2025"
    import datetime as _dt
    def _fmt_fecha(iso):
        if not iso:
            return ''
        try:
            dt = _dt.datetime.strptime(str(iso)[:10], '%Y-%m-%d')
        except Exception:
            return iso
        meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        return f'{dt.day:02d} / {meses[dt.month - 1]} / {dt.year}'

    ctx = {
        'lev': lev,
        'cliente_nombre': f1.get('cliente') or (lev.proyecto.cliente_nombre if lev.proyecto else ''),
        'elaboro':       f2.get('elaboro')       or (lev.creado_por.get_full_name() if lev.creado_por else ''),
        'doc_fecha_fmt': _fmt_fecha(f2.get('doc_fecha') or timezone.localdate().isoformat()),
        'solicitante':   f2.get('solicitante')   or f1.get('contacto') or '',
        'departamento':  f2.get('departamento')  or '',
        'planta':        f2.get('planta')        or f1.get('cliente') or '',
        'areas':         f2.get('areas')         or f1.get('area')    or '',
        'tipo_solucion': f2.get('tipo_solucion') or ', '.join(f1.get('servicios') or []),
        'fecha_inst':    f2.get('fecha_inst')    or 'Acordar con usuario',
        'desc_proyecto': f2.get('desc_proyecto') or f1.get('descripcion') or '',
        'especificaciones': [x for x in (f2.get('especificaciones') or []) if x and x.strip()],
        'comentarios_spec': [x for x in (f2.get('comentarios_spec') or []) if x and x.strip()],
        'notas_evidencia':  f2.get('notas_evidencia') or '',
        # Productos/Materiales — ahora se capturan en Fase 2 (con catálogo)
        # con fallback a fase1.productos para compat con levantamientos viejos.
        # Comentarios viven en fase2.comentarios como dict {idx: texto}; las
        # llaves pueden venir como string (JSON) o int — probamos ambas.
        'productos': [
            {
                'qty':    p.get('qty') or 1,
                'unidad': p.get('unidad') or 'PZA',
                'desc':   p.get('desc') or '',
                'marca':  p.get('marca') or '',
                'modelo': p.get('modelo') or '',
                'comentario': (
                    ((f2.get('comentarios') or {}).get(str(i))
                     or (f2.get('comentarios') or {}).get(i)
                     or '')
                ),
            }
            for i, p in enumerate(f2.get('productos') or f1.get('productos') or [])
        ],
        'evidencias': evidencias,
        # Recursos visuales — rutas filesystem para WeasyPrint
        'logo_url':          _static_file_url('images/iamet-logo.png'),
        'bajanet_hero_url':  _static_file_url('images/propuesta/bajanet_hero.jpeg'),
        'footer_logos_url':  _static_file_url('images/propuesta/footer_logos.png'),
        'watermark_url':     _static_file_url('images/propuesta/watermark.jpg'),
        # Datos de empresa (hardcoded por ahora; en el futuro pueden venir de
        # una configuración del sistema)
        'empresa_tel':       '664 000 0000',
        'empresa_web':       'www.iamet.mx',
        'empresa_direccion': 'Tijuana, B.C.',
    }

    html = render_to_string('crm/levantamiento_propuesta_pdf.html', ctx, request=request)

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Error generando PDF: {e}'}, status=500)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '_' for c in (lev.nombre or 'propuesta')).strip()[:80] or 'propuesta'
    filename = f'PropuestaTecnica_{safe_name}.pdf'
    disp = 'attachment' if request.GET.get('download') else 'inline'
    response['Content-Disposition'] = f'{disp}; filename="{filename}"'
    return response


# ─── VOLUMETRÍA: helpers compartidos (PDF + XLSX) ─────────────────
def _fmt_money(val, symbol=True):
    """Formato de moneda: $1,234.56 o 1,234.56."""
    try:
        n = float(val or 0)
    except (TypeError, ValueError):
        n = 0.0
    if abs(n) < 0.005 and n != 0:
        n = 0.0
    s = '{:,.2f}'.format(n)
    return ('$' + s) if symbol else s


def _fmt_qty(val):
    try:
        n = float(val or 0)
    except (TypeError, ValueError):
        return '0'
    if n == int(n):
        return str(int(n))
    return '{:,.2f}'.format(n)


def _build_volumetria_ctx(lev, sin_costos=False):
    """Construye el contexto de volumetría (usado por PDF y XLSX).

    Retorna un dict con materiales/mano_obra/gastos (filas formateadas) y totales.
    """
    f1 = lev.fase1_data or {}
    f2 = lev.fase2_data or {}
    f3 = lev.fase3_data or {}

    def _num(x):
        try:
            return float(x or 0)
        except (TypeError, ValueError):
            return 0.0

    # ── Materiales ──
    materiales = []
    tot_mat_venta = 0.0
    tot_mat_costo = 0.0
    for r in (f3.get('materiales') or []):
        qty = _num(r.get('qty'))
        costo_u = _num(r.get('costoUnit'))
        precio_l = _num(r.get('precioLista'))
        desc_c = _num(r.get('descCompra'))
        desc_v = _num(r.get('descVenta'))
        costo_real = costo_u * (1 - desc_c / 100)
        precio_real = precio_l * (1 - desc_v / 100)
        costo_total = costo_real * qty
        precio_venta = precio_real * qty
        tot_mat_venta += precio_venta
        tot_mat_costo += costo_total
        materiales.append({
            'qty': qty, 'qty_fmt': _fmt_qty(qty),
            'unid': r.get('unid') or 'PZA',
            'desc': r.get('desc') or '',
            'marca': r.get('marca') or '',
            'modelo': r.get('modelo') or '',
            'proveedor': r.get('proveedor') or '',
            'entrega': r.get('entrega') or '',
            'precio_lista': precio_l, 'precio_lista_fmt': _fmt_money(precio_l),
            'desc_venta': desc_v, 'desc_venta_fmt': ('{:.1f}%'.format(desc_v) if desc_v else '—'),
            'precio_unit': precio_real, 'precio_unit_fmt': _fmt_money(precio_real),
            'precio_venta': precio_venta, 'precio_venta_fmt': _fmt_money(precio_venta),
            'costo_unit': costo_real, 'costo_unit_fmt': _fmt_money(costo_real),
            'costo_total': costo_total, 'costo_total_fmt': _fmt_money(costo_total),
        })

    # ── Mano de obra ──
    mano_obra = []
    tot_mo = 0.0
    for r in (f3.get('manoObra') or []):
        qty = _num(r.get('qty'))
        p_unit = _num(r.get('precioUnit'))
        total = qty * p_unit
        tot_mo += total
        mano_obra.append({
            'qty': qty, 'qty_fmt': _fmt_qty(qty),
            'unid': r.get('unid') or 'SERV',
            'desc': r.get('desc') or '',
            'precio_unit': p_unit, 'precio_unit_fmt': _fmt_money(p_unit),
            'total': total, 'total_fmt': _fmt_money(total),
        })

    # ── Gastos ──
    gastos = []
    tot_gas = 0.0
    for r in (f3.get('gastos') or []):
        qty = _num(r.get('qty'))
        c_unit = _num(r.get('costoUnit'))
        total = qty * c_unit
        tot_gas += total
        gastos.append({
            'qty': qty, 'qty_fmt': _fmt_qty(qty),
            'unid': r.get('unid') or 'GLOB',
            'desc': r.get('desc') or '',
            'costo_unit': c_unit, 'costo_unit_fmt': _fmt_money(c_unit),
            'total': total, 'total_fmt': _fmt_money(total),
        })

    total_venta = tot_mat_venta + tot_mo + tot_gas
    total_costo = tot_mat_costo + tot_gas
    utilidad = total_venta - total_costo
    margen = (utilidad / total_venta * 100) if total_venta > 0 else 0.0

    import datetime as _dt
    def _fmt_fecha(iso):
        if not iso:
            return ''
        try:
            dt = _dt.datetime.strptime(str(iso)[:10], '%Y-%m-%d')
        except Exception:
            return iso
        meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        return f'{dt.day:02d} / {meses[dt.month - 1]} / {dt.year}'

    # Tipo de cambio — USD → MXN (se captura en Fase 3)
    tc_raw = f3.get('tipo_cambio')
    try:
        tc = float(tc_raw) if tc_raw else 0.0
    except (TypeError, ValueError):
        tc = 0.0
    if not tc or tc <= 0:
        tc = 19.5  # fallback razonable
    tc_fecha = f3.get('tipo_cambio_fecha') or ''

    return {
        'lev': lev,
        'sin_costos': sin_costos,
        'cliente_nombre': f1.get('cliente') or (lev.proyecto.cliente_nombre if lev.proyecto else ''),
        'elaboro': f2.get('elaboro') or (lev.creado_por.get_full_name() if lev.creado_por else ''),
        'doc_fecha_fmt': _fmt_fecha(f2.get('doc_fecha') or timezone.localdate().isoformat()),
        'solicitante': f2.get('solicitante') or f1.get('contacto') or '',
        'planta': f2.get('planta') or f1.get('cliente') or '',
        'areas': f2.get('areas') or f1.get('area') or '',
        'materiales': materiales,
        'mano_obra': mano_obra,
        'gastos': gastos,
        'tot_mat_venta': tot_mat_venta, 'tot_mat_venta_fmt': _fmt_money(tot_mat_venta),
        'tot_mat_costo': tot_mat_costo, 'tot_mat_costo_fmt': _fmt_money(tot_mat_costo),
        'tot_mo': tot_mo, 'tot_mo_fmt': _fmt_money(tot_mo),
        'tot_gas': tot_gas, 'tot_gas_fmt': _fmt_money(tot_gas),
        'total_venta': total_venta, 'total_venta_fmt': _fmt_money(total_venta),
        'total_costo': total_costo, 'total_costo_fmt': _fmt_money(total_costo),
        'utilidad': utilidad, 'utilidad_fmt': _fmt_money(utilidad),
        'margen_pct': '{:.1f}'.format(margen),
        # Conversión MXN con TC
        'tipo_cambio': tc,
        'tipo_cambio_fmt': '{:,.2f}'.format(tc),
        'tipo_cambio_fecha': tc_fecha,
        'tot_mat_venta_mxn_fmt': _fmt_money(tot_mat_venta * tc),
        'tot_mo_mxn_fmt': _fmt_money(tot_mo * tc),
        'tot_gas_mxn_fmt': _fmt_money(tot_gas * tc),
        'total_venta_mxn_fmt': _fmt_money(total_venta * tc),
        'total_costo_mxn_fmt': _fmt_money(total_costo * tc),
        'utilidad_mxn_fmt': _fmt_money(utilidad * tc),
    }


@login_required
@require_http_methods(["GET"])
def api_levantamiento_volumetria_pdf(request, levantamiento_id):
    """PDF de volumetría. Params:
      ?download=1      → fuerza descarga (attachment)
      ?sin_costos=1    → genera sin columnas de costo ni resumen financiero
    """
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.conf import settings
    import os
    try:
        lev = ProyectoLevantamiento.objects.select_related('proyecto', 'creado_por').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    def _static_file_url(rel_path):
        candidates = []
        if getattr(settings, 'STATIC_ROOT', None):
            candidates.append(os.path.join(settings.STATIC_ROOT, rel_path))
        candidates.append(os.path.join(settings.BASE_DIR, 'app', 'static', rel_path))
        for p in candidates:
            if os.path.exists(p):
                return 'file://' + p
        return ''

    sin_costos = bool(request.GET.get('sin_costos'))
    ctx = _build_volumetria_ctx(lev, sin_costos=sin_costos)
    ctx.update({
        'bajanet_hero_url': _static_file_url('images/propuesta/bajanet_hero.jpeg'),
        'footer_logos_url': _static_file_url('images/propuesta/footer_logos.png'),
        'empresa_tel': '664 000 0000',
        'empresa_web': 'www.iamet.mx',
        'empresa_direccion': 'Tijuana, B.C.',
    })

    html = render_to_string('crm/levantamiento_volumetria_pdf.html', ctx, request=request)
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Error generando PDF: {e}'}, status=500)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '_' for c in (lev.nombre or 'volumetria')).strip()[:80] or 'volumetria'
    suffix = '_SinCostos' if sin_costos else ''
    filename = f'Volumetria_{safe_name}{suffix}.pdf'
    disp = 'attachment' if request.GET.get('download') else 'inline'
    response['Content-Disposition'] = f'{disp}; filename="{filename}"'
    return response


@login_required
@require_http_methods(["GET"])
def api_levantamiento_volumetria_xlsx(request, levantamiento_id):
    """Exporta la volumetría en formato Excel replicando el layout tradicional
    (header con cliente/proyecto/contacto/elaboró, tabla de equipamiento agrupada,
    mano de obra y resumen)."""
    from django.http import HttpResponse
    try:
        lev = ProyectoLevantamiento.objects.select_related('proyecto', 'creado_por').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JsonResponse({'success': False, 'error': 'openpyxl no está instalado'}, status=500)

    ctx = _build_volumetria_ctx(lev, sin_costos=False)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Volumetria'

    # Estilos
    blue_fill = PatternFill('solid', fgColor='2563EB')
    band_fill = PatternFill('solid', fgColor='E8F1FB')
    section_fill = PatternFill('solid', fgColor='FFF7ED')
    header_fill = PatternFill('solid', fgColor='F1F5F9')
    total_fill = PatternFill('solid', fgColor='EFF6FF')
    white = Font(color='FFFFFF', bold=True, name='Calibri', size=11)
    band_font = Font(color='1D4ED8', bold=True, name='Calibri', size=10)
    header_font = Font(color='475569', bold=True, name='Calibri', size=10)
    bold = Font(bold=True)
    thin = Side(style='thin', color='DBE6F3')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    money_fmt = '"$"#,##0.00'

    # ── Header: Cliente / Proyecto / Fecha / Contacto / Elaboró ──
    ws['A1'] = 'Cliente'; ws['A1'].font = band_font; ws['A1'].fill = band_fill
    ws['B1'] = ctx['cliente_nombre']; ws['B1'].font = bold
    ws['D1'] = 'Contacto:'; ws['D1'].font = band_font; ws['D1'].fill = band_fill
    ws['E1'] = ctx['solicitante']
    ws['K1'] = 'Fecha:'; ws['K1'].font = band_font; ws['K1'].fill = band_fill
    ws['L1'] = ctx['doc_fecha_fmt']

    ws['A2'] = 'Proyecto'; ws['A2'].font = band_font; ws['A2'].fill = band_fill
    ws['B2'] = lev.nombre or ''; ws['B2'].font = bold
    ws['D2'] = 'Elaboró:'; ws['D2'].font = band_font; ws['D2'].fill = band_fill
    ws['E2'] = ctx['elaboro']
    ws['K2'] = 'Planta:'; ws['K2'].font = band_font; ws['K2'].fill = band_fill
    ws['L2'] = ctx['planta']

    # ── Equipamiento / Materiales ──
    # Fila 4: grupo título
    ws.cell(row=4, column=1, value='EQUIPAMIENTO / MATERIALES').font = Font(bold=True, color='9A3412', size=11)
    ws.cell(row=4, column=1).fill = section_fill
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=12)

    # Fila 5: encabezados
    headers = ['Marca', 'No. Parte', 'Cantidad', 'Unid', 'Descripción',
               'Precio Lista', 'Desc V%', 'Precio Venta', 'Costo Unit', 'Costo Total',
               'Proveedor', 'Entrega']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

    r = 6
    for m in ctx['materiales']:
        ws.cell(row=r, column=1, value=m['marca'])
        ws.cell(row=r, column=2, value=m['modelo'])
        ws.cell(row=r, column=3, value=m['qty'])
        ws.cell(row=r, column=4, value=m['unid'])
        ws.cell(row=r, column=5, value=m['desc'])
        c_pl = ws.cell(row=r, column=6, value=m['precio_lista']); c_pl.number_format = money_fmt
        ws.cell(row=r, column=7, value=m['desc_venta']).number_format = '0.0"%"'
        c_pv = ws.cell(row=r, column=8, value=m['precio_venta']); c_pv.number_format = money_fmt; c_pv.font = bold
        c_cu = ws.cell(row=r, column=9, value=m['costo_unit']); c_cu.number_format = money_fmt
        c_ct = ws.cell(row=r, column=10, value=m['costo_total']); c_ct.number_format = money_fmt
        ws.cell(row=r, column=11, value=m['proveedor'])
        ws.cell(row=r, column=12, value=m.get('entrega') or '')
        for col in range(1, 13):
            ws.cell(row=r, column=col).border = border
        r += 1

    # Total materiales
    if ctx['materiales']:
        ws.cell(row=r, column=1, value='Subtotal Materiales').font = bold
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c_tv = ws.cell(row=r, column=8, value=ctx['tot_mat_venta'])
        c_tv.number_format = money_fmt; c_tv.font = bold; c_tv.fill = total_fill
        c_tc = ws.cell(row=r, column=10, value=ctx['tot_mat_costo'])
        c_tc.number_format = money_fmt; c_tc.font = bold; c_tc.fill = total_fill
        r += 2
    else:
        r += 1

    # ── Mano de obra ──
    ws.cell(row=r, column=1, value='MANO DE OBRA').font = Font(bold=True, color='9A3412', size=11)
    ws.cell(row=r, column=1).fill = section_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 1
    mo_headers = ['', '', 'Cantidad', 'Unid', 'Descripción', 'Precio Unit', '', 'Total']
    for i, h in enumerate(mo_headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        if h:
            c.font = header_font; c.fill = header_fill; c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
    r += 1
    for mo in ctx['mano_obra']:
        ws.cell(row=r, column=3, value=mo['qty'])
        ws.cell(row=r, column=4, value=mo['unid'])
        ws.cell(row=r, column=5, value=mo['desc'])
        ws.cell(row=r, column=6, value=mo['precio_unit']).number_format = money_fmt
        c_t = ws.cell(row=r, column=8, value=mo['total'])
        c_t.number_format = money_fmt; c_t.font = bold
        for col in range(3, 9):
            ws.cell(row=r, column=col).border = border
        r += 1
    if ctx['mano_obra']:
        ws.cell(row=r, column=1, value='Subtotal Mano de Obra').font = bold
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c_tmo = ws.cell(row=r, column=8, value=ctx['tot_mo'])
        c_tmo.number_format = money_fmt; c_tmo.font = bold; c_tmo.fill = total_fill
        r += 2
    else:
        r += 1

    # ── Gastos ──
    if ctx['gastos']:
        ws.cell(row=r, column=1, value='GASTOS / VIÁTICOS').font = Font(bold=True, color='9A3412', size=11)
        ws.cell(row=r, column=1).fill = section_fill
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        r += 1
        for i, h in enumerate(mo_headers, start=1):
            c = ws.cell(row=r, column=i, value=h)
            if h:
                c.font = header_font; c.fill = header_fill; c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
        r += 1
        for g in ctx['gastos']:
            ws.cell(row=r, column=3, value=g['qty'])
            ws.cell(row=r, column=4, value=g['unid'])
            ws.cell(row=r, column=5, value=g['desc'])
            ws.cell(row=r, column=6, value=g['costo_unit']).number_format = money_fmt
            c_t = ws.cell(row=r, column=8, value=g['total'])
            c_t.number_format = money_fmt; c_t.font = bold
            for col in range(3, 9):
                ws.cell(row=r, column=col).border = border
            r += 1
        ws.cell(row=r, column=1, value='Subtotal Gastos').font = bold
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c_tg = ws.cell(row=r, column=8, value=ctx['tot_gas'])
        c_tg.number_format = money_fmt; c_tg.font = bold; c_tg.fill = total_fill
        r += 2

    # ── Análisis de costos ──
    ws.cell(row=r, column=1, value='ANÁLISIS DE COSTOS').font = Font(bold=True, color='FFFFFF', size=11)
    ws.cell(row=r, column=1).fill = blue_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 1
    summary_rows = [
        ('Subtotal Materiales (Venta)', ctx['tot_mat_venta']),
        ('Subtotal Mano de Obra', ctx['tot_mo']),
        ('Subtotal Gastos', ctx['tot_gas']),
        ('TOTAL VENTA', ctx['total_venta']),
        ('Total Costos', ctx['total_costo']),
        ('Utilidad Bruta', ctx['utilidad']),
    ]
    for lbl, val in summary_rows:
        ws.cell(row=r, column=1, value=lbl).font = bold
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        c_v = ws.cell(row=r, column=8, value=val)
        c_v.number_format = money_fmt; c_v.font = bold
        if lbl == 'TOTAL VENTA':
            c_v.fill = blue_fill; c_v.font = white
            ws.cell(row=r, column=1).fill = blue_fill; ws.cell(row=r, column=1).font = white
        elif lbl == 'Utilidad Bruta':
            c_v.fill = PatternFill('solid', fgColor='ECFDF5')
        r += 1
    ws.cell(row=r, column=1, value=f"Margen Bruto: {ctx['margen_pct']}%").font = Font(bold=True, color='047857', size=11)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    # Anchos de columna
    widths = {1: 14, 2: 16, 3: 10, 4: 8, 5: 46, 6: 14, 7: 9, 8: 14, 9: 13, 10: 14, 11: 14, 12: 12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    # Wrap en descripciones
    for row in ws.iter_rows(min_row=6, max_col=5):
        for cell in row:
            if cell.column == 5:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Response
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '_' for c in (lev.nombre or 'volumetria')).strip()[:80] or 'volumetria'
    filename = f'Volumetria_{safe_name}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_http_methods(["GET"])
def api_levantamiento_fragmento(request, levantamiento_id):
    """Devuelve el fragmento HTML (sólo cuerpo, con CSS scoped inline) de
    una fase del levantamiento — para insertarlo en el editor del reporte
    libre (Fase 5). El HTML incluye su propio <style> scoped.

    Query params:
      ?tipo=levantamiento|propuesta|volumetria|programa-obra
    """
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    import datetime as _dt

    try:
        lev = ProyectoLevantamiento.objects.select_related('proyecto', 'creado_por').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return HttpResponse('<div style="color:#DC2626;">Levantamiento no encontrado</div>', status=404)
    if not _check_access(request.user, lev.proyecto):
        return HttpResponse('<div style="color:#DC2626;">Sin acceso</div>', status=403)

    tipo = (request.GET.get('tipo') or '').strip()

    meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    def _fmt_fecha(iso):
        if not iso:
            return ''
        try:
            dt = _dt.datetime.strptime(str(iso)[:10], '%Y-%m-%d')
        except Exception:
            return iso
        return f'{dt.day:02d} / {meses[dt.month - 1]} / {dt.year}'

    f1 = lev.fase1_data or {}
    f2 = lev.fase2_data or {}

    if tipo == 'levantamiento':
        TIPOS_SERVICIO = ['CCTV', 'Alarma Intrusión', 'Control de Acceso', 'Cableado Estructurado', 'Voceo', 'Telefonía', 'Otros']
        COMPONENTES = [
            'Cámara IP', 'Cámara Analógica', 'NVR', 'DVR', 'Video Balum',
            'Tubería', 'Canalización', 'Escalerilla',
            'Cable Cat5e', 'Cable Cat6', 'Cable Cat6A', 'Fibra Óptica',
            'Gabinete/Rack', 'Serpentines', 'Controladora',
            'PBX', 'Teléfonos', 'Work Station', 'Servidor', 'Switches',
            'Antenas', 'Enlaces', 'Bocinas', 'Cableado',
            'Fuentes de Poder', 'Amplificadores', 'Access Point', 'UPS',
        ]
        selected_svc = set(f1.get('servicios') or [])
        selected_comp = set(f1.get('componentes') or [])
        productos = [{
            'qty': p.get('qty') or 1,
            'unidad': p.get('unidad') or 'PZA',
            'desc': p.get('desc') or '',
            'marca': p.get('marca') or '',
            'modelo': p.get('modelo') or '',
            'ubicacion': p.get('ubicacion') or '',
        } for p in (f1.get('productos') or [])]
        ctx = {
            'lev': lev,
            'cliente': f1.get('cliente') or (lev.proyecto.cliente_nombre if lev.proyecto else ''),
            'contacto': f1.get('contacto') or '',
            'area': f1.get('area') or '',
            'email': f1.get('email') or '',
            'telefono': f1.get('telefono') or '',
            'fecha_fmt': _fmt_fecha(f1.get('fecha') or timezone.localdate().isoformat()),
            'descripcion': f1.get('descripcion') or '',
            'tipos_servicio': [{'name': s, 'on': s in selected_svc} for s in TIPOS_SERVICIO],
            'componentes':    [{'name': c, 'on': c in selected_comp} for c in COMPONENTES],
            # Programa — leer de fase2.programa con fallback a fase1
            'fecha_inicio_fmt': _fmt_fecha((f2.get('programa') or {}).get('fecha_inicio') or f1.get('fecha_inicio')),
            'fecha_fin_fmt':    _fmt_fecha((f2.get('programa') or {}).get('fecha_fin')    or f1.get('fecha_fin')),
            'duracion':      (f2.get('programa') or {}).get('duracion')      or f1.get('duracion') or '',
            'turno':         (f2.get('programa') or {}).get('turno')         or f1.get('turno') or '',
            'apoyo_cliente': (f2.get('programa') or {}).get('apoyo_cliente') or f1.get('apoyo_cliente') or '',
            'personal_req':  (f2.get('programa') or {}).get('personal_req')  or f1.get('personal_req') or '',
            'elev_alto':     (f2.get('programa') or {}).get('elev_alto')     or f1.get('elev_alto') or '',
            'elev_ancho':    (f2.get('programa') or {}).get('elev_ancho')    or f1.get('elev_ancho') or '',
            'elev_modelo':   (f2.get('programa') or {}).get('elev_modelo')   or f1.get('elev_modelo') or '',
            'productos': productos,
        }
        template = 'crm/fragments/_frag_levantamiento.html'

    elif tipo == 'propuesta':
        ctx = {
            'lev': lev,
            'cliente_nombre': f1.get('cliente') or (lev.proyecto.cliente_nombre if lev.proyecto else ''),
            'elaboro':       f2.get('elaboro')       or (lev.creado_por.get_full_name() if lev.creado_por else ''),
            'doc_fecha_fmt': _fmt_fecha(f2.get('doc_fecha') or timezone.localdate().isoformat()),
            'solicitante':   f2.get('solicitante')   or f1.get('contacto') or '',
            'departamento':  f2.get('departamento')  or '',
            'planta':        f2.get('planta')        or f1.get('cliente') or '',
            'areas':         f2.get('areas')         or f1.get('area') or '',
            'tipo_solucion': f2.get('tipo_solucion') or ', '.join(f1.get('servicios') or []),
            'fecha_inst':    f2.get('fecha_inst')    or 'Acordar con usuario',
            'desc_proyecto': f2.get('desc_proyecto') or f1.get('descripcion') or '',
            'especificaciones':  [x for x in (f2.get('especificaciones') or []) if x and x.strip()],
            'comentarios_spec':  [x for x in (f2.get('comentarios_spec') or []) if x and x.strip()],
        }
        template = 'crm/fragments/_frag_propuesta.html'

    elif tipo == 'volumetria':
        ctx = _build_volumetria_ctx(lev, sin_costos=False)
        ctx['doc_fecha_fmt'] = _fmt_fecha(f2.get('doc_fecha') or timezone.localdate().isoformat())
        template = 'crm/fragments/_frag_volumetria.html'

    elif tipo == 'programa-obra':
        if not lev.proyecto_id:
            return HttpResponse('<div style="color:#DC2626;">Este levantamiento no está asociado a un proyecto.</div>', status=400)
        proyecto = lev.proyecto
        f_ini = proyecto.fecha_inicio
        f_fin = proyecto.fecha_fin
        if not f_ini or not f_fin:
            return HttpResponse('<div style="padding:20px;text-align:center;color:#94A3B8;border:1px dashed #CBD5E1;border-radius:8px;">Define fecha de inicio y fin del proyecto para generar el programa de obra.</div>', status=400)
        from .models import ProgramacionActividad
        monday_start = f_ini - _dt.timedelta(days=f_ini.weekday())
        acts_qs = ProgramacionActividad.objects.filter(
            proyecto_key=f'proy_{proyecto.id}'
        ).prefetch_related('responsables').order_by('fecha', 'hora_inicio')
        acts_by_date = {}
        for a in acts_qs:
            if a.fecha:
                acts_by_date.setdefault(a.fecha.isoformat(), []).append(a)
        def fmt_short(d):
            return f'{d.day:02d} {meses[d.month - 1]}'
        weeks = []
        cur = monday_start
        week_num = 1
        total_acts = 0
        total_done = 0
        while cur <= f_fin:
            week_end = cur + _dt.timedelta(days=6)
            has_weekend = False
            days = []
            for i in range(7):
                day = cur + _dt.timedelta(days=i)
                day_acts_raw = acts_by_date.get(day.isoformat(), [])
                day_acts = []
                for a in day_acts_raw:
                    total_acts += 1
                    if a.completada:
                        total_done += 1
                    resp_list = list(a.responsables.all())
                    resp_iniciales = []
                    for u in resp_list:
                        full = (u.get_full_name() or u.username or '').strip()
                        parts = full.split()
                        if len(parts) >= 2:
                            ini = (parts[0][0] + parts[1][0]).upper()
                        else:
                            ini = (full[:2] or '??').upper()
                        resp_iniciales.append({'iniciales': ini, 'nombre': full or u.username})
                    day_acts.append({
                        'titulo': a.titulo,
                        'hora_inicio': a.hora_inicio.strftime('%H:%M') if a.hora_inicio else '—',
                        'hora_fin':    a.hora_fin.strftime('%H:%M') if a.hora_fin else '—',
                        'responsables': resp_iniciales,
                        'completada': a.completada,
                    })
                if i >= 5 and day_acts:
                    has_weekend = True
                days.append({
                    'name': ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo'][i],
                    'date_short': fmt_short(day),
                    'is_weekend': i >= 5,
                    'activities': day_acts,
                })
            weeks.append({
                'num': week_num,
                'range_label': f'{fmt_short(cur)} — {fmt_short(week_end)}',
                'days': [d for d in days if (not d['is_weekend']) or has_weekend],
                'has_weekend': has_weekend,
            })
            cur = cur + _dt.timedelta(days=7)
            week_num += 1
        ctx = {
            'proyecto': proyecto,
            'cliente_nombre': getattr(proyecto, 'cliente_nombre', '') or '',
            'proyecto_fecha_inicio_fmt': _fmt_fecha(f_ini.isoformat()),
            'proyecto_fecha_fin_fmt': _fmt_fecha(f_fin.isoformat()),
            'elaboro': (lev.creado_por.get_full_name() or lev.creado_por.username) if lev.creado_por else '',
            'weeks': weeks,
            'total_actividades': total_acts,
            'total_completadas': total_done,
            'total_pendientes': total_acts - total_done,
        }
        template = 'crm/fragments/_frag_programa_obra.html'
    else:
        return HttpResponse('<div style="color:#DC2626;">Tipo inválido. Usa: levantamiento, propuesta, volumetria, programa-obra</div>', status=400)

    html = render_to_string(template, ctx, request=request)
    return HttpResponse(html, content_type='text/html; charset=utf-8')


@login_required
@require_http_methods(["GET"])
def api_levantamiento_reporte_pdf(request, levantamiento_id):
    """PDF del reporte libre (Fase 5) — el ingeniero escribe HTML rico
    y lo exportamos con el hero Bajanet arriba y footer estándar.

    Params:
      ?download=1 → attachment (por defecto inline para preview)
    """
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.conf import settings
    import os, datetime as _dt

    try:
        lev = ProyectoLevantamiento.objects.select_related('proyecto', 'creado_por').get(id=levantamiento_id)
    except ProyectoLevantamiento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Levantamiento no encontrado'}, status=404)
    if not _check_access(request.user, lev.proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    def _static_file_url(rel_path):
        candidates = []
        if getattr(settings, 'STATIC_ROOT', None):
            candidates.append(os.path.join(settings.STATIC_ROOT, rel_path))
        candidates.append(os.path.join(settings.BASE_DIR, 'app', 'static', rel_path))
        for p in candidates:
            if os.path.exists(p):
                return 'file://' + p
        return ''

    f1 = lev.fase1_data or {}
    f2 = lev.fase2_data or {}
    f5 = lev.fase5_data or {}
    reporte_html = f5.get('reporte_html') or ''

    meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    def _fmt_fecha(val):
        if not val:
            return ''
        try:
            dt = _dt.datetime.strptime(str(val)[:10], '%Y-%m-%d')
        except Exception:
            return val
        return f'{dt.day:02d} / {meses[dt.month - 1]} / {dt.year}'

    ctx = {
        'lev': lev,
        'cliente_nombre': f1.get('cliente') or (lev.proyecto.cliente_nombre if lev.proyecto else ''),
        'elaboro': f2.get('elaboro') or (lev.creado_por.get_full_name() if lev.creado_por else ''),
        'doc_fecha_fmt': _fmt_fecha(f2.get('doc_fecha') or timezone.localdate().isoformat()),
        'reporte_html': reporte_html,
        'bajanet_hero_url': _static_file_url('images/propuesta/bajanet_hero.jpeg'),
        'footer_logos_url': _static_file_url('images/propuesta/footer_logos.png'),
        'empresa_tel': '664 000 0000',
        'empresa_web': 'www.iamet.mx',
        'empresa_direccion': 'Tijuana, B.C.',
    }

    html = render_to_string('crm/levantamiento_reporte_pdf.html', ctx, request=request)
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Error generando PDF: {e}'}, status=500)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '_' for c in (lev.nombre or 'reporte')).strip()[:80] or 'reporte'
    filename = f'Reporte_{safe_name}.pdf'
    disp = 'attachment' if request.GET.get('download') else 'inline'
    response['Content-Disposition'] = f'{disp}; filename="{filename}"'
    return response


@login_required
@require_http_methods(["GET"])
def api_programa_obra_pdf(request, proyecto_id):
    """PDF del Programa de Obra de un proyecto.

    Lee las actividades de ProgramacionActividad (proyecto_key=proy_<id>),
    las agrupa por semana/día según fecha_inicio/fecha_fin del proyecto y
    las renderiza como calendario semanal.

    Params:
      ?download=1 → attachment (por defecto inline para preview)
    """
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from django.conf import settings
    from .models import ProyectoIAMET, ProgramacionActividad
    import os, datetime as _dt

    try:
        proyecto = ProyectoIAMET.objects.get(id=proyecto_id)
    except ProyectoIAMET.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado'}, status=404)
    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso'}, status=403)

    def _static_file_url(rel_path):
        candidates = []
        if getattr(settings, 'STATIC_ROOT', None):
            candidates.append(os.path.join(settings.STATIC_ROOT, rel_path))
        candidates.append(os.path.join(settings.BASE_DIR, 'app', 'static', rel_path))
        for p in candidates:
            if os.path.exists(p):
                return 'file://' + p
        return ''

    # Normalizar fechas
    f_ini = proyecto.fecha_inicio
    f_fin = proyecto.fecha_fin
    if not f_ini or not f_fin:
        return JsonResponse({
            'success': False,
            'error': 'Define fecha de inicio y fin del proyecto antes de generar el PDF.'
        }, status=400)

    # Lunes de la semana del inicio
    monday_start = f_ini - _dt.timedelta(days=f_ini.weekday())

    # Traer actividades
    acts_qs = ProgramacionActividad.objects.filter(
        proyecto_key=f'proy_{proyecto.id}'
    ).prefetch_related('responsables').order_by('fecha', 'hora_inicio')

    acts_by_date = {}
    for a in acts_qs:
        if not a.fecha:
            continue
        acts_by_date.setdefault(a.fecha.isoformat(), []).append(a)

    # Construir semanas
    meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    def fmt_short(d):
        return f'{d.day:02d} {meses[d.month - 1]}'
    def fmt_long(d):
        return f'{d.day:02d} / {meses[d.month - 1]} / {d.year}'

    weeks = []
    cur = monday_start
    week_num = 1
    total_actividades = 0
    total_completadas = 0
    while cur <= f_fin:
        week_end = cur + _dt.timedelta(days=6)
        # ¿Alguna actividad en sábado/domingo? Si sí, mostramos 7 días
        has_weekend = False
        days = []
        for i in range(7):
            day = cur + _dt.timedelta(days=i)
            day_acts_raw = acts_by_date.get(day.isoformat(), [])
            day_acts = []
            for a in day_acts_raw:
                total_actividades += 1
                if a.completada:
                    total_completadas += 1
                resp_list = list(a.responsables.all())
                resp_iniciales = []
                for u in resp_list:
                    full = (u.get_full_name() or u.username or '').strip()
                    parts = full.split()
                    if len(parts) >= 2:
                        ini = (parts[0][0] + parts[1][0]).upper()
                    else:
                        ini = (full[:2] or '??').upper()
                    resp_iniciales.append({'iniciales': ini, 'nombre': full or u.username})
                day_acts.append({
                    'id': a.id,
                    'titulo': a.titulo,
                    'descripcion': a.descripcion or '',
                    'hora_inicio': a.hora_inicio.strftime('%H:%M') if a.hora_inicio else '—',
                    'hora_fin': a.hora_fin.strftime('%H:%M') if a.hora_fin else '—',
                    'responsables': resp_iniciales,
                    'vehiculos': a.vehiculos or '',
                    'completada': a.completada,
                })
            if i >= 5 and day_acts:
                has_weekend = True
            days.append({
                'name': ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo'][i],
                'date': day,
                'date_short': fmt_short(day),
                'iso': day.isoformat(),
                'is_weekend': i >= 5,
                'activities': day_acts,
            })
        # Si la semana está completamente fuera del rango del proyecto (los
        # días laborales), skip — aunque rara vez pasa.
        weeks.append({
            'num': week_num,
            'start': cur,
            'end': week_end,
            'range_label': f'{fmt_short(cur)} — {fmt_short(week_end)}',
            'days': [d for d in days if (not d['is_weekend']) or has_weekend],
            'has_weekend': has_weekend,
        })
        cur = cur + _dt.timedelta(days=7)
        week_num += 1

    # Contexto
    elaboro = ''
    if getattr(proyecto, 'responsable', None):
        try:
            elaboro = proyecto.responsable.get_full_name() or proyecto.responsable.username
        except Exception:
            elaboro = ''
    if not elaboro:
        elaboro = (request.user.get_full_name() or request.user.username) if request.user.is_authenticated else ''

    ctx = {
        'proyecto': proyecto,
        'cliente_nombre': getattr(proyecto, 'cliente_nombre', '') or '',
        'proyecto_fecha_inicio_fmt': fmt_long(f_ini),
        'proyecto_fecha_fin_fmt': fmt_long(f_fin),
        'doc_fecha_fmt': fmt_long(timezone.localdate()),
        'elaboro': elaboro,
        'weeks': weeks,
        'total_actividades': total_actividades,
        'total_completadas': total_completadas,
        'total_pendientes': total_actividades - total_completadas,
        'bajanet_hero_url': _static_file_url('images/propuesta/bajanet_hero.jpeg'),
        'footer_logos_url': _static_file_url('images/propuesta/footer_logos.png'),
        'empresa_tel': '664 000 0000',
        'empresa_web': 'www.iamet.mx',
        'empresa_direccion': 'Tijuana, B.C.',
    }

    html = render_to_string('crm/programa_obra_pdf.html', ctx, request=request)
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Error generando PDF: {e}'}, status=500)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    safe_name = ''.join(c if c.isalnum() or c in ' -_' else '_' for c in (proyecto.nombre or 'programa')).strip()[:80] or 'programa'
    filename = f'ProgramaObra_{safe_name}.pdf'
    disp = 'attachment' if request.GET.get('download') else 'inline'
    response['Content-Disposition'] = f'{disp}; filename="{filename}"'
    return response


@login_required
@require_http_methods(["GET"])
def api_catalogo_productos(request):
    """Buscador universal de productos para Fase 1 / Fase 3 del wizard.

    Fuentes consultadas (en orden): CatalogoCableado, Producto (módulo Compras).
    Query: ?q=<texto>  (mínimo 1 char)  ?limit=<n>
    """
    from .models import CatalogoCableado, Producto
    q = (request.GET.get('q') or '').strip()
    try:
        limit = min(int(request.GET.get('limit') or 20), 50)
    except (ValueError, TypeError):
        limit = 20
    results = []
    if q:
        qs = CatalogoCableado.objects.filter(
            Q(descripcion__icontains=q) |
            Q(numero_parte__icontains=q) |
            Q(marca__icontains=q)
        )[:limit]
        for p in qs:
            results.append({
                'id': f'cab-{p.id}',
                'desc': p.descripcion,
                'marca': p.marca,
                'modelo': p.numero_parte,
                'unidad': 'PZA',
                'precio': float(p.precio_unitario or 0),
                'precio_proveedor': float(p.precio_proveedor or 0),
                'fuente': 'CatalogoCableado',
            })
        # Segunda fuente: catálogo global de productos (módulo Compras).
        # Sólo completamos hasta el cap total `limit` para no exceder lo solicitado.
        restante = limit - len(results)
        if restante > 0:
            qs_prod = Producto.objects.filter(estatus='activo').filter(
                Q(codigo__icontains=q) |
                Q(nombre__icontains=q) |
                Q(descripcion__icontains=q)
            ).select_related('unidad_cfdi')[:restante]
            for p in qs_prod:
                results.append({
                    'id': f'prod-{p.id}',
                    'desc': p.nombre or p.descripcion or p.codigo,
                    'marca': '',
                    'modelo': p.codigo,
                    'unidad': p.unidad_cfdi.clave if p.unidad_cfdi_id else 'PZA',
                    'precio': float(p.costo or 0),
                    'fuente': 'Producto',
                })
    return JsonResponse({'ok': True, 'data': results, 'q': q})


# ═══════════════════════════════════════════════════════════════
#  PWA LEVANTAMIENTOS — Endpoint de sync offline
# ═══════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["POST"])
def api_levantamiento_offline_sync(request):
    """Sube un levantamiento completo (metadata + 5 fases + fotos)
    en una sola llamada multipart. Idempotente: si se reintenta con
    el mismo idempotency_key devuelve el existente en vez de duplicar.

    Para el PWA offline: cuando el ingeniero recupera red, el cliente
    envía todo lo que capturó. Este endpoint es el único que el SW
    tiene que reintentar al dispararse 'online'.

    Campos esperados (multipart/form-data):
    - idempotency_key (str, requerido): UUID generado por el cliente.
    - proyecto_id     (int, requerido): proyecto al que pertenece.
    - nombre          (str, opcional): si no viene se autogenera.
    - status          (str, opcional): default 'borrador'.
    - fase_actual     (int, opcional): default 1.
    - fase1_data ... fase5_data (JSON strings, opcionales).
    - evidencias_meta (JSON array string, opcional):
        [{idempotency_key, comentario, producto_idx}, ...]
    - evidencia_<N>   (archivo): foto asociada al índice N de la meta.

    Respuesta:
    {
        success: true,
        created: bool,           // true si se creó; false si ya existía
        levantamiento_id: int,
        evidencias: [{client_key, server_id, created}, ...]
    }
    """
    idem = (request.POST.get('idempotency_key') or '').strip()
    if not idem or len(idem) < 8 or len(idem) > 64:
        return JsonResponse({'success': False, 'error': 'idempotency_key requerido (8-64 chars).'}, status=400)

    proyecto_id = request.POST.get('proyecto_id')
    try:
        proyecto_id = int(proyecto_id)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'proyecto_id inválido.'}, status=400)

    try:
        proyecto = Proyecto.objects.get(id=proyecto_id)
    except Proyecto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Proyecto no encontrado.'}, status=404)
    if not _check_access(request.user, proyecto):
        return JsonResponse({'success': False, 'error': 'Sin acceso al proyecto.'}, status=403)

    def _parse_fase(name):
        raw = request.POST.get(name)
        if not raw:
            return None
        try:
            v = json.loads(raw)
            return v if isinstance(v, dict) else None
        except (json.JSONDecodeError, ValueError):
            return None

    # ── 1. Idempotency: si ya existe, NO crear de nuevo. Solo
    #      actualizar con los datos más recientes que haya mandado
    #      el cliente (acepta reenvíos con más info).
    lev_created = False
    try:
        lev = ProyectoLevantamiento.objects.get(idempotency_key=idem)
        # Verificar que el proyecto coincida (por seguridad)
        if lev.proyecto_id != proyecto_id:
            return JsonResponse({'success': False, 'error': 'Conflicto: la clave ya pertenece a otro proyecto.'}, status=409)
    except ProyectoLevantamiento.DoesNotExist:
        nombre = (request.POST.get('nombre') or '').strip()
        if not nombre:
            nombre = f'Levantamiento {proyecto.levantamientos.count() + 1}'
        status = request.POST.get('status') or 'borrador'
        if status not in dict(ProyectoLevantamiento.STATUS_CHOICES):
            status = 'borrador'
        try:
            fase_actual = int(request.POST.get('fase_actual') or 1)
            if fase_actual < 1 or fase_actual > 5:
                fase_actual = 1
        except (TypeError, ValueError):
            fase_actual = 1
        lev = ProyectoLevantamiento.objects.create(
            proyecto=proyecto,
            nombre=nombre,
            status=status,
            fase_actual=fase_actual,
            fase1_data=_parse_fase('fase1_data') or {},
            fase2_data=_parse_fase('fase2_data') or {},
            fase3_data=_parse_fase('fase3_data') or {},
            fase4_data=_parse_fase('fase4_data') or {},
            fase5_data=_parse_fase('fase5_data') or {},
            creado_por=request.user,
            idempotency_key=idem,
        )
        lev_created = True

    if not lev_created:
        # Upsert de datos: mergear fases por si el cliente mandó info
        # nueva en un reintento. No borramos lo que ya está.
        dirty = False
        for k in ('fase1_data', 'fase2_data', 'fase3_data', 'fase4_data', 'fase5_data'):
            nueva = _parse_fase(k)
            if isinstance(nueva, dict) and nueva:
                setattr(lev, k, nueva)
                dirty = True
        if request.POST.get('fase_actual'):
            try:
                f = int(request.POST.get('fase_actual'))
                if 1 <= f <= 5 and f != lev.fase_actual:
                    lev.fase_actual = f
                    dirty = True
            except (TypeError, ValueError):
                pass
        if dirty:
            lev.save()

    # ── 2. Evidencias: dedupe por idempotency_key por foto ──
    try:
        evidencias_meta = json.loads(request.POST.get('evidencias_meta') or '[]')
        if not isinstance(evidencias_meta, list):
            evidencias_meta = []
    except (json.JSONDecodeError, ValueError):
        evidencias_meta = []

    resultado_evidencias = []
    for i, meta in enumerate(evidencias_meta):
        if not isinstance(meta, dict):
            continue
        ev_idem = (meta.get('idempotency_key') or '').strip()
        if not ev_idem:
            continue
        file_field = f'evidencia_{i}'

        # ¿Ya existe esta evidencia?
        existing = LevantamientoEvidencia.objects.filter(idempotency_key=ev_idem).first()
        if existing:
            resultado_evidencias.append({
                'client_key': ev_idem,
                'server_id': existing.id,
                'created': False,
            })
            continue

        archivo = request.FILES.get(file_field)
        if not archivo:
            # Meta sin archivo en este reintento (ya lo había subido antes
            # pero sin registrar el idempotency_key, o cliente mandó solo
            # la meta). No podemos crear la evidencia sin archivo.
            resultado_evidencias.append({
                'client_key': ev_idem,
                'server_id': None,
                'created': False,
                'error': 'archivo no presente',
            })
            continue

        try:
            producto_idx = meta.get('producto_idx')
            if producto_idx in (None, '', 'null'):
                producto_idx = None
            else:
                producto_idx = int(producto_idx)
        except (TypeError, ValueError):
            producto_idx = None

        ev = LevantamientoEvidencia.objects.create(
            levantamiento=lev,
            archivo=archivo,
            nombre_original=(archivo.name or '')[:255],
            comentario=(meta.get('comentario') or '')[:255],
            producto_idx=producto_idx,
            subido_por=request.user,
            idempotency_key=ev_idem,
        )
        resultado_evidencias.append({
            'client_key': ev_idem,
            'server_id': ev.id,
            'created': True,
        })

    return JsonResponse({
        'success': True,
        'created': lev_created,
        'levantamiento_id': lev.id,
        'evidencias': resultado_evidencias,
    })
