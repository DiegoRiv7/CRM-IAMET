# ----------------------------------------------------------------------
# views_compras.py — Backend del módulo Compras (Productos y Proveedores).
# ----------------------------------------------------------------------

import csv
import io
import json
import os
import re
from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods

from openpyxl import load_workbook

from .models import (
    Almacen,
    ClaveCFDI,
    Producto,
    Proveedor,
    UnidadCFDI,
)


# ─────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────

PAGE_SIZE_DEFAULT = 25
PAGE_SIZE_MAX = 200


def _json_body(request):
    """Devuelve el body parseado como dict; si falla regresa {}."""
    try:
        if not request.body:
            return {}
        return json.loads(request.body.decode('utf-8'))
    except Exception:
        return {}


def _paginate(qs, request):
    try:
        page = int(request.GET.get('page', '1') or '1')
    except ValueError:
        page = 1
    try:
        page_size = int(request.GET.get('page_size', str(PAGE_SIZE_DEFAULT)) or str(PAGE_SIZE_DEFAULT))
    except ValueError:
        page_size = PAGE_SIZE_DEFAULT
    if page < 1:
        page = 1
    if page_size < 1:
        page_size = PAGE_SIZE_DEFAULT
    if page_size > PAGE_SIZE_MAX:
        page_size = PAGE_SIZE_MAX
    total = qs.count()
    total_pages = (total + page_size - 1) // page_size if page_size else 1
    if total_pages == 0:
        total_pages = 1
    start = (page - 1) * page_size
    end = start + page_size
    return list(qs[start:end]), page, page_size, total, total_pages


def _to_decimal(value, default=Decimal('0')):
    if value is None or value == '':
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        raise ValidationError("Valor numérico inválido.")


def _serialize_almacen(a):
    return {
        'id': a.id,
        'nombre': a.nombre,
        'activo': a.activo,
    }


def _serialize_clave_cfdi(c):
    return {
        'id': c.id,
        'clave': c.clave,
        'descripcion': c.descripcion,
        'tipo': c.tipo,
    }


def _serialize_unidad_cfdi(u):
    return {
        'id': u.id,
        'clave': u.clave,
        'descripcion': u.descripcion,
    }


def _serialize_producto(p, include_almacenes=True):
    data = {
        'id': p.id,
        'codigo': p.codigo,
        'nombre': p.nombre,
        'descripcion': p.descripcion,
        'costo': str(p.costo),
        'moneda': p.moneda,
        'iva': str(p.iva),
        'clave_cfdi_id': p.clave_cfdi_id,
        'clave_cfdi': p.clave_cfdi.clave if p.clave_cfdi_id else None,
        'clave_cfdi_descripcion': p.clave_cfdi.descripcion if p.clave_cfdi_id else None,
        'unidad_cfdi_id': p.unidad_cfdi_id,
        'unidad_cfdi': p.unidad_cfdi.clave if p.unidad_cfdi_id else None,
        'unidad_cfdi_descripcion': p.unidad_cfdi.descripcion if p.unidad_cfdi_id else None,
        'tipo': p.tipo,
        'estatus': p.estatus,
        'created_at': p.created_at.isoformat() if p.created_at else None,
        'updated_at': p.updated_at.isoformat() if p.updated_at else None,
    }
    if include_almacenes:
        data['almacenes'] = [{'id': a.id, 'nombre': a.nombre} for a in p.almacenes.all()]
    return data


def _serialize_proveedor(p):
    return {
        'id': p.id,
        'razon_social': p.razon_social,
        'rfc': p.rfc,
        'tipo_persona': p.tipo_persona,
        'dias_credito': p.dias_credito,
        'monto_credito': str(p.monto_credito),
        'banco': p.banco,
        'cuenta_bancaria': p.cuenta_bancaria,
        'clabe': p.clabe,
        'cuenta_contable': p.cuenta_contable,
        'calle': p.calle,
        'numero': p.numero,
        'colonia': p.colonia,
        'ciudad': p.ciudad,
        'estado': p.estado,
        'cp': p.cp,
        'estatus': p.estatus,
        'created_at': p.created_at.isoformat() if p.created_at else None,
        'updated_at': p.updated_at.isoformat() if p.updated_at else None,
    }


# ─────────────────────────────────────────────────────────────────
# Productos
# ─────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(["GET", "POST"])
def productos_list_create(request):
    """GET → listado paginado y filtrable. POST → crear producto."""
    if request.method == 'GET':
        qs = Producto.objects.select_related('clave_cfdi', 'unidad_cfdi').prefetch_related('almacenes')
        q = (request.GET.get('q') or '').strip()
        if q:
            qs = qs.filter(Q(codigo__icontains=q) | Q(nombre__icontains=q) | Q(descripcion__icontains=q))
        tipo = (request.GET.get('tipo') or '').strip().upper()
        if tipo in ('PRODUCTO', 'SERVICIO'):
            qs = qs.filter(tipo=tipo)
        estatus = (request.GET.get('estatus') or '').strip().lower()
        if estatus in ('activo', 'inactivo'):
            qs = qs.filter(estatus=estatus)
        sort = (request.GET.get('sort') or '-created_at').strip()
        allowed = {'codigo', '-codigo', 'nombre', '-nombre', 'created_at', '-created_at', 'costo', '-costo'}
        if sort not in allowed:
            sort = '-created_at'
        qs = qs.order_by(sort)
        items, page, page_size, total, total_pages = _paginate(qs, request)
        return JsonResponse({
            'items': [_serialize_producto(p) for p in items],
            'page': page,
            'page_size': page_size,
            'total': total,
            'total_pages': total_pages,
        })

    # POST → crear
    data = _json_body(request)
    errors = {}

    codigo = (data.get('codigo') or '').strip()
    nombre = (data.get('nombre') or '').strip()
    descripcion = (data.get('descripcion') or '').strip()
    moneda = (data.get('moneda') or 'MXN').strip().upper()
    clave_cfdi_id = data.get('clave_cfdi_id') or data.get('clave_cfdi')
    unidad_cfdi_id = data.get('unidad_cfdi_id') or data.get('unidad_cfdi')
    almacenes_ids = data.get('almacenes') or data.get('almacenes_ids') or []
    estatus = (data.get('estatus') or 'activo').strip().lower()

    if not codigo:
        errors['codigo'] = 'El código es obligatorio.'
    if not nombre:
        errors['nombre'] = 'El nombre es obligatorio.'
    if not descripcion:
        errors['descripcion'] = 'La descripción es obligatoria.'
    if moneda not in ('MXN', 'USD'):
        errors['moneda'] = 'Moneda debe ser MXN o USD.'
    if not clave_cfdi_id:
        errors['clave_cfdi'] = 'Clave CFDI obligatoria.'
    if not unidad_cfdi_id:
        errors['unidad_cfdi'] = 'Unidad CFDI obligatoria.'
    if estatus not in ('activo', 'inactivo'):
        errors['estatus'] = 'Estatus inválido.'

    try:
        costo = _to_decimal(data.get('costo'), Decimal('0'))
        if costo < 0:
            errors['costo'] = 'El costo no puede ser negativo.'
    except ValidationError:
        errors['costo'] = 'Costo inválido.'
        costo = Decimal('0')

    try:
        iva = _to_decimal(data.get('iva', 16), Decimal('16'))
        if iva not in (Decimal('8'), Decimal('8.00'), Decimal('16'), Decimal('16.00')):
            errors['iva'] = 'IVA debe ser 8% o 16%.'
    except ValidationError:
        errors['iva'] = 'IVA inválido.'
        iva = Decimal('16')

    clave_cfdi_obj = None
    unidad_cfdi_obj = None
    if clave_cfdi_id and 'clave_cfdi' not in errors:
        clave_cfdi_obj = ClaveCFDI.objects.filter(pk=clave_cfdi_id).first()
        if not clave_cfdi_obj:
            errors['clave_cfdi'] = 'Clave CFDI no existe.'
    if unidad_cfdi_id and 'unidad_cfdi' not in errors:
        unidad_cfdi_obj = UnidadCFDI.objects.filter(pk=unidad_cfdi_id).first()
        if not unidad_cfdi_obj:
            errors['unidad_cfdi'] = 'Unidad CFDI no existe.'

    # Si servicio → forzar unidad E48
    if clave_cfdi_obj and clave_cfdi_obj.tipo == 'servicio':
        unidad_e48 = UnidadCFDI.objects.filter(clave='E48').first()
        if unidad_e48:
            unidad_cfdi_obj = unidad_e48

    if errors:
        return JsonResponse({'ok': False, 'errors': errors}, status=400)

    try:
        with transaction.atomic():
            p = Producto(
                codigo=codigo,
                nombre=nombre,
                descripcion=descripcion,
                costo=costo,
                moneda=moneda,
                iva=iva,
                clave_cfdi=clave_cfdi_obj,
                unidad_cfdi=unidad_cfdi_obj,
                estatus=estatus,
                created_by=request.user if request.user.is_authenticated else None,
            )
            p.save()
            if almacenes_ids:
                almacenes_qs = Almacen.objects.filter(pk__in=almacenes_ids)
                p.almacenes.set(almacenes_qs)
    except IntegrityError:
        return JsonResponse({'ok': False, 'errors': {'codigo': 'Ya existe un producto con ese código.'}}, status=400)
    except ValidationError as e:
        return JsonResponse({'ok': False, 'errors': getattr(e, 'message_dict', {'__all__': str(e)})}, status=400)
    except Exception as e:
        return JsonResponse({'ok': False, 'errors': {'__all__': str(e)}}, status=400)

    return JsonResponse({'ok': True, 'id': p.id, 'producto': _serialize_producto(p)})


@login_required
@require_http_methods(["GET", "PUT", "DELETE"])
def producto_detail(request, id):
    p = Producto.objects.select_related('clave_cfdi', 'unidad_cfdi').prefetch_related('almacenes', 'snapshots').filter(pk=id).first()
    if not p:
        return JsonResponse({'ok': False, 'errors': {'__all__': 'Producto no encontrado.'}}, status=404)

    if request.method == 'GET':
        return JsonResponse({'ok': True, 'producto': _serialize_producto(p)})

    if request.method == 'PUT':
        data = _json_body(request)
        errors = {}

        # codigo no editable
        if 'codigo' in data and (data.get('codigo') or '').strip() != p.codigo:
            errors['codigo'] = 'El código no es editable.'

        if 'nombre' in data:
            nombre = (data.get('nombre') or '').strip()
            if not nombre:
                errors['nombre'] = 'El nombre es obligatorio.'
            else:
                p.nombre = nombre
        if 'descripcion' in data:
            descripcion = (data.get('descripcion') or '').strip()
            if not descripcion:
                errors['descripcion'] = 'La descripción es obligatoria.'
            else:
                p.descripcion = descripcion
        if 'moneda' in data:
            moneda = (data.get('moneda') or '').strip().upper()
            if moneda not in ('MXN', 'USD'):
                errors['moneda'] = 'Moneda debe ser MXN o USD.'
            else:
                p.moneda = moneda
        if 'costo' in data:
            try:
                costo = _to_decimal(data.get('costo'), p.costo)
                if costo < 0:
                    errors['costo'] = 'El costo no puede ser negativo.'
                else:
                    p.costo = costo
            except ValidationError:
                errors['costo'] = 'Costo inválido.'
        if 'iva' in data:
            try:
                iva = _to_decimal(data.get('iva'), p.iva)
                if iva not in (Decimal('8'), Decimal('8.00'), Decimal('16'), Decimal('16.00')):
                    errors['iva'] = 'IVA debe ser 8% o 16%.'
                else:
                    p.iva = iva
            except ValidationError:
                errors['iva'] = 'IVA inválido.'
        if 'estatus' in data:
            estatus = (data.get('estatus') or '').strip().lower()
            if estatus not in ('activo', 'inactivo'):
                errors['estatus'] = 'Estatus inválido.'
            else:
                p.estatus = estatus

        # clave_cfdi: NO permite cambiar de tipo (Producto vs Servicio).
        if ('clave_cfdi_id' in data) or ('clave_cfdi' in data):
            new_clave_id = data.get('clave_cfdi_id') or data.get('clave_cfdi')
            if new_clave_id:
                new_clave = ClaveCFDI.objects.filter(pk=new_clave_id).first()
                if not new_clave:
                    errors['clave_cfdi'] = 'Clave CFDI no existe.'
                else:
                    derived = 'PRODUCTO' if new_clave.tipo == 'producto' else 'SERVICIO'
                    if p.tipo and p.tipo != derived:
                        errors['clave_cfdi'] = 'No se puede cambiar el tipo (Producto/Servicio).'
                    else:
                        p.clave_cfdi = new_clave
                        # Si ahora es servicio, forzar unidad E48
                        if new_clave.tipo == 'servicio':
                            e48 = UnidadCFDI.objects.filter(clave='E48').first()
                            if e48:
                                p.unidad_cfdi = e48

        if ('unidad_cfdi_id' in data) or ('unidad_cfdi' in data):
            new_unidad_id = data.get('unidad_cfdi_id') or data.get('unidad_cfdi')
            if new_unidad_id:
                new_unidad = UnidadCFDI.objects.filter(pk=new_unidad_id).first()
                if not new_unidad:
                    errors['unidad_cfdi'] = 'Unidad CFDI no existe.'
                else:
                    # Servicios siempre E48
                    if p.clave_cfdi and p.clave_cfdi.tipo == 'servicio' and new_unidad.clave != 'E48':
                        e48 = UnidadCFDI.objects.filter(clave='E48').first()
                        if e48:
                            p.unidad_cfdi = e48
                    else:
                        p.unidad_cfdi = new_unidad

        if errors:
            return JsonResponse({'ok': False, 'errors': errors}, status=400)

        try:
            with transaction.atomic():
                p.save()
                if 'almacenes' in data or 'almacenes_ids' in data:
                    almacenes_ids = data.get('almacenes') or data.get('almacenes_ids') or []
                    almacenes_qs = Almacen.objects.filter(pk__in=almacenes_ids)
                    p.almacenes.set(almacenes_qs)
        except IntegrityError as e:
            return JsonResponse({'ok': False, 'errors': {'__all__': str(e)}}, status=400)
        except ValidationError as e:
            return JsonResponse({'ok': False, 'errors': getattr(e, 'message_dict', {'__all__': str(e)})}, status=400)

        return JsonResponse({'ok': True, 'id': p.id, 'producto': _serialize_producto(p)})

    # DELETE → soft if has snapshots, hard otherwise
    has_snapshots = p.snapshots.exists()
    if has_snapshots:
        p.estatus = 'inactivo'
        p.save(update_fields=['estatus', 'updated_at'])
        return JsonResponse({'ok': True, 'mode': 'soft', 'id': id})
    else:
        p.delete()
        return JsonResponse({'ok': True, 'mode': 'hard', 'id': id})


@login_required
@require_http_methods(["POST"])
def productos_import(request):
    """
    POST archivo XLSX o CSV con productos. Upsert por `external_id` (UUID del Excel).

    Soporta:
    - Excel real del cliente: columnas en español (`ID`, `No Producto`, `Codigo`,
      `Titulo`, `Costo`, `Tipo de IVA.`, `Moneda`, `Clave CFDI`, `Unidad CFDI`,
      `Descripcion`, etc.).
    - CSV legacy (compat): columnas en inglés (`codigo`, `nombre`, ...).

    Reglas:
    - external_id presente y existe → UPDATE (sin pisar campos con valores vacíos).
    - external_id presente pero no existe → CREATE.
    - external_id ausente → CREATE con warning.
    - NUNCA hace DELETE: productos que no estén en el archivo se quedan tal cual.
    - costo solo se actualiza si Excel trae > 0 (no pisa costo capturado a mano).

    Form-data:
    - file: archivo .xlsx o .csv (requerido).
    - crear_cfdi_faltantes: '1'/'true' para crear ClaveCFDI/UnidadCFDI faltantes
      en lugar de fallar la fila.
    """
    f = request.FILES.get('file') or request.FILES.get('csv')
    if not f:
        return JsonResponse({'ok': False, 'errors': {'file': 'Archivo requerido.'}}, status=400)

    crear_cfdi_faltantes = (request.POST.get('crear_cfdi_faltantes') or '').strip().lower() in ('1', 'true', 'yes', 'on')

    nombre_archivo = (getattr(f, 'name', '') or '').lower()
    ext = os.path.splitext(nombre_archivo)[1]

    # ── Leer filas del archivo ────────────────────────────────────
    rows = []  # lista de dicts {col_normalizada: valor}
    wb = None

    def _norm_key(k):
        return re.sub(r'\s+', ' ', str(k or '')).strip().lower()

    try:
        if ext == '.xlsx':
            try:
                wb = load_workbook(filename=f, read_only=True, data_only=True)
            except Exception as e:
                return JsonResponse({'ok': False, 'errors': {'file': f'No se pudo leer XLSX: {e}'}}, status=400)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [_norm_key(c) for c in row]
                    continue
                if row is None:
                    continue
                # Saltar filas completamente vacías
                if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                    continue
                d = {}
                for h, v in zip(headers, row):
                    if not h:
                        continue
                    d[h] = v
                rows.append(d)
        elif ext == '.csv' or ext == '':
            try:
                raw = f.read().decode('utf-8-sig')
            except Exception:
                try:
                    f.seek(0)
                    raw = f.read().decode('latin-1')
                except Exception as e:
                    return JsonResponse({'ok': False, 'errors': {'file': f'No se pudo leer CSV: {e}'}}, status=400)
            reader = csv.DictReader(io.StringIO(raw))
            for r in reader:
                rows.append({_norm_key(k): v for k, v in r.items()})
        else:
            return JsonResponse({'ok': False, 'errors': {'file': f'Formato no soportado: {ext}. Usa .xlsx o .csv.'}}, status=400)
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

    # ── Aliases (todos en lower-case, espacios colapsados) ───────
    ALIASES = {
        'external_id': ['id', 'external_id'],
        'no_producto': ['no producto', 'no_producto', 'numero', 'número'],
        'codigo': ['codigo', 'código'],
        'nombre': ['titulo', 'título', 'nombre'],
        'descripcion': ['descripcion', 'descripción'],
        'costo': ['costo'],
        'moneda': ['moneda'],
        'iva': ['tipo de iva.', 'tipo de iva', 'iva'],
        'clave_cfdi': ['clave cfdi', 'clave_cfdi'],
        'unidad_cfdi': ['unidad cfdi', 'unidad_cfdi'],
        'unidad_medida': ['unidad de medida', 'unidad_medida'],
        'almacenes': ['almacenes'],
    }

    def _pick(row, field):
        for alias in ALIASES.get(field, []):
            if alias in row:
                v = row[alias]
                if v is None:
                    return ''
                return v
        return ''

    def _to_str(v):
        if v is None:
            return ''
        if isinstance(v, str):
            return v.strip()
        return str(v).strip()

    def _to_int(v):
        s = _to_str(v)
        if not s:
            return None
        # Permite floats tipo "12.0"
        try:
            return int(s)
        except (ValueError, TypeError):
            try:
                return int(float(s))
            except (ValueError, TypeError):
                return None

    def _parse_iva(v):
        s = _to_str(v)
        if not s:
            return Decimal('16'), False
        m = re.search(r'(\d+(?:\.\d+)?)', s)
        if not m:
            return Decimal('16'), True  # warning: no parseable
        try:
            n = Decimal(m.group(1))
        except InvalidOperation:
            return Decimal('16'), True
        if n in (Decimal('8'), Decimal('8.00'), Decimal('16'), Decimal('16.00')):
            return n, False
        return Decimal('16'), True  # fuera de catálogo → 16 con warning

    def _parse_costo(v):
        if v is None or v == '':
            return Decimal('0')
        # Si es numérico nativo (XLSX devuelve int/float)
        if isinstance(v, (int, float, Decimal)):
            try:
                return Decimal(str(v))
            except InvalidOperation:
                return Decimal('0')
        s = _to_str(v)
        if not s:
            return Decimal('0')
        # Limpia comas de miles y símbolos comunes
        s = s.replace(',', '').replace('$', '').strip()
        try:
            return Decimal(s)
        except InvalidOperation:
            return Decimal('0')

    # ── Caches para evitar N consultas ──────────────────────────
    cache_claves = {c.clave: c for c in ClaveCFDI.objects.all()}
    cache_unidades = {u.clave: u for u in UnidadCFDI.objects.all()}

    created = 0
    updated = 0
    skipped = 0
    failed = 0
    claves_cfdi_creadas = 0
    unidades_cfdi_creadas = 0
    errors_rows = []

    for idx, row in enumerate(rows, start=2):  # fila 1 = header
        row_errors = {}

        external_id = _to_str(_pick(row, 'external_id'))
        no_producto = _to_int(_pick(row, 'no_producto'))
        codigo = _to_str(_pick(row, 'codigo'))
        nombre = _to_str(_pick(row, 'nombre'))
        descripcion = _to_str(_pick(row, 'descripcion'))
        costo = _parse_costo(_pick(row, 'costo'))
        moneda = _to_str(_pick(row, 'moneda')).upper()
        iva, iva_warn = _parse_iva(_pick(row, 'iva'))
        clave_cfdi_str = _to_str(_pick(row, 'clave_cfdi'))
        unidad_cfdi_str = _to_str(_pick(row, 'unidad_cfdi'))
        unidad_medida_str = _to_str(_pick(row, 'unidad_medida'))
        almacenes_str = _to_str(_pick(row, 'almacenes'))

        # Saltar filas claramente vacías (sin código y sin título)
        if not codigo and not nombre and not external_id:
            skipped += 1
            continue

        # En CREATE necesitamos al menos código + nombre (la BD los exige
        # NOT NULL). En UPDATE el matched-by external_id puede traer solo
        # algunos campos para refrescar precios / CFDI; ahí la falta no es
        # error — solo no se sobrescribe.
        es_update = bool(external_id and Producto.objects.filter(external_id=external_id).exists())
        if not es_update:
            if not codigo:
                row_errors['codigo'] = 'Requerido para crear'
            if not nombre:
                row_errors['nombre'] = 'Requerido para crear'

        # Defaults / fallbacks
        if moneda not in ('MXN', 'USD'):
            moneda = 'MXN'
        if not descripcion:
            descripcion = nombre

        # Unidad CFDI: si > 5 chars o vacía → fallback a unidad_medida → 'H87'
        if not unidad_cfdi_str or len(unidad_cfdi_str) > 5:
            if unidad_medida_str and len(unidad_medida_str) <= 5:
                unidad_cfdi_str = unidad_medida_str
            else:
                unidad_cfdi_str = 'H87'

        # Clave CFDI: validar contenido y longitud
        if not clave_cfdi_str:
            row_errors['clave_cfdi'] = 'Requerida'
        elif len(clave_cfdi_str) > 8:
            row_errors['clave_cfdi'] = f'Longitud inválida (>8): {clave_cfdi_str}'

        # Buscar/crear ClaveCFDI
        clave_obj = None
        if clave_cfdi_str and 'clave_cfdi' not in row_errors:
            clave_obj = cache_claves.get(clave_cfdi_str)
            if not clave_obj:
                if crear_cfdi_faltantes:
                    try:
                        clave_obj = ClaveCFDI.objects.create(
                            clave=clave_cfdi_str,
                            descripcion='Importada desde Excel — revisar',
                            tipo='producto',
                        )
                        cache_claves[clave_cfdi_str] = clave_obj
                        claves_cfdi_creadas += 1
                    except Exception as e:
                        row_errors['clave_cfdi'] = f'No se pudo crear: {e}'
                else:
                    row_errors['clave_cfdi'] = 'no existe (sube con crear_cfdi_faltantes activado para crearla)'

        # Buscar/crear UnidadCFDI
        unidad_obj = None
        if unidad_cfdi_str and 'unidad_cfdi' not in row_errors:
            unidad_obj = cache_unidades.get(unidad_cfdi_str)
            if not unidad_obj:
                if crear_cfdi_faltantes:
                    try:
                        unidad_obj = UnidadCFDI.objects.create(
                            clave=unidad_cfdi_str,
                            descripcion='Importada desde Excel — revisar',
                        )
                        cache_unidades[unidad_cfdi_str] = unidad_obj
                        unidades_cfdi_creadas += 1
                    except Exception as e:
                        row_errors['unidad_cfdi'] = f'No se pudo crear: {e}'
                else:
                    row_errors['unidad_cfdi'] = 'no existe (sube con crear_cfdi_faltantes activado para crearla)'

        # Si servicio → forzar unidad E48
        if clave_obj and clave_obj.tipo == 'servicio':
            e48 = cache_unidades.get('E48') or UnidadCFDI.objects.filter(clave='E48').first()
            if e48:
                cache_unidades['E48'] = e48
                unidad_obj = e48

        if row_errors:
            failed += 1
            errors_rows.append({
                'row': idx,
                'codigo': codigo,
                'external_id': external_id,
                'errors': row_errors,
            })
            continue

        # ── Upsert ────────────────────────────────────────────
        try:
            with transaction.atomic():
                existing = None
                if external_id:
                    existing = Producto.objects.filter(external_id=external_id).first()

                if existing is not None:
                    # UPDATE — no pisar con vacíos
                    changed = False
                    if nombre:
                        if existing.nombre != nombre:
                            existing.nombre = nombre
                            changed = True
                    if descripcion:
                        if existing.descripcion != descripcion:
                            existing.descripcion = descripcion
                            changed = True
                    if codigo:
                        if existing.codigo != codigo:
                            existing.codigo = codigo
                            changed = True
                    if clave_obj and existing.clave_cfdi_id != clave_obj.id:
                        # OJO: si el tipo cambia, el save() del modelo lo bloquea.
                        existing.clave_cfdi = clave_obj
                        changed = True
                    if unidad_obj and existing.unidad_cfdi_id != unidad_obj.id:
                        existing.unidad_cfdi = unidad_obj
                        changed = True
                    if moneda and existing.moneda != moneda:
                        existing.moneda = moneda
                        changed = True
                    if iva and existing.iva != iva:
                        existing.iva = iva
                        changed = True
                    if costo and costo > 0 and existing.costo != costo:
                        existing.costo = costo
                        changed = True
                    if no_producto is not None and existing.no_producto != no_producto:
                        existing.no_producto = no_producto
                        changed = True
                    if changed:
                        existing.save()
                        updated += 1
                    else:
                        # Sin cambios reales — lo contamos como updated igualmente
                        # (la fila fue procesada y matcheada).
                        updated += 1
                else:
                    # CREATE — código y nombre garantizados por la validación
                    # de arriba (es_update == False).
                    p = Producto(
                        external_id=external_id or None,
                        no_producto=no_producto,
                        codigo=codigo,
                        nombre=nombre,
                        descripcion=descripcion or nombre,
                        costo=costo if costo and costo > 0 else Decimal('0'),
                        moneda=moneda,
                        iva=iva,
                        clave_cfdi=clave_obj,
                        unidad_cfdi=unidad_obj,
                        created_by=request.user if request.user.is_authenticated else None,
                    )
                    p.save()
                    if almacenes_str:
                        nombres = [n.strip() for n in almacenes_str.split('|') if n.strip()]
                        if nombres:
                            almacenes_qs = Almacen.objects.filter(nombre__in=nombres)
                            p.almacenes.set(almacenes_qs)
                    created += 1
                    if not external_id:
                        # Warning: fila sin UUID estable
                        errors_rows.append({
                            'row': idx,
                            'codigo': codigo,
                            'external_id': '',
                            'errors': {'external_id': 'WARNING: fila sin external_id; se creó pero re-imports duplicarán.'},
                        })
        except IntegrityError as e:
            failed += 1
            errors_rows.append({
                'row': idx,
                'codigo': codigo,
                'external_id': external_id,
                'errors': {'__all__': f'IntegrityError: {e}'},
            })
        except ValidationError as e:
            failed += 1
            errors_rows.append({
                'row': idx,
                'codigo': codigo,
                'external_id': external_id,
                'errors': getattr(e, 'message_dict', {'__all__': str(e)}),
            })
        except Exception as e:
            failed += 1
            errors_rows.append({
                'row': idx,
                'codigo': codigo,
                'external_id': external_id,
                'errors': {'__all__': str(e)},
            })

    return JsonResponse({
        'ok': True,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'failed': failed,
        'claves_cfdi_creadas': claves_cfdi_creadas,
        'unidades_cfdi_creadas': unidades_cfdi_creadas,
        'errors': errors_rows,
    })


@login_required
@require_http_methods(["GET"])
def productos_export(request):
    """Descarga CSV con todos los productos."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="productos.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'codigo', 'nombre', 'descripcion', 'costo', 'moneda', 'iva',
        'clave_cfdi', 'unidad_cfdi', 'tipo', 'estatus', 'almacenes',
    ])
    qs = Producto.objects.select_related('clave_cfdi', 'unidad_cfdi').prefetch_related('almacenes')
    for p in qs.iterator():
        almacenes_str = '|'.join(a.nombre for a in p.almacenes.all())
        writer.writerow([
            p.codigo, p.nombre, p.descripcion, str(p.costo), p.moneda, str(p.iva),
            p.clave_cfdi.clave if p.clave_cfdi_id else '',
            p.unidad_cfdi.clave if p.unidad_cfdi_id else '',
            p.tipo, p.estatus, almacenes_str,
        ])
    return response


# ─────────────────────────────────────────────────────────────────
# Proveedores
# ─────────────────────────────────────────────────────────────────

RFC_REGEX = re.compile(r'^[A-ZÑ&]{3,4}\d{6}[A-Z\d]{3}$')
CLABE_REGEX = re.compile(r'^\d{18}$')


def _validate_proveedor_payload(data, current=None):
    """Devuelve (cleaned_dict, errors_dict)."""
    errors = {}
    cleaned = {}

    razon_social = (data.get('razon_social') or '').strip()
    if 'razon_social' in data or current is None:
        if not razon_social:
            errors['razon_social'] = 'La razón social es obligatoria.'
        else:
            cleaned['razon_social'] = razon_social

    if 'rfc' in data or current is None:
        rfc_raw = (data.get('rfc') or '').strip().upper()
        if rfc_raw:
            if not RFC_REGEX.match(rfc_raw):
                errors['rfc'] = 'RFC inválido (formato SAT).'
            else:
                cleaned['rfc'] = rfc_raw
                cleaned['tipo_persona'] = 'moral' if len(rfc_raw) == 12 else 'fisica'
        else:
            cleaned['rfc'] = None
            cleaned['tipo_persona'] = ''

    if 'dias_credito' in data or current is None:
        dc = data.get('dias_credito', 0) or 0
        try:
            dc = int(dc)
            if dc < 0:
                errors['dias_credito'] = 'No negativo.'
            else:
                cleaned['dias_credito'] = dc
        except (TypeError, ValueError):
            errors['dias_credito'] = 'Inválido.'

    if 'monto_credito' in data or current is None:
        try:
            mc = _to_decimal(data.get('monto_credito', 0), Decimal('0'))
            if mc < 0:
                errors['monto_credito'] = 'No negativo.'
            else:
                cleaned['monto_credito'] = mc
        except ValidationError:
            errors['monto_credito'] = 'Inválido.'

    if 'clabe' in data or current is None:
        clabe = (data.get('clabe') or '').strip()
        if clabe and not CLABE_REGEX.match(clabe):
            errors['clabe'] = 'CLABE debe tener 18 dígitos.'
        else:
            cleaned['clabe'] = clabe

    if 'cuenta_contable' in data or current is None:
        cc = (data.get('cuenta_contable') or '').strip()
        if not cc:
            errors['cuenta_contable'] = 'La cuenta contable es obligatoria.'
        else:
            cleaned['cuenta_contable'] = cc

    # Plain string fields
    for f in ('banco', 'cuenta_bancaria', 'calle', 'numero', 'colonia', 'ciudad', 'estado', 'cp'):
        if f in data:
            cleaned[f] = (data.get(f) or '').strip()

    if 'estatus' in data:
        est = (data.get('estatus') or 'activo').strip().lower()
        if est not in ('activo', 'inactivo'):
            errors['estatus'] = 'Estatus inválido.'
        else:
            cleaned['estatus'] = est

    return cleaned, errors


@login_required
@require_http_methods(["GET", "POST"])
def proveedores_list_create(request):
    if request.method == 'GET':
        qs = Proveedor.objects.all()
        q = (request.GET.get('q') or '').strip()
        if q:
            qs = qs.filter(Q(razon_social__icontains=q) | Q(rfc__icontains=q) | Q(cuenta_contable__icontains=q))
        estatus = (request.GET.get('estatus') or '').strip().lower()
        if estatus in ('activo', 'inactivo'):
            qs = qs.filter(estatus=estatus)
        tipo_persona = (request.GET.get('tipo_persona') or '').strip().lower()
        if tipo_persona in ('moral', 'fisica'):
            qs = qs.filter(tipo_persona=tipo_persona)
        sort = (request.GET.get('sort') or '-created_at').strip()
        allowed = {'razon_social', '-razon_social', 'rfc', '-rfc', 'created_at', '-created_at'}
        if sort not in allowed:
            sort = '-created_at'
        qs = qs.order_by(sort)
        items, page, page_size, total, total_pages = _paginate(qs, request)
        return JsonResponse({
            'items': [_serialize_proveedor(p) for p in items],
            'page': page,
            'page_size': page_size,
            'total': total,
            'total_pages': total_pages,
        })

    data = _json_body(request)
    cleaned, errors = _validate_proveedor_payload(data, current=None)
    if errors:
        return JsonResponse({'ok': False, 'errors': errors}, status=400)

    try:
        with transaction.atomic():
            p = Proveedor(
                razon_social=cleaned.get('razon_social', ''),
                rfc=cleaned.get('rfc'),
                tipo_persona=cleaned.get('tipo_persona', ''),
                dias_credito=cleaned.get('dias_credito', 0),
                monto_credito=cleaned.get('monto_credito', Decimal('0')),
                banco=cleaned.get('banco', ''),
                cuenta_bancaria=cleaned.get('cuenta_bancaria', ''),
                clabe=cleaned.get('clabe', ''),
                cuenta_contable=cleaned.get('cuenta_contable', ''),
                calle=cleaned.get('calle', ''),
                numero=cleaned.get('numero', ''),
                colonia=cleaned.get('colonia', ''),
                ciudad=cleaned.get('ciudad', ''),
                estado=cleaned.get('estado', ''),
                cp=cleaned.get('cp', ''),
                estatus=cleaned.get('estatus', 'activo'),
                created_by=request.user if request.user.is_authenticated else None,
            )
            p.save()
    except IntegrityError:
        return JsonResponse({'ok': False, 'errors': {'rfc': 'Ya existe un proveedor con ese RFC.'}}, status=400)
    except ValidationError as e:
        return JsonResponse({'ok': False, 'errors': getattr(e, 'message_dict', {'__all__': str(e)})}, status=400)
    except Exception as e:
        return JsonResponse({'ok': False, 'errors': {'__all__': str(e)}}, status=400)

    return JsonResponse({'ok': True, 'id': p.id, 'proveedor': _serialize_proveedor(p)})


@login_required
@require_http_methods(["GET", "PUT", "DELETE"])
def proveedor_detail(request, id):
    p = Proveedor.objects.prefetch_related('snapshots').filter(pk=id).first()
    if not p:
        return JsonResponse({'ok': False, 'errors': {'__all__': 'Proveedor no encontrado.'}}, status=404)

    if request.method == 'GET':
        return JsonResponse({'ok': True, 'proveedor': _serialize_proveedor(p)})

    if request.method == 'PUT':
        data = _json_body(request)
        cleaned, errors = _validate_proveedor_payload(data, current=p)
        if errors:
            return JsonResponse({'ok': False, 'errors': errors}, status=400)
        try:
            with transaction.atomic():
                for k, v in cleaned.items():
                    setattr(p, k, v)
                p.save()
        except IntegrityError:
            return JsonResponse({'ok': False, 'errors': {'rfc': 'Ya existe un proveedor con ese RFC.'}}, status=400)
        except ValidationError as e:
            return JsonResponse({'ok': False, 'errors': getattr(e, 'message_dict', {'__all__': str(e)})}, status=400)
        except Exception as e:
            return JsonResponse({'ok': False, 'errors': {'__all__': str(e)}}, status=400)
        return JsonResponse({'ok': True, 'id': p.id, 'proveedor': _serialize_proveedor(p)})

    # DELETE
    has_snapshots = p.snapshots.exists()
    if has_snapshots:
        p.estatus = 'inactivo'
        p.save(update_fields=['estatus', 'updated_at'])
        return JsonResponse({'ok': True, 'mode': 'soft', 'id': id})
    else:
        p.delete()
        return JsonResponse({'ok': True, 'mode': 'hard', 'id': id})


@login_required
@require_http_methods(["POST"])
def proveedores_import(request):
    f = request.FILES.get('file') or request.FILES.get('csv')
    if not f:
        return JsonResponse({'ok': False, 'errors': {'file': 'Archivo requerido.'}}, status=400)

    try:
        raw = f.read().decode('utf-8-sig')
    except Exception:
        try:
            f.seek(0)
            raw = f.read().decode('latin-1')
        except Exception as e:
            return JsonResponse({'ok': False, 'errors': {'file': f'No se pudo leer el archivo: {e}'}}, status=400)

    reader = csv.DictReader(io.StringIO(raw))
    created = 0
    failed = 0
    errors_rows = []
    for idx, row in enumerate(reader, start=2):
        cleaned, errs = _validate_proveedor_payload(row, current=None)
        if errs:
            failed += 1
            errors_rows.append({'row': idx, 'razon_social': row.get('razon_social', ''), 'errors': errs})
            continue
        try:
            with transaction.atomic():
                p = Proveedor(
                    razon_social=cleaned.get('razon_social', ''),
                    rfc=cleaned.get('rfc'),
                    tipo_persona=cleaned.get('tipo_persona', ''),
                    dias_credito=cleaned.get('dias_credito', 0),
                    monto_credito=cleaned.get('monto_credito', Decimal('0')),
                    banco=cleaned.get('banco', ''),
                    cuenta_bancaria=cleaned.get('cuenta_bancaria', ''),
                    clabe=cleaned.get('clabe', ''),
                    cuenta_contable=cleaned.get('cuenta_contable', ''),
                    calle=cleaned.get('calle', ''),
                    numero=cleaned.get('numero', ''),
                    colonia=cleaned.get('colonia', ''),
                    ciudad=cleaned.get('ciudad', ''),
                    estado=cleaned.get('estado', ''),
                    cp=cleaned.get('cp', ''),
                    estatus=cleaned.get('estatus', 'activo'),
                    created_by=request.user if request.user.is_authenticated else None,
                )
                p.save()
            created += 1
        except IntegrityError:
            failed += 1
            errors_rows.append({'row': idx, 'razon_social': cleaned.get('razon_social', ''), 'errors': {'rfc': 'RFC duplicado.'}})
        except Exception as e:
            failed += 1
            errors_rows.append({'row': idx, 'razon_social': cleaned.get('razon_social', ''), 'errors': {'__all__': str(e)}})

    return JsonResponse({
        'ok': True,
        'created': created,
        'failed': failed,
        'errors': errors_rows,
    })


@login_required
@require_http_methods(["GET"])
def proveedores_export(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="proveedores.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'razon_social', 'rfc', 'tipo_persona', 'dias_credito', 'monto_credito',
        'banco', 'cuenta_bancaria', 'clabe', 'cuenta_contable',
        'calle', 'numero', 'colonia', 'ciudad', 'estado', 'cp', 'estatus',
    ])
    for p in Proveedor.objects.all().iterator():
        writer.writerow([
            p.razon_social, p.rfc or '', p.tipo_persona, p.dias_credito, str(p.monto_credito),
            p.banco, p.cuenta_bancaria, p.clabe, p.cuenta_contable,
            p.calle, p.numero, p.colonia, p.ciudad, p.estado, p.cp, p.estatus,
        ])
    return response


# ─────────────────────────────────────────────────────────────────
# Catálogos auxiliares
# ─────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(["GET"])
def claves_cfdi_search(request):
    q = (request.GET.get('q') or '').strip()
    qs = ClaveCFDI.objects.all()
    if q:
        qs = qs.filter(Q(clave__icontains=q) | Q(descripcion__icontains=q))
    tipo = (request.GET.get('tipo') or '').strip().lower()
    if tipo in ('producto', 'servicio'):
        qs = qs.filter(tipo=tipo)
    items = list(qs.order_by('clave')[:50])
    return JsonResponse({'items': [_serialize_clave_cfdi(c) for c in items]})


@login_required
@require_http_methods(["GET"])
def unidades_cfdi_list(request):
    items = list(UnidadCFDI.objects.all().order_by('clave'))
    return JsonResponse({'items': [_serialize_unidad_cfdi(u) for u in items]})


@login_required
@require_http_methods(["GET"])
def almacenes_list(request):
    items = list(Almacen.objects.filter(activo=True).order_by('nombre'))
    return JsonResponse({'items': [_serialize_almacen(a) for a in items]})
