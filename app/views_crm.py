# ----------------------------------------------------------------------
# views_crm.py — CRM home and oportunidades management.
# ----------------------------------------------------------------------

import json
import logging
import requests
import mimetypes
import os
from django.conf import settings
import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.db import models
from .models import TodoItem, Cliente, Cotizacion, DetalleCotizacion, UserProfile, Contacto, PendingFileUpload, OportunidadProyecto, Volumetria, DetalleVolumetria, CatalogoCableado, OportunidadActividad, OportunidadComentario, OportunidadArchivo, OportunidadEstado, Notificacion, Proyecto, ProyectoComentario, ProyectoArchivo, Tarea, TareaComentario, TareaArchivo, Actividad, CarpetaProyecto, ArchivoProyecto, CompartirArchivo, IntercambioNavidad, ParticipanteIntercambio, HistorialIntercambio, SolicitudAccesoProyecto, ArchivoFacturacion, ArchivoCobrado, AliasCliente, CarpetaOportunidad, ArchivoOportunidad, MensajeOportunidad, TareaOportunidad, ComentarioTareaOpp, PostMuro, ComentarioMuro, ProductoOportunidad, AsistenciaJornada, EficienciaMensual, SolicitudCambioPerfil, ProgramacionActividad, NovedadesConfig, EtapaPipeline
from . import views_exportar
from .views_tarea_comentarios import api_comentarios_tarea, api_agregar_comentario_tarea, api_editar_comentario_tarea, api_eliminar_comentario_tarea
from .forms import VentaForm, VentaFilterForm, CotizacionForm, ClienteForm, OportunidadModalForm, NuevaOportunidadForm
from django.db.models import Sum, Count, F, Q, Case, When, Value
from django.db.models.functions import Upper, Coalesce
from django.db.models import Value
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from django.utils import timezone
from decimal import Decimal
import decimal
from django.utils.html import json_script

# Helper function to detect lost opportunities
from .views_utils import *
from .views_grupos import get_usuarios_visibles_ids, get_clientes_visibles_q


def _get_etapas_pipeline_json():
    """Retorna JSON con etapas agrupadas por pipeline y lista de pipelines."""
    import json as _json
    from .models import EtapaPipeline
    etapas = EtapaPipeline.objects.filter(activo=True).order_by('pipeline', 'orden')
    result = {}
    for e in etapas:
        if e.pipeline not in result:
            result[e.pipeline] = []
        result[e.pipeline].append({'nombre': e.nombre, 'color': e.color})
    return _json.dumps(result, ensure_ascii=False)


@login_required
def get_oportunidades_por_cliente(request):
    cliente_id = request.GET.get('cliente_id')
    oportunidad_inicial_id = request.GET.get('oportunidad_inicial_id')  # Nueva línea para oportunidad específica
    
    print(f"DEBUG: get_oportunidades_por_cliente - cliente_id: {cliente_id}, oportunidad_inicial_id: {oportunidad_inicial_id}")

    if is_supervisor(request.user):
        if cliente_id:
            # Solo las 10 oportunidades más recientes del cliente para supervisores
            oportunidades = TodoItem.objects.filter(cliente_id=cliente_id).order_by('-fecha_creacion')[:10]
        else:
            # If no client_id, return the 20 most recent opportunities for supervisors
            oportunidades = TodoItem.objects.all().order_by('-fecha_creacion')[:20]
    else:
        ids_visibles = get_usuarios_visibles_ids(request.user)
        if ids_visibles and len(ids_visibles) > 1:
            user_filter = Q(usuario_id__in=ids_visibles)
        else:
            user_filter = Q(usuario=request.user)
        if cliente_id:
            oportunidades = TodoItem.objects.filter(Q(cliente_id=cliente_id) & user_filter).order_by('-fecha_creacion')[:10]
        else:
            oportunidades = TodoItem.objects.filter(user_filter).order_by('-fecha_creacion')[:20]

    # Si hay una oportunidad inicial específica, asegurar que esté incluida
    if oportunidad_inicial_id:
        try:
            # Limpiar el ID eliminando comas y espacios
            clean_id = oportunidad_inicial_id.replace(',', '').replace(' ', '').strip()
            print(f"DEBUG: Buscando oportunidad inicial con ID: {clean_id}")
            oportunidad_inicial = TodoItem.objects.get(id=int(clean_id))
            print(f"DEBUG: Oportunidad inicial encontrada: {oportunidad_inicial.oportunidad}")
            
            # Convertir queryset a lista para manipulación
            oportunidades_list = list(oportunidades)
            oportunidades_ids = [op.id for op in oportunidades_list]
            
            # Verificar si ya está en la lista por ID
            if oportunidad_inicial.id not in oportunidades_ids:
                print(f"DEBUG: Oportunidad inicial NO estaba en la lista, agregándola al principio")
                # Agregar la oportunidad específica al principio de la lista
                oportunidades_list.insert(0, oportunidad_inicial)
                oportunidades = oportunidades_list
            else:
                print(f"DEBUG: Oportunidad inicial YA estaba en la lista")
                # Moverla al principio si ya estaba presente
                oportunidades_list = [op for op in oportunidades_list if op.id != oportunidad_inicial.id]
                oportunidades_list.insert(0, oportunidad_inicial)
                oportunidades = oportunidades_list
        except (TodoItem.DoesNotExist, ValueError, TypeError) as e:
            print(f"DEBUG: Error procesando oportunidad inicial {oportunidad_inicial_id}: {e}")
            pass  # Si no existe o hay error de conversión, continuar con la lista normal

    data = [{'id': op.id, 'nombre': op.oportunidad} for op in oportunidades]

    return JsonResponse(data, safe=False)


@login_required
def get_bitrix_contacts_api(request):
    query = request.GET.get('query', '')
    company_id = request.GET.get('company_id', None)

    contacts = get_all_bitrix_contacts(request=request, company_id=company_id)

    # Filter contacts by query if provided
    if query:
        contacts = [c for c in contacts if query.lower() in (c.get('NAME', '') + ' ' + c.get('LAST_NAME', '')).lower()]

    data = []
    for contact in contacts:
        full_name = f"{contact.get('NAME', '')} {contact.get('LAST_NAME', '')}".strip()
        data.append({
            'ID': contact['ID'],
            'NAME': full_name,
            'COMPANY_ID': contact.get('COMPANY_ID'),
        })
    return JsonResponse(data, safe=False)


def _get_empleado_mes_data():
    """Obtiene datos del empleado del mes más reciente para el widget."""
    try:
        ganador = EficienciaMensual.objects.filter(
            empleado_del_mes=True
        ).order_by('-anio', '-mes').first()
        if not ganador:
            return None

        from datetime import date
        MESES_ES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                    7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

        user = ganador.usuario
        nombre = user.get_full_name() or user.username
        initials = ''.join(w[0].upper() for w in nombre.split() if w)[:2]
        profile = getattr(user, 'userprofile', None)
        avatar_url = profile.get_avatar_url() if profile and hasattr(profile, 'get_avatar_url') else ''

        return {
            'nombre': nombre,
            'username': f'@{user.username}',
            'initials': initials,
            'avatar_url': avatar_url,
            'eficiencia': int(ganador.promedio_eficiencia),
            'tareas': ganador.tareas_completadas,
            'cobradas': ganador.oportunidades_cobradas,
            'mes_nombre': f"{MESES_ES.get(int(ganador.mes), str(ganador.mes))} {ganador.anio}",
        }
    except Exception:
        return None


def _calcular_total_desglose(mes, anio):
    """Calcula el total facturado sumando datos_json — misma lógica que api_desglose_facturacion."""
    total = Decimal('0')
    try:
        if mes == 'todos':
            afs = list(ArchivoFacturacion.objects.filter(anio=anio))
        else:
            afs = [ArchivoFacturacion.objects.get(mes=mes, anio=anio)]
    except ArchivoFacturacion.DoesNotExist:
        return total
    for af in afs:
        raw = af.datos_json or {}
        for key, val in raw.items():
            if key == 'datos':
                continue
            try:
                if isinstance(val, dict) and 'monto' in val:
                    total += Decimal(str(val['monto']))
                else:
                    total += Decimal(str(val))
            except Exception:
                continue
    return total


@login_required
def crm_home(request):
    """
    Vista principal del CRM - tabla pivotada por cliente/producto.
    """
    from datetime import datetime
    user = request.user

    # Asegurar perfil
    profile, _ = UserProfile.objects.get_or_create(user=user)

    # Filtros mes/año — por defecto mes actual y año actual
    now = datetime.now()
    mes_filter = request.GET.get('mes', str(now.month).zfill(2))
    anio_filter = request.GET.get('anio', str(now.year))
    tab_activo = request.GET.get('tab', 'crm')

    # Mapeo de código mes a nombre en español (Bitrix guarda el nombre)
    MES_CODE_TO_NAME = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre',
    }
    MES_NAME_TO_CODE = {v: k for k, v in MES_CODE_TO_NAME.items()}

    try:
        anio_int = int(anio_filter)
    except ValueError:
        anio_int = now.year

    # MES_CHOICES para template (con opción "Todos" al inicio)
    mes_choices = [('todos', 'Todos')] + list(TodoItem.MES_CHOICES)
    mes_nombre = dict(mes_choices).get(mes_filter, '')
    mes_nombre_db = MES_CODE_TO_NAME.get(mes_filter, mes_filter)

    # ── Supervisor / Vendedor / Ingeniero logic ──
    es_supervisor = is_supervisor(user)
    es_ingeniero = (getattr(profile, 'rol', 'vendedor') == 'ingeniero')
    vendedores_filter = request.GET.get('vendedores', '')  # "1,2,3" or ""
    vendedores_ids = []
    if vendedores_filter:
        vendedores_ids = [int(x) for x in vendedores_filter.split(',') if x.strip().isdigit()]

    # Base queryset - oportunidades filtradas por anio_cierre/mes_cierre
    base_qs = TodoItem.objects.select_related('cliente', 'usuario', 'contacto', 'usuario__userprofile').filter(
        anio_cierre=anio_int
    )
    if mes_filter != 'todos':
        base_qs = base_qs.filter(mes_cierre=mes_filter)

    # Filtrado por visibilidad: supervisor global ve todo, grupos de trabajo amplían visibilidad
    if not es_supervisor:
        usuarios_visibles = get_usuarios_visibles_ids(user)
        if usuarios_visibles is None:
            pass  # ve todo (no debería llegar aquí)
        elif len(usuarios_visibles) == 1:
            base_qs = base_qs.filter(usuario=user)
        else:
            base_qs = base_qs.filter(usuario_id__in=usuarios_visibles)
    elif vendedores_ids:
        # Supervisor con filtro de vendedores específicos
        base_qs = base_qs.filter(usuario_id__in=vendedores_ids)

    # Lista de vendedores para el filtro (supervisores globales ven todos; supervisores de grupo ven su grupo)
    vendedores_list = []
    if es_supervisor:
        vendedores_list = User.objects.filter(
            is_active=True
        ).exclude(
            groups__name='Supervisores'
        ).order_by('first_name', 'last_name')
    else:
        # Usuarios de grupo visibles para el selector de vendedores
        usuarios_visibles_ids = get_usuarios_visibles_ids(user)
        if usuarios_visibles_ids and len(usuarios_visibles_ids) > 1:
            vendedores_list = User.objects.filter(
                id__in=usuarios_visibles_ids, is_active=True
            ).order_by('first_name', 'last_name')

    # ── Meta (calculada antes para usar en running meta) ──
    # Determinar qué campo de meta usar según el tab activo
    meta_field = 'meta_mensual'  # Default (Facturado)
    if tab_activo == 'crm':
        meta_field = 'meta_oportunidades'
    elif tab_activo == 'cotizado':
        meta_field = 'meta_cotizado'
    elif tab_activo == 'cobrado':
        meta_field = 'meta_cobrado'

    if es_supervisor:
        if vendedores_ids:
            meta = UserProfile.objects.filter(user_id__in=vendedores_ids).aggregate(
                t=Coalesce(Sum(meta_field), Value(Decimal('0')))
            )['t'] or Decimal('0')
        else:
            # Suma de todos los vendedores activos para el supervisor
            all_sellers_profiles = UserProfile.objects.filter(user__is_active=True).exclude(user__groups__name='Supervisores')
            meta = all_sellers_profiles.aggregate(t=Coalesce(Sum(meta_field), Value(Decimal('0'))))['t'] or Decimal('0')
    else:
        meta = getattr(profile, meta_field, Decimal('0')) or Decimal('0')

    # Si se seleccionó "Todos" los meses, la meta es anual (mensual × 12)
    if mes_filter == 'todos':
        meta = meta * 12

    # ── Tab CRM: Lista de oportunidades individuales ──
    if tab_activo == 'crm':
        tabla_data = base_qs.select_related('cliente', 'contacto', 'usuario').order_by('-fecha_actualizacion')

    # ── Tab Facturado: Datos del XLS por cliente + desglose por producto ──
    elif tab_activo == 'facturado':
        # Obtener datos de facturación del XLS subido
        facturado_por_cliente = {}  # {cliente_name: monto}
        if mes_filter == 'todos':
            # Sumar todos los meses del año
            for af in ArchivoFacturacion.objects.filter(anio=anio_int):
                for cname, monto_str in (af.datos_json or {}).items():
                    facturado_por_cliente[cname] = str(
                        Decimal(facturado_por_cliente.get(cname, '0')) + Decimal(str(monto_str))
                    )
        else:
            try:
                archivo_fact = ArchivoFacturacion.objects.get(mes=mes_filter, anio=anio_int)
                facturado_por_cliente = archivo_fact.datos_json or {}
            except ArchivoFacturacion.DoesNotExist:
                pass

        # Mapear nombres del XLS a objetos Cliente (matching flexible)
        facturado_por_cliente_obj = {}  # {cliente_id: monto}
        _all_clientes = list(Cliente.objects.all())
        for cliente_name, monto_str in facturado_por_cliente.items():
            monto_val = Decimal(str(monto_str))
            cn_upper = cliente_name.upper().strip()
            cliente_match = None
            # 1) Exact match
            for c in _all_clientes:
                if c.nombre_empresa and c.nombre_empresa.upper().strip() == cn_upper:
                    cliente_match = c
                    break
            # 2) XLS name contiene nombre CRM o viceversa
            if not cliente_match:
                for c in _all_clientes:
                    if not c.nombre_empresa:
                        continue
                    crm_upper = c.nombre_empresa.upper().strip()
                    if crm_upper in cn_upper or cn_upper in crm_upper:
                        cliente_match = c
                        break
            # 3) Match por primeras 2 palabras significativas
            if not cliente_match:
                palabras_xls = [w for w in cn_upper.split() if len(w) > 2 and w not in ('DE', 'DEL', 'LA', 'LAS', 'LOS', 'EL', 'SA', 'CV', 'SAS', 'INC', 'MEXICO')]
                if len(palabras_xls) >= 2:
                    for c in _all_clientes:
                        if not c.nombre_empresa:
                            continue
                        crm_upper = c.nombre_empresa.upper()
                        if palabras_xls[0] in crm_upper and palabras_xls[1] in crm_upper:
                            cliente_match = c
                            break
                elif len(palabras_xls) == 1 and len(palabras_xls[0]) >= 4:
                    for c in _all_clientes:
                        if not c.nombre_empresa:
                            continue
                        if palabras_xls[0] in c.nombre_empresa.upper():
                            cliente_match = c
                            break
            if cliente_match:
                facturado_por_cliente_obj[cliente_match.id] = (
                    facturado_por_cliente_obj.get(cliente_match.id, Decimal('0')) + monto_val
                )

        # Desglose por producto (de monto_facturacion en oportunidades)
        prod_data = base_qs.filter(monto_facturacion__gt=0).values('cliente').annotate(
            zebra=Coalesce(Sum('monto_facturacion', filter=Q(producto='ZEBRA')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            panduit=Coalesce(Sum('monto_facturacion', filter=Q(producto='PANDUIT')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            apc=Coalesce(Sum('monto_facturacion', filter=Q(producto='APC')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            avigilon=Coalesce(Sum('monto_facturacion', filter=Q(producto='AVIGILON') | Q(producto='AVIGILION')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            genetec=Coalesce(Sum('monto_facturacion', filter=Q(producto='GENETEC')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            axis=Coalesce(Sum('monto_facturacion', filter=Q(producto='AXIS')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            software=Coalesce(Sum('monto_facturacion', filter=Q(producto='SOFTWARE') | Q(producto='Desarrollo')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            runrate=Coalesce(Sum('monto_facturacion', filter=Q(producto='RUNRATE')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            poliza=Coalesce(Sum('monto_facturacion', filter=Q(producto='PÓLIZA') | Q(producto='POLIZA')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            total_prod=Coalesce(Sum('monto_facturacion'), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
        )
        prod_dict = {item['cliente']: item for item in prod_data}

        # Filtrar clientes según permisos
        if es_supervisor:
            if vendedores_ids:
                clientes_qs = Cliente.objects.filter(asignado_a_id__in=vendedores_ids).order_by('nombre_empresa')
            else:
                clientes_qs = Cliente.objects.all().order_by('nombre_empresa')
        else:
            clientes_qs = Cliente.objects.filter(get_clientes_visibles_q(user)).order_by('nombre_empresa')

        raw_data = []
        for c in clientes_qs:
            pdat = prod_dict.get(c.id, {})
            fact_monto = facturado_por_cliente_obj.get(c.id, Decimal('0'))
            zb = pdat.get('zebra', Decimal('0'))
            pa = pdat.get('panduit', Decimal('0'))
            ap = pdat.get('apc', Decimal('0'))
            av = pdat.get('avigilon', Decimal('0'))
            ge = pdat.get('genetec', Decimal('0'))
            ax = pdat.get('axis', Decimal('0'))
            so = pdat.get('software', Decimal('0'))
            rr = pdat.get('runrate', Decimal('0'))
            po = pdat.get('poliza', Decimal('0'))
            tp = pdat.get('total_prod', Decimal('0'))
            otros = tp - (zb + pa + ap + av + ge + ax + so + rr + po)
            raw_data.append({
                'cliente': c,
                'zebra': zb, 'panduit': pa, 'apc': ap, 'avigilon': av,
                'genetec': ge, 'axis': ax, 'software': so, 'runrate': rr,
                'poliza': po, 'otros': otros,
                'facturado': fact_monto,
            })

        # Ordenar por facturado descendente
        raw_data.sort(key=lambda x: x['facturado'], reverse=True)

        # Meta por cliente: meta individual del cliente - su facturado
        for item in raw_data:
            cliente_meta = item['cliente'].meta_mensual or Decimal('0')
            item['meta_cliente'] = cliente_meta
            item['meta_restante'] = cliente_meta - item['facturado']
            item['total'] = item['facturado'] # Para que el template use item.total
        tabla_data = raw_data

    # ── Tab Cotizado: Cotizaciones PDF generadas ──
    elif tab_activo == 'cotizado':
        opp_ids = base_qs.values_list('id', flat=True)
        cotizaciones_qs = (
            Cotizacion.objects
            .select_related('oportunidad', 'created_by', 'cliente')
            .filter(
                Q(oportunidad_id__in=opp_ids) |
                Q(oportunidad__isnull=True, fecha_creacion__year=anio_int)
            )
            .order_by('-fecha_creacion')
        )
        tabla_data = cotizaciones_qs

    # ── Tab Cobrado: Oportunidades con etapa Ganado/Pagado ──
    elif tab_activo == 'cobrado':
        tabla_data = base_qs.filter(etapa_corta__in=['Ganado', 'Pagado']).order_by('-monto')

    else:
        tabla_data = []

    # Stats generales
    total_general = base_qs.aggregate(t=Coalesce(Sum('monto'), Value(Decimal('0'))))['t']
    num_clientes = Cliente.objects.count() if es_supervisor else base_qs.values('cliente').distinct().count()
    num_deals = base_qs.count()
    num_cobradas = base_qs.filter(etapa_corta__in=['Ganado', 'Pagado']).count()

    # ── Total facturado: usar misma lógica que api_desglose_facturacion ──
    total_facturado = _calcular_total_desglose(mes_filter, anio_int)

    progreso = min(int((total_facturado / meta * 100)) if meta > 0 else 0, 100)

    # Stats para tab Cobrado
    total_cobrado = Decimal('0')
    if tab_activo == 'cobrado':
        total_cobrado = base_qs.filter(etapa_corta__in=['Ganado', 'Pagado']).aggregate(t=Coalesce(Sum('monto'), Value(Decimal('0'))))['t']

    # Stats para tab Cotizado
    num_cotizaciones = 0
    num_oportunidades_cotizadas = 0
    total_cotizado = Decimal('0')
    if tab_activo == 'cotizado':
        num_cotizaciones = cotizaciones_qs.count()
        num_oportunidades_cotizadas = cotizaciones_qs.exclude(oportunidad__isnull=True).values('oportunidad').distinct().count()
        total_cotizado = cotizaciones_qs.aggregate(t=Coalesce(Sum('total'), Value(Decimal('0'))))['t']

    # ── Widget Logic ──
    widget_label = 'Total Facturado'
    widget_metric = total_facturado 

    if tab_activo == 'crm':
        widget_label = 'Total Oportunidades'
        widget_metric = total_general
    elif tab_activo == 'clientes':
        widget_label = 'Total Facturado'
        widget_metric = total_facturado

    # Recalculate progress based on the correct metric vs correct meta (sin cap para mostrar > 100%)
    progreso = int((widget_metric / meta * 100)) if meta > 0 else 0

    context = {
        'widget_label': widget_label,
        'widget_metric': widget_metric,
        'tab_activo': tab_activo,
        'tabla_data': tabla_data,
        'mes_filter': mes_filter,
        'anio_filter': anio_filter,
        'anio_int': anio_int,
        'mes_nombre': mes_nombre,
        'mes_choices': mes_choices,
        'total_general': total_general,
        'total_facturado': total_facturado,
        'num_clientes': num_clientes,
        'num_deals': num_deals,
        'num_cobradas': num_cobradas,
        'meta': meta,
        'progreso': progreso,
        'progreso_visual': min(progreso, 100),
        'usuario': user,
        'years_range': range(2024, now.year + 2),
        'num_cotizaciones': num_cotizaciones,
        'num_oportunidades_cotizadas': num_oportunidades_cotizadas,
        'total_cotizado': total_cotizado,
        'total_cobrado': total_cobrado,
        'es_supervisor': es_supervisor,
        'es_ingeniero': es_ingeniero,
        'vendedores_list': vendedores_list,
        'vendedores_filter': vendedores_filter,
        'novedades_config': NovedadesConfig.get(),
        'empleado_mes_data': _get_empleado_mes_data(),
        'mis_grupos': _get_mis_grupos_ctx(user),
        'es_supervisor_de_grupo': _es_supervisor_de_grupo(user),
        'etapas_pipeline_json': _get_etapas_pipeline_json(),
        'pipelines_list': list(EtapaPipeline.objects.values_list('pipeline', flat=True).distinct().order_by('pipeline')) if EtapaPipeline.objects.exists() else ['runrate', 'proyecto'],
    }
    return render(request, 'crm_home.html', context)


def _get_mis_grupos_ctx(user):
    """Grupos activos del usuario para el contexto del template."""
    from .views_grupos import get_grupos_del_usuario
    from .views_utils import is_supervisor as _is_sup
    if _is_sup(user):
        from .models import GrupoTrabajo
        return list(GrupoTrabajo.objects.filter(activo=True).values('id', 'nombre', 'color'))
    grupos = get_grupos_del_usuario(user)
    return [{'id': g.id, 'nombre': g.nombre, 'color': g.color} for g in grupos]


def _es_supervisor_de_grupo(user):
    from .models import GrupoTrabajo
    return GrupoTrabajo.objects.filter(supervisor_grupo=user, activo=True).exists()


@login_required
def api_crm_table_data(request):
    """
    API endpoint que devuelve los datos de la tabla CRM en JSON
    para actualizar sin recargar la página.
    """
    from datetime import datetime
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)

    now = datetime.now()
    mes_filter = request.GET.get('mes', str(now.month).zfill(2))
    anio_filter = request.GET.get('anio', str(now.year))
    tab_activo = request.GET.get('tab', 'crm')
    desde_filter = request.GET.get('desde', '').strip()
    hasta_filter = request.GET.get('hasta', '').strip()
    usando_periodo = bool(desde_filter and hasta_filter)

    MES_CODE_TO_NAME = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre',
    }

    try:
        anio_int = int(anio_filter)
    except ValueError:
        anio_int = now.year

    mes_nombre_db = MES_CODE_TO_NAME.get(mes_filter, mes_filter)

    # Supervisor / Vendedor logic
    es_supervisor = is_supervisor(user)
    vendedores_filter = request.GET.get('vendedores', '')
    vendedores_ids = [int(x) for x in vendedores_filter.split(',') if x.strip().isdigit()] if vendedores_filter else []

    q_search = request.GET.get('q', '').strip()

    if q_search:
        # Búsqueda global: ignora mes/año, busca en nombre de oportunidad y cliente
        base_qs = TodoItem.objects.select_related('cliente', 'usuario', 'contacto', 'usuario__userprofile').filter(
            Q(oportunidad__icontains=q_search) | Q(cliente__nombre_empresa__icontains=q_search)
        )
    elif usando_periodo:
        base_qs = TodoItem.objects.select_related('cliente', 'usuario', 'contacto', 'usuario__userprofile').filter(
            fecha_creacion__date__gte=desde_filter,
            fecha_creacion__date__lte=hasta_filter,
        )
    else:
        base_qs = TodoItem.objects.select_related('cliente', 'usuario', 'contacto', 'usuario__userprofile').filter(
            anio_cierre=anio_int
        )
        if mes_filter != 'todos':
            base_qs = base_qs.filter(mes_cierre=mes_filter)

    if not es_supervisor:
        usuarios_visibles_r = get_usuarios_visibles_ids(user)
        if usuarios_visibles_r and len(usuarios_visibles_r) > 1:
            base_qs = base_qs.filter(usuario_id__in=usuarios_visibles_r)
        else:
            base_qs = base_qs.filter(usuario=user)
    elif vendedores_ids:
        base_qs = base_qs.filter(usuario_id__in=vendedores_ids)

    def format_money(val):
        if val is None:
            return '0'
        try:
            return '{:,.0f}'.format(val)
        except (ValueError, TypeError):
            return '0'

    # ── Calcular meta según tab (aplica a todos los tabs) ──
    meta_field_api = 'meta_mensual'
    if tab_activo == 'crm':
        meta_field_api = 'meta_oportunidades'
    elif tab_activo == 'cotizado':
        meta_field_api = 'meta_cotizado'
    elif tab_activo == 'cobrado':
        meta_field_api = 'meta_cobrado'
    elif tab_activo == 'clientes':
        _vista_cl = request.GET.get('vista', 'facturado')
        if _vista_cl == 'cobrado':
            meta_field_api = 'meta_cobrado'
        elif _vista_cl == 'oportunidades':
            meta_field_api = 'meta_oportunidades'
        elif _vista_cl == 'cotizado':
            meta_field_api = 'meta_cotizado'
        # else: facturado → meta_mensual (default)

    if es_supervisor:
        if vendedores_ids:
            api_meta = UserProfile.objects.filter(user_id__in=vendedores_ids).aggregate(
                t=Coalesce(Sum(meta_field_api), Value(Decimal('0')))
            )['t'] or Decimal('0')
        else:
            all_sellers_profiles = UserProfile.objects.filter(user__is_active=True).exclude(user__groups__name='Supervisores')
            api_meta = all_sellers_profiles.aggregate(t=Coalesce(Sum(meta_field_api), Value(Decimal('0'))))['t'] or Decimal('0')
    else:
        api_meta = getattr(profile, meta_field_api, Decimal('0')) or Decimal('0')

    if mes_filter == 'todos' and not usando_periodo:
        api_meta = api_meta * 12

    if tab_activo == 'crm':
        from django.utils import timezone as _tz
        from django.db.models import OuterRef, Exists as _Exists
        _now = _tz.now()
        _now_naive = _now.replace(tzinfo=None)
        _ESTADOS_ACTIVOS = ('pendiente', 'iniciada', 'en_progreso')

        # EXISTS subqueries — MySQL evalúa con short-circuit, 0 queries adicionales por fila
        _vencida_opp_sq = TareaOportunidad.objects.filter(
            oportunidad=OuterRef('pk'),
            estado__in=['pendiente', 'en_progreso'],
            fecha_limite__isnull=False,
            fecha_limite__lte=_now,
        )
        _vencida_tarea_sq = Tarea.objects.filter(
            oportunidad=OuterRef('pk'),
            estado__in=_ESTADOS_ACTIVOS,
            fecha_limite__isnull=False,
            fecha_limite__lte=_now,
        )
        _pendiente_opp_sq = TareaOportunidad.objects.filter(
            oportunidad=OuterRef('pk'),
        ).exclude(estado='completada')

        items = base_qs.select_related('cliente', 'contacto', 'usuario').annotate(
            _tiene_vencida=_Exists(_vencida_opp_sq) | _Exists(_vencida_tarea_sq),
            _tiene_pendiente=_Exists(_pendiente_opp_sq),
        ).order_by('-fecha_actualizacion')

        rows = []
        for item in items:
            tiene_vencida = item._tiene_vencida
            tiene_pendiente = item._tiene_pendiente
            es_bitrix = item.tipo_negociacion == 'bitrix_proyecto'
            rows.append({
                'id': item.id,
                'oportunidad': (item.oportunidad or '')[:35],
                'cliente': (item.cliente.nombre_empresa if item.cliente else '- Sin Cliente -')[:35],
                'cliente_id': item.cliente.id if item.cliente else None,
                'contacto': (item.contacto.nombre[:18] if item.contacto else '-'),
                'area': item.area or '-',
                'producto': item.producto or '',
                'monto': '0' if es_bitrix else format_money(item.monto),
                'fecha_iso': item.fecha_creacion.strftime('%Y-%m-%d'),
                'fecha_ts': int(item.fecha_actualizacion.timestamp()) if item.fecha_actualizacion else 0,
                'etapa': item.etapa_corta or '',
                'tiene_actividad_vencida': tiene_vencida,
                'sin_actividad_pendiente': not tiene_pendiente,
                'tipo_negociacion': item.tipo_negociacion or 'runrate',
            })
        # Ordenar: vencidas primero, luego por fecha_actualizacion más reciente
        rows.sort(key=lambda x: (not x['tiene_actividad_vencida'], -x.get('fecha_ts', 0)))
        # Stats
        total_general = base_qs.aggregate(t=Coalesce(Sum('monto'), Value(Decimal('0'))))['t']
        num_clientes = base_qs.values('cliente').distinct().count()
        num_deals = base_qs.count()

        api_progreso = int((total_general / api_meta * 100)) if api_meta > 0 else 0

        return JsonResponse({
            'tab': 'crm',
            'rows': rows,
            'footer': {
                'left': f'{num_clientes} clientes / {num_deals} Deals',
                'right': f'Total: ${format_money(total_general)}',
            },
            'total_facturado': format_money(total_general),
            'widget_label': 'Total Oportunidades',
            'vista_label': 'Oportunidades',
            'meta': format_money(api_meta),
            'progreso': api_progreso,
            'widget_left_stat': f'{num_deals} Oportunidades Creadas',
        })

    elif tab_activo == 'cotizado':
        opp_ids = base_qs.values_list('id', flat=True)
        if usando_periodo:
            cotizaciones_qs = (
                Cotizacion.objects
                .select_related('oportunidad', 'created_by', 'cliente')
                .filter(
                    Q(oportunidad_id__in=opp_ids) |
                    Q(oportunidad__isnull=True,
                      fecha_creacion__date__gte=desde_filter,
                      fecha_creacion__date__lte=hasta_filter)
                )
                .order_by('-fecha_creacion')
            )
        else:
            cotizaciones_qs = (
                Cotizacion.objects
                .select_related('oportunidad', 'created_by', 'cliente')
                .filter(
                    Q(oportunidad_id__in=opp_ids) |
                    Q(oportunidad__isnull=True, fecha_creacion__year=anio_int)
                )
                .order_by('-fecha_creacion')
            )
        rows = []
        for cot in cotizaciones_qs:
            rows.append({
                'id': cot.id,
                'oportunidad': (cot.oportunidad.oportunidad if cot.oportunidad else '—')[:35],
                'oportunidad_id': cot.oportunidad.id if cot.oportunidad else None,
                'cliente': (cot.cliente.nombre_empresa if cot.cliente else '- Sin Cliente -')[:35],
                'cliente_id': cot.cliente.id if cot.cliente else None,
                'usuario': (cot.created_by.get_full_name() or cot.created_by.username) if cot.created_by else '—',
                'subtotal': format_money(cot.subtotal),
                'total': format_money(cot.total),
                'pdf_url': f'/app/cotizacion/view/{cot.id}/',
                'fecha_iso': cot.fecha_creacion.strftime('%Y-%m-%d'),
            })
        num_cotizaciones = cotizaciones_qs.count()
        num_oportunidades_cotizadas = cotizaciones_qs.exclude(oportunidad__isnull=True).values('oportunidad').distinct().count()
        total_cotizado = cotizaciones_qs.aggregate(t=Coalesce(Sum('total'), Value(Decimal('0'))))['t']
        api_progreso_cot = int((total_cotizado / api_meta * 100)) if api_meta > 0 else 0
        return JsonResponse({
            'tab': 'cotizado',
            'rows': rows,
            'footer': {
                'left': f'{num_oportunidades_cotizadas} oportunidades / {num_cotizaciones} cotizaciones',
                'right': f'Total cotizado: ${format_money(total_cotizado)}',
            },
            'total_facturado': format_money(total_cotizado),
            'widget_label': 'Total Cotizado',
            'meta': format_money(api_meta),
            'progreso': api_progreso_cot,
            'widget_left_stat': f'{num_cotizaciones} Cotizaciones Creadas',
            'num_total_cotizaciones': num_cotizaciones,
        })

    elif tab_activo == 'cobrado':
        items = base_qs.filter(etapa_corta__in=['Ganado', 'Pagado']).order_by('-monto')
        rows = []
        for op in items:
            rows.append({
                'id': op.id,
                'oportunidad': (op.oportunidad or '')[:35],
                'cliente': (op.cliente.nombre_empresa if op.cliente else '- Sin Cliente -')[:35],
                'cliente_id': op.cliente.id if op.cliente else None,
                'producto_display': op.get_producto_display(),
                'usuario': (op.usuario.get_full_name() or op.usuario.username) if op.usuario else '—',
                'fecha': op.fecha_creacion.strftime('%d %b %Y'),
                'fecha_iso': op.fecha_creacion.strftime('%Y-%m-%d'),
                'monto': format_money(op.monto),
            })
        total_cobrado = items.aggregate(t=Coalesce(Sum('monto'), Value(Decimal('0'))))['t']
        num_deals = items.count()
        api_progreso_cob = int((total_cobrado / api_meta * 100)) if api_meta > 0 else 0
        return JsonResponse({
            'tab': 'cobrado',
            'rows': rows,
            'footer': {
                'left': f'{num_deals} Deals Cobrados',
                'right': f'Total cobrado: ${format_money(total_cobrado)}',
            },
            'total_facturado': format_money(total_cobrado),
            'widget_label': 'Total Cobrado',
            'meta': format_money(api_meta),
            'progreso': api_progreso_cob,
            'widget_left_stat': f'{num_deals} Oportunidades Cobradas',
        })

    elif tab_activo == 'facturado':
        # Reutilizamos la lógica del view principal para facturación
        facturado_por_cliente_obj = {}
        try:
            if usando_periodo:
                from datetime import date as _date
                cur = _date.fromisoformat(desde_filter).replace(day=1)
                end = _date.fromisoformat(hasta_filter).replace(day=1)
                while cur <= end:
                    mes_code = cur.strftime('%m')
                    try:
                        af = ArchivoFacturacion.objects.get(mes=mes_code, anio=cur.year)
                        data_af = af.datos_json.get('datos', {})
                        for c_name, val in data_af.items():
                            facturado_por_cliente_obj[c_name] = facturado_por_cliente_obj.get(c_name, Decimal('0')) + Decimal(str(val))
                    except ArchivoFacturacion.DoesNotExist:
                        pass
                    cur = (cur.replace(month=cur.month % 12 + 1, day=1) if cur.month < 12
                           else cur.replace(year=cur.year + 1, month=1, day=1))
            elif mes_filter == 'todos':
                for af in ArchivoFacturacion.objects.filter(anio=anio_int):
                    data_af = af.datos_json.get('datos', {})
                    for c_name, val in data_af.items():
                        facturado_por_cliente_obj[c_name] = facturado_por_cliente_obj.get(c_name, Decimal('0')) + Decimal(str(val))
            else:
                af = ArchivoFacturacion.objects.get(mes=mes_filter, anio=anio_int)
                data_af = af.datos_json.get('datos', {})
                for c_name, val in data_af.items():
                    facturado_por_cliente_obj[c_name] = Decimal(str(val))
        except ArchivoFacturacion.DoesNotExist:
            pass

        # Mapear nombres a IDs de clientes
        fact_by_id = {}
        for name, monto in facturado_por_cliente_obj.items():
            cliente = Cliente.objects.filter(nombre_empresa__icontains=name).first()
            if cliente:
                fact_by_id[cliente.id] = fact_by_id.get(cliente.id, Decimal('0')) + monto

        # Desglose prod
        prod_data = base_qs.filter(monto_facturacion__gt=0).values('cliente').annotate(
            zebra=Coalesce(Sum('monto_facturacion', filter=Q(producto='ZEBRA')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            panduit=Coalesce(Sum('monto_facturacion', filter=Q(producto='PANDUIT')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            apc=Coalesce(Sum('monto_facturacion', filter=Q(producto='APC')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            avigilon=Coalesce(Sum('monto_facturacion', filter=Q(producto='AVIGILON') | Q(producto='AVIGILION')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            genetec=Coalesce(Sum('monto_facturacion', filter=Q(producto='GENETEC')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            axis=Coalesce(Sum('monto_facturacion', filter=Q(producto='AXIS')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            software=Coalesce(Sum('monto_facturacion', filter=Q(producto='SOFTWARE') | Q(producto='Desarrollo')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            runrate=Coalesce(Sum('monto_facturacion', filter=Q(producto='RUNRATE')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            poliza=Coalesce(Sum('monto_facturacion', filter=Q(producto='PÓLIZA') | Q(producto='POLIZA')), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
            total_prod=Coalesce(Sum('monto_facturacion'), Value(0, output_field=models.DecimalField(max_digits=12, decimal_places=2))),
        )
        prod_dict = {item['cliente']: item for item in prod_data}

        if es_supervisor:
            clientes_qs = Cliente.objects.filter(asignado_a_id__in=vendedores_ids) if vendedores_ids else Cliente.objects.all()
        else:
            clientes_qs = Cliente.objects.filter(get_clientes_visibles_q(user))

        rows = []
        total_facturado_acum = _calcular_total_desglose(mes_filter, anio_int)

        for c in clientes_qs.order_by('nombre_empresa'):
            p = prod_dict.get(c.id, {})
            fact = fact_by_id.get(c.id, Decimal('0'))
            meta_c = c.meta_mensual or Decimal('0')
            if mes_filter == 'todos' and not usando_periodo:
                meta_c = meta_c * 12
            rows.append({
                'cliente_id': c.id,
                'cliente': c.nombre_empresa[:35],
                'zebra': format_money(p.get('zebra')),
                'panduit': format_money(p.get('panduit')),
                'apc': format_money(p.get('apc')),
                'avigilon': format_money(p.get('avigilon')),
                'genetec': format_money(p.get('genetec')),
                'axis': format_money(p.get('axis')),
                'software': format_money(p.get('software')),
                'runrate': format_money(p.get('runrate')),
                'poliza': format_money(p.get('poliza')),
                'otros': format_money(p.get('total_prod', Decimal('0')) - sum(p.get(k, 0) for k in ['zebra','panduit','apc','avigilon','genetec','axis','software','runrate','poliza'] if k in p)),
                'total': format_money(fact),
                'meta_cliente': format_money(meta_c),
                'meta_restante': format_money(meta_c - fact),
            })

        api_progreso_fact = int((total_facturado_acum / api_meta * 100)) if api_meta > 0 else 0
        num_clientes_fact = clientes_qs.count()
        return JsonResponse({
            'tab': 'facturado',
            'rows': rows,
            'footer': {
                'left': f'{num_clientes_fact} clientes',
                'right': f'Total facturado: ${format_money(total_facturado_acum)}',
            },
            'total_facturado': format_money(total_facturado_acum),
            'widget_label': 'Total Facturado',
            'meta': format_money(api_meta),
            'progreso': api_progreso_fact,
            'widget_left_stat': f'{num_clientes_fact} Clientes',
        })

    elif tab_activo == 'clientes':
        vista = request.GET.get('vista', 'facturado')
        _dec_field = models.DecimalField(max_digits=14, decimal_places=2)
        _zero = Value(Decimal('0'), output_field=_dec_field)

        if es_supervisor:
            clientes_qs = Cliente.objects.select_related('asignado_a').filter(asignado_a_id__in=vendedores_ids) if vendedores_ids else Cliente.objects.select_related('asignado_a').all()
        else:
            clientes_qs = Cliente.objects.select_related('asignado_a').filter(asignado_a=user)

        # Para oportunidades y cotizado en Clientes: filtrar por fecha_creacion
        # en vez de mes_cierre, así no muestra datos de meses futuros
        if vista in ('oportunidades', 'cotizado') and not q_search:
            if usando_periodo:
                base_qs = TodoItem.objects.select_related('cliente', 'usuario', 'contacto', 'usuario__userprofile').filter(
                    fecha_creacion__date__gte=desde_filter,
                    fecha_creacion__date__lte=hasta_filter,
                )
            else:
                base_qs = TodoItem.objects.select_related('cliente', 'usuario', 'contacto', 'usuario__userprofile').filter(
                    fecha_creacion__year=anio_int
                )
                if mes_filter != 'todos':
                    base_qs = base_qs.filter(fecha_creacion__month=int(mes_filter))
            # Re-aplicar filtros de usuario/vendedor
            if not es_supervisor:
                usuarios_visibles_r = get_usuarios_visibles_ids(user)
                if usuarios_visibles_r and len(usuarios_visibles_r) > 1:
                    base_qs = base_qs.filter(usuario_id__in=usuarios_visibles_r)
                else:
                    base_qs = base_qs.filter(usuario=user)
            elif vendedores_ids:
                base_qs = base_qs.filter(usuario_id__in=vendedores_ids)

        def _prod_annotate(qs, monto_field):
            return qs.values('cliente').annotate(
                zebra=Coalesce(Sum(monto_field, filter=Q(producto='ZEBRA')), _zero),
                panduit=Coalesce(Sum(monto_field, filter=Q(producto='PANDUIT')), _zero),
                apc=Coalesce(Sum(monto_field, filter=Q(producto='APC')), _zero),
                avigilon=Coalesce(Sum(monto_field, filter=Q(producto='AVIGILON') | Q(producto='AVIGILION')), _zero),
                genetec=Coalesce(Sum(monto_field, filter=Q(producto='GENETEC')), _zero),
                axis=Coalesce(Sum(monto_field, filter=Q(producto='AXIS')), _zero),
                software=Coalesce(Sum(monto_field, filter=Q(producto='SOFTWARE') | Q(producto='Desarrollo')), _zero),
                runrate=Coalesce(Sum(monto_field, filter=Q(producto='RUNRATE')), _zero),
                poliza=Coalesce(Sum(monto_field, filter=Q(producto='PÓLIZA') | Q(producto='POLIZA')), _zero),
                total_prod=Coalesce(Sum(monto_field), _zero),
            )

        _prev_by_id_cl = {}  # previous month totals per client (oportunidades only)
        _prev_sum = Decimal('0')  # previous month global total for this vista

        # Helper: compute prev month params
        def _get_prev_params():
            if mes_filter in ('todos',) or usando_periodo or q_search:
                return None, None
            try:
                pm = int(mes_filter)
                return str(pm - 1 if pm > 1 else 12).zfill(2), anio_int if pm > 1 else anio_int - 1
            except (ValueError, TypeError):
                return None, None

        _total_facturado_excel = Decimal('0')
        _total_cobrado_csv = Decimal('0')

        if vista == 'facturado':
            # Total desde ArchivoFacturacion
            facturado_por_cliente_obj = {}

            def _extract_entries(datos_json):
                """Extrae lista de {nombre, rfc, monto} del datos_json (soporta 3 formatos)"""
                if not datos_json:
                    return []
                entries = []
                for key, val in datos_json.items():
                    if key == 'datos':
                        continue
                    if isinstance(val, dict) and 'monto' in val:
                        # Formato nuevo: {rfc_or_name: {nombre, rfc, monto}}
                        entries.append({
                            'nombre': val.get('nombre', key),
                            'rfc': val.get('rfc', ''),
                            'monto': Decimal(str(val['monto'])),
                        })
                    else:
                        # Formato viejo: {nombre: monto_str}
                        try:
                            entries.append({'nombre': key, 'rfc': '', 'monto': Decimal(str(val))})
                        except Exception:
                            pass
                return entries

            # Acumular facturado por RFC o nombre
            _facturado_entries = []  # [{nombre, rfc, monto}]
            try:
                def _load_af(af):
                    return _extract_entries(af.datos_json)
                if usando_periodo:
                    from datetime import date as _date
                    cur = _date.fromisoformat(desde_filter).replace(day=1)
                    end = _date.fromisoformat(hasta_filter).replace(day=1)
                    while cur <= end:
                        try:
                            af = ArchivoFacturacion.objects.get(mes=cur.strftime('%m'), anio=cur.year)
                            _facturado_entries.extend(_load_af(af))
                        except ArchivoFacturacion.DoesNotExist:
                            pass
                        cur = cur.replace(month=cur.month % 12 + 1, day=1) if cur.month < 12 else cur.replace(year=cur.year + 1, month=1, day=1)
                elif mes_filter == 'todos':
                    for af in ArchivoFacturacion.objects.filter(anio=anio_int):
                        _facturado_entries.extend(_load_af(af))
                else:
                    af = ArchivoFacturacion.objects.get(mes=mes_filter, anio=anio_int)
                    _facturado_entries.extend(_load_af(af))
            except ArchivoFacturacion.DoesNotExist:
                pass

            # Match entries → Cliente: primero por RFC, luego por nombre
            _all_clientes_api = list(clientes_qs)
            _rfc_map = {c.rfc.upper().strip(): c for c in _all_clientes_api if c.rfc}
            total_by_id = {}
            for entry in _facturado_entries:
                c_obj = None
                rfc = entry['rfc'].upper().strip() if entry['rfc'] else ''
                nombre = entry['nombre']
                monto = entry['monto']
                # 1) Match por RFC (mas confiable)
                if rfc and rfc in _rfc_map:
                    c_obj = _rfc_map[rfc]
                # 2) Match por nombre (fallback)
                if not c_obj:
                    cn_upper = nombre.upper().strip()
                    for c in _all_clientes_api:
                        if c.nombre_empresa and c.nombre_empresa.upper().strip() == cn_upper:
                            c_obj = c; break
                    if not c_obj:
                        for c in _all_clientes_api:
                            if not c.nombre_empresa: continue
                            crm_u = c.nombre_empresa.upper().strip()
                            if crm_u in cn_upper or cn_upper in crm_u:
                                c_obj = c; break
                    if not c_obj:
                        pw = [w for w in cn_upper.split() if len(w) > 2 and w not in ('DE','DEL','LA','LAS','LOS','EL','SA','CV','SAS','INC','MEXICO')]
                        if len(pw) >= 2:
                            for c in _all_clientes_api:
                                if not c.nombre_empresa: continue
                                if pw[0] in c.nombre_empresa.upper() and pw[1] in c.nombre_empresa.upper():
                                    c_obj = c; break
                if c_obj:
                    total_by_id[c_obj.id] = total_by_id.get(c_obj.id, Decimal('0')) + monto
            # Total real del Excel (todos los entries, no solo los matcheados)
            _total_facturado_excel = sum(e['monto'] for e in _facturado_entries)
            prod_dict = {item['cliente']: item for item in _prod_annotate(base_qs.filter(monto_facturacion__gt=0), 'monto_facturacion')}
            meta_field_c = 'meta_mensual'
            vista_label = 'Facturado'
            # Prev month facturado sum
            _pm, _pa = _get_prev_params()
            if _pm:
                try:
                    _paf = ArchivoFacturacion.objects.get(mes=_pm, anio=_pa)
                    _prev_sum = sum(e['monto'] for e in _extract_entries(_paf.datos_json))
                except ArchivoFacturacion.DoesNotExist:
                    pass

        elif vista == 'cobrado':
            # Cobrado viene EXCLUSIVAMENTE del CSV (ArchivoCobrado), no de oportunidades
            _cobrado_entries = []  # [{nombre, monto}]
            _total_cobrado_csv = Decimal('0')
            try:
                def _extract_cobrado_entries(datos_json):
                    if not datos_json:
                        return []
                    entries = []
                    for key, val in datos_json.items():
                        if isinstance(val, dict) and 'monto' in val:
                            entries.append({
                                'nombre': val.get('nombre', key),
                                'monto': Decimal(str(val['monto'])),
                            })
                    return entries

                if usando_periodo:
                    from datetime import date as _date
                    cur = _date.fromisoformat(desde_filter).replace(day=1)
                    end = _date.fromisoformat(hasta_filter).replace(day=1)
                    while cur <= end:
                        try:
                            ac = ArchivoCobrado.objects.get(mes=cur.strftime('%m'), anio=cur.year)
                            _cobrado_entries.extend(_extract_cobrado_entries(ac.datos_json))
                        except ArchivoCobrado.DoesNotExist:
                            pass
                        cur = cur.replace(month=cur.month % 12 + 1, day=1) if cur.month < 12 else cur.replace(year=cur.year + 1, month=1, day=1)
                elif mes_filter == 'todos':
                    for ac in ArchivoCobrado.objects.filter(anio=anio_int):
                        _cobrado_entries.extend(_extract_cobrado_entries(ac.datos_json))
                else:
                    ac = ArchivoCobrado.objects.get(mes=mes_filter, anio=anio_int)
                    _cobrado_entries.extend(_extract_cobrado_entries(ac.datos_json))
            except ArchivoCobrado.DoesNotExist:
                pass

            # Cargar alias manuales
            alias_map = {a.palabra_clave.upper().strip(): a.buscar_como.upper().strip()
                         for a in AliasCliente.objects.all()}

            # Match entries → Cliente por nombre (usando misma lógica que facturado)
            _all_clientes_cob = list(clientes_qs)
            total_by_id = {}
            for entry in _cobrado_entries:
                c_obj = None
                nombre = entry['nombre']
                monto = entry['monto']
                cn_upper = nombre.upper().strip()
                # Aplicar alias si existe
                if cn_upper in alias_map:
                    cn_upper = alias_map[cn_upper]
                # Match por nombre
                for c in _all_clientes_cob:
                    if c.nombre_empresa and c.nombre_empresa.upper().strip() == cn_upper:
                        c_obj = c; break
                if not c_obj:
                    for c in _all_clientes_cob:
                        if not c.nombre_empresa: continue
                        crm_u = c.nombre_empresa.upper().strip()
                        if crm_u in cn_upper or cn_upper in crm_u:
                            c_obj = c; break
                if not c_obj:
                    pw = [w for w in cn_upper.split() if len(w) > 2 and w not in ('DE','DEL','LA','LAS','LOS','EL','SA','CV','SAS','INC','MEXICO')]
                    if len(pw) >= 2:
                        for c in _all_clientes_cob:
                            if not c.nombre_empresa: continue
                            if pw[0] in c.nombre_empresa.upper() and pw[1] in c.nombre_empresa.upper():
                                c_obj = c; break
                if c_obj:
                    total_by_id[c_obj.id] = total_by_id.get(c_obj.id, Decimal('0')) + monto
            _total_cobrado_csv = sum(e['monto'] for e in _cobrado_entries)
            prod_dict = {}
            meta_field_c = 'meta_cobrado'
            vista_label = 'Cobrado'
            # Prev month cobrado sum (también del CSV)
            _pm, _pa = _get_prev_params()
            if _pm:
                try:
                    _prev_ac = ArchivoCobrado.objects.get(mes=_pm, anio=_pa)
                    _prev_sum = sum(e['monto'] for e in _extract_cobrado_entries(_prev_ac.datos_json))
                except ArchivoCobrado.DoesNotExist:
                    pass

        elif vista == 'oportunidades':
            total_by_id = {item['cliente']: item['t'] for item in base_qs.values('cliente').annotate(t=Coalesce(Sum('monto'), _zero)) if item['cliente']}
            prod_dict = {item['cliente']: item for item in _prod_annotate(base_qs, 'monto')}
            meta_field_c = 'meta_oportunidades'
            vista_label = 'Oportunidades'
            # Previous month totals per client (for trend badge) — usa fecha_creacion
            _prev_by_id_cl = {}
            if mes_filter not in ('todos',) and not usando_periodo and not q_search:
                try:
                    _pm = int(mes_filter)
                    _prev_mes = _pm - 1 if _pm > 1 else 12
                    _prev_anio = anio_int if _pm > 1 else anio_int - 1
                    _prev_qs = TodoItem.objects.filter(fecha_creacion__year=_prev_anio, fecha_creacion__month=_prev_mes)
                    if not es_supervisor:
                        _gids = get_usuarios_visibles_ids(user)
                        _prev_qs = _prev_qs.filter(usuario_id__in=_gids) if _gids and len(_gids) > 1 else _prev_qs.filter(usuario=user)
                    elif vendedores_ids:
                        _prev_qs = _prev_qs.filter(usuario_id__in=vendedores_ids)
                    _prev_by_id_cl = {
                        item['cliente']: item['t']
                        for item in _prev_qs.values('cliente').annotate(t=Coalesce(Sum('monto'), _zero))
                        if item['cliente']
                    }
                except (ValueError, TypeError):
                    pass
            _prev_sum = sum(_prev_by_id_cl.values()) if _prev_by_id_cl else Decimal('0')

        elif vista == 'cotizado':
            opp_ids = base_qs.values_list('id', flat=True)
            if usando_periodo:
                cot_qs = Cotizacion.objects.select_related('cliente', 'oportunidad').filter(
                    Q(oportunidad_id__in=opp_ids) | Q(oportunidad__isnull=True, fecha_creacion__date__gte=desde_filter, fecha_creacion__date__lte=hasta_filter)
                )
            else:
                _cot_sueltas_q = Q(oportunidad__isnull=True, fecha_creacion__year=anio_int)
                if mes_filter != 'todos':
                    _cot_sueltas_q &= Q(fecha_creacion__month=int(mes_filter))
                cot_qs = Cotizacion.objects.select_related('cliente', 'oportunidad').filter(
                    Q(oportunidad_id__in=opp_ids) | _cot_sueltas_q
                )
            total_by_id = {}
            count_by_id = {}
            prod_dict_raw = {}
            PRODS = ['ZEBRA', 'PANDUIT', 'APC', 'AVIGILON', 'GENETEC', 'AXIS', 'SOFTWARE', 'RUNRATE', 'PÓLIZA']
            PROD_KEYS = ['zebra', 'panduit', 'apc', 'avigilon', 'genetec', 'axis', 'software', 'runrate', 'poliza']
            for cot in cot_qs:
                if not cot.cliente_id:
                    continue
                cid = cot.cliente_id
                t = cot.total or Decimal('0')
                total_by_id[cid] = total_by_id.get(cid, Decimal('0')) + t
                count_by_id[cid] = count_by_id.get(cid, 0) + 1
                prod = (cot.oportunidad.producto if cot.oportunidad else '') or ''
                prod_upper = prod.upper()
                if cid not in prod_dict_raw:
                    prod_dict_raw[cid] = {k: Decimal('0') for k in PROD_KEYS + ['total_prod']}
                prod_dict_raw[cid]['total_prod'] += t
                if 'ZEBRA' in prod_upper: prod_dict_raw[cid]['zebra'] += t
                elif 'PANDUIT' in prod_upper: prod_dict_raw[cid]['panduit'] += t
                elif prod_upper == 'APC': prod_dict_raw[cid]['apc'] += t
                elif 'AVIGILON' in prod_upper or 'AVIGILION' in prod_upper: prod_dict_raw[cid]['avigilon'] += t
                elif 'GENETEC' in prod_upper: prod_dict_raw[cid]['genetec'] += t
                elif 'AXIS' in prod_upper: prod_dict_raw[cid]['axis'] += t
                elif 'SOFTWARE' in prod_upper or 'DESARROLLO' in prod_upper: prod_dict_raw[cid]['software'] += t
                elif 'RUNRATE' in prod_upper: prod_dict_raw[cid]['runrate'] += t
                elif 'PÓLIZA' in prod_upper or 'POLIZA' in prod_upper: prod_dict_raw[cid]['poliza'] += t
            prod_dict = prod_dict_raw
            meta_field_c = 'meta_cotizado'
            vista_label = 'Cotizado'
            # Prev month cotizado sum — usa fecha_creacion
            _pm, _pa = _get_prev_params()
            if _pm:
                try:
                    _prev_opp_ids = TodoItem.objects.filter(fecha_creacion__year=_pa, fecha_creacion__month=int(_pm))
                    if not es_supervisor:
                        _gids = get_usuarios_visibles_ids(user)
                        _prev_opp_ids = _prev_opp_ids.filter(usuario_id__in=_gids) if _gids and len(_gids) > 1 else _prev_opp_ids.filter(usuario=user)
                    elif vendedores_ids:
                        _prev_opp_ids = _prev_opp_ids.filter(usuario_id__in=vendedores_ids)
                    _prev_cot_qs = Cotizacion.objects.filter(
                        Q(oportunidad_id__in=_prev_opp_ids) |
                        Q(oportunidad__isnull=True, fecha_creacion__year=_pa, fecha_creacion__month=int(_pm))
                    )
                    _prev_sum = _prev_cot_qs.aggregate(t=Coalesce(Sum('total'), _zero))['t'] or _zero
                except Exception:
                    pass
        elif vista == 'prospeccion':
          try:
            from .models import Prospecto, CampanaEnvio

            # Define clientes queryset for this scope
            if es_supervisor:
                clientes_qs = Cliente.objects.filter(asignado_a_id__in=vendedores_ids) if vendedores_ids else Cliente.objects.all()
            else:
                clientes_qs = Cliente.objects.filter(get_clientes_visibles_q(user))

            prospectos_qs = Prospecto.objects.filter(fecha_creacion__year=anio_int)
            if mes_filter and mes_filter != 'todos':
                try:
                    prospectos_qs = prospectos_qs.filter(fecha_creacion__month=int(mes_filter))
                except ValueError:
                    pass
            if not es_supervisor:
                _gids = get_usuarios_visibles_ids(user)
                prospectos_qs = prospectos_qs.filter(usuario_id__in=_gids) if _gids and len(_gids) > 1 else prospectos_qs.filter(usuario=user)
            elif vendedores_ids:
                prospectos_qs = prospectos_qs.filter(usuario_id__in=vendedores_ids)

            prosp_by_client = {}
            ganados_by_client = {}
            for p in prospectos_qs.select_related('cliente'):
                cid = p.cliente_id
                if not cid:
                    continue
                prosp_by_client[cid] = prosp_by_client.get(cid, 0) + 1
                if p.etapa == 'cerrado_ganado':
                    ganados_by_client[cid] = ganados_by_client.get(cid, 0) + 1

            # Campañas enviadas por cliente (graceful if table doesn't exist yet)
            camp_by_client = {}
            total_envios = 0
            total_respondidos = 0
            total_favorables = 0
            try:
                envios_qs = CampanaEnvio.objects.filter(fecha_envio__year=anio_int)
                if mes_filter and mes_filter != 'todos':
                    try:
                        envios_qs = envios_qs.filter(fecha_envio__month=int(mes_filter))
                    except ValueError:
                        pass
                if not es_supervisor:
                    envios_qs = envios_qs.filter(enviado_por=user)
                elif vendedores_ids:
                    envios_qs = envios_qs.filter(enviado_por_id__in=vendedores_ids)
                for env in envios_qs:
                    cid = env.cliente_id
                    if cid:
                        camp_by_client[cid] = camp_by_client.get(cid, 0) + 1
                total_envios = envios_qs.count()
                total_respondidos = envios_qs.filter(respondido=True).count()
                total_favorables = envios_qs.filter(respuesta_favorable=True).count()
            except Exception:
                pass  # Table may not exist yet

            # Ventas generadas desde prospeccion
            ventas_prosp = Decimal('0')
            try:
                opps_from_prosp = base_qs.filter(prospecto_origen__isnull=False)
                opps_vendidas = opps_from_prosp.filter(
                    Q(etapa_corta__icontains='vendido') | Q(etapa_corta__icontains='comprando') |
                    Q(etapa_corta__icontains='transito') | Q(etapa_corta__icontains='entregado') |
                    Q(etapa_corta__icontains='facturado') | Q(etapa_corta__icontains='cobrado') |
                    Q(probabilidad_cierre=100)
                )
                ventas_prosp = opps_vendidas.aggregate(t=Coalesce(Sum('monto'), Value(Decimal('0'))))['t'] or Decimal('0')
                total_opps_from_prosp_count = opps_from_prosp.count()
            except Exception:
                opps_from_prosp = TodoItem.objects.none()
                total_opps_from_prosp_count = 0

            # Chart data
            marca_counts = {}
            etapa_counts = {}
            for p in prospectos_qs:
                marca = (p.producto or 'OTRO').upper()
                marca_counts[marca] = marca_counts.get(marca, 0) + 1
                etapa_counts[p.etapa] = etapa_counts.get(p.etapa, 0) + 1

            all_client_ids = set(list(prosp_by_client.keys()) + list(camp_by_client.keys()))
            # Pre-load client names into dict for fast lookup
            client_map = {c.id: c for c in clientes_qs.filter(id__in=all_client_ids)}
            prosp_rows = []
            total_prosp = 0
            total_camp = 0
            total_ganados = 0
            for cid in sorted(all_client_ids, key=lambda x: (client_map.get(x).nombre_empresa if client_map.get(x) else '')):
                c_obj = client_map.get(cid)
                n_prosp = prosp_by_client.get(cid, 0)
                n_camp = camp_by_client.get(cid, 0)
                n_ganados = ganados_by_client.get(cid, 0)
                total_prosp += n_prosp
                total_camp += n_camp
                total_ganados += n_ganados
                prosp_rows.append({
                    'cliente_id': cid,
                    'cliente': (c_obj.nombre_empresa if c_obj else 'Sin nombre')[:35],
                    'vendedor': (c_obj.asignado_a.get_full_name() if c_obj and c_obj.asignado_a else ''),
                    'num_prospectos': n_prosp,
                    'num_campanas': n_camp,
                    'num_ganados': n_ganados,
                })

            num_clientes_prosp = len([r for r in prosp_rows if r['num_prospectos'] > 0])
            tasa_contacto = round(total_respondidos / total_envios * 100) if total_envios > 0 else 0

            return JsonResponse({
                'tab': 'clientes',
                'vista': 'prospeccion',
                'vista_label': 'Prospección',
                'rows': prosp_rows,
                'footer': {
                    'left': f'{num_clientes_prosp} clientes prospectados',
                    'right': f'{total_prosp} prospectos',
                },
                'total_prospectos': total_prosp,
                'total_campanas': total_camp,
                'total_ganados': total_ganados,
                'total_opps_from_prosp': total_opps_from_prosp_count,
                'ventas_generadas': format_money(ventas_prosp),
                'ventas_generadas_raw': float(ventas_prosp),
                'tasa_contacto': tasa_contacto,
                'total_envios': total_envios,
                'total_respondidos': total_respondidos,
                'total_favorables': total_favorables,
                'chart_marcas': marca_counts,
                'chart_etapas': etapa_counts,
                'meta': '0',
                'progreso': 0,
            })
          except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'tab': 'clientes', 'vista': 'prospeccion',
                'rows': [], 'footer': {'left': 'Error', 'right': str(e)},
                'total_prospectos': 0, 'total_campanas': 0, 'total_ganados': 0,
                'total_opps_from_prosp': 0, 'ventas_generadas': '0', 'ventas_generadas_raw': 0,
                'tasa_contacto': 0, 'total_envios': 0, 'total_respondidos': 0, 'total_favorables': 0,
                'chart_marcas': {}, 'chart_etapas': {}, 'meta': '0', 'progreso': 0,
            })

        else:
            return JsonResponse({'tab': 'clientes', 'rows': [], 'footer': {'left': '', 'right': ''}, 'vista': vista})

        rows = []
        total_acum = Decimal('0')
        for c in clientes_qs.order_by('nombre_empresa'):
            p = prod_dict.get(c.id, {})
            total_c = total_by_id.get(c.id, Decimal('0'))
            meta_c = getattr(c, meta_field_c, Decimal('0')) or Decimal('0')
            if mes_filter == 'todos' and not usando_periodo:
                meta_c = meta_c * 12
            faltante = meta_c - total_c
            vendedor_name = (c.asignado_a.get_full_name() or c.asignado_a.username) if c.asignado_a else ''
            rows.append({
                'cliente_id': c.id,
                'cliente': c.nombre_empresa[:35],
                'vendedor': vendedor_name,
                'zebra': format_money(p.get('zebra', 0)),
                'panduit': format_money(p.get('panduit', 0)),
                'apc': format_money(p.get('apc', 0)),
                'avigilon': format_money(p.get('avigilon', 0)),
                'genetec': format_money(p.get('genetec', 0)),
                'axis': format_money(p.get('axis', 0)),
                'software': format_money(p.get('software', 0)),
                'runrate': format_money(p.get('runrate', 0)),
                'poliza': format_money(p.get('poliza', 0)),
                'otros': format_money((p.get('total_prod', Decimal('0')) or Decimal('0')) - sum((p.get(k, Decimal('0')) or Decimal('0')) for k in ['zebra', 'panduit', 'apc', 'avigilon', 'genetec', 'axis', 'software', 'runrate', 'poliza'])),
                'total': format_money(total_c),
                '_total_raw': float(total_c),
                'meta': format_money(meta_c),
                'faltante': format_money(faltante),
                'prev_total': format_money(_prev_by_id_cl.get(c.id, Decimal('0'))),
                'num_cotizaciones': count_by_id.get(c.id, 0) if vista == 'cotizado' else 0,
            })
            total_acum += total_c

        # Ordenar por total de mayor a menor
        rows.sort(key=lambda r: r.get('_total_raw', 0), reverse=True)

        num_clientes_c = clientes_qs.count()

        # For cotizado vista, return number-based KPI
        if vista == 'cotizado':
            total_num_cot = sum(count_by_id.values())
            return JsonResponse({
                'tab': 'clientes',
                'vista': vista,
                'vista_label': vista_label,
                'rows': rows,
                'footer': {
                    'left': f'{num_clientes_c} clientes',
                    'right': f'Total: {total_num_cot} cotizaciones',
                },
                'total_facturado': str(total_num_cot),
                'total_monto_cotizado': format_money(total_acum),
                'widget_label': 'Cotizaciones Creadas',
                'meta': format_money(api_meta),
                'progreso': int((total_num_cot / int(api_meta) * 100)) if api_meta > 0 else 0,
                'widget_left_stat': f'{num_clientes_c} Clientes',
                'prev_sum': format_money(_prev_sum),
                'num_total_cotizaciones': total_num_cot,
            })

        # Para vista facturado/cobrado, usar el total real del archivo (incluye clientes sin match)
        if vista == 'facturado':
            _kpi_total = _total_facturado_excel
        elif vista == 'cobrado':
            _kpi_total = _total_cobrado_csv
        else:
            _kpi_total = total_acum
        return JsonResponse({
            'tab': 'clientes',
            'vista': vista,
            'vista_label': vista_label,
            'rows': rows,
            'footer': {
                'left': f'{num_clientes_c} clientes',
                'right': f'Total {vista_label}: ${format_money(_kpi_total)}',
            },
            'total_facturado': format_money(_kpi_total),
            'widget_label': f'Total {vista_label}',
            'meta': format_money(api_meta),
            'progreso': int((_kpi_total / api_meta * 100)) if api_meta > 0 else 0,
            'widget_left_stat': f'{num_clientes_c} Clientes',
            'prev_sum': format_money(_prev_sum),
        })

    return JsonResponse({'tab': tab_activo, 'rows': [], 'footer': {'left': '', 'right': ''}})


@login_required
@require_http_methods(["GET"])
def api_tendencia_mensual(request):
    """Devuelve totales mensuales de Facturado, Cobrado, Oportunidades y Cotizado
    para los últimos 6 meses, usado por la gráfica de tendencia."""
    from datetime import datetime
    from decimal import Decimal

    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)
    from .views_utils import is_supervisor as _is_sup
    es_supervisor = _is_sup(user)

    # Determinar vendedores a filtrar
    vendedores_ids = None
    if es_supervisor:
        vf = request.GET.get('vendedores', '')
        if vf and vf != 'todos':
            vendedores_ids = [int(x) for x in vf.split(',') if x.isdigit()]

    now = datetime.now()
    meses = []
    for i in range(5, -1, -1):
        m = now.month - i
        a = now.year
        while m <= 0:
            m += 12
            a -= 1
        meses.append((str(m).zfill(2), a))

    _zero = Decimal('0')
    labels = []
    data_fact = []
    data_cob = []
    data_opp = []
    data_cot = []

    nombres_mes = {
        '01': 'Ene', '02': 'Feb', '03': 'Mar', '04': 'Abr',
        '05': 'May', '06': 'Jun', '07': 'Jul', '08': 'Ago',
        '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dic'
    }

    for mes_str, anio in meses:
        labels.append(f"{nombres_mes[mes_str]} {anio}")

        # Facturado — desde ArchivoFacturacion
        fact_total = _zero
        try:
            af = ArchivoFacturacion.objects.get(mes=mes_str, anio=anio)
            datos = af.datos_json or {}
            for key, val in datos.items():
                if key == 'datos':
                    continue
                if isinstance(val, dict) and 'monto' in val:
                    fact_total += Decimal(str(val['monto']))
                else:
                    try:
                        fact_total += Decimal(str(val))
                    except Exception:
                        pass
        except ArchivoFacturacion.DoesNotExist:
            pass
        data_fact.append(float(fact_total))

        # Base queryset para oportunidades del mes
        opp_qs = TodoItem.objects.filter(anio_cierre=anio, mes_cierre=mes_str)
        if not es_supervisor:
            _gids = get_usuarios_visibles_ids(user)
            opp_qs = opp_qs.filter(usuario_id__in=_gids) if _gids and len(_gids) > 1 else opp_qs.filter(usuario=user)
        elif vendedores_ids:
            opp_qs = opp_qs.filter(usuario_id__in=vendedores_ids)

        # Cobrado — etapa Ganado/Pagado
        cob = opp_qs.filter(etapa_corta__in=['Ganado', 'Pagado']).aggregate(
            t=Coalesce(Sum('monto'), _zero))['t'] or _zero
        data_cob.append(float(cob))

        # Oportunidades — todas
        opp = opp_qs.aggregate(t=Coalesce(Sum('monto'), _zero))['t'] or _zero
        data_opp.append(float(opp))

        # Cotizado
        opp_ids = opp_qs.values_list('id', flat=True)
        cot = Cotizacion.objects.filter(
            Q(oportunidad_id__in=opp_ids) |
            Q(oportunidad__isnull=True, fecha_creacion__year=anio, fecha_creacion__month=int(mes_str))
        ).aggregate(t=Coalesce(Sum('total'), _zero))['t'] or _zero
        data_cot.append(float(cot))

    # Detectar puntos notables (cambios > 30% respecto al mes anterior)
    def find_annotations(data, label):
        annotations = []
        for i in range(1, len(data)):
            if data[i - 1] > 0:
                cambio = (data[i] - data[i - 1]) / data[i - 1] * 100
                if abs(cambio) >= 30:
                    annotations.append({
                        'index': i,
                        'cambio': round(cambio),
                        'label': label
                    })
        return annotations

    anotaciones = []
    anotaciones.extend(find_annotations(data_fact, 'Facturado'))
    anotaciones.extend(find_annotations(data_cob, 'Cobrado'))
    anotaciones.extend(find_annotations(data_opp, 'Oportunidades'))
    anotaciones.extend(find_annotations(data_cot, 'Cotizado'))

    return JsonResponse({
        'labels': labels,
        'facturado': data_fact,
        'cobrado': data_cob,
        'oportunidades': data_opp,
        'cotizado': data_cot,
        'anotaciones': anotaciones,
    })


@login_required
@require_http_methods(["GET"])
def api_desglose_facturacion(request):
    """Desglose completo de facturación del Excel por cliente (sin filtro de match)"""
    try:
        mes = request.GET.get('mes', 'todos')
        anio = int(request.GET.get('anio', 2026))
        acumulado = {}  # {key: {nombre, rfc, monto}}

        def _procesar_af(af):
            raw = af.datos_json or {}
            for key, val in raw.items():
                if key == 'datos':
                    continue
                if isinstance(val, dict) and 'monto' in val:
                    nombre = val.get('nombre', key)
                    rfc = val.get('rfc', '')
                    monto = float(Decimal(str(val['monto'])))
                else:
                    nombre = key
                    rfc = ''
                    try:
                        monto = float(Decimal(str(val)))
                    except Exception:
                        continue
                k = rfc if rfc else nombre
                if k in acumulado:
                    acumulado[k]['monto'] += monto
                else:
                    acumulado[k] = {'nombre': nombre, 'rfc': rfc, 'monto': monto}

        if mes == 'todos':
            for af in ArchivoFacturacion.objects.filter(anio=anio):
                _procesar_af(af)
        else:
            try:
                _procesar_af(ArchivoFacturacion.objects.get(mes=mes, anio=anio))
            except ArchivoFacturacion.DoesNotExist:
                pass

        rows = sorted(acumulado.values(), key=lambda x: -x['monto'])
        total = sum(r['monto'] for r in rows)
        return JsonResponse({'ok': True, 'rows': rows, 'total': total})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_subir_facturacion(request):
    """
    API para subir archivo XLS de facturación.
    Solo supervisores pueden subir.
    Parsea el XLS y extrae total de pagos por cliente.
    """
    if not is_supervisor(request.user):
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)

    from datetime import datetime as dt_now
    archivo = request.FILES.get('archivo')
    mes = request.POST.get('mes', '')
    anio = request.POST.get('anio', '')

    if not archivo:
        return JsonResponse({'success': False, 'error': 'Falta el archivo'})

    # Defaults si no vienen mes/anio
    now = dt_now.now()
    if not mes:
        mes = str(now.month).zfill(2)

    try:
        anio_int = int(anio) if anio else now.year
    except (ValueError, TypeError):
        anio_int = now.year

    try:
        import xlrd
        from xlrd import xldate_as_datetime
        content = archivo.read()
        wb = xlrd.open_workbook(file_contents=content)
        sheet = wb.sheet_by_index(0)

        # Agrupar por mes de emisión: {(mes, anio): {cliente: monto}}
        datos_por_mes = {}  # { 'MM': { 'YYYY': { cliente_name: monto_str } } }
        totales_por_mes = {}  # { (mes, anio): Decimal }

        for row_idx in range(1, sheet.nrows):
            try:
                estatus = str(sheet.cell_value(row_idx, 40)).strip().lower()
                if estatus == 'cancelada':
                    continue

                cliente_name = str(sheet.cell_value(row_idx, 5)).strip()
                nombre_comercial = str(sheet.cell_value(row_idx, 6)).strip()
                if not cliente_name:
                    continue

                # Extraer mes/año de la fecha de emisión (col D, idx 3)
                try:
                    date_val = sheet.cell_value(row_idx, 3)
                    fecha = xldate_as_datetime(date_val, wb.datemode)
                    row_mes = str(fecha.month).zfill(2)
                    row_anio = fecha.year
                except Exception:
                    continue  # Sin fecha válida, saltar

                # Facturado = Col L Subtotal (idx 11) - Col O Descuento (idx 14)
                # Los importes ya vienen en pesos, no se multiplica por T.C.
                subtotal_str = str(sheet.cell_value(row_idx, 11)).replace(',', '').strip()
                descuento_str = str(sheet.cell_value(row_idx, 14)).replace(',', '').strip()
                try:
                    subtotal = Decimal(subtotal_str) if subtotal_str else Decimal('0')
                except Exception:
                    subtotal = Decimal('0')
                try:
                    descuento = Decimal(descuento_str) if descuento_str else Decimal('0')
                except Exception:
                    descuento = Decimal('0')
                monto = subtotal - descuento

                # RFC del cliente (col G, idx 7)
                rfc_raw = str(sheet.cell_value(row_idx, 7)).strip()
                rfc = rfc_raw if rfc_raw and rfc_raw != '0.0' else ''

                if monto > 0:
                    key = (row_mes, row_anio)
                    if key not in datos_por_mes:
                        datos_por_mes[key] = {}
                        totales_por_mes[key] = Decimal('0')

                    clientes_mes = datos_por_mes[key]
                    # Usar RFC como key si existe, sino nombre
                    nombre_final = nombre_comercial if nombre_comercial else cliente_name
                    # Guardar con RFC para match preciso
                    entry_key = rfc if rfc else nombre_final
                    if entry_key in clientes_mes:
                        existing = clientes_mes[entry_key]
                        existing['monto'] = str(Decimal(existing['monto']) + monto)
                    else:
                        clientes_mes[entry_key] = {
                            'nombre': nombre_final,
                            'rfc': rfc,
                            'monto': str(monto),
                        }
                    totales_por_mes[key] += monto
            except (IndexError, ValueError):
                continue

        # Guardar un ArchivoFacturacion por cada mes encontrado
        archivo.seek(0)
        meses_guardados = []
        for (m, a), clientes_data in datos_por_mes.items():
            obj, created = ArchivoFacturacion.objects.update_or_create(
                mes=m, anio=a,
                defaults={
                    'archivo': archivo,
                    'total_facturado': totales_por_mes[(m, a)],
                    'datos_json': clientes_data,
                    'subido_por': request.user,
                }
            )
            meses_guardados.append(f"{m}/{a}")

        total_general = sum(totales_por_mes.values())
        return JsonResponse({
            'success': True,
            'total_facturado': str(total_general),
            'num_clientes': sum(len(v) for v in datos_por_mes.values()),
            'meses': meses_guardados,
            'created': True,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error procesando archivo: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def api_subir_cobrado(request):
    """
    API para subir CSV de ingresos (cobrado).
    Parsea el CSV y extrae cobros por cliente con detalle de facturas.
    datos_json: {cliente_name: {nombre, monto, facturas: [{factura, monto, fecha}]}}
    """
    if not is_supervisor(request.user):
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)

    from datetime import datetime as dt_now
    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'success': False, 'error': 'Falta el archivo'})

    try:
        import io
        content = archivo.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))

        datos_por_mes = {}   # {(mes, anio): {cliente: {nombre, monto, facturas:[]}}}
        totales_por_mes = {}

        for row in reader:
            try:
                cliente_name = (row.get('Cliente') or '').strip()
                if not cliente_name:
                    continue

                fecha_str = (row.get('Fecha') or '').strip()
                if not fecha_str:
                    continue
                try:
                    fecha = dt_now.strptime(fecha_str, '%d/%m/%Y %I:%M:%S %p')
                except ValueError:
                    try:
                        fecha = dt_now.strptime(fecha_str, '%d/%m/%Y %H:%M:%S')
                    except ValueError:
                        continue

                row_mes = str(fecha.month).zfill(2)
                row_anio = fecha.year

                total_str = (row.get('Total') or '0').replace(',', '').strip()
                try:
                    monto = Decimal(total_str)
                except Exception:
                    continue
                if monto <= 0:
                    continue

                facturas_str = (row.get('Facturas') or '').strip()
                fecha_corta = fecha.strftime('%d/%m/%Y')

                key = (row_mes, row_anio)
                if key not in datos_por_mes:
                    datos_por_mes[key] = {}
                    totales_por_mes[key] = Decimal('0')

                clientes_mes = datos_por_mes[key]
                if cliente_name in clientes_mes:
                    existing = clientes_mes[cliente_name]
                    existing['monto'] = str(Decimal(existing['monto']) + monto)
                    existing['facturas'].append({
                        'factura': facturas_str,
                        'monto': str(monto),
                        'fecha': fecha_corta,
                    })
                else:
                    clientes_mes[cliente_name] = {
                        'nombre': cliente_name,
                        'monto': str(monto),
                        'facturas': [{
                            'factura': facturas_str,
                            'monto': str(monto),
                            'fecha': fecha_corta,
                        }],
                    }
                totales_por_mes[key] += monto
            except (IndexError, ValueError, KeyError):
                continue

        archivo.seek(0)
        meses_guardados = []
        for (m, a), clientes_data in datos_por_mes.items():
            obj, created = ArchivoCobrado.objects.update_or_create(
                mes=m, anio=a,
                defaults={
                    'archivo': archivo,
                    'total_cobrado': totales_por_mes[(m, a)],
                    'datos_json': clientes_data,
                    'subido_por': request.user,
                }
            )
            meses_guardados.append(f"{m}/{a}")

        total_general = sum(totales_por_mes.values())
        return JsonResponse({
            'success': True,
            'total_cobrado': str(total_general),
            'num_clientes': sum(len(v) for v in datos_por_mes.values()),
            'meses': meses_guardados,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error procesando archivo: {str(e)}'})


def _match_clientes_cobrado(nombre_csv, clientes_list, alias_map=None):
    """
    Match inteligente: devuelve TODOS los clientes de la BD que matchean.
    Primero revisa alias manuales, luego match por nombre.
    alias_map: {PALABRA_CLAVE_UPPER: BUSCAR_COMO_UPPER}
    """
    cn_upper = nombre_csv.upper().strip()
    stop = {'DE', 'DEL', 'LA', 'LAS', 'LOS', 'EL', 'SA', 'CV', 'SAS', 'INC', 'MEXICO', 'S', 'RL', 'INDUSTRIES'}

    # 0) Alias manuales: si alguna palabra_clave está contenida en el nombre CSV, usar buscar_como
    if alias_map:
        for keyword, buscar in alias_map.items():
            if keyword in cn_upper:
                matches = []
                for c in clientes_list:
                    if not c.nombre_empresa:
                        continue
                    if buscar in c.nombre_empresa.upper():
                        matches.append(c)
                if matches:
                    return matches

    # 1) Exact match — devolver solo ese
    for c in clientes_list:
        if c.nombre_empresa and c.nombre_empresa.upper().strip() == cn_upper:
            return [c]

    # 2) Extraer palabra clave principal del CSV (primera palabra significativa de 4+ chars)
    palabras = [w for w in cn_upper.split() if len(w) >= 4 and w not in stop]
    if not palabras:
        palabras = [w for w in cn_upper.split() if len(w) >= 3 and w not in stop]

    # 3) Buscar TODOS los clientes que contengan la palabra clave principal
    matches = []
    if palabras:
        keyword = palabras[0]
        for c in clientes_list:
            if not c.nombre_empresa:
                continue
            if keyword in c.nombre_empresa.upper():
                matches.append(c)

    if matches:
        return matches

    # 4) Fallback: contiene o está contenido
    for c in clientes_list:
        if not c.nombre_empresa:
            continue
        crm_upper = c.nombre_empresa.upper().strip()
        if crm_upper in cn_upper or cn_upper in crm_upper:
            return [c]

    return []


@login_required
@require_http_methods(["GET"])
def api_desglose_cobrado(request):
    """Desglose de cobrado con match a clientes, vendedor, meta y facturas."""
    try:
        mes = request.GET.get('mes', 'todos')
        anio = int(request.GET.get('anio', 2026))
        acumulado = {}  # {nombre: {nombre, monto, facturas[]}}

        def _procesar(ac):
            raw = ac.datos_json or {}
            for key, val in raw.items():
                if isinstance(val, dict) and 'monto' in val:
                    nombre = val.get('nombre', key)
                    monto = float(Decimal(str(val['monto'])))
                    facturas = val.get('facturas', [])
                else:
                    continue
                if nombre in acumulado:
                    acumulado[nombre]['monto'] += monto
                    acumulado[nombre]['facturas'].extend(facturas)
                else:
                    acumulado[nombre] = {'nombre': nombre, 'monto': monto, 'facturas': facturas}

        if mes == 'todos':
            for ac in ArchivoCobrado.objects.filter(anio=anio):
                _procesar(ac)
        else:
            try:
                _procesar(ArchivoCobrado.objects.get(mes=mes, anio=anio))
            except ArchivoCobrado.DoesNotExist:
                pass

        # Cargar alias manuales
        alias_map = {a.palabra_clave.upper().strip(): a.buscar_como.upper().strip()
                     for a in AliasCliente.objects.all()}

        # Match con clientes de la BD — busca TODAS las variantes
        all_clientes = list(Cliente.objects.select_related('asignado_a').all())
        rows = []
        for entry in sorted(acumulado.values(), key=lambda x: -x['monto']):
            matches = _match_clientes_cobrado(entry['nombre'], all_clientes, alias_map)
            vendedor = ''
            meta_cobrado = 0
            if matches:
                # Vendedor: tomar del primer match que tenga asignado
                for m in matches:
                    if m.asignado_a:
                        vendedor = (m.asignado_a.get_full_name() or m.asignado_a.username)
                        break
                # Meta: sumar meta_cobrado de TODAS las variantes
                for m in matches:
                    mc = float(m.meta_cobrado or 0)
                    if mes == 'todos':
                        mc = mc * 12
                    meta_cobrado += mc
            faltante = meta_cobrado - entry['monto']
            rows.append({
                'nombre': entry['nombre'],
                'monto': entry['monto'],
                'vendedor': vendedor,
                'meta': meta_cobrado,
                'faltante': faltante,
                'facturas': entry.get('facturas', []),
            })

        total = sum(r['monto'] for r in rows)
        return JsonResponse({'ok': True, 'rows': rows, 'total': total})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@login_required
def api_cliente_oportunidades(request, cliente_id):
    """
    API que devuelve las oportunidades de un cliente específico en JSON.
    Formato compatible con buildCrmRow() del frontend.
    """
    from datetime import datetime
    user = request.user
    es_supervisor = is_supervisor(user)

    # Por defecto mostrar todo el historial del cliente, no solo el mes actual
    mes_filter = request.GET.get('mes', 'todos')
    anio_filter = request.GET.get('anio', 'todos')

    MES_CODE_TO_NAME = {
        '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
        '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
        '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre',
    }

    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cliente no encontrado'}, status=404)

    qs = TodoItem.objects.select_related('cliente', 'contacto', 'usuario').filter(cliente=cliente)

    # Filtrar por fecha_creacion si se indica (usado en tab Clientes)
    por_creacion = request.GET.get('por_creacion', '')

    # Aplicar filtros solo si no son 'todos'
    if anio_filter != 'todos':
        try:
            anio_int = int(anio_filter)
            if por_creacion:
                qs = qs.filter(fecha_creacion__year=anio_int)
            else:
                qs = qs.filter(anio_cierre=anio_int)
        except (ValueError, TypeError):
            pass

    if mes_filter != 'todos':
        if por_creacion:
            qs = qs.filter(fecha_creacion__month=int(mes_filter))
        else:
            qs = qs.filter(mes_cierre=mes_filter)

    if not es_supervisor:
        _gids = get_usuarios_visibles_ids(user)
        qs = qs.filter(usuario_id__in=_gids) if _gids and len(_gids) > 1 else qs.filter(usuario=user)

    # Filtrar por tipo si se especifica
    tipo = request.GET.get('tipo', '')
    if tipo == 'cobrado':
        qs = qs.filter(etapa_corta__in=['Ganado', 'Pagado'])

    qs = qs.order_by('-fecha_actualizacion')

    def format_money(val):
        if val is None:
            return '0'
        try:
            return '{:,.0f}'.format(val)
        except (ValueError, TypeError):
            return '0'

    rows = []
    for item in qs:
        rows.append({
            'id': item.id,
            'oportunidad': (item.oportunidad or '')[:35],
            'cliente': (item.cliente.nombre_empresa if item.cliente else '- Sin Cliente -')[:35],
            'cliente_id': item.cliente_id,
            'contacto': {'nombre': (item.contacto.nombre[:18] if item.contacto else '-')},
            'area': item.area or '-',
            'producto': item.producto or '',
            'producto_display': item.get_producto_display() if hasattr(item, 'get_producto_display') else (item.producto or ''),
            'usuario': item.usuario.get_full_name() or item.usuario.username if item.usuario else '—',
            'fecha': item.fecha_actualizacion.strftime('%d/%m/%y') if item.fecha_actualizacion else '—',
            'fecha_iso': item.fecha_actualizacion.isoformat() if item.fecha_actualizacion else '',
            'monto': format_money(item.monto),
            'monto_raw': float(item.monto or 0),
            'probabilidad_cierre': item.probabilidad_cierre,
        })

    # All contacts for this client (for filter dropdown)
    contactos_cliente = list(
        Contacto.objects.filter(cliente=cliente)
        .values_list('nombre', flat=True)
        .order_by('nombre')
    )

    return JsonResponse({
        'success': True,
        'cliente_nombre': cliente.nombre_empresa,
        'rows': rows,
        'contactos': contactos_cliente,
    })


@login_required
def editar_venta_todoitem(request, pk):
    # Supervisor puede editar cualquier venta, vendedor solo las suyas
    if is_supervisor(request.user):
        todo_item = get_object_or_404(TodoItem, pk=pk)
    else:
        todo_item = get_object_or_404(TodoItem, pk=pk, usuario=request.user)

    if request.method == 'POST':
        if 'delete' in request.POST:
            # Solo permite borrar si el usuario es supervisor o dueño
            if is_supervisor(request.user) or todo_item.usuario == request.user:
                todo_item.delete()
                messages.success(request, "Oportunidad eliminada con éxito.")
            else:
                messages.error(request, "No tienes permiso para eliminar esta oportunidad.")
            return redirect('todos')
        
        # Si no es delete, entonces es edición:
        form = VentaForm(request.POST, instance=todo_item, user=request.user if not is_supervisor(request.user) else None)
        if form.is_valid():
            cliente_nombre = form.cleaned_data['cliente_nombre']
            bitrix_company_id_from_form = form.cleaned_data.get('bitrix_company_id')

            cliente = None
            # Try to get client by bitrix_company_id if provided
            if bitrix_company_id_from_form:
                try:
                    cliente = Cliente.objects.get(bitrix_company_id=bitrix_company_id_from_form)
                except Cliente.DoesNotExist:
                    # If client doesn't exist locally, create it with the provided Bitrix ID
                    cliente = Cliente.objects.create(
                        nombre_empresa=cliente_nombre,
                        bitrix_company_id=bitrix_company_id_from_form
                    )
                except Exception as e:
                    messages.error(request, f"ERROR: No se pudo obtener o crear el cliente local por Bitrix ID: {e}")
                    form.add_error('cliente_nombre', 'Hubo un error al procesar el cliente.')
                    return render(request, 'editar_venta.html', {'form': form, 'todo_item': todo_item})
            else:
                # If no bitrix_company_id from form, try to find by name or create a new one without Bitrix ID
                try:
                    cliente, created = Cliente.objects.get_or_create(
                        nombre_empresa=cliente_nombre,
                        defaults={'bitrix_company_id': None}
                    )
                except Exception as e:
                    messages.error(request, f"ERROR: No se pudo obtener o crear el cliente local por nombre: {e}")
                    form.add_error('cliente_nombre', 'Hubo un error al procesar el cliente.')
                    return render(request, 'editar_venta.html', {'form': form, 'todo_item': todo_item})

            venta = form.save(commit=False)
            venta.cliente = cliente
            if is_supervisor(request.user):
                venta.usuario = form.cleaned_data['usuario']
            else:
                venta.usuario = request.user
            venta.save()

            # Actualizar en Bitrix si existe un bitrix_deal_id
            if venta.bitrix_deal_id:
                opportunity_data = {
                    'oportunidad': venta.oportunidad,
                    'monto': float(venta.monto),
                    'cliente': cliente.nombre_empresa,
                    'bitrix_company_id': cliente.bitrix_company_id,
                    'producto': venta.producto,
                    'area': venta.area,
                    'mes_cierre': venta.mes_cierre,
                    'probabilidad_cierre': venta.probabilidad_cierre,
                    'comentarios': venta.comentarios,
                    'bitrix_stage_id': venta.bitrix_stage_id,
                }
                # Obtener el bitrix_user_id del usuario asignado
                bitrix_assigned_by_id = None
                if venta.usuario and hasattr(venta.usuario, 'userprofile') and venta.usuario.userprofile.bitrix_user_id:
                    bitrix_assigned_by_id = venta.usuario.userprofile.bitrix_user_id
                opportunity_data['bitrix_assigned_by_id'] = bitrix_assigned_by_id
                bitrix_updated = update_opportunity_in_bitrix(venta.bitrix_deal_id, opportunity_data, request=request)
                if bitrix_updated:
                    messages.success(request, "Oportunidad actualizada en Bitrix24 con éxito.")
                else:
                    messages.error(request, "Error al actualizar la oportunidad en Bitrix24.")
            else:
                messages.warning(request, "La oportunidad no tiene un ID de Bitrix24 asociado. No se pudo actualizar en Bitrix24.")

            return redirect('todos')
    else:
        form = VentaForm(instance=todo_item, user=request.user if not is_supervisor(request.user) else None)

    return render(request, 'editar_venta.html', {'form': form, 'todo_item': todo_item})


@login_required
def reporte_ventas_por_cliente(request):
    from django.contrib.auth.models import User
    if is_supervisor(request.user):
        reporte_data = Cliente.objects.annotate(
            total_monto=Coalesce(
                Sum('oportunidades__monto', filter=Q(oportunidades__etapa_corta__in=['Ganado', 'Pagado'])),
                Value(Decimal('0.00'))
            )
        ).values(
            'id',
            'nombre_empresa',
            'total_monto'
        ).order_by('nombre_empresa')
        total_general = TodoItem.objects.filter(etapa_corta__in=['Ganado', 'Pagado']).aggregate(
            sum_monto=Sum('monto')
        )['sum_monto'] or Decimal('0.00')
        usuarios = User.objects.filter(is_active=True)
    else:
        _visible_ids = get_usuarios_visibles_ids(request.user)
        reporte_data = Cliente.objects.filter(get_clientes_visibles_q(request.user)).annotate(
            total_monto=Coalesce(
                Sum('oportunidades__monto', filter=Q(oportunidades__etapa_corta__in=['Ganado', 'Pagado'], oportunidades__usuario__in=_visible_ids) if _visible_ids else Q(oportunidades__etapa_corta__in=['Ganado', 'Pagado'])),
                Value(Decimal('0.00'))
            )
        ).values(
            'id',
            'nombre_empresa',
            'total_monto'
        ).order_by('nombre_empresa')
        if _visible_ids:
            total_general = TodoItem.objects.filter(
                usuario__in=_visible_ids, etapa_corta__in=['Ganado', 'Pagado']
            ).aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')
        else:
            total_general = TodoItem.objects.filter(
                etapa_corta__in=['Ganado', 'Pagado']
            ).aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')
        usuarios = None
    context = {
        'reporte_data': reporte_data,
        'total_general': total_general,
        'is_supervisor': is_supervisor(request.user),
        'usuarios': usuarios,
    }
    return render(request, 'reporte_ventas_por_cliente.html', context)


@login_required
def oportunidades_por_cliente(request, cliente_id):
    # Determinar qué clientes pueden ser vistos por el usuario
    if is_supervisor(request.user):
        cliente_seleccionado = get_object_or_404(Cliente, pk=cliente_id) # No filtrar por usuario
        oportunidades = TodoItem.objects.filter(cliente=cliente_seleccionado) # Todas las oportunidades del cliente
        print("DEBUG: Supervisor viendo oportunidades de cliente.")
    else:
        _visible_ids = get_usuarios_visibles_ids(request.user)
        _visible_q = get_clientes_visibles_q(request.user)
        cliente_seleccionado = get_object_or_404(Cliente, pk=cliente_id)
        # Verify the client is visible to this user
        if not Cliente.objects.filter(_visible_q, pk=cliente_id).exists():
            from django.http import Http404
            raise Http404
        oportunidades = TodoItem.objects.filter(cliente=cliente_seleccionado, usuario__in=_visible_ids) if _visible_ids else TodoItem.objects.filter(cliente=cliente_seleccionado)
        print(f"DEBUG: Vendedor {request.user.username} viendo oportunidades de grupo de cliente.")

    # El formulario de filtro no necesita el usuario para sus querysets de clientes en este contexto
    # ya que los clientes ya vienen filtrados por la vista o se obtienen todos.
    filter_form = VentaFilterForm(request.GET)

    if filter_form.is_valid():
        area = filter_form.cleaned_data.get('area')
        producto = filter_form.cleaned_data.get('producto')
        orden_monto = filter_form.cleaned_data.get('orden_monto')
        probabilidad_min = filter_form.cleaned_data.get('probabilidad_min')
        probabilidad_max = filter_form.cleaned_data.get('probabilidad_max')
        mes_cierre = filter_form.cleaned_data.get('mes_cierre')

        if area:
            oportunidades = oportunidades.filter(area=area)
        if producto:
            oportunidades = oportunidades.filter(producto=producto)
        if probabilidad_min is not None:
            oportunidades = oportunidades.filter(probabilidad_cierre__gte=probabilidad_min)
        if probabilidad_max is not None:
            oportunidades = oportunidades.filter(probabilidad_cierre__lte=probabilidad_max)
        if mes_cierre:
            oportunidades = oportunidades.filter(mes_cierre=mes_cierre)

        if orden_monto:
            if orden_monto == 'monto_asc':
                oportunidades = oportunidades.order_by('monto')
            elif orden_monto == 'monto_desc':
                oportunidades = oportunidades.order_by('-monto')
        else:
            oportunidades = oportunidades.order_by('-fecha_creacion')
    else:
        oportunidades = oportunidades.order_by('-fecha_creacion')


    context = {
        'cliente': cliente_seleccionado,
        'oportunidades': oportunidades,
        'filter_form': filter_form,
        'is_supervisor': is_supervisor(request.user),
    }
    return render(request, 'oportunidades_por_cliente.html', context)


@login_required
def producto_dashboard_detail(request, producto_val):
    print(f"DEBUG: producto_dashboard_detail - producto_val recibido RAW: {producto_val}")

    # Convertir a mayúsculas para asegurar que la comparación con PRODUCTO_CHOICES sea consistente
    producto_val_upper = producto_val.upper()
    print(f"DEBUG: producto_dashboard_detail - producto_val_upper: {producto_val_upper}")
    print(f"DEBUG: Keys de PRODUCTO_CHOICES: {list(dict(TodoItem.PRODUCTO_CHOICES).keys())}")

    # Verificar si el producto_val_upper es una clave válida en PRODUCTO_CHOICES
    if producto_val_upper not in dict(TodoItem.PRODUCTO_CHOICES):
        return redirect('dashboard')

    if is_supervisor(request.user):
        oportunidades = TodoItem.objects.filter(producto=producto_val_upper)
    else:
        oportunidades = TodoItem.objects.filter(producto=producto_val_upper, usuario=request.user)

    print(f"DEBUG: Oportunidades encontradas para {producto_val_upper} (antes de desglosar): {oportunidades.count()}")
    for op in oportunidades:
        print(f"DEBUG:   - ID: {op.id}, Oportunidad: {op.oportunidad}, Producto: {op.producto}, Usuario ID: {op.usuario.id}")

    # --- Ventas Cerradas (etapa Ganado/Pagado) para este producto ---
    ventas_cerradas = oportunidades.filter(etapa_corta__in=['Ganado', 'Pagado'])
    total_vendido_cerrado = ventas_cerradas.aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')
    total_vendido_cerrado_count = ventas_cerradas.count() # Conteo de oportunidades cerradas
    print(f"DEBUG: Ventas Cerradas (100%) para '{producto_val_upper}': {total_vendido_cerrado_count} oportunidades, Monto: {total_vendido_cerrado}")
    for venta in ventas_cerradas:
        print(f"DEBUG:   - Oportunidad: {venta.oportunidad}, Monto: {venta.monto}, Probabilidad: {venta.probabilidad_cierre}%")

    # --- Oportunidades Vigentes (probabilidad del 1% al 99%) para este producto ---
    oportunidades_vigentes = oportunidades.filter(
        probabilidad_cierre__gt=0, # Mayor que 0%
        probabilidad_cierre__lt=100 # Menor que 100%
    )
    total_monto_vigente = oportunidades_vigentes.aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')
    total_monto_vigente_count = oportunidades_vigentes.count() # Conteo de oportunidades vigentes
    print(f"DEBUG: Oportunidades Vigentes (0% < prob < 100%) para '{producto_val_upper}': {total_monto_vigente_count} oportunidades, Monto: {total_monto_vigente}")
    for op_vigente in oportunidades_vigentes:
        print(f"DEBUG:   - Oportunidad: {op_vigente.oportunidad}, Monto: {op_vigente.monto}, Probabilidad: {op_vigente.probabilidad_cierre}%")

    # --- Oportunidades Perdidas (probabilidad 0%) para este producto ---
    oportunidades_perdidas = oportunidades.filter(probabilidad_cierre=0)
    total_monto_perdido = oportunidades_perdidas.aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')
    total_monto_perdido_count = oportunidades_perdidas.count() # Conteo de oportunidades perdidas
    print(f"DEBUG: Oportunidades Perdidas (0%) para '{producto_val_upper}': {total_monto_perdido_count} oportunidades, Monto: {total_monto_perdido}")
    for op_perdida in oportunidades_perdidas:
        print(f"DEBUG:   - Oportunidad: {op_perdida.oportunidad}, Monto: {op_perdida.monto}, Probabilidad: {op_perdida.probabilidad_cierre}%")


    # Clientes involucrados en este producto
    clientes_involucrados = oportunidades.filter(cliente__isnull=False).values('cliente__id', 'cliente__nombre_empresa').distinct()

    # Meses involucrados en este producto (mes de cierre esperado)
    meses_involucrados = oportunidades.values('mes_cierre').distinct()

    # Mapear valores crudos de mes a sus nombres de visualización
    meses_display = []
    for m in meses_involucrados:
        # Aseguramos que la clave sea un string de dos dígitos para la búsqueda
        mes_key = str(m['mes_cierre']).zfill(2)
        meses_display.append(dict(TodoItem.MES_CHOICES).get(mes_key, mes_key))
    context = {
        'producto_val': producto_val_upper, # Aseguramos que la clave pasada sea la que usará el template
        'producto_display': dict(TodoItem.PRODUCTO_CHOICES).get(producto_val_upper, producto_val_upper),
        'total_vendido_cerrado': total_vendido_cerrado,
        'total_vendido_cerrado_count': total_vendido_cerrado_count, # AÑADIDO
        'total_monto_vigente': total_monto_vigente, # Nuevo: Monto oportunidades vigentes
        'total_monto_vigente_count': total_monto_vigente_count, # AÑADIDO
        'total_monto_perdido': total_monto_perdido, # Nuevo: Monto oportunidades perdidas
        'total_monto_perdido_count': total_monto_perdido_count, # AÑADIDO
        'clientes_involucrados': clientes_involucrados,
        'meses_involucrados_display': meses_display,
        'oportunidades': oportunidades, # Pasar todas las oportunidades para listarlas
        'is_supervisor': is_supervisor(request.user), # Pasamos si el usuario es supervisor al contexto
    }
    return redirect('/app/todos/')


@login_required
def mes_dashboard_detail(request, mes_val):
    # Asegúrate de que el mes_val recibido es uno de los choices válidos
    mes_val_padded = str(mes_val).zfill(2) # Asegurar que mes_val sea de dos dígitos para la validación
    if mes_val_padded not in dict(TodoItem.MES_CHOICES).keys():
        return redirect('home') # Redirige a home si el mes no es válido

    # Base queryset de oportunidades según el rol
    if is_supervisor(request.user):
        oportunidades_mes = TodoItem.objects.filter(mes_cierre=mes_val_padded)
    else:
        oportunidades_mes = TodoItem.objects.filter(usuario=request.user, mes_cierre=mes_val_padded)


    # Monto total esperado para este mes
    total_monto_esperado = oportunidades_mes.aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')

    # Monto POR COBRAR: oportunidades con probabilidad entre 1 and 99%
    por_cobrar_monto = oportunidades_mes.filter(probabilidad_cierre__gte=1, probabilidad_cierre__lte=99).aggregate(sum_monto=Sum('monto'))['sum_monto'] or Decimal('0.00')

    # Clientes involucrados en oportunidades para este mes
    clientes_involucrados = oportunidades_mes.filter(cliente__isnull=False).values('cliente__id', 'cliente__nombre_empresa').distinct()

    # Datos para la gráfica: Probabilidad de cierre vs. Monto
    graph_data_raw = oportunidades_mes.values('id', 'oportunidad', 'producto', 'monto', 'probabilidad_cierre', 'cliente__nombre_empresa')

    # Añadir 'get_producto_display' a cada item en graph_data
    graph_data_with_display = []
    for item in graph_data_raw:
        item_copy = item.copy()
        item_copy['get_producto_display'] = dict(TodoItem.PRODUCTO_CHOICES).get(item_copy['producto'], item_copy['producto'])
        graph_data_with_display.append(item_copy)

    context = {
        'mes_val': mes_val_padded, # Aseguramos que la clave pasada sea la que usará el template
        'mes_display': dict(TodoItem.MES_CHOICES).get(mes_val_padded, mes_val_padded),
        'total_monto_esperado': total_monto_esperado,
        'por_cobrar_monto': por_cobrar_monto,
        'clientes_involucrados': clientes_involucrados,
        'oportunidades': oportunidades_mes, # Pasar todas las oportunidades para listarlas
        'graph_data_json': graph_data_with_display, # Pasa los datos procesados con display_value
        'is_supervisor': is_supervisor(request.user),
    }
    return redirect('/app/todos/')


@login_required
def oportunidades_perdidas_detail(request):
    """
    Vista para mostrar todas las oportunidades con 0% de probabilidad de cierre.
    Considera el rol de supervisor.
    """
    if is_supervisor(request.user):
        oportunidades_perdidas = TodoItem.objects.filter(probabilidad_cierre=0).order_by('-fecha_creacion')
    else:
        oportunidades_perdidas = TodoItem.objects.filter(usuario=request.user, probabilidad_cierre=0).order_by('-fecha_creacion')

    total_perdido_monto = oportunidades_perdidas.aggregate(
        sum_monto=Sum('monto')
    )['sum_monto'] or Decimal('0.00')

    context = {
        'oportunidades': oportunidades_perdidas,
        'titulo': "Oportunidades Perdidas (0% Probabilidad)",
        'total_perdido_monto': total_perdido_monto,
        'is_supervisor': is_supervisor(request.user),
    }
    return redirect('/app/todos/')


@login_required
def exportar_oportunidades_csv(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import date
        OPENPYXL_AVAILABLE = True
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        OPENPYXL_AVAILABLE = False
        from datetime import date
    
    # 1. Get the base queryset with cotizations
    if is_supervisor(request.user):
        items = TodoItem.objects.select_related('usuario', 'cliente').prefetch_related('cotizaciones__detalles').all()
    else:
        _gids = get_usuarios_visibles_ids(request.user)
        if _gids and len(_gids) > 1:
            items = TodoItem.objects.select_related('usuario', 'cliente').prefetch_related('cotizaciones__detalles').filter(usuario_id__in=_gids)
        else:
            items = TodoItem.objects.select_related('usuario', 'cliente').prefetch_related('cotizaciones__detalles').filter(usuario=request.user)

    # 2. Apply filters (supporting multiple values)
    oportunidad_filter = request.GET.get('filterOportunidad', '').strip()
    if oportunidad_filter:
        items = items.filter(oportunidad__icontains=oportunidad_filter)
        
    cliente_filter = request.GET.get('filterCliente', '').strip()
    if cliente_filter:
        # Si es un número, buscar por ID; si no, buscar por nombre
        if cliente_filter.isdigit():
            items = items.filter(cliente__id=int(cliente_filter))
        else:
            items = items.filter(cliente__nombre_empresa__icontains=cliente_filter)
        
    # Filtro de tipo
    tipo_filter = request.GET.get('filterTipo', '').strip()
    if tipo_filter:
        items = items.filter(tipo_negociacion=tipo_filter)
        
    # Filtro de empleado múltiple
    empleado_filter = request.GET.get('empleado', '').strip()
    if empleado_filter:
        empleado_ids = [e.strip() for e in empleado_filter.split(',') if e.strip()]
        if empleado_ids:
            items = items.filter(usuario__id__in=empleado_ids)
            
    # Filtro de mes de cierre múltiple
    mes_cierre_filter = request.GET.get('filterMesCierre', '').strip()
    if mes_cierre_filter:
        meses = [m.strip() for m in mes_cierre_filter.split(',') if m.strip()]
        if meses:
            items = items.filter(mes_cierre__in=meses)
            
    # Filtro de etapa múltiple (estado)
    etapa_filter = request.GET.get('filterEtapa', '').strip()
    if etapa_filter:
        from django.db.models import Q
        etapas = [e.strip() for e in etapa_filter.split(',') if e.strip()]
        if etapas:
            etapa_conditions = Q()
            for etapa in etapas:
                if etapa == 'vigentes':
                    # Excluir cerradas (ganadas, perdidas y pagadas)
                    etapa_conditions |= ~Q(etapa_completa__icontains='ganado') & ~Q(etapa_completa__icontains='perdido') & ~Q(etapa_completa__icontains='pagado')
                elif etapa == 'ganadas':
                    # Solo cerradas ganadas, incluir pagado
                    etapa_conditions |= Q(etapa_completa__icontains='ganado') | Q(etapa_completa__icontains='pagado')
                elif etapa == 'perdidas':
                    # Solo cerradas perdidas
                    etapa_conditions |= Q(etapa_completa__icontains='perdido')
            items = items.filter(etapa_conditions)

    area_filter = request.GET.get('filterArea', '').strip()
    if area_filter:
        items = items.filter(area=area_filter)

    # 3. Apply sorting
    orden_monto = request.GET.get('orden_monto')
    if orden_monto:
        if orden_monto == 'monto_asc':
            items = items.order_by('monto')
        elif orden_monto == 'monto_desc':
            items = items.order_by('-monto')

    orden_probabilidad = request.GET.get('orden_probabilidad')
    if orden_probabilidad:
        if orden_probabilidad == 'prob_asc':
            items = items.order_by('probabilidad_cierre')
        elif orden_probabilidad == 'prob_desc':
            items = items.order_by('-probabilidad_cierre')

    # 4. Create report based on available libraries
    if OPENPYXL_AVAILABLE:
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Oportunidades"
        
        # --- START: Add Report Metadata ---
        bold_font = Font(bold=True)
        
        # Row 1: Report Author
        ws['A1'] = "Reporte generado por:"
        ws['A1'].font = bold_font
        ws['B1'] = request.user.get_full_name() or request.user.username
        
        # Row 2: Generation Date
        ws['A2'] = "Fecha de generación:"
        ws['A2'].font = bold_font
        ws['B2'] = date.today().strftime("%d/%m/%Y")
        
        # Row 3: Applied Filters
        ws['A3'] = "Filtros aplicados:"
        ws['A3'].font = bold_font
        
        filters_applied = []
        
        # Filtro de oportunidad
        if oportunidad_filter:
            filters_applied.append(f"Oportunidad: {oportunidad_filter}")
            
        # Filtro de cliente
        if cliente_filter:
            filters_applied.append(f"Cliente: {cliente_filter}")
            
        # Filtro de tipo
        if tipo_filter:
            tipo_display = "Runrate" if tipo_filter == "runrate" else "Proyecto"
            filters_applied.append(f"Tipo: {tipo_display}")
            
        # Filtro de empleado múltiple
        if empleado_filter:
            empleado_ids = [e.strip() for e in empleado_filter.split(',') if e.strip()]
            empleado_nombres = []
            for emp_id in empleado_ids:
                try:
                    from django.contrib.auth.models import User
                    empleado = User.objects.get(id=emp_id)
                    empleado_nombres.append(empleado.get_full_name() or empleado.username)
                except User.DoesNotExist:
                    empleado_nombres.append(f"ID:{emp_id}")
            if empleado_nombres:
                filters_applied.append(f"Empleado(s): {', '.join(empleado_nombres)}")
                
        # Filtro de mes de cierre múltiple
        if mes_cierre_filter:
            meses = [m.strip() for m in mes_cierre_filter.split(',') if m.strip()]
            mes_nombres = []
            mes_map = {
                '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
                '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
                '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
            }
            for mes in meses:
                mes_nombres.append(mes_map.get(mes, mes))
            if mes_nombres:
                filters_applied.append(f"Mes(es) de Cierre: {', '.join(mes_nombres)}")
                
        # Filtro de etapa múltiple
        if etapa_filter:
            etapas = [e.strip() for e in etapa_filter.split(',') if e.strip()]
            etapa_nombres = []
            etapa_map = {
                'vigentes': 'Solo vigentes',
                'ganadas': 'Solo cerradas ganadas', 
                'perdidas': 'Solo cerradas perdidas'
            }
            for etapa in etapas:
                etapa_nombres.append(etapa_map.get(etapa, etapa))
            if etapa_nombres:
                filters_applied.append(f"Estado(s): {', '.join(etapa_nombres)}")
                
        # Filtro de área
        if area_filter:
            filters_applied.append(f"Área: {area_filter}")
        
        if filters_applied:
            ws['B3'] = "; ".join(filters_applied)
        else:
            ws['B3'] = "Ninguno"
            
        # --- END: Add Report Metadata ---

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define all headers (brands + months)
        all_headers = [
            'OPORTUNIDAD', 'CLIENTE', 'AREA', 'CONTACTO', 'ZEBRA', 'PANDUIT', 'APC', 'AVIGILON', 
            'GENETEC', 'AXIS', 'SOFTWARE', 'RUNRATE', 'PÓLIZA', 'CISCO',
            'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEPT', 'OCT', 'NOV', 'DIC', 'ESTATUS', 'EMPLEADO'
        ]
        
        # Write all headers in a single row (starting at row 5)
        header_row_num = 5
        for col, header in enumerate(all_headers, 1):
            cell = ws.cell(row=header_row_num, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
    else:
        # Fallback to CSV - simplified structure
        headers = [
            'OPORTUNIDAD', 'CLIENTE', 'AREA', 'CONTACTO', 'ZEBRA', 'PANDUIT', 'APC', 'AVIGILON', 
            'GENETEC', 'AXIS', 'SOFTWARE', 'RUNRATE', 'PÓLIZA', 'CISCO',
            'ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEPT', 'OCT', 'NOV', 'DIC', 'ESTATUS', 'EMPLEADO'
        ]
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="reporte_cotizaciones_oportunidades.csv"'
        
        import csv
        writer = csv.writer(response)
        writer.writerow(headers)
        
    # Brand columns mapping - usar las marcas reales de PRODUCTO_CHOICES
    brand_columns = {
        'ZEBRA': 5, 'PANDUIT': 6, 'APC': 7, 'AVIGILON': 8,
        'GENETEC': 9, 'AXIS': 10, 'SOFTWARE': 11, 'RUNRATE': 12, 
        'PÓLIZA': 13, 'CISCO': 14
    }
        
    # Debug: Print item count and sample data
    items_count = items.count()
    print(f"DEBUG: Total items found: {items_count}")
    
    if items_count > 0:
        # Show first item fields for debugging
        first_item = items.first()
        print(f"DEBUG: First item fields:")
        print(f"  - oportunidad: '{first_item.oportunidad}'")
        print(f"  - mes_cierre: '{first_item.mes_cierre}' (type: {type(first_item.mes_cierre)})")
        print(f"  - probabilidad_cierre: {first_item.probabilidad_cierre} (type: {type(first_item.probabilidad_cierre)})")
        print(f"  - monto: {first_item.monto}")
        print(f"  - area: '{first_item.area}'")
        print(f"  - producto: '{first_item.producto}'")
    else:
        print("DEBUG: No items found - check your filters!")
    
    # Write data rows (start at row 6 for Excel since we now have metadata headers)
    row = 6
    for item in items:
        print(f"DEBUG: Processing item: {item.oportunidad}")
        # Get cotization details for this opportunity
        cotizaciones = item.cotizaciones.all()
        
        # Para cada oportunidad, el monto va en la columna del producto/área de la oportunidad
        # No en las cotizaciones, sino en el producto/área de la oportunidad misma
        oportunidad_producto = item.get_producto_display() if hasattr(item, 'get_producto_display') else ''
        oportunidad_monto = float(item.monto) if item.monto else 0
        
        # Mapear el producto de la oportunidad a las marcas disponibles
        brand_totals = {}
        if oportunidad_producto and oportunidad_monto > 0:
            # Detectar marca por el producto de la oportunidad
            producto_upper = oportunidad_producto.upper()
            marca_detectada = None
            
            for brand in brand_columns.keys():
                if brand.upper() in producto_upper:
                    marca_detectada = brand
                    break
            
            # Si no se detecta por nombre, usar mapeo específico
            if not marca_detectada:
                producto_mappings = {
                    'ZEBRA': ['ZEBRA'],
                    'PANDUIT': ['PANDUIT'],
                    'APC': ['APC', 'UPS'],
                    'AVIGILION': ['AVIGILION', 'CCTV', 'CAMARA'],
                    'GENETEC': ['GENETEC', 'SEGURIDAD'],
                    'AXIS': ['AXIS'],
                    'SOFTWARE': ['SOFTWARE', 'APP', 'DESARROLLO'],
                    'RUNRATE': ['RUNRATE', 'RUN RATE'],
                    'PÓLIZA': ['PÓLIZA', 'POLIZA', 'SEGURO'],
                    'CISCO': ['CISCO', 'NETWORKING', 'RED']
                }
                
                for brand, keywords in producto_mappings.items():
                    if any(keyword in producto_upper for keyword in keywords):
                        marca_detectada = brand
                        break
            
            # Si se detectó una marca, asignar el monto
            if marca_detectada:
                brand_totals[marca_detectada] = oportunidad_monto
        
        # Prepare row data
        row_data = [
            item.oportunidad,
            item.cliente.nombre_empresa if item.cliente else '',
            item.get_area_display(),
            str(item.contacto) if item.contacto else ''
        ]
        
        # Add brand amounts - usar las marcas correctas
        for brand in ['ZEBRA', 'PANDUIT', 'APC', 'AVIGILION', 'GENETEC', 'AXIS', 'SOFTWARE', 'RUNRATE', 'PÓLIZA', 'CISCO']:
            amount = brand_totals.get(brand, 0)
            row_data.append(f"${amount:,.2f}" if amount > 0 else '')
        
        # Add monthly data - LA PROBABILIDAD VA EN EL MES DE CIERRE
        months = [''] * 12
        
        # Extraer mes_cierre y probabilidad_cierre correctamente
        mes_cierre_valor = item.mes_cierre  # CharField con valores como '01', '02', etc.
        probabilidad_valor = item.probabilidad_cierre  # IntegerField
        
        print(f"DEBUG: Raw mes_cierre: '{mes_cierre_valor}' (type: {type(mes_cierre_valor)})")
        print(f"DEBUG: Raw probabilidad_cierre: {probabilidad_valor} (type: {type(probabilidad_valor)})")
        
        # Convertir mes_cierre a índice (0-11) para el array de meses
        if mes_cierre_valor and mes_cierre_valor.strip():
            # Crear mapeo de nombres de meses a números
            mes_nombres_a_numeros = {
                'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
                'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
                'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
            }
            
            mes_str = mes_cierre_valor.strip()
            
            # Intentar convertir usando el mapeo de nombres
            if mes_str in mes_nombres_a_numeros:
                mes = mes_nombres_a_numeros[mes_str]
                print(f"DEBUG: mes_cierre '{mes_str}' mapped to int: {mes}")
                
                # Formatear la probabilidad
                if probabilidad_valor is not None:
                    probabilidad_str = f"{probabilidad_valor}%"
                else:
                    probabilidad_str = "0%"  # Default si no hay probabilidad
                
                # Asignar al mes correspondiente (mes-1 porque el array es 0-indexed)
                month_names = ['ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEPT','OCT','NOV','DIC']
                months[mes - 1] = probabilidad_str
                print(f"DEBUG: Setting month {mes} ({month_names[mes-1]}) to {probabilidad_str}")
            else:
                # Intentar como número directo (fallback)
                try:
                    mes = int(mes_str)
                    if 1 <= mes <= 12:
                        if probabilidad_valor is not None:
                            probabilidad_str = f"{probabilidad_valor}%"
                        else:
                            probabilidad_str = "0%"
                        
                        month_names = ['ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEPT','OCT','NOV','DIC']
                        months[mes - 1] = probabilidad_str
                        print(f"DEBUG: Setting month {mes} ({month_names[mes-1]}) to {probabilidad_str}")
                    else:
                        print(f"DEBUG: mes_cierre {mes} is out of range (1-12)")
                except (ValueError, TypeError):
                    print(f"DEBUG: Unable to convert mes_cierre '{mes_str}' - not a recognized month name or number")
        else:
            print(f"DEBUG: mes_cierre is None, empty or whitespace: '{mes_cierre_valor}'")
        
        print(f"DEBUG: Final months array: {months}")
        row_data.extend(months)
        
        # Add estatus (etapa_corta) before empleado
        estatus = getattr(item, 'etapa_corta', '') or ''
        row_data.append(estatus)
        
        row_data.append(item.usuario.get_full_name() or item.usuario.username if item.usuario else '')
        
        if OPENPYXL_AVAILABLE:
            # Excel version
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                # Format currency columns (brand columns are 4-13)
                if col >= 4 and col <= 13 and value and value != '':
                    try:
                        numeric_value = float(value.replace('$', '').replace(',', ''))
                        cell.value = numeric_value
                        cell.number_format = '$#,##0.00'
                    except:
                        pass
        else:
            # CSV version
            writer.writerow(row_data)
        
        row += 1
    
    if OPENPYXL_AVAILABLE:
        # Auto-adjust column widths (14 main headers + 12 month columns + 1 estatus + 1 empleado = 28 total)
        total_columns = 14 + 12 + 1 + 1  # main headers + monthly columns + estatus + empleado
        for col in range(1, total_columns + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 12
            
        # Make first column wider for opportunity names
        ws.column_dimensions['A'].width = 30

        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_cotizaciones_oportunidades.xlsx"'
        
        wb.save(response)
    
    return response


@login_required
@require_http_methods(["POST"])
def editar_oportunidad_api(request, oportunidad_id):
    """
    API para editar información de una oportunidad
    """
    try:
        oportunidad = get_object_or_404(TodoItem, pk=oportunidad_id)
        
        # Verificar permisos - supervisores y compañeros de grupo pueden editar
        if not is_supervisor(request.user) and oportunidad.usuario != request.user:
            from .views_grupos import comparten_grupo
            if not (oportunidad.usuario and comparten_grupo(request.user, oportunidad.usuario)):
                return JsonResponse({'success': False, 'error': 'No tienes permisos para editar esta oportunidad'})
        
        updated_values = {}
        
        # Actualizar nombre (titulo) de la oportunidad
        if 'oportunidad' in request.POST and request.POST['oportunidad']:
            oportunidad.oportunidad = request.POST['oportunidad']
            updated_values['oportunidad'] = oportunidad.oportunidad

        # Actualizar cliente
        if 'cliente' in request.POST and request.POST['cliente']:
            try:
                cliente = Cliente.objects.get(id=request.POST['cliente'])
                oportunidad.cliente = cliente
                updated_values['cliente'] = cliente.nombre_empresa
            except Cliente.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Cliente no encontrado'})
        
        # Actualizar contacto (acepta ID de contacto)
        if 'contacto' in request.POST and request.POST['contacto']:
            try:
                contacto_obj = Contacto.objects.get(id=int(request.POST['contacto']))
                oportunidad.contacto = contacto_obj
                updated_values['contacto'] = f"{contacto_obj.nombre} {contacto_obj.apellido or ''}".strip()
            except (Contacto.DoesNotExist, ValueError):
                pass
        
        # Actualizar área
        if 'area' in request.POST and request.POST['area']:
            oportunidad.area = request.POST['area']
            updated_values['area'] = oportunidad.get_area_display()
        
        # Actualizar tipo de negociación
        if 'tipo_negociacion' in request.POST and request.POST['tipo_negociacion']:
            if request.POST['tipo_negociacion'] in ['runrate', 'proyecto']:
                oportunidad.tipo_negociacion = request.POST['tipo_negociacion']
                updated_values['tipo_negociacion'] = oportunidad.get_tipo_negociacion_display()
            else:
                return JsonResponse({'success': False, 'error': 'Tipo de negociación inválido'})
        
        # Actualizar producto
        if 'producto' in request.POST and request.POST['producto']:
            oportunidad.producto = request.POST['producto']
            updated_values['producto'] = oportunidad.get_producto_display()
        
        # Monto ya no es editable manualmente - se actualiza desde cotizaciones

        # Actualizar probabilidad
        if 'probabilidad' in request.POST:
            try:
                probabilidad = int(request.POST['probabilidad'])
                if 0 <= probabilidad <= 100:
                    oportunidad.probabilidad_cierre = probabilidad
                    updated_values['probabilidad'] = f"{probabilidad}%"
                else:
                    return JsonResponse({'success': False, 'error': 'Probabilidad debe estar entre 0 y 100'})
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'error': 'Probabilidad inválida'})
        
        # Actualizar mes de cierre
        if 'mes_cierre' in request.POST and request.POST['mes_cierre']:
            oportunidad.mes_cierre = request.POST['mes_cierre']
            updated_values['mes_cierre'] = oportunidad.get_mes_cierre_display()

        # Actualizar etapa (desde widget CRM)
        etapa_cambio = None
        if 'etapa_corta' in request.POST and request.POST['etapa_corta']:
            from .bitrix_integration import get_etapa_from_bitrix_stage
            nueva_etapa = request.POST['etapa_corta']
            etapa_cambio = nueva_etapa
            oportunidad.etapa_corta = nueva_etapa
            oportunidad.etapa_completa = nueva_etapa
            updated_values['etapa_corta'] = nueva_etapa

        # Actualizar usuario/vendedor
        if 'usuario' in request.POST and request.POST['usuario']:
            try:
                nuevo_usuario = User.objects.get(id=request.POST['usuario'])
                oportunidad.usuario = nuevo_usuario
                updated_values['usuario'] = nuevo_usuario.get_full_name() or nuevo_usuario.username
            except User.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Usuario no encontrado'})

        propietario_opp = oportunidad.usuario
        oportunidad.save()

        # Registrar en chat de grupo si es edición de oportunidad ajena
        try:
            if propietario_opp and propietario_opp != request.user:
                from .views_grupos import registrar_accion_grupo
                actor_nombre = request.user.get_full_name() or request.user.username
                prop_nombre = propietario_opp.get_full_name() or propietario_opp.username
                nombre_opp = oportunidad.oportunidad or 'oportunidad'
                if etapa_cambio:
                    msg = f'{actor_nombre} cambió la etapa de "{nombre_opp}" (de {prop_nombre}) a {etapa_cambio}'
                    accion = 'cambio_etapa'
                else:
                    msg = f'{actor_nombre} editó la oportunidad "{nombre_opp}" de {prop_nombre}'
                    accion = 'editar_oportunidad'
                registrar_accion_grupo(request.user, propietario_opp, accion, msg,
                                       objeto_tipo='oportunidad', objeto_id=oportunidad.id,
                                       objeto_titulo=nombre_opp)
        except Exception:
            pass

        # Ejecutar automatizaciones si hubo cambio de etapa
        tareas_auto = []
        if etapa_cambio:
            try:
                from .views_automatizacion import ejecutar_automatizaciones
                tareas_auto = ejecutar_automatizaciones(oportunidad, etapa_cambio, request.user)
            except Exception as e_auto:
                print(f'[Automatización] Error ejecutando reglas: {e_auto}')

        # Auto-crear ProyectoIAMET si la etapa es "Levantamiento" y es tipo proyecto
        if etapa_cambio and etapa_cambio == 'Levantamiento' and oportunidad.tipo_negociacion in ('proyecto', 'bitrix_proyecto'):
            try:
                from .models import ProyectoIAMET, ProyectoConfiguracion
                from datetime import timedelta
                # Solo crear si no existe ya un proyecto vinculado a esta oportunidad
                if not ProyectoIAMET.objects.filter(oportunidad=oportunidad).exists():
                    from datetime import date
                    hoy = date.today()
                    proyecto = ProyectoIAMET.objects.create(
                        usuario=oportunidad.usuario,
                        oportunidad=oportunidad,
                        nombre=oportunidad.oportunidad,
                        cliente_nombre=oportunidad.cliente.nombre_empresa if oportunidad.cliente else '',
                        descripcion=oportunidad.comentarios or oportunidad.oportunidad,
                        utilidad_presupuestada=oportunidad.monto or 0,
                        status='active',
                        fecha_inicio=hoy,
                        fecha_fin=hoy + timedelta(days=30),
                    )
                    ProyectoConfiguracion.objects.create(proyecto=proyecto)
                    # Sync existing tasks from the opportunity
                    from .models import ProyectoTarea
                    tareas_opp = oportunidad.tareas_oportunidad.all()
                    for t in tareas_opp:
                        ProyectoTarea.objects.create(
                            proyecto=proyecto,
                            titulo=t.titulo,
                            descripcion=t.descripcion or '',
                            status='completed' if t.estado == 'completada' else 'pending',
                            prioridad='high' if t.prioridad == 'alta' else 'medium',
                            asignado_a=t.responsable,
                            fecha_limite=t.fecha_limite.date() if t.fecha_limite else None,
                        )
            except Exception as e_proy:
                print(f'[ProyectoIAMET] Error auto-creando proyecto: {e_proy}')

        return JsonResponse({
            'success': True,
            'message': 'Oportunidad actualizada correctamente',
            'updated_values': updated_values,
            'tareas_automaticas': tareas_auto,
        })
        
    except Exception as e:
        print(f"Error editando oportunidad: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def oportunidad_detalle_api(request, id):
    """
    Devuelve los datos de una oportunidad en JSON para actualizar la fila tras edición.
    """
    try:
        todo = TodoItem.objects.get(pk=id)
        return JsonResponse({
            'id': todo.id,
            'oportunidad': todo.oportunidad,
            'monto': float(todo.monto),
            'probabilidad_cierre': todo.probabilidad_cierre,
            'cliente': str(todo.cliente) if todo.cliente else '',
            'mes_cierre': str(todo.get_mes_cierre_display()),
            'producto': str(todo.get_producto_display()),
            'area': str(todo.get_area_display()),
            'contacto': str(todo.contacto) if todo.contacto else '',
        })
    except TodoItem.DoesNotExist:
        return JsonResponse({'error': 'Oportunidad no encontrada'}, status=404)


@login_required
def actualizar_probabilidad(request, id):
    """
    API para actualizar la probabilidad_cierre de una oportunidad (TodoItem).
    URL: /api/oportunidad/<id>/probabilidad/
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    try:
        prob = int(request.POST.get('probabilidad', -1))
        if prob < 0 or prob > 100:
            return JsonResponse({'error': 'Probabilidad fuera de rango'}, status=400)
        todo = TodoItem.objects.get(pk=id)
        todo.probabilidad_cierre = prob
        todo.save(update_fields=['probabilidad_cierre'])
        return JsonResponse({'ok': True, 'probabilidad': prob})
    except TodoItem.DoesNotExist:
        return JsonResponse({'error': 'Oportunidad no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def actualizar_po(request, id):
    """API para actualizar los campos PO y Factura de una oportunidad."""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    try:
        todo = TodoItem.objects.get(pk=id)
        update_fields = []
        if 'po_number' in request.POST:
            todo.po_number = request.POST.get('po_number', '').strip()
            update_fields.append('po_number')
        if 'factura_numero' in request.POST:
            todo.factura_numero = request.POST.get('factura_numero', '').strip()
            update_fields.append('factura_numero')
        if update_fields:
            todo.save(update_fields=update_fields)
        return JsonResponse({'ok': True, 'po_number': todo.po_number, 'factura_numero': todo.factura_numero})
    except TodoItem.DoesNotExist:
        return JsonResponse({'error': 'Oportunidad no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@login_required

@login_required
def importar_oportunidades(request):
    clientes = Cliente.objects.all().order_by('nombre_empresa')

    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        tabla_json = request.POST.get('tabla_json')

        if not cliente_id or not tabla_json:
            return JsonResponse({'error': 'Faltan datos obligatorios (cliente o tabla_json)'}, status=400)

        try:
            cliente = Cliente.objects.get(pk=cliente_id)
            oportunidades_data = json.loads(tabla_json)
        except Cliente.DoesNotExist:
            return JsonResponse({'error': 'Cliente no encontrado'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Formato JSON inválido para la tabla'}, status=400)

        created_count = 0
        errors = []

        # Mapeo de nombres de meses a números de dos dígitos
        MONTH_MAPPING = {
            "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
            "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
            "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12"
        }

        # Columnas de producto (deben coincidir con los choices de TodoItem.PRODUCTO_CHOICES)
        PRODUCT_COLUMNS = [
            "zebra", "panduit", "apc", "avigilon", "genetec", "axis",
            "desarrollo_app", "runrate", "poliza", "cisco"
        ]

        for row_data in oportunidades_data:
            try:
                oportunidad_nombre = row_data.get('oportunidad', '')
                area = row_data.get('area', '')
                contacto = row_data.get('contacto', '')

                # Encontrar el producto y el monto/mes de cierre
                producto = None
                monto = Decimal('0.00')
                mes_cierre = None

                # Buscar el producto en las columnas de producto
                for prod_col in PRODUCT_COLUMNS:
                    if row_data.get(prod_col) and row_data.get(prod_col).strip() != '':
                        producto = row_data[prod_col].strip().upper() # Convertir a mayúsculas para coincidir con choices
                        break
                
                # Buscar el monto y mes de cierre en las columnas de meses
                for month_name, month_num in MONTH_MAPPING.items():
                    if row_data.get(month_name) and row_data.get(month_name).strip() != '':
                        try:
                            monto = Decimal(row_data[month_name].strip())
                            mes_cierre = month_num
                            break
                        except (ValueError, TypeError):
                            pass # Ignorar si el monto no es un número válido

                # Validaciones básicas
                if not oportunidad_nombre:
                    errors.append(f"Fila con oportunidad vacía: {row_data}")
                    continue
                if not producto:
                    errors.append(f"Fila '{oportunidad_nombre}': Producto no especificado o inválido.")
                    continue
                if not mes_cierre or monto == Decimal('0.00'):
                    errors.append(f"Fila '{oportunidad_nombre}': Mes de cierre o monto no especificado/inválido.")
                    continue

                # Crear el TodoItem
                TodoItem.objects.create(
                    oportunidad=oportunidad_nombre,
                    cliente=cliente,
                    area=area,
                    contacto=contacto,
                    producto=producto,
                    monto=monto,
                    mes_cierre=mes_cierre,
                    usuario=request.user # Asignar al usuario que importa
                )
                created_count += 1
            except Exception as e:
                errors.append(f"Error procesando fila {row_data}: {e}")

        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        else:
            return JsonResponse({'success': True, 'message': f'{created_count} oportunidades importadas con éxito.'})

    context = {
        'clientes': clientes,
    }
    return render(request, 'importar_oportunidades.html', context)


@csrf_exempt
@login_required
def crear_oportunidad_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'errors': 'Invalid request method'}, status=405)

    cliente_id = request.POST.get('cliente')
    oportunidad_nombre = request.POST.get('oportunidad') # Get opportunity name from POST
    monto = request.POST.get('monto')
    area = request.POST.get('area')
    mes_cierre = request.POST.get('mes_cierre')
    probabilidad_cierre = request.POST.get('probabilidad_cierre')
    
    # Set default values for fields not present in the simplified modal
    producto = request.POST.get('producto', '') # Default to empty string
    comentarios = request.POST.get('comentarios', '') # Default to empty string
    bitrix_stage_id = request.POST.get('bitrix_stage_id', 'UC_YUQKW6') # Default to 'Cotizando'

    if not cliente_id:
        return JsonResponse({'success': False, 'errors': {'cliente': 'ID de cliente es requerido.'}}, status=400)
    if not oportunidad_nombre:
        return JsonResponse({'success': False, 'errors': {'oportunidad': 'Nombre de oportunidad es requerido.'}}, status=400)
    if not monto:
        return JsonResponse({'success': False, 'errors': {'monto': 'Monto es requerido.'}}, status=400)
    if not area:
        return JsonResponse({'success': False, 'errors': {'area': 'Área es requerida.'}}, status=400)
    if not mes_cierre:
        return JsonResponse({'success': False, 'errors': {'mes_cierre': 'Mes de cierre es requerido.'}}, status=400)
    if not probabilidad_cierre:
        return JsonResponse({'success': False, 'errors': {'probabilidad_cierre': 'Probabilidad de cierre es requerida.'}}, status=400)

    try:
        cliente = Cliente.objects.get(id=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({'success': False, 'errors': {'cliente': 'Cliente seleccionado no encontrado.'}}, status=404)

    # Ensure the client has a bitrix_company_id
    if not cliente.bitrix_company_id:
        bitrix_company_id = get_or_create_bitrix_company(cliente.nombre_empresa, request=request)
        if bitrix_company_id:
            cliente.bitrix_company_id = bitrix_company_id
            cliente.save()
        else:
            return JsonResponse({'success': False, 'errors': {'cliente': 'No se pudo obtener o crear el ID de compañía de Bitrix para el cliente.'}}, status=400)
    
    # Create a dictionary for the TodoItem data
    todo_item_data = {
        'oportunidad': oportunidad_nombre,
        'monto': float(monto),
        'cliente': cliente,
        'usuario': request.user, # Assign the current user
        'area': area,
        'mes_cierre': mes_cierre,
        'anio_cierre': timezone.localdate().year,
        'probabilidad_cierre': int(probabilidad_cierre),
        'producto': producto, # Use default or provided
        'comentarios': comentarios, # Use default or provided
        'bitrix_stage_id': bitrix_stage_id, # Use default or provided
        'po_number': '', # Ensure PO is empty on creation
    }

    try:
        venta = TodoItem.objects.create(**todo_item_data)
        
        opportunity_data = {
            'oportunidad': venta.oportunidad,
            'monto': float(venta.monto),
            'cliente': cliente.nombre_empresa,
            'bitrix_company_id': cliente.bitrix_company_id,
            'producto': venta.producto,
            'area': venta.area,
            'mes_cierre': venta.mes_cierre,
            'probabilidad_cierre': venta.probabilidad_cierre,
            'comentarios': venta.comentarios,
            'bitrix_stage_id': venta.bitrix_stage_id,
            'bitrix_contact_id': None, # No contact from simplified form
        }
        
        bitrix_assigned_by_id = None
        if venta.usuario and hasattr(venta.usuario, 'userprofile') and venta.usuario.userprofile.bitrix_user_id:
            bitrix_assigned_by_id = venta.usuario.userprofile.bitrix_user_id
        opportunity_data['bitrix_assigned_by_id'] = bitrix_assigned_by_id
        
        bitrix_response = send_opportunity_to_bitrix(opportunity_data, request=request)
        
        if bitrix_response and bitrix_response.get('result'):
            venta.bitrix_deal_id = bitrix_response.get('result')
            venta.save(update_fields=['bitrix_deal_id'])
            
        return JsonResponse({
            'success': True,
            'oportunidad': {
                'id': venta.id,
                'nombre': venta.oportunidad,
            }
        })

    except Exception as e:
        print(f"ERROR: Falló la sincronización con Bitrix24 para la oportunidad {oportunidad_nombre}: {e}")
        return JsonResponse({
            'success': True, # Still return success for local creation
            'oportunidad': {
                'id': venta.id if 'venta' in locals() else None, # Return ID if created locally
                'nombre': oportunidad_nombre,
            },
            'warning': f'Oportunidad creada localmente, pero falló la sincronización con Bitrix24: {e}'
        })


@login_required
def check_new_local_opportunities(request):
    """
    API endpoint para detectar nuevas oportunidades creadas en nuestro sistema.
    Esta función verifica si hay oportunidades creadas después del último timestamp de verificación.
    """
    try:
        # Obtener timestamp de la última verificación desde el parámetro GET
        last_check_timestamp = request.GET.get('last_check')
        
        if not last_check_timestamp:
            return JsonResponse({
                'success': False,
                'error': 'Falta el parámetro last_check'
            })
        
        # Convertir timestamp a datetime
        from datetime import datetime
        last_check = datetime.fromtimestamp(int(last_check_timestamp) / 1000, tz=timezone.utc)
        
        # Primero verificar si hay una alerta inmediata en la sesión (Crown Jewel Feature)
        opportunities_data = []
        session_key = f'new_opportunity_alert_{request.user.id}'
        
        if session_key in request.session:
            # Hay una oportunidad que se acaba de crear - procesarla inmediatamente
            alert_data = request.session.pop(session_key)  # Remover después de leer
            opportunities_data.append(alert_data)
            print(f"DEBUG: Detectada oportunidad inmediata desde sesión: {alert_data}")
        
        # También buscar oportunidades nuevas creadas después del último check
        # Solo buscar las del usuario actual o todas si es supervisor
        if is_supervisor(request.user):
            new_opportunities = TodoItem.objects.filter(
                created_at__gt=last_check
            ).select_related('cliente').order_by('-created_at')[:5]
        else:
            new_opportunities = TodoItem.objects.filter(
                user=request.user,
                created_at__gt=last_check
            ).select_related('cliente').order_by('-created_at')[:5]
        
        # Agregar oportunidades de base de datos a la lista existente
        for opp in new_opportunities:
            opportunities_data.append({
                'id': opp.id,
                'titulo': opp.oportunidad,
                'cliente_id': opp.cliente.id if opp.cliente else None,
                'cliente_nombre': opp.cliente.nombre_empresa if opp.cliente else 'Sin cliente',
                'monto_estimado': str(opp.precio_estimado) if opp.precio_estimado else 'N/A',
                'probabilidad': opp.probabilidad_exito,
                'created_at': opp.created_at.isoformat(),
                'user': opp.usuario.username if opp.usuario else 'Sin usuario'
            })
        
        return JsonResponse({
            'success': True,
            'new_opportunities': opportunities_data,
            'count': len(opportunities_data),
            'last_check': last_check.isoformat()
        })
        
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en el formato del timestamp: {e}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error interno del servidor: {e}'
        })


@login_required
def check_new_bitrix_opportunities(request):
    """
    API endpoint para detectar nuevas oportunidades directamente desde Bitrix24.
    Esta función consulta la API de Bitrix para detectar oportunidades creadas recientemente.
    """
    from .bitrix_integration import get_all_bitrix_deals
    from datetime import datetime, timedelta
    from django.utils import timezone, translation
    from django.utils.translation import gettext_lazy as _
    from django.http import JsonResponse
    from django.utils.translation import activate, get_language

    try:
        # Obtener timestamp de la última verificación desde el parámetro GET
        last_check_timestamp = request.GET.get('last_check')
        if not last_check_timestamp:
            return JsonResponse({
                'success': False,
                'error': 'Falta el parámetro last_check'
            })
        
        # Convertir timestamp a datetime
        last_check = datetime.fromtimestamp(int(last_check_timestamp) / 1000, tz=django_timezone.utc)
        
        # Obtener el ID de usuario de Bitrix24 para este usuario de Django
        try:
            user_profile = request.user.userprofile
            user_bitrix_id = str(user_profile.bitrix_user_id) if user_profile.bitrix_user_id else None
        except:
            user_bitrix_id = None
            
        if not user_bitrix_id:
            return JsonResponse({
                'success': True,
                'new_opportunities': [],
                'count': 0,
                'message': 'Usuario no tiene ID de Bitrix24 configurado'
            })
        
        # Consultar todas las oportunidades desde Bitrix24
        bitrix_deals = get_all_bitrix_deals(request)
        
        if not bitrix_deals:
            return JsonResponse({
                'success': True,
                'new_opportunities': [],
                'count': 0,
                'message': 'No se pudieron obtener oportunidades de Bitrix24'
            })
        
        # Filtrar solo oportunidades asignadas a este usuario en Bitrix24
        user_deals = [deal for deal in bitrix_deals if deal.get('ASSIGNED_BY_ID') == user_bitrix_id]
        
        print(f"DEBUG Bot: Usuario Django {request.user.username} → Bitrix ID {user_bitrix_id}")
        print(f"DEBUG Bot: Encontradas {len(user_deals)} oportunidades para este usuario de {len(bitrix_deals)} totales")
        
        # Filtrar oportunidades nuevas (que no existen en nuestro sistema)
        recent_deals = []
        
        for deal in user_deals[:10]:  # Solo las 10 más recientes del usuario
            # Verificar si esta oportunidad ya existe en nuestro sistema
            deal_id = deal.get('ID')
            if deal_id:
                existing_opportunity = TodoItem.objects.filter(
                    bitrix_deal_id=deal_id
                ).first()
                
                if not existing_opportunity:
                    # Esta es una nueva oportunidad que no tenemos en nuestro sistema
                    # Obtener datos de la compañía si existe
                    company_name = 'Cliente por definir'
                    if deal.get('COMPANY_ID'):
                        try:
                            from .bitrix_integration import get_bitrix_company_details
                            company_data = get_bitrix_company_details(deal.get('COMPANY_ID'), request)
                            if company_data and company_data.get('TITLE'):
                                company_name = company_data.get('TITLE')
                        except Exception as e:
                            print(f"Error obteniendo datos de compañía: {e}")
                    
                    recent_deals.append({
                        'id': deal_id,
                        'bitrix_id': deal_id,
                        'titulo': deal.get('TITLE', 'Sin título'),
                        'monto_estimado': deal.get('OPPORTUNITY', '0'),
                        'company_id': deal.get('COMPANY_ID'),
                        'company_name': company_name,
                        'contact_id': deal.get('CONTACT_ID'),
                        'stage_id': deal.get('STAGE_ID'),
                        'comentarios': deal.get('COMMENTS', ''),
                        'assigned_by_id': deal.get('ASSIGNED_BY_ID'),
                        # Mapear campos personalizados
                        'producto_bitrix_id': deal.get('UF_CRM_1752859685662'),
                        'area_bitrix_id': deal.get('UF_CRM_1752859525038'),
                        'mes_cierre_bitrix_id': deal.get('UF_CRM_1752859877756'),
                        'probabilidad_bitrix_id': deal.get('UF_CRM_1752855787179'),
                        'is_from_bitrix': True,
                        'detected_at': django_timezone.now().isoformat()
                    })
        
        print(f"DEBUG Bot: Encontradas {len(recent_deals)} nuevas oportunidades desde Bitrix24")
        
        return JsonResponse({
            'success': True,
            'new_opportunities': recent_deals,
            'count': len(recent_deals),
            'last_check': django_timezone.now().timestamp() * 1000  # Nuevo timestamp
        })
        
    except Exception as e:
        print(f"ERROR Bot: Error al verificar oportunidades desde Bitrix24: {e}")
        return JsonResponse({
            'success': False,
            'error': f'Error al verificar oportunidades desde Bitrix24: {str(e)}'
        }, status=500)


@login_required
def nueva_oportunidad(request):
    """
    Vista optimizada para crear nuevas oportunidades con mejor UX y automatización.
    """
    if request.method == 'POST':
        form = NuevaOportunidadForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                oportunidad = form.save()
                messages.success(request, f'Oportunidad "{oportunidad.oportunidad}" creada exitosamente.')
                return redirect('todos')  # Redirigir a la lista de oportunidades
            except Exception as e:
                messages.error(request, f'Error al crear la oportunidad: {str(e)}')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = NuevaOportunidadForm(user=request.user)
    
    # Obtener lista de clientes para autocompletado
    clientes = Cliente.objects.all().order_by('nombre_empresa')
    
    context = {
        'form': form,
        'clientes': clientes,
        'title': 'Nueva Oportunidad'
    }
    
    return render(request, 'nueva_oportunidad.html', context)


@login_required
def api_crear_oportunidad(request):
    """
    API AJAX para crear oportunidad desde el widget CRM.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)

    try:
        import json
        data = json.loads(request.body) if request.content_type == 'application/json' else request.POST

        cliente_nombre = data.get('cliente_nombre', '').strip()
        if not cliente_nombre or len(cliente_nombre) < 2:
            return JsonResponse({'ok': False, 'error': 'El nombre del cliente es requerido (mín. 2 caracteres).'})

        oportunidad_nombre = data.get('oportunidad', '').strip()
        if not oportunidad_nombre:
            return JsonResponse({'ok': False, 'error': 'El nombre de la oportunidad es requerido.'})

        monto = data.get('monto', 0)
        try:
            monto = Decimal(str(monto))
        except Exception:
            monto = Decimal('0')

        # Buscar o crear cliente (filter evita error si hay duplicados por case)
        cliente = Cliente.objects.filter(nombre_empresa__iexact=cliente_nombre).order_by('id').first()
        if not cliente:
            cliente = Cliente.objects.create(nombre_empresa=cliente_nombre, asignado_a=request.user)

        # Buscar o crear contacto
        contacto = None
        contacto_nombre = data.get('contacto_nombre', '').strip()
        if contacto_nombre:
            nombre_parts = contacto_nombre.split(' ', 1)
            contacto, _ = Contacto.objects.get_or_create(
                nombre__iexact=nombre_parts[0],
                cliente=cliente,
                defaults={
                    'nombre': nombre_parts[0],
                    'apellido': nombre_parts[1] if len(nombre_parts) > 1 else '',
                    'cliente': cliente
                }
            )

        tipo_neg = data.get('tipo_negociacion', 'runrate')
        # Asignar etapa inicial desde la BD (primera etapa activa del pipeline)
        primera_etapa = EtapaPipeline.objects.filter(pipeline=tipo_neg, activo=True).order_by('orden').first()
        if primera_etapa:
            etapa_corta_init = primera_etapa.nombre
            etapa_completa_init = primera_etapa.nombre
            etapa_color_init = primera_etapa.color
        elif tipo_neg == 'proyecto':
            etapa_corta_init = 'Oportunidad'
            etapa_completa_init = 'Oportunidad'
            etapa_color_init = '#FFFFFF'
        else:
            etapa_corta_init = 'En Solicitud'
            etapa_completa_init = 'Solicitud de Cotizacion'
            etapa_color_init = '#FFFFFF'

        from datetime import datetime as dt_create
        now_dt = dt_create.now()
        mes_actual = str(now_dt.month).zfill(2)

        raw_mes = data.get('mes_cierre', '').strip()
        if not raw_mes or raw_mes == 'todos':
            raw_mes = mes_actual
        mes_cierre_val = raw_mes

        todo = TodoItem(
            usuario=request.user,
            oportunidad=oportunidad_nombre,
            cliente=cliente,
            contacto=contacto,
            monto=monto,
            probabilidad_cierre=int(data.get('probabilidad_cierre', 25)),
            mes_cierre=mes_cierre_val,
            anio_cierre=now_dt.year,
            area=data.get('area', 'SISTEMAS'),
            producto=data.get('producto', 'SOFTWARE'),
            tipo_negociacion=tipo_neg,
            comentarios=data.get('comentarios', ''),
            etapa_corta=etapa_corta_init,
            etapa_completa=etapa_completa_init,
            etapa_color=etapa_color_init,
            po_number='', # Ensure PO is empty on creation
        )
        todo.save()

        # Ejecutar automatizaciones para la etapa inicial
        try:
            from .views_automatizacion import ejecutar_automatizaciones
            ejecutar_automatizaciones(todo, etapa_corta_init, request.user)
        except Exception:
            pass  # No bloquear la creación si falla la automatización

        # If there are comments, add them as a chat message
        if todo.comentarios:
            from .models import MensajeOportunidad
            MensajeOportunidad.objects.create(
                oportunidad=todo,
                usuario=request.user,
                texto=todo.comentarios
            )

        # Notificar al chat de grupo si el cliente está asignado a otro miembro
        try:
            if cliente and hasattr(cliente, 'asignado_a') and cliente.asignado_a and cliente.asignado_a != request.user:
                from .views_grupos import registrar_accion_grupo
                actor_nombre = request.user.get_full_name() or request.user.username
                registrar_accion_grupo(
                    request.user, cliente.asignado_a, 'crear_oportunidad',
                    f'{actor_nombre} creó la oportunidad "{oportunidad_nombre}" en el cliente {cliente.nombre_empresa}',
                    objeto_tipo='oportunidad', objeto_id=todo.id, objeto_titulo=oportunidad_nombre,
                )
        except Exception:
            pass

        return JsonResponse({
            'ok': True,
            'message': f'Oportunidad "{oportunidad_nombre}" creada exitosamente.',
            'id': todo.id
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required
def api_oportunidad_detalle_crm(request, oportunidad_id):
    """
    API para obtener detalle completo de una oportunidad para el widget CRM.
    """
    try:
        todo = get_object_or_404(TodoItem, pk=oportunidad_id)

        # Obtener cotizaciones de esta oportunidad
        cotizaciones = Cotizacion.objects.filter(oportunidad=todo).order_by('-fecha_creacion')
        cots_list = []
        for cot in cotizaciones:
            cots_list.append({
                'id': cot.id,
                'titulo': cot.titulo or f'COT-{cot.id}',
                'fecha': cot.fecha_creacion.strftime('%d %b %Y') if cot.fecha_creacion else '',
                'total': float(cot.total) if cot.total else 0,
                'moneda': cot.moneda or 'MXN',
            })

        data = {
            'id': todo.id,
            'oportunidad': todo.oportunidad or '',
            'monto': float(todo.monto) if todo.monto else 0,
            'cliente': {
                'id': todo.cliente_id,
                'nombre': todo.cliente.nombre_empresa if todo.cliente else '',
            } if todo.cliente else None,
            'contacto': '',
            'contacto_id': todo.contacto_id,
            'producto': todo.producto or '',
            'area': todo.area or '',
            'probabilidad_cierre': todo.probabilidad_cierre or 0,
            'po_number': todo.po_number or '',
            'factura_numero': todo.factura_numero or '',
            'mes_cierre': todo.mes_cierre or '',
            'tipo_negociacion': todo.tipo_negociacion or 'runrate',
            'etapa_corta': todo.etapa_corta or '',
            'etapa_completa': todo.etapa_completa or '',
            'etapa_color': todo.etapa_color or '#FFFFFF',
            'usuario': todo.usuario.get_full_name() or todo.usuario.username if todo.usuario else '',
            'usuario_id': todo.usuario_id,
            'comentarios': todo.comentarios or '',
            'fecha_creacion': todo.fecha_creacion.strftime('%d/%m/%Y') if todo.fecha_creacion else '',
            'cotizaciones': cots_list,
            'productos_adicionales': [
                {'id': p.id, 'producto': p.producto, 'notas': p.notas}
                for p in todo.productos_adicionales.all()
            ],
        }

        # Contacto
        if todo.contacto:
            if hasattr(todo.contacto, 'nombre'):
                data['contacto'] = f"{todo.contacto.nombre} {todo.contacto.apellido or ''}".strip()
            else:
                data['contacto'] = str(todo.contacto)

        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
def api_oportunidad_productos(request, oportunidad_id):
    """Lista y agrega productos adicionales a una oportunidad."""
    todo = get_object_or_404(TodoItem, pk=oportunidad_id)
    if request.method == 'GET':
        prods = [{'id': p.id, 'producto': p.producto, 'notas': p.notas}
                 for p in todo.productos_adicionales.all()]
        return JsonResponse({'productos': prods})
    elif request.method == 'POST':
        data = json.loads(request.body)
        producto = data.get('producto', '').strip()
        if not producto:
            return JsonResponse({'error': 'Producto requerido.'}, status=400)
        p = ProductoOportunidad.objects.create(
            oportunidad=todo,
            producto=producto,
            notas=data.get('notas', ''),
        )
        return JsonResponse({'id': p.id, 'producto': p.producto, 'notas': p.notas}, status=201)
    return JsonResponse({'error': 'Método no permitido.'}, status=405)


@login_required
@csrf_exempt
def api_oportunidad_producto_delete(request, oportunidad_id, producto_id):
    """Elimina un producto adicional de una oportunidad."""
    p = get_object_or_404(ProductoOportunidad, pk=producto_id, oportunidad_id=oportunidad_id)
    if request.method == 'DELETE':
        p.delete()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'Método no permitido.'}, status=405)


@login_required
def api_buscar_clientes(request):
    """
    API para autocompletado de clientes en el formulario de nueva oportunidad.
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'clientes': []})
    
    # Buscar clientes que coincidan con el query (incluye grupo)
    from .views_grupos import get_clientes_visibles_q
    clientes = Cliente.objects.filter(
        Q(nombre_empresa__icontains=query) & get_clientes_visibles_q(request.user)
    ).order_by('nombre_empresa')[:10]
    
    clientes_data = []
    for cliente in clientes:
        clientes_data.append({
            'id': cliente.id,
            'nombre': cliente.nombre_empresa,
            'contacto_principal': cliente.contacto_principal or '',
            'email': cliente.email or '',
            'telefono': cliente.telefono or ''
        })
    
    return JsonResponse({'clientes': clientes_data})


@login_required  
def api_buscar_contactos(request):
    """
    API para autocompletado de contactos basado en el cliente seleccionado.
    """
    cliente_id = request.GET.get('cliente_id')
    query = request.GET.get('q', '').strip()
    
    if not cliente_id:
        return JsonResponse({'contactos': []})
    
    try:
        contactos = Contacto.objects.filter(cliente_id=cliente_id)
        
        if query:
            contactos = contactos.filter(
                Q(nombre__icontains=query) | Q(apellido__icontains=query)
            )
        
        contactos = contactos.order_by('nombre')[:10]
        
        contactos_data = []
        for contacto in contactos:
            contactos_data.append({
                'id': contacto.id,
                'nombre_completo': f"{contacto.nombre} {contacto.apellido or ''}".strip(),
                'nombre': contacto.nombre,
                'apellido': contacto.apellido or ''
            })
        
        return JsonResponse({'contactos': contactos_data})
        
    except Exception as e:
        return JsonResponse({'contactos': [], 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def cambiar_estado_oportunidad(request, oportunidad_id):
    """
    API para cambiar el estado CRM de una oportunidad
    """
    
    try:
        import json
        data = json.loads(request.body)
        estado = data.get('estado')
        
        if not estado:
            return JsonResponse({'error': 'Estado requerido'}, status=400)
        
        # Obtener la oportunidad
        oportunidad = get_object_or_404(TodoItem, id=oportunidad_id)
        
        # Validar que el usuario puede modificar esta oportunidad
        if not is_supervisor(request.user) and oportunidad.usuario != request.user:
            return JsonResponse({'error': 'No tienes permisos para modificar esta oportunidad'}, status=403)
        
        # Guardar estado anterior para actividad
        estado_anterior = oportunidad.estado_crm
        
        # Actualizar estado
        oportunidad.estado_crm = estado
        oportunidad.save()
        
        # Crear actividad en el timeline
        actividad = OportunidadActividad.objects.create(
            oportunidad=oportunidad,
            tipo='cambio_estado',
            titulo='Cambio de Estado',
            descripcion=f'Estado cambiado de "{estado_anterior}" a "{estado}"',
            usuario=request.user,
            estado_anterior=estado_anterior,
            estado_nuevo=estado
        )
        
        print(f"🔄 Actividad creada: {actividad.id}, estado_anterior: {actividad.estado_anterior}, estado_nuevo: {actividad.estado_nuevo}")
        
        # Preparar datos del timeline item para el frontend
        usuario_nombre = request.user.get_full_name() or request.user.username
        timeline_item = {
            'id': actividad.id,
            'tipo': 'cambio_estado',
            'titulo': 'Cambio de Estado',
            'descripcion': f'Estado cambiado de "{estado_anterior}" a "{estado}"',
            'usuario': usuario_nombre,
            'fecha': convert_to_tijuana_time(actividad.fecha_creacion).strftime('%d/%m/%Y %H:%M'),
            'icono': '🔄 Cambio de Estado',
            'estado_anterior': estado_anterior,
            'estado_nuevo': estado
        }
        
        return JsonResponse({
            'success': True,
            'nuevo_estado': estado,
            'message': f'Estado actualizado a {estado}',
            'timeline_item': timeline_item
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def agregar_comentario_oportunidad(request, oportunidad_id):
    """
    API para agregar comentarios con archivos a una oportunidad
    """
    try:
        # Obtener contenido del comentario
        contenido = request.POST.get('contenido', '').strip()
        
        # Obtener la oportunidad
        oportunidad = get_object_or_404(TodoItem, id=oportunidad_id)
        
        # Verificar permisos: supervisores ven todo, usuarios solo sus propias oportunidades
        if not is_supervisor(request.user) and oportunidad.usuario != request.user:
            return JsonResponse({'error': 'No tienes permisos para comentar en esta oportunidad'}, status=403)
        
        # Verificar que hay contenido o archivos
        archivos_subidos = []
        archivos_keys = [key for key in request.FILES.keys() if key.startswith('archivo_')]
        
        if not contenido and not archivos_keys:
            return JsonResponse({'error': 'Debe proporcionar contenido o archivos'}, status=400)
        
        # Crear comentario (puede estar vacío si solo hay archivos)
        comentario = OportunidadComentario.objects.create(
            oportunidad=oportunidad,
            usuario=request.user,
            contenido=contenido or "Archivo adjunto"
        )
        
        # Procesar archivos adjuntos
        print(f"📁 Procesando {len(archivos_keys)} archivos: {archivos_keys}")
        for key in archivos_keys:
            archivo = request.FILES[key]
            print(f"📄 Procesando archivo: {archivo.name}, tamaño: {archivo.size}, tipo: {archivo.content_type}")
            
            # Determinar tipo de archivo
            content_type = archivo.content_type.lower()
            if content_type.startswith('image/'):
                tipo_archivo = 'imagen'
            elif content_type in ['application/pdf']:
                tipo_archivo = 'documento'
            elif content_type in ['application/msword', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document']:
                tipo_archivo = 'documento'
            elif content_type in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'text/csv']:
                tipo_archivo = 'documento'
            else:
                tipo_archivo = 'otro'
            
            print(f"📋 Tipo determinado: {tipo_archivo}")
            
            try:
                # Crear registro de archivo
                archivo_obj = OportunidadArchivo.objects.create(
                    oportunidad=oportunidad,
                    usuario=request.user,
                    archivo=archivo,
                    nombre_original=archivo.name,
                    tipo=tipo_archivo,
                    tamaño=archivo.size,
                    descripcion=f"Adjuntado en comentario #{comentario.id}"
                )
                
                print(f"✅ Archivo guardado exitosamente: ID={archivo_obj.id}, URL={archivo_obj.archivo.url}")
                
                archivos_subidos.append({
                    'id': archivo_obj.id,
                    'nombre': archivo_obj.nombre_original,
                    'tipo': archivo_obj.tipo,
                    'tamaño': archivo_obj.tamaño,
                    'url': archivo_obj.archivo.url if archivo_obj.archivo else None
                })
                
            except Exception as e:
                print(f"❌ Error guardando archivo {archivo.name}: {e}")
                import traceback
                print(traceback.format_exc())
        
        # Crear actividad en el timeline con referencia al comentario
        descripcion_actividad = contenido[:200] + ('...' if len(contenido) > 200 else '')
        if archivos_subidos:
            if contenido:
                descripcion_actividad += f" ({len(archivos_subidos)} archivo{'s' if len(archivos_subidos) > 1 else ''} adjunto{'s' if len(archivos_subidos) > 1 else ''})"
            else:
                descripcion_actividad = f"Subió {len(archivos_subidos)} archivo{'s' if len(archivos_subidos) > 1 else ''}"
        
        # Agregar referencia al comentario en la descripción para linking directo
        descripcion_actividad += f" [COMENTARIO_ID:{comentario.id}]"
        
        print(f"🔥 Creando actividad - Usuario: {request.user}, Usuario ID: {request.user.id}, Nombre: {request.user.get_full_name()}")
        actividad_creada = OportunidadActividad.objects.create(
            oportunidad=oportunidad,
            tipo='comentario',
            titulo='Nuevo Comentario' + (' con archivos' if archivos_subidos else ''),
            descripcion=descripcion_actividad,
            usuario=request.user
        )
        print(f"💬 Actividad creada: ID={actividad_creada.id}, Usuario={actividad_creada.usuario}, Descripcion='{actividad_creada.descripcion}'")
        
        # ======================================
        # CREAR NOTIFICACIONES AUTOMÁTICAMENTE
        # ======================================
        
        # 1. Detectar menciones @usuario en el comentario
        if contenido:
            detectar_menciones_en_comentario(contenido, request.user, oportunidad, comentario)
        
        # 2. Notificar al dueño de la oportunidad (si no es el mismo que comenta)
        if oportunidad.usuario != request.user:
            mensaje_notif = f'{request.user.get_full_name() or request.user.username} comentó en tu oportunidad "{oportunidad.oportunidad}"'
            if contenido:
                mensaje_notif += f': {contenido[:100]}...' if len(contenido) > 100 else f': {contenido}'
            else:
                mensaje_notif += ' y adjuntó archivos'
                
            crear_notificacion(
                usuario_destinatario=oportunidad.usuario,
                tipo='comentario_oportunidad',
                titulo='Nuevo comentario en tu oportunidad',
                mensaje=mensaje_notif,
                oportunidad=oportunidad,
                comentario=comentario,
                usuario_remitente=request.user
            )
        
        # 3. Notificar a otros usuarios que han comentado en esta oportunidad (excepto el autor actual y el dueño)
        otros_comentaristas = User.objects.filter(
            oportunidadcomentario__oportunidad=oportunidad
        ).exclude(
            id__in=[request.user.id, oportunidad.usuario.id]
        ).distinct()
        
        for usuario in otros_comentaristas:
            mensaje_notif = f'{request.user.get_full_name() or request.user.username} también comentó en la oportunidad "{oportunidad.oportunidad}"'
            if contenido:
                mensaje_notif += f': {contenido[:100]}...' if len(contenido) > 100 else f': {contenido}'
            
            crear_notificacion(
                usuario_destinatario=usuario,
                tipo='comentario_oportunidad',
                titulo='Nuevo comentario en oportunidad que sigues',
                mensaje=mensaje_notif,
                oportunidad=oportunidad,
                comentario=comentario,
                usuario_remitente=request.user
            )
        
        return JsonResponse({
            'success': True,
            'comentario': {
                'id': comentario.id,
                'contenido': comentario.contenido,
                'usuario': request.user.get_full_name() or request.user.username,
                'fecha': convert_to_tijuana_time(comentario.fecha_creacion).strftime('%d/%m/%Y %H:%M'),
                'archivos': archivos_subidos
            }
        })
        
    except Exception as e:
        import traceback
        print(f"Error en agregar_comentario_oportunidad: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def timeline_oportunidad(request, oportunidad_id):
    """
    API para obtener el timeline completo de una oportunidad
    """
    from datetime import timedelta
    
    oportunidad = get_object_or_404(TodoItem, id=oportunidad_id)
    
    # Verificar permisos: supervisores ven todo, usuarios solo sus propias oportunidades
    if not is_supervisor(request.user) and oportunidad.usuario != request.user:
        return JsonResponse({'error': 'No tienes permisos para ver este timeline'}, status=403)
    
    try:
        # Limpiar actividades huérfanas antes de generar el timeline
        try:
            limpiar_actividades_huerfanas(oportunidad)
        except Exception as e:
            print(f"Error limpiando actividades huérfanas: {e}")
            # Continuar sin limpiar si hay error
        
        # Obtener todas las actividades
        actividades = oportunidad.actividades_crm.all().order_by('-fecha_creacion')
        
        timeline_data = []
        for actividad in actividades:
            # Obtener información del usuario
            usuario_nombre = 'Sistema'
            if actividad.usuario:
                usuario_nombre = actividad.usuario.get_full_name() or actividad.usuario.username
            else:
                usuario_nombre = 'Sistema'
            
            # Convertir fecha a zona horaria de Tijuana (Pacific Time)
            fecha_tijuana = convert_to_tijuana_time(actividad.fecha_creacion)
            
            item_data = {
                'id': actividad.id,
                'tipo': actividad.tipo,
                'titulo': actividad.titulo,
                'descripcion': actividad.descripcion,
                'usuario': usuario_nombre,
                'fecha': fecha_tijuana.strftime('%d/%m/%Y %H:%M'),
                'icono': dict(OportunidadActividad.TIPO_ACTIVIDAD_CHOICES).get(actividad.tipo, '⚙️')
            }
            
            # Agregar campos específicos según el tipo de actividad
            if actividad.tipo == 'cambio_estado':
                item_data['estado_anterior'] = actividad.estado_anterior
                item_data['estado_nuevo'] = actividad.estado_nuevo
            elif actividad.tipo == 'comentario':
                # Para comentarios, buscar el contenido real del comentario
                try:
                    # Nueva estrategia: buscar por ID directo en la descripción
                    comentario = None
                    
                    # Estrategia 1: Buscar por ID directo en la descripción
                    import re
                    match = re.search(r'\[COMENTARIO_ID:(\d+)\]', actividad.descripcion or '')
                    if match:
                        comentario_id = int(match.group(1))
                        try:
                            comentario = OportunidadComentario.objects.get(id=comentario_id)
                        except OportunidadComentario.DoesNotExist:
                            comentario = None
                    
                    # Estrategia 2 (fallback): Buscar por rango de tiempo
                    if not comentario:
                        comentarios_candidatos = OportunidadComentario.objects.filter(
                            oportunidad=oportunidad,
                            fecha_creacion__gte=actividad.fecha_creacion - timedelta(minutes=1),
                            fecha_creacion__lte=actividad.fecha_creacion + timedelta(minutes=1)
                        ).order_by('-fecha_creacion')
                        
                        if comentarios_candidatos.exists():
                            comentario = comentarios_candidatos.first()
                    
                    # Estrategia 3 (último recurso): Buscar por usuario y fecha cercana
                    if not comentario and actividad.usuario:
                        comentarios_por_usuario = OportunidadComentario.objects.filter(
                            oportunidad=oportunidad,
                            usuario=actividad.usuario
                        ).order_by('-fecha_creacion')
                        
                        for c in comentarios_por_usuario:
                            diff = abs((c.fecha_creacion - actividad.fecha_creacion).total_seconds())
                            if diff <= 300:  # 5 minutos
                                comentario = c
                                break
                    
                    if comentario:
                        item_data['contenido'] = comentario.contenido
                        item_data['comentario_id'] = comentario.id
                        
                        # Limpiar la descripción para mostrar solo el contenido real (sin el ID)
                        descripcion_limpia = re.sub(r' \[COMENTARIO_ID:\d+\]', '', actividad.descripcion or '')
                        item_data['descripcion'] = descripcion_limpia
                        item_data['puede_editar'] = comentario.usuario == request.user or is_supervisor(request.user)
                        
                        # Buscar archivos asociados a este comentario específico
                        # Usar la descripción del archivo que contiene el ID del comentario
                        archivos_asociados = OportunidadArchivo.objects.filter(
                            oportunidad=oportunidad,
                            descripcion__contains=f"Adjuntado en comentario #{comentario.id}"
                        )
                        
                        # Buscar archivos asociados específicamente por descripción
                        
                        # Si no encuentra por descripción exacta, NO usar fallback para evitar contaminación cruzada
                        # (El fallback por tiempo era lo que causaba que todos los archivos aparecieran en todos los comentarios)
                        
                        if archivos_asociados.exists():
                            item_data['archivos'] = []
                            for archivo in archivos_asociados:
                                item_data['archivos'].append({
                                    'id': archivo.id,
                                    'nombre': archivo.nombre_original,
                                    'tipo': archivo.tipo,
                                    'tamaño': archivo.tamaño_legible,
                                    'fecha': convert_to_tijuana_time(archivo.fecha_subida).strftime('%d/%m/%Y %H:%M'),
                                    'url': archivo.archivo.url if archivo.archivo else None
                                })
                        
                        # Actualizar usuario con el del comentario si está disponible
                        if comentario.usuario:
                            usuario_comentario = comentario.usuario.get_full_name() or comentario.usuario.username
                            item_data['usuario'] = usuario_comentario
                            item_data['usuario_id'] = comentario.usuario.id
                    else:
                        # Si no se encuentra comentario, saltar esta actividad para evitar huérfanas
                        continue
                        
                except Exception as e:
                    # Si hay error, saltar esta actividad
                    continue
            
            timeline_data.append(item_data)
        
        return JsonResponse({
            'success': True,
            'timeline': timeline_data
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
def editar_comentario_oportunidad(request, comentario_id):
    """
    API para editar un comentario específico de una oportunidad
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Obtener el comentario
        comentario = get_object_or_404(OportunidadComentario, id=comentario_id)
        
        # Verificar permisos: solo el autor del comentario o supervisores pueden editarlo
        if comentario.usuario != request.user and not is_supervisor(request.user):
            return JsonResponse({'error': 'No tienes permisos para editar este comentario'}, status=403)
        
        # Obtener el nuevo contenido
        nuevo_contenido = request.POST.get('contenido', '').strip()
        
        if not nuevo_contenido:
            return JsonResponse({'error': 'El contenido del comentario no puede estar vacío'}, status=400)
        
        # Actualizar el comentario
        comentario.contenido = nuevo_contenido
        comentario.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Comentario actualizado exitosamente',
            'nuevo_contenido': nuevo_contenido,
            'fecha_actualizacion': convert_to_tijuana_time(comentario.fecha_actualizacion).strftime('%d/%m/%Y %H:%M')
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt  
def eliminar_comentario_oportunidad(request, comentario_id):
    """
    API para eliminar un comentario específico de una oportunidad
    """
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        # Obtener el comentario
        comentario = get_object_or_404(OportunidadComentario, id=comentario_id)
        
        # Verificar permisos: solo el autor del comentario o supervisores pueden eliminarlo
        if comentario.usuario != request.user and not is_supervisor(request.user):
            return JsonResponse({'error': 'No tienes permisos para eliminar este comentario'}, status=403)
        
        # Guardar información antes de eliminar
        oportunidad_id = comentario.oportunidad.id
        oportunidad = comentario.oportunidad
        usuario_comentario = comentario.usuario
        fecha_comentario = comentario.fecha_creacion
        
        print(f"🗑️ Eliminando comentario ID={comentario_id}, usuario={usuario_comentario}, fecha={fecha_comentario}")
        
        # Buscar y eliminar TODAS las actividades que podrían estar apuntando a este comentario
        try:
            # Estrategia más amplia: buscar todas las actividades de comentario que podrían estar relacionadas
            actividades_candidatas = OportunidadActividad.objects.filter(
                oportunidad=oportunidad,
                tipo='comentario'
            )
            
            actividades_eliminadas = 0
            for actividad in actividades_candidatas:
                # Verificar si esta actividad apunta al comentario que vamos a eliminar
                # usando el nuevo sistema de IDs
                deberia_eliminar = False
                
                # Estrategia 1: Buscar por ID directo en la descripción (nuevo sistema)
                import re
                match = re.search(r'\[COMENTARIO_ID:(\d+)\]', actividad.descripcion or '')
                if match:
                    comentario_referenciado = int(match.group(1))
                    if comentario_referenciado == comentario_id:
                        deberia_eliminar = True
                        print(f"🎯 Actividad {actividad.id} apunta al comentario que se va a eliminar: {comentario_id}")
                else:
                    # Estrategia 2: Fallback para actividades del sistema viejo
                    # Verificar por rango de tiempo
                    diff_tiempo = abs((actividad.fecha_creacion - fecha_comentario).total_seconds())
                    if diff_tiempo <= 300:  # 5 minutos
                        deberia_eliminar = True
                        print(f"⏱️ Actividad {actividad.id} encontrada por tiempo: diff={diff_tiempo}s")
                    
                    # Verificar por usuario y descripción similar
                    if (actividad.usuario == usuario_comentario and 
                        comentario.contenido in actividad.descripcion):
                        deberia_eliminar = True
                        print(f"📝 Actividad {actividad.id} encontrada por contenido")
                
                if deberia_eliminar:
                    print(f"🗑️ Eliminando actividad relacionada ID={actividad.id}")
                    actividad.delete()
                    actividades_eliminadas += 1
            
            print(f"✅ Eliminadas {actividades_eliminadas} actividades relacionadas")
            
        except Exception as e:
            print(f"⚠️ Error eliminando actividades relacionadas: {e}")
            # No fallar si no se pueden eliminar las actividades, el comentario sí se debe eliminar
        
        # Eliminar el comentario
        comentario.delete()
        print(f"✅ Comentario ID={comentario_id} eliminado exitosamente")
        
        return JsonResponse({
            'success': True,
            'message': 'Comentario eliminado exitosamente',
            'oportunidad_id': oportunidad_id
        })
        
    except Exception as e:
        print(f"❌ Error eliminando comentario: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def descargar_archivo_oportunidad(request, archivo_id):
    """
    Vista para descargar archivos adjuntos de oportunidades
    """
    try:
        archivo = get_object_or_404(OportunidadArchivo, id=archivo_id)
        
        # Verificar permisos: supervisores ven todo, usuarios solo archivos de sus oportunidades
        if not is_supervisor(request.user) and archivo.oportunidad.usuario != request.user:
            return JsonResponse({'error': 'No tienes permisos para descargar este archivo'}, status=403)
        
        # Verificar que el archivo existe
        if not archivo.archivo:
            return JsonResponse({'error': 'Archivo no encontrado'}, status=404)
        
        # Importar las clases necesarias para la respuesta
        from django.http import HttpResponse, Http404
        from django.utils.encoding import smart_str
        import os
        import mimetypes
        
        # Obtener la ruta del archivo
        file_path = archivo.archivo.path
        
        if not os.path.exists(file_path):
            raise Http404("El archivo no existe en el servidor")
        
        # Determinar el tipo MIME
        content_type, _ = mimetypes.guess_type(file_path)
        if content_type is None:
            content_type = 'application/octet-stream'
        
        # Leer el archivo
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=content_type)
        
        # Configurar headers para descarga
        filename = smart_str(archivo.nombre_original)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = os.path.getsize(file_path)
        
        return response
        
    except Exception as e:
        print(f"❌ Error descargando archivo: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def vista_previa_archivo_oportunidad(request, archivo_id):
    """
    Vista para mostrar archivos en vista previa (inline)
    """
    try:
        archivo = get_object_or_404(OportunidadArchivo, id=archivo_id)
        
        # Verificar permisos: supervisores ven todo, usuarios solo archivos de sus oportunidades
        if not is_supervisor(request.user) and archivo.oportunidad.usuario != request.user:
            return HttpResponse('No tienes permisos para ver este archivo', status=403)
        
        # Verificar que el archivo existe
        if not archivo.archivo:
            return HttpResponse('Archivo no encontrado', status=404)
        
        # Importar las clases necesarias para la respuesta
        from django.http import HttpResponse, Http404
        from django.utils.encoding import smart_str
        import os
        import mimetypes
        
        # Obtener la ruta del archivo
        file_path = archivo.archivo.path
        
        if not os.path.exists(file_path):
            raise Http404("El archivo no existe en el servidor")
        
        # Determinar el tipo MIME
        content_type, _ = mimetypes.guess_type(file_path)
        if content_type is None:
            content_type = 'application/octet-stream'
        
        # Servir el archivo inline siempre (para vista previa en nueva pestaña)
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type=content_type)
        
        # Configurar headers para vista inline (no descarga)
        filename = smart_str(archivo.nombre_original)
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        response['Content-Length'] = os.path.getsize(file_path)
        
        return response
        
    except Exception as e:
        print(f"❌ Error en vista previa de archivo: {e}")
        return HttpResponse(f'Error al abrir archivo: {str(e)}', status=500)


# ── Novedades ──────────────────────────────────────────────────────────────

@login_required
def novedades_view(request):
    """Página de novedades estilo Apple Tips."""
    config = NovedadesConfig.get()
    return render(request, 'novedades.html', {
        'es_supervisor': is_supervisor(request.user),
        'novedades_config': config,
    })


@login_required
@require_http_methods(['POST'])
def api_toggle_novedades_widget(request):
    """Activa o desactiva el widget de novedades (solo supervisores)."""
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No permitido'}, status=403)
    data = json.loads(request.body)
    config = NovedadesConfig.get()
    activando = bool(data.get('activo', False))
    config.widget_activo = activando
    if activando:
        config.activation_count += 1  # Nueva clave localStorage → todos vuelven a ver el widget
    config.save()
    return JsonResponse({
        'ok': True,
        'activo': config.widget_activo,
        'activation_count': config.activation_count,
    })


@login_required
@require_http_methods(['POST'])
def api_toggle_empleado_mes_widget(request):
    """Activa o desactiva el widget de empleado del mes (solo supervisores)."""
    if not is_supervisor(request.user):
        return JsonResponse({'error': 'No permitido'}, status=403)
    data = json.loads(request.body)
    config = NovedadesConfig.get()
    activando = bool(data.get('activo', False))
    config.em_widget_activo = activando
    if activando:
        config.em_activation_count += 1
    config.save()
    return JsonResponse({
        'ok': True,
        'activo': config.em_widget_activo,
    })


@login_required
@require_http_methods(['POST'])
def api_quick_crear_cliente(request):
    """Crea un cliente rápido desde el formulario de nueva oportunidad."""
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    nombre = data.get('nombre_empresa', '').strip()
    if not nombre:
        return JsonResponse({'error': 'Nombre de empresa requerido'}, status=400)

    cliente, created = Cliente.objects.get_or_create(
        nombre_empresa__iexact=nombre,
        defaults={
            'nombre_empresa': nombre,
            'asignado_a': request.user,
        }
    )
    return JsonResponse({
        'success': True,
        'created': created,
        'id': cliente.id,
        'nombre': cliente.nombre_empresa,
    })


@login_required
@require_http_methods(['POST'])
def api_quick_crear_contacto(request):
    """Crea un contacto rápido desde el formulario de nueva oportunidad."""
    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    nombre = data.get('nombre', '').strip()
    if not nombre:
        return JsonResponse({'error': 'Nombre requerido'}, status=400)

    empresa_id = data.get('empresa_id')
    empresa = None
    if empresa_id:
        try:
            empresa = Cliente.objects.get(id=empresa_id)
        except Cliente.DoesNotExist:
            pass

    contacto = Contacto.objects.create(
        nombre=nombre,
        apellido=data.get('apellido', '').strip(),
        cliente=empresa,
    )
    return JsonResponse({
        'success': True,
        'id': contacto.id,
        'nombre_completo': f"{contacto.nombre} {contacto.apellido}".strip(),
    })


@login_required
@require_http_methods(["GET"])
def api_desglose_cotizaciones(request):
    """Desglose de numero de cotizaciones por cliente."""
    from datetime import datetime
    from collections import Counter
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)
    es_supervisor = is_supervisor(user)

    now = datetime.now()
    mes_filter = request.GET.get('mes', str(now.month).zfill(2))
    anio_filter = request.GET.get('anio', str(now.year))
    vendedores_filter = request.GET.get('vendedores', '')
    vendedores_ids = [int(x) for x in vendedores_filter.split(',') if x.strip().isdigit()] if vendedores_filter else []

    try:
        anio_int = int(anio_filter)
    except ValueError:
        anio_int = now.year

    # Base queryset
    base_qs = TodoItem.objects.filter(anio_cierre=anio_int)
    if mes_filter != 'todos':
        base_qs = base_qs.filter(mes_cierre=mes_filter)
    if not es_supervisor:
        usuarios_visibles = get_usuarios_visibles_ids(user)
        if usuarios_visibles and len(usuarios_visibles) > 1:
            base_qs = base_qs.filter(usuario_id__in=usuarios_visibles)
        else:
            base_qs = base_qs.filter(usuario=user)
    elif vendedores_ids:
        base_qs = base_qs.filter(usuario_id__in=vendedores_ids)

    opp_ids = base_qs.values_list('id', flat=True)
    cot_qs = Cotizacion.objects.select_related('cliente').filter(
        Q(oportunidad_id__in=opp_ids) | Q(oportunidad__isnull=True, fecha_creacion__year=anio_int)
    )

    # Contar por cliente
    count_by_client = Counter()
    monto_by_client = {}
    client_names = {}

    for cot in cot_qs:
        if not cot.cliente_id:
            continue
        count_by_client[cot.cliente_id] += 1
        monto_by_client[cot.cliente_id] = monto_by_client.get(cot.cliente_id, Decimal('0')) + (cot.total or Decimal('0'))
        if cot.cliente_id not in client_names:
            client_names[cot.cliente_id] = cot.cliente.nombre_empresa if cot.cliente else 'Sin Cliente'

    rows = []
    for cid, count in count_by_client.most_common():
        rows.append({
            'cliente': client_names.get(cid, 'Sin Cliente'),
            'cliente_id': cid,
            'num_cotizaciones': count,
            'monto_total': str(monto_by_client.get(cid, Decimal('0'))),
        })

    return JsonResponse({
        'ok': True,
        'rows': rows,
        'total': sum(count_by_client.values()),
    })
